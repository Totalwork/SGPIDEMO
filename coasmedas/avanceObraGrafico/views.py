from django.shortcuts import render,redirect,render_to_response
from django.urls import reverse
from .models import APeriodicidadG,BEsquemaCapitulosG,DReglasEstadoG,CEsquemaCapitulosActividadesG,EPresupuesto,FDetallePresupuesto,PDetalleCambio,RDiagramaGrahm
from .models import GCapa,HNodo,JCantidadesNodo,KCronograma,RPorcentajeApoyo,LProgramacion,LPorcentaje,IEnlace,QEjecucionProgramada,MEstadoCambio,NCambio,LPorcentajePresupuesto
from rest_framework import viewsets, serializers
from django.db.models import Q,Sum,Prefetch
from logs.models import Logs,Acciones
from usuario.models import Usuario
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import xlsxwriter
import json
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.deletion import ProtectedError

from proyecto.views import ProyectoSerializer
from proyecto.models import Proyecto,Proyecto_empresas

from contrato.models import EmpresaContrato,Contrato
from contrato.views import ContratoSerializer
from contrato.enumeration import tipoC

from .tasks import updateAsyncEstado,createAsyncEstado

from tipo.models import Tipo

from empresa.models import Empresa

from parametrizacion.models import Departamento
from proyecto.models import Proyecto

from django.db import connection

from datetime import timedelta
import time

import math

from usuario.views import UsuarioSerializer
from empresa.views import EmpresaSerializer

from tipo.views import TipoSerializer
from tipo.models import Tipo

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from django.conf import settings
from sinin4.functions import functions
from django.db import transaction
from parametrizacion.views import  MunicipioSerializer , DepartamentoSerializer

import openpyxl

# Create your views here.


#Api rest para periodicidad
class PeriodicidadSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = APeriodicidadG
		fields=('id','nombre','numero_dias')


class PeriodicidadGViewSet(viewsets.ModelViewSet):
	
	model=APeriodicidadG
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='avance_de_obra_grafico.periodicidad'
	queryset = model.objects.all()
	serializer_class = PeriodicidadSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PeriodicidadGViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PeriodicidadSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PeriodicidadSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para periodicidad


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista',)


#Api rest para Esquema de Capitulos
class EsquemaCapitulosSerializer(serializers.HyperlinkedModelSerializer):

	tipo=tipoC()
	macrocontrato=ContratoLiteSerializer(read_only=True)
	macrocontrato_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.filter(tipo_contrato=tipo.m_contrato))	

	class Meta:
		model = BEsquemaCapitulosG
		fields=('id','macrocontrato','macrocontrato_id','nombre')


class EsquemaCapitulosGViewSet(viewsets.ModelViewSet):
	
	model=BEsquemaCapitulosG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosGViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			macrocontrato_id= self.request.query_params.get('macrocontrato_id',None)
			lista_contrato=[]
			tipo=tipoC()
			mcontrato=EmpresaContrato.objects.filter(contrato__tipo_contrato=tipo.m_contrato,empresa=request.user.usuario.empresa.id,participa=1)

			if mcontrato is None:
				lista_contrato.append(0)
			else:
				for item in mcontrato:
					lista_contrato.append(item.contrato.id)

			qset=(Q(macrocontrato_id__in=lista_contrato))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if macrocontrato_id is not None and int(macrocontrato_id)>0:
				qset=qset&(Q(macrocontrato_id=macrocontrato_id))


			
			
			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(macrocontrato_id=request.DATA['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(macrocontrato_id=request.DATA['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos



#Api rest para Esquema de Capitulos de las actividades
class EsquemaCapitulosActividadesSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True, required = False)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())	

	
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','esquema','esquema_id','nombre','nivel','padre','peso')



class EsquemaCapitulosActividadesGViewSet(viewsets.ModelViewSet):
	
	model=CEsquemaCapitulosActividadesG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosActividadesSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos_actividades'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosActividadesGViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			niveles= self.request.query_params.get('niveles',None)
			padre_id= self.request.query_params.get('padre_id',None)
			qset=None


			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if esquema_id is not None and int(esquema_id)>0:
				qset=qset &(Q(esquema_id=esquema_id))


			if padre_id is not None and int(padre_id)>0:
				qset=qset &(Q(padre=padre_id))


			if niveles is not None:
				lista=niveles.split(',')
				qset=qset&(Q(nivel__in=lista))

			if qset is not None:				
				queryset = self.model.objects.filter(qset)
			


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosActividadesSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.DATA['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
					valor=round(float(model_actividad['peso__sum']),3)+round(float(request.DATA['peso']),3)
					
					if float(valor) <= 100: 
						serializer.save(esquema_id=request.DATA['esquema_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.DATA['padre']

						if padre>0:
							if int(request.DATA['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(request.DATA['peso'])
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()								
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(request.DATA['peso'])
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosActividadesSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.DATA['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']

					model_actividad2=CEsquemaCapitulosActividadesG.objects.get(pk=instance.id)
					valor_restante=float(request.DATA['peso']) - float(model_actividad2.peso)
					valor=float(model_actividad['peso__sum'])+valor_restante

					if float(valor) <= 100: 
						serializer.save(esquema_id=request.DATA['esquema_id'])					
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.DATA['padre']

						if padre>0:
							if int(request.DATA['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(valor_restante)
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(valor_restante)
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos de las actividades



#Api rest para regla de estado
class ReglaEstadoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())

	reglaAnterior=serializers.SerializerMethodField()	

	class Meta:
		model = DReglasEstadoG
		fields=('id','esquema','esquema_id','orden','operador','limite','nombre','reglaAnterior',)

	def get_reglaAnterior(self, obj):
		return DReglasEstadoG.objects.filter(orden__lt=obj.orden,esquema_id=obj.esquema_id).values('id','nombre').order_by('orden').last()


class ReglaEstadoGraficoViewSet(viewsets.ModelViewSet):
	
	model=DReglasEstadoG
	queryset = model.objects.all()
	serializer_class = ReglaEstadoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.regla'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ReglaEstadoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			qset=None


			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)

			if esquema_id and esquema_id>0:
				if dato:
					qset = qset &(
					Q(esquema_id=esquema_id)
					)
				else:
					qset =(
					Q(esquema_id=esquema_id)
					)

			if qset !=None:
				queryset = self.model.objects.filter(qset).order_by('orden')


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				reglas=DReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id'])
				if len(reglas) > 0:
					for item in list(reglas):
						if int(item.operador)==int(request.DATA['operador']) and float(item.limite)==float(request.DATA['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				orden=None
				orden=DReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id']).order_by('orden').last()
				
				if orden is None:
					request.DATA['orden']=1
				else:
					request.DATA['orden']=orden.orden+1

				if int(request.DATA['regla_anterior'])>0:
					reglas2=DReglasEstadoG.objects.get(pk=request.DATA['regla_anterior'])
					valor=int(reglas2.orden)+1
					request.DATA['orden']=valor

					if float(reglas2.limite)>=float(request.DATA['limite']) and int(reglas2.operador)==int(request.DATA['operador']):
						return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
			 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=DReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
					if len(reglas3)>0:
						for item in reglas3:
							if float(item.limite)<=float(request.DATA['limite']) and int(item.operador)==int(request.DATA['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

						reglas_actualizar=DReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
						for item2 in reglas_actualizar:
							estado=DReglasEstadoG.objects.get(pk=item2.id)
							estado.orden=estado.orden+1
							estado.save()
				
				serializer = ReglaEstadoGraficoSerializer(data=request.DATA,context={'request': request})
				if serializer.is_valid():
					serializer.save(esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.DATA['esquema_id'])

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				reglas=DReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id'])
				for item in list(reglas):
					if item.id!=instance.id:
						if int(item.operador)==int(request.DATA['operador']) and float(item.limite)==float(request.DATA['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				if int(request.DATA['regla_anterior'])>0:					
					valor=DReglasEstadoG.objects.filter(orden__lt=request.DATA['orden'],esquema_id=request.DATA['esquema_id']).values('id','nombre','orden').order_by('orden').last()
					if valor is None:
						valor=0

					reglas2=DReglasEstadoG.objects.get(pk=request.DATA['regla_anterior'])
					if float(reglas2.limite)>=float(request.DATA['limite']) and int(reglas2.operador)==int(request.DATA['operador']):
							return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=DReglasEstadoG.objects.filter(orden__gt=request.DATA['orden'],esquema_id=request.DATA['esquema_id']).values('id','nombre','orden','limite','operador').order_by('orden').first()
					if reglas3 is not None:
							if float(reglas3['limite'])<=float(request.DATA['limite']) and int(reglas3['operador'])==int(request.DATA['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
						 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					if int(request.DATA['regla_anterior'])!=int(valor['id']):
						valor=int(reglas2.orden)+1
						request.DATA['orden']=valor	
						if reglas3 is not None:
							reglas_actualizar=DReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
							for item2 in reglas_actualizar:
								estado=DReglasEstadoG.objects.get(pk=item2.id)
								estado.orden=estado.orden+1
								estado.save()


				serializer = ReglaEstadoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.DATA['esquema_id'])
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)				
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para regla de estado



# Serializer de proyecto
class ProyectoLiteSerializer(serializers.HyperlinkedModelSerializer):

	municipio = MunicipioSerializer(read_only = True)
	mcontrato=ContratoLiteSerializer(read_only=True)

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre','municipio','mcontrato')

#Api rest para Presupuesto
class PresupuestoSerializer(serializers.HyperlinkedModelSerializer):

	proyecto=ProyectoLiteSerializer(read_only=True)
	proyecto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all())	

	esquema=EsquemaCapitulosSerializer(read_only=True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())	

	class Meta:
		model = EPresupuesto
		fields=('id','proyecto','proyecto_id','esquema','esquema_id','nombre','cantidad_cronograma')


class PresupuestoViewSet(viewsets.ModelViewSet):
	
	model=EPresupuesto
	queryset = model.objects.all()
	serializer_class = PresupuestoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PresupuestoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			departamento_id= self.request.query_params.get('departamento_id',None)
			municipio_id= self.request.query_params.get('municipio_id',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			alcance= self.request.query_params.get('alcance',None)


			qset=(~Q(id=0))

			if alcance:
				listproyecto=Proyecto_empresas.objects.filter(empresa_id=request.user.usuario.empresa.id).values_list('proyecto_id').distinct()
				qset = qset &(
					Q(proyecto_id__in=listproyecto)
					)

			if proyecto_id and int(proyecto_id)>0:
				qset = qset &(
					Q(proyecto_id=proyecto_id)
					)

			if departamento_id and int(departamento_id)>0:
				qset = qset &(
					Q(proyecto__municipio__departamento__id=departamento_id)
					)

			if municipio_id and int(municipio_id)>0:
				qset = qset &(
					Q(proyecto__municipio__id=municipio_id)
					)

			if esquema_id and int(esquema_id)>0:
				qset = qset &(
					Q(esquema_id=esquema_id)
					)	

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)|
					Q(proyecto__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PresupuestoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(proyecto_id=request.DATA['proyecto_id'],esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PresupuestoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(proyecto_id=request.DATA['proyecto_id'],esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Presupuesto


#Api rest para Detalle Presupuesto
class DetallePresupuestoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	class Meta:
		model = FDetallePresupuesto
		fields=('id','presupuesto','presupuesto_id','actividad','actividad_id','disponibilidad_cantidad_apoyo','cantidad_apoyo','codigoUC','descripcionUC','valorManoObra','valorMaterial','valorGlobal','cantidad','nombre_padre')


class DetallePresupuestoViewSet(viewsets.ModelViewSet):
	
	model=FDetallePresupuesto
	queryset = model.objects.all()
	serializer_class = DetallePresupuestoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.detalle_presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetallePresupuestoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			hito_id= self.request.query_params.get('hito_id',None)
			actividad_id= self.request.query_params.get('actividad_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_apoyo= self.request.query_params.get('listado_apoyo',None)
			id_apoyo= self.request.query_params.get('id_apoyo',None)



			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if actividad_id:
				qset = qset &(
					Q(actividad_id=actividad_id)
					)

			if hito_id:
				qset = qset &(
					Q(actividad__padre=hito_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				presupuesto_id=cronograma.presupuesto.id
				qset = qset &(
					Q(presupuesto_id=cronograma.presupuesto.id)
					)

			lista=[]
			if listado_apoyo and presupuesto_id and int(presupuesto_id)>0 and id_apoyo:
				lista=HNodo.objects.filter(presupuesto_id=presupuesto_id).exclude(id=id_apoyo).values('id','nombre')


			if dato:
				qset = qset &(
					Q(codigoUC__icontains=dato)|
					Q(descripcionUC__icontains=dato)|
					Q(actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if listado_apoyo:						
						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'apoyos':lista}})
					else:
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
			

			serializer = self.get_serializer(queryset,many=True)
			if listado_apoyo:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'apoyos':lista}})			
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetallePresupuestoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],actividad_id=request.DATA['actividad_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetallePresupuestoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],actividad_id=request.DATA['actividad_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle presupuesto


#Api rest para Capa
class CapaSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = GCapa
		fields=('id','nombre','color')



#Api rest para Nodo
class NodoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	

	porcentajeNodo=serializers.SerializerMethodField()

	class Meta:
		model = HNodo
		fields=('id','presupuesto','porcentajeNodo','presupuesto_id','capa','capa_id','longitud','latitud','noProgramado','nombre')

	def get_porcentajeNodo(self, obj):
		linea=LProgramacion.objects.filter(cronograma_id=self.context['request'].query_params.get('cronograma_id'),cantidadesNodo__nodo__id=obj.id,cantidadesNodo__cantidad__gt=0,tipo_linea=2)
		total=0
		cant=0
		total_por=0

		for item in linea:
			cantidad=QEjecucionProgramada.objects.filter(programacion_id=item.id).aggregate(Sum('cantidadEjecutada'))
			cantidad['cantidadEjecutada__sum']=0 if cantidad['cantidadEjecutada__sum']==None else cantidad['cantidadEjecutada__sum']

			if float(item.cantidadesNodo.cantidad) > 0:
				valor=float(cantidad['cantidadEjecutada__sum'])/float(item.cantidadesNodo.cantidad)
				total=total+valor
				cant=cant+1


		if cant>0:
			total_por=(total/cant)*100

		if total_por> 100:
			total_por=100

		return round(total_por, 2)


class NodoViewSet(viewsets.ModelViewSet):
	
	model=HNodo
	queryset = model.objects.all()
	serializer_class = NodoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(NodoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_enlaces= self.request.query_params.get('listado_enlaces',None)
			ejecucion= self.request.query_params.get('ejecucion',None)
			programando= self.request.query_params.get('programando',None)
			apoyo_cambio= self.request.query_params.get('apoyo_cambio',None)
			estados= self.request.query_params.get('id_estados',None)


			qset=(~Q(id=0))

			if apoyo_cambio is not None:
				listado_estado=estados.split(',')

				listado_id=[]
				cambio=NCambio.objects.filter(cronograma_id=cronograma_id,estado_id__in=listado_estado)

				listado_id_ejecucion=[]
				if int(ejecucion)>0:

					listado_por=[]
					if int(ejecucion) == 2:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__gte=100)
						
					else:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__lt=100)


					for obj in listado_por:
						listado_id_ejecucion.append(obj.apoyo.id)

				if cambio is not None:
					for obj in cambio:
						for item in obj.nodos.all():
							listado_id.append(item.id)

				cronograma=KCronograma.objects.get(pk=cronograma_id)
				
				qset=(Q(presupuesto_id=cronograma.presupuesto.id)&(Q(eliminado=False)))

				if len(listado_id_ejecucion)==0:
					qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

				if len(listado_id_ejecucion)>0:
					qset=qset&(Q(id__in=listado_id_ejecucion))


				if dato!='':
					qset=qset&(Q(nombre__icontains=dato))
			else:

				if presupuesto_id:
					qset = qset &(
						Q(presupuesto_id=presupuesto_id)
						)

				if programando and programando!='':
					qset = qset &(
						Q(noProgramado=programando)
						)

				if cronograma_id:
					cronograma=KCronograma.objects.get(pk=cronograma_id)
					presupuesto_id=cronograma.presupuesto.id
					qset = qset &(
						Q(presupuesto_id=cronograma.presupuesto.id)
						)

				if dato:
					qset = qset &(
						Q(nombre__icontains=dato)
						)


				lista=[]
				if listado_enlaces and cronograma_id:
					lista=IEnlace.objects.filter(nodoOrigen__presupuesto__id=presupuesto_id).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')


				# if ejecucion:
				# 	programacion=LProgramacion.objects.filter(cronograma_id=cronograma_id)

			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if listado_enlaces:
						return self.get_paginated_response({'message':'','success':'ok',
							'data':{'datos':serializer.data,'enlace':lista}})
					else:	
						return self.get_paginated_response({'message':'','success':'ok',
							'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			if listado_enlaces:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'enlace':lista}})
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()

			if request.DATA['longitud'] == '':
				request.DATA['longitud']=None

			if request.DATA['latitud'] == '':
				request.DATA['latitud']=None

			try:
				serializer = NodoSerializer(data=request.DATA,context={'request': request})

				nombre_nodo=HNodo.objects.filter(nombre=request.DATA['nombre'],presupuesto_id=request.DATA['presupuesto_id'])

				if len(nombre_nodo)>0:
					return Response({'message':'El nombre del apoyo ya existe, digite otro nombre','success':'ok',
						'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],capa_id=request.DATA['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.DATA['longitud'] == '':
					request.DATA['longitud']=None

				if request.DATA['latitud'] == '':
					request.DATA['latitud']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = NodoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],capa_id=request.DATA['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Nodo


#Api rest para Cantidad de Nodo
class CantidadNodoSerializer(serializers.HyperlinkedModelSerializer):

	detallepresupuesto=DetallePresupuestoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())	

	nodo=NodoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	class Meta:
		model = JCantidadesNodo
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodo','nodo_id','cantidad')


class CantidadNodoViewSet(viewsets.ModelViewSet):
	
	model=JCantidadesNodo
	queryset = model.objects.all()
	serializer_class = CantidadNodoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.cantidad_nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CantidadNodoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detalle_presupuesto_id= self.request.query_params.get('detalle_presupuesto_id',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)



			qset=(~Q(id=0))

			if detalle_presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detalle_presupuesto_id)
					)

			if nodo_id:
				qset = qset &(
					Q(nodo_id=nodo_id)
					)

			if presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto__presupuesto__id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(nodo__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CantidadNodoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CantidadNodoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cantidad Nodo



#Api rest para Enlance
class EnlaceSerializer(serializers.HyperlinkedModelSerializer):

	nodoOrigen=NodoSerializer(read_only=True)
	nodoOrigen_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())

	nodoDestino=NodoSerializer(read_only=True)
	nodoDestino_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	detallepresupuesto=DetallePresupuestoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	

	class Meta:
		model = IEnlace
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodoOrigen','nodoOrigen_id','nodoDestino','nodoDestino_id','capa','capa_id')


class EnlaceGraficoViewSet(viewsets.ModelViewSet):
	
	model=IEnlace
	queryset = model.objects.all()
	serializer_class = EnlaceSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.enlace'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EnlaceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detalle_presupuesto_id= self.request.query_params.get('detalle_presupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			nodo_origen= self.request.query_params.get('nodo_origen',None)
			nodo_destino= self.request.query_params.get('nodo_destino',None)

			qset=(~Q(id=0))

			if detalle_presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detalle_presupuesto_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				qset=qset &(
						Q(nodoOrigen__presupuesto__id=cronograma.presupuesto.id)&
						Q(nodoDestino__presupuesto__id=cronograma.presupuesto.id)
					)

			if nodo_origen:
				qset = qset &(
					Q(nodoOrigen_id=nodo_origen)
					)

			if nodo_destino:
				qset = qset &(
					Q(nodoDestino_id=nodo_destino)
					)


			if dato:
				qset = qset &(
					Q(nodoOrigen__nombre__icontains=dato)|
					Q(nodoDestino__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EnlaceSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],capa_id=request.DATA['capa_id'],nodoOrigen_id=request.DATA['nodoOrigen_id'],nodoDestino_id=request.DATA['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EnlaceSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],capa_id=request.DATA['capa_id'],nodoOrigen_id=request.DATA['nodoOrigen_id'],nodoDestino_id=request.DATA['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cantidad Nodo



#Api rest para Diagrama Grahm
class DiagramaGrahmSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	
	class Meta:
		model = RDiagramaGrahm
		fields=('id','presupuesto','presupuesto_id','actividad','actividad_id','nombre_padre','fechaInicio','fechaFinal')


class DiagramaGrahmViewSet(viewsets.ModelViewSet):
	
	model=RDiagramaGrahm
	queryset = model.objects.all()
	serializer_class = DiagramaGrahmSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.diagrama'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DiagramaGrahmViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)


			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(actividad__nombre__icontains=dato)
					)

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DiagramaGrahmSerializer(data=request.DATA,context={'request': request})

				validacion_actividad=RDiagramaGrahm.objects.filter(actividad_id=request.DATA['actividad_id'],presupuesto_id=request.DATA['presupuesto_id'])

				if len(validacion_actividad)>0:
					return Response({'message':'La actividad ya esta registrada en el diagrama.','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(actividad_id=request.DATA['actividad_id'],presupuesto_id=request.DATA['presupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DiagramaGrahmSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
			

				if serializer.is_valid():				
					serializer.save(actividad_id=request.DATA['actividad_id'],presupuesto_id=request.DATA['presupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Diagrama Grahm


#Api rest para Cronograma
class CronogramaSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	estado=ReglaEstadoGraficoSerializer(read_only=True, allow_null = True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=DReglasEstadoG.objects.all(),allow_null = True)	

	class Meta:
		model = KCronograma
		fields=('id','presupuesto','presupuesto_id','fechaInicio','estado','estado_id','programacionCerrada','nombre')


class CronogramaGraficoViewSet(viewsets.ModelViewSet):
	
	model=KCronograma
	queryset = model.objects.all()
	serializer_class = CronogramaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.cronograma'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CronogramaGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)

			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if request.DATA['estado_id']==0:
					request.DATA['estado_id']=None

				serializer = CronogramaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.DATA['estado_id']==0:
					request.DATA['estado_id']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CronogramaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cronograma


#Api rest para Linea
class LineaSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=KCronograma.objects.all())	

	cantidadesNodo=CantidadNodoSerializer(read_only=True)
	cantidadesNodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=JCantidadesNodo.objects.all())	

	class Meta:
		model = LProgramacion
		fields=('id','cronograma','cronograma_id','cantidadesNodo','cantidadesNodo_id','fecha','tipo_linea','cantidad_ejecutadas')


class LineaGraficoViewSet(viewsets.ModelViewSet):
	
	model=LProgramacion
	queryset = model.objects.all()
	serializer_class = LineaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.linea'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(LineaGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)
			tipo_linea= self.request.query_params.get('tipo_linea',None)
			filtro_porcentaje= self.request.query_params.get('filtro_porcentaje',None)
			sin_cantidad= self.request.query_params.get('sin_cantidad',None)

			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(cronograma_id=cronograma_id)
					)

			if nodo_id:
				qset = qset &(
					Q(cantidadesNodo__nodo__id=nodo_id)
					)

			if sin_cantidad:
				qset = qset &(
					Q(cantidadesNodo__cantidad__gt=0)
					)

			if tipo_linea:
				qset = qset &(
					Q(tipo_linea=tipo_linea)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			datos_porcentaje=[]
			if filtro_porcentaje and cronograma_id and tipo_linea:
				listado_porcentaje=LPorcentaje.objects.filter(tipo_linea=tipo_linea,cronograma_id=cronograma_id).values('fecha').annotate(porcentaje=Sum('porcentaje')).distinct()
				
				porcentaje=0
				for item in listado_porcentaje:
					porcentaje=porcentaje+float(item['porcentaje'])
					datos_porcentaje.append({
							'fecha':item['fecha'],
							'porcentaje':porcentaje
							})


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if filtro_porcentaje:	
						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'porcentajes':list(datos_porcentaje)}})
					else:						
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})

		
			serializer = self.get_serializer(queryset,many=True)
			if filtro_porcentaje:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'porcentajes':list(datos_porcentaje)}})			
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				serializer = LineaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(cronograma_id=request.DATA['cronograma_id'],cantidadesNodo_id=request.DATA['cantidadesNodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = LineaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(cronograma_id=request.DATA['cronograma_id'],cantidadesNodo_id=request.DATA['cantidadesNodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Linea




#Api rest para Porcentaje
class PorcentajeSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=KCronograma.objects.all())

	class Meta:
		model = LPorcentaje
		fields=('id','cronograma','cronograma_id','fecha','porcentaje')


class PorcentajeGraficoViewSet(viewsets.ModelViewSet):
	
	model=LPorcentaje
	queryset = model.objects.all()
	serializer_class = PorcentajeSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.porcentaje'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PorcentajeGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)

			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(cronograma_id=cronograma_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				serializer = PorcentajeSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PorcentajeSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Linea


#Api rest para Estado de cambio
class EstadoCambioAvanceSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = MEstadoCambio
		fields=('id','nombre',)


class EstadoCambioAvanceViewSet(viewsets.ModelViewSet):
	
	model=MEstadoCambio
	queryset = model.objects.all()
	serializer_class = EstadoCambioAvanceSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.estado_cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EstadoCambioAvanceViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EstadoCambioAvanceSerializer(data=request.DATA,context={'request': request})
			
				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EstadoCambioAvanceSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para estado de cambio



#Api rest para Cambio
class CambioAvanceSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=KCronograma.objects.all())

	estado=EstadoCambioAvanceSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=MEstadoCambio.objects.all())	

	solicitante=UsuarioSerializer(read_only=True)
	solicitante_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	empresa_tecnica=EmpresaSerializer(read_only=True)
	empresa_tecnica_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	empresa_financiera=EmpresaSerializer(read_only=True)
	empresa_financiera_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	tipo_accion=TipoSerializer(read_only=True)
	tipo_accion_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Tipo.objects.filter(app='avanceObraGrafico'))

	class Meta:
		model = NCambio
		fields=('id','cronograma','cronograma_id','fecha_format','motivoRechazoTecnico','motivoRechazoFinanciero','motivoCancelacion','estado','estado_id','motivo','nombre','fecha','solicitante','solicitante_id','empresa_tecnica','empresa_tecnica_id','empresa_financiera','empresa_financiera_id','tipo_accion','tipo_accion_id')


class CambioAvanceViewSet(viewsets.ModelViewSet):
	
	model=NCambio
	queryset = model.objects.all()
	serializer_class = CambioAvanceSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CambioAvanceViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)

			qset=(~Q(id=0))

			if cronograma_id:
				qset=qset &(
						Q(cronograma_id=cronograma_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CambioAvanceSerializer(data=request.DATA,context={'request': request})
				tipoT=Tipo()
				request.DATA['tipo_accion_id']=tipoT.ObtenerID('avanceObraGrafico',request.DATA['tipo_accion_id'])

				if serializer.is_valid():
					serializer.save(tipo_accion_id=request.DATA['tipo_accion_id'],empresa_financiera_id=request.DATA['empresa_financiera_id'],cronograma_id=request.DATA['cronograma_id'],estado_id=request.DATA['estado_id'],solicitante_id=request.DATA['solicitante_id'],empresa_tecnica_id=request.DATA['empresa_tecnica_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CambioAvanceSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(tipo_accion_id=request.DATA['tipo_accion_id'],empresa_financiera_id=request.DATA['empresa_financiera_id'],cronograma_id=request.DATA['cronograma_id'],estado_id=request.DATA['estado_id'],solicitante_id=request.DATA['solicitante_id'],empresa_tecnica_id=request.DATA['empresa_tecnica_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para cambio

#Api rest para Ejecucion Programada
class EjecucionProgramadaSerializer(serializers.HyperlinkedModelSerializer):

	programacion=LineaSerializer(read_only=True)
	programacion_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=LProgramacion.objects.all())

	class Meta:
		model = QEjecucionProgramada
		fields=('id','programacion','programacion_id','fecha','cantidadEjecutada','observacion')


class EjecucionProgramadaViewSet(viewsets.ModelViewSet):
	
	model=QEjecucionProgramada
	queryset = model.objects.all()
	serializer_class = EjecucionProgramadaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.ejecucion_programada'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EjecucionProgramadaViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)

			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(programacion__cronograma__id=cronograma_id)
					)

			if nodo_id:
				qset = qset &(
					Q(programacion__cantidadesNodo__nodo__id=nodo_id)
					)

			if dato:
				qset = qset &(
					Q(observacion__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				serializer = EjecucionProgramadaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(programacion_id=request.DATA['programacion_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EjecucionProgramadaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(programacion_id=request.DATA['programacion_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Ejecucion Programacion


#Inicio eliminacion de un esquema
@transaction.atomic
def eliminar_esquema(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				DReglasEstadoG.objects.filter(esquema_id=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.filter(esquema_id=item['id']).delete()
				BEsquemaCapitulosG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.esquema',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un esquema

@login_required
@transaction.atomic
def eliminar_id_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EPresupuesto.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
def descargar_plantilla_presupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	esquema_id= request.GET['id_esquema']	
				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)
	worksheet.set_column('I:I', 20)

	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)
	worksheet.write('G1', 'Valor M.O', format1)
	worksheet.write('H1', 'Valor Material', format1)
	worksheet.write('I1', 'Valor Global', format1)

	for item in hitos:

		if item.nivel==1:
			actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id,padre=item.id)

			for item2 in actividades:

				worksheet.write(row, col, item2.id,format2)
				worksheet.write(row, col+1, item.nombre,format2)
				worksheet.write(row, col+2,item2.nombre,format2)
				worksheet.write(row, col+3,'',format2)
				worksheet.write(row, col+4,'',format2)
				worksheet.write(row, col+5,'',format2)
				worksheet.write(row, col+6,'',format2)
				worksheet.write(row, col+7,'',format2)
				worksheet.write(row, col+8,'',format2)

				row +=1


	workbook.close()

	return response
    #return response


@login_required
@transaction.atomic
def guardar_presupuesto_archivo(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte,data_only = True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		revisar_detalle=FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


		if len(revisar_detalle) > 0:
			FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).delete()


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						count=0
						j=0
						for fila2 in hoja.rows:
							if fila2[0].value:
								if j == 0:
									j=1
								else:
									if int(fila[0].value)==int(fila2[0].value):
										count = count+1

						
						actividad_porce=CEsquemaCapitulosActividadesG.objects.get(pk=fila[0].value)
						porcentaje=float(actividad_porce.peso)/count

						detalle=FDetallePresupuesto(porcentaje=porcentaje,presupuesto_id=presupuesto_id,actividad_id=fila[0].value,codigoUC=fila[3].value,descripcionUC=fila[4].value,valorManoObra=fila[6].value,valorMaterial=fila[7].value,valorGlobal=fila[8].value,cantidad=fila[5].value)
						detalle.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.detalle_presupuesto',id_manipulado=detalle.id)
						logs_model.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico.detalle_presupuesto')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def actualizar_cantidad(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.actualizar_cantidad')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def cierre_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()


		presupuesto=EPresupuesto.objects.get(pk=respuesta['id_presupuesto'])
		presupuesto.cerrar_presupuesto=True
		presupuesto.save()	
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.presupuesto',id_manipulado=respuesta['id_presupuesto'])
		logs_model.save()	

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cierre_presupuesto')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def descargar_plantilla_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', 'Apoyo', format1)
	worksheet.write('B1', 'Longitud', format1)
	worksheet.write('C1', 'Latitud', format1)


	workbook.close()

	return response
    #return response



@login_required
def descargar_plantilla_apoyo_sinposicion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', 'Apoyo', format1)

	workbook.close()

	return response
    #return response


@login_required
def guardar_apoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=fila[1].value,latitud=fila[2].value,noProgramado=False)
						nodo.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.nodo',id_manipulado=nodo.id)
						logs_model.save()
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def guardar_apoyo_archivo_sinposicion(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=None,latitud=None,noProgramado=False)
						nodo.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.nodo',id_manipulado=nodo.id)
						logs_model.save()
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def descargar_plantilla_cantidadApoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_cantidadApoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cantidad_Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:Z', 20)


	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)

	apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	row=0
	col=6

	for item_apoyo in apoyos:
		worksheet.write(row,col, item_apoyo.nombre, format1)

		col +=1


	row=1
	col=0

	hitos = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


	for item in hitos:
		worksheet.write(row, col, item.id,format2)
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)
		worksheet.write(row, col+1,str(capitulo.nombre),format2)
		worksheet.write(row, col+2,item.actividad.nombre,format2)
		worksheet.write(row, col+3,item.codigoUC,format2)	
		worksheet.write(row, col+4,item.descripcionUC,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		count=6
		for item_apoyo in apoyos:
			worksheet.write(row,col+count,0, format2)
			count +=1

		row +=1


	workbook.close()

	return response
    #return response


@login_required
def guardar_cantidadApoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')

						count=6
						for item_apoyo  in apoyos:
							consultar_nodo=JCantidadesNodo.objects.filter(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id)
							if len(consultar_nodo)==0:
								nodo=JCantidadesNodo(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id,cantidad=fila[count].value)
								nodo.save()
								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.cantidad_nodo',id_manipulado=nodo.id)
								logs_model.save()
							else:
								nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
								nodo.cantidad=fila[count].value
								nodo.save()
								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
								logs_model.save()

							count +=1
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cantidad_nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#Inicio eliminacion de un esquema
@transaction.atomic
def guardar_cantidad_apoyo(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			nodo=JCantidadesNodo.objects.get(pk=item['id'])
			nodo.cantidad=item['cantidad']
			nodo.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cantidad_apoyo',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#eliminacion de cronograma
@transaction.atomic
def eliminar_id_cronograma(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=KCronograma.objects.get(pk=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.cronograma',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cronograma')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de cronograma


@login_required
def descargar_plantilla_lineabase(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_linea.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Plantilla')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	cronograma_id= request.GET['cronograma_id']	
	tipo_linea= request.GET['tipo_linea']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:H', 30)


	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion UUCC', format1)
	worksheet.write('F1', 'Apoyo', format1)
	worksheet.write('G1', 'Cantidad', format1)
	worksheet.write('H1', 'Fecha (dd/mm/aaaa)', format1)

	row=1
	col=0

	linea = LProgramacion.objects.filter(cronograma_id=cronograma_id,tipo_linea=tipo_linea)


	for item in linea:
		worksheet.write(row, col, item.id,format2)

		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.cantidadesNodo.detallepresupuesto.actividad.padre)

		worksheet.write(row, col+1,str(capitulo.nombre),format2)
		worksheet.write(row, col+2,item.cantidadesNodo.detallepresupuesto.actividad.nombre,format2)
		worksheet.write(row, col+3,item.cantidadesNodo.detallepresupuesto.codigoUC,format2)	
		worksheet.write(row, col+4,item.cantidadesNodo.detallepresupuesto.descripcionUC,format2)
		worksheet.write(row, col+5,item.cantidadesNodo.nodo.nombre,format2)	
		worksheet.write(row, col+6,item.cantidadesNodo.cantidad,format2)
		fecha=""
		if item.fecha is not None:
			valor=str(item.fecha).split("-")
			fecha=str(valor[2])+"/"+str(valor[1])+"/"+str(valor[0])

		worksheet.write(row, col+7,fecha,format2)

		row +=1


	workbook.close()

	return response
    #return response



@login_required
def guardar_lineabase_fecha_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		cronograma_id= request.POST['id_cronograma']
		tipo_linea= request.POST['tipo_linea']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						if fila[0].value!='':
							linea=LProgramacion.objects.get(pk=fila[0].value)
							fecha=None
							if fila[7].value!='' and fila[7].value is not None:
								valor1=str(fila[7].value).split(' ')
								valor=str(valor1[0]).split('/')
								if len(valor)==3:
									fecha=str(valor[2])+"-"+str(valor[1])+"-"+str(valor[0])									
								else:
									fecha=valor[0]
								
							linea.fecha=fecha
							linea.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.linea',id_manipulado=fila[0].value)
							logs_model.save()

		cronograma=KCronograma.objects.get(pk=cronograma_id)
		listado=FDetallePresupuesto.objects.filter(presupuesto_id=cronograma.presupuesto.id)

		porcentaje_total=0		
		for item in listado:	
			programacion=LProgramacion.objects.filter(cantidadesNodo__detallepresupuesto__id=item.id,tipo_linea=tipo_linea,cronograma_id=cronograma_id).values('cantidadesNodo__detallepresupuesto__id','fecha','cantidadesNodo__detallepresupuesto__cantidad','cantidadesNodo__detallepresupuesto__actividad__peso').annotate(cantidades=Sum('cantidadesNodo__cantidad')).distinct()
			sumador=0
			porcentaje=0

			for item2 in programacion:
					if item2['fecha'] is not None and item2['fecha'] !='':
						sumador=float(item2['cantidades'])
						valor1=sumador/float(item2['cantidadesNodo__detallepresupuesto__cantidad'])
						porcentaje=valor1*float(item2['cantidadesNodo__detallepresupuesto__actividad__peso'])
						if porcentaje>100:
							porcentaje=100

						quePorcentaje=LPorcentaje.objects.filter(detallepresupuesto_id=item2['cantidadesNodo__detallepresupuesto__id'],tipo_linea=tipo_linea,cronograma_id=cronograma_id,fecha=item2['fecha'])

						if len(quePorcentaje)==0:
							queryporcentaje=LPorcentaje(cronograma_id=cronograma_id,fecha=item2['fecha'],porcentaje=porcentaje,detallepresupuesto_id=item2['cantidadesNodo__detallepresupuesto__id'],tipo_linea=tipo_linea)
							queryporcentaje.save()
							logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.porcentaje',id_manipulado=queryporcentaje.id)
							logs2_model.save()
						else:
							queryporcentaje=LPorcentaje.objects.get(pk=quePorcentaje[0].id)
							queryporcentaje.porcentaje=porcentaje
							queryporcentaje.save()
							logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.porcentaje',id_manipulado=quePorcentaje[0].id)
							logs2_model.save()

	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.linea')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@transaction.atomic
def guardar_linea(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		cronograma=KCronograma.objects.get(pk=respuesta['id_cronograma'])
		cronograma.programacionCerrada=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cronograma',id_manipulado=respuesta['id_cronograma'])
		logs_model.save()

		queryset = JCantidadesNodo.objects.filter(detallepresupuesto__presupuesto__id=cronograma.presupuesto.id)
		
		if queryset.count()>0:
			cantidadesNodo = queryset.values('id')
			for cn in cantidadesNodo:
				programacion_verificar=LProgramacion.objects.filter(cronograma_id=respuesta['id_cronograma'],tipo_linea=2,cantidadesNodo=JCantidadesNodo.objects.get(id=cn['id']))
				if len(programacion_verificar)==0:
					programacion = LProgramacion(
						cronograma_id=respuesta['id_cronograma'],
						tipo_linea=2,
						cantidadesNodo=JCantidadesNodo.objects.get(id=cn['id'])
					)
					programacion.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cronograma')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Inicio clonacion de los esquemas
@login_required
@transaction.atomic
def clonacion_esquema(request):
	sid = transaction.savepoint()
	
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		esquema=BEsquemaCapitulosG(nombre=respuesta['nombre_esquema'],macrocontrato_id=respuesta['id_macrocontrato'])
		esquema.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.esquema',id_manipulado=esquema.id)
		logs_model.save()

		capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=respuesta['id_etiqueta'],nivel=1)

		for item in capitulos:
			hitos=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item.peso,nombre=item.nombre,nivel=item.nivel,padre=item.padre)
			hitos.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.esquema_capitulos',id_manipulado=hitos.id)
			logs_model2.save()

			actividades=CEsquemaCapitulosActividadesG.objects.filter(padre=item.id)

			for item2 in actividades:
				actividad=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item2.peso,nombre=item2.nombre,nivel=item2.nivel,padre=hitos.id)
				actividad.save()
				logs_model3=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.esquema_capitulos',id_manipulado=actividad.id)
				logs_model3.save()

		transaction.savepoint_commit(sid)		
		return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})
		#return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
		#			'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin clonacion de los esquemas



@login_required
def informe_cantidad_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_cantidad_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Informe')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	presupuesto_id= request.GET['presupuesto_id']	
				
	detalle_presupuesto = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 30)

	worksheet.write('A1', 'Hitos', format1)
	worksheet.write('B1', 'Actividad', format1)
	worksheet.write('C1', 'Cod. UUCC', format1)
	worksheet.write('D1', 'Descripcion UUCC', format1)
	worksheet.write('E1', 'Cantidad Global', format1)
	worksheet.write('F1', 'Cantidad de Apoyos', format1)

	for item in detalle_presupuesto:

		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col, str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)
		worksheet.write(row, col+3,item.descripcionUC,format2)
		worksheet.write(row, col+4,item.cantidad,format2)

		porcentaje=JCantidadesNodo.objects.filter(detallepresupuesto_id=item.id).aggregate(Sum('cantidad'))
		porcentaje['cantidad__sum']=0 if porcentaje['cantidad__sum']==None else porcentaje['cantidad__sum']		

		worksheet.write(row, col+5,str(porcentaje['cantidad__sum']),format2)

		row +=1


	workbook.close()

	return response
    #return response



@login_required
@transaction.atomic
def guardar_nodos_nuevos(request):
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		nodo=HNodo(eliminado=False,noProgramado=respuesta['noProgramado'],presupuesto_id=respuesta['presupuesto_id'],nombre=respuesta['nombre'],capa_id=respuesta['capa_id'],latitud=respuesta['latitud'],longitud=respuesta['longitud'])
		nodo.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.nodo',id_manipulado=nodo.id)
		logs_model.save()
		cambio=NCambio.objects.get(pk=respuesta['id_cambio'])
		cambio.nodos.add(nodo.id)
		cambio.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cambio',id_manipulado=respuesta['id_cambio'])
		logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		listado_nodo=HNodo.objects.filter(id=nodo.id).values('id','nombre','longitud','latitud')
		#transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':list(listado_nodo)})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_cambio(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cambio=NCambio.objects.get(pk=respuesta['id_nodo'])

		if int(respuesta['tipo_accion'])==1:
			for item in cambio.nodos.all():
				LProgramacion.objects.filter(cantidadesNodo__nodo__id=item.id).delete()
				logs_model3=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.programacion_nodo',id_manipulado=item.id)
				logs_model3.save()
				JCantidadesNodo.objects.filter(nodo_id=item.id).delete()
				logs_model1=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.cantidad_nodo',id_manipulado=item.id)
				logs_model1.save()
				HNodo.objects.get(pk=item.id).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.nodo',id_manipulado=item.id)
				logs_model.save()

		cambio.delete()
		logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.cambio',id_manipulado=respuesta['id_nodo'])
		logs_model2.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_nodo_eliminacion(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cambio=NCambio.objects.get(pk=respuesta['id_cambio'])

		for item in respuesta['lista']:
			cambio.nodos.add(item['id'])
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cambio',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def consultar_detalle_cambio(request):
	sid = transaction.savepoint()
	try:
		id_cambio=request.GET['detalle_cambio']

		detalle_cambio=NCambio.objects.get(pk=id_cambio)
		tipo_accion=Tipo.objects.get(pk=detalle_cambio.tipo_accion.id)
		

		listado_id=[]

		listado=[]

		for item in detalle_cambio.nodos.all():
			listado_id.append(item.id)

		detalle_actividades=LProgramacion.objects.filter(cantidadesNodo__nodo_id__in=listado_id,cronograma_id=detalle_cambio.cronograma_id,tipo_linea=2)
		for p in detalle_actividades:
				cantidad_nueva=PDetalleCambio.objects.filter(nodo_id=p.cantidadesNodo.nodo.id,detallePresupuesto_id=p.cantidadesNodo.detallepresupuesto.id).last()
				valor=0
				if cantidad_nueva is not None:
					valor=cantidad_nueva.cantidad

				listado.append({
							'nombre_apoyo': p.cantidadesNodo.nodo.nombre,
							'nombre_actividad':p.cantidadesNodo.detallepresupuesto.actividad.nombre,
							'cantidad_actual':p.cantidadesNodo.cantidad,
							'cantidad_nueva':valor,
							'fecha':p.fecha,
							'color':tipo_accion.color,
							'icono':tipo_accion.icono,
							'tipo':tipo_accion.codigo
							})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':list(listado)})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_nuevas_cantidades(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)


		for item in respuesta['lista']:
			if item['cantidad'] > 0:
				cantidad=JCantidadesNodo(detallepresupuesto_id=item['id'],nodo_id=respuesta['id_nodo'],cantidad=item['cantidad'])
				cantidad.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.cantidadnodo',id_manipulado=cantidad.id)
				logs_model.save()
				if item['fecha'] == '':
					item['fecha']=None
				programacion=LProgramacion(cronograma_id=respuesta['id_cronograma'],tipo_linea=2,cantidadesNodo_id=cantidad.id,fecha=item['fecha'])
				programacion.save()
				logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.programacion',id_manipulado=programacion.id)
				logs_model2.save()


		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido registrado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def cambio_estado(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cambio=NCambio.objects.get(pk=respuesta['id_cambio'])
		cambio.estado_id=respuesta['id_estado']

		if int(respuesta['id_estado'])==5:
			cambio.motivoRechazoTecnico=respuesta['motivo']

		if int(respuesta['id_estado'])==4:
			cambio.motivoCancelacion=respuesta['motivo']

		if int(respuesta['id_estado'])==6:
			cambio.motivoRechazoFinanciero=respuesta['motivo']

		cambio.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cambio',id_manipulado=respuesta['id_cambio'])
		logs_model.save()

		if int(respuesta['id_estado'])==2 or int(respuesta['id_estado'])==3:
			tipo=Tipo.objects.get(pk=respuesta['id_estado'])

			if tipo.codigo==2:
				cambio=NCambio.objects.get(pk=respuesta['id_cambio'])
				for obj in cambio.nodos.all():
					nodo=HNodo.objects.get(pk=obj.id)
					nodo.eliminado=True
					nodo.save()


		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def consultar_detalle_modificacion(request):
	sid = transaction.savepoint()
	try:
		id_nodo=request.GET['nodo_id']
		listado=[]
		detalle=JCantidadesNodo.objects.filter(nodo_id=id_nodo)

		for p in detalle:
					lista=PDetalleCambio.objects.filter(nodo_id=id_nodo,detallePresupuesto_id=p.detallepresupuesto.id).last()
					valor=p.cantidad
					if lista and len(lista)>0:
						valor=lista.cantidad

					padre=CEsquemaCapitulosActividadesG.objects.get(pk=p.detallepresupuesto.actividad.padre)
					listado.append({
							'id':p.detallepresupuesto.id,
							'hito':padre.nombre,
							'nombre_actividad':p.detallepresupuesto.actividad.nombre,
							'codigoUC':p.detallepresupuesto.codigoUC,
							'descripcionUC':p.detallepresupuesto.descripcionUC,
							'cantidad':valor
							})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':list(listado)})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
@transaction.atomic
def guardar_cambio_cantidades(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			cambio=QEjecucionProgramada(programacion_id=item['programacion_id'],cantidadEjecutada=item['cantidad'],fecha=item['fecha'],observacion=item['observacion'])
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.cambio_ejecutada',id_manipulado=cambio.id)
			logs_model.save()

		detalle=QEjecucionProgramada.objects.filter(programacion__cronograma__id=respuesta['id_cronograma'],programacion__cantidadesNodo__nodo__id=respuesta['id_nodo']).values('programacion__cantidadesNodo__detallepresupuesto__actividad_id','programacion__cantidadesNodo__cantidad','programacion__cantidadesNodo__detallepresupuesto__actividad__peso').annotate(cantidades=Sum('cantidadEjecutada')).distinct()

		total=0
		por=0
		for item2 in detalle:
			valor=float(item2['cantidades'])/float(item2['programacion__cantidadesNodo__cantidad'])
			por=por+valor

		total=(por/len(detalle))*100

		if total>100:
			total=100

		porcentaje_apoyo=RPorcentajeApoyo.objects.filter(apoyo_id=respuesta['id_nodo'],cronograma_id=respuesta['id_cronograma'])
		
		if len(porcentaje_apoyo)==0:
			porcentaje=RPorcentajeApoyo(cronograma_id=respuesta['id_cronograma'],apoyo_id=respuesta['id_nodo'],porcentaje=total)
			porcentaje.save()
		else:
			porcentaje=RPorcentajeApoyo.objects.get(pk=porcentaje_apoyo[0].id)
			porcentaje.porcentaje=total
			porcentaje.save()

		detalle_por=QEjecucionProgramada.objects.filter(programacion__cronograma__id=respuesta['id_cronograma'],programacion__tipo_linea=2).values('programacion__cantidadesNodo__detallepresupuesto_id','fecha','programacion__cantidadesNodo__detallepresupuesto__cantidad','programacion__cantidadesNodo__detallepresupuesto__porcentaje','programacion__cantidadesNodo__detallepresupuesto__valorManoObra').annotate(cantidades=Sum('cantidadEjecutada')).distinct()

		for obj in detalle_por:
			if obj['fecha'] is not None and obj['fecha'] !='':
					sumador=float(obj['cantidades'])
					valor1=sumador/float(obj['programacion__cantidadesNodo__detallepresupuesto__cantidad'])
					porcentaje=valor1*float(obj['programacion__cantidadesNodo__detallepresupuesto__porcentaje'])
					if porcentaje>100:
						porcentaje=100

					quePorcentaje=LPorcentaje.objects.filter(detallepresupuesto_id=obj['programacion__cantidadesNodo__detallepresupuesto_id'],tipo_linea=3,cronograma_id=respuesta['id_cronograma'],fecha=obj['fecha'])

					if len(quePorcentaje)==0:
						queryporcentaje=LPorcentaje(cronograma_id=respuesta['id_cronograma'],fecha=obj['fecha'],porcentaje=porcentaje,detallepresupuesto_id=obj['programacion__cantidadesNodo__detallepresupuesto_id'],tipo_linea=3)
						queryporcentaje.save()
						logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.porcentaje',id_manipulado=queryporcentaje.id)
						logs2_model.save()
					else:
						queryporcentaje=LPorcentaje.objects.get(pk=quePorcentaje[0].id)
						queryporcentaje.porcentaje=porcentaje
						queryporcentaje.save()
						logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.porcentaje',id_manipulado=quePorcentaje[0].id)
						logs2_model.save()

						

					quePorcentajepre=LPorcentajePresupuesto.objects.filter(detallepresupuesto_id=obj['programacion__cantidadesNodo__detallepresupuesto_id'],cronograma_id=respuesta['id_cronograma'],fecha=obj['fecha'])

					valor_ganando=sumador*float(obj['programacion__cantidadesNodo__detallepresupuesto__valorManoObra'])
					if len(quePorcentajepre)==0:
						queryporcentaje=LPorcentajePresupuesto(cronograma_id=respuesta['id_cronograma'],fecha=obj['fecha'],valor_ganando=valor_ganando,detallepresupuesto_id=obj['programacion__cantidadesNodo__detallepresupuesto_id'])
						queryporcentaje.save()
						logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.porcentaje_presupuesto',id_manipulado=queryporcentaje.id)
						logs2_model.save()
					else:
						queryporcentaje=LPorcentajePresupuesto.objects.get(pk=quePorcentajepre[0].id)
						queryporcentaje.valor_ganando=valor_ganando
						queryporcentaje.save()
						logs2_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.porcentaje_presupuesto',id_manipulado=quePorcentaje[0].id)
						logs2_model.save()


		createAsyncEstado.delay(respuesta['id_cronograma'])

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cambio_ejecutada')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def guardar_modificacion(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			cambio=PDetalleCambio(fechaProgramada=None,cantidad=float(item['cantidad']),detallePresupuesto_id=item['id_presupuesto'],cambio_id=respuesta['id_cambio'],nodo_id=respuesta['id_nodo'])
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico.cambio_detalle',id_manipulado=cambio.id)
			logs_model.save()

		cambio2=NCambio.objects.get(pk=respuesta['id_cambio'])
		cambio2.nodos.add(respuesta['id_nodo'])
		cambio2.save()
		logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cambio',id_manipulado=respuesta['id_cambio'])
		logs_model2.save()


		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cambio_detalle')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})






@login_required
@transaction.atomic
def consultar_avance_obra(request):
	sid = transaction.savepoint()
	try:
		id_cronograma=request.GET['cronograma_id']
		programando=request.GET['programando']
		dato=request.GET['dato']
		estados=request.GET['id_estados']
		listado_estado=estados.split(',')
		ejecucion=request.GET['ejecucion']


		listado_id=[]
		cambio=NCambio.objects.filter(cronograma_id=id_cronograma,estado_id__in=listado_estado)

		listado_id_ejecucion=[]
		if int(ejecucion)>0:

			listado_por=[]
			if int(ejecucion) == 2:
				listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__gte=100)
				
			else:
				listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__lt=100)

			for obj in listado_por:
				listado_id_ejecucion.append(obj.apoyo.id)

		if cambio is not None:
			for obj in cambio:
				for item in obj.nodos.all():
					listado_id.append(item.id)

		
		cronograma=KCronograma.objects.get(pk=id_cronograma)
		
		qset=(Q(presupuesto_id=cronograma.presupuesto.id)&(Q(eliminado=False)))

		if len(listado_id_ejecucion)==0:
			qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

		if len(listado_id_ejecucion)>0:
			qset=qset&(Q(id__in=listado_id_ejecucion))


		if dato!='':
			qset=qset&(Q(nombre__icontains=dato))

		listado=HNodo.objects.filter(qset).values('id','nombre','longitud','latitud')

		qset1=Q(nodoOrigen__presupuesto__id=cronograma.presupuesto_id)

		if len(listado_id)>0:
			qset1=qset1&(Q(nodoOrigen__id__in=listado_id))

		if len(listado_id_ejecucion)>0:
			qset1=qset1&(Q(nodoOrigen__id__in=listado_id_ejecucion))

		listado_enlaces=IEnlace.objects.filter(qset1).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'datos':list(listado),'enlace':list(listado_enlaces)}})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

#eliminacion de regla de estado
@transaction.atomic
def eliminar_id_regla_estado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		esquema_id=None

		for item in respuesta['lista']:
			valor=DReglasEstadoG.objects.get(pk=item['id'])
			esquema_id=valor.esquema_id
			valor.delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.regla_estado',id_manipulado=item['id'])
			logs_model.save()	

		updateAsyncEstado.delay(esquema_id)	
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		functions.toLog(e,'avance_de_obra.regla_estado')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de regla de estado

@login_required
def informe_presupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']
	presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:H', 20)


	worksheet.write('A2', 'Presupuesto', format1)
	worksheet.write('B2', presupuesto.nombre, format1)

	worksheet.write('A3', 'Proyecto', format1)
	worksheet.write('B3', presupuesto.proyecto.nombre, format1)

	worksheet.write('A4', 'Esquema', format1)
	worksheet.write('B4', presupuesto.esquema.nombre, format1)

	worksheet.write('F6', 'Total Presupuesto', format1)

	worksheet.write('A7', 'Hitos', format1)
	worksheet.write('B7', 'Actividad', format1)
	worksheet.write('C7', 'codigo UUCC', format1)
	worksheet.write('D7', 'Descripcion UUCC', format1)
	worksheet.write('E7', 'Valor UUCC', format1)
	worksheet.write('F7', 'Cantidad', format1)
	worksheet.write('G7', 'Subtotal', format1)

	row=7
	col=0

	detalle = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)
	total=0

	for item in detalle:
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col,str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)	
		worksheet.write(row, col+3,item.descripcionUC,format2)

		valorManoObra=0
		if item.valorManoObra is not None:
			valorManoObra=item.valorManoObra

		valorMaterial=0
		if item.valorMaterial is not None:
			valorMaterial=item.valorMaterial

		valor=int(valorManoObra)+int(valorMaterial)

		worksheet.write(row, col+4,valor,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		subtotal=valor*int(item.cantidad)

		worksheet.write(row, col+6,subtotal,format2)

		total=total+subtotal

		row +=1


	worksheet.write('G6', total, format1)

	workbook.close()

	return response
    #return response


@login_required
@transaction.atomic
def eliminar_id_apoyo(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		HNodo.objects.get(id=respuesta['id']).delete()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.nodo',id_manipulado=respuesta['id'])
		logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.nodo')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




#Inicio eliminacion de varios apoyos
@transaction.atomic
def eliminar_apoyos(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				HNodo.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.apoyo',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de varios apoyos


#Inicio eliminacion de un diagrama
@transaction.atomic
def eliminar_diagrama(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				RDiagramaGrahm.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.diagrama',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un diagrama	


@login_required
def informe_diagrama(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_diagrama.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Diagrama Grannt')
	format1=workbook.add_format({'font_size':10,'bold':True})

	

	presupuesto_id= request.GET['presupuesto_id']	
	
	presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=presupuesto.esquema.id,nivel=1)

	worksheet.set_column('A:C', 25)

	cell_format = workbook.add_format()

	cell_format.set_pattern(1)  # This is optional when using a solid fill.
	cell_format.set_bg_color('green')

	#worksheet.write('A1', 'Id', format1)

	fecha_menor=RDiagramaGrahm.objects.filter(presupuesto_id=presupuesto_id).order_by('fechaInicio').first()	

	fecha_mayor=RDiagramaGrahm.objects.filter(presupuesto_id=presupuesto_id).order_by('fechaFinal').last()	

	date_nueva=fecha_mayor.fechaFinal-fecha_menor.fechaInicio
	semanas=math.ceil(date_nueva.days/7.0)

	worksheet.write(1, 0,'Capitulos / Actividades' ,format1)
	worksheet.write(1, 1, 'Fecha Inicio',format1)
	worksheet.write(1, 2, 'Fecha Final',format1)

	row=1
	col=3
	for i in range(int(semanas)):
		worksheet.write(row, col, "Semana "+str(i+1),format1)
		col +=1

	row=2
	col=0

	for item in hitos:

		actividades=RDiagramaGrahm.objects.filter(presupuesto_id=presupuesto_id,actividad__padre=item.id)

		if len(actividades)>0:
			worksheet.write(row, col, item.nombre,format1)

			fechacapitulo_menor=RDiagramaGrahm.objects.filter(presupuesto_id=presupuesto_id,actividad__padre=item.id).order_by('fechaInicio').first()
			fechacapitulo_mayor=RDiagramaGrahm.objects.filter(presupuesto_id=presupuesto_id,actividad__padre=item.id).order_by('fechaFinal').last()

			semana_inicio=0
			semana_final=0
			fecha_semana=fecha_menor.fechaInicio
			sw1=0
			sw2=0
			for i in range(int(semanas+1)):
				if fechacapitulo_menor.fechaInicio >=  fecha_semana:
					semana_inicio=i+1

				if fechacapitulo_mayor.fechaFinal <=  fecha_semana and sw2==0:
					semana_final=i+1
					sw2=1

				fecha_semana=fecha_semana + timedelta(days=7)

				
			for i in range(semana_inicio,semana_final):
				colu=col+2+semana_inicio
				worksheet.write(row, colu, "",cell_format)
				semana_inicio=semana_inicio+1


			row +=1		

			for item2 in actividades:

				worksheet.write(row, col, item2.actividad.nombre)
				worksheet.write(row, col+1, str(item2.fechaInicio))
				worksheet.write(row, col+2, str(item2.fechaFinal))

				semana_inicio=0
				semana_final=0
				fecha_semana=fecha_menor.fechaInicio
				sw1=0
				sw2=0
				for i in range(int(semanas+1)):
					if item2.fechaInicio >=  fecha_semana:
						semana_inicio=i+1

					if item2.fechaFinal <=  fecha_semana and sw2==0:
						semana_final=i+1
						sw2=1

					fecha_semana=fecha_semana + timedelta(days=7)

				
				for i in range(semana_inicio,semana_final):
					colu=col+2+semana_inicio
					worksheet.write(row, colu, "",cell_format)
					semana_inicio=semana_inicio+1


				row +=1


		

	workbook.close()

	return response
    #return response


@login_required
@transaction.atomic
def eliminar_id_nodo_destino(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			IEnlace.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.enlance',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
@transaction.atomic
def guardar_cambio_detalle(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)


		for item in respuesta['lista']:
			detalle=QEjecucionProgramada.objects.get(pk=item['id'])
			detalle.cantidadEjecutada=item['cantidad']
			detalle.fecha=item['fecha']
			detalle.observacion=item['observacion']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico.cambio_detalle',id_manipulado=detalle.id)
			logs_model.save()



		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido registrado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def consultar_graficos(request):
	sid = transaction.savepoint()
	try:
		cronograma_id=request.GET['cronograma_id']
		listado=[]
		listado_porcentaje=LPorcentaje.objects.filter(tipo_linea=1,cronograma_id=cronograma_id).values('fecha').annotate(porcentaje=Sum('porcentaje')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado.append({
					'fecha':item['fecha'],
					'porcentaje':round(porcentaje,2)
					})


		listado2=[]
		listado_porcentaje=LPorcentaje.objects.filter(tipo_linea=2,cronograma_id=cronograma_id).values('fecha').annotate(porcentaje=Sum('porcentaje')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado2.append({
					'fecha':item['fecha'],
					'porcentaje':round(porcentaje,2)
					})

		listado3=[]
		listado_porcentaje=LPorcentaje.objects.filter(tipo_linea=3,cronograma_id=cronograma_id).values('fecha').annotate(porcentaje=Sum('porcentaje')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado3.append({
					'fecha':item['fecha'],
					'porcentaje':round(porcentaje,2)
					})

		listado4=[]
		listado_prespuesto=LPorcentajePresupuesto.objects.filter(cronograma_id=cronograma_id).values('fecha').annotate(valor_ganando=Sum('valor_ganando')).distinct()
				
		valor_ganando=0
		for item in listado_prespuesto:
			valor_ganando=valor_ganando+float(item['valor_ganando'])
			
			listado4.append({
					'fecha':item['fecha'],
					'valor_ganando':valor_ganando
					})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'linea_base':list(listado),'programada':list(listado2),'avance':list(listado3),'presupuesto':list(listado4)}})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.grafico')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def avance_de_obra_grafico(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico/hitos.html',{'model':'esquemacapitulosg','app':'avanceObraGrafico','macrocontrato':ListMacro})


@login_required
def actividades(request,id_esquema):
	return render(request, 'avanceObraGrafico/actividades.html',{'model':'esquemacapitulosactividadesg','app':'avanceObraGrafico','id_esquema':id_esquema})


@login_required
def presupuesto(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico/presupuesto.html',{'model':'EPresupuesto','app':'avanceObraGrafico','macrocontrato':ListMacro})

@login_required
def presupuesto_encabezado(request,id_proyecto):
	nombre_proyecto=Proyecto.objects.get(pk=id_proyecto)
	ListEsquema=BEsquemaCapitulosG.objects.filter(macrocontrato_id=nombre_proyecto.mcontrato.id)
	return render(request, 'avanceObraGrafico/presupuesto_encabezado.html',{'model':'EPresupuesto','app':'avanceObraGrafico','proyecto_id':id_proyecto,'nombre_proyecto':nombre_proyecto.nombre,'esquema':ListEsquema})


@login_required
def presupuesto_detalle(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.esquema_id,nivel=1)
	return render(request, 'avanceObraGrafico/presupuesto_detalle.html',{'model':'EPresupuesto','app':'avanceObraGrafico','presupuesto_id':id_presupuesto,'nombre_presupuesto':nombre_presupuesto.nombre,'id_proyecto':nombre_presupuesto.proyecto.id,'nombre_proyecto':nombre_presupuesto.proyecto.nombre,'nombre_esquema':nombre_presupuesto.esquema.nombre,'esquema_id':nombre_presupuesto.esquema.id,'capitulos':capitulos,'cerrado':nombre_presupuesto.cerrar_presupuesto})


@login_required
def presupuesto_vista_detalle(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.esquema_id,nivel=1)
	return render(request, 'avanceObraGrafico/presupuesto_ver_detalle.html',{'model':'EPresupuesto','app':'avanceObraGrafico','presupuesto_id':id_presupuesto,'nombre_presupuesto':nombre_presupuesto.nombre,'id_proyecto':nombre_presupuesto.proyecto.id,'nombre_proyecto':nombre_presupuesto.proyecto.nombre,'nombre_esquema':nombre_presupuesto.esquema.nombre,'esquema_id':nombre_presupuesto.esquema.id,'capitulos':capitulos})


@login_required
def alcance(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico/alcance.html',{'model':'EPresupuesto','app':'avanceObraGrafico','macrocontrato':ListMacro})


@login_required
def apoyo(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico/apoyo.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'nombre_presupuesto':nombre_presupuesto.nombre,'nombre_proyecto':nombre_presupuesto.proyecto.nombre,'nombre_esquema':nombre_presupuesto.esquema.nombre})


@login_required
def cantidad_apoyo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.esquema_id,nivel=1)
	return render(request, 'avanceObraGrafico/cantidad_apoyo.html',{'model':'EPresupuesto','app':'avanceObraGrafico','presupuesto_id':id_presupuesto,'nombre_presupuesto':nombre_presupuesto.nombre,'nombre_proyecto':nombre_presupuesto.proyecto.nombre,'nombre_esquema':nombre_presupuesto.esquema.nombre,'capitulos':capitulos})


@login_required
def cantidad_apoyo_id(request,id_presupuesto,id_detalle):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	detalle_presupuesto=FDetallePresupuesto.objects.get(pk=id_detalle)
	return render(request, 'avanceObraGrafico/cantidad_apoyo_id.html',{'model':'EPresupuesto','app':'avanceObraGrafico','presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'detalle_presupuesto':detalle_presupuesto})


@login_required
def cronograma(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico/cronograma.html',{'model':'EPresupuesto','app':'avanceObraGrafico','macrocontrato':ListMacro})


@login_required
def cronograma_proyecto(request,id_presupuesto,id_proyecto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico/cronograma_proyecto.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_proyecto':id_proyecto,'id_presupuesto':id_presupuesto,'presupuesto':presupuesto})

@login_required
def actividades_lectura(request,id_presupuesto,id_proyecto,id_esquema):
	return render(request, 'avanceObraGrafico/actividades_lectura.html',{'model':'esquemacapitulosactividadesg','app':'avanceObraGrafico','id_esquema':id_esquema,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})

@login_required
def cantidad_ejecutar(request,id_presupuesto,id_proyecto):
	return render(request, 'avanceObraGrafico/cantidad_ejecutar.html',{'model':'esquemacapitulosactividadesg','app':'avanceObraGrafico','id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})


@login_required
def linea_base(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico/linea_base.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})


@login_required
def programacion(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico/programacion.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})


@login_required
def avance_obra(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	ListTipo=Tipo.objects.filter(app='AvanceObraGrafico')
	ListEmpresa=Empresa.objects.all()
	Listfiltro=MEstadoCambio.objects.all()
	return render(request, 'avanceObraGrafico/avance_obra.html',{'model':'EPresupuesto','app':'avanceObraGrafico','filtro':Listfiltro,'empresas':ListEmpresa,'id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto,'capa_id':capa.id,'tipos':ListTipo})


@login_required
def cambio(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico/cambios.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})


@login_required
def detalle_cambio(request,id_proyecto,id_cronograma,id_cambio):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	cambio=NCambio.objects.get(pk=id_cambio)
	return render(request, 'avanceObraGrafico/detalle_cambio.html',{'model':'EPresupuesto','app':'avanceObraGrafico','cambio':cambio,'id_cronograma':id_cronograma,'id_presupuesto':cronograma.presupuesto.id,'id_proyecto':id_proyecto,'id_cambio':id_cambio})

@login_required
def regla_estado(request,id_esquema):
	return render(request, 'avanceObraGrafico/regla_estado.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_esquema':id_esquema})


@login_required
def digrama_grahm(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=presupuesto.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico/diagrama.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_presupuesto':id_presupuesto,'id_proyecto':presupuesto.proyecto.id,'capitulos':capitulos})


@login_required
def grafico(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico/grafico.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto})


@login_required
def avance_obra_sin_gps(request,id_presupuesto,id_proyecto,id_cronograma):
	cronograma=KCronograma.objects.get(pk=id_cronograma)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	ListTipo=Tipo.objects.filter(app='AvanceObraGrafico')
	ListEmpresa=Empresa.objects.all()
	Listfiltro=MEstadoCambio.objects.all()
	return render(request, 'avanceObraGrafico/avance_obra_sin_gps.html',{'model':'EPresupuesto','app':'avanceObraGrafico','filtro':Listfiltro,'empresas':ListEmpresa,'id_cronograma':id_cronograma,'cronograma':cronograma,'id_presupuesto':id_presupuesto,'id_proyecto':id_proyecto,'capa_id':capa.id,'tipos':ListTipo})


@login_required
def apoyo_sin_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico/apoyo_sin_gps.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'nombre_presupuesto':nombre_presupuesto.nombre,'nombre_proyecto':nombre_presupuesto.proyecto.nombre,'nombre_esquema':nombre_presupuesto.esquema.nombre})
