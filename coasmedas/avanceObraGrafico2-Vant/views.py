from django.shortcuts import render,redirect,render_to_response
from django.urls import reverse
from .models import APeriodicidadG,BEsquemaCapitulosG,CEsquemaCapitulosActividadesG,CReglasEstadoG,Cronograma,EPresupuesto,FDetallePresupuesto
from .models import DiagramaGrahm,HNodo,GCapa,IEnlace,JCantidadesNodo,JReporteTrabajo,LCambio,KDetalleReporteTrabajo,MComentarioRechazo,LHistorialCambio
from .models import LDetalleCambio
from rest_framework import viewsets, serializers
from django.db.models import Q,Sum,Prefetch
from logs.models import Logs,Acciones
from usuario.models import Usuario
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import xlsxwriter
import json
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.deletion import ProtectedError

from proyecto.views import ProyectoLiteSerializer
from proyecto.models import Proyecto,Proyecto_empresas

from contrato.models import EmpresaContrato,Contrato
from contrato.views import ContratoSerializer
from contrato.enumeration import tipoC

from .tasks import updateAsyncEstado,agregarSinPoste

from tipo.models import Tipo

from empresa.models import Empresa

from parametrizacion.models import Departamento

from django.db import connection

from datetime import timedelta
import time

import math

from usuario.views import UsuarioSerializer
from empresa.views import EmpresaSerializer

from estado.models import Estado
from estado.views import EstadoSerializer

from tipo.views import TipoSerializer
from tipo.models import Tipo

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from django.conf import settings
from sinin4.functions import functions
from django.db import transaction
from parametrizacion.views import  MunicipioSerializer , DepartamentoSerializer

import openpyxl

# Create your views here.


#Api rest para periodicidad
class PeriodicidadSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = APeriodicidadG
		fields=('id','nombre','numero_dias')


class PeriodicidadGraficoViewSet(viewsets.ModelViewSet):
	
	model=APeriodicidadG
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='avance_de_obra_grafico2.periodicidad'
	queryset = model.objects.all()
	serializer_class = PeriodicidadSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PeriodicidadGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PeriodicidadSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PeriodicidadSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para periodicidad


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista',)



#Api rest para Esquema de Capitulos
class EsquemaCapitulosSerializer(serializers.HyperlinkedModelSerializer):

	tipo=tipoC()
	macrocontrato=ContratoLiteSerializer(read_only=True)
	macrocontrato_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.filter(tipo_contrato=tipo.m_contrato))	

	class Meta:
		model = BEsquemaCapitulosG
		fields=('id','macrocontrato','macrocontrato_id','nombre')


class EsquemaCapitulosGraficoViewSet(viewsets.ModelViewSet):
	
	model=BEsquemaCapitulosG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			macrocontrato_id= self.request.query_params.get('macrocontrato_id',None)
			lista_contrato=[]
			tipo=tipoC()
			mcontrato=EmpresaContrato.objects.filter(contrato__tipo_contrato=tipo.m_contrato,empresa=request.user.usuario.empresa.id,participa=1)

			if mcontrato is None:
				lista_contrato.append(0)
			else:
				for item in mcontrato:
					lista_contrato.append(item.contrato.id)

			qset=(Q(macrocontrato_id__in=lista_contrato))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if macrocontrato_id is not None and int(macrocontrato_id)>0:
				qset=qset&(Q(macrocontrato_id=macrocontrato_id))


			
			
			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(macrocontrato_id=request.DATA['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(macrocontrato_id=request.DATA['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos



#Api rest para Esquema de Capitulos de las actividades
class EsquemaCapitulosActividadesSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True, required = False)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())	

	
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','esquema','esquema_id','nombre','nivel','padre','peso')



class EsquemaCapitulosActividadesGraficoViewSet(viewsets.ModelViewSet):
	
	model=CEsquemaCapitulosActividadesG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosActividadesSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos_actividades'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosActividadesGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			niveles= self.request.query_params.get('niveles',None)
			padre_id= self.request.query_params.get('padre_id',None)
			qset=None


			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if esquema_id is not None and int(esquema_id)>0:
				qset=qset &(Q(esquema_id=esquema_id))


			if padre_id is not None and int(padre_id)>0:
				qset=qset &(Q(padre=padre_id))


			if niveles is not None:
				lista=niveles.split(',')
				qset=qset&(Q(nivel__in=lista))

			if qset is not None:				
				queryset = self.model.objects.filter(qset)
			


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosActividadesSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.DATA['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
					valor=round(float(model_actividad['peso__sum']),3)+round(float(request.DATA['peso']),3)
					
					if float(valor) <= 100: 
						serializer.save(esquema_id=request.DATA['esquema_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.DATA['padre']

						if padre>0:
							if int(request.DATA['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(request.DATA['peso'])
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()								
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(request.DATA['peso'])
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosActividadesSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.DATA['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']

					model_actividad2=CEsquemaCapitulosActividadesG.objects.get(pk=instance.id)
					valor_restante=float(request.DATA['peso']) - float(model_actividad2.peso)
					valor=float(model_actividad['peso__sum'])+valor_restante

					if float(valor) <= 100: 
						serializer.save(esquema_id=request.DATA['esquema_id'])					
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.DATA['padre']

						if padre>0:
							if int(request.DATA['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(valor_restante)
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(valor_restante)
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos de las actividades




#Api rest para regla de estado
class ReglaEstadoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())

	reglaAnterior=serializers.SerializerMethodField()	

	class Meta:
		model = CReglasEstadoG
		fields=('id','esquema','esquema_id','orden','operador','limite','nombre','reglaAnterior',)

	def get_reglaAnterior(self, obj):
		return CReglasEstadoG.objects.filter(orden__lt=obj.orden,esquema_id=obj.esquema_id).values('id','nombre').order_by('orden').last()


class ReglaEstadoAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=CReglasEstadoG
	queryset = model.objects.all()
	serializer_class = ReglaEstadoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.regla'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ReglaEstadoAvanceGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			qset=None


			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)

			if esquema_id and esquema_id>0:
				if dato:
					qset = qset &(
					Q(esquema_id=esquema_id)
					)
				else:
					qset =(
					Q(esquema_id=esquema_id)
					)

			if qset !=None:
				queryset = self.model.objects.filter(qset).order_by('orden')


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				reglas=CReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id'])
				if len(reglas) > 0:
					for item in list(reglas):
						if int(item.operador)==int(request.DATA['operador']) and float(item.limite)==float(request.DATA['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				orden=None
				orden=CReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id']).order_by('orden').last()
				
				if orden is None:
					request.DATA['orden']=1
				else:
					request.DATA['orden']=orden.orden+1

				if int(request.DATA['regla_anterior'])>0:
					reglas2=CReglasEstadoG.objects.get(pk=request.DATA['regla_anterior'])
					valor=int(reglas2.orden)+1
					request.DATA['orden']=valor

					if float(reglas2.limite)>=float(request.DATA['limite']) and int(reglas2.operador)==int(request.DATA['operador']):
						return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
			 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
					if len(reglas3)>0:
						for item in reglas3:
							if float(item.limite)<=float(request.DATA['limite']) and int(item.operador)==int(request.DATA['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

						reglas_actualizar=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
						for item2 in reglas_actualizar:
							estado=CReglasEstadoG.objects.get(pk=item2.id)
							estado.orden=estado.orden+1
							estado.save()
				
				serializer = ReglaEstadoGraficoSerializer(data=request.DATA,context={'request': request})
				if serializer.is_valid():
					serializer.save(esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.DATA['esquema_id'])

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				reglas=CReglasEstadoG.objects.filter(esquema_id=request.DATA['esquema_id'])
				for item in list(reglas):
					if item.id!=instance.id:
						if int(item.operador)==int(request.DATA['operador']) and float(item.limite)==float(request.DATA['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				if int(request.DATA['regla_anterior'])>0:					
					valor=CReglasEstadoG.objects.filter(orden__lt=request.DATA['orden'],esquema_id=request.DATA['esquema_id']).values('id','nombre','orden').order_by('orden').last()
					if valor is None:
						valor=0

					reglas2=CReglasEstadoG.objects.get(pk=request.DATA['regla_anterior'])
					if float(reglas2.limite)>=float(request.DATA['limite']) and int(reglas2.operador)==int(request.DATA['operador']):
							return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=CReglasEstadoG.objects.filter(orden__gt=request.DATA['orden'],esquema_id=request.DATA['esquema_id']).values('id','nombre','orden','limite','operador').order_by('orden').first()
					if reglas3 is not None:
							if float(reglas3['limite'])<=float(request.DATA['limite']) and int(reglas3['operador'])==int(request.DATA['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
						 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					if int(request.DATA['regla_anterior'])!=int(valor['id']):
						valor=int(reglas2.orden)+1
						request.DATA['orden']=valor	
						if reglas3 is not None:
							reglas_actualizar=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.DATA['esquema_id'])
							for item2 in reglas_actualizar:
								estado=CReglasEstadoG.objects.get(pk=item2.id)
								estado.orden=estado.orden+1
								estado.save()


				serializer = ReglaEstadoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.DATA['esquema_id'])
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)				
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para regla de estado

#serializer lite de proyecto
class ProyectoLite2Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	municipio = MunicipioSerializer(read_only = True)
	contrato = ContratoLiteSerializer(read_only = True, many = True)

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'mcontrato' ,
				 'municipio' ,
				 'contrato',
				 )


#Api rest para Cronograma
class CronogramaSerializer(serializers.HyperlinkedModelSerializer):

	proyecto=ProyectoLite2Serializer(read_only=True)
	proyecto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all())	

	estado=ReglaEstadoGraficoSerializer(read_only=True, allow_null = True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CReglasEstadoG.objects.all(),allow_null = True)	

	periodicidad=PeriodicidadSerializer(read_only=True, allow_null = True)
	periodicidad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=APeriodicidadG.objects.all(),allow_null = True)	

	esquema=EsquemaCapitulosSerializer(read_only=True, allow_null = True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all(),allow_null = True)	


	class Meta:
		model = Cronograma
		fields=('id','proyecto','proyecto_id','periodicidad','esquema_id','esquema','periodicidad_id','estado','estado_id','programacionCerrada','nombre')


class CronogramaAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=Cronograma
	queryset = model.objects.all()
	serializer_class = CronogramaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cronograma'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CronogramaAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			esquema_id= self.request.query_params.get('esquema_id',None)


			qset=(~Q(id=0))

			if proyecto_id:
				qset = qset &(
					Q(proyecto_id=proyecto_id)
					)

			if esquema_id:
				qset = qset &(
					Q(esquema_id=esquema_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if request.DATA['estado_id']==0:
					request.DATA['estado_id']=None

				serializer = CronogramaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(periodicidad_id=request.DATA['periodicidad_id'],proyecto_id=request.DATA['proyecto_id'],estado_id=request.DATA['estado_id'],esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.DATA['estado_id']==0:
					request.DATA['estado_id']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CronogramaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(periodicidad_id=request.DATA['periodicidad_id'],proyecto_id=request.DATA['proyecto_id'],estado_id=request.DATA['estado_id'],esquema_id=request.DATA['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cronograma


#Api rest para Presupuesto
class PresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())	

	totalCambios=serializers.SerializerMethodField()

	class Meta:
		model = EPresupuesto
		fields=('id','cronograma','cronograma_id','cerrar_presupuesto','nombre','totalCambios','sin_poste')


	def get_totalCambios(self, obj):
		return LCambio.objects.filter(presupuesto_id=obj.id).count()


class PresupuestoAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=EPresupuesto
	queryset = model.objects.all()
	serializer_class = PresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PresupuestoAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			cerrado_presupuesto= self.request.query_params.get('cerrado_presupuesto',None)


			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(cronograma_id=cronograma_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if cerrado_presupuesto:
				qset = qset &(
					Q(cerrar_presupuesto=cerrado_presupuesto)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PresupuestoGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PresupuestoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Presupuesto




#Api rest para Detalle Presupuesto
class DetallePresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	sumaPresupuesto=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','sumaPresupuesto','presupuesto','presupuesto_id','actividad','disponibilidad_cantidad_apoyo','cantidad_apoyo','actividad_id','codigoUC','descripcionUC','valorManoObra','valorMaterial','valorGlobal','cantidad','porcentaje','nombre_padre')


	def get_sumaPresupuesto(self, obj):
		sumatoria=0
		suma=FDetallePresupuesto.objects.filter(presupuesto_id=obj.presupuesto.id)	
		valor=0
		for item in suma:
			valor=float(item.valorMaterial)+float(item.valorManoObra)
			total=valor*float(item.cantidad)

			sumatoria=sumatoria+total

		return round(sumatoria,3)


class DetallePresupuestoGraficoViewSet(viewsets.ModelViewSet):
	
	model=FDetallePresupuesto
	queryset = model.objects.all()
	serializer_class = DetallePresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetallePresupuestoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			hito_id= self.request.query_params.get('hito_id',None)
			actividad_id= self.request.query_params.get('actividad_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_apoyo= self.request.query_params.get('listado_apoyo',None)
			id_apoyo= self.request.query_params.get('id_apoyo',None)



			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if actividad_id:
				qset = qset &(
					Q(actividad_id=actividad_id)
					)

			if hito_id:
				qset = qset &(
					Q(actividad__padre=hito_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				presupuesto_id=cronograma.presupuesto.id
				qset = qset &(
					Q(presupuesto_id=cronograma.presupuesto.id)
					)

			lista=[]
			if listado_apoyo and presupuesto_id and int(presupuesto_id)>0 and id_apoyo:
				lista=HNodo.objects.filter(presupuesto_id=presupuesto_id).exclude(id=id_apoyo).values('id','nombre')


			if dato:
				qset = qset &(
					Q(codigoUC__icontains=dato)|
					Q(descripcionUC__icontains=dato)|
					Q(actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if listado_apoyo:						
						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'apoyos':lista}})
					else:
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
			

			serializer = self.get_serializer(queryset,many=True)
			if listado_apoyo:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'apoyos':lista}})			
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetallePresupuestoGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],actividad_id=request.DATA['actividad_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetallePresupuestoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],actividad_id=request.DATA['actividad_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle presupuesto


#Api rest para Diagrama Grahm
class DiagramaGrahmGraficoSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	
	class Meta:
		model = DiagramaGrahm
		fields=('id','cronograma','cronograma_id','actividad','actividad_id','nombre_padre','fechaInicio','fechaFinal','actividad_inicial')


class DiagramaGrahmGraficoViewSet(viewsets.ModelViewSet):
	
	model=DiagramaGrahm
	queryset = model.objects.all()
	serializer_class = DiagramaGrahmGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.diagrama'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DiagramaGrahmGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)


			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(actividad__nombre__icontains=dato)
					)

			if cronograma_id:
				qset = qset &(
					Q(cronograma_id=cronograma_id)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DiagramaGrahmGraficoSerializer(data=request.DATA,context={'request': request})

				validacion_actividad=DiagramaGrahm.objects.filter(actividad_id=request.DATA['actividad_id'],cronograma_id=request.DATA['cronograma_id'])

				if len(validacion_actividad)>0:
					return Response({'message':'La actividad ya esta registrada en el diagrama.','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(actividad_id=request.DATA['actividad_id'],cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DiagramaGrahmGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
			

				if serializer.is_valid():				
					serializer.save(actividad_id=request.DATA['actividad_id'],cronograma_id=request.DATA['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Diagrama Grahm


#Api rest para Capa
class CapaSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = GCapa
		fields=('id','nombre','color')

#Api rest para Nodo
class NodoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	


	class Meta:
		model = HNodo
		fields=('id','presupuesto','porcentajeAcumulado','presupuesto_id','eliminado','capa','capa_id','longitud','latitud','noProgramado','nombre')


class NodoGraficoViewSet(viewsets.ModelViewSet):
	
	model=HNodo
	queryset = model.objects.all()
	serializer_class = NodoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(NodoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_enlaces= self.request.query_params.get('listado_enlaces',None)
			ejecucion= self.request.query_params.get('ejecucion',None)
			programando= self.request.query_params.get('programando',None)
			filtro_gps= self.request.query_params.get('filtro_gps',None)
			eliminado= self.request.query_params.get('eliminado',None)
			apoyo_cambio= self.request.query_params.get('apoyo_cambio',None)
			estados= self.request.query_params.get('id_estados',None)


			qset=(~Q(id=0))

			if filtro_gps is not None and int(filtro_gps)>0:
				if int(filtro_gps)==2:
					qset=qset&(Q(latitud=None))&(Q(longitud=None))
				if int(filtro_gps)==1:
					qset=qset&(~Q(latitud=None))&(~Q(longitud=None))


			if apoyo_cambio is not None:
				listado_estado=estados.split(',')

				listado_id=[]
				cambio=LCambio.objects.filter(cronograma_id=cronograma_id,estado_id__in=listado_estado)

				listado_id_ejecucion=[]
				if int(ejecucion)>0:

					listado_por=[]
					if int(ejecucion) == 2:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__gte=100)
						
					else:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__lt=100)


					for obj in listado_por:
						listado_id_ejecucion.append(obj.apoyo.id)

				if cambio is not None:
					for obj in cambio:
						for item in obj.nodos.all():
							listado_id.append(item.id)

				cronograma=Cronograma.objects.get(pk=cronograma_id)
				
				qset=(Q(presupuesto_id=cronograma.presupuesto.id)&(Q(eliminado=False)))

				if len(listado_id_ejecucion)==0:
					qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

				if len(listado_id_ejecucion)>0:
					qset=qset&(Q(id__in=listado_id_ejecucion))


				if dato!='':
					qset=qset&(Q(nombre__icontains=dato))
			else:

				if presupuesto_id:
					qset = qset &(
						Q(presupuesto_id=presupuesto_id)
						)

				if programando and programando!='':
					qset = qset &(
						Q(noProgramado=programando)
						)

				if eliminado and eliminado!='':
					qset = qset &(
						Q(eliminado=eliminado)
						)

				if cronograma_id:
					cronograma=Cronograma.objects.get(pk=cronograma_id)
					presupuesto_id=cronograma.presupuesto.id
					qset = qset &(
						Q(presupuesto_id=cronograma.presupuesto.id)
						)

				if dato:
					qset = qset &(
						Q(nombre__icontains=dato)
						)


				lista=[]
				if listado_enlaces and cronograma_id:
					lista=IEnlace.objects.filter(nodoOrigen__presupuesto__id=presupuesto_id).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')


				# if ejecucion:
				# 	programacion=LProgramacion.objects.filter(cronograma_id=cronograma_id)

			if qset is not None:
				#print qset
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if listado_enlaces:
						return self.get_paginated_response({'message':'','success':'ok',
							'data':{'datos':serializer.data,'enlace':lista}})
					else:	
						return self.get_paginated_response({'message':'','success':'ok',
							'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			if listado_enlaces:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'enlace':lista}})
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()

			if request.DATA['longitud'] == '' or request.DATA['latitud'] == '':
				request.DATA['longitud']=None
				request.DATA['latitud']=None

			try:
				serializer = NodoGraficoSerializer(data=request.DATA,context={'request': request})

				nombre_nodo=HNodo.objects.filter(nombre=request.DATA['nombre'],presupuesto_id=request.DATA['presupuesto_id'])

				if len(nombre_nodo)>0:
					return Response({'message':'El nombre del apoyo ya existe, digite otro nombre','success':'ok',
						'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],capa_id=request.DATA['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					presupuesto=EPresupuesto.objects.get(pk=request.DATA['presupuesto_id'])

					if presupuesto.sin_poste == True:
						agregarSinPoste.delay(presupuesto.id,request.user.usuario.id,serializer.data['id'])	


					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.DATA['longitud'] == '' or request.DATA['latitud'] == '':
					request.DATA['longitud']=None
					request.DATA['latitud']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = NodoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],capa_id=request.DATA['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Nodo



#Api rest para Cantidad de Nodo
class CantidadNodoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())	

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	class Meta:
		model = JCantidadesNodo
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodo','nodo_id','cantidad')


class CantidadNodoGraficoViewSet(viewsets.ModelViewSet):
	
	model=JCantidadesNodo
	queryset = model.objects.all()
	serializer_class = CantidadNodoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cantidad_nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CantidadNodoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detalle_presupuesto_id= self.request.query_params.get('detalle_presupuesto_id',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)



			qset=(~Q(id=0))

			if detalle_presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detalle_presupuesto_id)
					)

			if nodo_id:
				qset = qset &(
					Q(nodo_id=nodo_id)
					)

			if presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto__presupuesto__id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(nodo__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CantidadNodoGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CantidadNodoGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cantidad Nodo



#Api rest para Enlance
class EnlaceGraficoSerializer(serializers.HyperlinkedModelSerializer):

	nodoOrigen=NodoGraficoSerializer(read_only=True)
	nodoOrigen_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())

	nodoDestino=NodoGraficoSerializer(read_only=True)
	nodoDestino_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	

	class Meta:
		model = IEnlace
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodoOrigen','nodoOrigen_id','nodoDestino','nodoDestino_id','capa','capa_id')


class EnlaceAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=IEnlace
	queryset = model.objects.all()
	serializer_class = EnlaceGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.enlace'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EnlaceAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detallepresupuesto_id= self.request.query_params.get('detallepresupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			nodo_origen= self.request.query_params.get('nodo_origen',None)
			nodo_destino= self.request.query_params.get('nodo_destino',None)

			qset=(~Q(id=0))

			if detallepresupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detallepresupuesto_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				qset=qset &(
						Q(nodoOrigen__presupuesto__id=cronograma.presupuesto.id)&
						Q(nodoDestino__presupuesto__id=cronograma.presupuesto.id)
					)

			if nodo_origen:
				qset = qset &(
					Q(nodoOrigen_id=nodo_origen)
					)

			if nodo_destino:
				qset = qset &(
					Q(nodoDestino_id=nodo_destino)
					)


			if dato:
				qset = qset &(
					Q(nodoOrigen__nombre__icontains=dato)|
					Q(nodoDestino__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EnlaceGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],capa_id=request.DATA['capa_id'],nodoOrigen_id=request.DATA['nodoOrigen_id'],nodoDestino_id=request.DATA['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EnlaceGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],capa_id=request.DATA['capa_id'],nodoOrigen_id=request.DATA['nodoOrigen_id'],nodoDestino_id=request.DATA['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cantidad Nodo


#Api rest para Reporte de Trabajo
class ReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())

	usuario_registro=UsuarioSerializer(read_only=True)
	usuario_registro_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())
	
	usuario_aprueba=UsuarioSerializer(read_only=True, allow_null = True)
	usuario_aprueba_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all(), allow_null = True)

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app="Avance_obra_grafico"))

	empresa=EmpresaSerializer(read_only=True)
	empresa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	totalRechazados=serializers.SerializerMethodField()

	class Meta:
		model = JReporteTrabajo
		fields=('id','fecha_format','totalRechazados','reporteCerrado','empresa','empresa_id','presupuesto','presupuesto_id','estado','estado_id','fechaTrabajo','usuario_registro','usuario_registro_id','valor_ganando_acumulado','avance_obra_acumulado','sinAvance','motivoSinAvance','fecharevision','soporteAprobacion','usuario_aprueba','usuario_aprueba_id')


	def get_totalRechazados(self, obj):
		return MComentarioRechazo.objects.filter(reporte_trabajo_id=obj.id).count()



class ReporteTrabajoGraficoViewSet(viewsets.ModelViewSet):
	
	model=JReporteTrabajo
	queryset = model.objects.all()
	serializer_class = ReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ReporteTrabajoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			empresa_id= self.request.query_params.get('empresa_id',None)
			registrado= self.request.query_params.get('registrado',None)


			qset=(~Q(id=0))


			if empresa_id:
				lista=[]				

				if registrado:
					reportes=JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=proyecto_id)
					sw=0
					
					revision=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
					aprobado=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
					corregido=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
					rechazados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)

					for item in reportes:
						for valor in item.presupuesto.cronograma.proyecto.contrato.all():
							if int(valor.tipo_contrato.id)==9 and int(valor.contratista.id)==int(empresa_id):
								sw=1

						if sw==1:
							if int(registrado)==2:
								if int(item.estado.id)==int(corregido[0].id):
									lista.append(item.id)
							elif int(registrado)==3:
								if int(item.estado.id)==int(rechazados[0].id):
									lista.append(item.id)
							else:
								if int(item.estado.id)==int(revision[0].id) or int(item.estado.id)==int(aprobado[0].id):
									lista.append(item.id)						

					if sw == 0:
						qset = qset &(
							Q(empresa_id=empresa_id)
							)
					else:
						qset=qset & (Q(id__in=lista))


				else:
					reportes=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id)
					sw=0
					estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)
					for item in reportes:
						if sw==1:
							if int(item.estado.id)!=int(estados[0].id):
								lista.append(item.id)

						for valor in item.presupuesto.cronograma.proyecto.contrato.all():
							if int(valor.tipo_contrato.id)==9 and int(valor.contratista.id)==int(empresa_id):
								sw=1

					if sw == 0:
						qset = qset &(
							Q(empresa_id=empresa_id)
							)
					else:
						qset=qset & (Q(id__in=lista))


			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)		

			if proyecto_id:
				qset = qset &(
					Q(presupuesto__cronograma__proyecto__id=proyecto_id)
					)		


			if dato:
				qset = qset &(
					Q(presupuesto__nombre__icontains=dato)|
					Q(presupuesto__cronograma__nombre__icontains=dato)|
					Q(usuario_registro__persona__nombres__icontains=dato)|
					Q(usuario_registro__persona__apellidos__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-fechaTrabajo')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = ReporteTrabajoSerializer(data=request.DATA,context={'request': request})

				verifica_reporte=JReporteTrabajo.objects.filter(presupuesto_id=request.DATA['presupuesto_id'],fechaTrabajo=request.DATA['fechaTrabajo'])

				if len(verifica_reporte)>0:
					return Response({'message':'No se puede registrar dos reportes de trabajos con la misma fecha','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)

				if request.DATA['usuario_aprueba_id'] == '':
					request.DATA['usuario_aprueba_id']=None

				if serializer.is_valid():
					serializer.save(empresa_id=request.DATA['empresa_id'],presupuesto_id=request.DATA['presupuesto_id'],soporteAprobacion=self.request.FILES.get('soporteAprobacion') if self.request.FILES.get('soporteAprobacion') is not None else None,usuario_registro_id=request.DATA['usuario_registro_id'],usuario_aprueba_id=request.DATA['usuario_aprueba_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = ReporteTrabajoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				verifica_reporte=JReporteTrabajo.objects.filter((~Q(id=instance.id))&(Q(presupuesto_id=request.DATA['presupuesto_id']))&(Q(fechaTrabajo=request.DATA['fechaTrabajo'])))

				if len(verifica_reporte)>0:
					return Response({'message':'No se puede registrar dos reportes de trabajos con la misma fecha','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
				
				if serializer.is_valid():				
					serializer.save(empresa_id=request.DATA['empresa_id'],presupuesto_id=request.DATA['presupuesto_id'],usuario_registro_id=request.DATA['usuario_registro_id'],usuario_aprueba_id=request.DATA['usuario_aprueba_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Reporte


#Api rest para Detalle Reporte de Trabajo
class DetalleReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):

	reporte_trabajo=ReporteTrabajoSerializer(read_only=True)
	reporte_trabajo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=JReporteTrabajo.objects.all())

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())
	
	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(queryset=FDetallePresupuesto.objects.all())

	class Meta:
		model = KDetalleReporteTrabajo
		fields=('id','reporte_trabajo','reporte_trabajo_id','nodo','nodo_id','detallepresupuesto','detallepresupuesto_id','cantidadEjecutada')


class DetalleReporteTrabajoGraficoViewSet(viewsets.ModelViewSet):
	
	model=KDetalleReporteTrabajo
	queryset = model.objects.all()
	serializer_class = DetalleReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleReporteTrabajoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			reporte_id= self.request.query_params.get('reporte_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)
			detallepresupuesto_id= self.request.query_params.get('detallepresupuesto_id',None)

			qset=(~Q(id=0))

			if reporte_id:
				qset = qset &(
					Q(reporte_trabajo_id=reporte_id)
					)

			if nodo_id:
				qset = qset &(
					Q(nodo_id=nodo_id)
					)

			if detallepresupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detallepresupuesto_id)
					)		


			if dato:
				qset = qset &(
					Q(detallepresupuesto__actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetalleReporteTrabajoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(reporte_trabajo_id=request.DATA['reporte_trabajo_id'],nodo_id=request.DATA['nodo_id'],detallepresupuesto_id=request.DATA['detallepresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleReporteTrabajoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(reporte_trabajo_id=request.DATA['reporte_trabajo_id'],nodo_id=request.DATA['nodo_id'],detallepresupuesto_id=request.DATA['detallepresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Reporte


#Api rest para Cambio
class CambioGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())

	empresaSolicitante=EmpresaSerializer(read_only=True)
	empresaSolicitante_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())
	
	empresaTecnica=EmpresaSerializer(read_only=True)
	empresaTecnica_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	empresaFinanciera=EmpresaSerializer(read_only=True)
	empresaFinanciera_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app="Avance_obra_grafico_cambio"))

	mensajeCancelado=serializers.SerializerMethodField()

	class Meta:
		model = LCambio
		fields=('id','mensajeCancelado','presupuesto','presupuesto_id','estado','estado_id','descripcion','motivo','empresaSolicitante_id','empresaSolicitante','empresaTecnica','empresaTecnica_id','empresaFinanciera','empresaFinanciera_id')

	def get_mensajeCancelado(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
		historial=LHistorialCambio.objects.filter(estado_id=estados[0].id,cambio_id=obj.id).order_by('-id')
		mensaje=''

		if len(historial) > 0:
			mensaje=historial[0].motivoEstado
		return mensaje


class CambioGraficoViewSet(viewsets.ModelViewSet):
	
	model=LCambio
	queryset = model.objects.all()
	serializer_class = CambioGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CambioGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)

			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(descripcion__icontains=dato)|
					Q(motivo__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CambioGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],empresaSolicitante_id=request.DATA['empresaSolicitante_id'],empresaTecnica_id=request.DATA['empresaTecnica_id'],empresaFinanciera_id=request.DATA['empresaFinanciera_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					historial=LHistorialCambio(cambio_id=serializer.data['id'],usuario_registro_id=request.user.usuario.id,estado_id=request.DATA['estado_id'])
					historial.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='avanceObraGrafico2.historial_cambio',id_manipulado=historial.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CambioGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.DATA['presupuesto_id'],empresaSolicitante_id=request.DATA['empresaSolicitante_id'],empresaTecnica_id=request.DATA['empresaTecnica_id'],empresaFinanciera_id=request.DATA['empresaFinanciera_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cambio


#Api rest para mensaje de rechazo
class MensajeRechazoSerializer(serializers.HyperlinkedModelSerializer):

	reporte_trabajo=ReporteTrabajoSerializer(read_only=True)
	reporte_trabajo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=JReporteTrabajo.objects.all())

	class Meta:
		model = MComentarioRechazo
		fields=('id','reporte_trabajo','reporte_trabajo_id','fecha_format','fecha_hora','motivoRechazo')


class MensajeRechazoViewSet(viewsets.ModelViewSet):
	
	model=MComentarioRechazo
	queryset = model.objects.all()
	serializer_class = MensajeRechazoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.mensaje_rechazo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(MensajeRechazoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			reporte_trabajo_id= self.request.query_params.get('reporte_trabajo_id',None)

			qset=(~Q(id=0))

			if reporte_trabajo_id:
				qset = qset &(
					Q(reporte_trabajo_id=reporte_trabajo_id)
					)

			if dato:
				qset = qset &(
					Q(motivoRechazo__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-fecha_hora')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = MensajeRechazoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(reporte_trabajo_id=request.DATA['reporte_trabajo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = MensajeRechazoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(reporte_trabajo_id=request.DATA['reporte_trabajo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cambio


#Api rest para Detalle de cambio
class DetalleCambioGraficoSerializer(serializers.HyperlinkedModelSerializer):

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all(),allow_null = True)

	cambio=CambioGraficoSerializer(read_only=True)
	cambio_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=LCambio.objects.all())

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())
	
	class Meta:
		model = LDetalleCambio
		fields=('id','nodo','nodo_id','operacion','cantidadPropuesta','detallepresupuesto','detallepresupuesto_id','codigoUC','descripcionUC','valorManoObra','valorMaterial','valorGlobal','cambio','cambio_id')


class DetalleCambioGraficoViewSet(viewsets.ModelViewSet):
	
	model=LDetalleCambio
	queryset = model.objects.all()
	serializer_class = DetalleCambioGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleCambioGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cambio_id= self.request.query_params.get('cambio_id',None)


			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(detallepresupuesto__actividad__nombre__icontains=dato)
					)

			if cambio_id:
				qset = qset &(
					Q(cambio_id=cambio_id)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if int(request.DATA['detallepresupuesto_id'])==0:
					request.DATA['detallepresupuesto_id']=None

				serializer = DetalleCambioGraficoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],cambio_id=request.DATA['cambio_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleCambioGraficoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
			

				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.DATA['detallepresupuesto_id'],cambio_id=request.DATA['cambio_id'],nodo_id=request.DATA['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle de cambio



#Inicio eliminacion de un esquema
@transaction.atomic
def eliminar_esquema(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				CReglasEstadoG.objects.filter(esquema_id=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.filter(esquema_id=item['id']).delete()
				BEsquemaCapitulosG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.esquema',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un esquema



#eliminacion de capitulo/actividad de un esquema
@transaction.atomic
def eliminar_id_capitulo_actividad_esquema(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])

			if int(valor.padre)==0:
				CEsquemaCapitulosActividadesG.objects.filter(padre=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=item['id'])
				logs_model.save()

			else:
				modelhijos=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])
				model=CEsquemaCapitulosActividadesG.objects.get(pk=modelhijos.padre)
				model.peso=float(model.peso) - float(modelhijos.peso)
				model.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=model.id)
				logs_model.save()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=item['id'])
				logs_model2.save()

		transaction.savepoint_commit(sid)			
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

			
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra.capitulos_esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de capitulo/actividad de un esquema


#Inicio clonacion de los esquemas
@login_required
@transaction.atomic
def clonacion_esquema(request):
	sid = transaction.savepoint()
	
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		esquema=BEsquemaCapitulosG(nombre=respuesta['nombre_esquema'],macrocontrato_id=respuesta['id_macrocontrato'])
		esquema.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema',id_manipulado=esquema.id)
		logs_model.save()

		capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=respuesta['id_etiqueta'],nivel=1)

		for item in capitulos:
			hitos=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item.peso,nombre=item.nombre,nivel=item.nivel,padre=item.padre)
			hitos.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema_capitulos',id_manipulado=hitos.id)
			logs_model2.save()

			actividades=CEsquemaCapitulosActividadesG.objects.filter(padre=item.id)

			for item2 in actividades:
				actividad=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item2.peso,nombre=item2.nombre,nivel=item2.nivel,padre=hitos.id)
				actividad.save()
				logs_model3=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema_capitulos',id_manipulado=actividad.id)
				logs_model3.save()

		transaction.savepoint_commit(sid)		
		return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})
		#return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
		#			'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin clonacion de los esquemas


#eliminacion de regla de estado
@transaction.atomic
def eliminar_id_regla_estado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		esquema_id=None

		for item in respuesta['lista']:
			valor=CReglasEstadoG.objects.get(pk=item['id'])
			esquema_id=valor.esquema_id
			valor.delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.regla_estado',id_manipulado=item['id'])
			logs_model.save()	

		updateAsyncEstado.delay(esquema_id)	
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		print(e) 
		functions.toLog(e,'avance_de_obra.regla_estado')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de regla de estado


#eliminacion de cronograma
@transaction.atomic
def eliminar_id_cronograma(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=Cronograma.objects.get(pk=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.cronograma',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cronograma')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de cronograma


@login_required
@transaction.atomic
def eliminar_id_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EPresupuesto.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
def descargar_plantilla_presupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	esquema_id= request.GET['id_esquema']	
				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)
	worksheet.set_column('I:I', 20)

	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)
	worksheet.write('G1', 'Valor M.O', format1)
	worksheet.write('H1', 'Valor Material', format1)
	worksheet.write('I1', 'Valor Global', format1)

	for item in hitos:

		if item.nivel==1:
			actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id,padre=item.id)

			for item2 in actividades:

				worksheet.write(row, col, item2.id,format2)
				worksheet.write(row, col+1, item.nombre,format2)
				worksheet.write(row, col+2,item2.nombre,format2)
				worksheet.write(row, col+3,'',format2)
				worksheet.write(row, col+4,'',format2)
				worksheet.write(row, col+5,'',format2)
				worksheet.write(row, col+6,'',format2)
				worksheet.write(row, col+7,'',format2)
				worksheet.write(row, col+8,'',format2)

				row +=1


	workbook.close()

	return response
    #return response


@login_required
@transaction.atomic
def guardar_presupuesto_archivo(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte,data_only = True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		revisar_detalle=FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


		if len(revisar_detalle) > 0:
			FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).delete()


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						count=0
						j=0
						for fila2 in hoja.rows:
							if fila2[0].value:
								if j == 0:
									j=1
								else:
									if int(fila[0].value)==int(fila2[0].value):
										count = count+1

						
						actividad_porce=CEsquemaCapitulosActividadesG.objects.get(pk=fila[0].value)
						porcentaje=float(actividad_porce.peso)/count

						detalle=FDetallePresupuesto(porcentaje=porcentaje,presupuesto_id=presupuesto_id,actividad_id=fila[0].value,codigoUC=fila[3].value,descripcionUC=fila[4].value,valorManoObra=fila[6].value,valorMaterial=fila[7].value,valorGlobal=fila[8].value,cantidad=fila[5].value)
						detalle.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=detalle.id)
						logs_model.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.detalle_presupuesto')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def actualizar_cantidad(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.actualizar_cantidad')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def cierre_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()


		presupuesto=EPresupuesto.objects.get(pk=respuesta['id_presupuesto'])
		presupuesto.cerrar_presupuesto=True
		presupuesto.save()	
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.presupuesto',id_manipulado=respuesta['id_presupuesto'])
		logs_model.save()	

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico.cierre_presupuesto')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de un diagrama
@transaction.atomic
def eliminar_diagrama(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				DiagramaGrahm.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.diagrama',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.diagrama')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un diagrama	


@login_required
def informe_diagrama(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_diagrama.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Diagrama Grannt')
	format1=workbook.add_format({'font_size':10,'bold':True})

	

	cronograma_id= request.GET['cronograma_id']	
	
	cronograma=Cronograma.objects.get(pk=cronograma_id)				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=cronograma.esquema.id,nivel=1)

	worksheet.set_column('A:C', 25)

	cell_format = workbook.add_format()

	cell_format.set_pattern(1)  # This is optional when using a solid fill.
	cell_format.set_bg_color('green')

	#worksheet.write('A1', 'Id', format1)

	fecha_menor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id).order_by('fechaInicio').first()	

	fecha_mayor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id).order_by('fechaFinal').last()	

	date_nueva=fecha_mayor.fechaFinal-fecha_menor.fechaInicio
	semanas=math.ceil(date_nueva.days/7.0)

	worksheet.write(1, 0,'Capitulos / Actividades' ,format1)
	worksheet.write(1, 1, 'Fecha Inicio',format1)
	worksheet.write(1, 2, 'Fecha Final',format1)

	row=1
	col=3
	for i in range(int(semanas)):
		worksheet.write(row, col, "Semana "+str(i+1),format1)
		col +=1

	row=2
	col=0

	for item in hitos:

		actividades=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id)

		if len(actividades)>0:
			worksheet.write(row, col, item.nombre,format1)

			fechacapitulo_menor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id).order_by('fechaInicio').first()
			fechacapitulo_mayor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id).order_by('fechaFinal').last()

			semana_inicio=0
			semana_final=0
			fecha_semana=fecha_menor.fechaInicio
			sw1=0
			sw2=0
			sw3=0
			date_nueva2=fechacapitulo_mayor.fechaFinal-fechacapitulo_menor.fechaInicio
			dias=math.floor(date_nueva2.days/7.0)

			for i in range(int(semanas)):

				if fechacapitulo_menor.fechaInicio == fechacapitulo_mayor.fechaFinal and sw2==0:
					semana_inicio=i+1
					semana_final=i+2
					sw2=1
					sw1=1
					sw3=1

				if fechacapitulo_menor.fechaInicio >=  fecha_semana and sw1==0:
					semana_inicio=i+1

					# if item2.fechaFinal <=  fecha_semana and sw2==0:
					# 	semana_final=i+1
					# 	sw2=1

				fecha_semana=fecha_semana + timedelta(days=7)

			if sw3==0:
				semana_final=semana_inicio+int(dias)+1

				
			for i in range(semana_inicio,semana_final):
				colu=col+2+semana_inicio
				worksheet.write(row, colu, "",cell_format)
				semana_inicio=semana_inicio+1


			row +=1		

			for item2 in actividades:

				worksheet.write(row, col, item2.actividad.nombre)
				worksheet.write(row, col+1, str(item2.fechaInicio))
				worksheet.write(row, col+2, str(item2.fechaFinal))

				semana_inicio=0
				semana_final=0
				fecha_semana=fecha_menor.fechaInicio
				sw1=0
				sw2=0
				sw3=0

				date_nueva2=item2.fechaFinal-item2.fechaInicio
				dias=math.floor(date_nueva2.days/7.0)

				for i in range(int(semanas)):

					if fechacapitulo_menor.fechaInicio == fechacapitulo_mayor.fechaFinal and sw2==0:
						semana_inicio=i+1
						semana_final=i+2
						sw2=1
						sw1=1
						sw3=1

					if item2.fechaInicio >=  fecha_semana and sw1==0:
						semana_inicio=i+1

					# if item2.fechaFinal <=  fecha_semana and sw2==0:
					# 	semana_final=i+1
					# 	sw2=1

					fecha_semana=fecha_semana + timedelta(days=7)

				if sw3==0:
					semana_final=semana_inicio+int(dias)+1

								
				for i in range(semana_inicio,semana_final):
					colu=col+2+semana_inicio
					worksheet.write(row, colu, "",cell_format)
					semana_inicio=semana_inicio+1


				row +=1


		

	workbook.close()

	return response
    #return response


@login_required
def descargar_plantilla_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', 'Apoyo', format1)
	worksheet.write('B1', 'Longitud', format1)
	worksheet.write('C1', 'Latitud', format1)


	workbook.close()

	return response
    #return response



@login_required
def guardar_apoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=fila[1].value,latitud=fila[2].value,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
						nodo.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=nodo.id)
						logs_model.save()
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de varios apoyos
@transaction.atomic
def eliminar_apoyos(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				HNodo.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de varios apoyos



@login_required
def descargar_plantilla_apoyo_sinposicion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', '*Los campos latitud y longitud no son campos obligatorio*', format1)
	worksheet.write('A2', 'Apoyo', format1)
	worksheet.write('B2', 'Latitud', format1)
	worksheet.write('C2', 'Longitud', format1)

	workbook.close()

	return response
    #return response


	
@login_required
@transaction.atomic
def guardar_apoyo_archivo_sinposicion(request):

	sid = transaction.savepoint()
	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0 or i==1:
						i=i+1
					else:
						veri_nodo=HNodo.objects.filter(presupuesto_id=presupuesto_id,nombre=fila[0].value)

						if len(veri_nodo) == 0:
							latitud=None
							longitud=None

							if fila[1].value is not None and  fila[2].value is not None:
								latitud=fila[1].value
								longitud=fila[2].value

							nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=longitud,latitud=latitud,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
							nodo.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=nodo.id)
							logs_model.save()
						else:
							transaction.savepoint_rollback(sid)
							return JsonResponse({'message':'No se puede repetir el nombre del nodo en este presupuesto','success':'error',
							'data':''})

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def descargar_plantilla_cantidadApoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_cantidadApoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cantidad_Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:Z', 20)


	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)

	apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	row=0
	col=6

	for item_apoyo in apoyos:
		worksheet.write(row,col, item_apoyo.nombre, format1)

		col +=1


	row=1
	col=0

	hitos = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


	for item in hitos:
		worksheet.write(row, col, item.id,format2)
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)
		worksheet.write(row, col+1,str(capitulo.nombre),format2)
		worksheet.write(row, col+2,item.actividad.nombre,format2)
		worksheet.write(row, col+3,item.codigoUC,format2)	
		worksheet.write(row, col+4,item.descripcionUC,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		count=6
		for item_apoyo in apoyos:
			worksheet.write(row,col+count,0, format2)
			count +=1

		row +=1


	workbook.close()

	return response
    #return response



@login_required
def guardar_cantidadApoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')

						count=6
						for item_apoyo  in apoyos:
							consultar_nodo=JCantidadesNodo.objects.filter(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id)
							if len(consultar_nodo)==0:
								nodo=JCantidadesNodo(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id,cantidad=fila[count].value)
								nodo.save()
								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=nodo.id)
								logs_model.save()
							else:
								nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
								nodo.cantidad=fila[count].value
								nodo.save()
								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
								logs_model.save()

							count +=1
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_nodo')
		#transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def informe_cantidad_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_cantidad_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Informe')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	presupuesto_id= request.GET['presupuesto_id']	
				
	detalle_presupuesto = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 30)

	worksheet.write('A1', 'Hitos', format1)
	worksheet.write('B1', 'Actividad', format1)
	worksheet.write('C1', 'Cod. UUCC', format1)
	worksheet.write('D1', 'Descripcion UUCC', format1)
	worksheet.write('E1', 'Cantidad Total', format1)
	worksheet.write('F1', 'Cantidad en Apoyos', format1)

	for item in detalle_presupuesto:

		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col, str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)
		worksheet.write(row, col+3,item.descripcionUC,format2)
		worksheet.write(row, col+4,item.cantidad,format2)

		porcentaje=JCantidadesNodo.objects.filter(detallepresupuesto_id=item.id).aggregate(Sum('cantidad'))
		porcentaje['cantidad__sum']=0 if porcentaje['cantidad__sum']==None else porcentaje['cantidad__sum']		

		worksheet.write(row, col+5,str(porcentaje['cantidad__sum']),format2)

		row +=1


	workbook.close()

	return response
    #return response


@transaction.atomic
def guardar_cantidad_apoyo(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			nodo=JCantidadesNodo.objects.get(pk=item['id'])
			nodo.cantidad=item['cantidad']
			nodo.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cantidad_apoyo',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@login_required
@transaction.atomic
def consultar_avance_obra(request):
	sid = transaction.savepoint()
	try:
		id_cronograma=request.GET['cronograma_id']
		presupuesto_id=request.GET['presupuesto_id']
		programando=request.GET['programando']
		dato=request.GET['dato']
		# estados=request.GET['id_estados']
		# listado_estado=estados.split(',')
		# ejecucion=request.GET['ejecucion']


		# listado_id=[]
		# cambio=LCambio.objects.filter(cronograma_id=id_cronograma,estado_id__in=listado_estado)

		# listado_id_ejecucion=[]
		# if int(ejecucion)>0:

		# 	listado_por=[]
		# 	if int(ejecucion) == 2:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__gte=100)
				
		# 	else:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__lt=100)

		# 	for obj in listado_por:
		# 		listado_id_ejecucion.append(obj.apoyo.id)

		# if cambio is not None:
		# 	for obj in cambio:
		# 		for item in obj.nodos.all():
		# 			listado_id.append(item.id)

		
		# cronograma=KCronograma.objects.get(pk=id_cronograma)
		
		

		# if len(listado_id_ejecucion)==0:
		# 	qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

		# if len(listado_id_ejecucion)>0:
		# 	qset=qset&(Q(id__in=listado_id_ejecucion))
		qset=(Q(presupuesto_id=presupuesto_id)&(Q(eliminado=False)))


		if dato!='':
			qset=qset&(Q(nombre__icontains=dato))

		listado=HNodo.objects.filter(qset).values('id','nombre','longitud','latitud','capa__color')

		qset1=Q(nodoOrigen__presupuesto__id=presupuesto_id)

		# if len(listado_id)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id))

		# if len(listado_id_ejecucion)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id_ejecucion))

		listado_enlaces=IEnlace.objects.filter(qset1).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'datos':list(listado),'enlace':list(listado_enlaces)}})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def consultar_ingresos_datos(request):
	sid = transaction.savepoint()
	try:
		nodo_id=request.GET['nodo_id']
		reporte_id=request.GET['reporte_id']

		nodo=HNodo.objects.get(pk=nodo_id)
		datos=[]

		if nodo.noProgramado == False:
			listado=[]
			print (nodo.presupuesto.sin_poste)
			if nodo.presupuesto.sin_poste == False:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id,cantidad__gt=0)
			else:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id)

			for item in listado:
				cantidad_ejecutada=KDetalleReporteTrabajo.objects.filter(nodo_id=nodo_id,detallepresupuesto_id=item.detallepresupuesto.id).aggregate(Sum('cantidadEjecutada'))
				cantidad_ejecutada['cantidadEjecutada__sum']=0 if cantidad_ejecutada['cantidadEjecutada__sum']==None else cantidad_ejecutada['cantidadEjecutada__sum']
				datos.append({
							'id_detalle':item.detallepresupuesto.id,
							'codigo':item.detallepresupuesto.codigoUC,
							'descripcion':item.detallepresupuesto.descripcionUC,
							'cantidad':item.cantidad,
							'cantidad_ejecutada':cantidad_ejecutada['cantidadEjecutada__sum']
					})

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':datos})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_cambio_cantidades(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			cambio=KDetalleReporteTrabajo(reporte_trabajo_id=respuesta['id_reporte'],nodo_id=respuesta['id_nodo'],detallepresupuesto_id=item['id_detalle'],cantidadEjecutada=item['cantidad'])
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.detalle_reporte',id_manipulado=cambio.id)
			logs_model.save()



		reporte=JReporteTrabajo.objects.get(pk=respuesta['id_reporte'])

		if reporte.presupuesto.sin_poste==False:
			detalle=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=respuesta['id_reporte'],nodo_id=respuesta['id_nodo']).values('detallepresupuesto_id').annotate(cantidades=Sum('cantidadEjecutada')).distinct()

			total=0
			por=0
			for item2 in detalle:
				cantidad_nodo=JCantidadesNodo.objects.filter(detallepresupuesto_id=item2['detallepresupuesto_id'],nodo_id=respuesta['id_nodo'])
				valor=float(item2['cantidades'])/float(cantidad_nodo[0].cantidad)
				por=por+valor

			total=(por/len(detalle))*100

			if total>100:
				total=100

			porcentaje_apoyo=HNodo.objects.get(pk=respuesta['id_nodo'])
			porcentaje_apoyo.porcentajeAcumulado=total
			porcentaje_apoyo.save()

		#createAsyncEstado.delay(respuesta['id_cronograma'])

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cambio_ejecutada')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_cambio_detalle(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)


		for item in respuesta['lista']:
			detalle=KDetalleReporteTrabajo.objects.get(pk=item['id'])
			detalle.cantidadEjecutada=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cambio_detalle',id_manipulado=detalle.id)
			logs_model.save()



		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido registrado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_id_nodo_destino(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			IEnlace.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.enlance',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@login_required
@transaction.atomic
def cierre_programacion(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cronograma=Cronograma.objects.get(id=respuesta['id_cronograma'])
		cronograma.programacionCerrada=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cronograma',id_manipulado=respuesta['id_cronograma'])
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def informe_detallepresupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']
	presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:H', 20)


	worksheet.write('A2', 'Presupuesto', format1)
	worksheet.write('B2', presupuesto.nombre, format1)

	worksheet.write('A3', 'Proyecto', format1)
	worksheet.write('B3', presupuesto.cronograma.proyecto.nombre, format1)

	worksheet.write('A4', 'Esquema', format1)
	worksheet.write('B4', presupuesto.cronograma.esquema.nombre, format1)

	worksheet.write('F6', 'Total Presupuesto', format1)

	worksheet.write('A7', 'Hitos', format1)
	worksheet.write('B7', 'Actividad', format1)
	worksheet.write('C7', 'codigo UUCC', format1)
	worksheet.write('D7', 'Descripcion UUCC', format1)
	worksheet.write('E7', 'Valor UUCC', format1)
	worksheet.write('F7', 'Cantidad', format1)
	worksheet.write('G7', 'Subtotal', format1)

	row=7
	col=0

	detalle = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)
	total=0

	for item in detalle:
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col,str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)	
		worksheet.write(row, col+3,item.descripcionUC,format2)

		valorManoObra=0
		if item.valorManoObra is not None:
			valorManoObra=item.valorManoObra

		valorMaterial=0
		if item.valorMaterial is not None:
			valorMaterial=item.valorMaterial

		valor=int(valorManoObra)+int(valorMaterial)

		worksheet.write(row, col+4,valor,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		subtotal=valor*int(item.cantidad)

		worksheet.write(row, col+6,subtotal,format2)

		total=total+subtotal

		row +=1


	worksheet.write('G6', total, format1)

	workbook.close()

	return response
    #return response


@login_required
def menu_gps(request,id_presupuesto):

	try:
		nodo=HNodo.objects.filter(presupuesto_id=id_presupuesto)
		
		if len(nodo) == 0:	
			return JsonResponse({'message':'No tiene ningun punto de gps registrado','success':'ok',
					'data':['1']})
		else:
			sqlnodo=HNodo.objects.filter(presupuesto_id=id_presupuesto,latitud=None,longitud=None)

			if len(sqlnodo) == 0:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['2']})
			else:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['3']})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def guardar_reporte_trabajo(request,id_reporte):

	sid = transaction.savepoint()
	try:
		reporte=JReporteTrabajo.objects.get(pk=id_reporte)
		#detalle=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)
		detalles=FDetallePresupuesto.objects.filter(presupuesto_id=reporte.presupuesto.id)

		# model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
		# valor=round(float(model_actividad['peso__sum']),3)+round(float(request.DATA['peso']),3)
		#import pdb; pdb.set_trace()
		detalle_reporte=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)

		if len(detalle_reporte) == 0:
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Debe ingresar alguna cantidad ejecutada en el reporte','success':'error',
			'data':''})

		total_porcentaje=0
		total_pagado=0
		for item in detalles:
			sql_detalletrabajo=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte,detallepresupuesto_id=item.id).aggregate(Sum('cantidadEjecutada'))
			sql_detalletrabajo['cantidadEjecutada__sum']=0 if sql_detalletrabajo['cantidadEjecutada__sum']==None else sql_detalletrabajo['cantidadEjecutada__sum']
			#valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])+float(item.cantidad)
			valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])/float(item.cantidad)
			item.porcentaje=0 if item.porcentaje==None else item.porcentaje
			
			#porcentaje=(valor*float(item.porcentaje))/100
			porcentaje=(valor*float(item.porcentaje))
			total_porcentaje=total_porcentaje+porcentaje

			valor2=float(item.valorManoObra)+float(item.valorMaterial)
			pagado=float(sql_detalletrabajo['cantidadEjecutada__sum'])*float(valor2)
			total_pagado=total_pagado+pagado

		if total_porcentaje>100:
			total_porcentaje=100

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		estado_rechazo=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		estado_corregido=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)

		reporte.valor_ganando_acumulado=round(total_pagado,3)
		reporte.avance_obra_acumulado=round(total_porcentaje,3)
		reporte.reporteCerrado=True

		if int(reporte.estado_id)==int(estado_rechazo[0].id):			
			reporte.estado_id=estado_corregido[0].id
		else:
			reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.reporte',id_manipulado=id_reporte)
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#serializer de empresa
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')



#serializer lite de proyecto empresa
class ProyectoEmpresaLite2Serializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLite2Serializer(read_only = True )
	empresa = EmpresaLiteSerializer(read_only = True  )

	totalCorrregidos=serializers.SerializerMethodField()

	totalAprobados=serializers.SerializerMethodField()

	totalRegistrados=serializers.SerializerMethodField()

	totalRechazados=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id' ,
				'proyecto',
 				'empresa',
 				'totalCorrregidos',	
 				'totalAprobados',	
 				'totalRegistrados',
 				'totalRechazados'
				)

	def get_totalCorrregidos(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalAprobados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRegistrados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRechazados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()


#Api lite de empresa proyecto
class ProyectoReporteAvanceViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las categorias de administrador de fotos
	"""
	model=Proyecto_empresas
	queryset = model.objects.all()
	serializer_class = ProyectoEmpresaLite2Serializer
	nombre_modulo='administrador_foto.Proyecto_empresa'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ProyectoReporteAvanceViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			empresa = self.request.query_params.get('empresa', None)
			mcontrato = self.request.query_params.get('mcontrato', None)
			departamento = self.request.query_params.get('departamento', None)
			municipio = self.request.query_params.get('municipio', None)
			contratista = self.request.query_params.get('contratista', None)
			corregido = self.request.query_params.get('corregido', None)
			registrado = self.request.query_params.get('registrado', None)
			aprobado = self.request.query_params.get('aprobado', None)
			rechazados = self.request.query_params.get('rechazados', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			qset=(~Q(id=0))

			if corregido is not None:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			if rechazados is not None:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			
			if registrado is not None and int(registrado)==1:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))

			if aprobado is not None and int(aprobado)==1:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))


			if(dato or empresa or mcontrato or departamento or municipio):

				if dato:
					qset = qset & ( Q(empresa__nombre__icontains = dato) |
								Q(empresa__nit__icontains = dato) |
								Q(proyecto__nombre__icontains = dato))					
			

				if empresa and int(empresa)>0:
					qset = qset & (Q(empresa__id = empresa))					


				if mcontrato and int(mcontrato)>0:
					qset = qset & (Q(proyecto__mcontrato = mcontrato))
					
				if departamento and int(departamento)>0:
					qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
					
				if municipio and int(municipio)>0:
					qset = qset & (Q(proyecto__municipio = municipio))
					

				if contratista and int(contratista)>0:
					qset = qset & (Q(proyecto__contrato__contratista__id = contratista))	

			
			if qset != '':
				queryset = self.model.objects.filter(qset).distinct()

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@login_required
@transaction.atomic
def guardar_archivo_aprobacion(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		reporte_id= request.POST['reporte_id']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		
		reporte=JReporteTrabajo.objects.get(pk=reporte_id)
		reporte.soporteAprobacion=soporte
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.soporte_aprobacion',id_manipulado=reporte_id)
		logs_model.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.soporte_aprobacion')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_rechazo_reporte(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		reporte_id= respuesta['reporte_id']
		mensaje= respuesta['mensaje']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		
		reporte=JReporteTrabajo.objects.get(pk=reporte_id)
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.reporteCerrado=False
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.rechazo_reporte',id_manipulado=reporte_id)
		logs_model.save()

		motivo=MComentarioRechazo(reporte_trabajo_id=reporte_id,motivoRechazo=mensaje)
		motivo.save()
		logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.motivo_rechazo',id_manipulado=motivo.id)
		logs_model2.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.rechazo_reporte')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def consultar_graficos(request):
	sid = transaction.savepoint()
	try:
		presupuesto_id=request.GET['presupuesto_id']
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		listado=[]
		listado_porcentaje=DiagramaGrahm.objects.filter(cronograma_id=presupuesto.cronograma.id).values('fechaFinal').annotate(porcentaje=Sum('actividad__peso')).distinct()
		print(listado_porcentaje)
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado.append({
					'fecha':item['fechaFinal'],
					'porcentaje':round(porcentaje,2)
					})

		print(listado)


		listado3=[]
		listado_porcentaje=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(porcentaje=Sum('avance_obra_acumulado')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado3.append({
					'fecha':item['fechaTrabajo'],
					'porcentaje':round(porcentaje,2)
					})

		listado4=[]
		listado_prespuesto=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(valor_ganando=Sum('valor_ganando_acumulado')).distinct()
				
		valor_ganando=0
		for item in listado_prespuesto:
			valor_ganando=valor_ganando+float(item['valor_ganando'])
			
			listado4.append({
					'fecha':item['fechaTrabajo'],
					'valor_ganando':round(valor_ganando,2)
					})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'programada':list(listado),'avance':list(listado3),'presupuesto':list(listado4)}})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.grafico')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def consultar_cantidad_reportes(request):

	try:

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		cantidad_corregidos=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		cantidad_registrado=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		cantidad_rechazados=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()
		

		return JsonResponse({'message':'','success':'ok',
					'data':{'cantidad_corregidos':cantidad_corregidos,'cantidad_registrado':cantidad_registrado,'cantidad_rechazados':cantidad_rechazados}})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_reportes')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def actualizar_cancelado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
		for item in respuesta['lista']:
			cambio=LCambio.objects.get(id=item['id'])
			cambio.estado_id=estado[0].id
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cambio',id_manipulado=item['id'])
			logs_model.save()

			historial=LHistorialCambio(cambio_id=item['id'],usuario_registro_id=request.user.usuario.id,motivoEstado=respuesta['motivo'],estado_id=estado[0].id)
			historial.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.historial_cambio',id_manipulado=historial.id)
			logs_model2.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.actualizar_cantidad')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
@transaction.atomic
def reportar_sin_poste(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		presupuesto_id= respuesta['presupuesto_id']
		
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		presupuesto.sin_poste=True
		presupuesto.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.sin_poste',id_manipulado=presupuesto_id)
		logs_model.save()

		# agregarSinPoste.delay(presupuesto_id,request.user.usuario.id)
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'avance_de_obra_grafico2.sin_poste')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def avance_de_obra_grafico2(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico2/hitos.html',{'model':'esquemacapitulosgrafico','app':'avanceObraGrafico2','macrocontrato':ListMacro})


@login_required
def actividades(request,id_esquema):
	return render(request, 'avanceObraGrafico2/actividad.html',{'model':'esquemacapitulosactividadesgrafico','app':'avanceObraGrafico2','id_esquema':id_esquema})


@login_required
def regla_estado(request,id_esquema):
	return render(request, 'avanceObraGrafico2/regla.html',{'model':'EPresupuesto','app':'avanceObraGrafico','id_esquema':id_esquema})


@login_required
def cronograma(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico2/cronograma.html',{'model':'cronograma','app':'avanceObraGrafico2','macrocontrato':ListMacro})



@login_required
def cronograma_proyecto(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	esquema=BEsquemaCapitulosG.objects.filter(macrocontrato_id=proyecto.mcontrato)
	periodo=APeriodicidadG.objects.all()
	return render(request, 'avanceObraGrafico2/cronograma_proyecto.html',{'model':'cronograma','app':'avanceObraGrafico2','id_proyecto':id_proyecto,'proyecto':proyecto,'esquema':esquema,'periodo':periodo})


@login_required
def presupuesto(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico2/presupuesto.html',{'model':'EPresupuesto','app':'avanceObraGrafico2','cronograma':querycronograma,'cronograma_id':id_cronograma,'proyecto_id':querycronograma.proyecto.id})


@login_required
def presupuesto_detalle(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=presupuesto.cronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico2/detalle_presupuesto.html',{'model':'EPresupuesto','app':'avanceObraGrafico','presupuesto_id':id_presupuesto,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'capitulos':capitulos,'cronograma_id':presupuesto.cronograma.id})


@login_required
def programacion(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=querycronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico2/diagrama.html',{'model':'diagramaGrahm','app':'avanceObraGrafico2','capitulos':capitulos,'cronograma':querycronograma,'cronograma_id':id_cronograma,'proyecto_id':querycronograma.proyecto.id})



@login_required
def apoyo_con_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/apoyo_con_gps.html',{'model':'nodo','app':'avanceObraGrafico2','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def apoyo_sin_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/apoyo_sin_gps.html',{'model':'nodo','app':'avanceObraGrafico2','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def cantidad_apoyo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.cronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico2/cantidad_apoyo.html',{'model':'EPresupuesto','app':'avanceObraGrafico2','cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'capitulos':capitulos})


@login_required
def cantidad_apoyo_id(request,id_presupuesto,id_detalle):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	detalle_presupuesto=FDetallePresupuesto.objects.get(pk=id_detalle)
	return render(request, 'avanceObraGrafico2/cantidad_apoyo_id.html',{'model':'EPresupuesto','app':'avanceObraGrafico2','presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'detalle_presupuesto':detalle_presupuesto,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'cronograma_id':nombre_presupuesto.cronograma.id})


@login_required
def reporte_trabajo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
	estados2=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)

	empresa_id=0
	if nombre_presupuesto is not None:
		for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
			if int(valor.tipo_contrato.id)==9:
				empresa_id=valor.contratista.id	
				
	usuarios=Usuario.objects.filter(empresa_id=empresa_id,user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraGrafico2/reporte.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2','Usuarios':usuarios,'estado_id_procesado':estados2[0].id,'estado_id_registrado':estados[0].id,'cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto})


@login_required
def avance_con_gps(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	
	return render(request, 'avanceObraGrafico2/avance_con_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte})


@login_required
def avance_sin_gps(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	estados_cambios=Estado.objects.filter(app='Avance_obra_grafico')
	return render(request, 'avanceObraGrafico2/avance_sin_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','estados':estados_cambios,'capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte})



@login_required
def aprobacion(request):
	return render(request, 'avanceObraGrafico2/aprobacion.html',{'model':'reporte','app':'avanceObraGrafico2'})


@login_required
def corregido(request):
	return render(request, 'avanceObraGrafico2/corregido.html',{'model':'reporte','app':'avanceObraGrafico2'})


@login_required
def registrado(request):
	return render(request, 'avanceObraGrafico2/registrado.html',{'model':'reporte','app':'avanceObraGrafico2'})


@login_required
def rechazados(request):
	return render(request, 'avanceObraGrafico2/rechazados.html',{'model':'reporte','app':'avanceObraGrafico2'})


@login_required
def reporte_trabajo_registrado(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_registrado.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_registrado(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraGrafico2/detalle_registrado.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def reporte_trabajo_corregido(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_corregido.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_corregido(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraGrafico2/detalle_corregido.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def grafico(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/grafico.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'cronograma_id':presupuesto.cronograma.id,'presupuesto':presupuesto})


@login_required
def reporte_trabajo_rechazados(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_rechazados.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def cambios(request):
	return render(request, 'avanceObraGrafico2/cambios.html',{'model':'cambio','app':'avanceObraGrafico2'})


@login_required
def index_cambio(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	empresa=Empresa.objects.filter(~Q(id=request.user.usuario.empresa.id))
	estados=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=101)
	cancelado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
	return render(request, 'avanceObraGrafico2/index_cambio.html',{'model':'cambio','app':'avanceObraGrafico2','estado_cancelado':cancelado[0].id,'estado_id':estados[0].id,'presupuesto_id':id_presupuesto,'presupuesto':presupuesto,'empresa':empresa})


@login_required
def detalle_cambio(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	return render(request, 'avanceObraGrafico2/detalle_cambio.html',{'model':'cambio','app':'avanceObraGrafico2','cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def agregar_detalle(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	apoyos=HNodo.objects.filter(presupuesto_id=cambio.presupuesto.id)
	return render(request, 'avanceObraGrafico2/agregar_detalle.html',{'model':'cambio','app':'avanceObraGrafico2','apoyos':apoyos,'cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def aprobacion_cambio(request):
	return render(request, 'avanceObraGrafico2/aprobacion_cambio.html',{'model':'cambio','app':'avanceObraGrafico2'})


@login_required
def autorizacion_cambio(request):
	return render(request, 'avanceObraGrafico2/autorizacion_cambio.html',{'model':'cambio','app':'avanceObraGrafico2'})
