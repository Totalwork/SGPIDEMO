from django.shortcuts import render,redirect
from .models import APeriodicidadG,BEsquemaCapitulosG,CEsquemaCapitulosActividadesG,CReglasEstadoG,Cronograma,EPresupuesto,FDetallePresupuesto
from .models import DiagramaGrahm,HNodo,GCapa,IEnlace,JCantidadesNodo,JReporteTrabajo,LCambio,KDetalleReporteTrabajo,MComentarioRechazo,LHistorialCambio
from .models import LDetalleCambio, FotoNodo, DesgloceManoDeObra, EReformado
from .models import UnidadConstructiva, DesgloceManoDeObra, DesgloceMaterial, EReformadoDetalle, ManoDeObra
from rest_framework import viewsets, serializers
from django.db.models import Q,Sum,Prefetch
from logs.models import Logs,Acciones
from usuario.models import Usuario
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import xlsxwriter
import json
from datetime import date, datetime
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.deletion import ProtectedError

from proyecto.views import ProyectoLiteSerializer
from proyecto.models import Proyecto,Proyecto_empresas

from contrato.models import EmpresaContrato,Contrato
from contrato.views import ContratoSerializer
from contrato.enumeration import tipoC

from .tasks import updateAsyncEstado,agregarSinPoste

from tipo.models import Tipo

from empresa.models import Empresa

from parametrizacion.models import Departamento

from django.db import connection

from datetime import timedelta
import time

import math

from usuario.views import UsuarioSerializer
from empresa.views import EmpresaSerializer

from estado.models import Estado
from estado.views import EstadoSerializer

from tipo.views import TipoSerializer
from tipo.models import Tipo

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view

from django.conf import settings
from sinin4.functions import functions
from django.db import transaction
from parametrizacion.views import  MunicipioSerializer , DepartamentoSerializer
from puntos_gps.views import PuntosGps
from control_cambios.views import CCambio
from p_p_construccion.models import ALote
from django.db.models import Q, Sum, Max
from avance_de_obra.models import BCronograma
from .models import MLiquidacionUUCC, CatalogoUnidadConstructiva, Material
import locale
# locale.setlocale( locale.LC_ALL, 'en_US' )

import openpyxl
#from utilidades.funciones import funciones
# Create your views here.

#serializer de empresa
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista')


#serializer lite de proyecto
class ProyectoLite2Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	municipio = MunicipioSerializer(read_only = True)
	contrato = ContratoLiteSerializer(read_only = True, many = True)
	conteo_puntos=serializers.SerializerMethodField()
	cantidad_cambio=serializers.SerializerMethodField()
	cantidad_lote=serializers.SerializerMethodField()

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'mcontrato' ,
				 'municipio' ,
				 'contrato',
				 'conteo_puntos',
 				 'cantidad_cambio',
				 'cantidad_lote'
				 )

	def get_conteo_puntos(self, obj):
		return PuntosGps.objects.filter(proyecto_id=obj.id).count()

	def get_cantidad_cambio(self, obj):
		return CCambio.objects.filter(proyecto_id=obj.id).count()

	def get_cantidad_lote(self, obj):
		return ALote.objects.filter(proyecto_id=obj.id).count()
	
#serializer lite de proyecto empresa
class ProyectoEmpresaLite4Serializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLite2Serializer(read_only = True )
	empresa = EmpresaLiteSerializer(read_only = True  )

	totalCronograma=serializers.SerializerMethodField()

	totalPresupuesto=serializers.SerializerMethodField()

	totalCronogramaGrafico=serializers.SerializerMethodField()

	porcentajeAvance=serializers.SerializerMethodField()

	porcentajeAvanceFinanciero=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id' ,
				'proyecto',
 				'empresa',
 				'totalCronograma',
 				'totalPresupuesto',
 				'totalCronogramaGrafico',		 
 				'porcentajeAvance',
 				'porcentajeAvanceFinanciero'
				)

	def get_totalCronograma(self, obj):
		return BCronograma.objects.filter(proyecto_id=obj.proyecto.id).count()

	def get_totalCronogramaGrafico(self, obj):
		return Cronograma.objects.filter(proyecto_id=obj.proyecto.id).count()

	def get_totalPresupuesto(self, obj):
		#from avanceObraGrafico.models import EPresupuesto
		return EPresupuesto.objects.filter(cronograma__proyecto_id=obj.proyecto.id).count()

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []
		avanzo = False
		corte = 'Sin avance'
		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['detallepresupuesto__actividad__peso']
				})
		##import pdb; pdb.set_trace()
		actividades = porcentajes
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		# obtenemos la fecha corte a la cual le restaremos la periodicidad para determinar
		# si hubo avance en el proyecto	
		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__presupuesto__cronograma__id').annotate(
			maxima=Max('reporte_trabajo__fechaTrabajo'))

		if queryset_ejecutada:	
			corte = queryset_ejecutada[0]['maxima']

		# Determinamos si aumento o no aumento el avance tomando como referencia del calculo
		# la periodicidad del cronograma.
		avance_anterior=0
		porcentajes = []
		##import pdb; pdb.set_trace()
		if corte != 'Sin avance':
			#fecha = datetime.strptime(corte, '%Y-%m-%d').date()
			#fecha = corte
			fecha = datetime.today().date()
			cronograma = Cronograma.objects.filter(
				proyecto__id = obj.proyecto.id).values('periodicidad__numero_dias')
			if cronograma:
				fecha = fecha - timedelta(days=int(cronograma[0]['periodicidad__numero_dias']))
				corte1 = fecha.strftime('%Y-%m-%d')
				queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
					detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
					cantidadEjecutada__gt=0,
					reporte_trabajo__fechaTrabajo__lte=corte1).values(
					'detallepresupuesto__actividad__id',
					'detallepresupuesto__actividad__nombre'
					).annotate(total=Sum('cantidadEjecutada'))

				for Aejecutar in queryset_a_ejecutar:
					agregado = False
					for ejecutado in queryset_ejecutada:
						if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
							agregado = True
							porcentajes.append({
								'id' : Aejecutar['detallepresupuesto__actividad__id'],
								'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
								'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
								'peso': Aejecutar['detallepresupuesto__actividad__peso']
							})
						if agregado:
							break
					if agregado == False:
						porcentajes.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': 0,
							'peso': Aejecutar['detallepresupuesto__actividad__peso']
						})
				##import pdb; pdb.set_trace()
				for p in porcentajes:
					avance_anterior = avance_anterior + (float(p['porcentaje']) * float(float(p['peso'])/100))

		if avance_anterior < avance:
			avanzo = True	

		return {'avance': round(avance,2), 'avanzo' : avanzo, 'corte':corte, 'actividades':actividades}

	def get_porcentajeAvanceFinanciero(self,obj):
		totalPresupuesto = 0
		totalEjecutado = 0
		avance = 0
		cantidadesEjecutar = FDetallePresupuesto.objects.filter(
			presupuesto__cronograma__proyecto__id=obj.proyecto.id).values(
			'cantidad','codigoUC','cantidad',
			'catalogoUnidadConstructiva__id', 'id')

		for rowCantidadesEjecutar in cantidadesEjecutar:
			totalMo = 0
			totalMat = 0
			queryset = DesgloceManoDeObra.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']
				).values(
				'manoDeObra__valorHora','rendimiento')
			
			if queryset:
				for row in queryset:
					totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

			queryset = DesgloceMaterial.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']).values(
				'material__valorUnitario','cantidad')
			
			if queryset:
				for row in queryset:
					totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))

			totalPresupuesto = totalPresupuesto + \
			((totalMo * float(rowCantidadesEjecutar['cantidad'])) + \
			(totalMat * float(rowCantidadesEjecutar['cantidad'])))

			cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__id=rowCantidadesEjecutar['id'],
			cantidadEjecutada__gte=0).values(
			'detallepresupuesto__id').annotate(total=Sum('cantidadEjecutada'))

			if cantidadesEjecutadas:
				totalEjecutado = totalEjecutado + \
				(float(cantidadesEjecutadas[0]['total']) * totalMo) + \
				(float(cantidadesEjecutadas[0]['total']) * totalMat)

		if totalPresupuesto > 0:		
			avance = (totalEjecutado / totalPresupuesto) * 100
		else:
			avance = 0

		return round(avance,2)


class ProyectoEmpresaLite3Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = serializers.SerializerMethodField('_mcontrato',read_only=True)
	porcentajeAvance=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('mcontrato',		 
 				'porcentajeAvance',
				)

	def _mcontrato(self,obj):
		return  {'id': obj.proyecto.mcontrato.id,'nombre':obj.proyecto.mcontrato.nombre}

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []
		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['detallepresupuesto__actividad__peso']
				})
		##import pdb; pdb.set_trace()
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		# obtenemos la fecha corte a la cual le restaremos la periodicidad para determinar
		# si hubo avance en el proyecto	
		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__presupuesto__cronograma__id').annotate(
			maxima=Max('reporte_trabajo__fechaTrabajo'))

		return {'avance': round(avance,2)}


#Api lite de empresa proyecto
class ProyectoEmpresaLiteViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las categorias de administrador de fotos
	"""
	model=Proyecto_empresas
	queryset = model.objects.all()
	serializer_class = ProyectoEmpresaLite4Serializer
	nombre_modulo='avanceObraGrafico2.Proyecto_empresa'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ProyectoEmpresaLiteViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			empresa = self.request.query_params.get('empresa', None)
			mcontrato = self.request.query_params.get('mcontrato', None)
			departamento = self.request.query_params.get('departamento', None)
			municipio = self.request.query_params.get('municipio', None)
			contratista = self.request.query_params.get('contratista', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			homepage = self.request.query_params.get('homepage', None)
			qset=(~Q(id=0))

			if(dato or empresa or mcontrato or departamento or municipio):

				if dato:
					qset = qset & ( Q(empresa__nombre__icontains = dato) |
								Q(empresa__nit__icontains = dato) |
								Q(proyecto__nombre__icontains = dato))					
			

				if empresa and int(empresa)>0:
					qset = qset & (Q(empresa__id = empresa))					


				if mcontrato and int(mcontrato)>0:
					qset = qset & (Q(proyecto__mcontrato = mcontrato))
					
				if departamento and int(departamento)>0:
					qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
					
				if municipio and int(municipio)>0:
					qset = qset & (Q(proyecto__municipio = municipio))
					

				if contratista and int(contratista)>0:
					qset = qset & (Q(proyecto__contrato__contratista__id = contratista))	

			
			if qset != '':
				queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if homepage:
						serializer = ProyectoEmpresaLite3Serializer(page,many=True, context={'request': request})
					else:
						serializer = self.get_serializer(page,many=True)
					
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#Api rest para periodicidad
class PeriodicidadSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = APeriodicidadG
		fields=('id','nombre','numero_dias')


class PeriodicidadGraficoViewSet(viewsets.ModelViewSet):
	
	model=APeriodicidadG
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='avance_de_obra_grafico2.periodicidad'
	queryset = model.objects.all()
	serializer_class = PeriodicidadSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PeriodicidadGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PeriodicidadSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PeriodicidadSerializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para periodicidad


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista',)



#Api rest para Esquema de Capitulos
class EsquemaCapitulosSerializer(serializers.HyperlinkedModelSerializer):

	tipo=tipoC()
	macrocontrato=ContratoLiteSerializer(read_only=True)
	macrocontrato_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.filter(tipo_contrato=tipo.m_contrato))	

	class Meta:
		model = BEsquemaCapitulosG
		fields=('id','macrocontrato','macrocontrato_id','nombre')


class EsquemaCapitulosGraficoViewSet(viewsets.ModelViewSet):
	
	model=BEsquemaCapitulosG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			macrocontrato_id= self.request.query_params.get('macrocontrato_id',None)
			lista_contrato=[]
			tipo=tipoC()
			mcontrato=EmpresaContrato.objects.filter(contrato__tipo_contrato=tipo.m_contrato,empresa=request.user.usuario.empresa.id,participa=1)

			if mcontrato is None:
				lista_contrato.append(0)
			else:
				for item in mcontrato:
					lista_contrato.append(item.contrato.id)

			qset=(Q(macrocontrato_id__in=lista_contrato))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if macrocontrato_id is not None and int(macrocontrato_id)>0:
				qset=qset&(Q(macrocontrato_id=macrocontrato_id))


			
			
			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosSerializer(data=request.data,
					context={'request': request})

				if serializer.is_valid():
					serializer.save(macrocontrato_id=request.data['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,
						accion=self.model_acciones.accion_crear,
						nombre_modelo=self.nombre_modulo,
						id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)			
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(macrocontrato_id=request.data['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos



#Api rest para Esquema de Capitulos de las actividades
class EsquemaCapitulosActividadesSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True, required = False)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())	

	
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','esquema','esquema_id','nombre','nivel','padre','peso')


class EsquemaCapitulosActividadesLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','nombre')



class EsquemaCapitulosActividadesGraficoViewSet(viewsets.ModelViewSet):
	
	model=CEsquemaCapitulosActividadesG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosActividadesSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.esquema_capitulos_actividades'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosActividadesGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			niveles= self.request.query_params.get('niveles',None)
			padre_id= self.request.query_params.get('padre_id',None)
			qset=None


			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if esquema_id is not None and int(esquema_id)>0:
				qset=qset &(Q(esquema_id=esquema_id))


			if padre_id is not None and int(padre_id)>0:
				qset=qset &(Q(padre=padre_id))


			if niveles is not None:
				lista=niveles.split(',')
				qset=qset&(Q(nivel__in=lista))

			if qset is not None:				
				queryset = self.model.objects.filter(qset)
			


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosActividadesSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.data['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
					valor=round(float(model_actividad['peso__sum']),3)+round(float(request.data['peso']),3)
					
					if float(valor) <= 100: 
						serializer.save(esquema_id=request.data['esquema_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.data['padre']
						
						if int(padre)>0:
							if int(request.data['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(request.data['peso'])
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()								
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(request.data['peso'])
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosActividadesSerializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.data['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']

					model_actividad2=CEsquemaCapitulosActividadesG.objects.get(pk=instance.id)
					valor_restante=float(request.data['peso']) - float(model_actividad2.peso)
					valor=float(model_actividad['peso__sum'])+valor_restante

					if float(valor) <= 100: 
						serializer.save(esquema_id=request.data['esquema_id'])					
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.data['padre']

						if int(padre)>0:
							if int(request.data['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(valor_restante)
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(valor_restante)
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos de las actividades




#Api rest para regla de estado
class ReglaEstadoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())

	reglaAnterior=serializers.SerializerMethodField()	

	class Meta:
		model = CReglasEstadoG
		fields=('id','esquema','esquema_id','orden','operador','limite','nombre','reglaAnterior',)

	def get_reglaAnterior(self, obj):
		return CReglasEstadoG.objects.filter(orden__lt=obj.orden,esquema_id=obj.esquema_id).values('id','nombre').order_by('orden').last()


class ReglaEstadoAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=CReglasEstadoG
	queryset = model.objects.all()
	serializer_class = ReglaEstadoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico.regla'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ReglaEstadoAvanceGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			qset=None


			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)

			if esquema_id and esquema_id>0:
				if dato:
					qset = qset &(
					Q(esquema_id=esquema_id)
					)
				else:
					qset =(
					Q(esquema_id=esquema_id)
					)

			if qset !=None:
				queryset = self.model.objects.filter(qset).order_by('orden')


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				reglas=CReglasEstadoG.objects.filter(esquema_id=request.data['esquema_id'])
				if len(reglas) > 0:
					for item in list(reglas):
						if int(item.operador)==int(request.data['operador']) and float(item.limite)==float(request.data['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				orden=None
				orden=CReglasEstadoG.objects.filter(esquema_id=request.data['esquema_id']).order_by('orden').last()
				
				if orden is None:
					request.data['orden']=1
				else:
					request.data['orden']=orden.orden+1

				if int(request.data['regla_anterior'])>0:
					reglas2=CReglasEstadoG.objects.get(pk=request.data['regla_anterior'])
					valor=int(reglas2.orden)+1
					request.data['orden']=valor

					if float(reglas2.limite)>=float(request.data['limite']) and int(reglas2.operador)==int(request.data['operador']):
						return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
			 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.data['esquema_id'])
					if len(reglas3)>0:
						for item in reglas3:
							if float(item.limite)<=float(request.data['limite']) and int(item.operador)==int(request.data['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

						reglas_actualizar=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.data['esquema_id'])
						for item2 in reglas_actualizar:
							estado=CReglasEstadoG.objects.get(pk=item2.id)
							estado.orden=estado.orden+1
							estado.save()
				
				serializer = ReglaEstadoGraficoSerializer(data=request.data,context={'request': request})
				if serializer.is_valid():
					serializer.save(esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.data['esquema_id'])

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				reglas=CReglasEstadoG.objects.filter(esquema_id=request.data['esquema_id'])
				for item in list(reglas):
					if item.id!=instance.id:
						if int(item.operador)==int(request.data['operador']) and float(item.limite)==float(request.data['limite']):
							return Response({'message':'Debe registrar otro operador o limite, el limite y el operador ya existe','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

				if int(request.data['regla_anterior'])>0:					
					valor=CReglasEstadoG.objects.filter(orden__lt=request.data['orden'],esquema_id=request.data['esquema_id']).values('id','nombre','orden').order_by('orden').last()
					if valor is None:
						valor=0

					reglas2=CReglasEstadoG.objects.get(pk=request.data['regla_anterior'])
					if float(reglas2.limite)>=float(request.data['limite']) and int(reglas2.operador)==int(request.data['operador']):
							return Response({'message':'Debe registrar otro limite, el limite es menor que el limite anterior','success':'fail',
				 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					reglas3=CReglasEstadoG.objects.filter(orden__gt=request.data['orden'],esquema_id=request.data['esquema_id']).values('id','nombre','orden','limite','operador').order_by('orden').first()
					if reglas3 is not None:
							if float(reglas3['limite'])<=float(request.data['limite']) and int(reglas3['operador'])==int(request.data['operador']):
								return Response({'message':'Debe registrar otro limite, el limite es mayor que el limite siguiente','success':'fail',
						 					'data':''},status=status.HTTP_400_BAD_REQUEST)

					if int(request.data['regla_anterior'])!=int(valor['id']):
						valor=int(reglas2.orden)+1
						request.data['orden']=valor	
						if reglas3 is not None:
							reglas_actualizar=CReglasEstadoG.objects.filter(orden__gt=reglas2.orden,esquema_id=request.data['esquema_id'])
							for item2 in reglas_actualizar:
								estado=CReglasEstadoG.objects.get(pk=item2.id)
								estado.orden=estado.orden+1
								estado.save()


				serializer = ReglaEstadoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()
					updateAsyncEstado.delay(request.data['esquema_id'])
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)				
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para regla de estado

#serializer lite de proyecto
class ProyectoLite2Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	municipio = MunicipioSerializer(read_only = True)
	contrato = ContratoLiteSerializer(read_only = True, many = True)

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'mcontrato' ,
				 'municipio' ,
				 'contrato',
				 )


#Api rest para Cronograma

class CronogramaSerializerLite(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = Cronograma
		fields = ('id','nombre')

class CronogramaSerializer(serializers.HyperlinkedModelSerializer):

	proyecto=ProyectoLite2Serializer(read_only=True)
	proyecto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all())	

	estado=ReglaEstadoGraficoSerializer(read_only=True, allow_null = True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CReglasEstadoG.objects.all(),allow_null = True)	

	periodicidad=PeriodicidadSerializer(read_only=True, allow_null = True)
	periodicidad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=APeriodicidadG.objects.all(),allow_null = True)	

	esquema=EsquemaCapitulosSerializer(read_only=True, allow_null = True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all(),allow_null = True)	

	porcentajeAvance=serializers.SerializerMethodField()

	porcentajeAvanceFinanciero=serializers.SerializerMethodField()

	class Meta:
		model = Cronograma
		fields=('id','proyecto','proyecto_id','periodicidad',
			'esquema_id','esquema','periodicidad_id','estado',
			'estado_id','programacionCerrada','nombre', 'porcentajeAvance',
			'porcentajeAvanceFinanciero')

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []

		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['detallepresupuesto__actividad__peso']
				})
		##import pdb; pdb.set_trace()
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		return round(avance,2)

	def get_porcentajeAvanceFinanciero(self,obj):
		totalPresupuesto = 0
		totalEjecutado = 0
		avance = 0
		cantidadesEjecutar = FDetallePresupuesto.objects.filter(
			presupuesto__cronograma__proyecto__id=obj.proyecto.id).values(
			'cantidad','codigoUC','cantidad',
			'catalogoUnidadConstructiva__id', 'id')

		for rowCantidadesEjecutar in cantidadesEjecutar:
			totalMo = 0
			totalMat = 0
			queryset = DesgloceManoDeObra.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']
				).values(
				'manoDeObra__valorHora','rendimiento')
			
			if queryset:
				for row in queryset:
					totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

			queryset = DesgloceMaterial.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']).values(
				'material__valorUnitario','cantidad')
			
			if queryset:
				for row in queryset:
					totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))

			totalPresupuesto = totalPresupuesto + \
			((totalMo * float(rowCantidadesEjecutar['cantidad'])) + \
			(totalMat * float(rowCantidadesEjecutar['cantidad'])))

			cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__id=rowCantidadesEjecutar['id'],
			cantidadEjecutada__gte=0).values(
			'detallepresupuesto__id').annotate(total=Sum('cantidadEjecutada'))

			if cantidadesEjecutadas:
				totalEjecutado = totalEjecutado + \
				(float(cantidadesEjecutadas[0]['total']) * totalMo) + \
				(float(cantidadesEjecutadas[0]['total']) * totalMat)

		if totalPresupuesto > 0:		
			avance = (totalEjecutado / totalPresupuesto) * 100
		else:
			avance = 0

		return round(avance,2)


class CronogramaAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=Cronograma
	queryset = model.objects.all()
	serializer_class = CronogramaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cronograma'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CronogramaAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			esquema_id= self.request.query_params.get('esquema_id',None)


			qset=(~Q(id=0))

			if proyecto_id:
				qset = qset &(
					Q(proyecto_id=proyecto_id)
					)

			if esquema_id:
				qset = qset &(
					Q(esquema_id=esquema_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if request.data['estado_id']=='0':
					request.data._mutable = True
					request.data['estado_id']=None
					request.data._mutable = False

				serializer = CronogramaSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(periodicidad_id=request.data['periodicidad_id'],proyecto_id=request.data['proyecto_id'],estado_id=request.data['estado_id'],esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					##import pdb; pdb.set_trace()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.data['estado_id']==0:
					request.data['estado_id']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CronogramaSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(periodicidad_id=request.data['periodicidad_id'],proyecto_id=request.data['proyecto_id'],estado_id=request.data['estado_id'],esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cronograma


#Api rest para Presupuesto
class PresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())	

	totalCambios=serializers.SerializerMethodField()

	class Meta:
		model = EPresupuesto
		fields=('id','cronograma','cronograma_id','cerrar_presupuesto','nombre','totalCambios','sin_poste')


	def get_totalCambios(self, obj):
		return LCambio.objects.filter(presupuesto_id=obj.id).count()


class PresupuestoAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=EPresupuesto
	queryset = model.objects.all()
	serializer_class = PresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PresupuestoAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			cerrado_presupuesto= self.request.query_params.get('cerrado_presupuesto',None)


			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(cronograma__id=cronograma_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if cerrado_presupuesto:
				qset = qset &(
					Q(cerrar_presupuesto=cerrado_presupuesto)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PresupuestoGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PresupuestoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Presupuesto




#Api rest para Detalle Presupuesto
from .models import CatalogoUnidadConstructiva
class CatalogoUnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	mcontrato_id = serializers.PrimaryKeyRelatedField(read_only=True)
	
	class Meta:
		model=CatalogoUnidadConstructiva
		fields=('id','nombre','ano','activo','mcontrato', 'mcontrato_id')

class MaterialSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=Material
		fields=('id','codigo','descripcion','valorUnitario', 'unidadMedida', 'catalogo_id', 'catalogo')

class MaterialSerializerLite(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model=Material
		fields=('id','descripcion','codigo')

class MaterialSerializerLite2(serializers.HyperlinkedModelSerializer):

	class Meta:
		model=Material
		fields=('id','codigo','descripcion','valorUnitario', 'unidadMedida')



class ManoObraSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion','valorHora', 'catalogo_id', 'catalogo')

class ManoObraSerializerLite2(serializers.HyperlinkedModelSerializer):

	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion','valorHora')		

class ManoObraSerializerLite(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion')


class DetallePresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	catalogoUnidadConstructiva = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogoUnidadConstructiva_id=serializers.PrimaryKeyRelatedField(write_only=True,
		queryset=CatalogoUnidadConstructiva.objects.all())
	sumaPresupuesto=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','sumaPresupuesto','presupuesto','presupuesto_id',
			'actividad','disponibilidad_cantidad_apoyo','cantidad_apoyo',
			'actividad_id','codigoUC','descripcionUC','valorManoObra','valorMaterial',
			'valorGlobal','cantidad','porcentaje','nombre_padre',
			'catalogoUnidadConstructiva','catalogoUnidadConstructiva_id')


	def get_sumaPresupuesto(self, obj):
		sumatoria=0
		suma=FDetallePresupuesto.objects.filter(presupuesto_id=obj.presupuesto.id)	
		valor=0
		for item in suma:
			valor=float(item.valorMaterial)+float(item.valorManoObra)
			total=valor*float(item.cantidad)

			sumatoria=sumatoria+total

		return round(sumatoria,3)
		
class DetallePresupuestoGraficoLiteSerializer(serializers.HyperlinkedModelSerializer):
	actividad=EsquemaCapitulosActividadesLiteSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	sumaPresupuesto=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','sumaPresupuesto','actividad','cantidad_apoyo','actividad_id','codigoUC','descripcionUC',
		'valorGlobal','cantidad','nombre_padre','disponibilidad_cantidad_apoyo')


	def get_sumaPresupuesto(self, obj):
		sumatoria=0
		suma=FDetallePresupuesto.objects.filter(presupuesto_id=obj.presupuesto.id)	
		valor=0
		for item in suma:
			valor=float(item.valorMaterial)+float(item.valorManoObra)
			total=valor*float(item.cantidad)

			sumatoria=sumatoria+total

		return round(sumatoria,3)


class DetallePresupuestoGraficoViewSet(viewsets.ModelViewSet):
	
	model=FDetallePresupuesto
	queryset = model.objects.all()
	serializer_class = DetallePresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetallePresupuestoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			hito_id= self.request.query_params.get('hito_id',None)
			actividad_id= self.request.query_params.get('actividad_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_apoyo= self.request.query_params.get('listado_apoyo',None)
			id_apoyo= self.request.query_params.get('id_apoyo',None)
			lite2= self.request.query_params.get('lite2',None)



			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if actividad_id:
				qset = qset &(
					Q(actividad_id=actividad_id)
					)

			if hito_id:
				qset = qset &(
					Q(actividad__padre=hito_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				presupuesto_id=cronograma.presupuesto.id
				qset = qset &(
					Q(presupuesto_id=cronograma.presupuesto.id)
					)

			lista=[]
			if listado_apoyo and presupuesto_id and int(presupuesto_id)>0 and id_apoyo:
				lista=HNodo.objects.filter(presupuesto_id=presupuesto_id).exclude(id=id_apoyo).values('id','nombre')


			if dato:
				qset = qset &(
					Q(codigoUC__icontains=dato)|
					Q(descripcionUC__icontains=dato)|
					Q(actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite2:
						serializer = DetallePresupuestoGraficoLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)
					if listado_apoyo:						
						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'apoyos':lista}})
					else:
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
			

			serializer = self.get_serializer(queryset,many=True)
			if listado_apoyo:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'apoyos':lista}})			
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetallePresupuestoGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.data['presupuesto_id'],
						actividad_id=request.data['actividad_id'],
						catalogoUnidadConstructiva_id=request.data['catalogoUnidadConstructiva_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetallePresupuestoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.data['presupuesto_id'],
						actividad_id=request.data['actividad_id'],
						catalogoUnidadConstructiva_id=request.data['catalogoUnidadConstructiva_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle presupuesto


#Api rest para Diagrama Grahm
class DiagramaGrahmGraficoSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	
	class Meta:
		model = DiagramaGrahm
		fields=('id','cronograma','cronograma_id','actividad','actividad_id',
			'nombre_padre','fechaInicio','fechaFinal','actividad_inicial')

class DiagramaGrahmGraficoLiteSerializer(serializers.HyperlinkedModelSerializer):

	actividad = EsquemaCapitulosActividadesLiteSerializer(read_only=True)
	cronograma = CronogramaSerializerLite(read_only=True)

	class Meta:
		model = DiagramaGrahm
		fields = ('id','cronograma','actividad','nombre_padre','fechaInicio',
			'fechaFinal','actividad_inicial')

class DiagramaGrahmGraficoViewSet(viewsets.ModelViewSet):
	
	model=DiagramaGrahm
	queryset = model.objects.all()
	serializer_class = DiagramaGrahmGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.diagrama'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DiagramaGrahmGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			lite = self.request.query_params.get('lite',None)


			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(actividad__nombre__icontains=dato)
					)

			if cronograma_id:
				qset = qset &(
					Q(cronograma_id=cronograma_id)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = DiagramaGrahmGraficoLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
			if lite:
				serializer = DiagramaGrahmGraficoLiteSerializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:				

				if request.data['fechaInicio']>request.data['fechaFinal']:
					return JsonResponse({
						'message':'La fecha de inicio es posterior a la fecha de fin' + \
						'. Sugerimos corregir la informacion ingresada',
						'success':'error',
						'data':''})

				serializer = DiagramaGrahmGraficoSerializer(data=request.data,context={'request': request})

				validacion_actividad=DiagramaGrahm.objects.filter(actividad_id=request.data['actividad_id'],cronograma_id=request.data['cronograma_id'])

				if len(validacion_actividad)>0:
					return Response({'message':'La actividad ya esta registrada en el diagrama.','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(actividad_id=request.data['actividad_id'],cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DiagramaGrahmGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
			

				if serializer.is_valid():				
					serializer.save(actividad_id=request.data['actividad_id'],cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Diagrama Grahm


#Api rest para Capa
class CapaSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = GCapa
		fields=('id','nombre','color')

#Api rest para Nodo
class NodoGraficoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = HNodo
		fields = ('id','nombre')

class NodoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	


	class Meta:
		model = HNodo
		fields=('id','presupuesto','porcentajeAcumulado','presupuesto_id','eliminado','capa','capa_id','longitud','latitud','noProgramado','nombre')

class NodoGraficoLite2Serializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = HNodo
		fields = ('id','nombre','porcentajeAcumulado','longitud','latitud')


class NodoGraficoViewSet(viewsets.ModelViewSet):
	
	model=HNodo
	queryset = model.objects.all()
	serializer_class = NodoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(NodoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_enlaces= self.request.query_params.get('listado_enlaces',None)
			ejecucion= self.request.query_params.get('ejecucion',None)
			programando= self.request.query_params.get('programando',None)
			filtro_gps= self.request.query_params.get('filtro_gps',None)
			eliminado= self.request.query_params.get('eliminado',None)
			apoyo_cambio= self.request.query_params.get('apoyo_cambio',None)
			estados= self.request.query_params.get('id_estados',None)
			lite = self.request.query_params.get('lite',None)


			qset=(~Q(id=0))

			if filtro_gps is not None and int(filtro_gps)>0:
				if int(filtro_gps)==2:
					qset=qset&(Q(latitud=None))&(Q(longitud=None))
				if int(filtro_gps)==1:
					qset=qset&(~Q(latitud=None))&(~Q(longitud=None))


			if apoyo_cambio is not None:
				listado_estado=estados.split(',')

				listado_id=[]
				cambio=LCambio.objects.filter(cronograma_id=cronograma_id,estado_id__in=listado_estado)

				listado_id_ejecucion=[]
				if int(ejecucion)>0:

					listado_por=[]
					if int(ejecucion) == 2:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__gte=100)
						
					else:
						listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=cronograma_id,porcentaje__lt=100)


					for obj in listado_por:
						listado_id_ejecucion.append(obj.apoyo.id)

				if cambio is not None:
					for obj in cambio:
						for item in obj.nodos.all():
							listado_id.append(item.id)

				cronograma=Cronograma.objects.get(pk=cronograma_id)
				
				qset=(Q(presupuesto_id=cronograma.presupuesto.id)&(Q(eliminado=False)))

				if len(listado_id_ejecucion)==0:
					qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

				if len(listado_id_ejecucion)>0:
					qset=qset&(Q(id__in=listado_id_ejecucion))


				if dato!='':
					qset=qset&(Q(nombre__icontains=dato))
			else:

				if presupuesto_id:
					qset = qset &(
						Q(presupuesto_id=presupuesto_id)
						)

				if programando and programando!='':
					qset = qset &(
						Q(noProgramado=programando)
						)

				if eliminado and eliminado!='':
					qset = qset &(
						Q(eliminado=eliminado)
						)

				if cronograma_id:
					cronograma=Cronograma.objects.get(pk=cronograma_id)
					presupuesto_id=cronograma.presupuesto.id
					qset = qset &(
						Q(presupuesto_id=cronograma.presupuesto.id)
						)

				if dato:
					qset = qset &(
						Q(nombre__icontains=dato)
						)


				lista=[]
				if listado_enlaces and cronograma_id:
					lista=IEnlace.objects.filter(nodoOrigen__presupuesto__id=presupuesto_id).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')


				# if ejecucion:
				# 	programacion=LProgramacion.objects.filter(cronograma_id=cronograma_id)

			if qset is not None:
				#print ("texto")
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					if listado_enlaces:
						return self.get_paginated_response({'message':'','success':'ok',
							'data':{'datos':serializer.data,'enlace':lista}})
					else:	
						return self.get_paginated_response({'message':'','success':'ok',
							'data':serializer.data})
			if lite:
				serializer = NodoGraficoLite2Serializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)
				
			if listado_enlaces:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'enlace':lista}})
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()

			if request.data['longitud'] == '' or request.data['latitud'] == '':
				request.data['longitud']=None
				request.data['latitud']=None

			try:
				serializer = NodoGraficoSerializer(data=request.data,context={'request': request})

				nombre_nodo=HNodo.objects.filter(nombre=request.data['nombre'],presupuesto_id=request.data['presupuesto_id'])

				if len(nombre_nodo)>0:
					return Response({'message':'El nombre del apoyo ya existe, digite otro nombre','success':'ok',
						'data':''},status=status.HTTP_400_BAD_REQUEST)


				if serializer.is_valid():
					serializer.save(presupuesto_id=request.data['presupuesto_id'],capa_id=request.data['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					presupuesto=EPresupuesto.objects.get(pk=request.data['presupuesto_id'])

					if presupuesto.sin_poste == True:
						agregarSinPoste.delay(presupuesto.id,request.user.usuario.id,serializer.data['id'])	


					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.data['longitud'] == '' or request.data['latitud'] == '':
					request.data['longitud']=None
					request.data['latitud']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = NodoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.data['presupuesto_id'],capa_id=request.data['capa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Nodo



#Api rest para Cantidad de Nodo
class CantidadNodoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())	

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	class Meta:
		model = JCantidadesNodo
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodo','nodo_id','cantidad')


class CantidadNodoGraficoViewSet(viewsets.ModelViewSet):
	
	model=JCantidadesNodo
	queryset = model.objects.all()
	serializer_class = CantidadNodoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cantidad_nodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CantidadNodoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detalle_presupuesto_id= self.request.query_params.get('detalle_presupuesto_id',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)



			qset=(~Q(id=0))

			if detalle_presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detalle_presupuesto_id)
					)

			if nodo_id:
				qset = qset &(
					Q(nodo_id=nodo_id)
					)

			if presupuesto_id:
				qset = qset &(
					Q(detallepresupuesto__presupuesto__id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(nodo__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CantidadNodoGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],nodo_id=request.data['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CantidadNodoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],nodo_id=request.data['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cantidad Nodo



#Api rest para Enlance
class EnlaceGraficoSerializer(serializers.HyperlinkedModelSerializer):

	nodoOrigen=NodoGraficoSerializer(read_only=True)
	nodoOrigen_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())

	nodoDestino=NodoGraficoSerializer(read_only=True)
	nodoDestino_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())	

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())

	capa=CapaSerializer(read_only=True)
	capa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=GCapa.objects.all())	

	class Meta:
		model = IEnlace
		fields=('id','detallepresupuesto','detallepresupuesto_id','nodoOrigen','nodoOrigen_id','nodoDestino','nodoDestino_id','capa','capa_id')


class EnlaceAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=IEnlace
	queryset = model.objects.all()
	serializer_class = EnlaceGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.enlace'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EnlaceAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			detallepresupuesto_id= self.request.query_params.get('detallepresupuesto_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			nodo_origen= self.request.query_params.get('nodo_origen',None)
			nodo_destino= self.request.query_params.get('nodo_destino',None)

			qset=(~Q(id=0))

			if detallepresupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detallepresupuesto_id)
					)

			if cronograma_id:
				cronograma=KCronograma.objects.get(pk=cronograma_id)
				qset=qset &(
						Q(nodoOrigen__presupuesto__id=cronograma.presupuesto.id)&
						Q(nodoDestino__presupuesto__id=cronograma.presupuesto.id)
					)

			if nodo_origen:
				qset = qset &(
					Q(nodoOrigen_id=nodo_origen)
					)

			if nodo_destino:
				qset = qset &(
					Q(nodoDestino_id=nodo_destino)
					)


			if dato:
				qset = qset &(
					Q(nodoOrigen__nombre__icontains=dato)|
					Q(nodoDestino__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EnlaceGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],capa_id=request.data['capa_id'],nodoOrigen_id=request.data['nodoOrigen_id'],nodoDestino_id=request.data['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EnlaceGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],capa_id=request.data['capa_id'],nodoOrigen_id=request.data['nodoOrigen_id'],nodoDestino_id=request.data['nodoDestino_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin Api rest para Cantidad Nodo


#Api rest para Reporte de Trabajo
class ReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())

	usuario_registro=UsuarioSerializer(read_only=True)
	usuario_registro_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())
	
	usuario_aprueba=UsuarioSerializer(read_only=True, allow_null = True)
	usuario_aprueba_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all(), allow_null = True)

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app="Avance_obra_grafico"))

	empresa=EmpresaSerializer(read_only=True)
	empresa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	totalRechazados=serializers.SerializerMethodField()

	class Meta:
		model = JReporteTrabajo
		fields=('id','fecha_format','totalRechazados','reporteCerrado','empresa','empresa_id','presupuesto','presupuesto_id','estado','estado_id','fechaTrabajo','usuario_registro','usuario_registro_id','valor_ganando_acumulado','avance_obra_acumulado','sinAvance','motivoSinAvance','fecharevision','soporteAprobacion','usuario_aprueba','usuario_aprueba_id')


	def get_totalRechazados(self, obj):
		return MComentarioRechazo.objects.filter(reporte_trabajo_id=obj.id).count()

class ReformadoSerializer(serializers.HyperlinkedModelSerializer):			
	usuario_registro=UsuarioSerializer(read_only=True)
	usuario_registro_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())
	
	class Meta:
		model = EReformado
		fields=('id','fecha_registro','usuario_registro','usuario_registro_id')

class ReformadoDetalleSerializer(serializers.HyperlinkedModelSerializer):			
	apoyo = NodoGraficoLiteSerializer(read_only=True)
	apoyo_id = serializers.PrimaryKeyRelatedField(queryset=HNodo.objects.all(), write_only=True)
	reformado = ReformadoSerializer(read_only=True)
	reformado_id = serializers.PrimaryKeyRelatedField(queryset=EReformado.objects.all(), write_only=True)	
	
	class Meta:
		model = EReformadoDetalle
		fields=('id','codigo_uucc','descripcion_uucc','cantidad_anterior','cantidad_final','diferencia','apoyo','apoyo_id','reformado','reformado_id')


class ReporteTrabajoGraficoViewSet(viewsets.ModelViewSet):
	
	model=JReporteTrabajo
	queryset = model.objects.all()
	serializer_class = ReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		
		try:
			queryset = super(ReporteTrabajoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)			
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			empresa_id= self.request.query_params.get('empresa_id',None)
			registrado= self.request.query_params.get('registrado',None)
			reformado = self.request.query_params.get('reformado',None)
			reformado_id = self.request.query_params.get('reformado_id',None)

			qset=(~Q(id=0))
			
			if empresa_id:
				lista=[]				

				if registrado:
					reportes=JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=proyecto_id)
					sw=0
					
					revision=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
					aprobado=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
					corregido=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
					rechazados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)

					for item in reportes:
						for valor in item.presupuesto.cronograma.proyecto.contrato.all():
							if int(valor.tipo_contrato.id)==9 and int(valor.contratista.id)==int(empresa_id):
								sw=1

						if sw==1:
							if int(registrado)==2:
								if int(item.estado.id)==int(corregido[0].id):
									lista.append(item.id)
							elif int(registrado)==3:
								if int(item.estado.id)==int(rechazados[0].id):
									lista.append(item.id)
							else:
								if int(item.estado.id)==int(revision[0].id) or int(item.estado.id)==int(aprobado[0].id):
									lista.append(item.id)											
					if sw == 0:
						qset = qset &(
							Q(empresa_id=empresa_id)
							)
					else:
						qset=qset & (Q(id__in=lista))


				else:
					reportes=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id)
					# sw=0
					# estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)
					# for item in reportes:
					# 	if sw==1:
					# 		if int(item.estado.id)!=int(estados[0].id):
					# 			lista.append(item.id)

					# 	for valor in item.presupuesto.cronograma.proyecto.contrato.all():
					# 		if int(valor.tipo_contrato.id)==9 and int(valor.contratista.id)==int(empresa_id):
					# 			sw=1

					# if sw == 0:
					# 	qset = qset &(
					# 		Q(empresa_id=empresa_id)
					# 		)
					# else:
					# 	qset=qset & (Q(id__in=lista))

			#import pdb; pdb.set_trace()			
			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)		

			if proyecto_id:
				qset = qset &(
					Q(presupuesto__cronograma__proyecto__id=proyecto_id)
					)		


			if dato:
				qset = qset &(
					Q(presupuesto__nombre__icontains=dato)|
					Q(presupuesto__cronograma__nombre__icontains=dato)|
					Q(usuario_registro__persona__nombres__icontains=dato)|
					Q(usuario_registro__persona__apellidos__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-fechaTrabajo')

			##import pdb; pdb.set_trace()
			if reformado is not None:
				model=EReformado
				queryset = model.objects.all()
				serializer_class = ReformadoSerializer	
				model_log=Logs
				model_acciones=Acciones	
				nombre_modulo='avance_de_obra_grafico2.reformado'				
				
				qsetDetalle = Q(apoyo__presupuesto__id = presupuesto_id)
				ListRefor = EReformadoDetalle.objects.filter(qsetDetalle).values('reformado_id')#.order_by('-id')
				if ListRefor:
					qset2 = Q(id__in = ListRefor)
					if dato:
						qset2 = qset2 &(Q(id__icontains=dato))					
					queryset = EReformado.objects.filter(qset2).order_by('-id')
					if sin_paginacion is None:
						page = self.paginate_queryset(queryset)
						if page is not None:
							serializer = ReformadoSerializer(page,many=True)	
							return self.get_paginated_response({'message':'','success':'ok',
							'data':serializer.data})
				else:
					if sin_paginacion is None:
						page = self.paginate_queryset(queryset)
						if page is not None:
							serializer = ReformadoSerializer(page,many=True)	
							return self.get_paginated_response({'message':'','success':'ok',
							'data':serializer.data})					

			##import pdb; pdb.set_trace()
			if reformado_id is not None:
				model=EReformadoDetalle
				queryset = model.objects.all()
				serializer_class = ReformadoDetalleSerializer	
				model_log=Logs
				model_acciones=Acciones	
				nombre_modulo='avance_de_obra_grafico2.reformado'				
				
				qsetDetalle = Q(reformado__id = reformado_id)
				queryset = EReformadoDetalle.objects.filter(qsetDetalle).order_by('-id')
				if sin_paginacion is None:
					page = self.paginate_queryset(queryset)
					if page is not None:
						serializer = ReformadoDetalleSerializer(page,many=True)	
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})				
				

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = ReporteTrabajoSerializer(data=request.data,context={'request': request})

				verifica_reporte=JReporteTrabajo.objects.filter(presupuesto_id=request.data['presupuesto_id'],fechaTrabajo=request.data['fechaTrabajo'])

				if len(verifica_reporte)>0:
					return Response({'message':'No se puede registrar dos reportes de trabajos con la misma fecha','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)

				if request.data['usuario_aprueba_id'] == '':
					request.data['usuario_aprueba_id']=None

				if serializer.is_valid():
					serializer.save(empresa_id=request.data['empresa_id'],presupuesto_id=request.data['presupuesto_id'],soporteAprobacion=self.request.FILES.get('soporteAprobacion') if self.request.FILES.get('soporteAprobacion') is not None else None,usuario_registro_id=request.data['usuario_registro_id'],usuario_aprueba_id=request.data['usuario_aprueba_id'],estado_id=request.data['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					errors = funciones.erroresSerializer(serializer.errors)
					functions.toDeatilLog(errors, self.nombre_modulo + ".create")
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = ReporteTrabajoSerializer(instance,data=request.data,context={'request': request},partial=partial)

				verifica_reporte=JReporteTrabajo.objects.filter((~Q(id=instance.id))&(Q(presupuesto_id=request.data['presupuesto_id']))&(Q(fechaTrabajo=request.data['fechaTrabajo'])))

				if len(verifica_reporte)>0:
					return Response({'message':'No se puede registrar dos reportes de trabajos con la misma fecha','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
				
				if serializer.is_valid():				
					serializer.save(empresa_id=request.data['empresa_id'],presupuesto_id=request.data['presupuesto_id'],usuario_registro_id=request.data['usuario_registro_id'],usuario_aprueba_id=request.data['usuario_aprueba_id'],estado_id=request.data['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Reporte


#Api rest para Detalle Reporte de Trabajo
class DetalleReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):

	reporte_trabajo=ReporteTrabajoSerializer(read_only=True)
	reporte_trabajo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=JReporteTrabajo.objects.all())

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())
	
	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(queryset=FDetallePresupuesto.objects.all())

	class Meta:
		model = KDetalleReporteTrabajo
		fields=('id','reporte_trabajo','reporte_trabajo_id','nodo','nodo_id','detallepresupuesto','detallepresupuesto_id','cantidadEjecutada')


class DetalleReporteTrabajoGraficoViewSet(viewsets.ModelViewSet):
	
	model=KDetalleReporteTrabajo
	queryset = model.objects.all()
	serializer_class = DetalleReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleReporteTrabajoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			reporte_id= self.request.query_params.get('reporte_id',None)
			nodo_id= self.request.query_params.get('nodo_id',None)
			detallepresupuesto_id= self.request.query_params.get('detallepresupuesto_id',None)

			qset=(~Q(id=0))

			if reporte_id:
				qset = qset &(
					Q(reporte_trabajo_id=reporte_id)
					)

			if nodo_id:
				qset = qset &(
					Q(nodo_id=nodo_id)
					)

			if detallepresupuesto_id:
				qset = qset &(
					Q(detallepresupuesto_id=detallepresupuesto_id)
					)		


			if dato:
				qset = qset &(
					Q(detallepresupuesto__actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetalleReporteTrabajoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(reporte_trabajo_id=request.data['reporte_trabajo_id'],nodo_id=request.data['nodo_id'],detallepresupuesto_id=request.data['detallepresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleReporteTrabajoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(reporte_trabajo_id=request.data['reporte_trabajo_id'],nodo_id=request.data['nodo_id'],detallepresupuesto_id=request.data['detallepresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Reporte


#Api rest para Cambio
class CambioGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())

	empresaSolicitante=EmpresaSerializer(read_only=True)
	empresaSolicitante_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())
	
	empresaTecnica=EmpresaSerializer(read_only=True)
	empresaTecnica_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	empresaFinanciera=EmpresaSerializer(read_only=True)
	empresaFinanciera_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app="Avance_obra_grafico_cambio"))

	mensajeCancelado=serializers.SerializerMethodField()

	class Meta:
		model = LCambio
		fields=('id','mensajeCancelado','presupuesto','presupuesto_id','estado','estado_id','descripcion','motivo','empresaSolicitante_id','empresaSolicitante','empresaTecnica','empresaTecnica_id','empresaFinanciera','empresaFinanciera_id')

	def get_mensajeCancelado(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
		historial=LHistorialCambio.objects.filter(estado_id=estados[0].id,cambio_id=obj.id).order_by('-id')
		mensaje=''

		if len(historial) > 0:
			mensaje=historial[0].motivoEstado
		return mensaje


class CambioGraficoViewSet(viewsets.ModelViewSet):
	
	model=LCambio
	queryset = model.objects.all()
	serializer_class = CambioGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CambioGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)

			qset=(~Q(id=0))

			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if dato:
				qset = qset &(
					Q(descripcion__icontains=dato)|
					Q(motivo__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CambioGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.data['presupuesto_id'],empresaSolicitante_id=request.data['empresaSolicitante_id'],empresaTecnica_id=request.data['empresaTecnica_id'],empresaFinanciera_id=request.data['empresaFinanciera_id'],estado_id=request.data['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					historial=LHistorialCambio(cambio_id=serializer.data['id'],usuario_registro_id=request.user.usuario.id,estado_id=request.data['estado_id'])
					historial.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='avanceObraGrafico2.historial_cambio',id_manipulado=historial.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CambioGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.data['presupuesto_id'],empresaSolicitante_id=request.data['empresaSolicitante_id'],empresaTecnica_id=request.data['empresaTecnica_id'],empresaFinanciera_id=request.data['empresaFinanciera_id'],estado_id=request.data['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cambio


#Api rest para mensaje de rechazo
class MensajeRechazoSerializer(serializers.HyperlinkedModelSerializer):

	reporte_trabajo=ReporteTrabajoSerializer(read_only=True)
	reporte_trabajo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=JReporteTrabajo.objects.all())

	class Meta:
		model = MComentarioRechazo
		fields=('id','reporte_trabajo','reporte_trabajo_id','fecha_format','fecha_hora','motivoRechazo')


class MensajeRechazoViewSet(viewsets.ModelViewSet):
	
	model=MComentarioRechazo
	queryset = model.objects.all()
	serializer_class = MensajeRechazoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.mensaje_rechazo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(MensajeRechazoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			reporte_trabajo_id= self.request.query_params.get('reporte_trabajo_id',None)

			qset=(~Q(id=0))

			if reporte_trabajo_id:
				qset = qset &(
					Q(reporte_trabajo_id=reporte_trabajo_id)
					)

			if dato:
				qset = qset &(
					Q(motivoRechazo__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-fecha_hora')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = MensajeRechazoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(reporte_trabajo_id=request.data['reporte_trabajo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = MensajeRechazoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(reporte_trabajo_id=request.data['reporte_trabajo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
			
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Cambio


#Api rest para Detalle de cambio
class DetalleCambioGraficoSerializer(serializers.HyperlinkedModelSerializer):

	detallepresupuesto=DetallePresupuestoGraficoSerializer(read_only=True)
	detallepresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all(),allow_null = True)

	cambio=CambioGraficoSerializer(read_only=True)
	cambio_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=LCambio.objects.all())

	nodo=NodoGraficoSerializer(read_only=True)
	nodo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=HNodo.objects.all())
	
	class Meta:
		model = LDetalleCambio
		fields=('id','nodo','nodo_id','operacion','cantidadPropuesta','detallepresupuesto','detallepresupuesto_id','codigoUC','descripcionUC','valorManoObra','valorMaterial','valorGlobal','cambio','cambio_id')


class DetalleCambioGraficoViewSet(viewsets.ModelViewSet):
	
	model=LDetalleCambio
	queryset = model.objects.all()
	serializer_class = DetalleCambioGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avance_de_obra_grafico2.detalle_cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleCambioGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cambio_id= self.request.query_params.get('cambio_id',None)


			qset=(~Q(id=0))

			if dato:
				qset = qset &(
					Q(detallepresupuesto__actividad__nombre__icontains=dato)
					)

			if cambio_id:
				qset = qset &(
					Q(cambio_id=cambio_id)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if int(request.data['detallepresupuesto_id'])==0:
					request.data['detallepresupuesto_id']=None

				serializer = DetalleCambioGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],cambio_id=request.data['cambio_id'],nodo_id=request.data['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleCambioGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
			

				if serializer.is_valid():				
					serializer.save(detallepresupuesto_id=request.data['detallepresupuesto_id'],cambio_id=request.data['cambio_id'],nodo_id=request.data['nodo_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle de cambio



#Inicio eliminacion de un esquema
@transaction.atomic
def eliminar_esquema(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				CReglasEstadoG.objects.filter(esquema_id=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.filter(esquema_id=item['id']).delete()
				BEsquemaCapitulosG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.esquema',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un esquema



#eliminacion de capitulo/actividad de un esquema
@transaction.atomic
def eliminar_id_capitulo_actividad_esquema(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])

			if int(valor.padre)==0:
				CEsquemaCapitulosActividadesG.objects.filter(padre=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=item['id'])
				logs_model.save()

			else:
				modelhijos=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])
				model=CEsquemaCapitulosActividadesG.objects.get(pk=modelhijos.padre)
				model.peso=float(model.peso) - float(modelhijos.peso)
				model.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=model.id)
				logs_model.save()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.capitulos_esquema',id_manipulado=item['id'])
				logs_model2.save()

		transaction.savepoint_commit(sid)			
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra.capitulos_esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de capitulo/actividad de un esquema


#Inicio clonacion de los esquemas
@login_required
@transaction.atomic
def clonacion_esquema(request):
	sid = transaction.savepoint()
	##import pdb; pdb.set_trace()
	try:
		# lista=request.POST['_content']
		# respuesta= json.loads(lista)

		esquema=BEsquemaCapitulosG(nombre=request.POST['nombre_esquema'],macrocontrato_id=int(request.POST['id_macrocontrato']))
		esquema.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema',id_manipulado=esquema.id)
		logs_model.save()

		capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=int(request.POST['id_etiqueta']),nivel=1)

		for item in capitulos:
			hitos=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item.peso,nombre=item.nombre,nivel=item.nivel,padre=item.padre)
			hitos.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema_capitulos',id_manipulado=hitos.id)
			logs_model2.save()

			actividades=CEsquemaCapitulosActividadesG.objects.filter(padre=item.id)

			for item2 in actividades:
				actividad=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item2.peso,nombre=item2.nombre,nivel=item2.nivel,padre=hitos.id)
				actividad.save()
				logs_model3=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.esquema_capitulos',id_manipulado=actividad.id)
				logs_model3.save()

		transaction.savepoint_commit(sid)		
		return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})
		#return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
		#			'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin clonacion de los esquemas


#eliminacion de regla de estado
@transaction.atomic
def eliminar_id_regla_estado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		esquema_id=None

		for item in respuesta['lista']:
			valor=CReglasEstadoG.objects.get(pk=item['id'])
			esquema_id=valor.esquema_id
			valor.delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.regla_estado',id_manipulado=item['id'])
			logs_model.save()	

		updateAsyncEstado.delay(esquema_id)	
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		print(e) 
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra.regla_estado')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de regla de estado


#eliminacion de cronograma
@transaction.atomic
def eliminar_id_cronograma(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=Cronograma.objects.get(pk=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.cronograma',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.cronograma')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de cronograma

@login_required
@transaction.atomic
def eliminar_id_reportetrabajo(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			JReporteTrabajo.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@login_required
@transaction.atomic
def eliminar_id_uucc(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			UnidadConstructiva.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.UnidadConstructiva',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_materiales(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			Material.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_mano_obra(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			ManoDeObra.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_desgl_mat(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DesgloceMaterial.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_desgl_mo(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DesgloceManoDeObra.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
@transaction.atomic
def eliminar_id_presupuesto(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EPresupuesto.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@login_required
@transaction.atomic
def eliminar_id_reformado(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EReformado.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.reformado',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.reformado')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
def descargar_plantilla_presupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	esquema_id= request.GET['id_esquema']	
				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)
	worksheet.set_column('I:I', 20)

	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)
	worksheet.write('G1', 'Valor M.O', format1)
	worksheet.write('H1', 'Valor Material', format1)
	worksheet.write('I1', 'Valor Global', format1)

	for item in hitos:

		if item.nivel==1:
			actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id,padre=item.id)

			for item2 in actividades:

				worksheet.write(row, col, item2.id,format2)
				worksheet.write(row, col+1, item.nombre,format2)
				worksheet.write(row, col+2,item2.nombre,format2)
				worksheet.write(row, col+3,'',format2)
				worksheet.write(row, col+4,'',format2)
				worksheet.write(row, col+5,'',format2)
				worksheet.write(row, col+6,'',format2)
				worksheet.write(row, col+7,'',format2)
				worksheet.write(row, col+8,'',format2)

				row +=1


	workbook.close()

	return response
    #return response


@login_required
def guardar_presupuesto_archivo(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte,data_only = True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		revisar_detalle=FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


		if len(revisar_detalle) > 0:
			FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).delete()


		contador=hoja.max_row - 1
		numeroFila = 1
		if int(contador) > 0:
			
			for fila in hoja.rows:				
				if numeroFila > 1:
					
					if (fila[0].value == None or fila[3].value == None or \
					fila[4].value == None or fila[6].value == None or \
					fila[7].value == None or fila[8].value == None or \
					fila[5].value == None):
						return JsonResponse({
							'message':'Se encontraron celdas vacias en la fila ' + str(numeroFila),
							'success':'error',
							'data':''})
					else:
						##import pdb; pdb.set_trace()
						if UnidadConstructiva.objects.filter(
							catalogo_id=request.POST['catalogoUnidadConstructiva_id'],
							codigo=fila[3].value).count() == 0:
							import pdb; pdb.set_trace()
							return JsonResponse({
								'message':'La unidad constructiva en la fila ' + str(numeroFila) + \
								' no se encuentra registrada en el catalogo seleccionado.',
								'success':'error',
								'data':''})

				numeroFila = numeroFila + 1	
									

			
			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						count=0
						j=0
						for fila2 in hoja.rows:
							if fila2[0].value:
								if j == 0:
									j=1
								else:
									if int(fila[0].value)==int(fila2[0].value):
										count = count+1

						
						actividad_porce=CEsquemaCapitulosActividadesG.objects.get(pk=fila[0].value)
						porcentaje=float(actividad_porce.peso)/count

						detalle=FDetallePresupuesto(porcentaje=porcentaje,
							presupuesto_id=presupuesto_id,
							actividad_id=fila[0].value,
							codigoUC=fila[3].value,
							descripcionUC=fila[4].value,
							valorManoObra=fila[6].value,
							valorMaterial=fila[7].value,
							valorGlobal=fila[8].value,
							cantidad=fila[5].value,
							catalogoUnidadConstructiva_id=request.POST['catalogoUnidadConstructiva_id'])
						detalle.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=detalle.id)
						logs_model.save()

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.detalle_presupuesto')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def actualizar_cantidad(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.actualizar_cantidad')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def cierre_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()


		presupuesto=EPresupuesto.objects.get(pk=respuesta['id_presupuesto'])
		presupuesto.cerrar_presupuesto=True
		presupuesto.save()	
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.presupuesto',id_manipulado=respuesta['id_presupuesto'])
		logs_model.save()	

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.cierre_presupuesto')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de un diagrama
@transaction.atomic
def eliminar_diagrama(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				DiagramaGrahm.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.diagrama',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.diagrama')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un diagrama	


@login_required
def informe_diagrama(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_diagrama.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Diagrama Grannt')
	format1=workbook.add_format({'font_size':10,'bold':True})

	

	cronograma_id= request.GET['cronograma_id']	
	
	cronograma=Cronograma.objects.get(pk=cronograma_id)				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=cronograma.esquema.id,nivel=1)

	worksheet.set_column('A:C', 25)

	cell_format = workbook.add_format()

	cell_format.set_pattern(1)  # This is optional when using a solid fill.
	cell_format.set_bg_color('green')

	#worksheet.write('A1', 'Id', format1)

	fecha_menor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id).order_by('fechaInicio').first()	

	fecha_mayor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id).order_by('fechaFinal').last()	

	date_nueva=fecha_mayor.fechaFinal-fecha_menor.fechaInicio
	semanas=math.ceil(date_nueva.days/7.0)

	worksheet.write(1, 0,'Capitulos / Actividades' ,format1)
	worksheet.write(1, 1, 'Fecha Inicio',format1)
	worksheet.write(1, 2, 'Fecha Final',format1)

	row=1
	col=3
	for i in range(int(semanas)):
		worksheet.write(row, col, "Semana "+str(i+1),format1)
		col +=1

	row=2
	col=0

	for item in hitos:

		actividades=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id)

		if len(actividades)>0:
			worksheet.write(row, col, item.nombre,format1)

			fechacapitulo_menor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id).order_by('fechaInicio').first()
			fechacapitulo_mayor=DiagramaGrahm.objects.filter(cronograma_id=cronograma_id,actividad__padre=item.id).order_by('fechaFinal').last()

			semana_inicio=0
			semana_final=0
			fecha_semana=fecha_menor.fechaInicio
			sw1=0
			sw2=0
			sw3=0
			date_nueva2=fechacapitulo_mayor.fechaFinal-fechacapitulo_menor.fechaInicio
			dias=math.floor(date_nueva2.days/7.0)

			for i in range(int(semanas)):

				if fechacapitulo_menor.fechaInicio == fechacapitulo_mayor.fechaFinal and sw2==0:
					semana_inicio=i+1
					semana_final=i+2
					sw2=1
					sw1=1
					sw3=1

				if fechacapitulo_menor.fechaInicio >=  fecha_semana and sw1==0:
					semana_inicio=i+1

					# if item2.fechaFinal <=  fecha_semana and sw2==0:
					# 	semana_final=i+1
					# 	sw2=1

				fecha_semana=fecha_semana + timedelta(days=7)

			if sw3==0:
				semana_final=semana_inicio+int(dias)+1

				
			for i in range(semana_inicio,semana_final):
				colu=col+2+semana_inicio
				worksheet.write(row, colu, "",cell_format)
				semana_inicio=semana_inicio+1


			row +=1		

			for item2 in actividades:

				worksheet.write(row, col, item2.actividad.nombre)
				worksheet.write(row, col+1, str(item2.fechaInicio))
				worksheet.write(row, col+2, str(item2.fechaFinal))

				semana_inicio=0
				semana_final=0
				fecha_semana=fecha_menor.fechaInicio
				sw1=0
				sw2=0
				sw3=0

				date_nueva2=item2.fechaFinal-item2.fechaInicio
				dias=math.floor(date_nueva2.days/7.0)

				for i in range(int(semanas)):

					if fechacapitulo_menor.fechaInicio == fechacapitulo_mayor.fechaFinal and sw2==0:
						semana_inicio=i+1
						semana_final=i+2
						sw2=1
						sw1=1
						sw3=1

					if item2.fechaInicio >=  fecha_semana and sw1==0:
						semana_inicio=i+1

					# if item2.fechaFinal <=  fecha_semana and sw2==0:
					# 	semana_final=i+1
					# 	sw2=1

					fecha_semana=fecha_semana + timedelta(days=7)

				if sw3==0:
					semana_final=semana_inicio+int(dias)+1

								
				for i in range(semana_inicio,semana_final):
					colu=col+2+semana_inicio
					worksheet.write(row, colu, "",cell_format)
					semana_inicio=semana_inicio+1


				row +=1


		

	workbook.close()

	return response
    #return response


@login_required
def descargar_plantilla_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', 'Apoyo', format1)
	worksheet.write('B1', 'Longitud', format1)
	worksheet.write('C1', 'Latitud', format1)


	workbook.close()

	return response
    #return response



@login_required
def guardar_apoyo_archivo(request):
	
	sid = transaction.savepoint()
	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:			
			nombreNodos = []
			for fila in hoja.rows:
				if fila[0].value:
					
					if nombreNodos:
						if fila[0].value not in nombreNodos:
							nombreNodos.append(fila[0].value)
						else:
							transaction.savepoint_rollback(sid)
							return JsonResponse({'message':'El nodo ' + fila[0].value + ' se encuentra repetido, por favor corregir y volver a intentarlo.','success':'error',
								'data':''})							
					else:
						nombreNodos.append(fila[0].value)
					
					nodo_data = HNodo.objects.filter(nombre=fila[0].value, presupuesto_id=presupuesto_id).values('id')
					if i == 0:
						i=1
					else:

						longitud = str(fila[1].value)
						if "," in longitud:
							longitud = float(longitud.replace(",","."))
						
						latitud = str(fila[2].value)
						if "," in latitud:
							latitud = float(latitud.replace(",","."))

						if nodo_data:							
							nodo_act = HNodo.objects.get(pk=nodo_data[0]['id'])
							nodo_act.longitud = longitud
							nodo_act.latitud = latitud
							nodo_act.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=nodo_data[0]['id'])
							logs_model.save()							
						else:
							nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=longitud,latitud=latitud,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
							nodo.save()							
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=nodo.id)
							logs_model.save()
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:		
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de varios apoyos
@transaction.atomic
def eliminar_apoyos(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				HNodo.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de varios apoyos



@login_required
def descargar_plantilla_apoyo_sinposicion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', '*Los campos latitud y longitud no son campos obligatorio*', format1)
	worksheet.write('A2', 'Apoyo', format1)
	worksheet.write('B2', 'Latitud', format1)
	worksheet.write('C2', 'Longitud', format1)

	workbook.close()

	return response
    #return response


	
@login_required
@transaction.atomic
def guardar_apoyo_archivo_sinposicion(request):

	sid = transaction.savepoint()
	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0 or i==1:
						i=i+1
					else:
						veri_nodo=HNodo.objects.filter(presupuesto_id=presupuesto_id,nombre=fila[0].value)

						if len(veri_nodo) == 0:
							latitud=None
							longitud=None

							if fila[1].value is not None and  fila[2].value is not None:
								latitud=fila[1].value
								longitud=fila[2].value

							nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=longitud,latitud=latitud,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
							nodo.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.nodo',id_manipulado=nodo.id)
							logs_model.save()
						else:
							transaction.savepoint_rollback(sid)
							return JsonResponse({'message':'No se puede repetir el nombre del nodo en este presupuesto','success':'error',
							'data':''})

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def descargar_plantilla_cantidadApoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_cantidadApoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cantidad_Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:Z', 20)


	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)

	apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	row=0
	col=6

	for item_apoyo in apoyos:
		worksheet.write(row,col, item_apoyo.nombre, format1)

		col +=1


	row=1
	col=0

	hitos = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


	for item in hitos:
		worksheet.write(row, col, item.id,format2)
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)
		worksheet.write(row, col+1,str(capitulo.nombre),format2)
		worksheet.write(row, col+2,item.actividad.nombre,format2)
		worksheet.write(row, col+3,item.codigoUC,format2)	
		worksheet.write(row, col+4,item.descripcionUC,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		count=6
		for item_apoyo in apoyos:
			worksheet.write(row,col+count,0, format2)
			count +=1

		row +=1


	workbook.close()

	return response
    #return response


@login_required
def guardar_cantidadApoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1
		numeroFila = 1
		apoyos = []
		if int(contador) > 0:
			sid = transaction.savepoint()
			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						if FDetallePresupuesto.objects.filter(id=fila[0].value).count() == 0:

							cantidadesNodoGuardadas = JCantidadesNodo.objects.filter(
								detallepresupuesto__presupuesto__id=presupuesto_id)
							cantidadesNodoGuardadas.delete()							

							return JsonResponse({
								'message':'No se encontro el Id indicado en la fila No.' + str(numeroFila) + \
								'. Sugerimos descargar la plantilla para verificar el Id en esta fila',
								'success':'error',
								'data':''})
						elif UnidadConstructiva.objects.filter(
							catalogo_id=FDetallePresupuesto.objects.get(id=fila[0].value).catalogoUnidadConstructiva,
							codigo=fila[3].value).count() == 0:

							cantidadesNodoGuardadas = JCantidadesNodo.objects.filter(
								detallepresupuesto__presupuesto__id=presupuesto_id)
							cantidadesNodoGuardadas.delete()

							return JsonResponse({
								'message':'No se encontro la UUCC indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos descargar la plantilla para verificar el codigo de UUCC en esta fila',
								'success':'error',
								'data':''})
						else:
							#codigo para guardar las cantidades:
							if len(apoyos) == 0:
								apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
							
							count = 6
							for item_apoyo in apoyos:
								consultar_nodo=JCantidadesNodo.objects.filter(
									detallepresupuesto_id=fila[0].value,
									nodo_id=item_apoyo.id)

								if consultar_nodo.count() == 0:
									
									nodo=JCantidadesNodo(
										detallepresupuesto_id=fila[0].value,
										nodo_id=item_apoyo.id,cantidad=fila[count].value)
									nodo.save()
									logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=nodo.id)
									logs_model.save()

								else:
									nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
									nodo.cantidad=fila[count].value
									nodo.save()
									logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
									logs_model.save()
								count +=1



					numeroFila = numeroFila + 1		
			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
			# i=0
			# apoyos=[]
			# for fila in hoja.rows:
			# 	if fila[0].value:
			# 		if i == 0:
			# 			i=1
			# 		else:
			# 			######################################################################################
			# 			if len(apoyos) == 0:
			# 				apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')

			# 			count=6
			# 			for item_apoyo  in apoyos:
						
			# 				consultar_nodo=JCantidadesNodo.objects.filter(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id)
			# 				if len(consultar_nodo)==0:
								
			# 						nodo=JCantidadesNodo(
			# 							detallepresupuesto_id=fila[0].value,
			# 							nodo_id=item_apoyo.id,cantidad=fila[count].value)
			# 						nodo.save()
			# 						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=nodo.id)
			# 						logs_model.save()

			# 				else:
			# 					nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
			# 					nodo.cantidad=fila[count].value
			# 					nodo.save()
			# 					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
			# 					logs_model.save()
			# 				count +=1
			# transaction.savepoint_commit(sid)
			# return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
			# 	'data':''})
		else:
			return JsonResponse({
				'message':'No se encontraron datos en el archivo cargado.',
				'success':'error',
				'data':''})			
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def informe_cantidad_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_cantidad_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Informe')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	presupuesto_id= request.GET['presupuesto_id']	
				
	detalle_presupuesto = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 30)

	worksheet.write('A1', 'Hitos', format1)
	worksheet.write('B1', 'Actividad', format1)
	worksheet.write('C1', 'Cod. UUCC', format1)
	worksheet.write('D1', 'Descripcion UUCC', format1)
	worksheet.write('E1', 'Cantidad Total', format1)
	worksheet.write('F1', 'Cantidad en Apoyos', format1)

	for item in detalle_presupuesto:

		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col, str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)
		worksheet.write(row, col+3,item.descripcionUC,format2)
		worksheet.write(row, col+4,item.cantidad,format2)

		porcentaje=JCantidadesNodo.objects.filter(detallepresupuesto_id=item.id).aggregate(Sum('cantidad'))
		porcentaje['cantidad__sum']=0 if porcentaje['cantidad__sum']==None else porcentaje['cantidad__sum']		

		worksheet.write(row, col+5,str(porcentaje['cantidad__sum']),format2)

		row +=1


	workbook.close()

	return response
    #return response


@transaction.atomic
@api_view(['POST',])
def guardar_cantidad_apoyo(request):

	sid = transaction.savepoint()
	try:
		respuesta=request.data
		#respuesta= json.loads(lista)
		##import pdb; pdb.set_trace()											
		#respuesta = json.loads(request.data['_content']) if request.data['_content']  else request.data		

		today = date.today()			
		uucc = FDetallePresupuesto.objects.get(pk=int(respuesta['detalle_presupuesto_id']))
		array = []				
		for item in respuesta['lista']:
			nodo=JCantidadesNodo.objects.get(pk=item['id'])
			if(float(nodo.cantidad) != float(item['cantidad'])):					
				array.append([{'cantidad_nueva':float(item['cantidad']),
							'apoyo':nodo.nodo.id,
							'cantidad_anterior':nodo.cantidad}])				
			nodo.cantidad=item['cantidad']
			nodo.save()				
						
		if array:
			reformado = EReformado(fecha_registro = today.strftime("%Y-%m-%d"), 
								usuario_registro_id = request.user.usuario.id,									
								)
			reformado.save()

			##import pdb; pdb.set_trace()
			for x in array:
				reformado_detalle = EReformadoDetalle(
									apoyo_id = int(x[0]['apoyo']),
									reformado_id = reformado.id,
									codigo_uucc = uucc.codigoUC,
									descripcion_uucc = uucc.descripcionUC,
									cantidad_anterior = float(x[0]['cantidad_anterior']),
									cantidad_final = float(x[0]['cantidad_nueva']),
									diferencia = float(x[0]['cantidad_nueva']) - float(x[0]['cantidad_anterior'])
									)			
				reformado_detalle.save()				
			

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cantidad_apoyo',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return Response({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@api_view(['GET',])
def consultar_avance_obra(request):
	#sid = transaction.savepoint()
	try:
		id_cronograma=request.GET['cronograma_id']
		presupuesto_id=request.GET['presupuesto_id']
		programando=request.GET['programando']
		dato=request.GET['dato']
		# estados=request.GET['id_estados']
		# listado_estado=estados.split(',')
		# ejecucion=request.GET['ejecucion']


		# listado_id=[]
		# cambio=LCambio.objects.filter(cronograma_id=id_cronograma,estado_id__in=listado_estado)

		# listado_id_ejecucion=[]
		# if int(ejecucion)>0:

		# 	listado_por=[]
		# 	if int(ejecucion) == 2:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__gte=100)
				
		# 	else:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__lt=100)

		# 	for obj in listado_por:
		# 		listado_id_ejecucion.append(obj.apoyo.id)

		# if cambio is not None:
		# 	for obj in cambio:
		# 		for item in obj.nodos.all():
		# 			listado_id.append(item.id)

		
		# cronograma=KCronograma.objects.get(pk=id_cronograma)
		
		

		# if len(listado_id_ejecucion)==0:
		# 	qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

		# if len(listado_id_ejecucion)>0:
		# 	qset=qset&(Q(id__in=listado_id_ejecucion))
		qset=(Q(presupuesto_id=presupuesto_id)&(Q(eliminado=False)))


		if dato!='':
			qset=qset&(Q(nombre__icontains=dato))

		listado=HNodo.objects.filter(qset).values(
			'id','nombre','longitud','latitud','capa__color',
			'porcentajeAcumulado')

		qset1=Q(nodoOrigen__presupuesto__id=presupuesto_id)

		# if len(listado_id)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id))

		# if len(listado_id_ejecucion)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id_ejecucion))

		listado_enlaces=IEnlace.objects.filter(qset1).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')

		#transaction.savepoint_commit(sid)
		return Response({'message':'','success':'ok',
				'data':{'datos':list(listado),'enlace':list(listado_enlaces)}})
		
	except Exception as e:
		#print(e)
		#transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#@login_required
#@transaction.atomic
@api_view(['GET',])
def consultar_ingresos_datos(request):
	#sid = transaction.savepoint()
	try:
		nodo_id=request.GET['nodo_id']
		reporte_id=request.GET['reporte_id']

		nodo=HNodo.objects.get(pk=nodo_id)
		datos=[]

		if nodo.noProgramado == False:
			listado=[]
			#print (nodo.presupuesto.sin_poste)
			if nodo.presupuesto.sin_poste == False:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id,cantidad__gt=0)
			else:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id)

			uucc = []
			materiales = []
			for item in listado:
				cantidad_ejecutada=KDetalleReporteTrabajo.objects.filter(
					nodo_id=nodo_id,
					detallepresupuesto_id=item.detallepresupuesto.id
					).aggregate(Sum('cantidadEjecutada'))
				cantidad_ejecutada['cantidadEjecutada__sum']=0 if cantidad_ejecutada['cantidadEjecutada__sum']==None else cantidad_ejecutada['cantidadEjecutada__sum']
				editable = False
				if cantidad_ejecutada['cantidadEjecutada__sum'] < item.cantidad:
					editable = True
				uucc.append({
							'id_detalle':item.detallepresupuesto.id,
							'codigo':item.detallepresupuesto.codigoUC,
							'descripcion':item.detallepresupuesto.descripcionUC,
							'cantidad':item.cantidad,
							'cantidad_ejecutada':cantidad_ejecutada['cantidadEjecutada__sum'],
							'editable': editable
					})
				#Extraer los materiales:
				uc = UnidadConstructiva.objects.filter(
					codigo=item.detallepresupuesto.codigoUC,
					catalogo__id=item.detallepresupuesto.catalogoUnidadConstructiva.id).values('id')

				desgloceMat = DesgloceMaterial.objects.filter(
					unidadConstructiva__id=uc[0]['id']).values(
					'material__codigo','material__descripcion', 'cantidad')


				for obj in desgloceMat:
					materiales.append({
						'codigo' : obj['material__codigo'],
						'descripcion' : obj['material__descripcion'],
						'cantidad' : float(obj['cantidad']) * float(item.cantidad),
						'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['cantidadEjecutada__sum'])
					})
				#unificar cantidades de materialales:
				materialesAgrupados=[]
				
				for mat in materiales:
					agregar = True
					for obj in materialesAgrupados:
						if mat['codigo'] == obj['codigo']:
							obj['cantidad']	+= mat['cantidad']
							obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
							agregar = False
					if agregar:		
						materialesAgrupados.append(mat)				

			#Extraer datos para grafica del nodo:
			
			porcentajes = []
			queryset_a_ejecutar = JCantidadesNodo.objects.filter(
				nodo__id=nodo_id,cantidad__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre').annotate(total=Sum('cantidad'))			

			queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
				nodo__id=nodo_id,cantidadEjecutada__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre'
				).annotate(total=Sum('cantidadEjecutada'))

			

			for Aejecutar in queryset_a_ejecutar:
				agregado = False
				for ejecutado in queryset_ejecutada:
					if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
						agregado = True
						porcentajes.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
						})
					if agregado:
						break
				if agregado == False:
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': 0
					})	

			
				

			if listado:
				datos={
					'unidadesConstructivas': uucc,
					'materiales': materialesAgrupados,
					'porcentajeEjecucion': porcentajes
				}
			else:
				datos={
					'unidadesConstructivas': uucc,
					'materiales': [],
					'porcentajeEjecucion': porcentajes
				}	
		
		#transaction.savepoint_commit(sid)
		return Response({'message':'','success':'ok',
				'data':datos})
		
	except Exception as e:
		print(e)
		#transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@api_view(['POST',])
@transaction.atomic
def guardar_cambio_cantidades(request):
	sid = transaction.savepoint()
	try:
		# Visbal:
		respuesta=request.data
		# respuesta=request.data

		# # respuesta= json.loads(lista)
		# ##import pdb; pdb.set_trace()
		# for item in respuesta['lista']:

		# Didi:
		# if request.data['_content']:
			
		# lista=request.data['_content']
		# respuesta= json.loads(lista)

		#respuesta = json.loads(request.data['_content']) if request.data['_content']  else request.data
		##import pdb; pdb.set_trace()
		for item in respuesta['lista']:
			
			qset_cantidadAejecutar = JCantidadesNodo.objects.filter(
				detallepresupuesto_id= item['id_detalle'],
				nodo_id=respuesta['id_nodo']
				).values('cantidad')
			cantidadAejecutar = float(qset_cantidadAejecutar[0]['cantidad'])

			qset_cantidadEjecutada = KDetalleReporteTrabajo.objects.filter(
				detallepresupuesto_id=item['id_detalle'],
				nodo_id=respuesta['id_nodo']
				).values('detallepresupuesto_id').annotate(
				cantidades=Sum('cantidadEjecutada')).distinct()

			if qset_cantidadEjecutada:	
				cantidadEjecutada = float(qset_cantidadEjecutada[0]['cantidades'])
			else:
				cantidadEjecutada = 0

			if (cantidadEjecutada + float(item['cantidad'])) <= cantidadAejecutar:

				cambio=KDetalleReporteTrabajo(
					reporte_trabajo_id=respuesta['id_reporte'],
					nodo_id=respuesta['id_nodo'],
					detallepresupuesto_id=item['id_detalle'],
					cantidadEjecutada=item['cantidad'])
				cambio.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,
					accion=Acciones.accion_crear,
					nombre_modelo='avance_de_obra_grafico2.detalle_reporte',
					id_manipulado=cambio.id)
				logs_model.save()

			else:
				dp = FDetallePresupuesto.objects.get(pk=item['id_detalle'])
				transaction.savepoint_rollback(sid)
				return Response({
					'message':'La cantidad a reportar en la unidad constructiva '+ dp.descripcionUC + \
					' sobrepasa la cantidad a ejecutar registrada en el sistema para este apoyo.',
					'success':'error',
					'data':''})



		reporte=JReporteTrabajo.objects.get(pk=respuesta['id_reporte'])
		
		if reporte.presupuesto.sin_poste==False:
			porcentajes = []
			queryset_a_ejecutar = JCantidadesNodo.objects.filter(
				nodo__id=respuesta['id_nodo'],cantidad__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre').annotate(total=Sum('cantidad'))			

			queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
				nodo__id=respuesta['id_nodo'],cantidadEjecutada__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre'
				).annotate(total=Sum('cantidadEjecutada'))

			

			for Aejecutar in queryset_a_ejecutar:
				agregado = False
				for ejecutado in queryset_ejecutada:
					if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
						agregado = True
						pr = (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
						if pr > 100:
							pr=100
						porcentajes.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							#'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': pr
						})
					if agregado:
						break
				if agregado == False:
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						#'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': 0
					})
			total=0
			for p in porcentajes:
				total = total + float(p['porcentaje'])

			########################################################################################
			# detalle=KDetalleReporteTrabajo.objects.filter(
			# 	reporte_trabajo_id=respuesta['id_reporte'],
			# 	nodo_id=respuesta['id_nodo']).values(
			# 	'detallepresupuesto_id').annotate(cantidades=Sum('cantidadEjecutada')).distinct()

			# total=0
			# por=0
			# for item2 in detalle:
			# 	cantidad_nodo=JCantidadesNodo.objects.filter(
			# 		detallepresupuesto_id=item2['detallepresupuesto_id'],
			# 		nodo_id=respuesta['id_nodo'])
			# 	valor=float(item2['cantidades'])/float(cantidad_nodo[0].cantidad)
			# 	por=por+valor

			# total=(por/len(detalle))*100

			# if total>100:
			# 	total=100
			########################################################################################
			porcentaje_apoyo=HNodo.objects.get(pk=respuesta['id_nodo'])
			porcentaje_apoyo.porcentajeAcumulado=round((float(total)/len(porcentajes)),2)
			porcentaje_apoyo.save()


		#createAsyncEstado.delay(respuesta['id_cronograma'])

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.cambio_ejecutada')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@api_view(['POST',])
@transaction.atomic
def guardar_cambio_detalle(request):
	sid = transaction.savepoint()
	try:
		respuesta=request.data
		#respuesta= json.loads(lista)


		for item in respuesta['lista']:
			detalle=KDetalleReporteTrabajo.objects.get(pk=item['id'])
			detalle.cantidadEjecutada=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cambio_detalle',id_manipulado=detalle.id)
			logs_model.save()



		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return Response({'message':'El registro ha sido registrado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.cambio')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_id_nodo_destino(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			IEnlace.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avance_de_obra_grafico.enlance',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@login_required
@transaction.atomic
def cierre_programacion(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cronograma=Cronograma.objects.get(id=respuesta['id_cronograma'])
		cronograma.programacionCerrada=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cronograma',id_manipulado=respuesta['id_cronograma'])
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def informe_detallepresupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']
	presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:H', 20)


	worksheet.write('A2', 'Presupuesto', format1)
	worksheet.write('B2', presupuesto.nombre, format1)

	worksheet.write('A3', 'Proyecto', format1)
	worksheet.write('B3', presupuesto.cronograma.proyecto.nombre, format1)

	worksheet.write('A4', 'Esquema', format1)
	worksheet.write('B4', presupuesto.cronograma.esquema.nombre, format1)

	worksheet.write('F6', 'Total Presupuesto', format1)

	worksheet.write('A7', 'Hitos', format1)
	worksheet.write('B7', 'Actividad', format1)
	worksheet.write('C7', 'codigo UUCC', format1)
	worksheet.write('D7', 'Descripcion UUCC', format1)
	worksheet.write('E7', 'Valor UUCC', format1)
	worksheet.write('F7', 'Cantidad', format1)
	worksheet.write('G7', 'Subtotal', format1)

	row=7
	col=0

	detalle = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)
	total=0

	for item in detalle:
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col,str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)	
		worksheet.write(row, col+3,item.descripcionUC,format2)

		valorManoObra=0
		if item.valorManoObra is not None:
			valorManoObra=item.valorManoObra

		valorMaterial=0
		if item.valorMaterial is not None:
			valorMaterial=item.valorMaterial

		valor=(valorManoObra)+(valorMaterial)

		worksheet.write(row, col+4,valor,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		subtotal=valor*(item.cantidad)

		worksheet.write(row, col+6,subtotal,format2)

		total=total+subtotal

		row +=1


	worksheet.write('G6', total, format1)

	workbook.close()

	return response
    #return response


@login_required
def menu_gps(request,id_presupuesto):

	try:
		nodo=HNodo.objects.filter(presupuesto_id=id_presupuesto)
		
		if len(nodo) == 0:	
			return JsonResponse({'message':'No tiene ningun punto de gps registrado','success':'ok',
					'data':['1']})
		else:
			sqlnodo=HNodo.objects.filter(presupuesto_id=id_presupuesto,latitud=None,longitud=None)

			if len(sqlnodo) == 0:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['2']})
			else:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['3']})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def guardar_reporte_trabajo(request,id_reporte):

	sid = transaction.savepoint()
	try:
		reporte=JReporteTrabajo.objects.get(pk=id_reporte)
		#detalle=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)
		detalles=FDetallePresupuesto.objects.filter(presupuesto_id=reporte.presupuesto.id)

		# model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
		# valor=round(float(model_actividad['peso__sum']),3)+round(float(request.data['peso']),3)
		##import pdb; pdb.set_trace()
		detalle_reporte=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)

		if len(detalle_reporte) == 0:
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Debe ingresar alguna cantidad ejecutada en el reporte','success':'error',
			'data':''})

		total_porcentaje=0
		total_pagado=0
		for item in detalles:
			sql_detalletrabajo=KDetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte,detallepresupuesto_id=item.id).aggregate(Sum('cantidadEjecutada'))
			sql_detalletrabajo['cantidadEjecutada__sum']=0 if sql_detalletrabajo['cantidadEjecutada__sum']==None else sql_detalletrabajo['cantidadEjecutada__sum']
			#valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])+float(item.cantidad)
			valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])/float(item.cantidad)
			item.porcentaje=0 if item.porcentaje==None else item.porcentaje
			
			#porcentaje=(valor*float(item.porcentaje))/100
			porcentaje=(valor*float(item.porcentaje))
			total_porcentaje=total_porcentaje+porcentaje

			valor2=float(item.valorManoObra)+float(item.valorMaterial)
			pagado=float(sql_detalletrabajo['cantidadEjecutada__sum'])*float(valor2)
			total_pagado=total_pagado+pagado

		if total_porcentaje>100:
			total_porcentaje=100

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		estado_rechazo=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		estado_corregido=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)

		reporte.valor_ganando_acumulado=round(total_pagado,3)
		reporte.avance_obra_acumulado=round(total_porcentaje,3)
		reporte.reporteCerrado=True

		if int(reporte.estado_id)==int(estado_rechazo[0].id):			
			reporte.estado_id=estado_corregido[0].id
		else:
			reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.reporte',id_manipulado=id_reporte)
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.cierre_programacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#serializer de empresa
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')



#serializer lite de proyecto empresa
class ProyectoEmpresaLite2Serializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLite2Serializer(read_only = True )
	empresa = EmpresaLiteSerializer(read_only = True  )

	totalCorrregidos=serializers.SerializerMethodField()

	totalAprobados=serializers.SerializerMethodField()

	totalRegistrados=serializers.SerializerMethodField()

	totalRechazados=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id' ,
				'proyecto',
 				'empresa',
 				'totalCorrregidos',	
 				'totalAprobados',	
 				'totalRegistrados',
 				'totalRechazados'
				)

	def get_totalCorrregidos(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalAprobados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRegistrados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRechazados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		return JReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()


#Api lite de empresa proyecto
class ProyectoReporteAvanceViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las categorias de administrador de fotos
	"""
	model=Proyecto_empresas
	queryset = model.objects.all()
	serializer_class = ProyectoEmpresaLite2Serializer
	nombre_modulo='administrador_foto.Proyecto_empresa'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ProyectoReporteAvanceViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			empresa = self.request.query_params.get('empresa', None)
			mcontrato = self.request.query_params.get('mcontrato', None)
			departamento = self.request.query_params.get('departamento', None)
			municipio = self.request.query_params.get('municipio', None)
			contratista = self.request.query_params.get('contratista', None)
			corregido = self.request.query_params.get('corregido', None)
			registrado = self.request.query_params.get('registrado', None)
			aprobado = self.request.query_params.get('aprobado', None)
			rechazados = self.request.query_params.get('rechazados', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			qset=(~Q(id=0))

			if corregido is not None:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			if rechazados is not None:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			
			if registrado is not None and int(registrado)==1:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))

			if aprobado is not None and int(aprobado)==1:
				reporte=JReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))


			if(dato or empresa or mcontrato or departamento or municipio):

				if dato:
					qset = qset & ( Q(empresa__nombre__icontains = dato) |
								Q(empresa__nit__icontains = dato) |
								Q(proyecto__nombre__icontains = dato))					
			

				if empresa and int(empresa)>0:
					qset = qset & (Q(empresa__id = empresa))					


				if mcontrato and int(mcontrato)>0:
					qset = qset & (Q(proyecto__mcontrato = mcontrato))
					
				if departamento and int(departamento)>0:
					qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
					
				if municipio and int(municipio)>0:
					qset = qset & (Q(proyecto__municipio = municipio))
					

				if contratista and int(contratista)>0:
					qset = qset & (Q(proyecto__contrato__contratista__id = contratista))	

			
			if qset != '':
				queryset = self.model.objects.filter(qset).distinct()

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@login_required
@transaction.atomic
def guardar_archivo_aprobacion(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		reporte_id= request.POST['reporte_id']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		
		reporte=JReporteTrabajo.objects.get(pk=reporte_id)
		reporte.soporteAprobacion=soporte
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.soporte_aprobacion',id_manipulado=reporte_id)
		logs_model.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.soporte_aprobacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_rechazo_reporte(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		reporte_id= respuesta['reporte_id']
		mensaje= respuesta['mensaje']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		
		reporte=JReporteTrabajo.objects.get(pk=reporte_id)
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.reporteCerrado=False
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.rechazo_reporte',id_manipulado=reporte_id)
		logs_model.save()

		motivo=MComentarioRechazo(reporte_trabajo_id=reporte_id,motivoRechazo=mensaje)
		motivo.save()
		logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.motivo_rechazo',id_manipulado=motivo.id)
		logs_model2.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.rechazo_reporte')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def consultar_graficos(request):
	sid = transaction.savepoint()
	try:
		presupuesto_id=request.GET['presupuesto_id']
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		listado=[]
		listado_porcentaje=DiagramaGrahm.objects.filter(cronograma_id=presupuesto.cronograma.id).values('fechaFinal').annotate(porcentaje=Sum('actividad__peso')).distinct()
		print(listado_porcentaje)
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado.append({
					'fecha':item['fechaFinal'],
					'porcentaje':round(porcentaje,2)
					})

		print(listado)


		listado3=[]
		listado_porcentaje=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(porcentaje=Sum('avance_obra_acumulado')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado3.append({
					'fecha':item['fechaTrabajo'],
					'porcentaje':round(porcentaje,2)
					})

		listado4=[]
		listado_prespuesto=JReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(valor_ganando=Sum('valor_ganando_acumulado')).distinct()
				
		valor_ganando=0
		for item in listado_prespuesto:
			valor_ganando=valor_ganando+float(item['valor_ganando'])
			
			listado4.append({
					'fecha':item['fechaTrabajo'],
					'valor_ganando':round(valor_ganando,2)
					})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'programada':list(listado),'avance':list(listado3),'presupuesto':list(listado4)}})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.grafico')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def consultar_cantidad_reportes(request):

	try:

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		cantidad_corregidos=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		cantidad_registrado=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		cantidad_rechazados=JReporteTrabajo.objects.filter(estado__id=estados[0].id).count()
		

		return JsonResponse({'message':'','success':'ok',
					'data':{'cantidad_corregidos':cantidad_corregidos,'cantidad_registrado':cantidad_registrado,'cantidad_rechazados':cantidad_rechazados}})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avance_de_obra_grafico2.cantidad_reportes')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def actualizar_cancelado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
		for item in respuesta['lista']:
			cambio=LCambio.objects.get(id=item['id'])
			cambio.estado_id=estado[0].id
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.cambio',id_manipulado=item['id'])
			logs_model.save()

			historial=LHistorialCambio(cambio_id=item['id'],usuario_registro_id=request.user.usuario.id,motivoEstado=respuesta['motivo'],estado_id=estado[0].id)
			historial.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.historial_cambio',id_manipulado=historial.id)
			logs_model2.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.actualizar_cantidad')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
@transaction.atomic
def reportar_sin_poste(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		presupuesto_id= respuesta['presupuesto_id']
		
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		presupuesto.sin_poste=True
		presupuesto.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.sin_poste',id_manipulado=presupuesto_id)
		logs_model.save()

		# agregarSinPoste.delay(presupuesto_id,request.user.usuario.id)
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.sin_poste')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def avance_de_obra_grafico2(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico2/hitos.html',{'model':'besquemacapitulosg','app':'avanceObraGrafico2','macrocontrato':ListMacro})

@login_required
def catalogo(request):	
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)		
	return render(request,'avanceObraGrafico2/catalogo.html',{'model':'catalogounidadconstructiva','app':'avanceObraGrafico2','macrocontrato':ListMacro})

@login_required
def uucc(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraGrafico2/uucc.html',{'model':'unidadconstructiva','app':'avanceObraGrafico2', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def materiales(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraGrafico2/materiales.html',{'model':'material','app':'avanceObraGrafico2', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def mano_obra(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraGrafico2/manoObra.html',{'model':'manodeobra','app':'avanceObraGrafico2', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def desgloce_mat(request,uucc_id):
	#import pdb; pdb.set_trace()		
	uucc = UnidadConstructiva.objects.filter(id=uucc_id)
	uucc = uucc[0]
	if uucc.catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraGrafico2/desgloce_mat.html',{'model':'desglocematerial','app':'avanceObraGrafico2', 'uucc_id':int(uucc_id), 'catalogo': uucc.catalogo, 'catalogo_activo': catalogo_activo, 'uucc': uucc })

@login_required
def desgloce_mo(request,uucc_id):
	#import pdb; pdb.set_trace()		
	uucc = UnidadConstructiva.objects.filter(id=uucc_id)
	uucc = uucc[0]
	if uucc.catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraGrafico2/desgloce_mo.html',{'model':'desglocemanodeobra','app':'avanceObraGrafico2', 'uucc_id':int(uucc_id), 'catalogo': uucc.catalogo, 'catalogo_activo': catalogo_activo, 'uucc': uucc })


@login_required
def actividades(request,id_esquema):
	return render(request, 'avanceObraGrafico2/actividad.html',{'model':'cesquemacapitulosactividadesg','app':'avanceObraGrafico2','id_esquema':id_esquema})


@login_required
def regla_estado(request,id_esquema):
	return render(request, 'avanceObraGrafico2/regla.html',{'model':'epresupuesto','app':'avanceObraGrafico','id_esquema':id_esquema})


@login_required
def cronograma(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraGrafico2/cronograma.html',{'model':'cronograma','app':'avanceObraGrafico2','macrocontrato':ListMacro})



@login_required
def cronograma_proyecto(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	esquema=BEsquemaCapitulosG.objects.filter(macrocontrato_id=proyecto.mcontrato)
	periodo=APeriodicidadG.objects.all()
	return render(request, 'avanceObraGrafico2/cronograma_proyecto.html',{'model':'cronograma','app':'avanceObraGrafico2','id_proyecto':id_proyecto,'proyecto':proyecto,'esquema':esquema,'periodo':periodo})

@login_required
def tablero_contrato(request,id_mcontrato):
	mcontrato=Contrato.objects.get(pk=id_mcontrato)
	return render(request, 'avanceObraGrafico2/tablero_contrato.html',
		{'model':'cronograma',
		'app':'avanceObraGrafico2',
		'mcontrato':mcontrato})



@login_required
def presupuesto(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	return render(request, 'avanceObraGrafico2/presupuesto.html',{'model':'epresupuesto','app':'avanceObraGrafico2','cronograma':querycronograma,'cronograma_id':id_cronograma,'proyecto_id':querycronograma.proyecto.id})


@login_required
def presupuesto_detalle(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=presupuesto.cronograma.esquema.id,nivel=1)
	catalogos = CatalogoUnidadConstructiva.objects.filter(activo=True)
	return render(request, 'avanceObraGrafico2/detalle_presupuesto.html',
		{'model':'fdetallepresupuesto',
		'app':'avanceObraGrafico',
		'presupuesto_id':id_presupuesto,
		'presupuesto':presupuesto,
		'proyecto_id':presupuesto.cronograma.proyecto.id,
		'capitulos':capitulos,
		'cronograma_id':presupuesto.cronograma.id,
		'catalogos':catalogos})


@login_required
def programacion(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=querycronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico2/diagrama.html',{'model':'diagramagrahm','app':'avanceObraGrafico2','capitulos':capitulos,'cronograma':querycronograma,'cronograma_id':id_cronograma,'proyecto_id':querycronograma.proyecto.id})



@login_required
def apoyo_con_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/apoyo_con_gps.html',{'model':'hnodo','app':'avanceObraGrafico2','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def apoyo_sin_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/apoyo_sin_gps.html',{'model':'hnodo','app':'avanceObraGrafico2','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def cantidad_apoyo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.cronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraGrafico2/cantidad_apoyo.html',{'model':'epresupuesto','app':'avanceObraGrafico2','cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'capitulos':capitulos})


@login_required
def cantidad_apoyo_id(request,id_presupuesto,id_detalle):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	detalle_presupuesto=FDetallePresupuesto.objects.get(pk=id_detalle)
	return render(request, 'avanceObraGrafico2/cantidad_apoyo_id.html',{'model':'epresupuesto','app':'avanceObraGrafico2','presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'detalle_presupuesto':detalle_presupuesto,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'cronograma_id':nombre_presupuesto.cronograma.id})


@login_required
def reporte_trabajo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
	estados2=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)

	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	

				
	usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraGrafico2/reporte.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2','Usuarios':usuarios,'estado_id_procesado':estados2[0].id,'estado_id_registrado':estados[0].id,'cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto})

@login_required
def reformado(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
	estados2=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)

	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	
				
	usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	
	return render(request, 'avanceObraGrafico2/reformado.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2','Usuarios':usuarios,'estado_id_procesado':estados2[0].id,'estado_id_registrado':estados[0].id,'cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto})

@login_required
def reformadoDetalle(request,id_reformado):
	##import pdb; pdb.set_trace()
	reformado=EReformado.objects.get(pk=int(id_reformado))	
	detalle = EReformadoDetalle.objects.filter(reformado__id = int(reformado.id)).values('apoyo__presupuesto_id').distinct()
	detalle = detalle[0]['apoyo__presupuesto_id']
	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	
	
	return render(request, 'avanceObraGrafico2/reformadoDetalle.html',{'model':'ereformado','app':'avanceObraGrafico2','reformado':reformado, 'presupuesto_id': detalle})

@login_required
def avance_con_gps(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	valoraciones = Tipo.objects.filter(app='no_conformidad_valoracion')
	tipos = Tipo.objects.filter(app='no_conformidad')
	return render(request, 'avanceObraGrafico2/avance_con_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte,'tipos': tipos, 'valoraciones': valoraciones})


@login_required
def avance_sin_gps(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	estados_cambios=Estado.objects.filter(app='Avance_obra_grafico')
	return render(request, 'avanceObraGrafico2/avance_sin_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','estados':estados_cambios,'capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte})



@login_required
def aprobacion(request):
	return render(request, 'avanceObraGrafico2/aprobacion.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2'})


@login_required
def corregido(request):
	return render(request, 'avanceObraGrafico2/corregido.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2'})


@login_required
def registrado(request):
	return render(request, 'avanceObraGrafico2/registrado.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2'})


@login_required
def rechazados(request):
	return render(request, 'avanceObraGrafico2/rechazados.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2'})


@login_required
def reporte_trabajo_registrado(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_registrado.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_registrado(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraGrafico2/detalle_registrado.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def reporte_trabajo_corregido(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_corregido.html',{'model':'jreportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_corregido(request,id_reporte):
	reporte=JReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraGrafico2/detalle_corregido.html',{'model':'kdetallereportetrabajo','app':'avanceObraGrafico2','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def grafico(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraGrafico2/grafico.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'cronograma_id':presupuesto.cronograma.id,'presupuesto':presupuesto})


@login_required
def reporte_trabajo_rechazados(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraGrafico2/reporte_rechazados.html',{'model':'reportetrabajo','app':'avanceObraGrafico2','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def cambios(request):
	return render(request, 'avanceObraGrafico2/cambios.html',{'model':'lcambio','app':'avanceObraGrafico2'})


@login_required
def index_cambio(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	empresa=Empresa.objects.filter(~Q(id=request.user.usuario.empresa.id))
	estados=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=101)
	cancelado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
	return render(request, 'avanceObraGrafico2/index_cambio.html',{'model':'lcambio','app':'avanceObraGrafico2','estado_cancelado':cancelado[0].id,'estado_id':estados[0].id,'presupuesto_id':id_presupuesto,'presupuesto':presupuesto,'empresa':empresa})


@login_required
def detalle_cambio(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	return render(request, 'avanceObraGrafico2/detalle_cambio.html',{'model':'lcambio','app':'avanceObraGrafico2','cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def agregar_detalle(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	apoyos=HNodo.objects.filter(presupuesto_id=cambio.presupuesto.id)
	return render(request, 'avanceObraGrafico2/agregar_detalle.html',{'model':'lcambio','app':'avanceObraGrafico2','apoyos':apoyos,'cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def aprobacion_cambio(request):
	return render(request, 'avanceObraGrafico2/aprobacion_cambio.html',{'model':'lcambio','app':'avanceObraGrafico2'})


@login_required
def autorizacion_cambio(request):
	return render(request, 'avanceObraGrafico2/autorizacion_cambio.html',{'model':'lcambio','app':'avanceObraGrafico2'})




#API Manejo de unidades constructivas
#Tipos de unidades constructivas
from .models import TipoUnidadConstructiva
class TipoUnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model=TipoUnidadConstructiva
		fields=('id','nombre','activa')

class TipoUnidadConstructivaViewSet(viewsets.ModelViewSet):
	model=TipoUnidadConstructiva
	queryset = model.objects.all()
	serializer_class = TipoUnidadConstructivaSerializer
	nombre_modulo='avance_de_obra_grafico2.tipoUnidadConstructiva'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(TipoUnidadConstructivaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			activa = self.request.query_params.get('activa', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))

			if dato:
				qset = qset & (Q(nombre__icontains=dato))
			if activa:
				qset = 	qset & (Q(activa=activa))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#catalogos de unidades constructivas


class CatalogoUnidadConstructivaViewSet(viewsets.ModelViewSet):
	model = CatalogoUnidadConstructiva
	queryset = model.objects.all()
	serializer_class = CatalogoUnidadConstructivaSerializer
	nombre_modulo='avance_de_obra_grafico2.catalogoUnidadConstructiva'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:	
			#import pdb; pdb.set_trace()		
			queryset = super(CatalogoUnidadConstructivaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			activo = self.request.query_params.get('activo', None)
			ano = self.request.query_params.get('ano', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''


			qsetMcontrato = (Q(empresa_id=request.user.usuario.empresa.id))
			listContratos = EmpresaContrato.objects.filter(qsetMcontrato).values('contrato_id').distinct()
			qset = (~Q(id=0))
			if request.user.usuario.empresa.id != 4:
				qset = qset & (Q(mcontrato_id__in=listContratos))
			
			if dato:
				qset = qset & (Q(nombre__icontains=dato))
			if activo:
				qset = qset & (Q(activo=activo))
			if ano:
				qset = qset & (Q(ano=ano))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				validator = self.model.objects.filter(nombre=request.data['nombre'])

				if not validator:

					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(mcontrato_id=int(request.data['mcontrato_id']))
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
							
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un catalogo con el nombre suministrado.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				catalogo = self.model.objects.get(pk=int(request.data['id']))
				if catalogo:
					catalogo.ano = request.data['ano']
					catalogo.mcontrato_id = int(request.data['mcontrato_id'])
					catalogo.nombre = request.data['nombre']
					catalogo.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  
#unidades constructivas

class MaterialesViewSet(viewsets.ModelViewSet):
	model = Material
	queryset = model.objects.all()
	serializer_class = MaterialSerializer
	nombre_modulo='avance_de_obra_grafico2.materiales'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()			
			queryset = super(MaterialesViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			pk = self.request.query_params.get('pk', None)
			catalogo_id = self.request.query_params.get('catalogo', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			lite = self.request.query_params.get('lite',None)
			mensaje=''

			qset = (~Q(id=0))
			if catalogo_id:
				qset = qset & (Q(catalogo_id=catalogo_id))
			if dato:
				qset = qset & (Q(descripcion__icontains=dato) | Q(codigo__icontains=dato))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					if pk:
						queryset = 	self.model.objects.filter(id=int(pk))				
					serializer = MaterialSerializerLite(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)								
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo'])) & Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()
				if not validator:
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(catalogo_id=request.data['catalogo_id'],valorUnitario=request.data['valorUnitario'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)

				else:
					return Response({'message':'Ya existe un Material con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)						
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  

class ManoObraViewSet(viewsets.ModelViewSet):
	model = ManoDeObra
	queryset = model.objects.all()
	serializer_class = ManoObraSerializer
	nombre_modulo='avance_de_obra_grafico2.ManoDeObra'
	model_log=Logs
	model_acciones=Acciones	

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()			
			queryset = super(ManoObraViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			catalogo_id = self.request.query_params.get('catalogo', None)			
			lite = self.request.query_params.get('lite',None)
			pk = self.request.query_params.get('pk', None)
			mensaje=''

			qset = (~Q(id=0))
			if catalogo_id:
				qset = qset & (Q(catalogo_id=catalogo_id))
			if dato:
				qset = qset & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					if pk:
						queryset = 	self.model.objects.filter(id=int(pk))				
					serializer = ManoObraSerializerLite(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)								
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})								
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo']))& Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()
				if not validator:
					
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(catalogo_id=request.data['catalogo_id'],valorHora=request.data['valorHora'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un registro de Mano de Obra con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  

class UnidadConstructivaLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)

	class Meta:
		model=UnidadConstructiva
		fields=('id','catalogo','tipoUnidadConstructiva',
			'codigo','descripcion')

class UnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	totalManoDeObra = serializers.SerializerMethodField()
	totalMateriales = serializers.SerializerMethodField()

	class Meta:
		model=UnidadConstructiva
		fields=('id','catalogo','catalogo_id','tipoUnidadConstructiva','tipoUnidadConstructiva_id',
			'codigo','descripcion','totalManoDeObra','totalMateriales')

	def get_totalManoDeObra(self, obj):	
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'manoDeObra__valorHora','rendimiento')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		return round(total,2)

	def get_totalMateriales(self,obj):
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'material__valorUnitario','cantidad')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		return round(total,2)

class UnidadConstructivaSerializerLite(serializers.HyperlinkedModelSerializer):

	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	totalManoDeObra = serializers.SerializerMethodField()
	totalMateriales = serializers.SerializerMethodField()

	class Meta:
		model=UnidadConstructiva
		fields=('id','tipoUnidadConstructiva','tipoUnidadConstructiva_id',
			'codigo','descripcion','totalManoDeObra','totalMateriales')

	def get_totalManoDeObra(self, obj):	
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'manoDeObra__valorHora','rendimiento')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		return round(total,2)

	def get_totalMateriales(self,obj):
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'material__valorUnitario','cantidad')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		return round(total,2)

class DesgloceMaterialSerializer(serializers.HyperlinkedModelSerializer):
	
	material = MaterialSerializer(read_only=True)
	material_id = serializers.PrimaryKeyRelatedField(read_only=True)

	unidadConstructiva = UnidadConstructivaSerializer(read_only=True)
	unidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=DesgloceMaterial
		fields=('id','cantidad','material', 'material_id','unidadConstructiva','unidadConstructiva_id')		

class DesgloceMaterialViewSet(viewsets.ModelViewSet):
	model = DesgloceMaterial
	queryset = model.objects.all()
	serializer_class = DesgloceMaterialSerializer
	nombre_modulo='avance_de_obra_grafico2.DesgloceMaterial'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()		
			queryset = super(DesgloceMaterialViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			uucc_id = self.request.query_params.get('uucc_id', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))
			qset = qset & (Q(unidadConstructiva_id=uucc_id))
			
			if dato:
				qset = qset & (Q(material__descripcion__icontains=dato) | Q(material__codigo__icontains=dato)) 

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = (Q(material_id=request.data['material_id']) & Q(unidadConstructiva_id=request.data['uucc_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:				
					cantidad = request.data['cantidad']
					if ',' in cantidad:
						request.data['cantidad'] = cantidad.replace(',','.')

					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(material_id=request.data['material_id'], unidadConstructiva_id=request.data['uucc_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)

				else:
					return Response({'message':'Ya existe un Desglose de Material con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)						
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				data = self.model.objects.get(pk=int(request.data['id']))
				if data:
					cantidad = request.data['cantidad']
					if ',' in cantidad:
						request.data['cantidad'] = cantidad.replace(',','.')					
					data.material_id = int(request.data['material_id'])
					data.cantidad = request.data['cantidad']
					data.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)

				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				# if serializer.is_valid():
					# self.perform_update(serializer)

				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  

class DesgloceManoDeObraSerializer(serializers.HyperlinkedModelSerializer):
	
	manoDeObra = ManoObraSerializer(read_only=True)
	manoDeObra_id = serializers.PrimaryKeyRelatedField(read_only=True)

	unidadConstructiva = UnidadConstructivaSerializer(read_only=True)
	unidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=DesgloceManoDeObra
		fields=('id','rendimiento','manoDeObra', 'manoDeObra_id','unidadConstructiva','unidadConstructiva_id')

class DesgloceManoDeObraViewSet(viewsets.ModelViewSet):
	model = DesgloceManoDeObra
	queryset = model.objects.all()
	serializer_class = DesgloceManoDeObraSerializer
	nombre_modulo='avance_de_obra_grafico2.DesgloceManoDeObra'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()		
			queryset = super(DesgloceManoDeObraViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			uucc_id = self.request.query_params.get('uucc_id', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))
			qset = qset & (Q(unidadConstructiva_id=uucc_id))
			
			if dato:
				qset = qset & (Q(manoDeObra__descripcion__icontains=dato) | Q(manoDeObra__codigo__icontains=dato)) 

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = (Q(manoDeObra_id=request.data['manoDeObra_id']) & Q(unidadConstructiva_id=request.data['uucc_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:				
					rendimiento = request.data['rendimiento']
					if ',' in rendimiento:
						request.data['rendimiento'] = rendimiento.replace(',','.')				
					
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(manoDeObra_id=request.data['manoDeObra_id'], unidadConstructiva_id=request.data['uucc_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un Desglose de Mano de Obra con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)	

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				data = self.model.objects.get(pk=int(request.data['id']))
				if data:
					rendimiento = request.data['rendimiento']
					if ',' in rendimiento:
						request.data['rendimiento'] = rendimiento.replace(',','.')	

					data.manoDeObra_id = int(request.data['manoDeObra_id'])
					data.rendimiento = request.data['rendimiento']
					data.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)

				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				# if serializer.is_valid():
					# self.perform_update(serializer)

				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  


class UnidadConstructivaAOViewSet(viewsets.ModelViewSet):
	model = UnidadConstructiva
	queryset = model.objects.all()
	serializer_class = UnidadConstructivaSerializer
	nombre_modulo='avance_de_obra_grafico2.UnidadConstructiva'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(UnidadConstructivaAOViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			codigo = self.request.query_params.get('codigo', None)
			catalogo = self.request.query_params.get('catalogo', None)
			tipoUnidadConstructiva = self.request.query_params.get('tipoUnidadConstructiva', None)
			lite = self.request.query_params.get('lite', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))

			if dato:
				qset = qset & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))
			if catalogo:
				qset = qset & (Q(catalogo__id=catalogo))
			if tipoUnidadConstructiva:
				qset = qset & (Q(tipoUnidadConstructiva__id=tipoUnidadConstructiva))
			if codigo:
				qset = qset & (Q(codigo=codigo))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					serializer = UnidadConstructivaLiteSerializer(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = UnidadConstructivaLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})



		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo'])) & Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(
							catalogo_id = request.data['catalogo_id'],
							tipoUnidadConstructiva_id = request.data['tipoUnidadConstructiva_id']
						)
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un registro con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
				
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)	
#Permite obtener el APU de una Unidad constructiva
def get_apuUnidadConstructiva(request,idUnidadConstructiva):
	try:
		
		uc = UnidadConstructiva.objects.filter(
			id=idUnidadConstructiva).values(
			'id','codigo','descripcion')

		desgloceMo = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=idUnidadConstructiva).values(
			'id','manoDeObra__codigo','manoDeObra__descripcion',
			'manoDeObra__valorHora','rendimiento')
		
		totalMo = 0	
		if desgloceMo:	
			for row in desgloceMo:
				totalMo = float(totalMo) + \
				 (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		##import pdb; pdb.set_trace()
		desgloceMat = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=idUnidadConstructiva).values(
			'id','material__codigo','material__descripcion',
			'material__valorUnitario','cantidad')

		totalMat = 0		 
		if desgloceMat:
			for row in desgloceMat:
				totalMat = float(totalMat) + \
				 (float(row['material__valorUnitario']) * float(row['cantidad']))

		data = {
			'unidadConstructiva' : uc[0],
			'totalManoDeObra' : round(totalMo,2),
			'totalMateriales' : round(totalMat,2),
			'desgloceManoDeObra' : list(desgloceMo),
			'desgloceMateriales' : list(desgloceMat)
		}
		
		return JsonResponse({'message':'','success':'ok','data': data})
	except Exception as e:
		functions.toLog(e,'avanceObraGrafico2.getUnidadConstructiva')
		return Response(
			{'message':'Se presentaron errores de comunicacion con el servidor',
			'success':'fail','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#Permite obtener el listado de UUCC y materiales requeridas en un nodo con un presupuesto
def get_cantidadesDeNodo(request,nodo_id,presupuesto_id):
	unidadesConstructivas = []
	materiales = []
	try:
		
		condicion = Q(nodo__id=nodo_id)
		condicion = condicion & (~Q(cantidad=0))
		queryset = JCantidadesNodo.objects.filter(condicion).values(
			'detallepresupuesto__codigoUC','detallepresupuesto__catalogoUnidadConstructiva__id',
			'cantidad'
			)
		for row in queryset:
			querysetUUCC = UnidadConstructiva.objects.filter(
				catalogo__id=row['detallepresupuesto__catalogoUnidadConstructiva__id'],
				codigo=row['detallepresupuesto__codigoUC']).values('id','descripcion')
			for rowUC in querysetUUCC:
				unidadesConstructivas.append(
					{
						'id':rowUC['id'],
						'codigo':row['detallepresupuesto__codigoUC'],
						'descripcion':rowUC['descripcion'],
						'cantidad':row['cantidad'],
						'catalogo_id': row['detallepresupuesto__catalogoUnidadConstructiva__id']
					}
				)
		
		
		for unidad in unidadesConstructivas:
			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=unidad['id']).values(
				'material__codigo','material__descripcion','cantidad')		
			for mat in desgloceMat:
				materiales.append({
					'codigo': mat['material__codigo'],
					'descripcion': mat['material__descripcion'],
					'cantidad': float(mat['cantidad']) * float(unidad['cantidad'])
					})
		##import pdb; pdb.set_trace()		
		materialesAgrupados=[]
		
		for mat in materiales:
			agregar = True
			for obj in materialesAgrupados:
				if mat['codigo'] == obj['codigo']:
					obj['cantidad']	+= mat['cantidad']
					agregar = False
			if agregar:		
				materialesAgrupados.append(mat)



		data = {
			'unidadesConstructivas': unidadesConstructivas,
			'materiales' : materialesAgrupados
		}
		return JsonResponse({'message':'','success':'ok','data': data})
	except Exception as e:
		functions.toLog(e,'avanceObraGrafico2.getCantidadesEnNodo')
		return Response(
			{'message':'Se presentaron errores de comunicacion con el servidor',
			'success':'fail','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FotoNodoSerializer(serializers.HyperlinkedModelSerializer):
	
	nodo = NodoGraficoLiteSerializer(read_only=True)
	nodo_id = serializers.PrimaryKeyRelatedField(
		queryset=HNodo.objects.all(), write_only=True)

	class Meta:
		model=FotoNodo
		fields=('id','nodo','nodo_id',
			'fecha','foto_publica','comentario')


class FotoNodoViewSet(viewsets.ModelViewSet):
	model = FotoNodo
	model_acciones=Acciones
	model_log=Logs
	queryset = model.objects.all()
	serializer_class = FotoNodoSerializer
	nombre_modulo = 'avance_de_obra_grafico2.FotoNodo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','success':'ok','data': serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(FotoNodoViewSet, self).get_queryset()
			page = self.request.query_params.get('page', None)
			nodo = self.request.query_params.get('nodo', None)
			desde = self.request.query_params.get('desde', None)
			hasta = self.request.query_params.get('hasta', None)
			ignorePagination = self.request.query_params.get('ignorePagination', None)

			qset=(~Q(id=0))

			if nodo:
				qset = qset & (Q(nodo__id=nodo))
				if desde:
					qset = 	qset & (Q(fecha__gte=desde))
				if hasta:
					qset = qset & (Q(fecha__lte=hasta))

				queryset = self.model.objects.filter(qset).order_by('-fecha')
				mensaje = 'No se encontraron registros con los' + \
						' criterios de busqueda ingresados.' if queryset.count()==0 else ''

				ignorePagination= self.request.query_params.get('ignorePagination',None)
				if ignorePagination is None:
					page = self.paginate_queryset(queryset)
					if page is not None:
						serializer = self.get_serializer(page,many=True)
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})

				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
					'data':serializer.data})

						
			else:
				return Response({'message':'Debe especificar el nodo','success':'error',
				'data':''},status=status.HTTP_400_BAD_REQUEST)


		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar los datos',
				'success':'error',
				'data':''},status=status.HTTP_400_BAD_REQUEST)


	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				lista=self.request.FILES.getlist('archivo[]')
				for item in lista:
					serializer = FotoNodoSerializer(data = request.data,
						context = {'request': request})
					if serializer.is_valid():
						serializer.save(
							nodo_id=request.data['nodo_id'],
							ruta=item)
						logs_model=Logs(
							usuario_id=request.user.usuario.id,
							accion=Acciones.accion_crear,
							nombre_modelo='avanceObraGrafico2.fotoNodo',
							id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)

						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
					else:						
						return Response({
							'message':'datos requeridos no fueron recibidos',
							'success':'fail',
							'data':''},
							status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos',
					'success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			sid = transaction.savepoint()
			instance = self.get_object()
			id = instance.id
			#archivo = instance.ruta.url
			self.perform_destroy(instance)
			#functions.eliminarArchivoS3(archivo)
			logs_model=Logs(
				usuario_id=request.user.usuario.id,
				accion=Acciones.accion_borrar,
				nombre_modelo='avanceObraGrafico2.fotoNodo',
				id_manipulado=id)
			logs_model.save()
			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			transaction.savepoint_rollback(sid)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

@login_required
def verFoto(request):
	if request.method == 'GET':
		try:
			
			foto = FotoNodo.objects.get(pk=request.GET['id'])			
			return functions.exportarArchivoS3(str(foto.ruta))

		except Exception as e:
			functions.toLog(e,'constrato.VerSoporte')
			return JsonResponse({
				'message':'Se presentaron errores de comunicacion con el servidor',
				'status':'error','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)	


def graficaCronograma(request,id):
	try:
			
		porcentajes = []
		avanceObra = []
		avanceFinanciero = []
		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__id=id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__id=id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0
				})

	#codigo para obtener el avance de obra en el tiempo
		
		reportesTrabajo = JReporteTrabajo.objects.filter(
			presupuesto__cronograma__id=id
			).order_by('fechaTrabajo').values('fechaTrabajo').distinct()

		for fechaReporteTrabajo in reportesTrabajo:
			avanceFisico = 0
			porcentajesAvanceFisico = []
			queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
				detallepresupuesto__presupuesto__cronograma__id=id,
				cantidadEjecutada__gt=0,
				reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo']).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre'
				).annotate(total=Sum('cantidadEjecutada'))

			for Aejecutar in queryset_a_ejecutar:
				agregado = False
				for ejecutado in queryset_ejecutada:
					if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
						agregado = True
						porcentajesAvanceFisico.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
							'peso': Aejecutar['detallepresupuesto__actividad__peso']
						})
					if agregado:
						break
				if agregado == False:
					porcentajesAvanceFisico.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': 0,
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})

			
			for p in porcentajesAvanceFisico:
				avanceFisico = avanceFisico + ((float(p['porcentaje']) * float(float(p['peso'])/100)))

			#if cantidadDeProyectos > 0:
					#Codigo para llenar la curva programada:
			
			
			qsetP = DiagramaGrahm.objects.filter(
				fechaFinal__lte=fechaReporteTrabajo['fechaTrabajo'],
				cronograma__id=id).values('cronograma__id').annotate(
				porcentaje=Sum('actividad__peso'))

			if qsetP:	

				avanceObra.append ( {
					'fecha' : fechaReporteTrabajo['fechaTrabajo'],
					'avance' : round(avanceFisico,2),
					'avance_proyectado': qsetP[0]['porcentaje']
					})
			else:
				avanceObra.append ( {
					'fecha' : fechaReporteTrabajo['fechaTrabajo'],
					'avance' : round(avanceFisico,2),
					'avance_proyectado': 0
					})


		if avanceObra:		
			fechamax = avanceObra[-1]['fecha']	
			avanceMax = avanceObra[-1]['avance']
			avanceMaxPr = avanceObra[-1]['avance_proyectado']
			qsetP = DiagramaGrahm.objects.filter(
					fechaFinal__gte=fechamax,
					cronograma__id=id).order_by('fechaFinal').values(
					'fechaFinal').annotate(
					porcentaje=Sum('actividad__peso'))	

			proy = avanceMaxPr
			for rowD in qsetP:
				proy = proy + rowD['porcentaje']
				avanceObra.append({
					'fecha': rowD['fechaFinal'],
					'avance':round(avanceMax,2),
					'avance_proyectado': proy
					})
		else:
			qsetP = DiagramaGrahm.objects.filter(
					cronograma__id=id).order_by('fechaFinal').values(
					'fechaFinal').annotate(
					porcentaje=Sum('actividad__peso'))
			proy = 0
			for rowD in qsetP:
				proy = proy + rowD['porcentaje']
				avanceObra.append({
					'fecha': rowD['fechaFinal'],
					'avance':0,
					'avance_proyectado': proy
					})
					

		#codigo para obtener el avance financiero en el tiempo
		totalPresupuesto = 0
		avanceF=0
		cantidadesEjecutar = FDetallePresupuesto.objects.filter(
			presupuesto__cronograma__id=id).values(
			'cantidad','codigoUC','cantidad',
			'catalogoUnidadConstructiva__id', 'id')			
		for rowCantidadesEjecutar in cantidadesEjecutar:
			totalMo = 0
			totalMat = 0
			queryset = DesgloceManoDeObra.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']
				).values(
				'manoDeObra__valorHora','rendimiento')
			
			if queryset:
				for row in queryset:
					totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

			queryset = DesgloceMaterial.objects.filter(
				unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
				unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']).values(
				'material__valorUnitario','cantidad')
			
			if queryset:
				for row in queryset:
					totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))

			totalPresupuesto = totalPresupuesto + \
			((totalMo * float(rowCantidadesEjecutar['cantidad'])) + \
			(totalMat * float(rowCantidadesEjecutar['cantidad'])))

		##import pdb; pdb.set_trace()	
		for fechaReporteTrabajo in reportesTrabajo:
			totalEjecutado = 0
			avanceF=0
			cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
				cantidadEjecutada__gt=0,
				reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo'],
				#detallepresupuesto__id=rowCantidadesEjecutar['id']
				).values('detallepresupuesto__id',
				'detallepresupuesto__codigoUC',
				'detallepresupuesto__catalogoUnidadConstructiva__id').annotate(
				total=Sum('cantidadEjecutada'))

				
			if cantidadesEjecutadas:
				for cantidadEjecutada in cantidadesEjecutadas:
					totalMo = 0
					totalMat = 0

					queryset = DesgloceManoDeObra.objects.filter(
						unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
						unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']
						).values(
						'manoDeObra__valorHora','rendimiento')
					
					if queryset:
						for row in queryset:
							totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

					queryset = DesgloceMaterial.objects.filter(
						unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
						unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']).values(
						'material__valorUnitario','cantidad')
					
					if queryset:
						for row in queryset:
							totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))


					totalEjecutado = totalEjecutado + \
					(float(cantidadEjecutada['total']) * totalMo) + \
					(float(cantidadEjecutada['total']) * totalMat)

			if totalPresupuesto > 0:
				avanceF = avanceF + round((totalEjecutado / totalPresupuesto) * 100,2)		

			#if cantidadDeProyectos > 0:
			avanceF = round(avanceF,2)
			#else:
			#	avanceF = 0

			avanceFinanciero.append({
				'fecha': fechaReporteTrabajo['fechaTrabajo'],
				'avance': avanceF,
				'monto': round(totalEjecutado) #"$ {:,.2f}".format(round(totalEjecutado,2))
			})





		datos = {
			#'avanceObra' : avance,
			'porHito' : porcentajes,
			'curvaAvanceObra' : avanceObra,
			'curvaAvanceFinanciero' : avanceFinanciero
			#'obrasPorEstado' : obrasPorEstado
		}


		return JsonResponse({'message':'','success':'ok','data':datos})					

	except Exception as e:
		functions.toLog(e,'avanceObraGrafico2.graficacronograma')
		return Response({
			'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def graficasTablero(request,id):
	try:
		cantidadDeProyectos = 0
		porcentajes = []
		avanceObra = []
		avanceFinanciero = []
		obrasPorEstado = []
		#codigo para obtener el avance por hito
		avance = 0
		porcentajes = []
		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))	

		queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['detallepresupuesto__actividad__peso']
				})
		
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		##import pdb; pdb.set_trace()
		cantidadDeProyectos = Proyecto.objects.filter(
			mcontrato__id=id).count()	
		if cantidadDeProyectos > 0:
			avance = round(avance,2)
		else:
			avance = 0

		#codigo para obtener el avance de obra en el tiempo
		
		# reportesTrabajo = JReporteTrabajo.objects.filter(
		# 	presupuesto__cronograma__proyecto__mcontrato__id=id
		# 	).order_by('fechaTrabajo').values('fechaTrabajo').distinct()

		# for fechaReporteTrabajo in reportesTrabajo:
		# 	avanceFisico = 0
		# 	porcentajesAvanceFisico = []
		# 	queryset_ejecutada = KDetalleReporteTrabajo.objects.filter(
		# 		detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
		# 		cantidadEjecutada__gt=0,
		# 		reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo']).values(
		# 		'detallepresupuesto__actividad__id',
		# 		'detallepresupuesto__actividad__nombre'
		# 		).annotate(total=Sum('cantidadEjecutada'))

		# 	for Aejecutar in queryset_a_ejecutar:
		# 		agregado = False
		# 		for ejecutado in queryset_ejecutada:
		# 			if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
		# 				agregado = True
		# 				porcentajesAvanceFisico.append({
		# 					'id' : Aejecutar['detallepresupuesto__actividad__id'],
		# 					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
		# 					'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
		# 					'peso': Aejecutar['detallepresupuesto__actividad__peso']
		# 				})
		# 			if agregado:
		# 				break
		# 		if agregado == False:
		# 			porcentajesAvanceFisico.append({
		# 				'id' : Aejecutar['detallepresupuesto__actividad__id'],
		# 				'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
		# 				'porcentaje': 0,
		# 				'peso': Aejecutar['detallepresupuesto__actividad__peso']
		# 			})
			
		# 	for p in porcentajesAvanceFisico:
		# 		avanceFisico = avanceFisico + ((float(p['porcentaje']) * float(float(p['peso'])/100)))

		# 	if cantidadDeProyectos > 0:

		# 		avanceObra.append ( {
		# 			'fecha' : fechaReporteTrabajo['fechaTrabajo'],
		# 			'avance' : round(avanceFisico,2)
		# 			})

		# #codigo para obtener el avance financiero en el tiempo
		# totalPresupuesto = 0
		# avanceF=0
		# cantidadesEjecutar = FDetallePresupuesto.objects.filter(
		# 	presupuesto__cronograma__proyecto__mcontrato__id=id).values(
		# 	'cantidad','codigoUC','cantidad',
		# 	'catalogoUnidadConstructiva__id', 'id')			
		# for rowCantidadesEjecutar in cantidadesEjecutar:
		# 	totalMo = 0
		# 	totalMat = 0
		# 	queryset = DesgloceManoDeObra.objects.filter(
		# 		unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
		# 		unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']
		# 		).values(
		# 		'manoDeObra__valorHora','rendimiento')
			
		# 	if queryset:
		# 		for row in queryset:
		# 			totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		# 	queryset = DesgloceMaterial.objects.filter(
		# 		unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
		# 		unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']).values(
		# 		'material__valorUnitario','cantidad')
			
		# 	if queryset:
		# 		for row in queryset:
		# 			totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		# 	totalPresupuesto = totalPresupuesto + \
		# 	((totalMo * float(rowCantidadesEjecutar['cantidad'])) + \
		# 	(totalMat * float(rowCantidadesEjecutar['cantidad'])))

		# ##import pdb; pdb.set_trace()	
		# for fechaReporteTrabajo in reportesTrabajo:
		# 	totalEjecutado = 0
		# 	avanceF=0
		# 	cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
		# 		cantidadEjecutada__gt=0,
		# 		reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo'],
		# 		#detallepresupuesto__id=rowCantidadesEjecutar['id']
		# 		).values('detallepresupuesto__id',
		# 		'detallepresupuesto__codigoUC',
		# 		'detallepresupuesto__catalogoUnidadConstructiva__id').annotate(
		# 		total=Sum('cantidadEjecutada'))

				
		# 	if cantidadesEjecutadas:
		# 		for cantidadEjecutada in cantidadesEjecutadas:
		# 			totalMo = 0
		# 			totalMat = 0

		# 			queryset = DesgloceManoDeObra.objects.filter(
		# 				unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
		# 				unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']
		# 				).values(
		# 				'manoDeObra__valorHora','rendimiento')
					
		# 			if queryset:
		# 				for row in queryset:
		# 					totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		# 			queryset = DesgloceMaterial.objects.filter(
		# 				unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
		# 				unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']).values(
		# 				'material__valorUnitario','cantidad')
					
		# 			if queryset:
		# 				for row in queryset:
		# 					totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))


		# 			totalEjecutado = totalEjecutado + \
		# 			(float(cantidadEjecutada['total']) * totalMo) + \
		# 			(float(cantidadEjecutada['total']) * totalMat)

		# 	if totalPresupuesto > 0:
		# 		avanceF = avanceF + round((totalEjecutado / totalPresupuesto) * 100,2)		

		# 	if cantidadDeProyectos > 0:
		# 		avanceF = round(avanceF,2)
		# 	else:
		# 		avanceF = 0

		# 	avanceFinanciero.append({
		# 		'fecha': fechaReporteTrabajo['fechaTrabajo'],
		# 		'avance': avanceF,
		# 		'monto': round(totalEjecutado) #"$ {:,.2f}".format(round(totalEjecutado,2))
		# 	})


		datos = {
			'avanceObra' : avance,
			'porHito' : porcentajes,
			# 'curvaAvanceObra' : avanceObra,
			# 'curvaAvanceFinanciero' : avanceFinanciero,
			# 'obrasPorEstado' : obrasPorEstado
		}

		return JsonResponse({'message':'','success':'ok','data':datos})
	except Exception as e:
		return Response({
			'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)


api_view(['GET',])
def cantidadesAliquidar(request, id):
	try:
		#identificar cantidades a ejecutar:
		datos = []
		cantidadesEjecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__id=id,
			cantidad__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(
			total=Sum('cantidad'))

		cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			#'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))


		# cantidadesLiquidadas = MLiquidacionUUCC.objects.filter(
		# 	presupuesto__id=id).values(
		# 	'detallereportetrabajo__detallepresupuesto__id',
		# 	'detallereportetrabajo__detallepresupuesto__codigoUC',
		# 	'detallereportetrabajo__detallepresupuesto__descripcionUC',
		# 	'detallereportetrabajo__detallepresupuesto__catalogoUnidadConstructiva__id'
		# 	).annotate(total = Sum('detallereportetrabajo__cantidadEjecutada'))	

		##import pdb; pdb.set_trace()	
		for cantidadEjecutar in cantidadesEjecutar:
			cejecutar = 0
			cliquidada = 0
			liquidada = False
			for cantidadEjecutada in cantidadesEjecutadas:
				if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
					cejecutar = cantidadEjecutada['total']
					#if cantidadEjecutada['detallepresupuesto__liquidacion__id']:
					#liquidada = cantidadEjecutada['liquidada']
					break

			# for cantidadLiquidada in cantidadesLiquidadas:
			# 	if cantidadLiquidada['detallereportetrabajo__detallepresupuesto__codigoUC'] == \
			# 	cantidadEjecutar['detallepresupuesto__codigoUC']:
			# 		cliquidada = cantidadLiquidada['total']
			# 		break

			# liquidable = True

			# if cantidadEjecutar['total'] == cejecutar and cantidadEjecutar['total'] == cliquidada:
			# 	liquidable = False

			datos.append(
				{
					'hito': cantidadEjecutar['detallepresupuesto__actividad__nombre'],
					'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
					'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
					'ejecutar': cantidadEjecutar['total'],
					'ejecutado': cejecutar,
					'pendiente': int(cantidadEjecutar['total']) - int(cejecutar)
					#'liquidado' : cliquidada,
					#'liquidada': liquidada
				}
			)
				

			
		return JsonResponse({'message':'','success':'ok','data':datos})	
	except Exception as e:
		functions.toLog(e,'avanceObraGRafico2.cantidadesAliquidar')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def seguimientoCantidades(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
				
	return render(request, 'avanceObraGrafico2/seguimientocantidades.html',
		{'model':'jreportetrabajo',
		'app':'avanceObraGrafico2',
		'presupuesto':presupuesto})


@login_required
def liquidacionuucc(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraGrafico2/liquidacion_uucc.html',{
		'model':'mliquidacionuucc',
		'app':'avanceObraGrafico2',
		'Usuarios':usuarios,
		'cronograma':nombre_presupuesto.cronograma,
		'cronograma_id':nombre_presupuesto.cronograma.id,
		'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,
		'presupuesto_id':id_presupuesto,
		'presupuesto':nombre_presupuesto})

class LiquidacionUUCCSerializer(serializers.HyperlinkedModelSerializer):
	valor_liquidacion = serializers.SerializerMethodField('_valor_liquidacion',read_only=True)
	cantidaduucc = serializers.SerializerMethodField('_cantidaduucc',read_only=True)

	class Meta:
		model=MLiquidacionUUCC
		fields=('id','estado','fecha','cantidaduucc','valor_liquidacion')

	def _cantidaduucc(self,obj):
		###import pdb; pdb.set_trace()		
		aux = obj.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=obj.id).values('kdetallereportetrabajo__detallepresupuesto__id').distinct()
		return float(aux.count())

	def _valor_liquidacion(self,obj):
		model=MLiquidacionUUCC.objects.get(pk=obj.id)

		list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=obj.id).values('kdetallereportetrabajo_id')

		###import pdb; pdb.set_trace()
		subtotal_manoobra_acumulado=0
		subtotal_material_acumulado=0

		for x in list_detallereporte:
			model_detallereporte=KDetalleReporteTrabajo.objects.get(pk=int(x['kdetallereportetrabajo_id']))

			subtotal_manoobra_acumulado+=float(model_detallereporte.detallepresupuesto.valorManoObra*model_detallereporte.cantidadEjecutada)
			subtotal_material_acumulado+=float(model_detallereporte.detallepresupuesto.valorMaterial*model_detallereporte.cantidadEjecutada)

		return float(subtotal_manoobra_acumulado+subtotal_material_acumulado)


class LiquidacionUUCCLiteSerializer(serializers.HyperlinkedModelSerializer):
	presupuesto = PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id = serializers.PrimaryKeyRelatedField(queryset=EPresupuesto.objects.all(), write_only=True)
	

	class Meta:
		model=MLiquidacionUUCC
		fields=('id','estado','presupuesto','presupuesto_id','fecha')

	

class LiquidacionUUCCViewSet(viewsets.ModelViewSet):
	
	model=MLiquidacionUUCC
	queryset = model.objects.all()
	serializer_class = LiquidacionUUCCLiteSerializer
	nombre_modulo='avance_de_obra_grafico2.liquidacion'	
	model_log=Logs
	model_acciones=Acciones	

	
	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	
	def list(self, request, *args, **kwargs):
		if request.method == 'GET':			
			try:
				###import pdb; pdb.set_trace()
				queryset = super(LiquidacionUUCCViewSet, self).get_queryset()			
				sin_paginacion= self.request.query_params.get('sin_paginacion',None)
				presupuesto_id = self.request.query_params.get('presupuesto_id',None)

				qset=(~Q(id=0))
				if  presupuesto_id:
					qset = qset & (Q(presupuesto__id=int(presupuesto_id)))

				queryset = self.model.objects.filter(qset).order_by('-id')

				if sin_paginacion is None:
					page = self.paginate_queryset(queryset)
					if page is not None:
						serializer = LiquidacionUUCCSerializer(page,many=True)	
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
		
				serializer = LiquidacionUUCCSerializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})			
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#@transaction.atomic
api_view(['GET',])
def consultar_uucc_ejecutados(request):	
	if request.method == 'GET':			
		try:
			###import pdb; pdb.set_trace()
			presupuesto_id=request.GET['presupuesto_id']
			aux= MLiquidacionUUCC.objects.filter(presupuesto__id=int(request.GET['presupuesto_id']))
			list_excluir=[]

			for a in aux:				
				list_aux=a.detallereportetrabajo.through.objects.filter().values('kdetallereportetrabajo_id')
				for b in list_aux:
					list_excluir.append(b['kdetallereportetrabajo_id'])

			
			qset = (Q(detallepresupuesto__presupuesto__id=int(presupuesto_id))) & (~Q(id__in=list_excluir))
			reporte=KDetalleReporteTrabajo.objects.filter(qset)

			list_return=[]

			for x in reporte:
				list_return.append({
					'id':x.id,
					'codigo':x.detallepresupuesto.codigoUC,
					'descripcion':x.detallepresupuesto.descripcionUC,
					'nodo':x.nodo.nombre,
					})

			###import pdb; pdb.set_trace()
			return JsonResponse({'message':'','success':'ok','data':list_return})

		except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
def anularReporteLiquidacion(request):
	
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			model = MLiquidacionUUCC.objects.get(pk=int(respuesta['liquidacion_id']))
			model.estado=3
			model.motivo_anulacion=respuesta['motivo_anular']
			model.save()

			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraGrafico2.AnularReporteLiquidacion'
									,id_manipulado=model.id)
			logs_model.save()


			list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=model.id)
			aux = model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=model.id).values('kdetallereportetrabajo_id')

			##import pdb; pdb.set_trace()
			for y in aux:
				model_detallereporte=KDetalleReporteTrabajo.objects.get(pk=y['kdetallereportetrabajo_id'])
				model_detallereporte.liquidada=False
				model_detallereporte.save()

				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraGrafico2.KDetalleReporteTrabajo'
									,id_manipulado=model_detallereporte.id)
				logs_model.save()


			for x in list_detallereporte:				
				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_borrar
									,nombre_modelo='avanceObraGrafico2.AnularReporteLiquidacion'
									,id_manipulado=x.id)				
				model.detallereportetrabajo.remove(x.kdetallereportetrabajo_id)

				logs_model.save()

		

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'La liquidacion se ha anulado correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraGRafico2.liquidacionuucc')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

def cerrar_liquidacionuucc(request):
	###import pdb; pdb.set_trace()
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			model = MLiquidacionUUCC.objects.get(pk=int(respuesta['liquidacion_id']))
			model.estado=2
			model.save()

			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraGrafico2.liquidacionuucc'
									,id_manipulado=model.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'La liquidacion se ha cerrado correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraGRafico2.liquidacionuucc')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



def guardar_liquidacionuucc(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			##import pdb; pdb.set_trace()
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			liquidacionuucc=MLiquidacionUUCC(presupuesto_id=int(respuesta['presupuesto_id']))
			liquidacionuucc.save()


			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_crear
									,nombre_modelo='avanceObraGrafico2.liquidacionuucc'
									,id_manipulado=liquidacionuucc.id)
			logs_model.save()
			transaction.savepoint_commit(sid)

			list_detallepresupuesto=respuesta['lista']
			list_aux=[]
			for d in list_detallepresupuesto:
				list_detallereporte = KDetalleReporteTrabajo.objects.filter(liquidada=False,detallepresupuesto__id=d['id']).values('id')
				for e in list_detallereporte:
					list_aux.append(e['id'])
			##import pdb; pdb.set_trace()

			list_temporal=[]
			for x in list_aux:
				liquidacionuucc.detallereportetrabajo.add(x)
				list_temporal.append(x)

				model_detallert=KDetalleReporteTrabajo.objects.get(pk=x)
				model_detallert.liquidada=True
				model_detallert.save()

				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraGrafico2.detallereportetrabajo(liquidacion)'
									,id_manipulado=x)
				logs_model.save()

			
			list_liq_detalle = liquidacionuucc.detallereportetrabajo.through.objects.filter(kdetallereportetrabajo_id__in=list_temporal)

			insert_list = []			
			for a in list_liq_detalle:
				insert_list.append(Logs(usuario_id=request.user.usuario.id
										,accion=Acciones.accion_crear
										,nombre_modelo='avanceObraGrafico2.liquidacionuucc.detallereportetrabajo'
										,id_manipulado=a.id)
										)
			Logs.objects.bulk_create(insert_list)

			transaction.savepoint_commit(sid)
			

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	

# Exporta a excel contrato 
def exportReporteLiquidacion(request):
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte_liquidacion.xls"'

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Liquidacion de UUCC')

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
		format1.set_align('center')
		format1.set_align('vcenter')
		format2=workbook.add_format({'border':0})
		format3=workbook.add_format({'border':0,'font_size':12})
		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')

		worksheet.set_column('A:A', 15)
		worksheet.set_column('B:B', 35)
		worksheet.set_column('C:H', 20)

		row=1
		col=0

		data_return=[]

		model=MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))

		list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=int(request.GET['liquidacion_id'])).values('kdetallereportetrabajo__detallepresupuesto__id').annotate(cantidad_ejecutada=Sum('kdetallereportetrabajo__cantidadEjecutada')).distinct()

		
		subtotal_manoobra_acumulado=0
		subtotal_material_acumulado=0
		total_acumulado = 0

		for x in list_detallereporte:
			model_detalle_presupuesto=FDetallePresupuesto.objects.get(pk=int(x['kdetallereportetrabajo__detallepresupuesto__id']))

			###import pdb; pdb.set_trace()

			subtotal_manoobra = float(model_detalle_presupuesto.valorManoObra*x['cantidad_ejecutada'])
			subtotal_material = float(model_detalle_presupuesto.valorMaterial*x['cantidad_ejecutada'])
			total = float(subtotal_manoobra+subtotal_material)

			subtotal_manoobra_acumulado+=subtotal_manoobra
			subtotal_material_acumulado+=subtotal_material
			total_acumulado = total_acumulado + total

			data_return.append({
				'id':model_detalle_presupuesto.id,
				'codigoUC':model_detalle_presupuesto.codigoUC,
				'descripcionUC':model_detalle_presupuesto.descripcionUC,
				'cantidad_ejecutada':float(x['cantidad_ejecutada']),
				'valorManoObra':model_detalle_presupuesto.valorManoObra,
				'valorMaterial':model_detalle_presupuesto.valorMaterial,
				'Subtotal_mano_obra':float(subtotal_manoobra),
				'Subtotal_materiales':float(subtotal_material),
				'Total':total,
				})

		data_return.append({
			'id':' ',
			'codigoUC':' ',
			'descripcionUC':' ',
			'cantidad_ejecutada':' ',
			'valorManoObra':' ',
			'valorMaterial':' ',
			'Subtotal_mano_obra':float(subtotal_manoobra_acumulado),
			'Subtotal_materiales':float(subtotal_material_acumulado),
			'Total':float(total_acumulado),
			})

		if data_return:			
			worksheet.write('A1', 'Codigo UUCC', format1)
			worksheet.write('B1', 'Descripcion UUCC', format1)
			worksheet.write('C1', 'Cantidad ejecutada', format1)
			worksheet.write('D1', 'Vlr Mano obra', format1)
			worksheet.write('E1', 'Vlr Material', format1)
			worksheet.write('F1', 'Subtotal mano obra', format1)
			worksheet.write('G1', 'Subtotal materiales', format1)
			worksheet.write('H1', 'Valor Liquidacion', format1)

			##import pdb; pdb.set_trace()
			for x in data_return:
				worksheet.write(row, col,x['codigoUC'],format2)
				worksheet.write(row, col+1,x['descripcionUC'],format2)
				worksheet.write(row, col+2,x['cantidad_ejecutada'],format2)
				worksheet.write(row, col+3,x['valorManoObra'],format2)
				worksheet.write(row, col+4,x['valorMaterial'],format2)
				worksheet.write(row, col+5,x['Subtotal_mano_obra'],format2)
				worksheet.write(row, col+6,x['Subtotal_materiales'],format2)
				worksheet.write(row, col+7,x['Total'],format2)
				row +=1
			workbook.close()
			return response


	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	


api_view(['GET',])
def consultar_detallereporte_liquidacion(request):
	if request.method == 'GET':
		try:
			##import pdb; pdb.set_trace()
			data_return=[]
			model=MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))

			list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=int(request.GET['liquidacion_id'])).values('kdetallereportetrabajo__detallepresupuesto__id').annotate(cantidad_ejecutada=Sum('kdetallereportetrabajo__cantidadEjecutada')).distinct()

			###import pdb; pdb.set_trace()
			subtotal_manoobra_acumulado=0
			subtotal_material_acumulado=0
			for x in list_detallereporte:
				model_detalle_presupuesto=FDetallePresupuesto.objects.get(pk=int(x['kdetallereportetrabajo__detallepresupuesto__id']))

				###import pdb; pdb.set_trace()
				subtotal_manoobra = float(model_detalle_presupuesto.valorManoObra*x['cantidad_ejecutada'])
				subtotal_material = float(model_detalle_presupuesto.valorMaterial*x['cantidad_ejecutada'])

				subtotal_manoobra_acumulado+=subtotal_manoobra
				subtotal_material_acumulado+=subtotal_material
			
				data_return.append({
					'id':model_detalle_presupuesto.id,
					'codigoUC':model_detalle_presupuesto.codigoUC,
					'descripcionUC':model_detalle_presupuesto.descripcionUC,
					'cantidad_ejecutada':float(x['cantidad_ejecutada']),
					'valorManoObra':model_detalle_presupuesto.valorManoObra,
					'valorMaterial':model_detalle_presupuesto.valorMaterial,
					'Subtotal_mano_obra':float(subtotal_manoobra),
					'Subtotal_materiales':float(subtotal_material),
					})

			data_return.append({
				'id':' ',
				'codigoUC':' ',
				'descripcionUC':' ',
				'cantidad_ejecutada':' ',
				'valorManoObra':' ',
				'valorMaterial':' ',
				'Subtotal_mano_obra':float(subtotal_manoobra_acumulado),
				'Subtotal_materiales':float(subtotal_material_acumulado),
				})
		
			return JsonResponse({'message':'','success':'ok','data':data_return})
		except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	

api_view(['GET',])
def cantidadesAliquidarlite(request):
	try:
		#identificar cantidades a ejecutar:
		datos = []
		##import pdb; pdb.set_trace()
		# cantidadesEjecutar = JCantidadesNodo.objects.filter(
		# 	detallepresupuesto__presupuesto__id=int(request.GET['presupuesto_id']),
		# 	cantidad__gt=0).values(
		# 	'detallepresupuesto__id',
		# 	'detallepresupuesto__codigoUC',
		# 	'detallepresupuesto__descripcionUC',
		# 	'detallepresupuesto__catalogoUnidadConstructiva__id'
		# 	).annotate(
		# 	total=Sum('cantidad'))

		cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = int(request.GET['presupuesto_id']),
			cantidadEjecutada__gt=0, liquidada=False).values(
			'detallepresupuesto__id',
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))



		# 	
		# for cantidadEjecutar in cantidadesEjecutar:
		# 	cejecutar = 0
		# 	cliquidada = 0
		# 	liquidada = False
		# 	for cantidadEjecutada in cantidadesEjecutadas:
		# 		if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
		# 			cejecutar = cantidadEjecutada['total']
		# 			#if cantidadEjecutada['detallepresupuesto__liquidacion__id']:
		# 			liquidada = cantidadEjecutada['liquidada']
		# 			break


		# 	datos.append(
		# 		{	'id': cantidadEjecutar['detallepresupuesto__id'],
		# 			'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
		# 			'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
		# 			'ejecutar': float(cantidadEjecutar['total']),
		# 			'ejecutado': float(cejecutar),
		# 			#'liquidado' : cliquidada,
		# 			'liquidada': liquidada
		# 		}
		# 	)

		return JsonResponse({'message':'','success':'ok','data':list(cantidadesEjecutadas)})
	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
api_view(['GET',])
def consultarmotivoanulacion(request):
	try:
		##import pdb; pdb.set_trace()
		model = MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))
		data=[]
		data.append({
			'id':model.id,
			'motivo_anulacion':model.motivo_anulacion
			})
		return JsonResponse({'message':'','success':'ok','data':data})
	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})

def exportCronograma(request):	
	
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte_cronogramas.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Cronogramas')
		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:G', 25)
		worksheet.set_column('H:AQ', 20)

		row=1
		col=0

		dato = request.GET['dato'] if request.GET['dato'] else None
		mcontrato = request.GET['macrocontrato_id'] if request.GET['macrocontrato_id'] else None
		departamento = request.GET['departamento_id'] if request.GET['departamento_id'] else None
		municipio = request.GET['municipio_id'] if request.GET['municipio_id'] else None
		contratista = request.GET['contratista_id'] if request.GET['contratista_id'] else None
		empresa = request.user.usuario.empresa.id

		

		qset=(~Q(id=0))
		if dato:
			qset = qset & ( Q(empresa__nombre__icontains = dato) |
						Q(empresa__nit__icontains = dato) |
						Q(proyecto__nombre__icontains = dato))

		if empresa and int(empresa)>0:
			qset = qset & (Q(empresa__id = empresa))

		if mcontrato and int(mcontrato)>0:
			qset = qset & (Q(proyecto__mcontrato__id = mcontrato))
			
		if departamento and int(departamento)>0:
			qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
			
		if municipio and int(municipio)>0:
			qset = qset & (Q(proyecto__municipio__id = municipio))
			

		if contratista and int(contratista)>0:
			qset = qset & (Q(proyecto__contrato__contratista__id = contratista))

		queryset = Proyecto_empresas.objects.filter(qset)

		serializer_context = {
			'request': request
		}

		serializer = ProyectoEmpresaLite4Serializer(queryset,many=True,context=serializer_context)

		if serializer:
			worksheet.write('A1', 'Macrocontrato', format1)
			worksheet.write('B1', 'Departamento', format1)
			worksheet.write('C1', 'Municipio', format1)
			worksheet.write('D1', 'Proyecto', format1)
			worksheet.write('E1', '% Avance de obra', format1)
			worksheet.write('F1', '% Avance financiero', format1)
			worksheet.write('G1', 'Fecha corte', format1)

			
			

			for item in serializer.data:
				#import pdb; pdb.set_trace()
				worksheet.write(row, col,item['proyecto']['mcontrato']['nombre'],format3)
				worksheet.write(row, col+1,item['proyecto']['municipio']['departamento']['nombre'],format3)
				worksheet.write(row, col+2,item['proyecto']['municipio']['nombre'],format3)
				worksheet.write(row, col+3,item['proyecto']['nombre'],format3)
				porcentajeAvance = item['porcentajeAvance']['avance']
				porcentajeAvanceFinanciero = item['porcentajeAvanceFinanciero']
				worksheet.write(row, col+4,str(porcentajeAvance)+'%',format2)
				worksheet.write(row, col+5,str(porcentajeAvanceFinanciero)+'%',format2)
				worksheet.write(row, col+6,item['porcentajeAvance']['corte'],format5)
				n_col=7
				
				if mcontrato and int(mcontrato)>0:
					qset = (Q(esquema__macrocontrato__id=int(mcontrato)))&(~Q(padre=0))&(~Q(nombre='Replanteo'))&(~Q(nombre='replanteo'))&(~Q(nombre='Liquidacion'))&(~Q(nombre='liquidacion'))
					capitulos_model= CEsquemaCapitulosActividadesG.objects.filter(qset).exists()
					if capitulos_model:
						capitulos_model= CEsquemaCapitulosActividadesG.objects.filter(qset)
						
						for acts in capitulos_model:
							worksheet.write(0, n_col,acts.nombre,format1)
							if item['porcentajeAvance']['actividades']:
								#import pdb; pdb.set_trace()

								# act_model = item['porcentajeAvance']['actividades']

								# act_obj = act_model.objects.filter(id=int(acts.id)).exists()

								# if act_obj:
								# 	act_obj = act_model.objects.filter(id=int(acts.id))
								# 	worksheet.write(row, n_col,str( round(act_obj.porcentaje,2))+'%',format2)

								for act in item['porcentajeAvance']['actividades']:
									if int(acts.id)==int(act['id']):																						
										worksheet.write(row, n_col,str( round(act['porcentaje'],2))+'%',format2)			
							else:
								worksheet.write(row, n_col,'0%',format2)

							n_col+=1
					

				row +=1

			

		workbook.close()
		return response
	except Exception as e:
		#print(e)
		functions.toLog(e,'avanceObraGrafico2')
		return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def seguimientoMateriales(request,id_presupuesto):
	#import pdb; pdb.set_trace()
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	#usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraGrafico2/seguimientomateriales.html',{
		'model':'jreportetrabajo',
		'app':'avanceObraGrafico2',		
		'cronograma':nombre_presupuesto.cronograma,
		'cronograma_id':nombre_presupuesto.cronograma.id,
		'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,
		'presupuesto_id':id_presupuesto,
		'presupuesto':nombre_presupuesto})

@login_required
def materialesAliquidarlite(request, id):
	try:		
			
		listado=FDetallePresupuesto.objects.filter(presupuesto__id=id, cantidad__gt=0)
		materiales = []

		
		for item in listado:		

			cantidad_ejecutada=KDetalleReporteTrabajo.objects.filter(	
				detallepresupuesto_id=item.id
				).aggregate(sum_cantidad_ejecutada=Sum('cantidadEjecutada'))


			cantidad_ejecutada['sum_cantidad_ejecutada']=0 if cantidad_ejecutada['sum_cantidad_ejecutada']==None else cantidad_ejecutada['sum_cantidad_ejecutada']
			
			
			#Extraer los materiales:
			uc = UnidadConstructiva.objects.filter(
				codigo=item.codigoUC,
				catalogo__id=item.catalogoUnidadConstructiva.id).values('id')

			#import pdb; pdb.set_trace()

			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=uc[0]['id']).values(
				'material__codigo','material__descripcion', 'cantidad')


			for obj in desgloceMat:
				materiales.append({
					'codigo' : obj['material__codigo'],
					'hito': item.actividad.nombre,
					'descripcion' : obj['material__descripcion'],
					'cantidad' : float(obj['cantidad']) * float(item.cantidad),
					'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['sum_cantidad_ejecutada']),
					'pendiente': 0
				})
			#unificar cantidades de materialales:
			materialesAgrupados=[]
			
			for mat in materiales:
				agregar = True
				for obj in materialesAgrupados:
					if mat['codigo'] == obj['codigo'] and mat['hito'] == obj['hito']:
						obj['cantidad']	+= mat['cantidad']
						obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
						agregar = False
				if agregar:		
					materialesAgrupados.append(mat)				

			for obj in materialesAgrupados:
				obj['cantidad'] = round(obj['cantidad'],2)
				obj['cantidad_ejecutada'] = round(obj['cantidad_ejecutada'],2)
				obj['pendiente'] = round(obj['cantidad'] - obj['cantidad_ejecutada'],2)
		

		if listado:
			datos=materialesAgrupados
			
		else:
			datos=[]

		
		#import pdb; pdb.set_trace()
		return JsonResponse({'message':'','success':'ok','data':datos})
	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.cantidadesAliquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
@login_required
def exportarMaterialesaLiquidar(request):
	try:
		presupuesto_id= request.GET['presupuesto_id'] if request.GET['presupuesto_id'] else None
		listado=FDetallePresupuesto.objects.filter(presupuesto__id=int(presupuesto_id), cantidad__gt=0)
		materiales = []		
		for item in listado:		

			cantidad_ejecutada=KDetalleReporteTrabajo.objects.filter(	
				detallepresupuesto_id=item.id
				).aggregate(sum_cantidad_ejecutada=Sum('cantidadEjecutada'))

			cantidad_ejecutada['sum_cantidad_ejecutada']=0 if cantidad_ejecutada['sum_cantidad_ejecutada']==None else cantidad_ejecutada['sum_cantidad_ejecutada']
				
			#Extraer los materiales:
			uc = UnidadConstructiva.objects.filter(
				codigo=item.codigoUC,
				catalogo__id=item.catalogoUnidadConstructiva.id).values('id')
			#import pdb; pdb.set_trace()
			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=uc[0]['id']).values(
				'material__codigo','material__descripcion', 'cantidad')

			for obj in desgloceMat:
				materiales.append({
					'codigo' : obj['material__codigo'],
					'hito': item.actividad.nombre,
					'descripcion' : obj['material__descripcion'],
					'cantidad' : float(obj['cantidad']) * float(item.cantidad),
					'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['sum_cantidad_ejecutada']),
					'pendiente': 0
				})
			#unificar cantidades de materialales:
			materialesAgrupados=[]
			
			for mat in materiales:
				agregar = True
				for obj in materialesAgrupados:
					if mat['codigo'] == obj['codigo'] and mat['hito'] == obj['hito']:
						obj['cantidad']	+= mat['cantidad']
						obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
						agregar = False
				if agregar:		
					materialesAgrupados.append(mat)				

			for obj in materialesAgrupados:
				obj['cantidad'] = round(obj['cantidad'],2)
				obj['cantidad_ejecutada'] = round(obj['cantidad_ejecutada'],2)
				obj['pendiente'] = round(obj['cantidad'] - obj['cantidad_ejecutada'],2)
		

		if listado:
			datos=materialesAgrupados			
		else:
			datos=[]		
		
		#import pdb; pdb.set_trace()
		proyecto = request.GET['proyecto'] if request.GET['proyecto'] else 0
		proyecto_model = Proyecto.objects.get(pk=int(proyecto))

		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte seguimiento materiales -'+proyecto_model.mcontrato.nombre+'.xls"'
		
		

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		
		
		if len(proyecto_model.nombre)<30:
			worksheet = workbook.add_worksheet(proyecto_model.nombre)
		else:
			worksheet = workbook.add_worksheet(proyecto_model.municipio.nombre)

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:B', 25)
		worksheet.set_column('C:C', 40)
		worksheet.set_column('D:F', 25)
		row=1
		col=0

		worksheet.write('A1', 'Hito', format1)
		worksheet.write('B1', 'Codigo', format1)
		worksheet.write('C1', 'Descripcion', format1)
		worksheet.write('D1', 'Cantidad a ejecutar', format1)
		worksheet.write('E1', 'Cantidad ejecutada', format1)
		worksheet.write('F1', 'Cantidad pendiente', format1)

		if datos:
			for item in datos:
				worksheet.write(row, col,item['hito'],format3)
				worksheet.write(row, col+1,item['codigo'],format3)
				worksheet.write(row, col+2,item['descripcion'],format3)
				worksheet.write(row, col+3,item['cantidad'],format3)
				worksheet.write(row, col+4,item['cantidad_ejecutada'],format3)
				worksheet.write(row, col+5,item['pendiente'],format3)
				row +=1

		workbook.close()
		return response
	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.exportarMaterialesaLiquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
@login_required
def exportarCantidadesaLiquidar(request):
	try:
		datos = []
		presupuesto_id= request.GET['presupuesto_id'] if request.GET['presupuesto_id'] else None
		cantidadesEjecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__id=int(presupuesto_id),
			cantidad__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(
			total=Sum('cantidad'))

		cantidadesEjecutadas = KDetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = int(presupuesto_id),
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			#'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))



		##import pdb; pdb.set_trace()	
		for cantidadEjecutar in cantidadesEjecutar:
			cejecutar = 0
			cliquidada = 0
			liquidada = False
			for cantidadEjecutada in cantidadesEjecutadas:
				if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
					cejecutar = cantidadEjecutada['total']					
					break


			datos.append(
				{
					'hito': cantidadEjecutar['detallepresupuesto__actividad__nombre'],
					'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
					'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
					'ejecutar': cantidadEjecutar['total'],
					'ejecutado': cejecutar,
					'pendiente': int(cantidadEjecutar['total']) - int(cejecutar)
				}
			)
		
		#import pdb; pdb.set_trace()
		proyecto = request.GET['proyecto'] if request.GET['proyecto'] else 0
		proyecto_model = Proyecto.objects.get(pk=int(proyecto))

		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte seguimiento cantidades -'+proyecto_model.mcontrato.nombre+'.xls"'
		
		

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		if len(proyecto_model.nombre)<30:
			worksheet = workbook.add_worksheet(proyecto_model.nombre)
		else:
			worksheet = workbook.add_worksheet(proyecto_model.municipio.nombre)

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:B', 25)
		worksheet.set_column('C:C', 40)
		worksheet.set_column('D:F', 25)
		row=1
		col=0

		worksheet.write('A1', 'Hito', format1)
		worksheet.write('B1', 'Codigo UU CC', format1)
		worksheet.write('C1', 'Descripcion UU CC', format1)
		worksheet.write('D1', 'Cantidad a ejecutar', format1)
		worksheet.write('E1', 'Cantidad ejecutada', format1)
		worksheet.write('F1', 'Cantidad pendiente', format1)

		if datos:
			for item in datos:
				worksheet.write(row, col,item['hito'],format3)
				worksheet.write(row, col+1,item['codigoUC'],format3)
				worksheet.write(row, col+2,item['descripcion'],format3)
				worksheet.write(row, col+3,item['ejecutar'],format3)
				worksheet.write(row, col+4,item['ejecutado'],format3)
				worksheet.write(row, col+5,item['pendiente'],format3)
				row +=1

		workbook.close()
		return response
	except Exception as e:
			functions.toLog(e,'avanceObraGrafico2.exportarCantidadesaLiquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})

@login_required
def descargar_plantilla_programacion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_programacion.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Programacion')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})
	format5=workbook.add_format({'border':1})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_align('center')

	esquema_id= request.GET['esquema_id']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 50)
	worksheet.set_column('C:C', 50)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 20)
	worksheet.set_column('F:F', 20)

	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Actividad inicial', format1)
	worksheet.write('E1', 'Fecha de Inicio', format1)
	worksheet.write('F1', 'Fecha Final', format1)


	actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id,padre__gte=1).order_by('id')
	row=1

	for item_act in actividades:
		worksheet.write(row,0, item_act.id, format2)		

		padre = CEsquemaCapitulosActividadesG.objects.get(pk=item_act.padre)
		worksheet.write(row,1, padre.nombre, format2)

		worksheet.write(row,2, item_act.nombre, format2)
		if row==1:
			worksheet.write(row,3, 'Si', format2)
		else:
			worksheet.write(row,3, 'No', format2)
		worksheet.write(row,4, '', format5)
		worksheet.write(row,5, '', format5)
		
		row+=1
	workbook.close()

	return response

@login_required
def guardar_programacion_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		esquema_id= request.POST['esquema_id']
		cronograma_id= request.POST['cronograma_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1
		numeroFila = 1
		
		if int(contador) > 0:
			#sid = transaction.savepoint()
			

			validad_positiva_actividad_inicial = False
			cont_si=0
			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						
						if CEsquemaCapitulosActividadesG.objects.filter(id=fila[0].value,esquema__id=esquema_id).count() == 0:
							return JsonResponse({
								'message':'No se encontro el Id indicado en la fila No.' + str(numeroFila) + \
								'. Sugerimos descargar nuevamente la plantilla para verificar el Id en esta fila',
								'success':'error',
								'data':''})

						act_model = CEsquemaCapitulosActividadesG.objects.get(id=fila[0].value,esquema__id=esquema_id)

						if act_model.nombre!=fila[2].value:
							return JsonResponse({
								'message':'No coindicen entre si el Id y la Actividad en la fila No.' + str(numeroFila) + \
								'. Sugerimos descargar nuevamente la plantilla para verificar el Id y la Actividad en esta fila',
								'success':'error',
								'data':''})

						

						if CEsquemaCapitulosActividadesG.objects.filter(nombre=fila[1].value,esquema__id=esquema_id).count() == 0:
							return JsonResponse({
								'message':'No se encontro el Hito indicado en la fila No.' + str(numeroFila) + \
								'. Sugerimos descargar nuevamente la plantilla para verificar el Hito en esta fila',
								'success':'error',
								'data':''})
						else:
							if CEsquemaCapitulosActividadesG.objects.filter(id=act_model.padre,esquema__id=esquema_id).count() >0:
								padre_model = CEsquemaCapitulosActividadesG.objects.get(id=act_model.padre,esquema__id=esquema_id)
							
								if padre_model.nombre != fila[1].value:
									return JsonResponse({
										'message':'No coincide el Hito con la Actividad en la fila No.' + str(numeroFila) + \
										'. Sugerimos descargar nuevamente la plantilla para verificar el Hito y la Actividad en esta fila',
										'success':'error',
										'data':''})


						actividad_inicial = fila[3].value
						if actividad_inicial=='Si' or actividad_inicial=='SI' or actividad_inicial=='sI':
							validad_positiva_actividad_inicial = True
							cont_si +=1

						if str(actividad_inicial)!='Si' and str(actividad_inicial)!='No' and str(actividad_inicial)!='SI' and str(actividad_inicial)!='NO' and str(actividad_inicial)!='sI' and str(actividad_inicial)!='nO':
							return JsonResponse({
								'message':'Solo se puede recibir "Si" o "No" en la actividad inicial indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos corregir la plantilla en esta fila y volver a subirla al sistema',
								'success':'error',
								'data':''})

						#import pdb; pdb.set_trace()
						try:
							fecha_inicio = fila[4].value
							aux = fecha_inicio.strftime('%Y-%m-%d')

						except Exception as e:
							return JsonResponse({
								'message':'Hubo un error con la fecha de inicio indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos corregir la plantilla en esta fila y volver a subirla al sistema',
								'success':'error',
								'data':''})

						try:
							fecha_fin = fila[5].value
							aux = fecha_fin.strftime('%Y-%m-%d')

						except Exception as e:
							return JsonResponse({
								'message':'Hubo un error con la fecha de final indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos corregir la plantilla en esta fila y volver a subirla al sistema',
								'success':'error',
								'data':''})

						if fecha_inicio>fecha_fin:
							return JsonResponse({
								'message':'La fecha de inicio es posterior a la fecha de fin indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos corregir la plantilla en esta fila y volver a subirla al sistema',
								'success':'error',
								'data':''})


					numeroFila = numeroFila + 1

			numeroFila = 1
			if validad_positiva_actividad_inicial and cont_si==1:
				i=0
				for fila in hoja.rows:
					if fila[0].value:
						if i == 0:
							i=1
						else:
							act_model = CEsquemaCapitulosActividadesG.objects.get(id=fila[0].value,esquema__id=esquema_id)
							act_inicial = ''
							if fila[3].value=='Si' or fila[3].value=='SI' or fila[3].value=='sI':
								act_inicial = True
							else:
								act_inicial = False

							if DiagramaGrahm.objects.filter(actividad__id=fila[0].value,cronograma__id=cronograma_id).count()==0:
								

								diagrama_model = DiagramaGrahm(
									fechaInicio=fila[4].value,
									fechaFinal=fila[5].value,
									actividad_inicial=act_inicial,
									actividad_id=int(fila[0].value),
									cronograma_id=int(cronograma_id))
								diagrama_model.save()

								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avance_de_obra_grafico2.DiagramaGrahm(carga_masiva)',id_manipulado=diagrama_model.id)
								logs_model.save()

							elif DiagramaGrahm.objects.filter(actividad__id=fila[0].value,cronograma__id=cronograma_id).count()==1:
								diagrama_model = DiagramaGrahm.objects.get(actividad__id=fila[0].value,cronograma__id=cronograma_id)
								diagrama_model.fechaInicio=fila[4].value
								diagrama_model.fechaFinal=fila[5].value
								diagrama_model.actividad_inicial=act_inicial
								diagrama_model.save()

								logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avance_de_obra_grafico2.DiagramaGrahm(carga_masiva)',id_manipulado=diagrama_model.id)
								logs_model.save()

							else:
								return JsonResponse({
									'message':'Dentro de esta programacion se encontro mas de una vez la actividad indicada en la fila No. ' + str(numeroFila) + \
									'. Sugerimos ponerse en contacto con el soporte tecnico de SININ',
									'success':'error',
									'data':''})
							
							

						numeroFila = numeroFila + 1
								
			elif validad_positiva_actividad_inicial and cont_si>1:
				return JsonResponse({
					'message':'Se especifico mas de una actividad inicial' + \
					'. Sugerimos especificar solo una actividad inicial y subir la plantilla nuevamente al sistema',
					'success':'error',
					'data':''})
			else:
				return JsonResponse({
					'message':'No se especifico ninguna actividad inicial' + \
					'. Sugerimos especificar la actividad inicial, con un "Si" en la columna D y subir la plantilla nuevamente al sistema',
					'success':'error',
					'data':''})
		#transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		#transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.DiagramaGrahm(carga_masiva)')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
def excel_catalogo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_catalogo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Catalogos')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 12)	

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	dato = request.GET['dato'] if 'dato' in request.GET else None
	#import pdb; pdb.set_trace()	
	if dato:
		qsetFilter = qsetFilter & (Q(nombre__icontains=dato))	

	qset = CatalogoUnidadConstructiva.objects.filter(qsetFilter)	

	formato_fecha = "%Y-%m-%d"

	if qset:
		worksheet.write('A1', 'Nombre', format1)
		worksheet.write('B1', 'Anio', format1)

	for catalogo in qset:
		worksheet.write(row, col, catalogo.nombre,format2)
		worksheet.write(row, col+1, catalogo.ano,format2)					
		row +=1

	workbook.close()
	return response

@login_required
def excel_uucc(request):
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_uucc.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('UUCC')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format4=workbook.add_format()
	format4.set_num_format("{num_format':'$#,##0.00'}")
	format4.set_align('center')
	format4.set_align('vcenter')	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 30)	

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None	
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))			
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = UnidadConstructiva.objects.filter(qsetFilter)	
	serializer = UnidadConstructivaSerializerLite(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Total MO', format1)
		worksheet.write('D1', 'Valor Total Mat', format1)
		worksheet.write('E1', 'Tipo UUCC', format1)
		#import pdb; pdb.set_trace()	
		for uucc in serializer.data:
			worksheet.write(row, col, uucc['codigo'],format2)
			worksheet.write(row, col+1, uucc['descripcion'],format2)					
			worksheet.write(row, col+2, float(uucc['totalManoDeObra']),format2)					
			worksheet.write(row, col+3, float(uucc['totalMateriales']),format2)					
			worksheet.write(row, col+4, uucc['tipoUnidadConstructiva']['nombre'],format2)							
			row +=1

	workbook.close()
	return response

def excel_mat(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_materiales.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Materiales')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))	
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = Material.objects.filter(qsetFilter)	
	serializer = MaterialSerializerLite2(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Unitario', format1)		
		worksheet.write('D1', 'Unidad de Medida', format1)		
		for material in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, material['codigo'],format2)
			worksheet.write(row, col+1, material['descripcion'],format2)	
			if material['valorUnitario']:
				worksheet.write(row, col+2, float(material['valorUnitario']),format2)								
			else:
				worksheet.write(row, col+2, '0.00',format2)								
			worksheet.write(row, col+3, material['unidadMedida'],format2)							
			row +=1

	workbook.close()
	return response

def excel_mo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_MO.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Mano de Obra')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))	
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = ManoDeObra.objects.filter(qsetFilter)	
	serializer = ManoObraSerializerLite2(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Hora', format1)						
		for material in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, material['codigo'],format2)
			worksheet.write(row, col+1, material['descripcion'],format2)	
			if material['valorHora']:
				worksheet.write(row, col+2, float(material['valorHora']),format2)								
			else:
				worksheet.write(row, col+2, '0.00',format2)																	
			row +=1

	workbook.close()
	return response

def excel_desgl_mat(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_Desgl_Mat.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Desglose Material')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	uucc_id = request.GET['uucc_id'] if 'uucc_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(material__descripcion__icontains=dato) | Q(material__codigo__icontains=dato))
	if uucc_id:
		qsetFilter = qsetFilter & (Q(unidadConstructiva_id=uucc_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = DesgloceMaterial.objects.filter(qsetFilter)	
	serializer = DesgloceMaterialSerializer(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo UUCC', format1)
		worksheet.write('B1', 'Descripcion UUCC', format1)
		worksheet.write('C1', 'Codigo Material', format1)
		worksheet.write('D1', 'Descripcion Material', format1)	
		worksheet.write('E1', 'Cantidad', format1)	

		for dm in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, dm['unidadConstructiva']['codigo'],format2)
			worksheet.write(row, col+1, dm['unidadConstructiva']['descripcion'],format2)	
			worksheet.write(row, col+2, dm['material']['codigo'],format2)											
			worksheet.write(row, col+3, dm['material']['descripcion'],format2)	
			worksheet.write(row, col+4, dm['cantidad'],format2)											
			row +=1

	workbook.close()
	return response

def excel_desgl_mo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_desgl_MO.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Desglose Material')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	uucc_id = request.GET['uucc_id'] if 'uucc_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(manoDeObra__descripcion__icontains=dato) | Q(manoDeObra__codigo__icontains=dato)) 
	if uucc_id:
		qsetFilter = qsetFilter & (Q(unidadConstructiva_id=uucc_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = DesgloceManoDeObra.objects.filter(qsetFilter)	
	serializer = DesgloceManoDeObraSerializer(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo UUCC', format1)
		worksheet.write('B1', 'Descripcion UUCC', format1)
		worksheet.write('C1', 'Codigo MO', format1)
		worksheet.write('D1', 'Descripcion MO', format1)	
		worksheet.write('E1', 'Rendimiento', format1)	

		for dmo in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, dmo['unidadConstructiva']['codigo'],format2)
			worksheet.write(row, col+1, dmo['unidadConstructiva']['descripcion'],format2)	
			worksheet.write(row, col+2, dmo['manoDeObra']['codigo'],format2)											
			worksheet.write(row, col+3, dmo['manoDeObra']['descripcion'],format2)	
			worksheet.write(row, col+4, dmo['rendimiento'],format2)											
			row +=1

	workbook.close()
	return response


@login_required
def inactivar_catalogo(request):
	sid = transaction.savepoint()
	try:		
		#import pdb; pdb.set_trace()	
		id_catalogo = request.GET['id'] if request.GET['id'] else None
		if id_catalogo:
			catalogo = CatalogoUnidadConstructiva.objects.get(pk=id_catalogo)
			if catalogo:
				catalogo.activo = 0
				catalogo.save()		
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok','data':''})		
		else:
			return JsonResponse({'message':'No se encontraron registros','success':'ok','data':''})			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.inactivar_catalogo')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})			

@login_required
def activar_catalogo(request):
	sid = transaction.savepoint()
	try:		
		#import pdb; pdb.set_trace()	
		id_catalogo = request.GET['id'] if request.GET['id'] else None
		if id_catalogo:
			catalogo = CatalogoUnidadConstructiva.objects.get(pk=id_catalogo)
			if catalogo:
				catalogo.activo = 1
				catalogo.save()		
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok','data':''})		
		else:
			return JsonResponse({'message':'No se encontraron registros','success':'ok','data':''})			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra_grafico2.activar_catalogo')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})					


@login_required
def descargar_plantilla_masiva(request):
	return functions.exportarArchivoS3('plantillas/catalogos/PlantillaCargaMasivaCatalogo.xlsx')



@transaction.atomic
def cargar_excel_catalogo_masivo(request):	
	#import pdb; pdb.set_trace()		
	try:
		array_uucc = []
		array_mat = []
		array_mo = []	
		catalogo_id= request.POST['catalogo_id']
		soporte= request.FILES['archivo']
		doc = openpyxl.load_workbook(soporte)
		
		#Validacion desgl_mat y desgl_mo
		index=0
		hoja = doc.get_sheet_by_name('UUCC')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_uucc.append(fila[0].value)
			index+=1


		index=0
		hoja = doc.get_sheet_by_name('Material')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_mat.append(fila[0].value)
			index+=1

		index=0
		hoja = doc.get_sheet_by_name('Mano de Obra')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_mo.append(fila[0].value)	
			index+=1	

		index=0		
		hoja = doc.get_sheet_by_name('Desgl_Material')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None:
					if fila[0].value not in array_uucc:
						return JsonResponse({'message':'La UUCC con codigo '+str(fila[0].value)+' en Desgl_Material, no se encuentra dentro del listado de UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
					
					if fila[1].value not in array_mat:
						return JsonResponse({'message':'El material con codigo '+str(fila[1].value)+' en Desgl_Material, no se encuentra dentro del listado de Material','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			
											
			index+=1
		
		index=0		
		hoja = doc.get_sheet_by_name('Desgl_MO')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None:
					if fila[0].value not in array_uucc:
						return JsonResponse({'message':'La UUCC con codigo '+str(fila[0].value)+' en Desgl_MO, no se encuentra dentro del listado de UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
					
					if fila[1].value not in array_mo:
						return JsonResponse({'message':'La Mano de Obra con codigo '+str(fila[1].value)+' en Desgl_MO, no se encuentra dentro del listado de Mano de Obra','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			
											
			index+=1													

		#Fin Validacion

		#UUCC		
		hoja = doc.get_sheet_by_name('UUCC')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc = UnidadConstructiva.objects.filter(qset).first()

					if uucc is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						tipoUUCC = fila[2].value
						uucc2 = UnidadConstructiva(codigo=codigo,descripcion=descripcion,tipoUnidadConstructiva_id=tipoUUCC,catalogo_id=catalogo_id)
						uucc2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1

		
		#Material
		hoja = doc.get_sheet_by_name('Material')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None and fila[3].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					mat = Material.objects.filter(qset).first()

					if mat is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						valorUnitario = fila[2].value
						UnidadDeMedida = str(fila[3].value)
						mat2 = Material(codigo=codigo,descripcion=descripcion,valorUnitario=valorUnitario,unidadMedida=UnidadDeMedida,catalogo_id=catalogo_id)
						mat2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1


		#Mo
		hoja = doc.get_sheet_by_name('Mano de Obra')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					mo = ManoDeObra.objects.filter(qset).first()

					if mo is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						valorHora = fila[2].value
						mo2 = ManoDeObra(codigo=codigo,descripcion=descripcion,valorHora=valorHora,catalogo_id=catalogo_id)
						mo2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1


		#Desgl_Material
		hoja = doc.get_sheet_by_name('Desgl_Material')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qsetUUCC = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc_id = UnidadConstructiva.objects.filter(qsetUUCC).values('id').first()
					
					qsetMat = (Q(codigo = str(fila[1].value)) & Q(catalogo_id = catalogo_id))
					mat_id = Material.objects.filter(qsetMat).values('id').first()

					if uucc_id is not None and mat_id is not None:
						qsetDegl_mat = (Q(material_id = mat_id['id']) & Q(unidadConstructiva_id = uucc_id['id']))						
						desgl_mat = DesgloceMaterial.objects.filter(qsetDegl_mat)

						if any(desgl_mat) is False:
							desgl_mat2 = DesgloceMaterial(unidadConstructiva_id=uucc_id['id'],material_id=mat_id['id'],cantidad=fila[2].value)
							desgl_mat2.save()

					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1						

		#Desgl_MO
		hoja = doc.get_sheet_by_name('Desgl_MO')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qsetUUCC = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc_id = UnidadConstructiva.objects.filter(qsetUUCC).values('id').first()
					
					qsetMo = (Q(codigo = str(fila[1].value)) & Q(catalogo_id = catalogo_id))
					mo_id = ManoDeObra.objects.filter(qsetMo).values('id').first()

					if uucc_id is not None and mo_id is not None:
						qsetDegl_mo = (Q(manoDeObra_id = mo_id['id']) & Q(unidadConstructiva_id = uucc_id['id']))						
						desgl_mo = DesgloceManoDeObra.objects.filter(qsetDegl_mo)

						if any(desgl_mo) is False:
							desgl_mo2 = DesgloceManoDeObra(unidadConstructiva_id=uucc_id['id'],manoDeObra_id=mo_id['id'],rendimiento=fila[2].value)
							desgl_mo2.save()

					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1						
		
		return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response

