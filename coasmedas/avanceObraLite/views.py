# -*- coding: utf-8 -*- 
from django.shortcuts import render,redirect
from .models import APeriodicidadG,BEsquemaCapitulosG,CEsquemaCapitulosActividadesG,Cronograma
from .models import EPresupuesto,FDetallePresupuesto, PeriodoProgramacion, DetallePeriodoProgramacion, ReporteTrabajo, DetallePeriodoProgramacion
from .models import UnidadConstructiva, ManoDeObra, Material, DesgloceManoDeObra, DesgloceMaterial
from .models import TipoUnidadConstructiva, CatalogoUnidadConstructiva, DetalleReporteTrabajo

# from descargo.models import Permiso_validez
from rest_framework import viewsets, serializers
from django.db.models import Q,Sum,Prefetch
from logs.models import Logs,Acciones
from usuario.models import Usuario
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination

import xlsxwriter
from xlsxwriter.utility import xl_col_to_name

import json
from datetime import date, datetime, timedelta
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.deletion import ProtectedError

from proyecto.views import ProyectoLiteSerializer
from proyecto.models import Proyecto,Proyecto_empresas

from contrato.models import EmpresaContrato,Contrato
from contrato.views import ContratoSerializer
from contrato.enumeration import tipoC

from .tasks import updateAsyncEstado,agregarSinPoste

from tipo.models import Tipo

from empresa.models import Empresa

from parametrizacion.models import Departamento

from django.db import connection

from datetime import timedelta
import time

from datetime import date
from datetime import datetime



import math

from usuario.views import UsuarioSerializer, UsuarioLiteSerializer
from empresa.views import EmpresaSerializer

from estado.models import Estado
from estado.views import EstadoSerializer, EstadoLiteSerializer

from tipo.views import TipoSerializer
from tipo.models import Tipo

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view

from django.conf import settings
from sinin4.functions import functions
from django.db import transaction
from parametrizacion.views import  MunicipioSerializer , DepartamentoSerializer
from puntos_gps.views import PuntosGps
from control_cambios.views import CCambio
from p_p_construccion.models import ALote
from django.db.models import Q, Sum, Max
from avance_de_obra.models import BCronograma
from .models import CatalogoUnidadConstructiva, Material
import locale
# locale.setlocale( locale.LC_ALL, 'en_US' )

import openpyxl
#from utilidades.funciones import funciones
# Create your views here.

#serializer de empresa
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista')


#serializer lite de proyecto
class ProyectoLite2Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	municipio = MunicipioSerializer(read_only = True)
	contrato = ContratoLiteSerializer(read_only = True, many = True)
	conteo_puntos=serializers.SerializerMethodField()
	cantidad_cambio=serializers.SerializerMethodField()
	cantidad_lote=serializers.SerializerMethodField()

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'mcontrato' ,
				 'municipio' ,
				 'contrato',
				 'conteo_puntos',
 				 'cantidad_cambio',
				 'cantidad_lote'
				 )

	def get_conteo_puntos(self, obj):
		return PuntosGps.objects.filter(proyecto_id=obj.id).count()

	def get_cantidad_cambio(self, obj):
		return CCambio.objects.filter(proyecto_id=obj.id).count()

	def get_cantidad_lote(self, obj):
		return ALote.objects.filter(proyecto_id=obj.id).count()
	
#serializer lite de proyecto empresa
class ProyectoEmpresaLite4Serializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLite2Serializer(read_only = True )
	empresa = EmpresaLiteSerializer(read_only = True  )

	totalCronograma=serializers.SerializerMethodField()

	totalPresupuesto=serializers.SerializerMethodField()

	totalCronogramaGrafico=serializers.SerializerMethodField()

	porcentajeAvance=serializers.SerializerMethodField()

	porcentajeAvanceFinanciero=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id' ,
				'proyecto',
 				'empresa',
 				'totalCronograma',
 				'totalPresupuesto',
 				'totalCronogramaGrafico',		 
 				'porcentajeAvance',
 				'porcentajeAvanceFinanciero'
				)

	def get_totalCronograma(self, obj):
		return Cronograma.objects.filter(proyecto_id=obj.proyecto.id).count()

	def get_totalCronogramaGrafico(self, obj):
		return Cronograma.objects.filter(proyecto_id=obj.proyecto.id).count()

	def get_totalPresupuesto(self, obj):
		#from avanceObraGrafico.models import EPresupuesto
		return EPresupuesto.objects.filter(cronograma__proyecto_id=obj.proyecto.id).count()

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []
		avanzo = False
		corte = 'Sin avance'

		# import pdb; pdb.set_trace()
		queryset_a_ejecutar = FDetallePresupuesto.objects.filter(
			presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'actividad__id',
			'actividad__nombre',
			'actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
			detallePresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'detallePresupuesto__actividad__id',
			'detallePresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidad'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['actividad__id'],
						'actividad': Aejecutar['actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['actividad__id'],
					'actividad': Aejecutar['actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['actividad__peso']
				})
		
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		# obtenemos la fecha corte a la cual le restaremos la periodicidad para determinar
		# si hubo avance en el proyecto	
		aux = ReporteTrabajo.objects.filter(periodoProgramacion__cronograma__proyecto__id=obj.proyecto.id).order_by('fechaReporte').last()
		if aux:
			corte = aux.fechaReporte
		# Determinamos si aumento o no aumento el avance tomando como referencia del calculo
		# la periodicidad del cronograma.
		avance_anterior=0
		porcentajes = []
		#import pdb; pdb.set_trace()
		if corte != 'Sin avance':
			fecha = datetime.today().date()
			cronograma = Cronograma.objects.filter(
				proyecto__id = obj.proyecto.id).values('periodicidad__numero_dias')
			if cronograma:
				fecha = fecha - timedelta(days=int(cronograma[0]['periodicidad__numero_dias']))
				corte1 = fecha.strftime('%Y-%m-%d')
				queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
					detallePresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
					cantidad__gt=0,
					reporteTrabajo__fechaReporte__lte=corte1).values(
					'detallePresupuesto__actividad__id',
					'detallePresupuesto__actividad__nombre'
					).annotate(total=Sum('cantidad'))
				for Aejecutar in queryset_a_ejecutar:
					agregado = False
					for ejecutado in queryset_ejecutada:
						if Aejecutar['actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
							agregado = True
							porcentajes.append({
								'id' : Aejecutar['actividad__id'],
								'actividad': Aejecutar['actividad__nombre'],
								'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
								'peso': Aejecutar['actividad__peso']
							})
						if agregado:
							break
					if agregado == False:
						porcentajes.append({
							'id' : Aejecutar['actividad__id'],
							'actividad': Aejecutar['actividad__nombre'],
							'porcentaje': 0,
							'peso': Aejecutar['actividad__peso']
						})

				for p in porcentajes:
					avance_anterior = avance_anterior + (float(p['porcentaje']) * float(float(p['peso'])/100))

		if avance_anterior < avance:
			avanzo = True

		# return round(avance,2)
		return {'avance': round(avance,2), 'avanzo' : avanzo, 'corte':corte,}
		# return 0

	def get_porcentajeAvanceFinanciero(self,obj):		
		avance = 0		

		presupuesto = EPresupuesto.objects.filter(cronograma__proyecto__id=obj.proyecto.id,cerrar_presupuesto=True)
		aiu = 1
		if presupuesto:
			aiu = presupuesto.first().aiu
			presupuesto = presupuesto.first().suma_presupuesto()			
		else:
			presupuesto = 0


		queryset_a_ejecutar = DetalleReporteTrabajo.objects.filter(
			detallePresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values('cantidad','detallePresupuesto__valorGlobal')


		for item in queryset_a_ejecutar:

			subtotal_aux = float(item['cantidad'])*float(item['detallePresupuesto__valorGlobal'])

			peso_aux = (float(subtotal_aux)*float(aiu))/float(presupuesto*aiu)			

			avance += peso_aux
			# avance += item.peso()
			# Aqui ando
		
		return round(avance*100,2)


class ProyectoEmpresaLite3Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = serializers.SerializerMethodField('_mcontrato',read_only=True)
	porcentajeAvance=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('mcontrato',		 
 				'porcentajeAvance',
				)

	def _mcontrato(self,obj):
		return  {'id': obj.proyecto.mcontrato.id,'nombre':obj.proyecto.mcontrato.nombre}

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []

		# import pdb; pdb.set_trace()
		queryset_a_ejecutar = FDetallePresupuesto.objects.filter(
			presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'actividad__id',
			'actividad__nombre',
			'actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
			detallePresupuesto__presupuesto__cronograma__proyecto__id=obj.proyecto.id,
			cantidad__gt=0).values(
			'detallePresupuesto__actividad__id',
			'detallePresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidad'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['actividad__id'],
						'actividad': Aejecutar['actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['actividad__id'],
					'actividad': Aejecutar['actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['actividad__peso']
				})
		
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		return round(avance,2)
		# return 0



class ProyectoEmpresaLite5Serializer(serializers.HyperlinkedModelSerializer):
	proyecto = ProyectoLite2Serializer(read_only = True )	
	# cantSoportes = serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id',		 
 				'proyecto',
				# 'cantSoportes'
				)

	# def get_cantSoportes(self,obj):		
	# 	cant = Permiso_validez.objects.filter(proyecto_id=obj.proyecto.id).count()							
	# 	return cant





#Api lite de empresa proyecto
class ProyectoEmpresaLiteViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las categorias de administrador de fotos
	"""
	model=Proyecto_empresas
	queryset = model.objects.all()
	serializer_class = ProyectoEmpresaLite4Serializer
	nombre_modulo='avanceObraLite.Proyecto_empresa'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		try:
			queryset = super(ProyectoEmpresaLiteViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			empresa = self.request.query_params.get('empresa', None)
			mcontrato = self.request.query_params.get('mcontrato', None)
			departamento = self.request.query_params.get('departamento', None)
			municipio = self.request.query_params.get('municipio', None)
			contratista = self.request.query_params.get('contratista', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			isLite = self.request.query_params.get('isLite', None)
			homepage = self.request.query_params.get('homepage', None)
			qset=(~Q(id=0))

			if(dato or empresa or mcontrato or departamento or municipio):

				if dato:
					qset = qset & ( Q(empresa__nombre__icontains = dato) |
								Q(empresa__nit__icontains = dato) |
								Q(proyecto__nombre__icontains = dato))					
			

				if empresa and int(empresa)>0:
					qset = qset & (Q(empresa__id = empresa))					


				if mcontrato and int(mcontrato)>0:
					qset = qset & (Q(proyecto__mcontrato = mcontrato))
					
				if departamento and int(departamento)>0:
					qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
					
				if municipio and int(municipio)>0:
					qset = qset & (Q(proyecto__municipio = municipio))
					

				if contratista and int(contratista)>0:
					qset = qset & (Q(proyecto__contrato__contratista__id = contratista))	

			
			if qset != '':
				queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if isLite:
						serializer = ProyectoEmpresaLite5Serializer(page,many=True, context={'request': request})
					elif homepage:
						serializer = ProyectoEmpresaLite3Serializer(page,many=True, context={'request': request})
					else:
						serializer = self.get_serializer(page,many=True)
					
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#Api rest para periodicidad
class PeriodicidadSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = APeriodicidadG
		fields=('id','nombre','numero_dias')


class PeriodicidadGraficoViewSet(viewsets.ModelViewSet):
	
	model=APeriodicidadG
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='avanceObraLite.periodicidad'
	queryset = model.objects.all()
	serializer_class = PeriodicidadSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PeriodicidadGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PeriodicidadSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PeriodicidadSerializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para periodicidad


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre', 'contratista',)



#Api rest para Esquema de Capitulos
class EsquemaCapitulosSerializer(serializers.HyperlinkedModelSerializer):

	tipo=tipoC()
	macrocontrato=ContratoLiteSerializer(read_only=True)
	macrocontrato_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.filter(tipo_contrato=tipo.m_contrato))	

	class Meta:
		model = BEsquemaCapitulosG
		fields=('id','macrocontrato','macrocontrato_id','nombre')


class EsquemaCapitulosGraficoViewSet(viewsets.ModelViewSet):
	
	model=BEsquemaCapitulosG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.esquema_capitulos'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			macrocontrato_id= self.request.query_params.get('macrocontrato_id',None)
			lista_contrato=[]
			tipo=tipoC()
			mcontrato=EmpresaContrato.objects.filter(contrato__tipo_contrato=tipo.m_contrato,empresa=request.user.usuario.empresa.id,participa=1)

			if mcontrato is None:
				lista_contrato.append(0)
			else:
				for item in mcontrato:
					lista_contrato.append(item.contrato.id)

			qset=(Q(macrocontrato_id__in=lista_contrato))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if macrocontrato_id is not None and int(macrocontrato_id)>0:
				qset=qset&(Q(macrocontrato_id=macrocontrato_id))


			
			
			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosSerializer(data=request.data,
					context={'request': request})

				if serializer.is_valid():
					serializer.save(macrocontrato_id=request.data['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,
						accion=self.model_acciones.accion_crear,
						nombre_modelo=self.nombre_modulo,
						id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)			
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(macrocontrato_id=request.data['macrocontrato_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos



#Api rest para Esquema de Capitulos de las actividades
class EsquemaCapitulosActividadesSerializer(serializers.HyperlinkedModelSerializer):

	esquema=EsquemaCapitulosSerializer(read_only=True, required = False)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all())	

	
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','esquema','esquema_id','nombre','nivel','padre','peso')


class EsquemaCapitulosActividadesLiteSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = CEsquemaCapitulosActividadesG
		fields=('id','nombre')


class EsquemaCapitulosActividadesGraficoViewSet(viewsets.ModelViewSet):
	
	model=CEsquemaCapitulosActividadesG
	queryset = model.objects.all()
	serializer_class = EsquemaCapitulosActividadesSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.esquema_capitulos_actividades'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EsquemaCapitulosActividadesGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			niveles= self.request.query_params.get('niveles',None)
			padre_id= self.request.query_params.get('padre_id',None)
			qset=None


			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if esquema_id is not None and int(esquema_id)>0:
				qset=qset &(Q(esquema_id=esquema_id))


			if padre_id is not None and int(padre_id)>0:
				qset=qset &(Q(padre=padre_id))


			if niveles is not None:
				lista=niveles.split(',')
				qset=qset&(Q(nivel__in=lista))

			if qset is not None:				
				queryset = self.model.objects.filter(qset)
			


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EsquemaCapitulosActividadesSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.data['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
					valor=round(float(model_actividad['peso__sum']),3)+round(float(request.data['peso']),3)
					
					if float(valor) <= 100: 
						serializer.save(esquema_id=request.data['esquema_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.data['padre']
						
						if int(padre)>0:
							if int(request.data['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(request.data['peso'])
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()								
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(request.data['peso'])
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EsquemaCapitulosActividadesSerializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					model_actividad=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=request.data['esquema_id'],nivel=1).aggregate(Sum('peso'))
					model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']

					model_actividad2=CEsquemaCapitulosActividadesG.objects.get(pk=instance.id)
					valor_restante=float(request.data['peso']) - float(model_actividad2.peso)
					valor=float(model_actividad['peso__sum'])+valor_restante

					if float(valor) <= 100: 
						serializer.save(esquema_id=request.data['esquema_id'])					
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
						padre=request.data['padre']

						if int(padre)>0:
							if int(request.data['nivel'])==3:
								valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
								valor=float(valores.peso)+float(valor_restante)
								valores.peso=valor
								valores.save()
								logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
								logs_model.save()
								padre=valores.padre

							valores=CEsquemaCapitulosActividadesG.objects.get(pk=padre)
							valor=float(valores.peso)+float(valor_restante)
							valores.peso=valor
							valores.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=padre)
							logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Esquema de Capitulos de las actividades



#serializer lite de proyecto
class ProyectoLite2Serializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	municipio = MunicipioSerializer(read_only = True)
	contrato = ContratoLiteSerializer(read_only = True, many = True)

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'mcontrato' ,
				 'municipio' ,
				 'contrato',
				 )


#Api rest para Cronograma

class CronogramaSerializerLite(serializers.HyperlinkedModelSerializer):
	detalles_cargados=serializers.SerializerMethodField()	
	periodicidad=PeriodicidadSerializer(read_only=True)

	class Meta:
		model = Cronograma
		fields = ('id','nombre','fechaInicio','fechaFinal','confirmarFechas','detalles_cargados','periodicidad')

	def get_detalles_cargados(self,obj):
		return DetallePeriodoProgramacion.objects.filter(periodoProgramacion__cronograma_id=obj.id).exists() 

class CronogramaSerializer(serializers.HyperlinkedModelSerializer):

	proyecto=ProyectoLite2Serializer(read_only=True)
	proyecto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all())	

	# estado=ReglaEstadoGraficoSerializer(read_only=True, allow_null = True)
	# estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CReglasEstadoG.objects.all(),allow_null = True)	

	periodicidad=PeriodicidadSerializer(read_only=True, allow_null = True)
	periodicidad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=APeriodicidadG.objects.all(),allow_null = True)	

	esquema=EsquemaCapitulosSerializer(read_only=True, allow_null = True)
	esquema_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=BEsquemaCapitulosG.objects.all(),allow_null = True)	

	porcentajeAvance=serializers.SerializerMethodField()

	porcentajeAvanceFinanciero=serializers.SerializerMethodField()

	class Meta:
		model = Cronograma
		fields=('id','proyecto','proyecto_id','periodicidad',
			'esquema_id','esquema','periodicidad_id',
			'programacionCerrada','nombre', 'porcentajeAvance',
			'porcentajeAvanceFinanciero','fechaInicio','fechaFinal','confirmarFechas')

	def get_porcentajeAvance(self,obj):
		avance = 0
		porcentajes = []

		qset1 = (Q(presupuesto__cronograma__id=obj.id)& Q(cantidad__gt=0) & (~Q(actividad__nombre='No aplica')))
		qset2 = (Q(detallePresupuesto__presupuesto__cronograma__id=obj.id)& Q(cantidad__gt=0) & (~Q(detallePresupuesto__actividad__nombre='No aplica')))

		# import pdb; pdb.set_trace()
		queryset_a_ejecutar = FDetallePresupuesto.objects.filter(
			# presupuesto__cronograma__id=obj.id,
			# cantidad__gt=0
			qset1).values(
			'actividad__id',
			'actividad__nombre',
			'actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
			# detallePresupuesto__presupuesto__cronograma__id=obj.id,
			# cantidad__gt=0
			qset2).values(
			'detallePresupuesto__actividad__id',
			'detallePresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidad'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['actividad__id'],
						'actividad': Aejecutar['actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100,
						'peso': Aejecutar['actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['actividad__id'],
					'actividad': Aejecutar['actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['actividad__peso']
				})
		
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		return round(avance,2)
		# return 0

	def get_porcentajeAvanceFinanciero(self,obj):
		avance = 0		

		presupuesto = EPresupuesto.objects.filter(cronograma_id=obj.id,cerrar_presupuesto=True)
		aiu = 1
		if presupuesto:
			aiu = presupuesto.first().aiu
			presupuesto = presupuesto.first().suma_presupuesto()			
		else:
			presupuesto = 0


		queryset_a_ejecutar = DetalleReporteTrabajo.objects.filter(
			detallePresupuesto__presupuesto__cronograma__id=obj.id,
			cantidad__gt=0).values('cantidad','detallePresupuesto__valorGlobal')


		for item in queryset_a_ejecutar:

			subtotal_aux = float(item['cantidad'])*float(item['detallePresupuesto__valorGlobal'])

			peso_aux = (float(subtotal_aux)*float(aiu))/float(presupuesto*aiu)			

			avance += peso_aux
			# avance += item.peso()
			# Aqui ando
		
		return round(avance*100,2)

@transaction.atomic
def GenerarPeriodosDelCronograma(cronograma_id):	
	# import pdb; pdb.set_trace()		
	try:
		cronograma = Cronograma.objects.get(pk=cronograma_id)
		valide = False
		fecha_inicio = cronograma.fechaInicio

		while valide==False:			
			fecha_final = fecha_inicio + timedelta(days=cronograma.periodicidad.numero_dias)
			if fecha_final>=cronograma.fechaFinal:
				fecha_final = cronograma.fechaFinal
			periodo = PeriodoProgramacion(cronograma_id=cronograma_id,fechaDesde=fecha_inicio,fechaHasta=fecha_final)
			periodo.save()

			if fecha_final==cronograma.fechaFinal:		
				valide = True
			else:
				fecha_inicio = fecha_final + timedelta(days=1)

	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		


class CronogramaAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=Cronograma
	queryset = model.objects.all()
	serializer_class = CronogramaSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.cronograma'

	def retrieve(self,request,*args, **kwargs):
		try:
			lite = self.request.query_params.get('lite', None)
			instance = self.get_object()

			if lite:
				serializer = CronogramaSerializerLite(instance)
			else:
				serializer = self.get_serializer(instance)
			
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CronogramaAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			proyecto_id= self.request.query_params.get('proyecto_id',None)
			esquema_id= self.request.query_params.get('esquema_id',None)
			programacionCerrada= self.request.query_params.get('programacionCerrada',None)

			qset=(~Q(id=0))

			if proyecto_id:
				qset = qset &(
					Q(proyecto_id=proyecto_id)
					)

			if esquema_id:
				qset = qset &(
					Q(esquema_id=esquema_id)
					)

			if programacionCerrada:
				qset = qset &(
					Q(programacionCerrada=True)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				# import pdb; pdb.set_trace()
				
				serializer = CronogramaSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(periodicidad_id=request.data['periodicidad_id'],proyecto_id=request.data['proyecto_id'],esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					GenerarPeriodosDelCronograma(serializer.data['id']);

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					##import pdb; pdb.set_trace()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.data['estado_id']==0:
					request.data['estado_id']=None

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CronogramaSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(periodicidad_id=request.data['periodicidad_id'],proyecto_id=request.data['proyecto_id'],estado_id=request.data['estado_id'],esquema_id=request.data['esquema_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	




#Fin Api rest para Cronograma


#Api rest para Presupuesto
class PresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	cronograma=CronogramaSerializer(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())	

	# totalCambios=serializers.SerializerMethodField()

	class Meta:
		model = EPresupuesto
		fields=('id','cronograma','cronograma_id','cerrar_presupuesto','nombre','aiu')


	# def get_totalCambios(self, obj):
	# 	return LCambio.objects.filter(presupuesto_id=obj.id).count()

class PresupuestoGraficoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = EPresupuesto
		fields = ('id','nombre','aiu')


class PresupuestoAvanceGraficoViewSet(viewsets.ModelViewSet):
	
	model=EPresupuesto
	queryset = model.objects.all()
	serializer_class = PresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PresupuestoAvanceGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			cerrado_presupuesto= self.request.query_params.get('cerrado_presupuesto',None)


			qset=(~Q(id=0))

			if cronograma_id:
				qset = qset &(
					Q(cronograma__id=cronograma_id)
					)

			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if cerrado_presupuesto:
				qset = qset &(
					Q(cerrar_presupuesto=cerrado_presupuesto)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				# import pdb; pdb.set_trace()
				serializer = PresupuestoGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:

				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PresupuestoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Presupuesto




#Api rest para Detalle Presupuesto
from .models import CatalogoUnidadConstructiva
class CatalogoUnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	mcontrato = ContratoLiteSerializer(read_only = True)
	mcontrato_id = serializers.PrimaryKeyRelatedField(read_only=True)
	
	class Meta:
		model=CatalogoUnidadConstructiva
		fields=('id','nombre','ano','activo','mcontrato', 'mcontrato_id')

class MaterialSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=Material
		fields=('id','codigo','descripcion','valorUnitario', 'unidadMedida', 'catalogo_id', 'catalogo')

class MaterialSerializerLite(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model=Material
		fields=('id','descripcion','codigo')

class MaterialSerializerLite2(serializers.HyperlinkedModelSerializer):

	class Meta:
		model=Material
		fields=('id','codigo','descripcion','valorUnitario', 'unidadMedida')



class ManoObraSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion','valorHora', 'catalogo_id', 'catalogo')

class ManoObraSerializerLite2(serializers.HyperlinkedModelSerializer):

	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion','valorHora')		

class ManoObraSerializerLite(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model=ManoDeObra
		fields=('id','codigo','descripcion')


class DetallePresupuestoGraficoSerializer(serializers.HyperlinkedModelSerializer):

	presupuesto=PresupuestoGraficoSerializer(read_only=True)
	presupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EPresupuesto.objects.all())	

	actividad=EsquemaCapitulosActividadesSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	catalogoUnidadConstructiva = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogoUnidadConstructiva_id=serializers.PrimaryKeyRelatedField(write_only=True,
		queryset=CatalogoUnidadConstructiva.objects.all())
	sumaPresupuesto=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','sumaPresupuesto','presupuesto','presupuesto_id',
			'actividad',
			'actividad_id','codigoUC','descripcionUC','valorManoObra','valorMaterial',
			'valorGlobal','cantidad','porcentaje','nombre_padre',
			'catalogoUnidadConstructiva','catalogoUnidadConstructiva_id')


	def get_sumaPresupuesto(self, obj):
		sumatoria=0
		suma=FDetallePresupuesto.objects.filter(presupuesto_id=obj.presupuesto.id)	
		valor=0
		for item in suma:
			valor=float(item.valorMaterial)+float(item.valorManoObra)
			total=valor*float(item.cantidad)

			sumatoria=sumatoria+total

		return round(sumatoria,3)
		
class DetallePresupuestoGraficoLiteSerializer(serializers.HyperlinkedModelSerializer):
	actividad=EsquemaCapitulosActividadesLiteSerializer(read_only=True)
	actividad_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CEsquemaCapitulosActividadesG.objects.all())

	sumaPresupuesto=serializers.SerializerMethodField()
	# peso=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','sumaPresupuesto','actividad','actividad_id','codigoUC','descripcionUC','peso',
		'valorGlobal','cantidad','nombre_padre')


	def get_sumaPresupuesto(self, obj):
		sumatoria=0
		suma=FDetallePresupuesto.objects.filter(presupuesto_id=obj.presupuesto.id)	
		valor=0
		for item in suma:
			valor=float(item.valorMaterial)+float(item.valorManoObra)
			total=valor*float(item.cantidad)

			sumatoria=sumatoria+total

		return round(sumatoria,3)

	# def get_peso(self, obj):
	# 	return obj.peso

class DetallePresupuestoGraficoLite2Serializer(serializers.HyperlinkedModelSerializer):
	actividad=EsquemaCapitulosActividadesLiteSerializer(read_only=True)
	cantidad_programada=serializers.SerializerMethodField()
	# peso=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','actividad','codigoUC','descripcionUC','cantidad','cantidad_programada','nombre_padre')


	def get_cantidad_programada(self, obj):
		aux = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0

class DetallePresupuestoGraficoLite3Serializer(serializers.HyperlinkedModelSerializer):
	actividad=EsquemaCapitulosActividadesLiteSerializer(read_only=True)
	cantidad_ejecutada=serializers.SerializerMethodField()
	cantidad_aejecutar=serializers.SerializerMethodField()
	# peso=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','actividad','codigoUC','descripcionUC',
			# 'cantidad',
			'cantidad_aejecutar',
			'cantidad_ejecutada','nombre_padre')


	def get_cantidad_ejecutada(self, obj):
		# import pdb; pdb.set_trace()
		aux = DetalleReporteTrabajo.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0

	def get_cantidad_aejecutar(self, obj):
		aux = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0

class DetallePresupuestoGraficoLite4Serializer(serializers.HyperlinkedModelSerializer):
	cantidad_aejecutar=serializers.SerializerMethodField()
	# peso=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id','cantidad','cantidad_aejecutar')


	def get_cantidad_aejecutar(self, obj):
		aux = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0

class DetallePresupuestoGraficoLite5Serializer(serializers.HyperlinkedModelSerializer):
	cantidad_ejecutada=serializers.SerializerMethodField()
	cantidad_aejecutar=serializers.SerializerMethodField()
	# peso=serializers.SerializerMethodField()

	class Meta:
		model = FDetallePresupuesto
		fields=('id',
			# 'cantidad',
			'cantidad_aejecutar',
			'cantidad_ejecutada')


	def get_cantidad_ejecutada(self, obj):
		# import pdb; pdb.set_trace()
		aux = DetalleReporteTrabajo.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0

	def get_cantidad_aejecutar(self, obj):
		aux = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=obj.id).aggregate(Sum('cantidad'))
		if aux['cantidad__sum'] is not None:
			return round(float(aux['cantidad__sum']),4)
		else:
			return 0




class DetallePresupuestoGraficoViewSet(viewsets.ModelViewSet):
	
	model=FDetallePresupuesto
	queryset = model.objects.all()
	serializer_class = DetallePresupuestoGraficoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.detalle_presupuesto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetallePresupuestoGraficoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			presupuesto_id= self.request.query_params.get('presupuesto_id',None)
			hito_id= self.request.query_params.get('hito_id',None)
			actividad_id= self.request.query_params.get('actividad_id',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)
			listado_apoyo= self.request.query_params.get('listado_apoyo',None)
			id_apoyo= self.request.query_params.get('id_apoyo',None)
			lite2= self.request.query_params.get('lite2',None)
			lite3= self.request.query_params.get('lite3',None)
			lite4= self.request.query_params.get('lite4',None)

			periodoProgramacion_id= self.request.query_params.get('periodoProgramacion_id',None)
			reporteTrabajo_id= self.request.query_params.get('reporteTrabajo_id',None)
			
			qset=(~Q(id=0))
			if periodoProgramacion_id and cronograma_id:
				no_ids= DetallePeriodoProgramacion.objects.filter(detallePresupuesto__presupuesto__cronograma__id=cronograma_id,periodoProgramacion_id=periodoProgramacion_id).values('detallePresupuesto__id')
				qset = qset &(
					~Q(id__in=no_ids)
					)

			if reporteTrabajo_id and presupuesto_id:
				# import pdb; pdb.set_trace()

				no_ids= DetalleReporteTrabajo.objects.filter(detallePresupuesto__presupuesto__id=presupuesto_id,reporteTrabajo_id=reporteTrabajo_id).values('detallePresupuesto__id')
				
				# ids_restantes= FDetallePresupuesto.objects.filter( (~Q(id__in=no_ids) ) & (Q(presupuesto__id=presupuesto_id)) )
				# serializer_aux2 = DetallePresupuestoGraficoLite5Serializer(ids_restantes,many=True)
				# ids_validos = []
				# for item in serializer_aux2.data:
				# 	if float(item['cantidad_aejecutar'])>float(item['cantidad_ejecutada']):
				# 		ids_validos.append(item['id'])

				qset = qset &(
					# Q(id__in=ids_validos)
					~Q(id__in=no_ids)
					)


			if presupuesto_id:
				qset = qset &(
					Q(presupuesto_id=presupuesto_id)
					)

			if actividad_id:
				qset = qset &(
					Q(actividad_id=actividad_id)
					)

			if hito_id:
				qset = qset &(
					Q(actividad__padre=hito_id)
					)

			if cronograma_id:
				# cronograma=KCronograma.objects.get(pk=cronograma_id)
				# presupuesto_id=cronograma.presupuesto.id
				# qset = qset &(
				# 	Q(presupuesto_id=cronograma.presupuesto.id)
				# 	)
				qset = qset &(
					Q(presupuesto__cronograma__id=cronograma_id)
					)

			lista=[]
			if listado_apoyo and presupuesto_id and int(presupuesto_id)>0 and id_apoyo:
				lista=HNodo.objects.filter(presupuesto_id=presupuesto_id).exclude(id=id_apoyo).values('id','nombre')


			if dato:
				qset = qset &(
					Q(codigoUC__icontains=dato)|
					Q(descripcionUC__icontains=dato)|
					Q(actividad__nombre__icontains=dato)
					)


			if qset is not None:
				queryset = self.model.objects.filter(qset)	
			# import pdb; pdb.set_trace()
			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite2:
						serializer = DetallePresupuestoGraficoLiteSerializer(page,many=True)
					elif lite3:
						serializer = DetallePresupuestoGraficoLite2Serializer(page,many=True)
						serializer_aux = DetallePresupuestoGraficoLite4Serializer(queryset,many=True)
						cant_superior = 0
						cant_inferior= 0
						# import pdb; pdb.set_trace()
						for item in serializer_aux.data:
							if float(item['cantidad'])<float(item['cantidad_aejecutar']):
								cant_superior+= 1
							if float(item['cantidad'])>float(item['cantidad_aejecutar']):
								cant_inferior+= 1

						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'cant_superior':cant_superior,'cant_inferior':cant_inferior}})
					elif lite4:
						serializer = DetallePresupuestoGraficoLite3Serializer(page,many=True)
						serializer_aux = DetallePresupuestoGraficoLite5Serializer(queryset,many=True)
						cant_superior = 0
						cant_inferior= 0
						for item in serializer_aux.data:
							if float(item['cantidad_aejecutar'])<float(item['cantidad_ejecutada']):
								cant_superior+= 1
							if float(item['cantidad_aejecutar'])>float(item['cantidad_ejecutada']):
								cant_inferior+= 1

						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'cant_superior':cant_superior,'cant_inferior':cant_inferior}})
					else:
						serializer = self.get_serializer(page,many=True)
					if listado_apoyo:						
						return self.get_paginated_response({'message':'','success':'ok',
						'data':{'datos':serializer.data,'apoyos':lista}})
					else:
						return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
			
			if lite2:
				serializer = DetallePresupuestoGraficoLiteSerializer(queryset,many=True)
			elif lite3:
				serializer = DetallePresupuestoGraficoLite2Serializer(queryset,many=True)
			elif lite4:
				serializer = DetallePresupuestoGraficoLite3Serializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)

			if listado_apoyo:
				return Response({'message':'','success':'ok',
					'data':{'datos':serializer.data,'apoyos':lista}})			
			else:
				return Response({'message':'','success':'ok',
					'data':serializer.data})
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetallePresupuestoGraficoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():
					serializer.save(presupuesto_id=request.data['presupuesto_id'],
						actividad_id=request.data['actividad_id'],
						catalogoUnidadConstructiva_id=request.data['catalogoUnidadConstructiva_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetallePresupuestoGraficoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(presupuesto_id=request.data['presupuesto_id'],
						actividad_id=request.data['actividad_id'],
						catalogoUnidadConstructiva_id=request.data['catalogoUnidadConstructiva_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para Detalle presupuesto

class PeriodoProgramacionLiteSerializer(serializers.HyperlinkedModelSerializer):
	cronograma=CronogramaSerializerLite(read_only=True)

	class Meta:
		model = PeriodoProgramacion
		fields=('id','cronograma','fechaDesde','fechaHasta')


class DetallePeriodoProgramacionSerializer(serializers.HyperlinkedModelSerializer):

	periodoProgramacion=PeriodoProgramacionLiteSerializer(read_only=True)
	periodoProgramacion_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=PeriodoProgramacion.objects.all())	

	detallePresupuesto=DetallePresupuestoGraficoLite2Serializer(read_only=True)
	detallePresupuesto_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())

	class Meta:
		model = DetallePeriodoProgramacion
		fields=('id','periodoProgramacion','periodoProgramacion_id','detallePresupuesto','detallePresupuesto_id','cantidad')


		
class DetallePeriodoProgramacionLiteSerializer(serializers.HyperlinkedModelSerializer):
	periodoProgramacion=PeriodoProgramacionLiteSerializer(read_only=True)
	detallePresupuesto=DetallePresupuestoGraficoLite2Serializer(read_only=True)

	class Meta:
		model = DetallePeriodoProgramacion
		fields=('id','periodoProgramacion','detallePresupuesto','cantidad')


	

class DetallePeriodoProgramacionViewSet(viewsets.ModelViewSet):
	
	model=DetallePeriodoProgramacion
	queryset = model.objects.all()
	serializer_class = DetallePeriodoProgramacionSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.detalle_PeriodoProgramacion'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetallePeriodoProgramacionViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			periodoProgramacion_id= self.request.query_params.get('periodoProgramacion_id',None)
			detallePresupuesto_id= self.request.query_params.get('detallePresupuesto_id',None)

			lite= self.request.query_params.get('lite',None)


			qset=(~Q(id=0))

			if periodoProgramacion_id:
				qset = qset &(
					Q(periodoProgramacion_id=periodoProgramacion_id)
					)

			if detallePresupuesto_id:
				qset = qset &(
					Q(detallePresupuesto_id=detallePresupuesto_id)
					)



			if qset is not None:
				queryset = self.model.objects.filter(qset)	

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = DetallePeriodoProgramacionLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
						'data':serializer.data})
			

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})
		
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetallePeriodoProgramacionSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():					
					serializer.save(periodoProgramacion_id=request.data['periodoProgramacion_id'],
						detallePresupuesto_id=request.data['detallePresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetallePeriodoProgramacionSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(periodoProgramacion_id=request.data['periodoProgramacion_id'],
						detallePresupuesto_id=request.data['detallePresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Inicio eliminacion de un esquema
@transaction.atomic
def eliminar_esquema(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				CReglasEstadoG.objects.filter(esquema_id=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.filter(esquema_id=item['id']).delete()
				BEsquemaCapitulosG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.esquema',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un esquema



#eliminacion de capitulo/actividad de un esquema
@transaction.atomic
def eliminar_id_capitulo_actividad_esquema(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])

			if int(valor.padre)==0:
				CEsquemaCapitulosActividadesG.objects.filter(padre=item['id']).delete()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.capitulos_esquema',id_manipulado=item['id'])
				logs_model.save()

			else:
				modelhijos=CEsquemaCapitulosActividadesG.objects.get(pk=item['id'])
				model=CEsquemaCapitulosActividadesG.objects.get(pk=modelhijos.padre)
				model.peso=float(model.peso) - float(modelhijos.peso)
				model.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.capitulos_esquema',id_manipulado=model.id)
				logs_model.save()
				CEsquemaCapitulosActividadesG.objects.get(pk=item['id']).delete()
				logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.capitulos_esquema',id_manipulado=item['id'])
				logs_model2.save()

		transaction.savepoint_commit(sid)			
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra.capitulos_esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de capitulo/actividad de un esquema


#Inicio clonacion de los esquemas
@login_required
@transaction.atomic
def clonacion_esquema(request):
	sid = transaction.savepoint()
	##import pdb; pdb.set_trace()
	try:
		# lista=request.POST['_content']
		# respuesta= json.loads(lista)

		esquema=BEsquemaCapitulosG(nombre=request.POST['nombre_esquema'],macrocontrato_id=int(request.POST['id_macrocontrato']))
		esquema.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.esquema',id_manipulado=esquema.id)
		logs_model.save()

		capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=int(request.POST['id_etiqueta']),nivel=1)

		for item in capitulos:
			hitos=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item.peso,nombre=item.nombre,nivel=item.nivel,padre=item.padre)
			hitos.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.esquema_capitulos',id_manipulado=hitos.id)
			logs_model2.save()

			actividades=CEsquemaCapitulosActividadesG.objects.filter(padre=item.id)

			for item2 in actividades:
				actividad=CEsquemaCapitulosActividadesG(esquema_id=esquema.id,peso=item2.peso,nombre=item2.nombre,nivel=item2.nivel,padre=hitos.id)
				actividad.save()
				logs_model3=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.esquema_capitulos',id_manipulado=actividad.id)
				logs_model3.save()

		transaction.savepoint_commit(sid)		
		return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})
		#return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
		#			'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.esquema')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin clonacion de los esquemas


#eliminacion de regla de estado
@transaction.atomic
def eliminar_id_regla_estado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		esquema_id=None

		for item in respuesta['lista']:
			valor=CReglasEstadoG.objects.get(pk=item['id'])
			esquema_id=valor.esquema_id
			valor.delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.regla_estado',id_manipulado=item['id'])
			logs_model.save()	

		updateAsyncEstado.delay(esquema_id)	
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		print(e) 
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avance_de_obra.regla_estado')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de regla de estado


#eliminacion de cronograma
@transaction.atomic
def eliminar_id_cronograma(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			valor=Cronograma.objects.get(pk=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.cronograma',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cronograma')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

#Fin eliminacion de cronograma

@login_required
@transaction.atomic
def eliminar_id_reportetrabajo(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			ReporteTrabajo.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@login_required
@transaction.atomic
def eliminar_id_uucc(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			UnidadConstructiva.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.UnidadConstructiva',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_materiales(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			Material.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_mano_obra(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			ManoDeObra.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_desgl_mat(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DesgloceMaterial.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@transaction.atomic
def eliminar_id_desgl_mo(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DesgloceManoDeObra.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.Material',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.UnidadConstructiva')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
@transaction.atomic
def eliminar_id_presupuesto(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EPresupuesto.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.presupuesto',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	

@login_required
@transaction.atomic
def eliminar_id_reformado(request):

	sid = transaction.savepoint()
	try:
		##import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			EReformado.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.reformado',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.reformado')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
def descargar_plantilla_presupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	esquema_id= request.GET['id_esquema']	
				
	hitos = CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)
	worksheet.set_column('I:I', 20)

	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)
	# worksheet.write('G1', 'Valor M.O', format1)
	# worksheet.write('H1', 'Valor Material', format1)
	# worksheet.write('I1', 'Valor Global', format1)

	for item in hitos:

		if item.nivel==1:
			actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=esquema_id,padre=item.id)

			for item2 in actividades:

				worksheet.write(row, col, item2.id,format2)
				worksheet.write(row, col+1, item.nombre,format2)
				worksheet.write(row, col+2,item2.nombre,format2)
				worksheet.write(row, col+3,'',format2)
				worksheet.write(row, col+4,'',format2)
				worksheet.write(row, col+5,'',format2)
				# worksheet.write(row, col+6,'',format2)
				# worksheet.write(row, col+7,'',format2)
				# worksheet.write(row, col+8,'',format2)

				row +=1


	workbook.close()

	return response
    #return response


@login_required
def guardar_presupuesto_archivo(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte,data_only = True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		revisar_detalle=FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


		if len(revisar_detalle) > 0:
			FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).delete()


		contador=hoja.max_row - 1
		numeroFila = 1
		if int(contador) > 0:
			
			for fila in hoja.rows:				
				if numeroFila > 1:
					
					if (fila[0].value == None or fila[1].value == None or \
					fila[2].value == None or fila[3].value == None or  \
					fila[4].value == None or fila[5].value == None):
					
						return JsonResponse({
							'message':'Se encontraron celdas vacias en la fila ' + str(numeroFila),
							'success':'error',
							'data':''})
					else:
						##import pdb; pdb.set_trace()
						if UnidadConstructiva.objects.filter(
							catalogo_id=request.POST['catalogoUnidadConstructiva_id'],
							codigo=fila[3].value).count() == 0:
							#import pdb; pdb.set_trace()
							return JsonResponse({
								'message':'La unidad constructiva en la fila ' + str(numeroFila) + \
								' no se encuentra registrada en el catalogo seleccionado.',
								'success':'error',
								'data':''})

				numeroFila = numeroFila + 1	
									

			
			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						count=0
						j=0
						for fila2 in hoja.rows:
							if fila2[0].value:
								if j == 0:
									j=1
								else:
									if int(fila[0].value)==int(fila2[0].value):
										count = count+1

						
						actividad_porce=CEsquemaCapitulosActividadesG.objects.get(pk=fila[0].value)
						porcentaje=float(actividad_porce.peso)/count

						# import pdb; pdb.set_trace()
						
						qset = UnidadConstructiva.objects.filter(catalogo__id=request.POST['catalogoUnidadConstructiva_id'], codigo=fila[3].value)	
						serializer = UnidadConstructivaSerializerLite(qset,many=True)

						if qset:
							detalle=FDetallePresupuesto(porcentaje=porcentaje,
								presupuesto_id=presupuesto_id,
								actividad_id=fila[0].value,
								codigoUC=fila[3].value,
								descripcionUC=fila[4].value,
								valorManoObra=serializer.data[0]['totalManoDeObra'],
								valorMaterial=serializer.data[0]['totalMateriales'],
								valorGlobal=float(serializer.data[0]['totalMateriales'])+float(serializer.data[0]['totalManoDeObra']),
								cantidad=fila[5].value,
								catalogoUnidadConstructiva_id=request.POST['catalogoUnidadConstructiva_id'])
							detalle.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.detalle_presupuesto',id_manipulado=detalle.id)
							logs_model.save()

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.detalle_presupuesto')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def actualizar_cantidad(request):

	sid = transaction.savepoint()
	try:
		# import pdb; pdb.set_trace()
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.valorGlobal=item['valorUC'] if 'valorUC' in item else detalle.valorGlobal
			detalle.save()

			# ActualizarPorcentajesDetallesPresupuesto(detalle.actividad.id,detalle.actividad.peso)

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.actualizar_cantidad')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@transaction.atomic
def ActualizarPorcentajesDetallesPresupuesto(actividad_id,peso):	
	# import pdb; pdb.set_trace()		
	try:
		detalles=FDetallePresupuesto.filter(actividad_id=actividad_id)
		for item in detalles:
			item.porcentaje = peso/len(detalles)
			item.save()

	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		


@login_required
@transaction.atomic
def cierre_presupuesto(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			detalle=FDetallePresupuesto.objects.get(id=item['id'])
			detalle.cantidad=item['cantidad']
			detalle.valorGlobal=item['valorUC']
			detalle.save()

			# ActualizarPorcentajesDetallesPresupuesto(detalle.actividad.id,detalle.actividad.peso)

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.detalle_presupuesto',id_manipulado=item['id'])
			logs_model.save()


		presupuesto=EPresupuesto.objects.get(pk=respuesta['id_presupuesto'])
		presupuesto.cerrar_presupuesto=True
		presupuesto.save()	
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.presupuesto',id_manipulado=respuesta['id_presupuesto'])
		logs_model.save()	

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cierre_presupuesto')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de un diagrama
@transaction.atomic
def eliminar_diagrama(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				PeriodoProgramacion.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.PeriodoProgramacion',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.diagrama')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de un diagrama	


@login_required
def informe_diagrama(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Avance de Obra - Informe de programación.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Programacion semanal')

	format0=workbook.add_format({'border':1})
	format0.set_bg_color('#D5D8DC')

	format1=workbook.add_format({'border':1,'font_size':12,'bold':True})
	format1.set_text_wrap()


	format2=workbook.add_format({'border':1,'font_size':8})
	format2.set_text_wrap()

	format3=workbook.add_format({'border':1,'font_size':8})
	format3.set_align('center')

	format4=workbook.add_format({'border':1,'font_size':8})
	format4.set_align('center')
	format4.set_locked(True)

	currency_format = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8})
	currency_format.set_align('center')

	format4_GREY=workbook.add_format({'border':1,'font_size':8,'bold':True})
	format4_GREY.set_align('center')
	format4_GREY.set_bg_color('#D5D8DC')
	format4_GREY.set_locked(True)
	
	currency_format_GREY = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8,'bold':True})
	currency_format_GREY.set_align('center')
	currency_format_GREY.set_bg_color('#D5D8DC')
	
	format5=workbook.add_format({'border':1,'font_size':8})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_text_wrap()


	format6=workbook.add_format({'font_size':12,'bold':True})
	format6.set_align('center')
	format6.set_text_wrap()
	
	format7=workbook.add_format({'border':1, 'num_format': '0.00"%"','font_size':8})
	format7.set_align('center')
	format7.set_locked(True)
	# format7.set_text_wrap()



	format_red = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_red.set_bg_color('#E6B0AA')
	format_red.set_font_color('#C70039')
	format_red.set_align('center')
	format_red.set_locked(True)

	format_yellow = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_yellow.set_bg_color('#F9E79F')
	format_yellow.set_font_color('#FFC300')
	format_yellow.set_align('center')
	format_yellow.set_locked(True)

	format_green =  workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_green.set_bg_color('#ABEBC6')
	format_green.set_font_color('#28B463')
	format_green.set_align('center')
	format_green.set_locked(True)


	worksheet.set_column('A:A', 5)
	worksheet.set_column('B:B', 12)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 12)
	worksheet.set_column('E:E', 16)
	worksheet.set_column('F:F', 16)

	merge_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})

	merge_format_horizontal = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center'})    
	merge_format_horizontal.set_bg_color('#D5D8DC')


	cronograma_id= request.GET['cronograma_id']	

	presupuesto_id = EPresupuesto.objects.filter(cronograma_id=cronograma_id,cerrar_presupuesto=True).last().id
	
	

	


  

	worksheet.merge_range('A1:A3', 'Merged Range', merge_format)
	worksheet.merge_range('B1:B3', 'Merged Range', merge_format)
	worksheet.merge_range('C1:C3', 'Merged Range', merge_format)
	worksheet.merge_range('D1:D3', 'Merged Range', merge_format)
	worksheet.merge_range('E1:E3', 'Merged Range', merge_format)
	worksheet.merge_range('F1:F3', 'Merged Range', merge_format)
	# worksheet.merge_range('G1:G3', 'Merged Range', merge_format)
	

	worksheet.write('A1', 'Item', format1)
	# worksheet.write('B1', 'Hitos', format1)
	# worksheet.write('B1', 'Actividad', format1)
	worksheet.write('B1', 'Cod. UUCC', format1)
	worksheet.write('C1', 'Descripcion UUCC', format1)
	worksheet.write('D1', 'Cant.', format1)	
	worksheet.write('E1', 'Valor Unitario', format1)
	worksheet.write('F1', 'Valor TOTAL', format1)

	worksheet.freeze_panes(3, 6)
	
	border_format=workbook.add_format({
                            'border':2
                           })

	# import pdb; pdb.set_trace()
	col=6
	p=1
	periodos = PeriodoProgramacion.objects.filter(cronograma_id=cronograma_id).order_by('fechaDesde')
	for semana in periodos:	
		# import pdb; pdb.set_trace()
		fechaDesde = Str_fecha(semana.fechaDesde)
		fechaHasta = Str_fecha(semana.fechaHasta)

		# fechaDesde = '{0}/{1}/{2}'.format(str(semana.fechaDesde.day),str(semana.fechaDesde.month),str(semana.fechaDesde.year))
		# fechaHasta = '{0}/{1}/{2}'.format(str(semana.fechaHasta.day),str(semana.fechaHasta.month),str(semana.fechaHasta.year))

		worksheet.merge_range('{0}1:{1}1'.format(xl_col_to_name(col),xl_col_to_name(col+1)), 'Periodo '+str(p), merge_format)


		worksheet.write(1,col,fechaDesde, format6)		
		worksheet.write(2,col,'Cantidad', format6)
		col+=1
		
		worksheet.write(1,col,fechaHasta, format6)	
		worksheet.write(2,col,'Valor', format6)
		p+=1
		col+=1

	worksheet.merge_range('{0}1:{0}3'.format(xl_col_to_name(col)), 'Merged Range', merge_format)
	worksheet.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+1)), 'Merged Range', merge_format)
	worksheet.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+2)), 'Merged Range', merge_format)

	worksheet.write('{0}1'.format(xl_col_to_name(col)), 'Cant.', format1)
	worksheet.write('{0}1'.format(xl_col_to_name(col+1)), 'Valor', format1)
	worksheet.write('{0}1'.format(xl_col_to_name(col+2)), '%', format1)

	worksheet.set_column(col+1,col+1, 16)
	worksheet.set_column(6,col-1, 16)

	

	actividades = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	capitulos_id = actividades.filter().values('actividad__id').distinct()
	capitulos = CEsquemaCapitulosActividadesG.objects.filter(pk__in=capitulos_id).order_by('id')

	row=3
	item=1
	for cap in capitulos:
		# import pdb; pdb.set_trace()
		worksheet.write(row,0, '', format0)
		# worksheet.merge_range('B{0}:{1}{0}'.format(row,xl_col_to_name(col)), 'Merged Range', merge_format)
		worksheet.merge_range('A{0}:F{0}'.format(row+1), cap.nombre, merge_format_horizontal)

		
		worksheet.merge_range('{0}{1}:{2}{1}'.format(xl_col_to_name(6),row+1,xl_col_to_name(col+2)), '', merge_format_horizontal)
		

		row+=1
		actividades_cap = actividades.filter(actividad__id=cap.id).order_by('id')

		#----------UUCC ACTIVIDADES------------------------
		for item_act in actividades_cap:
			if item_act.id==actividades_cap[0].id:
				row_periodo  = row+1

			worksheet.write(row,0, item, format2)
			item+=1
			# padre = CEsquemaCapitulosActividadesG.objects.get(pk=item_act.actividad.padre)
			# worksheet.write(row,1, padre.nombre, format2)
			# worksheet.write(row,1, item_act.actividad.nombre, format2)
			worksheet.write(row,1, item_act.codigoUC, format2)
			worksheet.write(row,2, item_act.descripcionUC, format2)
			worksheet.write(row,3, round(item_act.cantidad, 4), format4)	
			
			
			worksheet.write(row,4, item_act.valorGlobal, currency_format)
			worksheet.write(row,5, '=E{0}*D{0}'.format(str(row+1)), currency_format)

			#----------Periodos------------------------
			col=6			
			str_suma_cantidades = '='
			str_suma_valores  = '='
			for semana in periodos:	

				str_suma_cantidades = str_suma_cantidades + '+{0}{1}'.format(xl_col_to_name(col),row+1)				
				cantidad = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=item_act.id,periodoProgramacion_id=semana.id).aggregate(Sum('cantidad'))
				worksheet.write(row,col,cantidad['cantidad__sum'], format4)
				col+=1

				str_suma_valores = str_suma_valores + '+{0}{1}'.format(xl_col_to_name(col),row+1)
				worksheet.write(row,col,'=E{0}*{1}{0}'.format(row+1,xl_col_to_name(col-1)), currency_format)
				col+=1

			#----------Fin Periodos------------------------

			#----------Valores programados------------------------
			

			worksheet.write(row,col, str_suma_cantidades, format4)
			worksheet.write(row,col+1, str_suma_valores, currency_format)
			worksheet.write(row,col+2, '=({1}{0}/D{0})*100'.format(str(row+1),xl_col_to_name(col)),format7)					

			worksheet.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '>',
		                                        'value':    100,
		                                        'format':   format_red})
			worksheet.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': 'between',
		                                        'minimum': 0,
		                                        'maximum': 99.99999,
		                                        'format':   format_yellow})
			worksheet.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '==',
		                                        'value':    100,
		                                        'format':   format_green})

			row+=1
			#----------Fin Valores programados------------------------
		#----------Fin UUCC ACTIVIDADES------------------------		
		worksheet.merge_range('A{0}:E{0}'.format(row+1), 'SUBTOTAL '+cap.nombre, merge_format_horizontal)

		sumatoria_valor = 'F{1}:F{0}'.format(row,row_periodo)
		sumatoria_cant_extra = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)
		sumatoria_valor_extra = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col+1),row_periodo)

		worksheet.write(row,5, '=SUM('+sumatoria_valor+')', currency_format_GREY)
		worksheet.write(row,col, '=SUM('+sumatoria_cant_extra+')', format4_GREY)
		worksheet.write(row,col+1, '=SUM('+sumatoria_valor_extra+')', currency_format_GREY)
		worksheet.write(row,col+2, '', format0)

		col=6
		# import pdb; pdb.set_trace()
		for semana in periodos:	
			sumatoria_cant_periodo = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)		
			worksheet.write(row,col, '=SUM('+sumatoria_cant_periodo+')', format4_GREY)
			col+=1

			sumatoria_valor_periodo = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)
			worksheet.write(row,col, '=SUM('+sumatoria_valor_periodo+')', currency_format_GREY)
			col+=1


		row+=1


	workbook.close()

	return response


@login_required
def descargar_plantilla_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', 'Apoyo', format1)
	worksheet.write('B1', 'Longitud', format1)
	worksheet.write('C1', 'Latitud', format1)


	workbook.close()

	return response
    #return response



@login_required
def guardar_apoyo_archivo(request):
	
	sid = transaction.savepoint()
	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:			
			nombreNodos = []
			for fila in hoja.rows:
				if fila[0].value:
					
					if nombreNodos:
						if fila[0].value not in nombreNodos:
							nombreNodos.append(fila[0].value)
						else:
							transaction.savepoint_rollback(sid)
							return JsonResponse({'message':'El nodo ' + fila[0].value + ' se encuentra repetido, por favor corregir y volver a intentarlo.','success':'error',
								'data':''})							
					else:
						nombreNodos.append(fila[0].value)
					
					nodo_data = HNodo.objects.filter(nombre=fila[0].value, presupuesto_id=presupuesto_id).values('id')
					if i == 0:
						i=1
					else:

						longitud = str(fila[1].value)
						if "," in longitud:
							longitud = float(longitud.replace(",","."))
						
						latitud = str(fila[2].value)
						if "," in latitud:
							latitud = float(latitud.replace(",","."))

						if nodo_data:							
							nodo_act = HNodo.objects.get(pk=nodo_data[0]['id'])
							nodo_act.longitud = longitud
							nodo_act.latitud = latitud
							nodo_act.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.nodo',id_manipulado=nodo_data[0]['id'])
							logs_model.save()							
						else:
							nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=longitud,latitud=latitud,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
							nodo.save()							
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.nodo',id_manipulado=nodo.id)
							logs_model.save()
	
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:		
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.nodo')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#Inicio eliminacion de varios apoyos
@transaction.atomic
def eliminar_apoyos(request):

	sid = transaction.savepoint()
	try:
	
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
				HNodo.objects.get(pk=item['id']).delete()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.nodo',id_manipulado=item['id'])
				logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
					'data':''})


	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	

	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


#Fin eliminacion de varios apoyos



@login_required
def descargar_plantilla_apoyo_sinposicion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.write('A1', '*Los campos latitud y longitud no son campos obligatorio*', format1)
	worksheet.write('A2', 'Apoyo', format1)
	worksheet.write('B2', 'Latitud', format1)
	worksheet.write('C2', 'Longitud', format1)

	workbook.close()

	return response
    #return response


	
@login_required
@transaction.atomic
def guardar_apoyo_archivo_sinposicion(request):

	sid = transaction.savepoint()
	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		capa_id= request.POST['capa_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value:
					if i == 0 or i==1:
						i=i+1
					else:
						veri_nodo=HNodo.objects.filter(presupuesto_id=presupuesto_id,nombre=fila[0].value)

						if len(veri_nodo) == 0:
							latitud=None
							longitud=None

							if fila[1].value is not None and  fila[2].value is not None:
								latitud=fila[1].value
								longitud=fila[2].value

							nodo=HNodo(presupuesto_id=presupuesto_id,nombre=fila[0].value,capa_id=capa_id,longitud=longitud,latitud=latitud,noProgramado=False,eliminado=False,porcentajeAcumulado=0)
							nodo.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.nodo',id_manipulado=nodo.id)
							logs_model.save()
						else:
							transaction.savepoint_rollback(sid)
							return JsonResponse({'message':'No se puede repetir el nombre del nodo en este presupuesto','success':'error',
							'data':''})

		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def descargar_plantilla_cantidadApoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_cantidadApoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cantidad_Apoyo')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']	
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:Z', 20)


	worksheet.write('A1', 'Id', format1)
	worksheet.write('B1', 'Hitos', format1)
	worksheet.write('C1', 'Actividad', format1)
	worksheet.write('D1', 'Cod. UUCC', format1)
	worksheet.write('E1', 'Descripcion', format1)
	worksheet.write('F1', 'Cantidad', format1)

	apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	row=0
	col=6

	for item_apoyo in apoyos:
		worksheet.write(row,col, item_apoyo.nombre, format1)

		col +=1


	row=1
	col=0

	hitos = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)


	for item in hitos:
		worksheet.write(row, col, item.id,format2)
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)
		worksheet.write(row, col+1,str(capitulo.nombre),format2)
		worksheet.write(row, col+2,item.actividad.nombre,format2)
		worksheet.write(row, col+3,item.codigoUC,format2)	
		worksheet.write(row, col+4,item.descripcionUC,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		count=6
		for item_apoyo in apoyos:
			worksheet.write(row,col+count,0, format2)
			count +=1

		row +=1


	workbook.close()

	return response
    #return response


@login_required
def guardar_cantidadApoyo_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		presupuesto_id= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1
		numeroFila = 1
		apoyos = []
		if int(contador) > 0:
			sid = transaction.savepoint()
			for fila in hoja.rows:
				if fila[0].value:
					if i == 0:
						i=1
					else:
						if FDetallePresupuesto.objects.filter(id=fila[0].value).count() == 0:

							cantidadesNodoGuardadas = JCantidadesNodo.objects.filter(
								detallepresupuesto__presupuesto__id=presupuesto_id)
							cantidadesNodoGuardadas.delete()							

							return JsonResponse({
								'message':'No se encontro el Id indicado en la fila No.' + str(numeroFila) + \
								'. Sugerimos descargar la plantilla para verificar el Id en esta fila',
								'success':'error',
								'data':''})
						elif UnidadConstructiva.objects.filter(
							catalogo_id=FDetallePresupuesto.objects.get(id=fila[0].value).catalogoUnidadConstructiva,
							codigo=fila[3].value).count() == 0:

							cantidadesNodoGuardadas = JCantidadesNodo.objects.filter(
								detallepresupuesto__presupuesto__id=presupuesto_id)
							cantidadesNodoGuardadas.delete()

							return JsonResponse({
								'message':'No se encontro la UUCC indicada en la fila No. ' + str(numeroFila) + \
								'. Sugerimos descargar la plantilla para verificar el codigo de UUCC en esta fila',
								'success':'error',
								'data':''})
						else:
							#codigo para guardar las cantidades:
							if len(apoyos) == 0:
								apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
							
							count = 6
							for item_apoyo in apoyos:
								consultar_nodo=JCantidadesNodo.objects.filter(
									detallepresupuesto_id=fila[0].value,
									nodo_id=item_apoyo.id)

								if consultar_nodo.count() == 0:
									
									nodo=JCantidadesNodo(
										detallepresupuesto_id=fila[0].value,
										nodo_id=item_apoyo.id,cantidad=fila[count].value)
									nodo.save()
									logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.cantidad_nodo',id_manipulado=nodo.id)
									logs_model.save()

								else:
									nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
									nodo.cantidad=fila[count].value
									nodo.save()
									logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
									logs_model.save()
								count +=1



					numeroFila = numeroFila + 1		
			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
			# i=0
			# apoyos=[]
			# for fila in hoja.rows:
			# 	if fila[0].value:
			# 		if i == 0:
			# 			i=1
			# 		else:
			# 			######################################################################################
			# 			if len(apoyos) == 0:
			# 				apoyos=HNodo.objects.filter(presupuesto_id=presupuesto_id).order_by('id')

			# 			count=6
			# 			for item_apoyo  in apoyos:
						
			# 				consultar_nodo=JCantidadesNodo.objects.filter(detallepresupuesto_id=fila[0].value,nodo_id=item_apoyo.id)
			# 				if len(consultar_nodo)==0:
								
			# 						nodo=JCantidadesNodo(
			# 							detallepresupuesto_id=fila[0].value,
			# 							nodo_id=item_apoyo.id,cantidad=fila[count].value)
			# 						nodo.save()
			# 						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.cantidad_nodo',id_manipulado=nodo.id)
			# 						logs_model.save()

			# 				else:
			# 					nodo=JCantidadesNodo.objects.get(pk=consultar_nodo[0].id)
			# 					nodo.cantidad=fila[count].value
			# 					nodo.save()
			# 					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cantidad_nodo',id_manipulado=consultar_nodo[0].id)
			# 					logs_model.save()
			# 				count +=1
			# transaction.savepoint_commit(sid)
			# return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
			# 	'data':''})
		else:
			return JsonResponse({
				'message':'No se encontraron datos en el archivo cargado.',
				'success':'error',
				'data':''})			
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cantidad_nodo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def informe_cantidad_apoyo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="informe_cantidad_apoyo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Informe')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	presupuesto_id= request.GET['presupuesto_id']	
				
	detalle_presupuesto = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 30)

	worksheet.write('A1', 'Hitos', format1)
	worksheet.write('B1', 'Actividad', format1)
	worksheet.write('C1', 'Cod. UUCC', format1)
	worksheet.write('D1', 'Descripcion UUCC', format1)
	worksheet.write('E1', 'Cantidad Total', format1)
	worksheet.write('F1', 'Cantidad en Apoyos', format1)

	for item in detalle_presupuesto:

		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col, str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)
		worksheet.write(row, col+3,item.descripcionUC,format2)
		worksheet.write(row, col+4,item.cantidad,format2)

		porcentaje=JCantidadesNodo.objects.filter(detallepresupuesto_id=item.id).aggregate(Sum('cantidad'))
		porcentaje['cantidad__sum']=0 if porcentaje['cantidad__sum']==None else porcentaje['cantidad__sum']		

		worksheet.write(row, col+5,str(porcentaje['cantidad__sum']),format2)

		row +=1


	workbook.close()

	return response
    #return response


@transaction.atomic
@api_view(['POST',])
def guardar_cantidad_apoyo(request):

	sid = transaction.savepoint()
	try:
		respuesta=request.data
		#respuesta= json.loads(lista)
		##import pdb; pdb.set_trace()											
		#respuesta = json.loads(request.data['_content']) if request.data['_content']  else request.data		

		today = date.today()			
		uucc = FDetallePresupuesto.objects.get(pk=int(respuesta['detalle_presupuesto_id']))
		array = []				
		for item in respuesta['lista']:
			nodo=JCantidadesNodo.objects.get(pk=item['id'])
			if(float(nodo.cantidad) != float(item['cantidad'])):					
				array.append([{'cantidad_nueva':float(item['cantidad']),
							'apoyo':nodo.nodo.id,
							'cantidad_anterior':nodo.cantidad}])				
			nodo.cantidad=item['cantidad']
			nodo.save()				
						
		if array:
			reformado = EReformado(fecha_registro = today.strftime("%Y-%m-%d"), 
								usuario_registro_id = request.user.usuario.id,									
								)
			reformado.save()

			##import pdb; pdb.set_trace()
			for x in array:
				reformado_detalle = EReformadoDetalle(
									apoyo_id = int(x[0]['apoyo']),
									reformado_id = reformado.id,
									codigo_uucc = uucc.codigoUC,
									descripcion_uucc = uucc.descripcionUC,
									cantidad_anterior = float(x[0]['cantidad_anterior']),
									cantidad_final = float(x[0]['cantidad_nueva']),
									diferencia = float(x[0]['cantidad_nueva']) - float(x[0]['cantidad_anterior'])
									)			
				reformado_detalle.save()				
			

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cantidad_apoyo',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return Response({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})
			
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cantidad_nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@api_view(['GET',])
def consultar_avance_obra(request):
	#sid = transaction.savepoint()
	try:
		id_cronograma=request.GET['cronograma_id']
		presupuesto_id=request.GET['presupuesto_id']
		programando=request.GET['programando']
		dato=request.GET['dato']
		# estados=request.GET['id_estados']
		# listado_estado=estados.split(',')
		# ejecucion=request.GET['ejecucion']


		# listado_id=[]
		# cambio=LCambio.objects.filter(cronograma_id=id_cronograma,estado_id__in=listado_estado)

		# listado_id_ejecucion=[]
		# if int(ejecucion)>0:

		# 	listado_por=[]
		# 	if int(ejecucion) == 2:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__gte=100)
				
		# 	else:
		# 		listado_por=RPorcentajeApoyo.objects.filter(cronograma_id=id_cronograma,porcentaje__lt=100)

		# 	for obj in listado_por:
		# 		listado_id_ejecucion.append(obj.apoyo.id)

		# if cambio is not None:
		# 	for obj in cambio:
		# 		for item in obj.nodos.all():
		# 			listado_id.append(item.id)

		
		# cronograma=KCronograma.objects.get(pk=id_cronograma)
		
		

		# if len(listado_id_ejecucion)==0:
		# 	qset=qset&((Q(noProgramado=programando))|(Q(id__in=listado_id)))

		# if len(listado_id_ejecucion)>0:
		# 	qset=qset&(Q(id__in=listado_id_ejecucion))
		qset=(Q(presupuesto_id=presupuesto_id)&(Q(eliminado=False)))


		if dato!='':
			qset=qset&(Q(nombre__icontains=dato))

		listado=HNodo.objects.filter(qset).values(
			'id','nombre','longitud','latitud','capa__color',
			'porcentajeAcumulado')

		qset1=Q(nodoOrigen__presupuesto__id=presupuesto_id)

		# if len(listado_id)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id))

		# if len(listado_id_ejecucion)>0:
		# 	qset1=qset1&(Q(nodoOrigen__id__in=listado_id_ejecucion))

		listado_enlaces=IEnlace.objects.filter(qset1).values('id','capa__color','nodoOrigen__nombre','nodoOrigen__latitud','nodoOrigen__longitud','nodoDestino__nombre','nodoDestino__latitud','nodoDestino__longitud')

		#transaction.savepoint_commit(sid)
		return Response({'message':'','success':'ok',
				'data':{'datos':list(listado),'enlace':list(listado_enlaces)}})
		
	except Exception as e:
		#print(e)
		#transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#@login_required
#@transaction.atomic
@api_view(['GET',])
def consultar_ingresos_datos(request):
	#sid = transaction.savepoint()
	try:
		nodo_id=request.GET['nodo_id']
		reporte_id=request.GET['reporte_id']

		nodo=HNodo.objects.get(pk=nodo_id)
		datos=[]

		if nodo.noProgramado == False:
			listado=[]
			#print (nodo.presupuesto.sin_poste)
			if nodo.presupuesto.sin_poste == False:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id,cantidad__gt=0)
			else:
				listado=JCantidadesNodo.objects.filter(nodo_id=nodo.id)

			uucc = []
			materiales = []
			for item in listado:
				cantidad_ejecutada=DetalleReporteTrabajo.objects.filter(
					nodo_id=nodo_id,
					detallepresupuesto_id=item.detallepresupuesto.id
					).aggregate(Sum('cantidadEjecutada'))
				cantidad_ejecutada['cantidadEjecutada__sum']=0 if cantidad_ejecutada['cantidadEjecutada__sum']==None else cantidad_ejecutada['cantidadEjecutada__sum']
				editable = False
				if cantidad_ejecutada['cantidadEjecutada__sum'] < item.cantidad:
					editable = True
				uucc.append({
							'id_detalle':item.detallepresupuesto.id,
							'codigo':item.detallepresupuesto.codigoUC,
							'descripcion':item.detallepresupuesto.descripcionUC,
							'cantidad':item.cantidad,
							'cantidad_ejecutada':cantidad_ejecutada['cantidadEjecutada__sum'],
							'editable': editable
					})
				#Extraer los materiales:
				uc = UnidadConstructiva.objects.filter(
					codigo=item.detallepresupuesto.codigoUC,
					catalogo__id=item.detallepresupuesto.catalogoUnidadConstructiva.id).values('id')

				desgloceMat = DesgloceMaterial.objects.filter(
					unidadConstructiva__id=uc[0]['id']).values(
					'material__codigo','material__descripcion', 'cantidad')


				for obj in desgloceMat:
					materiales.append({
						'codigo' : obj['material__codigo'],
						'descripcion' : obj['material__descripcion'],
						'cantidad' : float(obj['cantidad']) * float(item.cantidad),
						'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['cantidadEjecutada__sum'])
					})
				#unificar cantidades de materialales:
				materialesAgrupados=[]
				
				for mat in materiales:
					agregar = True
					for obj in materialesAgrupados:
						if mat['codigo'] == obj['codigo']:
							obj['cantidad']	+= mat['cantidad']
							obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
							agregar = False
					if agregar:		
						materialesAgrupados.append(mat)				

			#Extraer datos para grafica del nodo:
			
			porcentajes = []
			queryset_a_ejecutar = JCantidadesNodo.objects.filter(
				nodo__id=nodo_id,cantidad__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre').annotate(total=Sum('cantidad'))			

			queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
				nodo__id=nodo_id,cantidadEjecutada__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre'
				).annotate(total=Sum('cantidadEjecutada'))

			

			for Aejecutar in queryset_a_ejecutar:
				agregado = False
				for ejecutado in queryset_ejecutada:
					if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
						agregado = True
						porcentajes.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
						})
					if agregado:
						break
				if agregado == False:
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': 0
					})	

			
				

			if listado:
				datos={
					'unidadesConstructivas': uucc,
					'materiales': materialesAgrupados,
					'porcentajeEjecucion': porcentajes
				}
			else:
				datos={
					'unidadesConstructivas': uucc,
					'materiales': [],
					'porcentajeEjecucion': porcentajes
				}	
		
		#transaction.savepoint_commit(sid)
		return Response({'message':'','success':'ok',
				'data':datos})
		
	except Exception as e:
		print(e)
		#transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.nodo')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@api_view(['POST',])
@transaction.atomic
def guardar_cambio_cantidades(request):
	sid = transaction.savepoint()
	try:
		# Visbal:
		respuesta=request.data
		# respuesta=request.data

		# # respuesta= json.loads(lista)
		# ##import pdb; pdb.set_trace()
		# for item in respuesta['lista']:

		# Didi:
		# if request.data['_content']:
			
		# lista=request.data['_content']
		# respuesta= json.loads(lista)

		#respuesta = json.loads(request.data['_content']) if request.data['_content']  else request.data
		##import pdb; pdb.set_trace()
		for item in respuesta['lista']:
			
			qset_cantidadAejecutar = JCantidadesNodo.objects.filter(
				detallepresupuesto_id= item['id_detalle'],
				nodo_id=respuesta['id_nodo']
				).values('cantidad')
			cantidadAejecutar = float(qset_cantidadAejecutar[0]['cantidad'])

			qset_cantidadEjecutada = DetalleReporteTrabajo.objects.filter(
				detallepresupuesto_id=item['id_detalle'],
				nodo_id=respuesta['id_nodo']
				).values('detallepresupuesto_id').annotate(
				cantidades=Sum('cantidadEjecutada')).distinct()

			if qset_cantidadEjecutada:	
				cantidadEjecutada = float(qset_cantidadEjecutada[0]['cantidades'])
			else:
				cantidadEjecutada = 0

			if (cantidadEjecutada + float(item['cantidad'])) <= cantidadAejecutar:

				cambio=DetalleReporteTrabajo(
					reporte_trabajo_id=respuesta['id_reporte'],
					nodo_id=respuesta['id_nodo'],
					detallepresupuesto_id=item['id_detalle'],
					cantidadEjecutada=item['cantidad'])
				cambio.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,
					accion=Acciones.accion_crear,
					nombre_modelo='avanceObraLite.detalle_reporte',
					id_manipulado=cambio.id)
				logs_model.save()

			else:
				dp = FDetallePresupuesto.objects.get(pk=item['id_detalle'])
				transaction.savepoint_rollback(sid)
				return Response({
					'message':'La cantidad a reportar en la unidad constructiva '+ dp.descripcionUC + \
					' sobrepasa la cantidad a ejecutar registrada en el sistema para este apoyo.',
					'success':'error',
					'data':''})



		reporte=ReporteTrabajo.objects.get(pk=respuesta['id_reporte'])
		
		if reporte.presupuesto.sin_poste==False:
			porcentajes = []
			queryset_a_ejecutar = JCantidadesNodo.objects.filter(
				nodo__id=respuesta['id_nodo'],cantidad__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre').annotate(total=Sum('cantidad'))			

			queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
				nodo__id=respuesta['id_nodo'],cantidadEjecutada__gt=0).values(
				'detallepresupuesto__actividad__id',
				'detallepresupuesto__actividad__nombre'
				).annotate(total=Sum('cantidadEjecutada'))

			

			for Aejecutar in queryset_a_ejecutar:
				agregado = False
				for ejecutado in queryset_ejecutada:
					if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
						agregado = True
						pr = (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
						if pr > 100:
							pr=100
						porcentajes.append({
							'id' : Aejecutar['detallepresupuesto__actividad__id'],
							#'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
							'porcentaje': pr
						})
					if agregado:
						break
				if agregado == False:
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						#'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': 0
					})
			total=0
			for p in porcentajes:
				total = total + float(p['porcentaje'])

			########################################################################################
			# detalle=DetalleReporteTrabajo.objects.filter(
			# 	reporte_trabajo_id=respuesta['id_reporte'],
			# 	nodo_id=respuesta['id_nodo']).values(
			# 	'detallepresupuesto_id').annotate(cantidades=Sum('cantidadEjecutada')).distinct()

			# total=0
			# por=0
			# for item2 in detalle:
			# 	cantidad_nodo=JCantidadesNodo.objects.filter(
			# 		detallepresupuesto_id=item2['detallepresupuesto_id'],
			# 		nodo_id=respuesta['id_nodo'])
			# 	valor=float(item2['cantidades'])/float(cantidad_nodo[0].cantidad)
			# 	por=por+valor

			# total=(por/len(detalle))*100

			# if total>100:
			# 	total=100
			########################################################################################
			porcentaje_apoyo=HNodo.objects.get(pk=respuesta['id_nodo'])
			porcentaje_apoyo.porcentajeAcumulado=round((float(total)/len(porcentajes)),2)
			porcentaje_apoyo.save()


		#createAsyncEstado.delay(respuesta['id_cronograma'])

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cambio_ejecutada')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@api_view(['POST',])
@transaction.atomic
def guardar_cambio_detalle(request):
	sid = transaction.savepoint()
	try:
		respuesta=request.data
		#respuesta= json.loads(lista)


		for item in respuesta['lista']:
			detalle=DetalleReporteTrabajo.objects.get(pk=item['id'])
			detalle.cantidadEjecutada=item['cantidad']
			detalle.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cambio_detalle',id_manipulado=detalle.id)
			logs_model.save()



		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return Response({'message':'El registro ha sido registrado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cambio')
		
		return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_id_nodo_destino(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print ("texto")

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			IEnlace.objects.get(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='avanceObraLite.enlance',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.presupuesto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@login_required
@transaction.atomic
def cierre_programacion(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cronograma=Cronograma.objects.get(id=respuesta['id_cronograma'])
		cronograma.programacionCerrada=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cronograma',id_manipulado=respuesta['id_cronograma'])
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cierre_programacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def informe_detallepresupuesto(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="presupuesto.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Presupuesto')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	

	presupuesto_id= request.GET['presupuesto_id']
	presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
				
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 30)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:D', 20)
	worksheet.set_column('E:E', 30)
	worksheet.set_column('F:F', 20)
	worksheet.set_column('G:H', 20)


	worksheet.write('A2', 'Presupuesto', format1)
	worksheet.write('B2', presupuesto.nombre, format1)

	worksheet.write('A3', 'Proyecto', format1)
	worksheet.write('B3', presupuesto.cronograma.proyecto.nombre, format1)

	worksheet.write('A4', 'Esquema', format1)
	worksheet.write('B4', presupuesto.cronograma.esquema.nombre, format1)

	worksheet.write('F6', 'Total Presupuesto', format1)

	worksheet.write('A7', 'Hitos', format1)
	worksheet.write('B7', 'Actividad', format1)
	worksheet.write('C7', 'codigo UUCC', format1)
	worksheet.write('D7', 'Descripcion UUCC', format1)
	worksheet.write('E7', 'Valor UUCC', format1)
	worksheet.write('F7', 'Cantidad', format1)
	worksheet.write('G7', 'Subtotal', format1)

	row=7
	col=0

	detalle = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id)
	total=0

	for item in detalle:
		capitulo=CEsquemaCapitulosActividadesG.objects.get(pk=item.actividad.padre)

		worksheet.write(row, col,str(capitulo.nombre),format2)
		worksheet.write(row, col+1,item.actividad.nombre,format2)
		worksheet.write(row, col+2,item.codigoUC,format2)	
		worksheet.write(row, col+3,item.descripcionUC,format2)

		valorManoObra=0
		if item.valorManoObra is not None:
			valorManoObra=item.valorManoObra

		valorMaterial=0
		if item.valorMaterial is not None:
			valorMaterial=item.valorMaterial

		valor=(valorManoObra)+(valorMaterial)

		worksheet.write(row, col+4,valor,format2)	
		worksheet.write(row, col+5,item.cantidad,format2)

		subtotal=valor*(item.cantidad)

		worksheet.write(row, col+6,subtotal,format2)

		total=total+subtotal

		row +=1


	worksheet.write('G6', total, format1)

	workbook.close()

	return response
    #return response


@login_required
def menu_gps(request,id_presupuesto):

	try:
		nodo=HNodo.objects.filter(presupuesto_id=id_presupuesto)
		
		if len(nodo) == 0:	
			return JsonResponse({'message':'No tiene ningun punto de gps registrado','success':'ok',
					'data':['1']})
		else:
			sqlnodo=HNodo.objects.filter(presupuesto_id=id_presupuesto,latitud=None,longitud=None)

			if len(sqlnodo) == 0:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['2']})
			else:
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':['3']})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avanceObraLite.cierre_programacion')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def guardar_reporte_trabajo(request,id_reporte):

	sid = transaction.savepoint()
	try:
		reporte=ReporteTrabajo.objects.get(pk=id_reporte)
		#detalle=DetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)
		detalles=FDetallePresupuesto.objects.filter(presupuesto_id=reporte.presupuesto.id)

		# model_actividad['peso__sum']=0 if model_actividad['peso__sum']==None else model_actividad['peso__sum']
		# valor=round(float(model_actividad['peso__sum']),3)+round(float(request.data['peso']),3)
		##import pdb; pdb.set_trace()
		detalle_reporte=DetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte)

		if len(detalle_reporte) == 0:
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Debe ingresar alguna cantidad ejecutada en el reporte','success':'error',
			'data':''})

		total_porcentaje=0
		total_pagado=0
		for item in detalles:
			sql_detalletrabajo=DetalleReporteTrabajo.objects.filter(reporte_trabajo_id=id_reporte,detallepresupuesto_id=item.id).aggregate(Sum('cantidadEjecutada'))
			sql_detalletrabajo['cantidadEjecutada__sum']=0 if sql_detalletrabajo['cantidadEjecutada__sum']==None else sql_detalletrabajo['cantidadEjecutada__sum']
			#valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])+float(item.cantidad)
			valor=float(sql_detalletrabajo['cantidadEjecutada__sum'])/float(item.cantidad)
			item.porcentaje=0 if item.porcentaje==None else item.porcentaje
			
			#porcentaje=(valor*float(item.porcentaje))/100
			porcentaje=(valor*float(item.porcentaje))
			total_porcentaje=total_porcentaje+porcentaje

			valor2=float(item.valorManoObra)+float(item.valorMaterial)
			pagado=float(sql_detalletrabajo['cantidadEjecutada__sum'])*float(valor2)
			total_pagado=total_pagado+pagado

		if total_porcentaje>100:
			total_porcentaje=100

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		estado_rechazo=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		estado_corregido=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)

		reporte.valor_ganando_acumulado=round(total_pagado,3)
		reporte.avance_obra_acumulado=round(total_porcentaje,3)
		reporte.reporteCerrado=True

		if int(reporte.estado_id)==int(estado_rechazo[0].id):			
			reporte.estado_id=estado_corregido[0].id
		else:
			reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.reporte',id_manipulado=id_reporte)
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cierre_programacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#serializer de empresa
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')



#serializer lite de proyecto empresa
class ProyectoEmpresaLite2Serializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLite2Serializer(read_only = True )
	empresa = EmpresaLiteSerializer(read_only = True  )

	totalCorrregidos=serializers.SerializerMethodField()

	totalAprobados=serializers.SerializerMethodField()

	totalRegistrados=serializers.SerializerMethodField()

	totalRechazados=serializers.SerializerMethodField()

	class Meta:
		model = Proyecto_empresas
		fields=('id' ,
				'proyecto',
 				'empresa',
 				'totalCorrregidos',	
 				'totalAprobados',	
 				'totalRegistrados',
 				'totalRechazados'
				)

	def get_totalCorrregidos(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		return ReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalAprobados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		return ReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRegistrados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		return ReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()

	def get_totalRechazados(self, obj):
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		return ReporteTrabajo.objects.filter(presupuesto__cronograma__proyecto__id=obj.proyecto.id,estado__id=estados[0].id).count()


#Api lite de empresa proyecto
class ProyectoReporteAvanceViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las categorias de administrador de fotos
	"""
	model=Proyecto_empresas
	queryset = model.objects.all()
	serializer_class = ProyectoEmpresaLite2Serializer
	nombre_modulo='administrador_foto.Proyecto_empresa'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ProyectoReporteAvanceViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			empresa = self.request.query_params.get('empresa', None)
			mcontrato = self.request.query_params.get('mcontrato', None)
			departamento = self.request.query_params.get('departamento', None)
			municipio = self.request.query_params.get('municipio', None)
			contratista = self.request.query_params.get('contratista', None)
			corregido = self.request.query_params.get('corregido', None)
			registrado = self.request.query_params.get('registrado', None)
			aprobado = self.request.query_params.get('aprobado', None)
			rechazados = self.request.query_params.get('rechazados', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			qset=(~Q(id=0))

			if corregido is not None:
				reporte=ReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			if rechazados is not None:
				reporte=ReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				qset= qset & (Q(proyecto__id__in = lista))

			
			if registrado is not None and int(registrado)==1:
				reporte=ReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))

			if aprobado is not None and int(aprobado)==1:
				reporte=ReporteTrabajo.objects.all()
				estado=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
				lista=[]
				for item in reporte:
					if int(item.estado_id)==int(estado[0].id):
						lista.append(item.presupuesto.cronograma.proyecto.id)

				if len(lista) > 0:
					qset= qset & (Q(proyecto__id__in = lista))


			if(dato or empresa or mcontrato or departamento or municipio):

				if dato:
					qset = qset & ( Q(empresa__nombre__icontains = dato) |
								Q(empresa__nit__icontains = dato) |
								Q(proyecto__nombre__icontains = dato))					
			

				if empresa and int(empresa)>0:
					qset = qset & (Q(empresa__id = empresa))					


				if mcontrato and int(mcontrato)>0:
					qset = qset & (Q(proyecto__mcontrato = mcontrato))
					
				if departamento and int(departamento)>0:
					qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
					
				if municipio and int(municipio)>0:
					qset = qset & (Q(proyecto__municipio = municipio))
					

				if contratista and int(contratista)>0:
					qset = qset & (Q(proyecto__contrato__contratista__id = contratista))	

			
			if qset != '':
				queryset = self.model.objects.filter(qset).distinct()

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@login_required
@transaction.atomic
def guardar_archivo_aprobacion(request):

	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		reporte_id= request.POST['reporte_id']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=2)
		
		reporte=ReporteTrabajo.objects.get(pk=reporte_id)
		reporte.soporteAprobacion=soporte
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.soporte_aprobacion',id_manipulado=reporte_id)
		logs_model.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.soporte_aprobacion')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def guardar_rechazo_reporte(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		reporte_id= respuesta['reporte_id']
		mensaje= respuesta['mensaje']
		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		
		reporte=ReporteTrabajo.objects.get(pk=reporte_id)
		reporte.usuario_aprueba_id=request.user.usuario.id
		reporte.estado_id=estados[0].id
		reporte.reporteCerrado=False
		reporte.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.rechazo_reporte',id_manipulado=reporte_id)
		logs_model.save()

		motivo=MComentarioRechazo(reporte_trabajo_id=reporte_id,motivoRechazo=mensaje)
		motivo.save()
		logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.motivo_rechazo',id_manipulado=motivo.id)
		logs_model2.save()
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.rechazo_reporte')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def consultar_graficos(request):
	sid = transaction.savepoint()
	try:
		presupuesto_id=request.GET['presupuesto_id']
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		listado=[]
		listado_porcentaje=DiagramaGrahm.objects.filter(cronograma_id=presupuesto.cronograma.id).values('fechaFinal').annotate(porcentaje=Sum('actividad__peso')).distinct()
		#print(listado_porcentaje)
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado.append({
					'fecha':item['fechaFinal'],
					'porcentaje':round(porcentaje,2)
					})

		#print(listado)


		listado3=[]
		listado_porcentaje=ReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(porcentaje=Sum('avance_obra_acumulado')).distinct()
				
		porcentaje=0
		for item in listado_porcentaje:
			porcentaje=porcentaje+float(item['porcentaje'])
			if porcentaje>100:
				porcentaje=100
			listado3.append({
					'fecha':item['fechaTrabajo'],
					'porcentaje':round(porcentaje,2)
					})

		listado4=[]
		listado_prespuesto=ReporteTrabajo.objects.filter(presupuesto_id=presupuesto_id).values('fechaTrabajo').annotate(valor_ganando=Sum('valor_ganando_acumulado')).distinct()
				
		valor_ganando=0
		for item in listado_prespuesto:
			valor_ganando=valor_ganando+float(item['valor_ganando'])
			
			listado4.append({
					'fecha':item['fechaTrabajo'],
					'valor_ganando':round(valor_ganando,2)
					})

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'','success':'ok',
				'data':{'programada':list(listado),'avance':list(listado3),'presupuesto':list(listado4)}})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.grafico')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def consultar_cantidad_reportes(request):

	try:

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=5)
		cantidad_corregidos=ReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
		cantidad_registrado=ReporteTrabajo.objects.filter(estado__id=estados[0].id).count()

		estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=3)
		cantidad_rechazados=ReporteTrabajo.objects.filter(estado__id=estados[0].id).count()
		

		return JsonResponse({'message':'','success':'ok',
					'data':{'cantidad_corregidos':cantidad_corregidos,'cantidad_registrado':cantidad_registrado,'cantidad_rechazados':cantidad_rechazados}})

		
	except Exception as e:
		print(e)
		functions.toLog(e,'avanceObraLite.cantidad_reportes')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def actualizar_cancelado(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
		for item in respuesta['lista']:
			cambio=LCambio.objects.get(id=item['id'])
			cambio.estado_id=estado[0].id
			cambio.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cambio',id_manipulado=item['id'])
			logs_model.save()

			historial=LHistorialCambio(cambio_id=item['id'],usuario_registro_id=request.user.usuario.id,motivoEstado=respuesta['motivo'],estado_id=estado[0].id)
			historial.save()
			logs_model2=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='avanceObraLite.historial_cambio',id_manipulado=historial.id)
			logs_model2.save()
			

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.actualizar_cantidad')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
@transaction.atomic
def reportar_sin_poste(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		presupuesto_id= respuesta['presupuesto_id']
		
		presupuesto=EPresupuesto.objects.get(pk=presupuesto_id)
		presupuesto.sin_poste=True
		presupuesto.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.sin_poste',id_manipulado=presupuesto_id)
		logs_model.save()

		# agregarSinPoste.delay(presupuesto_id,request.user.usuario.id)
		
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.sin_poste')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def avanceObraLite(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraLite/hitos.html',{'model':'besquemacapitulosg','app':'avanceObraLite','macrocontrato':ListMacro})

@login_required
def catalogo(request):	
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)		
	return render(request,'avanceObraLite/catalogo.html',{'model':'catalogounidadconstructiva','app':'avanceObraLite','macrocontrato':ListMacro})

@login_required
def uucc(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraLite/uucc.html',{'model':'unidadconstructiva','app':'avanceObraLite', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def materiales(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraLite/materiales.html',{'model':'material','app':'avanceObraLite', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def mano_obra(request,catalogo_id):
	#import pdb; pdb.set_trace()		
	catalogo = CatalogoUnidadConstructiva.objects.filter(id=catalogo_id)
	catalogo = catalogo[0]
	if catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraLite/manoObra.html',{'model':'manodeobra','app':'avanceObraLite', 'catalogo_id':int(catalogo_id), 'catalogo': catalogo, 'catalogo_activo': catalogo_activo })

@login_required
def desgloce_mat(request,uucc_id):
	#import pdb; pdb.set_trace()		
	uucc = UnidadConstructiva.objects.filter(id=uucc_id)
	uucc = uucc[0]
	if uucc.catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraLite/desgloce_mat.html',{'model':'desglocematerial','app':'avanceObraLite', 'uucc_id':int(uucc_id), 'catalogo': uucc.catalogo, 'catalogo_activo': catalogo_activo, 'uucc': uucc })

@login_required
def desgloce_mo(request,uucc_id):
	#import pdb; pdb.set_trace()		
	uucc = UnidadConstructiva.objects.filter(id=uucc_id)
	uucc = uucc[0]
	if uucc.catalogo.activo is True:
		catalogo_activo = 'Activo'
	else:
		catalogo_activo = 'Inactivo'
	return render(request,'avanceObraLite/desgloce_mo.html',{'model':'desglocemanodeobra','app':'avanceObraLite', 'uucc_id':int(uucc_id), 'catalogo': uucc.catalogo, 'catalogo_activo': catalogo_activo, 'uucc': uucc })


@login_required
def actividades(request,id_esquema):
	return render(request, 'avanceObraLite/actividad.html',{'model':'cesquemacapitulosactividadesg','app':'avanceObraLite','id_esquema':id_esquema})


@login_required
def regla_estado(request,id_esquema):
	return render(request, 'avanceObraLite/regla.html',{'model':'epresupuesto','app':'avanceObraGrafico','id_esquema':id_esquema})


@login_required
def cronograma(request):
	tipo=tipoC()
	qset = (Q(contrato__tipo_contrato=tipo.m_contrato)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
	ListMacro = EmpresaContrato.objects.filter(qset)
	return render(request, 'avanceObraLite/cronograma.html',{'model':'cronograma','app':'avanceObraLite','macrocontrato':ListMacro})



@login_required
def cronograma_proyecto(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	esquema=BEsquemaCapitulosG.objects.filter(macrocontrato_id=proyecto.mcontrato)
	periodo=APeriodicidadG.objects.all()
	return render(request, 'avanceObraLite/cronograma_proyecto.html',{'model':'cronograma','app':'avanceObraLite','id_proyecto':id_proyecto,'proyecto':proyecto,'esquema':esquema,'periodo':periodo})

@login_required
def tablero_contrato(request,id_mcontrato):
	mcontrato=Contrato.objects.get(pk=id_mcontrato)
	return render(request, 'avanceObraLite/tablero_contrato.html',
		{'model':'cronograma',
		'app':'avanceObraLite',
		'mcontrato':mcontrato})



@login_required
def presupuesto(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=querycronograma.esquema.id,nivel=1)
	catalogos = CatalogoUnidadConstructiva.objects.filter(activo=True)
	return render(request, 'avanceObraLite/presupuesto.html',
		{'model':'epresupuesto',
		'app':'avanceObraLite',
		'cronograma':querycronograma,
		'cronograma_id':id_cronograma,
		'capitulos':capitulos,
		'proyecto_id':querycronograma.proyecto.id, 
		'catalogos':catalogos})


@login_required
def presupuesto_detalle(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=presupuesto.cronograma.esquema.id,nivel=1)
	catalogos = CatalogoUnidadConstructiva.objects.filter(activo=True)
	return render(request, 'avanceObraLite/detalle_presupuesto.html',
		{'model':'fdetallepresupuesto',
		'app':'avanceObraGrafico',
		'presupuesto_id':id_presupuesto,
		'presupuesto':presupuesto,
		'proyecto_id':presupuesto.cronograma.proyecto.id,
		'capitulos':capitulos,
		'cronograma_id':presupuesto.cronograma.id,
		'catalogos':catalogos})


@login_required
def programacion(request,id_cronograma):
	querycronograma=Cronograma.objects.get(pk=id_cronograma)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=querycronograma.esquema.id,nivel=1)
	actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=querycronograma.esquema.id,nivel=2)
	presupuestos = EPresupuesto.objects.filter(cronograma_id=id_cronograma,cerrar_presupuesto=True)
	periodicidades=APeriodicidadG.objects.all()
	return render(request, 'avanceObraLite/diagrama.html',
		{'model':'diagramagrahm',
		'app':'avanceObraLite',
		'presupuestos':presupuestos,
		'cronograma':querycronograma,
		'cronograma_id':id_cronograma,
		'proyecto_id':querycronograma.proyecto.id,
		'capitulos':capitulos,
		'actividades': actividades,
		'periodicidades':periodicidades,
		})



@login_required
def apoyo_con_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraLite/apoyo_con_gps.html',{'model':'hnodo','app':'avanceObraLite','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def apoyo_sin_gps(request,id_presupuesto):
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	capa2=GCapa.objects.filter(nombre__icontains='archivo').last()
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraLite/apoyo_sin_gps.html',{'model':'hnodo','app':'avanceObraLite','id_presupuesto':id_presupuesto,'id_capa_manual':capa.id,'id_capa_archivo':capa2.id,'presupuesto':presupuesto,'proyecto_id':presupuesto.cronograma.proyecto.id,'cronograma_id':presupuesto.cronograma.id})


@login_required
def cantidad_apoyo(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.cronograma.esquema.id,nivel=1)
	return render(request, 'avanceObraLite/cantidad_apoyo.html',{'model':'epresupuesto','app':'avanceObraLite','cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'capitulos':capitulos})


@login_required
def cantidad_apoyo_id(request,id_presupuesto,id_detalle):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	detalle_presupuesto=FDetallePresupuesto.objects.get(pk=id_detalle)
	return render(request, 'avanceObraLite/cantidad_apoyo_id.html',{'model':'epresupuesto','app':'avanceObraLite','presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto,'detalle_presupuesto':detalle_presupuesto,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'cronograma_id':nombre_presupuesto.cronograma.id})


@login_required
def reporte_trabajo(request,id_cronograma):
	nombre_presupuesto=EPresupuesto.objects.filter(cronograma_id=id_cronograma,cerrar_presupuesto=True).last()
	estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
	estados2=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)

	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	
	capitulos=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.cronograma.esquema.id,nivel=1)
	actividades=CEsquemaCapitulosActividadesG.objects.filter(esquema_id=nombre_presupuesto.cronograma.esquema.id,nivel=2)		
	usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraLite/reporte.html',{
		'model':'jreportetrabajo',
		'app':'avanceObraLite',
		'Usuarios':usuarios,
		'estado_id_procesado':estados2[0].id,
		'estado_id_registrado':estados[0].id,
		'cronograma_id':nombre_presupuesto.cronograma.id,
		'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,
		'presupuesto_id':nombre_presupuesto.id,
		'presupuesto':nombre_presupuesto,
		'actividades':actividades,
		'capitulos':capitulos,})

@login_required
def reformado(request,id_presupuesto):
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	estados=Estado.objects.filter(app='Avance_obra_grafico',codigo=1)
	estados2=Estado.objects.filter(app='Avance_obra_grafico',codigo=4)

	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	
				
	usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	
	return render(request, 'avanceObraLite/reformado.html',{'model':'jreportetrabajo','app':'avanceObraLite','Usuarios':usuarios,'estado_id_procesado':estados2[0].id,'estado_id_registrado':estados[0].id,'cronograma_id':nombre_presupuesto.cronograma.id,'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'presupuesto':nombre_presupuesto})

@login_required
def reformadoDetalle(request,id_reformado):
	##import pdb; pdb.set_trace()
	reformado=EReformado.objects.get(pk=int(id_reformado))	
	detalle = EReformadoDetalle.objects.filter(reformado__id = int(reformado.id)).values('apoyo__presupuesto_id').distinct()
	detalle = detalle[0]['apoyo__presupuesto_id']
	# empresa_id=0
	# if nombre_presupuesto is not None:
	# 	for valor in nombre_presupuesto.cronograma.proyecto.contrato.all():
	# 		if int(valor.tipo_contrato.id)==9:
	# 			empresa_id=valor.contratista.id	
	
	return render(request, 'avanceObraLite/reformadoDetalle.html',{'model':'ereformado','app':'avanceObraLite','reformado':reformado, 'presupuesto_id': detalle})

@login_required
def avance_con_gps(request,id_reporte):
	reporte=ReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	valoraciones = Tipo.objects.filter(app='no_conformidad_valoracion')
	tipos = Tipo.objects.filter(app='no_conformidad')
	return render(request, 'avanceObraLite/avance_con_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraLite','capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte,'tipos': tipos, 'valoraciones': valoraciones})


@login_required
def avance_sin_gps(request,id_reporte):
	reporte=ReporteTrabajo.objects.get(pk=id_reporte)
	capa=GCapa.objects.filter(nombre__icontains='manual').last()
	estados_cambios=Estado.objects.filter(app='Avance_obra_grafico')
	return render(request, 'avanceObraLite/avance_sin_gps.html',{'model':'kdetallereportetrabajo','app':'avanceObraLite','estados':estados_cambios,'capa_id':capa.id,'cronograma_id':reporte.presupuesto.cronograma.id,'proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'presupuesto_id':reporte.presupuesto.id,'reporte':reporte,'reporte_id':id_reporte})



@login_required
def aprobacion(request):
	return render(request, 'avanceObraLite/aprobacion.html',{'model':'jreportetrabajo','app':'avanceObraLite'})


@login_required
def corregido(request):
	return render(request, 'avanceObraLite/corregido.html',{'model':'jreportetrabajo','app':'avanceObraLite'})


@login_required
def registrado(request):
	return render(request, 'avanceObraLite/registrado.html',{'model':'jreportetrabajo','app':'avanceObraLite'})


@login_required
def rechazados(request):
	return render(request, 'avanceObraLite/rechazados.html',{'model':'jreportetrabajo','app':'avanceObraLite'})


@login_required
def reporte_trabajo_registrado(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraLite/reporte_registrado.html',{'model':'jreportetrabajo','app':'avanceObraLite','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_registrado(request,id_reporte):
	reporte=ReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraLite/detalle_registrado.html',{'model':'kdetallereportetrabajo','app':'avanceObraLite','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def reporte_trabajo_corregido(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraLite/reporte_corregido.html',{'model':'jreportetrabajo','app':'avanceObraLite','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def detalle_corregido(request,id_reporte):
	reporte=ReporteTrabajo.objects.get(pk=id_reporte)
	return render(request, 'avanceObraLite/detalle_corregido.html',{'model':'kdetallereportetrabajo','app':'avanceObraLite','proyecto_id':reporte.presupuesto.cronograma.proyecto.id,'reporte':reporte,'id_reporte':id_reporte})


@login_required
def grafico(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	return render(request, 'avanceObraLite/grafico.html',{'model':'reportetrabajo','app':'avanceObraLite','proyecto_id':presupuesto.cronograma.proyecto.id,'presupuesto_id':id_presupuesto,'cronograma_id':presupuesto.cronograma.id,'presupuesto':presupuesto})


@login_required
def reporte_trabajo_rechazados(request,id_proyecto):
	proyecto=Proyecto.objects.get(pk=id_proyecto)
	return render(request, 'avanceObraLite/reporte_rechazados.html',{'model':'reportetrabajo','app':'avanceObraLite','proyecto_id':id_proyecto,'proyecto':proyecto})


@login_required
def cambios(request):
	return render(request, 'avanceObraLite/cambios.html',{'model':'lcambio','app':'avanceObraLite'})


@login_required
def index_cambio(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	empresa=Empresa.objects.filter(~Q(id=request.user.usuario.empresa.id))
	estados=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=101)
	cancelado=Estado.objects.filter(app='Avance_obra_grafico_cambio',codigo=105)
	return render(request, 'avanceObraLite/index_cambio.html',{'model':'lcambio','app':'avanceObraLite','estado_cancelado':cancelado[0].id,'estado_id':estados[0].id,'presupuesto_id':id_presupuesto,'presupuesto':presupuesto,'empresa':empresa})


@login_required
def detalle_cambio(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	return render(request, 'avanceObraLite/detalle_cambio.html',{'model':'lcambio','app':'avanceObraLite','cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def agregar_detalle(request,id_cambio):
	cambio=LCambio.objects.get(pk=id_cambio)
	apoyos=HNodo.objects.filter(presupuesto_id=cambio.presupuesto.id)
	return render(request, 'avanceObraLite/agregar_detalle.html',{'model':'lcambio','app':'avanceObraLite','apoyos':apoyos,'cambio_id':id_cambio,'cambio':cambio,'presupuesto_id':cambio.presupuesto.id})


@login_required
def aprobacion_cambio(request):
	return render(request, 'avanceObraLite/aprobacion_cambio.html',{'model':'lcambio','app':'avanceObraLite'})


@login_required
def autorizacion_cambio(request):
	return render(request, 'avanceObraLite/autorizacion_cambio.html',{'model':'lcambio','app':'avanceObraLite'})




#API Manejo de unidades constructivas
#Tipos de unidades constructivas
from .models import TipoUnidadConstructiva
class TipoUnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model=TipoUnidadConstructiva
		fields=('id','nombre','activa')

class TipoUnidadConstructivaViewSet(viewsets.ModelViewSet):
	model=TipoUnidadConstructiva
	queryset = model.objects.all()
	serializer_class = TipoUnidadConstructivaSerializer
	nombre_modulo='avanceObraLite.tipoUnidadConstructiva'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(TipoUnidadConstructivaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			activa = self.request.query_params.get('activa', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))

			if dato:
				qset = qset & (Q(nombre__icontains=dato))
			if activa:
				qset = 	qset & (Q(activa=activa))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#catalogos de unidades constructivas


class CatalogoUnidadConstructivaViewSet(viewsets.ModelViewSet):
	model = CatalogoUnidadConstructiva
	queryset = model.objects.all()
	serializer_class = CatalogoUnidadConstructivaSerializer
	nombre_modulo='avanceObraLite.catalogoUnidadConstructiva'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:	
			#import pdb; pdb.set_trace()		
			queryset = super(CatalogoUnidadConstructivaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			activo = self.request.query_params.get('activo', None)
			ano = self.request.query_params.get('ano', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''


			qsetMcontrato = (Q(empresa_id=request.user.usuario.empresa.id))
			listContratos = EmpresaContrato.objects.filter(qsetMcontrato).values('contrato_id').distinct()
			qset = (~Q(id=0))
			if request.user.usuario.empresa.id != 4:
				qset = qset & (Q(mcontrato_id__in=listContratos))
			
			if dato:
				qset = qset & (Q(nombre__icontains=dato))
			if activo:
				qset = qset & (Q(activo=activo))
			if ano:
				qset = qset & (Q(ano=ano))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				validator = self.model.objects.filter(nombre=request.data['nombre'])

				if not validator:

					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(mcontrato_id=int(request.data['mcontrato_id']))
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
							
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un catalogo con el nombre suministrado.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				catalogo = self.model.objects.get(pk=int(request.data['id']))
				if catalogo:
					catalogo.ano = request.data['ano']
					catalogo.mcontrato_id = int(request.data['mcontrato_id'])
					catalogo.nombre = request.data['nombre']
					catalogo.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  
#unidades constructivas

class MaterialesViewSet(viewsets.ModelViewSet):
	model = Material
	queryset = model.objects.all()
	serializer_class = MaterialSerializer
	nombre_modulo='avanceObraLite.materiales'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()			
			queryset = super(MaterialesViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			pk = self.request.query_params.get('pk', None)
			catalogo_id = self.request.query_params.get('catalogo', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			lite = self.request.query_params.get('lite',None)
			mensaje=''

			qset = (~Q(id=0))
			if catalogo_id:
				qset = qset & (Q(catalogo_id=catalogo_id))
			if dato:
				qset = qset & (Q(descripcion__icontains=dato) | Q(codigo__icontains=dato))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					if pk:
						queryset = 	self.model.objects.filter(id=int(pk))				
					serializer = MaterialSerializerLite(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)								
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo'])) & Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()
				if not validator:
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(catalogo_id=request.data['catalogo_id'],valorUnitario=request.data['valorUnitario'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)

				else:
					return Response({'message':'Ya existe un Material con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)						
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  

class ManoObraViewSet(viewsets.ModelViewSet):
	model = ManoDeObra
	queryset = model.objects.all()
	serializer_class = ManoObraSerializer
	nombre_modulo='avanceObraLite.ManoDeObra'
	model_log=Logs
	model_acciones=Acciones	

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()			
			queryset = super(ManoObraViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			catalogo_id = self.request.query_params.get('catalogo', None)			
			lite = self.request.query_params.get('lite',None)
			pk = self.request.query_params.get('pk', None)
			mensaje=''

			qset = (~Q(id=0))
			if catalogo_id:
				qset = qset & (Q(catalogo_id=catalogo_id))
			if dato:
				qset = qset & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					if pk:
						queryset = 	self.model.objects.filter(id=int(pk))				
					serializer = ManoObraSerializerLite(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)								
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})								
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo']))& Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()
				if not validator:
					
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(catalogo_id=request.data['catalogo_id'],valorHora=request.data['valorHora'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un registro de Mano de Obra con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  

class UnidadConstructivaLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)

	class Meta:
		model=UnidadConstructiva
		fields=('id','catalogo','tipoUnidadConstructiva',
			'codigo','descripcion')

class UnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):
	catalogo = CatalogoUnidadConstructivaSerializer(read_only=True)
	catalogo_id = serializers.PrimaryKeyRelatedField(read_only=True)

	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	totalManoDeObra = serializers.SerializerMethodField()
	totalMateriales = serializers.SerializerMethodField()

	class Meta:
		model=UnidadConstructiva
		fields=('id','catalogo','catalogo_id','tipoUnidadConstructiva','tipoUnidadConstructiva_id',
			'codigo','descripcion','totalManoDeObra','totalMateriales')

	def get_totalManoDeObra(self, obj):	
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'manoDeObra__valorHora','rendimiento')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		return round(total,2)

	def get_totalMateriales(self,obj):
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'material__valorUnitario','cantidad')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		return round(total,2)

class UnidadConstructivaSerializerLite(serializers.HyperlinkedModelSerializer):

	tipoUnidadConstructiva = TipoUnidadConstructivaSerializer(read_only=True)
	tipoUnidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	totalManoDeObra = serializers.SerializerMethodField()
	totalMateriales = serializers.SerializerMethodField()

	class Meta:
		model=UnidadConstructiva
		fields=('id','tipoUnidadConstructiva','tipoUnidadConstructiva_id',
			'codigo','descripcion','totalManoDeObra','totalMateriales')

	def get_totalManoDeObra(self, obj):	
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'manoDeObra__valorHora','rendimiento')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		return round(total,2)

	def get_totalMateriales(self,obj):
		total = 0
		##import pdb; pdb.set_trace()
		queryset = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=obj.id).values(
			'material__valorUnitario','cantidad')
		
		if queryset:
			for row in queryset:
				total = float(total) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		return round(total,2)

class DesgloceMaterialSerializer(serializers.HyperlinkedModelSerializer):
	
	material = MaterialSerializer(read_only=True)
	material_id = serializers.PrimaryKeyRelatedField(read_only=True)

	unidadConstructiva = UnidadConstructivaSerializer(read_only=True)
	unidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=DesgloceMaterial
		fields=('id','cantidad','material', 'material_id','unidadConstructiva','unidadConstructiva_id')		

class DesgloceMaterialViewSet(viewsets.ModelViewSet):
	model = DesgloceMaterial
	queryset = model.objects.all()
	serializer_class = DesgloceMaterialSerializer
	nombre_modulo='avanceObraLite.DesgloceMaterial'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()		
			queryset = super(DesgloceMaterialViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			uucc_id = self.request.query_params.get('uucc_id', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))
			qset = qset & (Q(unidadConstructiva_id=uucc_id))
			
			if dato:
				qset = qset & (Q(material__descripcion__icontains=dato) | Q(material__codigo__icontains=dato)) 

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = (Q(material_id=request.data['material_id']) & Q(unidadConstructiva_id=request.data['uucc_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:				
					cantidad = request.data['cantidad']
					if ',' in cantidad:
						request.data['cantidad'] = cantidad.replace(',','.')

					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(material_id=request.data['material_id'], unidadConstructiva_id=request.data['uucc_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)

				else:
					return Response({'message':'Ya existe un Desglose de Material con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)						
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				data = self.model.objects.get(pk=int(request.data['id']))
				if data:
					cantidad = request.data['cantidad']
					if ',' in cantidad:
						request.data['cantidad'] = cantidad.replace(',','.')					
					data.material_id = int(request.data['material_id'])
					data.cantidad = request.data['cantidad']
					data.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)

				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				# if serializer.is_valid():
					# self.perform_update(serializer)

				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  



class DesgloceManoDeObraSerializer(serializers.HyperlinkedModelSerializer):
	
	manoDeObra = ManoObraSerializer(read_only=True)
	manoDeObra_id = serializers.PrimaryKeyRelatedField(read_only=True)

	unidadConstructiva = UnidadConstructivaSerializer(read_only=True)
	unidadConstructiva_id = serializers.PrimaryKeyRelatedField(read_only=True)

	class Meta:
		model=DesgloceManoDeObra
		fields=('id','rendimiento','manoDeObra', 'manoDeObra_id','unidadConstructiva','unidadConstructiva_id')

class DesgloceManoDeObraViewSet(viewsets.ModelViewSet):
	model = DesgloceManoDeObra
	queryset = model.objects.all()
	serializer_class = DesgloceManoDeObraSerializer
	nombre_modulo='avance_de_obra_grafico2.DesgloceManoDeObra'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			#import pdb; pdb.set_trace()		
			queryset = super(DesgloceManoDeObraViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			uucc_id = self.request.query_params.get('uucc_id', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))
			qset = qset & (Q(unidadConstructiva_id=uucc_id))
			
			if dato:
				qset = qset & (Q(manoDeObra__descripcion__icontains=dato) | Q(manoDeObra__codigo__icontains=dato)) 

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = (Q(manoDeObra_id=request.data['manoDeObra_id']) & Q(unidadConstructiva_id=request.data['uucc_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:				
					rendimiento = request.data['rendimiento']
					if ',' in rendimiento:
						request.data['rendimiento'] = rendimiento.replace(',','.')				
					
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(manoDeObra_id=request.data['manoDeObra_id'], unidadConstructiva_id=request.data['uucc_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un Desglose de Mano de Obra con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)	

			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				data = self.model.objects.get(pk=int(request.data['id']))
				if data:
					rendimiento = request.data['rendimiento']
					if ',' in rendimiento:
						request.data['rendimiento'] = rendimiento.replace(',','.')	

					data.manoDeObra_id = int(request.data['manoDeObra_id'])
					data.rendimiento = request.data['rendimiento']
					data.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':''},status=status.HTTP_201_CREATED)

				# serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				# if serializer.is_valid():
					# self.perform_update(serializer)

				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)					  


class PeriodoProgramacionSerializer(serializers.HyperlinkedModelSerializer):
	cronograma=CronogramaSerializerLite(read_only=True)
	cronograma_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cronograma.objects.all())
	utilizado = serializers.SerializerMethodField()

	class Meta:
		model = PeriodoProgramacion
		fields=('id','cronograma','cronograma_id','fechaDesde','fechaHasta','utilizado')

	def get_utilizado(self, obj):		
		return DetallePeriodoProgramacion.objects.filter(periodoProgramacion_id=obj.id).exists()

class PeriodoProgramacionLiteSerializer(serializers.HyperlinkedModelSerializer):
	class Meta:
		model = PeriodoProgramacion
		fields=('id','fechaDesde','fechaHasta',)

class PeriodoProgramacionViewSet(viewsets.ModelViewSet):
	model = PeriodoProgramacion
	queryset = model.objects.all()
	serializer_class = PeriodoProgramacionSerializer
	nombre_modulo='avanceObraLite.PeriodoProgramacion'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
					
			queryset = super(PeriodoProgramacionViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			cronograma_id = self.request.query_params.get('cronograma_id', None)
			detallePresupuesto_id = self.request.query_params.get('detallePresupuesto_id', None)
			lite = self.request.query_params.get('lite', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))
			
			if dato:
				qset = qset & (Q(cronograma__nombre__icontains=dato)) 

			if detallePresupuesto_id:
				ids= DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=detallePresupuesto_id).values('periodoProgramacion_id')
				qset = qset & (~Q(id__in=ids)) 

			if cronograma_id:
				qset = qset & (Q(cronograma__id=cronograma_id)) 

			queryset = self.model.objects.filter(qset).order_by('fechaDesde')

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			# import pdb; pdb.set_trace()
			if ignorePagination:
				if lite:
					serializer = PeriodoProgramacionLiteSerializer(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['fechaDesde']) | Q(codigo=request.data['fechaHasta'])) & Q(catalogo_id=request.data['cronograma_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(
							cronograma_id = request.data['cronograma_id'],
						)
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un registro con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
				
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		# import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(cronograma_id=request.data['cronograma_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)	

class UnidadConstructivaAOViewSet(viewsets.ModelViewSet):
	model = UnidadConstructiva
	queryset = model.objects.all()
	serializer_class = UnidadConstructivaSerializer
	nombre_modulo='avanceObraLite.UnidadConstructiva'
	model_log=Logs
	model_acciones=Acciones

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},
				status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(UnidadConstructivaAOViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			codigo = self.request.query_params.get('codigo', None)
			catalogo = self.request.query_params.get('catalogo', None)
			tipoUnidadConstructiva = self.request.query_params.get('tipoUnidadConstructiva', None)
			lite = self.request.query_params.get('lite', None)
			ignorePagination= self.request.query_params.get('ignorePagination',None)
			mensaje=''

			qset = (~Q(id=0))

			if dato:
				qset = qset & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))
			if catalogo:
				qset = qset & (Q(catalogo__id=catalogo))
			if tipoUnidadConstructiva:
				qset = qset & (Q(tipoUnidadConstructiva__id=tipoUnidadConstructiva))
			if codigo:
				qset = qset & (Q(codigo=codigo))

			queryset = self.model.objects.filter(qset)

			if queryset.count()==0:
				mensaje='No se encontraron Items con los criterios de busqueda ingresados'

			if ignorePagination:
				if lite:
					serializer = UnidadConstructivaLiteSerializer(queryset,many=True)
				else:
					serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
				'data':serializer.data})				
			else:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = UnidadConstructivaLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':mensaje,'success':'ok',
					'data':serializer.data})
				
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':mensaje,'success':'ok',
					'data':serializer.data})



		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'success':'fail','data':''},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				qsetValidator = ((Q(descripcion=request.data['descripcion']) | Q(codigo=request.data['codigo'])) & Q(catalogo_id=request.data['catalogo_id']))
				validator = self.model.objects.filter(qsetValidator).first()								
				if not validator:
					serializer = self.get_serializer(data=request.data,context={'request': request})

					if serializer.is_valid():
						serializer.save(
							catalogo_id = request.data['catalogo_id'],
							tipoUnidadConstructiva_id = request.data['tipoUnidadConstructiva_id']
						)
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
					else:
						return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)
				else:
					return Response({'message':'Ya existe un registro con la informacion suministrada.','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
				
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		#import pdb; pdb.set_trace()
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = self.get_serializer(instance,data=request.data,context={'request': request},partial=partial)
				if serializer.is_valid():
					self.perform_update(serializer)
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)	
#Permite obtener el APU de una Unidad constructiva
def get_apuUnidadConstructiva(request,idUnidadConstructiva):
	try:
		
		uc = UnidadConstructiva.objects.filter(
			id=idUnidadConstructiva).values(
			'id','codigo','descripcion')

		desgloceMo = DesgloceManoDeObra.objects.filter(
			unidadConstructiva__id=idUnidadConstructiva).values(
			'id','manoDeObra__codigo','manoDeObra__descripcion',
			'manoDeObra__valorHora','rendimiento')
		
		totalMo = 0	
		if desgloceMo:	
			for row in desgloceMo:
				totalMo = float(totalMo) + \
				 (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		##import pdb; pdb.set_trace()
		desgloceMat = DesgloceMaterial.objects.filter(
			unidadConstructiva__id=idUnidadConstructiva).values(
			'id','material__codigo','material__descripcion',
			'material__valorUnitario','cantidad')

		totalMat = 0		 
		if desgloceMat:
			for row in desgloceMat:
				totalMat = float(totalMat) + \
				 (float(row['material__valorUnitario']) * float(row['cantidad']))

		data = {
			'unidadConstructiva' : uc[0],
			'totalManoDeObra' : round(totalMo,2),
			'totalMateriales' : round(totalMat,2),
			'desgloceManoDeObra' : list(desgloceMo),
			'desgloceMateriales' : list(desgloceMat)
		}
		
		return JsonResponse({'message':'','success':'ok','data': data})
	except Exception as e:
		functions.toLog(e,'avanceObraLite.getUnidadConstructiva')
		return Response(
			{'message':'Se presentaron errores de comunicacion con el servidor',
			'success':'fail','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#Permite obtener el listado de UUCC y materiales requeridas en un nodo con un presupuesto
def get_cantidadesDeNodo(request,nodo_id,presupuesto_id):
	unidadesConstructivas = []
	materiales = []
	try:
		
		condicion = Q(nodo__id=nodo_id)
		condicion = condicion & (~Q(cantidad=0))
		queryset = JCantidadesNodo.objects.filter(condicion).values(
			'detallepresupuesto__codigoUC','detallepresupuesto__catalogoUnidadConstructiva__id',
			'cantidad'
			)
		for row in queryset:
			querysetUUCC = UnidadConstructiva.objects.filter(
				catalogo__id=row['detallepresupuesto__catalogoUnidadConstructiva__id'],
				codigo=row['detallepresupuesto__codigoUC']).values('id','descripcion')
			for rowUC in querysetUUCC:
				unidadesConstructivas.append(
					{
						'id':rowUC['id'],
						'codigo':row['detallepresupuesto__codigoUC'],
						'descripcion':rowUC['descripcion'],
						'cantidad':row['cantidad'],
						'catalogo_id': row['detallepresupuesto__catalogoUnidadConstructiva__id']
					}
				)
		
		
		for unidad in unidadesConstructivas:
			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=unidad['id']).values(
				'material__codigo','material__descripcion','cantidad')		
			for mat in desgloceMat:
				materiales.append({
					'codigo': mat['material__codigo'],
					'descripcion': mat['material__descripcion'],
					'cantidad': float(mat['cantidad']) * float(unidad['cantidad'])
					})
		##import pdb; pdb.set_trace()		
		materialesAgrupados=[]
		
		for mat in materiales:
			agregar = True
			for obj in materialesAgrupados:
				if mat['codigo'] == obj['codigo']:
					obj['cantidad']	+= mat['cantidad']
					agregar = False
			if agregar:		
				materialesAgrupados.append(mat)



		data = {
			'unidadesConstructivas': unidadesConstructivas,
			'materiales' : materialesAgrupados
		}
		return JsonResponse({'message':'','success':'ok','data': data})
	except Exception as e:
		functions.toLog(e,'avanceObraLite.getCantidadesEnNodo')
		return Response(
			{'message':'Se presentaron errores de comunicacion con el servidor',
			'success':'fail','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def seguimientoContratual(request,id):
	try:
		seguimiento = []
		porcentajes = []
		ejecutado=0
		programado=0
		desviacion=0
		# import pdb; pdb.set_trace()
		qset = (~Q(actividad__nombre='No aplica')) & (Q(presupuesto__cronograma__id=id))

		actividades = FDetallePresupuesto.objects.filter(qset).values(
			'actividad__id',
			'actividad__nombre',
			'actividad__peso').annotate(total=Sum('cantidad'))

		fecha_actual = datetime.today().date()

		for act in actividades:
			
			ejecutar = DetalleReporteTrabajo.objects.filter(
				detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))

			programar= DetallePeriodoProgramacion.objects.filter(
				periodoProgramacion__fechaDesde__lte=fecha_actual,
				detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))

			if programar['cantidad__sum'] :
				desviacion = programar['cantidad__sum']
			else:
				desviacion = 0

			if ejecutar['cantidad__sum']:
				desviacion = desviacion -ejecutar['cantidad__sum']
			

			if ejecutar['cantidad__sum']:
				porcent_ejecutado = float(ejecutar['cantidad__sum']/act['total'])*act['actividad__peso']
			else:
				porcent_ejecutado = 0

			if programar['cantidad__sum']:
				porcent_program = float(programar['cantidad__sum']/act['total'])*act['actividad__peso']
			else:
				porcent_program = 0

			

			seguimiento.append({
				'actividad_nombre':act['actividad__nombre'],
				'actividad_peso':str(act['actividad__peso'])+'%',
				'actividad_cant':act['total'],

				'obra_ejecutada_acumulada':ejecutar['cantidad__sum'] if ejecutar['cantidad__sum'] else 0,
				'obra_ejecutada_porcentual':str(round(porcent_ejecutado,3))+'%',

				'obra_programada_acumulada':programar['cantidad__sum'] if programar['cantidad__sum'] else 0,
				'obra_programada_porcentual':str(round(porcent_program,3))+'%',

				'obra_desviacion_acumulada':desviacion,
				'obra_desviacion_porcentual':str(round(porcent_ejecutado-porcent_program,3))+'%',

				})

			ejecutado += porcent_ejecutado
			programado += porcent_program


		

		porcentajes.append({	
					'id':0,				
					'nombre': 'Obra ejecutada',
					'porcentaje': round(ejecutado,2),
					'color':'primary',
					'valor': round(ejecutado,2),
				})

		porcentajes.append({	
					'id':1,				
					'nombre': 'Obra programada',
					'porcentaje': round(programado,2),
					'color':'success',
					'valor':round(programado,2),
				})

		porcentajes.append({	
					'id':2,				
					'nombre': '% Desviación',
					'porcentaje': abs(round(ejecutado-programado,2)),
					'color':'warning',
					'valor':round(ejecutado-programado,2),
				})

		datos = {
			'seguimiento' : seguimiento,
			'porcentajes' : porcentajes,
		}

		return JsonResponse({'message':'','success':'ok','data':datos})


	except Exception as e:
		functions.toLog(e,'avanceObraLite.seguimientoContratual')
		return Response({
			'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def graficaCronograma(request,id):
	try:
			
		porcentajes = []
		avanceObra = []
		avanceFinanciero = []	

		qset1 = (Q(presupuesto__cronograma__id=id)& Q(cantidad__gt=0) & (~Q(actividad__nombre='No aplica')))
		qset2 = (Q(detallePresupuesto__presupuesto__cronograma__id=id)& Q(cantidad__gt=0) & (~Q(detallePresupuesto__actividad__nombre='No aplica')))

		queryset_a_ejecutar = FDetallePresupuesto.objects.filter(
			# presupuesto__cronograma__id=id,
			# cantidad__gt=0
			qset1).values(
			'actividad__id',
			'actividad__nombre',
			'actividad__peso').annotate(total=Sum('cantidad'))			

		queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
			# detallePresupuesto__presupuesto__cronograma__id=id,
			# cantidad__gt=0
			qset2).values(
			'detallePresupuesto__actividad__id',
			'detallePresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidad'))

		

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['actividad__id'],
						'actividad': Aejecutar['actividad__nombre'],
						'porcentaje': (float(ejecutado['total']) / float(Aejecutar['total'])) * 100
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['actividad__id'],
					'actividad': Aejecutar['actividad__nombre'],
					'porcentaje': 0
				})

	# #codigo para obtener el avance de obra en el tiempo
		
		# reportesTrabajo = ReporteTrabajo.objects.filter(
		# 	periodoProgramacion__cronograma__id=id
		# 	).order_by('fechaReporte').values('fechaReporte').distinct()

		# for fechaReporteTrabajo in reportesTrabajo:
		# 	avanceFisico = 0
		# 	porcentajesAvanceFisico = []
		# 	queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
		# 		detallePresupuesto__presupuesto__cronograma__id=id,
		# 		cantidad__gt=0,
		# 		reporteTrabajo__fechaReporte__lte=fechaReporteTrabajo['fechaReporte']).values(
		# 		'detallePresupuesto__actividad__id',
		# 		'detallePresupuesto__actividad__nombre'
		# 		).annotate(total=Sum('cantidad'))

		# 	for Aejecutar in queryset_a_ejecutar:
		# 		agregado = False
		# 		for ejecutado in queryset_ejecutada:
		# 			if Aejecutar['detallePresupuesto__actividad__id'] == ejecutado['detallePresupuesto__actividad__id']:
		# 				agregado = True
		# 				porcentajesAvanceFisico.append({
		# 					'id' : Aejecutar['detallePresupuesto__actividad__id'],
		# 					'actividad': Aejecutar['detallePresupuesto__actividad__nombre'],
		# 					'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
		# 					'peso': Aejecutar['detallePresupuesto__actividad__peso']
		# 				})
		# 			if agregado:
		# 				break
		# 		if agregado == False:
		# 			porcentajesAvanceFisico.append({
		# 				'id' : Aejecutar['detallePresupuesto__actividad__id'],
		# 				'actividad': Aejecutar['detallePresupuesto__actividad__nombre'],
		# 				'porcentaje': 0,
		# 				'peso': Aejecutar['detallePresupuesto__actividad__peso']
		# 			})

			
		# 	for p in porcentajesAvanceFisico:
		# 		avanceFisico = avanceFisico + ((float(p['porcentaje']) * float(float(p['peso'])/100)))

		# 	#if cantidadDeProyectos > 0:
		# 			#Codigo para llenar la curva programada:
			
			
		# 	qsetP = DetallePeriodoProgramacion.objects.filter(
		# 		periodoProgramacion__fechaHasta__lte=fechaReporteTrabajo['fechaReporte'],
		# 		detallePresupuesto__presupuesto__cronograma__id=id)

		# 	sumatoria_qsetP = 0
		# 	for item_qsetP in qsetP:
		# 		sumatoria_qsetP += float(item_qsetP.cantidad / item_qsetP.detallePresupuesto.cantidad) * float(item_qsetP.detallePresupuesto.porcentaje)


		# 	if qsetP:	

		# 		avanceObra.append ( {
		# 			'fecha' : fechaReporteTrabajo['fechaReporte'],
		# 			'avance' : round(avanceFisico,2),
		# 			'avance_proyectado': sumatoria_qsetP
		# 			})
		# 	else:
		# 		avanceObra.append ( {
		# 			'fecha' : fechaReporteTrabajo['fechaReporte'],
		# 			'avance' : round(avanceFisico,2),
		# 			'avance_proyectado': 0
		# 			})


		# if avanceObra:		
		# 	fechamax = avanceObra[-1]['fecha']	
		# 	avanceMax = avanceObra[-1]['avance']
		# 	avanceMaxPr = avanceObra[-1]['avance_proyectado']
		# 	qsetP = DetallePeriodoProgramacion.objects.filter(
		# 			periodoProgramacion__fechaHasta__gte=fechamax,
		# 			detallePresupuesto__presupuesto__cronograma__id=id).order_by('periodoProgramacion__fechaHasta')

		# 	sumatoria_qsetP = 0
		# 	for item_qsetP in qsetP:
		# 		sumatoria_qsetP += float(item_qsetP.cantidad / item_qsetP.detallePresupuesto.cantidad) * float(item_qsetP.detallePresupuesto.porcentaje)

		# 	proy = avanceMaxPr
		# 	for rowD in qsetP:
		# 		proy = proy + sumatoria_qsetP
		# 		avanceObra.append({
		# 			'fecha': rowD.periodoProgramacion.fechaHasta,
		# 			'avance':round(avanceMax,2),
		# 			'avance_proyectado': proy
		# 			})
		# else:
		# 	qsetP = DetallePeriodoProgramacion.objects.filter(
		# 			detallePresupuesto__presupuesto__cronograma__id=id).order_by('periodoProgramacion__fechaHasta')

		# 	sumatoria_qsetP = 0
		# 	for item_qsetP in qsetP:
		# 		sumatoria_qsetP += float(item_qsetP.cantidad / item_qsetP.detallePresupuesto.cantidad) * float(item_qsetP.detallePresupuesto.porcentaje)

		# 	proy = 0
		# 	for rowD in qsetP:
		# 		proy = proy + sumatoria_qsetP
		# 		avanceObra.append({
		# 			'fecha': rowD.periodoProgramacion.fechaHasta,
		# 			'avance':0,
		# 			'avance_proyectado': proy
		# 			})
			
			
				


		#codigo para obtener el avance financiero en el tiempo
		
		avanceF=0
		totalEjecutado = 0

		reportesTrabajo = DetalleReporteTrabajo.objects.filter(
			reporteTrabajo__periodoProgramacion__cronograma__id=id,cantidad__gt=0).values(
			'reporteTrabajo__id',
			'reporteTrabajo__fechaReporte').order_by('reporteTrabajo__fechaReporte').distinct()

		detalles_r = DetalleReporteTrabajo.objects.filter(reporteTrabajo__periodoProgramacion__cronograma__id=id,cantidad__gt=0)


		presupuesto = EPresupuesto.objects.filter(cronograma_id=id,cerrar_presupuesto=True)
		aiu = 1
		if presupuesto:
			aiu = presupuesto.first().aiu
			presupuesto = presupuesto.first().suma_presupuesto()			
		else:
			presupuesto = 0

		for fechaReporteTrabajo in reportesTrabajo:
			detalles_reporte = detalles_r.filter(reporteTrabajo__id=int(fechaReporteTrabajo['reporteTrabajo__id'])).values(
				'cantidad','detallePresupuesto__valorGlobal')
			# .aggregate(Sum('cantidad'),Sum('detallePresupuesto__valorGlobal'))

			sumatoria_pesos = 0
			sumatoria_totalejecutado = 0

			# subtotal_aux = float(detalles_reporte['cantidad__sum'])*float(detalles_reporte['detallePresupuesto__valorGlobal__sum'])
			# print(subtotal_aux)
			# sumatoria_totalejecutado+= subtotal_aux 
			# sumatoria_pesos += float(subtotal_aux*aiu)/float(presupuesto*aiu)

			for item in detalles_reporte:				
				# sumatoria_totalejecutado += item.subtotal()
				# sumatoria_pesos += item.peso()

				subtotal_aux = float(item['cantidad'])*float(item['detallePresupuesto__valorGlobal'])
				sumatoria_totalejecutado += subtotal_aux

				peso_aux = (float(subtotal_aux)*float(aiu))/float(presupuesto*aiu)

				sumatoria_pesos += peso_aux

			# import pdb; pdb.set_trace()
			#if cantidadDeProyectos > 0:
			avanceF += round(sumatoria_pesos*100,2)
			totalEjecutado += sumatoria_totalejecutado
			#else:
			#	avanceF = 0

			avanceFinanciero.append({
				'fecha': fechaReporteTrabajo['reporteTrabajo__fechaReporte'],
				'avance': avanceF,
				'monto': round(totalEjecutado), #"$ {:,.2f}".format(round(totalEjecutado,2))
				# 'monto_neto': round(totalEjecutado*aiu)
			})



		

		# avanceObra_agrupado = []
		# avanceObra_agrupado.append({
		# 			'fecha': avanceObra[0]['fecha'],
		# 			'avance':avanceObra[0]['avance'],
		# 			'avance_proyectado': avanceObra[0]['avance_proyectado']
		# 			})
		# i=1
		# # import pdb; pdb.set_trace()
		# for item in avanceObra:				
		# 	if item['fecha']==avanceObra_agrupado[i-1]['fecha']:
		# 		avanceObra_agrupado[i-1]['avance']+=item['avance']
		# 		avanceObra_agrupado[i-1]['avance_proyectado']+=item['avance_proyectado']
		# 	else:
		# 		avanceObra_agrupado.append({
		# 			'fecha': item['fecha'],
		# 			'avance':item['avance'],
		# 			'avance_proyectado': item['avance_proyectado']
		# 			})
		# 		i+=1



		datos = {
			#'avanceObra' : avance,
			'porHito' : porcentajes,
			'curvaAvanceObra' : [],
			'curvaAvanceFinanciero' : avanceFinanciero,
			'sumaPresupuesto': presupuesto*(aiu+0.0245),
			'sumaPresupuestoDirecto':presupuesto,
			#'obrasPorEstado' : obrasPorEstado
		}


		return JsonResponse({'message':'','success':'ok','data':datos})					

	except Exception as e:
		functions.toLog(e,'avanceObraLite.graficacronograma')
		return Response({
			'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def graficasTablero(request,id):
	try:
		cantidadDeProyectos = 0
		porcentajes = []
		avanceObra = []
		avanceFinanciero = []
		obrasPorEstado = []
		#codigo para obtener el avance por hito
		avance = 0
		porcentajes = []
		queryset_a_ejecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
			cantidad__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre',
			'detallepresupuesto__actividad__peso').annotate(total=Sum('cantidad'))	

		queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__actividad__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(total=Sum('cantidadEjecutada'))

		for Aejecutar in queryset_a_ejecutar:
			agregado = False
			for ejecutado in queryset_ejecutada:
				if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
					agregado = True
					porcentajes.append({
						'id' : Aejecutar['detallepresupuesto__actividad__id'],
						'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
						'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
						'peso': Aejecutar['detallepresupuesto__actividad__peso']
					})
				if agregado:
					break
			if agregado == False:
				porcentajes.append({
					'id' : Aejecutar['detallepresupuesto__actividad__id'],
					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
					'porcentaje': 0,
					'peso': Aejecutar['detallepresupuesto__actividad__peso']
				})
		
		for p in porcentajes:
			avance = avance + (float(p['porcentaje']) * float(float(p['peso'])/100))

		##import pdb; pdb.set_trace()
		cantidadDeProyectos = Proyecto.objects.filter(
			mcontrato__id=id).count()	
		if cantidadDeProyectos > 0:
			avance = round(avance,2)
		else:
			avance = 0

		#codigo para obtener el avance de obra en el tiempo
		
		# reportesTrabajo = ReporteTrabajo.objects.filter(
		# 	presupuesto__cronograma__proyecto__mcontrato__id=id
		# 	).order_by('fechaTrabajo').values('fechaTrabajo').distinct()

		# for fechaReporteTrabajo in reportesTrabajo:
		# 	avanceFisico = 0
		# 	porcentajesAvanceFisico = []
		# 	queryset_ejecutada = DetalleReporteTrabajo.objects.filter(
		# 		detallepresupuesto__presupuesto__cronograma__proyecto__mcontrato__id=id,
		# 		cantidadEjecutada__gt=0,
		# 		reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo']).values(
		# 		'detallepresupuesto__actividad__id',
		# 		'detallepresupuesto__actividad__nombre'
		# 		).annotate(total=Sum('cantidadEjecutada'))

		# 	for Aejecutar in queryset_a_ejecutar:
		# 		agregado = False
		# 		for ejecutado in queryset_ejecutada:
		# 			if Aejecutar['detallepresupuesto__actividad__id'] == ejecutado['detallepresupuesto__actividad__id']:
		# 				agregado = True
		# 				porcentajesAvanceFisico.append({
		# 					'id' : Aejecutar['detallepresupuesto__actividad__id'],
		# 					'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
		# 					'porcentaje': round((float(ejecutado['total']) / float(Aejecutar['total'])) * 100,2),
		# 					'peso': Aejecutar['detallepresupuesto__actividad__peso']
		# 				})
		# 			if agregado:
		# 				break
		# 		if agregado == False:
		# 			porcentajesAvanceFisico.append({
		# 				'id' : Aejecutar['detallepresupuesto__actividad__id'],
		# 				'actividad': Aejecutar['detallepresupuesto__actividad__nombre'],
		# 				'porcentaje': 0,
		# 				'peso': Aejecutar['detallepresupuesto__actividad__peso']
		# 			})
			
		# 	for p in porcentajesAvanceFisico:
		# 		avanceFisico = avanceFisico + ((float(p['porcentaje']) * float(float(p['peso'])/100)))

		# 	if cantidadDeProyectos > 0:

		# 		avanceObra.append ( {
		# 			'fecha' : fechaReporteTrabajo['fechaTrabajo'],
		# 			'avance' : round(avanceFisico,2)
		# 			})

		# #codigo para obtener el avance financiero en el tiempo
		# totalPresupuesto = 0
		# avanceF=0
		# cantidadesEjecutar = FDetallePresupuesto.objects.filter(
		# 	presupuesto__cronograma__proyecto__mcontrato__id=id).values(
		# 	'cantidad','codigoUC','cantidad',
		# 	'catalogoUnidadConstructiva__id', 'id')			
		# for rowCantidadesEjecutar in cantidadesEjecutar:
		# 	totalMo = 0
		# 	totalMat = 0
		# 	queryset = DesgloceManoDeObra.objects.filter(
		# 		unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
		# 		unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']
		# 		).values(
		# 		'manoDeObra__valorHora','rendimiento')
			
		# 	if queryset:
		# 		for row in queryset:
		# 			totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		# 	queryset = DesgloceMaterial.objects.filter(
		# 		unidadConstructiva__codigo=rowCantidadesEjecutar['codigoUC'],
		# 		unidadConstructiva__catalogo__id=rowCantidadesEjecutar['catalogoUnidadConstructiva__id']).values(
		# 		'material__valorUnitario','cantidad')
			
		# 	if queryset:
		# 		for row in queryset:
		# 			totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))

		# 	totalPresupuesto = totalPresupuesto + \
		# 	((totalMo * float(rowCantidadesEjecutar['cantidad'])) + \
		# 	(totalMat * float(rowCantidadesEjecutar['cantidad'])))

		# ##import pdb; pdb.set_trace()	
		# for fechaReporteTrabajo in reportesTrabajo:
		# 	totalEjecutado = 0
		# 	avanceF=0
		# 	cantidadesEjecutadas = DetalleReporteTrabajo.objects.filter(
		# 		cantidadEjecutada__gt=0,
		# 		reporte_trabajo__fechaTrabajo__lte=fechaReporteTrabajo['fechaTrabajo'],
		# 		#detallepresupuesto__id=rowCantidadesEjecutar['id']
		# 		).values('detallepresupuesto__id',
		# 		'detallepresupuesto__codigoUC',
		# 		'detallepresupuesto__catalogoUnidadConstructiva__id').annotate(
		# 		total=Sum('cantidadEjecutada'))

				
		# 	if cantidadesEjecutadas:
		# 		for cantidadEjecutada in cantidadesEjecutadas:
		# 			totalMo = 0
		# 			totalMat = 0

		# 			queryset = DesgloceManoDeObra.objects.filter(
		# 				unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
		# 				unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']
		# 				).values(
		# 				'manoDeObra__valorHora','rendimiento')
					
		# 			if queryset:
		# 				for row in queryset:
		# 					totalMo = float(totalMo) + (float(row['manoDeObra__valorHora']) * float(row['rendimiento']))

		# 			queryset = DesgloceMaterial.objects.filter(
		# 				unidadConstructiva__codigo=cantidadEjecutada['detallepresupuesto__codigoUC'],
		# 				unidadConstructiva__catalogo__id=cantidadEjecutada['detallepresupuesto__catalogoUnidadConstructiva__id']).values(
		# 				'material__valorUnitario','cantidad')
					
		# 			if queryset:
		# 				for row in queryset:
		# 					totalMat = float(totalMat) + (float(row['material__valorUnitario']) * float(row['cantidad']))


		# 			totalEjecutado = totalEjecutado + \
		# 			(float(cantidadEjecutada['total']) * totalMo) + \
		# 			(float(cantidadEjecutada['total']) * totalMat)

		# 	if totalPresupuesto > 0:
		# 		avanceF = avanceF + round((totalEjecutado / totalPresupuesto) * 100,2)		

		# 	if cantidadDeProyectos > 0:
		# 		avanceF = round(avanceF,2)
		# 	else:
		# 		avanceF = 0

		# 	avanceFinanciero.append({
		# 		'fecha': fechaReporteTrabajo['fechaTrabajo'],
		# 		'avance': avanceF,
		# 		'monto': round(totalEjecutado) #"$ {:,.2f}".format(round(totalEjecutado,2))
		# 	})


		datos = {
			'avanceObra' : avance,
			'porHito' : porcentajes,
			# 'curvaAvanceObra' : avanceObra,
			# 'curvaAvanceFinanciero' : avanceFinanciero,
			# 'obrasPorEstado' : obrasPorEstado
		}

		return JsonResponse({'message':'','success':'ok','data':datos})
	except Exception as e:
		return Response({
			'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},
			status=status.HTTP_500_INTERNAL_SERVER_ERROR)


api_view(['GET',])
def cantidadesAliquidar(request, id):
	try:
		#identificar cantidades a ejecutar:
		datos = []
		cantidadesEjecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__id=id,
			cantidad__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(
			total=Sum('cantidad'))

		cantidadesEjecutadas = DetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = id,
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			#'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))


		# cantidadesLiquidadas = MLiquidacionUUCC.objects.filter(
		# 	presupuesto__id=id).values(
		# 	'detallereportetrabajo__detallepresupuesto__id',
		# 	'detallereportetrabajo__detallepresupuesto__codigoUC',
		# 	'detallereportetrabajo__detallepresupuesto__descripcionUC',
		# 	'detallereportetrabajo__detallepresupuesto__catalogoUnidadConstructiva__id'
		# 	).annotate(total = Sum('detallereportetrabajo__cantidadEjecutada'))	

		##import pdb; pdb.set_trace()	
		for cantidadEjecutar in cantidadesEjecutar:
			cejecutar = 0
			cliquidada = 0
			liquidada = False
			for cantidadEjecutada in cantidadesEjecutadas:
				if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
					cejecutar = cantidadEjecutada['total']
					#if cantidadEjecutada['detallepresupuesto__liquidacion__id']:
					#liquidada = cantidadEjecutada['liquidada']
					break

			# for cantidadLiquidada in cantidadesLiquidadas:
			# 	if cantidadLiquidada['detallereportetrabajo__detallepresupuesto__codigoUC'] == \
			# 	cantidadEjecutar['detallepresupuesto__codigoUC']:
			# 		cliquidada = cantidadLiquidada['total']
			# 		break

			# liquidable = True

			# if cantidadEjecutar['total'] == cejecutar and cantidadEjecutar['total'] == cliquidada:
			# 	liquidable = False

			datos.append(
				{
					'hito': cantidadEjecutar['detallepresupuesto__actividad__nombre'],
					'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
					'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
					'ejecutar': cantidadEjecutar['total'],
					'ejecutado': cejecutar,
					'pendiente': int(cantidadEjecutar['total']) - int(cejecutar)
					#'liquidado' : cliquidada,
					#'liquidada': liquidada
				}
			)
				

			
		return JsonResponse({'message':'','success':'ok','data':datos})	
	except Exception as e:
		functions.toLog(e,'avanceObraLite.cantidadesAliquidar')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def seguimientoCantidades(request,id_presupuesto):
	presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
				
	return render(request, 'avanceObraLite/seguimientocantidades.html',
		{'model':'jreportetrabajo',
		'app':'avanceObraLite',
		'presupuesto':presupuesto})


#@transaction.atomic
api_view(['GET',])
def consultar_uucc_ejecutados(request):	
	if request.method == 'GET':			
		try:
			###import pdb; pdb.set_trace()
			presupuesto_id=request.GET['presupuesto_id']
			aux= MLiquidacionUUCC.objects.filter(presupuesto__id=int(request.GET['presupuesto_id']))
			list_excluir=[]

			for a in aux:				
				list_aux=a.detallereportetrabajo.through.objects.filter().values('kdetallereportetrabajo_id')
				for b in list_aux:
					list_excluir.append(b['kdetallereportetrabajo_id'])

			
			qset = (Q(detallepresupuesto__presupuesto__id=int(presupuesto_id))) & (~Q(id__in=list_excluir))
			reporte=DetalleReporteTrabajo.objects.filter(qset)

			list_return=[]

			for x in reporte:
				list_return.append({
					'id':x.id,
					'codigo':x.detallepresupuesto.codigoUC,
					'descripcion':x.detallepresupuesto.descripcionUC,
					'nodo':x.nodo.nombre,
					})

			###import pdb; pdb.set_trace()
			return JsonResponse({'message':'','success':'ok','data':list_return})

		except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
def anularReporteLiquidacion(request):
	
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			model = MLiquidacionUUCC.objects.get(pk=int(respuesta['liquidacion_id']))
			model.estado=3
			model.motivo_anulacion=respuesta['motivo_anular']
			model.save()

			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraLite.AnularReporteLiquidacion'
									,id_manipulado=model.id)
			logs_model.save()


			list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=model.id)
			aux = model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=model.id).values('kdetallereportetrabajo_id')

			##import pdb; pdb.set_trace()
			for y in aux:
				model_detallereporte=DetalleReporteTrabajo.objects.get(pk=y['kdetallereportetrabajo_id'])
				model_detallereporte.liquidada=False
				model_detallereporte.save()

				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraLite.DetalleReporteTrabajo'
									,id_manipulado=model_detallereporte.id)
				logs_model.save()


			for x in list_detallereporte:				
				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_borrar
									,nombre_modelo='avanceObraLite.AnularReporteLiquidacion'
									,id_manipulado=x.id)				
				model.detallereportetrabajo.remove(x.kdetallereportetrabajo_id)

				logs_model.save()

		

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'La liquidacion se ha anulado correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

def cerrar_liquidacionuucc(request):
	###import pdb; pdb.set_trace()
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			model = MLiquidacionUUCC.objects.get(pk=int(respuesta['liquidacion_id']))
			model.estado=2
			model.save()

			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraLite.liquidacionuucc'
									,id_manipulado=model.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'La liquidacion se ha cerrado correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



def guardar_liquidacionuucc(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			##import pdb; pdb.set_trace()
			lista=request.POST['_content']
			respuesta= json.loads(lista)

			liquidacionuucc=MLiquidacionUUCC(presupuesto_id=int(respuesta['presupuesto_id']))
			liquidacionuucc.save()


			logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_crear
									,nombre_modelo='avanceObraLite.liquidacionuucc'
									,id_manipulado=liquidacionuucc.id)
			logs_model.save()
			transaction.savepoint_commit(sid)

			list_detallepresupuesto=respuesta['lista']
			list_aux=[]
			for d in list_detallepresupuesto:
				list_detallereporte = DetalleReporteTrabajo.objects.filter(liquidada=False,detallepresupuesto__id=d['id']).values('id')
				for e in list_detallereporte:
					list_aux.append(e['id'])
			##import pdb; pdb.set_trace()

			list_temporal=[]
			for x in list_aux:
				liquidacionuucc.detallereportetrabajo.add(x)
				list_temporal.append(x)

				model_detallert=DetalleReporteTrabajo.objects.get(pk=x)
				model_detallert.liquidada=True
				model_detallert.save()

				logs_model=Logs(usuario_id=request.user.usuario.id
									,accion=Acciones.accion_actualizar
									,nombre_modelo='avanceObraLite.detallereportetrabajo(liquidacion)'
									,id_manipulado=x)
				logs_model.save()

			
			list_liq_detalle = liquidacionuucc.detallereportetrabajo.through.objects.filter(kdetallereportetrabajo_id__in=list_temporal)

			insert_list = []			
			for a in list_liq_detalle:
				insert_list.append(Logs(usuario_id=request.user.usuario.id
										,accion=Acciones.accion_crear
										,nombre_modelo='avanceObraLite.liquidacionuucc.detallereportetrabajo'
										,id_manipulado=a.id)
										)
			Logs.objects.bulk_create(insert_list)

			transaction.savepoint_commit(sid)
			

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha guardando correctamente','success':'ok',
						'data':''})

		except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	

# Exporta a excel contrato 
def exportReporteLiquidacion(request):
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte_liquidacion.xls"'

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Liquidacion de UUCC')

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
		format1.set_align('center')
		format1.set_align('vcenter')
		format2=workbook.add_format({'border':0})
		format3=workbook.add_format({'border':0,'font_size':12})
		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')

		worksheet.set_column('A:A', 15)
		worksheet.set_column('B:B', 35)
		worksheet.set_column('C:H', 20)

		row=1
		col=0

		data_return=[]

		model=MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))

		list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=int(request.GET['liquidacion_id'])).values('kdetallereportetrabajo__detallepresupuesto__id').annotate(cantidad_ejecutada=Sum('kdetallereportetrabajo__cantidadEjecutada')).distinct()

		
		subtotal_manoobra_acumulado=0
		subtotal_material_acumulado=0
		total_acumulado = 0

		for x in list_detallereporte:
			model_detalle_presupuesto=FDetallePresupuesto.objects.get(pk=int(x['kdetallereportetrabajo__detallepresupuesto__id']))

			###import pdb; pdb.set_trace()

			subtotal_manoobra = float(model_detalle_presupuesto.valorManoObra*x['cantidad_ejecutada'])
			subtotal_material = float(model_detalle_presupuesto.valorMaterial*x['cantidad_ejecutada'])
			total = float(subtotal_manoobra+subtotal_material)

			subtotal_manoobra_acumulado+=subtotal_manoobra
			subtotal_material_acumulado+=subtotal_material
			total_acumulado = total_acumulado + total

			data_return.append({
				'id':model_detalle_presupuesto.id,
				'codigoUC':model_detalle_presupuesto.codigoUC,
				'descripcionUC':model_detalle_presupuesto.descripcionUC,
				'cantidad_ejecutada':float(x['cantidad_ejecutada']),
				'valorManoObra':model_detalle_presupuesto.valorManoObra,
				'valorMaterial':model_detalle_presupuesto.valorMaterial,
				'Subtotal_mano_obra':float(subtotal_manoobra),
				'Subtotal_materiales':float(subtotal_material),
				'Total':total,
				})

		data_return.append({
			'id':' ',
			'codigoUC':' ',
			'descripcionUC':' ',
			'cantidad_ejecutada':' ',
			'valorManoObra':' ',
			'valorMaterial':' ',
			'Subtotal_mano_obra':float(subtotal_manoobra_acumulado),
			'Subtotal_materiales':float(subtotal_material_acumulado),
			'Total':float(total_acumulado),
			})

		if data_return:			
			worksheet.write('A1', 'Codigo UUCC', format1)
			worksheet.write('B1', 'Descripcion UUCC', format1)
			worksheet.write('C1', 'Cantidad ejecutada', format1)
			worksheet.write('D1', 'Vlr Mano obra', format1)
			worksheet.write('E1', 'Vlr Material', format1)
			worksheet.write('F1', 'Subtotal mano obra', format1)
			worksheet.write('G1', 'Subtotal materiales', format1)
			worksheet.write('H1', 'Valor Liquidacion', format1)

			##import pdb; pdb.set_trace()
			for x in data_return:
				worksheet.write(row, col,x['codigoUC'],format2)
				worksheet.write(row, col+1,x['descripcionUC'],format2)
				worksheet.write(row, col+2,x['cantidad_ejecutada'],format2)
				worksheet.write(row, col+3,x['valorManoObra'],format2)
				worksheet.write(row, col+4,x['valorMaterial'],format2)
				worksheet.write(row, col+5,x['Subtotal_mano_obra'],format2)
				worksheet.write(row, col+6,x['Subtotal_materiales'],format2)
				worksheet.write(row, col+7,x['Total'],format2)
				row +=1
			workbook.close()
			return response


	except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	


api_view(['GET',])
def consultar_detallereporte_liquidacion(request):
	if request.method == 'GET':
		try:
			##import pdb; pdb.set_trace()
			data_return=[]
			model=MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))

			list_detallereporte=model.detallereportetrabajo.through.objects.filter(mliquidacionuucc_id=int(request.GET['liquidacion_id'])).values('kdetallereportetrabajo__detallepresupuesto__id').annotate(cantidad_ejecutada=Sum('kdetallereportetrabajo__cantidadEjecutada')).distinct()

			###import pdb; pdb.set_trace()
			subtotal_manoobra_acumulado=0
			subtotal_material_acumulado=0
			for x in list_detallereporte:
				model_detalle_presupuesto=FDetallePresupuesto.objects.get(pk=int(x['kdetallereportetrabajo__detallepresupuesto__id']))

				###import pdb; pdb.set_trace()
				subtotal_manoobra = float(model_detalle_presupuesto.valorManoObra*x['cantidad_ejecutada'])
				subtotal_material = float(model_detalle_presupuesto.valorMaterial*x['cantidad_ejecutada'])

				subtotal_manoobra_acumulado+=subtotal_manoobra
				subtotal_material_acumulado+=subtotal_material
			
				data_return.append({
					'id':model_detalle_presupuesto.id,
					'codigoUC':model_detalle_presupuesto.codigoUC,
					'descripcionUC':model_detalle_presupuesto.descripcionUC,
					'cantidad_ejecutada':float(x['cantidad_ejecutada']),
					'valorManoObra':model_detalle_presupuesto.valorManoObra,
					'valorMaterial':model_detalle_presupuesto.valorMaterial,
					'Subtotal_mano_obra':float(subtotal_manoobra),
					'Subtotal_materiales':float(subtotal_material),
					})

			data_return.append({
				'id':' ',
				'codigoUC':' ',
				'descripcionUC':' ',
				'cantidad_ejecutada':' ',
				'valorManoObra':' ',
				'valorMaterial':' ',
				'Subtotal_mano_obra':float(subtotal_manoobra_acumulado),
				'Subtotal_materiales':float(subtotal_material_acumulado),
				})
		
			return JsonResponse({'message':'','success':'ok','data':data_return})
		except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})	

api_view(['GET',])
def cantidadesAliquidarlite(request):
	try:
		#identificar cantidades a ejecutar:
		datos = []
		##import pdb; pdb.set_trace()
		# cantidadesEjecutar = JCantidadesNodo.objects.filter(
		# 	detallepresupuesto__presupuesto__id=int(request.GET['presupuesto_id']),
		# 	cantidad__gt=0).values(
		# 	'detallepresupuesto__id',
		# 	'detallepresupuesto__codigoUC',
		# 	'detallepresupuesto__descripcionUC',
		# 	'detallepresupuesto__catalogoUnidadConstructiva__id'
		# 	).annotate(
		# 	total=Sum('cantidad'))

		cantidadesEjecutadas = DetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = int(request.GET['presupuesto_id']),
			cantidadEjecutada__gt=0, liquidada=False).values(
			'detallepresupuesto__id',
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))



		# 	
		# for cantidadEjecutar in cantidadesEjecutar:
		# 	cejecutar = 0
		# 	cliquidada = 0
		# 	liquidada = False
		# 	for cantidadEjecutada in cantidadesEjecutadas:
		# 		if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
		# 			cejecutar = cantidadEjecutada['total']
		# 			#if cantidadEjecutada['detallepresupuesto__liquidacion__id']:
		# 			liquidada = cantidadEjecutada['liquidada']
		# 			break


		# 	datos.append(
		# 		{	'id': cantidadEjecutar['detallepresupuesto__id'],
		# 			'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
		# 			'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
		# 			'ejecutar': float(cantidadEjecutar['total']),
		# 			'ejecutado': float(cejecutar),
		# 			#'liquidado' : cliquidada,
		# 			'liquidada': liquidada
		# 		}
		# 	)

		return JsonResponse({'message':'','success':'ok','data':list(cantidadesEjecutadas)})
	except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
api_view(['GET',])
def consultarmotivoanulacion(request):
	try:
		##import pdb; pdb.set_trace()
		model = MLiquidacionUUCC.objects.get(pk=int(request.GET['liquidacion_id']))
		data=[]
		data.append({
			'id':model.id,
			'motivo_anulacion':model.motivo_anulacion
			})
		return JsonResponse({'message':'','success':'ok','data':data})
	except Exception as e:
			functions.toLog(e,'avanceObraLite.liquidacionuucc')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})

def exportCronograma(request):	
	
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte_cronogramas.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Cronogramas')
		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:G', 25)
		worksheet.set_column('H:AQ', 20)

		row=1
		col=0

		dato = request.GET['dato'] if request.GET['dato'] else None
		mcontrato = request.GET['macrocontrato_id'] if request.GET['macrocontrato_id'] else None
		departamento = request.GET['departamento_id'] if request.GET['departamento_id'] else None
		municipio = request.GET['municipio_id'] if request.GET['municipio_id'] else None
		contratista = request.GET['contratista_id'] if request.GET['contratista_id'] else None
		empresa = request.user.usuario.empresa.id

		

		qset=(~Q(id=0))
		if dato:
			qset = qset & ( Q(empresa__nombre__icontains = dato) |
						Q(empresa__nit__icontains = dato) |
						Q(proyecto__nombre__icontains = dato))

		if empresa and int(empresa)>0:
			qset = qset & (Q(empresa__id = empresa))

		if mcontrato and int(mcontrato)>0:
			qset = qset & (Q(proyecto__mcontrato__id = mcontrato))
			
		if departamento and int(departamento)>0:
			qset = qset & (Q(proyecto__municipio__departamento__id = departamento))
			
		if municipio and int(municipio)>0:
			qset = qset & (Q(proyecto__municipio__id = municipio))
			

		if contratista and int(contratista)>0:
			qset = qset & (Q(proyecto__contrato__contratista__id = contratista))

		queryset = Proyecto_empresas.objects.filter(qset)

		serializer_context = {
			'request': request
		}

		serializer = ProyectoEmpresaLite4Serializer(queryset,many=True,context=serializer_context)

		if serializer:
			worksheet.write('A1', 'Macrocontrato', format1)
			worksheet.write('B1', 'Departamento', format1)
			worksheet.write('C1', 'Municipio', format1)
			worksheet.write('D1', 'Proyecto', format1)
			worksheet.write('E1', '% Avance de obra', format1)
			worksheet.write('F1', '% Avance financiero', format1)
			worksheet.write('G1', 'Fecha corte', format1)

			
			

			for item in serializer.data:
				#import pdb; pdb.set_trace()
				worksheet.write(row, col,item['proyecto']['mcontrato']['nombre'],format3)
				worksheet.write(row, col+1,item['proyecto']['municipio']['departamento']['nombre'],format3)
				worksheet.write(row, col+2,item['proyecto']['municipio']['nombre'],format3)
				worksheet.write(row, col+3,item['proyecto']['nombre'],format3)
				porcentajeAvance = item['porcentajeAvance']['avance']
				porcentajeAvanceFinanciero = item['porcentajeAvanceFinanciero']
				worksheet.write(row, col+4,str(porcentajeAvance)+'%',format2)
				worksheet.write(row, col+5,str(porcentajeAvanceFinanciero)+'%',format2)
				worksheet.write(row, col+6,item['porcentajeAvance']['corte'],format5)
				n_col=7
				
				if mcontrato and int(mcontrato)>0:
					qset = (Q(esquema__macrocontrato__id=int(mcontrato)))&(~Q(padre=0))&(~Q(nombre='Replanteo'))&(~Q(nombre='replanteo'))&(~Q(nombre='Liquidacion'))&(~Q(nombre='liquidacion'))
					capitulos_model= CEsquemaCapitulosActividadesG.objects.filter(qset).exists()
					if capitulos_model:
						capitulos_model= CEsquemaCapitulosActividadesG.objects.filter(qset)
						
						for acts in capitulos_model:
							worksheet.write(0, n_col,acts.nombre,format1)
							if item['porcentajeAvance']['actividades']:
								#import pdb; pdb.set_trace()

								# act_model = item['porcentajeAvance']['actividades']

								# act_obj = act_model.objects.filter(id=int(acts.id)).exists()

								# if act_obj:
								# 	act_obj = act_model.objects.filter(id=int(acts.id))
								# 	worksheet.write(row, n_col,str( round(act_obj.porcentaje,2))+'%',format2)

								for act in item['porcentajeAvance']['actividades']:
									if int(acts.id)==int(act['id']):																						
										worksheet.write(row, n_col,str( round(act['porcentaje'],2))+'%',format2)			
							else:
								worksheet.write(row, n_col,'0%',format2)

							n_col+=1
					

				row +=1

			

		workbook.close()
		return response
	except Exception as e:
		#print(e)
		functions.toLog(e,'avanceObraLite')
		return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def seguimientoMateriales(request,id_presupuesto):
	#import pdb; pdb.set_trace()
	nombre_presupuesto=EPresupuesto.objects.get(pk=id_presupuesto)
	#usuarios=Usuario.objects.filter(user__is_active=True).order_by('persona__nombres')
	return render(request, 'avanceObraLite/seguimientomateriales.html',{
		'model':'jreportetrabajo',
		'app':'avanceObraLite',		
		'cronograma':nombre_presupuesto.cronograma,
		'cronograma_id':nombre_presupuesto.cronograma.id,
		'proyecto_id':nombre_presupuesto.cronograma.proyecto.id,
		'presupuesto_id':id_presupuesto,
		'presupuesto':nombre_presupuesto})

@login_required
def materialesAliquidarlite(request, id):
	try:		
			
		listado=FDetallePresupuesto.objects.filter(presupuesto__id=id, cantidad__gt=0)
		materiales = []

		
		for item in listado:		

			cantidad_ejecutada=DetalleReporteTrabajo.objects.filter(	
				detallepresupuesto_id=item.id
				).aggregate(sum_cantidad_ejecutada=Sum('cantidadEjecutada'))


			cantidad_ejecutada['sum_cantidad_ejecutada']=0 if cantidad_ejecutada['sum_cantidad_ejecutada']==None else cantidad_ejecutada['sum_cantidad_ejecutada']
			
			
			#Extraer los materiales:
			uc = UnidadConstructiva.objects.filter(
				codigo=item.codigoUC,
				catalogo__id=item.catalogoUnidadConstructiva.id).values('id')

			#import pdb; pdb.set_trace()

			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=uc[0]['id']).values(
				'material__codigo','material__descripcion', 'cantidad')


			for obj in desgloceMat:
				materiales.append({
					'codigo' : obj['material__codigo'],
					'hito': item.actividad.nombre,
					'descripcion' : obj['material__descripcion'],
					'cantidad' : float(obj['cantidad']) * float(item.cantidad),
					'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['sum_cantidad_ejecutada']),
					'pendiente': 0
				})
			#unificar cantidades de materialales:
			materialesAgrupados=[]
			
			for mat in materiales:
				agregar = True
				for obj in materialesAgrupados:
					if mat['codigo'] == obj['codigo'] and mat['hito'] == obj['hito']:
						obj['cantidad']	+= mat['cantidad']
						obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
						agregar = False
				if agregar:		
					materialesAgrupados.append(mat)				

			for obj in materialesAgrupados:
				obj['cantidad'] = round(obj['cantidad'],2)
				obj['cantidad_ejecutada'] = round(obj['cantidad_ejecutada'],2)
				obj['pendiente'] = round(obj['cantidad'] - obj['cantidad_ejecutada'],2)
		

		if listado:
			datos=materialesAgrupados
			
		else:
			datos=[]

		
		#import pdb; pdb.set_trace()
		return JsonResponse({'message':'','success':'ok','data':datos})
	except Exception as e:
			functions.toLog(e,'avanceObraLite.cantidadesAliquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
@login_required
def exportarMaterialesaLiquidar(request):
	try:
		presupuesto_id= request.GET['presupuesto_id'] if request.GET['presupuesto_id'] else None
		listado=FDetallePresupuesto.objects.filter(presupuesto__id=int(presupuesto_id), cantidad__gt=0)
		materiales = []		
		for item in listado:		

			cantidad_ejecutada=DetalleReporteTrabajo.objects.filter(	
				detallepresupuesto_id=item.id
				).aggregate(sum_cantidad_ejecutada=Sum('cantidadEjecutada'))

			cantidad_ejecutada['sum_cantidad_ejecutada']=0 if cantidad_ejecutada['sum_cantidad_ejecutada']==None else cantidad_ejecutada['sum_cantidad_ejecutada']
				
			#Extraer los materiales:
			uc = UnidadConstructiva.objects.filter(
				codigo=item.codigoUC,
				catalogo__id=item.catalogoUnidadConstructiva.id).values('id')
			#import pdb; pdb.set_trace()
			desgloceMat = DesgloceMaterial.objects.filter(
				unidadConstructiva__id=uc[0]['id']).values(
				'material__codigo','material__descripcion', 'cantidad')

			for obj in desgloceMat:
				materiales.append({
					'codigo' : obj['material__codigo'],
					'hito': item.actividad.nombre,
					'descripcion' : obj['material__descripcion'],
					'cantidad' : float(obj['cantidad']) * float(item.cantidad),
					'cantidad_ejecutada': float(obj['cantidad']) * float(cantidad_ejecutada['sum_cantidad_ejecutada']),
					'pendiente': 0
				})
			#unificar cantidades de materialales:
			materialesAgrupados=[]
			
			for mat in materiales:
				agregar = True
				for obj in materialesAgrupados:
					if mat['codigo'] == obj['codigo'] and mat['hito'] == obj['hito']:
						obj['cantidad']	+= mat['cantidad']
						obj['cantidad_ejecutada'] += mat['cantidad_ejecutada']
						agregar = False
				if agregar:		
					materialesAgrupados.append(mat)				

			for obj in materialesAgrupados:
				obj['cantidad'] = round(obj['cantidad'],2)
				obj['cantidad_ejecutada'] = round(obj['cantidad_ejecutada'],2)
				obj['pendiente'] = round(obj['cantidad'] - obj['cantidad_ejecutada'],2)
		

		if listado:
			datos=materialesAgrupados			
		else:
			datos=[]		
		
		#import pdb; pdb.set_trace()
		proyecto = request.GET['proyecto'] if request.GET['proyecto'] else 0
		proyecto_model = Proyecto.objects.get(pk=int(proyecto))

		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte seguimiento materiales -'+proyecto_model.mcontrato.nombre+'.xls"'
		
		

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		
		
		if len(proyecto_model.nombre)<30:
			worksheet = workbook.add_worksheet(proyecto_model.nombre)
		else:
			worksheet = workbook.add_worksheet(proyecto_model.municipio.nombre)

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:B', 25)
		worksheet.set_column('C:C', 40)
		worksheet.set_column('D:F', 25)
		row=1
		col=0

		worksheet.write('A1', 'Hito', format1)
		worksheet.write('B1', 'Codigo', format1)
		worksheet.write('C1', 'Descripcion', format1)
		worksheet.write('D1', 'Cantidad a ejecutar', format1)
		worksheet.write('E1', 'Cantidad ejecutada', format1)
		worksheet.write('F1', 'Cantidad pendiente', format1)

		if datos:
			for item in datos:
				worksheet.write(row, col,item['hito'],format3)
				worksheet.write(row, col+1,item['codigo'],format3)
				worksheet.write(row, col+2,item['descripcion'],format3)
				worksheet.write(row, col+3,item['cantidad'],format3)
				worksheet.write(row, col+4,item['cantidad_ejecutada'],format3)
				worksheet.write(row, col+5,item['pendiente'],format3)
				row +=1

		workbook.close()
		return response
	except Exception as e:
			functions.toLog(e,'avanceObraLite.exportarMaterialesaLiquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})
@login_required
def exportarCantidadesaLiquidar(request):
	try:
		datos = []
		presupuesto_id= request.GET['presupuesto_id'] if request.GET['presupuesto_id'] else None
		cantidadesEjecutar = JCantidadesNodo.objects.filter(
			detallepresupuesto__presupuesto__id=int(presupuesto_id),
			cantidad__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			'detallepresupuesto__actividad__nombre'
			).annotate(
			total=Sum('cantidad'))

		cantidadesEjecutadas = DetalleReporteTrabajo.objects.filter(
			detallepresupuesto__presupuesto__id = int(presupuesto_id),
			cantidadEjecutada__gt=0).values(
			'detallepresupuesto__codigoUC',
			'detallepresupuesto__descripcionUC',
			'detallepresupuesto__catalogoUnidadConstructiva__id',
			#'liquidada'
			).annotate(total = Sum('cantidadEjecutada'))



		##import pdb; pdb.set_trace()	
		for cantidadEjecutar in cantidadesEjecutar:
			cejecutar = 0
			cliquidada = 0
			liquidada = False
			for cantidadEjecutada in cantidadesEjecutadas:
				if cantidadEjecutar['detallepresupuesto__codigoUC'] == cantidadEjecutada['detallepresupuesto__codigoUC']:
					cejecutar = cantidadEjecutada['total']					
					break


			datos.append(
				{
					'hito': cantidadEjecutar['detallepresupuesto__actividad__nombre'],
					'codigoUC': cantidadEjecutar['detallepresupuesto__codigoUC'],
					'descripcion': cantidadEjecutar['detallepresupuesto__descripcionUC'],
					'ejecutar': cantidadEjecutar['total'],
					'ejecutado': cejecutar,
					'pendiente': int(cantidadEjecutar['total']) - int(cejecutar)
				}
			)
		
		#import pdb; pdb.set_trace()
		proyecto = request.GET['proyecto'] if request.GET['proyecto'] else 0
		proyecto_model = Proyecto.objects.get(pk=int(proyecto))

		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Reporte seguimiento cantidades -'+proyecto_model.mcontrato.nombre+'.xls"'
		
		

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		if len(proyecto_model.nombre)<30:
			worksheet = workbook.add_worksheet(proyecto_model.nombre)
		else:
			worksheet = workbook.add_worksheet(proyecto_model.municipio.nombre)

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})		
		format2=workbook.add_format({'border':0})
		format1.set_align('center')
		format2.set_align('center')
		format2.set_align('vcenter')
		format3=workbook.add_format({'border':0,'font_size':12})
		# format3.set_align('center')
		# format3.set_align('vcenter')

		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format5.set_align('center')
		worksheet.set_column('A:B', 25)
		worksheet.set_column('C:C', 40)
		worksheet.set_column('D:F', 25)
		row=1
		col=0

		worksheet.write('A1', 'Hito', format1)
		worksheet.write('B1', 'Codigo UU CC', format1)
		worksheet.write('C1', 'Descripcion UU CC', format1)
		worksheet.write('D1', 'Cantidad a ejecutar', format1)
		worksheet.write('E1', 'Cantidad ejecutada', format1)
		worksheet.write('F1', 'Cantidad pendiente', format1)

		if datos:
			for item in datos:
				worksheet.write(row, col,item['hito'],format3)
				worksheet.write(row, col+1,item['codigoUC'],format3)
				worksheet.write(row, col+2,item['descripcion'],format3)
				worksheet.write(row, col+3,item['ejecutar'],format3)
				worksheet.write(row, col+4,item['ejecutado'],format3)
				worksheet.write(row, col+5,item['pendiente'],format3)
				row +=1

		workbook.close()
		return response
	except Exception as e:
			functions.toLog(e,'avanceObraLite.exportarCantidadesaLiquidar')
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})

def Str_fecha(date):
	
	my_time = datetime.min.time()
	my_datetime = datetime.combine(date, my_time)
	my_datetime = my_datetime.strftime("%d %b %Y")

	return my_datetime
	

@login_required
def descargar_plantilla_programacion(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_programacion.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Programacion')

	format0=workbook.add_format({'border':1})
	format0.set_bg_color('#D5D8DC')

	format1=workbook.add_format({'border':1,'font_size':13,'bold':True})
	format1.set_text_wrap()

	format2=workbook.add_format({'border':1})
	format2.set_text_wrap()

	format3=workbook.add_format({'border':1})
	format3.set_align('center')

	format4=workbook.add_format({'border':1})
	format4.set_align('center')
	format4.set_locked(True)


	format5=workbook.add_format({'border':1})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_text_wrap()


	format6=workbook.add_format({'font_size':13,'bold':True})
	format6.set_align('center')
	format6.set_text_wrap()
	
	format7=workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format7.set_align('center')
	format7.set_locked(True)
	# format7.set_text_wrap()



	format_red = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_red.set_bg_color('#E6B0AA')
	format_red.set_font_color('#C70039')
	format_red.set_align('center')
	format_red.set_locked(True)

	format_yellow = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_yellow.set_bg_color('#F9E79F')
	format_yellow.set_font_color('#FFC300')
	format_yellow.set_align('center')
	format_yellow.set_locked(True)

	format_green =  workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_green.set_bg_color('#ABEBC6')
	format_green.set_font_color('#28B463')
	format_green.set_align('center')
	format_green.set_locked(True)


	worksheet.set_column('A:A', 5)
	worksheet.set_column('B:B', 15)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:F', 16)


	cronograma_id= request.GET['cronograma_id']	
	presupuesto_id= request.GET['presupuesto_id']	
	
	

	merge_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})

	merge_format_horizontal = workbook.add_format({
    'bold': 1,
    'border': 1})    
	merge_format_horizontal.set_bg_color('#D5D8DC')

	worksheet.merge_range('A1:A3', 'Merged Range', merge_format)
	worksheet.merge_range('B1:B3', 'Merged Range', merge_format)
	worksheet.merge_range('C1:C3', 'Merged Range', merge_format)
	worksheet.merge_range('D1:D3', 'Merged Range', merge_format)
	worksheet.merge_range('E1:E3', 'Merged Range', merge_format)
	worksheet.merge_range('F1:F3', 'Merged Range', merge_format)
	# worksheet.merge_range('G1:G3', 'Merged Range', merge_format)
	# worksheet.merge_range('H1:H3', 'Merged Range', merge_format)

	worksheet.write('A1', 'Id', format1)
	# worksheet.write('B1', 'Hitos', format1)
	# worksheet.write('B1', 'Actividad', format1)
	worksheet.write('B1', 'Cod. UUCC', format1)
	worksheet.write('C1', 'Descripcion UUCC', format1)
	worksheet.write('D1', 'Cant. presupuestada', format1)
	worksheet.write('E1', 'Cant. programada', format1)
	worksheet.write('F1', '% Cant. Reportada', format1)

	worksheet.freeze_panes(3, 6)
	

	# import pdb; pdb.set_trace()
	col=6
	p=1 
	periodos = PeriodoProgramacion.objects.filter(cronograma_id=cronograma_id).order_by('fechaDesde')
	for semana in periodos:	
		# import pdb; pdb.set_trace()
		fechaDesde = Str_fecha(semana.fechaDesde)
		fechaHasta = Str_fecha(semana.fechaHasta)

		# fechaDesde = '{0}/{1}/{2}'.format(str(semana.fechaDesde.day),str(semana.fechaDesde.month),str(semana.fechaDesde.year))
		# fechaHasta = '{0}/{1}/{2}'.format(str(semana.fechaHasta.day),str(semana.fechaHasta.month),str(semana.fechaHasta.year))

		worksheet.write(0,col,'Periodo '+str(p), format6)
		worksheet.write(1,col,'('+fechaDesde+' - '+fechaHasta+')', format6)		
		worksheet.write(2,col,'Cantidad', format6)

		p+=1
		col+=1

	worksheet.set_column(6,col, 16)

	

	uucc_actividades = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	actividades = uucc_actividades.filter().values('actividad__id').distinct()
	actividades_cap = CEsquemaCapitulosActividadesG.objects.filter(pk__in=actividades).order_by('id')

	row=3

	for cap in actividades_cap:
		# import pdb; pdb.set_trace()
		worksheet.write(row,0, '', format0)
		# worksheet.merge_range('B{0}:{1}{0}'.format(row,xl_col_to_name(col)), 'Merged Range', merge_format)
		worksheet.merge_range('B{0}:F{0}'.format(row+1), cap.nombre, merge_format_horizontal)

		col=6
		for semana in periodos:	
			worksheet.write(row,col,'', format0)
			col+=1

		row+=1

		uucc = uucc_actividades.filter(actividad__id=cap.id).order_by('id')

		
		for item_act in uucc:
			worksheet.write(row,0, item_act.id, format2)
			# padre = CEsquemaCapitulosActividadesG.objects.get(pk=item_act.actividad.padre)
			# worksheet.write(row,1, padre.nombre, format2)
			# worksheet.write(row,1, item_act.actividad.nombre, format2)
			worksheet.write(row,1, item_act.codigoUC, format2)
			worksheet.write(row,2, item_act.descripcionUC, format2)
			worksheet.write(row,3, round(item_act.cantidad, 4), format4)	
			worksheet.write(row,4, '=SUM({0}{1}:{2}{1})'.format(xl_col_to_name(6),str(row+1),xl_col_to_name(col-1)), format4)
			worksheet.write(row,5, '=(E{0}/D{0})*100'.format(str(row+1)),format7)

			col=6
			for semana in periodos:	
				worksheet.write(row,col,'', format3)
				col+=1

			row+=1
		

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': '>',
		                                        'value':    100,
		                                        'format':   format_red})

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': 'between',
		                                        'minimum': 0,
		                                        'maximum': 99.99999,
		                                        'format':   format_yellow})

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': '==',
		                                        'value':    100,
		                                        'format':   format_green})



	workbook.close()

	return response

@login_required
@transaction.atomic
def guardar_programacion_archivo(request):

	try:		
		soporte= request.FILES['archivo']
		esquema_id= request.POST['esquema_id']
		cronograma_id= request.POST['cronograma_id']
		id_presupuesto= request.POST['presupuesto_id']
		doc = openpyxl.load_workbook(soporte,data_only=True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
		contador=hoja.max_row - 1
		numeroFila = 1
		
		mensaje_acumulado = ''
		if int(contador) > 0:
			sid = transaction.savepoint()
			

			validad_positiva_actividad_inicial = False
			cont_si=0
			for fila in hoja.rows:
				if fila[0].value and numeroFila>3:

					if FDetallePresupuesto.objects.filter(id=fila[0].value,presupuesto__id=id_presupuesto,codigoUC=fila[1].value).count() == 0:
						return JsonResponse({
							'message':'El Id indicado en la fila No.' + str(numeroFila) + ' no coincide con la UUCC'+ \
							'. Sugerimos descargar nuevamente la plantilla para verificar el Id en esta fila',
							'success':'error',
							'data':''})

					try:
						# import pdb; pdb.set_trace()
						if  float(fila[5].value)>0:
							mensaje_acumulado = mensaje_acumulado + '<br>El % Cant. reportada en la fila No ' + str(numeroFila) + ' es superior al 100%'
					except Exception as e:
						return JsonResponse({
							'message':'Hubo un error en fila No.' + str(numeroFila) + \
							'. Fue ingresado una letra en lugar de un numero en las cantidades de algun periodo en dicha fila '
							'. Sugerimos corregir la planilla gestionada y volverla a subir en SININ',
							'success':'error',
							'data':''})



				numeroFila+= 1

			transaction.savepoint_commit(sid)
			periodos = PeriodoProgramacion.objects.filter(cronograma_id=cronograma_id).order_by('fechaDesde')

			DetallePeriodoProgramacion.objects.filter(detallePresupuesto__presupuesto__id=id_presupuesto).delete()

			numeroFila = 1
			for fila in hoja.rows:
				if fila[0].value and numeroFila>3:

					c = 0
					for column in fila:
						if c>5 and column.value:
							if float(column.value)>0:
								# import pdb; pdb.set_trace()
								detalle_programacion = DetallePeriodoProgramacion(
									detallePresupuesto_id=fila[0].value,
									periodoProgramacion_id=periodos[c-6].id,
									cantidad=column.value)
								detalle_programacion.save()
						c+=1

				numeroFila+= 1
								
			
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.DiagramaGrahm(carga_masiva)')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
def excel_catalogo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_catalogo.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Catalogos')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 12)	

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	dato = request.GET['dato'] if 'dato' in request.GET else None
	#import pdb; pdb.set_trace()	
	if dato:
		qsetFilter = qsetFilter & (Q(nombre__icontains=dato))	

	qset = CatalogoUnidadConstructiva.objects.filter(qsetFilter)	

	formato_fecha = "%Y-%m-%d"

	if qset:
		worksheet.write('A1', 'Nombre', format1)
		worksheet.write('B1', 'Anio', format1)

	for catalogo in qset:
		worksheet.write(row, col, catalogo.nombre,format2)
		worksheet.write(row, col+1, catalogo.ano,format2)					
		row +=1

	workbook.close()
	return response

@login_required
def excel_uucc(request):
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_uucc.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('UUCC')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format4=workbook.add_format()
	format4.set_num_format("{num_format':'$#,##0.00'}")
	format4.set_align('center')
	format4.set_align('vcenter')	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 30)	

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None	
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))			
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = UnidadConstructiva.objects.filter(qsetFilter)	
	serializer = UnidadConstructivaSerializerLite(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Total MO', format1)
		worksheet.write('D1', 'Valor Total Mat', format1)
		worksheet.write('E1', 'Tipo UUCC', format1)
		#import pdb; pdb.set_trace()	
		for uucc in serializer.data:
			worksheet.write(row, col, uucc['codigo'],format2)
			worksheet.write(row, col+1, uucc['descripcion'],format2)					
			worksheet.write(row, col+2, float(uucc['totalManoDeObra']),format2)					
			worksheet.write(row, col+3, float(uucc['totalMateriales']),format2)					
			worksheet.write(row, col+4, uucc['tipoUnidadConstructiva']['nombre'],format2)							
			row +=1

	workbook.close()
	return response

def excel_mat(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_materiales.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Materiales')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))	
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = Material.objects.filter(qsetFilter)	
	serializer = MaterialSerializerLite2(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Unitario', format1)		
		worksheet.write('D1', 'Unidad de Medida', format1)		
		for material in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, material['codigo'],format2)
			worksheet.write(row, col+1, material['descripcion'],format2)	
			if material['valorUnitario']:
				worksheet.write(row, col+2, float(material['valorUnitario']),format2)								
			else:
				worksheet.write(row, col+2, '0.00',format2)								
			worksheet.write(row, col+3, material['unidadMedida'],format2)							
			row +=1

	workbook.close()
	return response

def excel_mo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_MO.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Mano de Obra')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	catalogo_id = request.GET['catalogo_id'] if 'catalogo_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(descripcion__icontains=dato)|Q(codigo__icontains=dato))	
	if catalogo_id:
		qsetFilter = qsetFilter & (Q(catalogo_id=catalogo_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = ManoDeObra.objects.filter(qsetFilter)	
	serializer = ManoObraSerializerLite2(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo', format1)
		worksheet.write('B1', 'Descripcion', format1)
		worksheet.write('C1', 'Valor Hora', format1)						
		for material in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, material['codigo'],format2)
			worksheet.write(row, col+1, material['descripcion'],format2)	
			if material['valorHora']:
				worksheet.write(row, col+2, float(material['valorHora']),format2)								
			else:
				worksheet.write(row, col+2, '0.00',format2)																	
			row +=1

	workbook.close()
	return response

def excel_desgl_mat(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_Desgl_Mat.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Desglose Material')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	uucc_id = request.GET['uucc_id'] if 'uucc_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(material__descripcion__icontains=dato) | Q(material__codigo__icontains=dato))
	if uucc_id:
		qsetFilter = qsetFilter & (Q(unidadConstructiva_id=uucc_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = DesgloceMaterial.objects.filter(qsetFilter)	
	serializer = DesgloceMaterialSerializer(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo UUCC', format1)
		worksheet.write('B1', 'Descripcion UUCC', format1)
		worksheet.write('C1', 'Codigo Material', format1)
		worksheet.write('D1', 'Descripcion Material', format1)	
		worksheet.write('E1', 'Cantidad', format1)	

		for dm in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, dm['unidadConstructiva']['codigo'],format2)
			worksheet.write(row, col+1, dm['unidadConstructiva']['descripcion'],format2)	
			worksheet.write(row, col+2, dm['material']['codigo'],format2)											
			worksheet.write(row, col+3, dm['material']['descripcion'],format2)	
			worksheet.write(row, col+4, dm['cantidad'],format2)											
			row +=1

	workbook.close()
	return response

def excel_desgl_mo(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_desgl_MO.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Desglose Material')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})	
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 30)	
	worksheet.set_column('C:C', 30)	
	worksheet.set_column('D:D', 30)	
	worksheet.set_column('E:E', 17)		

	row=1
	col=0

	qsetFilter = (~Q(id=0))
	
	dato = request.GET['dato'] if 'dato' in request.GET else None	
	uucc_id = request.GET['uucc_id'] if 'uucc_id' in request.GET else None		
	if dato:
		qsetFilter = qsetFilter & (Q(manoDeObra__descripcion__icontains=dato) | Q(manoDeObra__codigo__icontains=dato)) 
	if uucc_id:
		qsetFilter = qsetFilter & (Q(unidadConstructiva_id=uucc_id))
	#import pdb; pdb.set_trace()
	formato_fecha = "%Y-%m-%d"	
	qset = DesgloceManoDeObra.objects.filter(qsetFilter)	
	serializer = DesgloceManoDeObraSerializer(qset,many=True)	

	if serializer:
		worksheet.write('A1', 'Codigo UUCC', format1)
		worksheet.write('B1', 'Descripcion UUCC', format1)
		worksheet.write('C1', 'Codigo MO', format1)
		worksheet.write('D1', 'Descripcion MO', format1)	
		worksheet.write('E1', 'Rendimiento', format1)	

		for dmo in serializer.data:
			#import pdb; pdb.set_trace()	
			worksheet.write(row, col, dmo['unidadConstructiva']['codigo'],format2)
			worksheet.write(row, col+1, dmo['unidadConstructiva']['descripcion'],format2)	
			worksheet.write(row, col+2, dmo['manoDeObra']['codigo'],format2)											
			worksheet.write(row, col+3, dmo['manoDeObra']['descripcion'],format2)	
			worksheet.write(row, col+4, dmo['rendimiento'],format2)											
			row +=1

	workbook.close()
	return response


@login_required
def inactivar_catalogo(request):
	sid = transaction.savepoint()
	try:		
		#import pdb; pdb.set_trace()	
		id_catalogo = request.GET['id'] if request.GET['id'] else None
		if id_catalogo:
			catalogo = CatalogoUnidadConstructiva.objects.get(pk=id_catalogo)
			if catalogo:
				catalogo.activo = 0
				catalogo.save()		
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok','data':''})		
		else:
			return JsonResponse({'message':'No se encontraron registros','success':'ok','data':''})			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.inactivar_catalogo')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})			

@login_required
def activar_catalogo(request):
	sid = transaction.savepoint()
	try:		
		#import pdb; pdb.set_trace()	
		id_catalogo = request.GET['id'] if request.GET['id'] else None
		if id_catalogo:
			catalogo = CatalogoUnidadConstructiva.objects.get(pk=id_catalogo)
			if catalogo:
				catalogo.activo = 1
				catalogo.save()		
				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok','data':''})		
		else:
			return JsonResponse({'message':'No se encontraron registros','success':'ok','data':''})			
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.activar_catalogo')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})					


@login_required
def descargar_plantilla_masiva(request):
	return functions.exportarArchivoS3('plantillas/catalogos/PlantillaCargaMasivaCatalogo.xlsx')



@transaction.atomic
def cargar_excel_catalogo_masivo(request):	
	#import pdb; pdb.set_trace()		
	try:
		array_uucc = []
		array_mat = []
		array_mo = []	
		catalogo_id= request.POST['catalogo_id']
		soporte= request.FILES['archivo']
		doc = openpyxl.load_workbook(soporte)
		
		#Validacion desgl_mat y desgl_mo
		index=0
		hoja = doc.get_sheet_by_name('UUCC')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_uucc.append(fila[0].value)
			index+=1


		index=0
		hoja = doc.get_sheet_by_name('Material')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_mat.append(fila[0].value)
			index+=1

		index=0
		hoja = doc.get_sheet_by_name('Mano de Obra')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None:
					array_mo.append(fila[0].value)	
			index+=1	

		index=0		
		hoja = doc.get_sheet_by_name('Desgl_Material')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None:
					if fila[0].value not in array_uucc:
						return JsonResponse({'message':'La UUCC con codigo '+str(fila[0].value)+' en Desgl_Material, no se encuentra dentro del listado de UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
					
					if fila[1].value not in array_mat:
						return JsonResponse({'message':'El material con codigo '+str(fila[1].value)+' en Desgl_Material, no se encuentra dentro del listado de Material','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			
											
			index+=1
		
		index=0		
		hoja = doc.get_sheet_by_name('Desgl_MO')		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None:
					if fila[0].value not in array_uucc:
						return JsonResponse({'message':'La UUCC con codigo '+str(fila[0].value)+' en Desgl_MO, no se encuentra dentro del listado de UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
					
					if fila[1].value not in array_mo:
						return JsonResponse({'message':'La Mano de Obra con codigo '+str(fila[1].value)+' en Desgl_MO, no se encuentra dentro del listado de Mano de Obra','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			
											
			index+=1													

		#Fin Validacion

		#UUCC		
		hoja = doc.get_sheet_by_name('UUCC')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc = UnidadConstructiva.objects.filter(qset).first()

					if uucc is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						tipoUUCC = fila[2].value
						uucc2 = UnidadConstructiva(codigo=codigo,descripcion=descripcion,tipoUnidadConstructiva_id=tipoUUCC,catalogo_id=catalogo_id)
						uucc2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1

		
		#Material
		hoja = doc.get_sheet_by_name('Material')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None and fila[3].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					mat = Material.objects.filter(qset).first()

					if mat is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						valorUnitario = fila[2].value
						UnidadDeMedida = str(fila[3].value)
						mat2 = Material(codigo=codigo,descripcion=descripcion,valorUnitario=valorUnitario,unidadMedida=UnidadDeMedida,catalogo_id=catalogo_id)
						mat2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1


		#Mo
		hoja = doc.get_sheet_by_name('Mano de Obra')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qset = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					mo = ManoDeObra.objects.filter(qset).first()

					if mo is None:
						codigo = str(fila[0].value)
						descripcion = str(fila[1].value)
						valorHora = fila[2].value
						mo2 = ManoDeObra(codigo=codigo,descripcion=descripcion,valorHora=valorHora,catalogo_id=catalogo_id)
						mo2.save()				
					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1


		#Desgl_Material
		hoja = doc.get_sheet_by_name('Desgl_Material')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qsetUUCC = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc_id = UnidadConstructiva.objects.filter(qsetUUCC).values('id').first()
					
					qsetMat = (Q(codigo = str(fila[1].value)) & Q(catalogo_id = catalogo_id))
					mat_id = Material.objects.filter(qsetMat).values('id').first()

					if uucc_id is not None and mat_id is not None:
						qsetDegl_mat = (Q(material_id = mat_id['id']) & Q(unidadConstructiva_id = uucc_id['id']))						
						desgl_mat = DesgloceMaterial.objects.filter(qsetDegl_mat)

						if any(desgl_mat) is False:
							desgl_mat2 = DesgloceMaterial(unidadConstructiva_id=uucc_id['id'],material_id=mat_id['id'],cantidad=fila[2].value)
							desgl_mat2.save()

					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1						

		#Desgl_MO
		hoja = doc.get_sheet_by_name('Desgl_MO')

		index=0				
		#import pdb; pdb.set_trace()		
		for fila in hoja.rows:
			if index>=1:
				if fila[0].value is not None and fila[1].value is not None and fila[2].value is not None:
					qsetUUCC = (Q(codigo = str(fila[0].value)) & Q(catalogo_id = catalogo_id))
					uucc_id = UnidadConstructiva.objects.filter(qsetUUCC).values('id').first()
					
					qsetMo = (Q(codigo = str(fila[1].value)) & Q(catalogo_id = catalogo_id))
					mo_id = ManoDeObra.objects.filter(qsetMo).values('id').first()

					if uucc_id is not None and mo_id is not None:
						qsetDegl_mo = (Q(manoDeObra_id = mo_id['id']) & Q(unidadConstructiva_id = uucc_id['id']))						
						desgl_mo = DesgloceManoDeObra.objects.filter(qsetDegl_mo)

						if any(desgl_mo) is False:
							desgl_mo2 = DesgloceManoDeObra(unidadConstructiva_id=uucc_id['id'],manoDeObra_id=mo_id['id'],rendimiento=fila[2].value)
							desgl_mo2.save()

					# return JsonResponse({'message':'No se admiten campos vacios en UUCC','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		
			index+=1						
		
		return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response



@login_required
@transaction.atomic
def confirmar_fechas(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cronograma=Cronograma.objects.get(id=respuesta['id_cronograma'])
		cronograma.confirmarFechas=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cronograma',id_manipulado=respuesta['id_cronograma'])
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.confirmar_fechas')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@transaction.atomic
def ActualizarPeriodoPrincipal(request):	
	# import pdb; pdb.set_trace()	
	sid = transaction.savepoint()	
	try:
		cronograma = Cronograma.objects.get(pk=request.POST['id_cronograma'])

		if DetallePeriodoProgramacion.objects.filter(periodoProgramacion__cronograma_id=cronograma.id).count()==0 and not cronograma.programacionCerrada and not cronograma.confirmarFechas:
			cronograma.periodicidad_id = request.POST['periodicidad_id']
			cronograma.fechaInicio = request.POST['fechaInicio']
			cronograma.fechaFinal = request.POST['fechaFinal']
			cronograma.save()

			PeriodoProgramacion.objects.filter(cronograma_id=cronograma.id).delete()
			GenerarPeriodosDelCronograma(cronograma.id);

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.cronograma',id_manipulado=cronograma.id)
			logs_model.save()
			transaction.savepoint_commit(sid)

			return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})

		else:
			if not cronograma.programacionCerrada:
				return JsonResponse({'message':'Este cronograma ya cerró la programación','success':'error',
				'data':''})
			elif not cronograma.confirmarFechas:
				return JsonResponse({'message':'Este cronograma ya confirmó las fechas de la programación','success':'error',
				'data':''})
			else:
				return JsonResponse({'message':'Este cronograma ya cuenta con detalles sobre los periodos semanales','success':'error',
				'data':''})
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.confirmar_fechas')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		



class ReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):
	periodoProgramacion=PeriodoProgramacionLiteSerializer(read_only=True)
	periodoProgramacion_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=PeriodoProgramacion.objects.all())


	usuario_registro=UsuarioSerializer(read_only=True)
	usuario_registro_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())
	
	

	
	class Meta:
		model = ReporteTrabajo
		fields=('id','reporteCerrado','periodoProgramacion','periodoProgramacion_id', \
			'fechaReporte',\
			'usuario_registro','usuario_registro_id',\
			'sinAvance','motivoSinAvance','reporteCerrado')



class ReporteTrabajoLiteSerializer(serializers.HyperlinkedModelSerializer):
	periodoProgramacion=PeriodoProgramacionLiteSerializer(read_only=True)
	usuario_registro = UsuarioLiteSerializer(read_only=True)
	class Meta:
		model = ReporteTrabajo
		fields = ('id','fechaReporte','periodoProgramacion',
			'reporteCerrado',
			'usuario_registro','sinAvance','motivoSinAvance')

	


class ReporteTrabajoGraficoViewSet(viewsets.ModelViewSet):
	
	model=ReporteTrabajo
	queryset = model.objects.all()
	serializer_class = ReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		
		try:
			queryset = super(ReporteTrabajoGraficoViewSet, self).get_queryset()
			dato= self.request.query_params.get('dato',None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)			
			proyecto_id= self.request.query_params.get('proyecto_id',None)			
			registrado= self.request.query_params.get('registrado',None)			
			lite = self.request.query_params.get('lite',None)

			sinAvance = self.request.query_params.get('sinAvance',None)
			reporteCerrado = self.request.query_params.get('reporteCerrado',None)
			fechaReporte = self.request.query_params.get('fechaReporte',None)

			qset=(~Q(id=0))
			

			#import pdb; pdb.set_trace()			
			if cronograma_id:
				qset = qset &(
					Q(periodoProgramacion__cronograma_id=cronograma_id)
					)		

			if proyecto_id:
				qset = qset &(
					Q(periodoProgramacion__cronograma__proyecto__id=proyecto_id)
					)		


			if dato:
				qset = qset &(
					Q(periodoProgramacion__cronograma__nombre__icontains=dato)|
					Q(usuario_registro__persona__nombres__icontains=dato)|
					Q(usuario_registro__persona__apellidos__icontains=dato)
					)

			if sinAvance:
				qset = qset &(
					Q(sinAvance=sinAvance)
					)

			if reporteCerrado:
				qset = qset &(
					Q(reporteCerrado=reporteCerrado)
					)

			if fechaReporte:
				qset = qset &(
					Q(fechaReporte=fechaReporte)
					)

			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-fechaReporte')

			##import pdb; pdb.set_trace()
			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = ReporteTrabajoLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			if lite:
				serializer = ReporteTrabajoLiteSerializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = ReporteTrabajoSerializer(data=request.data,context={'request': request})

				verifica_reporte=ReporteTrabajo.objects.filter(periodoProgramacion__cronograma_id=request.data['cronograma_id'],fechaReporte=request.data['fechaReporte'])

				if len(verifica_reporte)>0:
					return Response({'message':'No se puede registrar dos reportes de trabajos con la misma fecha','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)

					
				# 
				
				periodoprogramacion = PeriodoProgramacion.objects.filter(cronograma__id=request.data['cronograma_id'],fechaDesde__lte=request.data['fechaReporte'] ,fechaHasta__gte=request.data['fechaReporte'])
				

				if periodoprogramacion:
					request.data['periodoProgramacion_id'] = periodoprogramacion.last().id

				else:
					return Response({'message':'La fecha ingresada no se encuentra dentro de la programación establecida','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)

				

				# import pdb; pdb.set_trace()
				if serializer.is_valid():
					serializer.save(
						periodoProgramacion_id=request.data['periodoProgramacion_id'],
						usuario_registro_id=request.data['usuario_registro_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					# errors = funciones.erroresSerializer(serializer.errors)
					functions.toLog(serializer.errors, self.nombre_modulo + ".create")
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'El registro se está siendo utilizado en otro lugar del sistema','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	




class DetalleReporteTrabajoSerializer(serializers.HyperlinkedModelSerializer):
	reporteTrabajo =ReporteTrabajoLiteSerializer(read_only=True)
	detallePresupuesto =DetallePresupuestoGraficoLite2Serializer(read_only=True)

	reporteTrabajo_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=ReporteTrabajo.objects.all())
	detallePresupuesto_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=FDetallePresupuesto.objects.all())
	
	class Meta:
		model = DetalleReporteTrabajo
		fields=('id','detallePresupuesto','detallePresupuesto_id', \
			'reporteTrabajo','reporteTrabajo_id',\
			'cantidad')



class DetalleReporteTrabajoLiteSerializer(serializers.HyperlinkedModelSerializer): 
	reporteTrabajo =ReporteTrabajoLiteSerializer(read_only=True)
	detallePresupuesto =DetallePresupuestoGraficoLite2Serializer(read_only=True)

	class Meta:
		model = DetalleReporteTrabajo
		fields = ('id','reporteTrabajo','detallePresupuesto','cantidad')

	


class DetalleReporteTrabajoViewSet(viewsets.ModelViewSet):
	
	model=DetalleReporteTrabajo
	queryset = model.objects.all()
	serializer_class = DetalleReporteTrabajoSerializer	
	model_log=Logs
	model_acciones=Acciones	
	nombre_modulo='avanceObraLite.detalle_reporte_trabajo'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		
		try:
			queryset = super(DetalleReporteTrabajoViewSet, self).get_queryset()			
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			cronograma_id= self.request.query_params.get('cronograma_id',None)			
			reporteTrabajo_id= self.request.query_params.get('reporteTrabajo_id',None)			
			detallePresupuesto_id= self.request.query_params.get('detallePresupuesto_id',None)
			lite = self.request.query_params.get('lite',None)

			qset=(~Q(id=0))
			

			#import pdb; pdb.set_trace()			
			if reporteTrabajo_id:
				qset = qset &(
					Q(reporteTrabajo_id=reporteTrabajo_id)
					)		

			if reporteTrabajo_id:
				qset = qset &(
					Q(reporteTrabajo_id=reporteTrabajo_id)
					)		



			if qset is not None:
				queryset = self.model.objects.filter(qset)

			##import pdb; pdb.set_trace()
			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = DetalleReporteTrabajoLiteSerializer(page,many=True)
					else:
						serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			if lite:
				serializer = DetalleReporteTrabajoLiteSerializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)			
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
	@transaction.atomic		
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetalleReporteTrabajoSerializer(data=request.data,context={'request': request})

				if serializer.is_valid():					
					serializer.save(reporteTrabajo_id=request.data['reporteTrabajo_id'],
						detallePresupuesto_id=request.data['detallePresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleReporteTrabajoSerializer(instance,data=request.data,context={'request': request},partial=partial)
				
				if serializer.is_valid():				
					serializer.save(reporteTrabajo_id=request.data['reporteTrabajo_id'],
						detallePresupuesto_id=request.data['detallePresupuesto_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


@login_required
@transaction.atomic
def cerrar_reportetrabajo(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		cronograma=ReporteTrabajo.objects.get(id=respuesta['id'])
		cronograma.reporteCerrado=True
		cronograma.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='avanceObraLite.reportetrabajo',id_manipulado=respuesta['id'])
		logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.cierre_reportetrabajo')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



def cantidad_maxima_detallePresupuesto(request):
	try:
		
		detallePresupuesto_id=request.GET['detallePresupuesto_id']
		DetalleReporteTrabajo_id=request.GET['DetalleReporteTrabajo_id']

		diferencia = 0
		if DetalleReporteTrabajo_id:
			# import pdb; pdb.set_trace()
			object_detalle = FDetallePresupuesto.objects.get(pk=detallePresupuesto_id)
			serializer = DetallePresupuestoGraficoLite5Serializer(object_detalle)
			if serializer.data:
				diferencia = float(serializer.data['cantidad_aejecutar'])-float(serializer.data['cantidad_ejecutada'])

				diferencia+=float(DetalleReporteTrabajo.objects.get(pk=DetalleReporteTrabajo_id).cantidad)

		else:
			object_detalle = FDetallePresupuesto.objects.get(pk=detallePresupuesto_id)
			serializer = DetallePresupuestoGraficoLite5Serializer(object_detalle)
			if serializer.data:
				diferencia = float(serializer.data['cantidad_aejecutar'])-float(serializer.data['cantidad_ejecutada'])

		
		return JsonResponse({'message':'','success':'ok','data':{'cantidad_maxima':diferencia}})


		
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:		
		functions.toLog(e,'avanceObraLite.cantidad_maxima_detallePresupuesto')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
def seguimientoSemanal(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Avance de Obra - Seguimiento Semanal.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})	

	format0=workbook.add_format({'border':1})
	format0.set_bg_color('#D5D8DC')

	format1=workbook.add_format({'border':1,'font_size':12,'bold':True})
	format1.set_text_wrap()

	format_date = workbook.add_format(
		{'border':1, 'num_format': 'd-mmm-yy',
		'align': 'center','valign': 'vcenter'})
	format_date.set_text_wrap()

	format2=workbook.add_format({'border':1,'font_size':8})
	format2.set_text_wrap()

	format3=workbook.add_format({'border':1,'font_size':8})
	format3.set_align('center')

	format4=workbook.add_format({'border':1,'font_size':8})
	format4.set_align('center')
	format4.set_locked(True)

	currency_format = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8})
	currency_format.set_align('center')

	format4_GREY=workbook.add_format({'border':1,'font_size':8,'bold':True})
	format4_GREY.set_align('center')
	format4_GREY.set_bg_color('#D5D8DC')
	format4_GREY.set_locked(True)
	
	currency_format_GREY = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8,'bold':True})
	currency_format_GREY.set_align('center')
	currency_format_GREY.set_bg_color('#D5D8DC')
	
	format5=workbook.add_format({'border':1,'font_size':8})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_text_wrap()


	format6=workbook.add_format({'font_size':12,'bold':True})
	format6.set_align('center')
	format6.set_text_wrap()
	
	format7=workbook.add_format({'border':1, 'num_format': '0.00"%"','font_size':8})
	format7.set_align('center')
	format7.set_locked(True)
	# format7.set_text_wrap()



	format_red = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_red.set_bg_color('#E6B0AA')
	format_red.set_font_color('#C70039')
	format_red.set_align('center')
	format_red.set_locked(True)

	

	format_yellow = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_yellow.set_bg_color('#F9E79F')
	format_yellow.set_font_color('#FFC300')
	format_yellow.set_align('center')
	format_yellow.set_locked(True)

	format_green =  workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_green.set_bg_color('#ABEBC6')
	format_green.set_font_color('#28B463')
	format_green.set_align('center')
	format_green.set_locked(True)

	merge_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})
	merge_format_horizontal = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})    
	merge_format_horizontal.set_bg_color('#D5D8DC')
	merge_format_horizontal.set_text_wrap()

	border_format=workbook.add_format({'border':2})


	cronograma_id= request.GET['cronograma_id']	
	presupuesto_id = EPresupuesto.objects.filter(cronograma_id=cronograma_id,cerrar_presupuesto=True).last().id	

	actividades = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	capitulos_id = actividades.filter().values('actividad__id').distinct()
	capitulos = CEsquemaCapitulosActividadesG.objects.filter(pk__in=capitulos_id).order_by('id')
	periodos = PeriodoProgramacion.objects.filter(cronograma_id=cronograma_id).order_by('fechaDesde')

	worksheet1 = workbook.add_worksheet('PROGRAMACION SEMANAL')
	worksheet2 = workbook.add_worksheet('CANTIDADES SEMANAL')
	worksheet3 = workbook.add_worksheet('SEGUIMIENTO SEMANAL')


	#--------------------Hoja 1-----------------------------------------------------------------
	#-------------------------------------------------------------------------------------------
	

	worksheet1.set_column('A:A', 5)
	worksheet1.set_column('B:B', 12)
	worksheet1.set_column('C:C', 30)
	worksheet1.set_column('D:D', 12)
	worksheet1.set_column('E:E', 16)
	worksheet1.set_column('F:F', 16)


	worksheet1.merge_range('A1:A3', 'Merged Range', merge_format)
	worksheet1.merge_range('B1:B3', 'Merged Range', merge_format)
	worksheet1.merge_range('C1:C3', 'Merged Range', merge_format)
	worksheet1.merge_range('D1:D3', 'Merged Range', merge_format)
	worksheet1.merge_range('E1:E3', 'Merged Range', merge_format)
	worksheet1.merge_range('F1:F3', 'Merged Range', merge_format)
	

	worksheet1.write('A1', 'Item', format1)
	worksheet1.write('B1', 'Cod. UUCC', format1)
	worksheet1.write('C1', 'Descripcion UUCC', format1)
	worksheet1.write('D1', 'Cant.', format1)	
	worksheet1.write('E1', 'Valor Unitario', format1)
	worksheet1.write('F1', 'Valor TOTAL', format1)

	
	worksheet1.freeze_panes(3, 6)
	
	border_format=workbook.add_format({
                            'border':2
                           })

	# import pdb; pdb.set_trace()
	col=6
	p=1

	#----------------- Inicio tabla principal-----------------------
	str_sumatorias_periodo = []
	for semana in periodos:	
		str_sumatorias_periodo.append('=')
		# import pdb; pdb.set_trace()
		fechaDesde = Str_fecha(semana.fechaDesde)
		fechaHasta = Str_fecha(semana.fechaHasta)

		worksheet1.merge_range('{0}1:{1}1'.format(xl_col_to_name(col),xl_col_to_name(col+1)), 'Periodo '+str(p), merge_format)
		worksheet1.write(1,col,fechaDesde, format6)		
		worksheet1.write(2,col,'Cantidad', format6)

		worksheet2.merge_range('{0}1:{1}1'.format(xl_col_to_name(col),xl_col_to_name(col+1)), 'Periodo '+str(p), merge_format)
		worksheet2.write(1,col,fechaDesde, format6)		
		worksheet2.write(2,col,'Cantidad', format6)
		col+=1
		
		worksheet1.write(1,col,fechaHasta, format6)	
		worksheet1.write(2,col,'Valor', format6)

		worksheet2.write(1,col,fechaHasta, format6)	
		worksheet2.write(2,col,'Valor', format6)
		p+=1
		col+=1


	worksheet1.merge_range('{0}1:{0}3'.format(xl_col_to_name(col)), 'Merged Range', merge_format)
	worksheet1.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+1)), 'Merged Range', merge_format)
	worksheet1.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+2)), 'Merged Range', merge_format)	
	worksheet2.merge_range('{0}1:{0}3'.format(xl_col_to_name(col)), 'Merged Range', merge_format)
	worksheet2.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+1)), 'Merged Range', merge_format)
	worksheet2.merge_range('{0}1:{0}3'.format(xl_col_to_name(col+2)), 'Merged Range', merge_format)

	worksheet1.write('{0}1'.format(xl_col_to_name(col)), 'Cant.', format1)
	worksheet1.write('{0}1'.format(xl_col_to_name(col+1)), 'Valor', format1)
	worksheet1.write('{0}1'.format(xl_col_to_name(col+2)), '%', format1)
	worksheet1.set_column(col+1,col+1, 16)
	worksheet1.set_column(6,col-1, 16)	
	worksheet2.write('{0}1'.format(xl_col_to_name(col)), 'Cant.', format1)
	worksheet2.write('{0}1'.format(xl_col_to_name(col+1)), 'Valor', format1)
	worksheet2.write('{0}1'.format(xl_col_to_name(col+2)), '%', format1)
	worksheet2.set_column(col+1,col+1, 16)
	worksheet2.set_column(6,col-1, 16)	

	row=3
	item=1

	str_suma_subtotales = '='
	str_suma_subtotales_periodos = '='
	for cap in capitulos:
		# import pdb; pdb.set_trace()
		worksheet1.write(row,0, '', format0)
		worksheet1.merge_range('A{0}:F{0}'.format(row+1), cap.nombre, merge_format_horizontal)		
		worksheet1.merge_range('{0}{1}:{2}{1}'.format(xl_col_to_name(6),row+1,xl_col_to_name(col+2)), '', merge_format_horizontal)
		
		worksheet2.write(row,0, '', format0)
		worksheet2.merge_range('A{0}:F{0}'.format(row+1), cap.nombre, merge_format_horizontal)		
		worksheet2.merge_range('{0}{1}:{2}{1}'.format(xl_col_to_name(6),row+1,xl_col_to_name(col+2)), '', merge_format_horizontal)

		row+=1
		actividades_cap = actividades.filter(actividad__id=cap.id).order_by('id')

		#----------UUCC ACTIVIDADES------------------------
		for item_act in actividades_cap:
			if item_act.id==actividades_cap[0].id:
				row_periodo  = row+1

			worksheet1.write(row,0, item, format2)			
			worksheet1.write(row,1, item_act.codigoUC, format2)
			worksheet1.write(row,2, item_act.descripcionUC, format2)
			worksheet1.write(row,3, round(item_act.cantidad, 4), format4)
			worksheet1.write(row,4, item_act.valorGlobal, currency_format)
			worksheet1.write(row,5, '=E{0}*D{0}'.format(str(row+1)), currency_format)

			worksheet2.write(row,0, item, format2)
			worksheet2.write(row,1, item_act.codigoUC, format2)
			worksheet2.write(row,2, item_act.descripcionUC, format2)
			worksheet2.write(row,3, round(item_act.cantidad, 4), format4)	
			worksheet2.write(row,4, item_act.valorGlobal, currency_format)
			worksheet2.write(row,5, '=E{0}*D{0}'.format(str(row+1)), currency_format)

			item+=1
			#----------Periodos------------------------
			col=6			
			str_suma_cantidades = '='
			str_suma_valores  = '='
			for semana in periodos:	
				str_suma_cantidades = str_suma_cantidades + '+{0}{1}'.format(xl_col_to_name(col),row+1)		
				# -------------------		
				cantidad = DetallePeriodoProgramacion.objects.filter(detallePresupuesto_id=item_act.id,periodoProgramacion_id=semana.id).aggregate(Sum('cantidad'))
				if cantidad:
					worksheet1.write(row,col,cantidad['cantidad__sum'], format4)
				else:
					worksheet1.write(row,col,0, format4)

				cantidad = DetalleReporteTrabajo.objects.filter(detallePresupuesto_id=item_act.id,reporteTrabajo__periodoProgramacion_id=semana.id).aggregate(Sum('cantidad'))
				if cantidad:
					worksheet2.write(row,col,cantidad['cantidad__sum'], format4)
				else:
					worksheet2.write(row,col,0, format4)

				col+=1

				str_suma_valores = str_suma_valores + '+{0}{1}'.format(xl_col_to_name(col),row+1)
				worksheet1.write(row,col,'=E{0}*{1}{0}'.format(row+1,xl_col_to_name(col-1)), currency_format)
				worksheet2.write(row,col,'=E{0}*{1}{0}'.format(row+1,xl_col_to_name(col-1)), currency_format)
				col+=1

			#----------Fin Periodos------------------------

			#----------Valores programados------------------------
			

			worksheet1.write(row,col, str_suma_cantidades, format4)
			worksheet1.write(row,col+1, str_suma_valores, currency_format)
			worksheet1.write(row,col+2, '=({1}{0}/D{0})*100'.format(str(row+1),xl_col_to_name(col)),format7)
			worksheet1.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '>',
		                                        'value':    100,
		                                        'format':   format_red})
			worksheet1.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': 'between',
		                                        'minimum': 0,
		                                        'maximum': 99.99999,
		                                        'format':   format_yellow})
			worksheet1.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '==',
		                                        'value':    100,
		                                        'format':   format_green})


			worksheet2.write(row,col, str_suma_cantidades, format4)
			worksheet2.write(row,col+1, str_suma_valores, currency_format)
			worksheet2.write(row,col+2, '=({1}{0}/D{0})*100'.format(str(row+1),xl_col_to_name(col)),format7)
			worksheet2.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '>',
		                                        'value':    100,
		                                        'format':   format_red})
			worksheet2.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': 'between',
		                                        'minimum': 0,
		                                        'maximum': 99.99999,
		                                        'format':   format_yellow})
			worksheet2.conditional_format('{1}{0}:{1}{0}'.format(row+1,xl_col_to_name(col+2)), {
												'type':     'cell',
		                                        'criteria': '==',
		                                        'value':    100,
		                                        'format':   format_green})

			row+=1
			#----------Fin Valores programados------------------------
		#----------Fin UUCC ACTIVIDADES------------------------		
		worksheet1.merge_range('A{0}:E{0}'.format(row+1), 'SUBTOTAL '+cap.nombre, merge_format_horizontal)
		worksheet2.merge_range('A{0}:E{0}'.format(row+1), 'SUBTOTAL '+cap.nombre, merge_format_horizontal)

		sumatoria_valor = 'F{1}:F{0}'.format(row,row_periodo)
		sumatoria_cant_extra = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)
		sumatoria_valor_extra = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col+1),row_periodo)

		str_suma_subtotales = str_suma_subtotales + '+F{0}'.format(row+1)
		str_suma_subtotales_periodos = str_suma_subtotales_periodos + '+{1}{0}'.format(row+1,xl_col_to_name(col+1))
		worksheet1.write(row,5, '=SUM('+sumatoria_valor+')', currency_format_GREY)
		worksheet1.write(row,col, '=SUM('+sumatoria_cant_extra+')', format4_GREY)
		worksheet1.write(row,col+1, '=SUM('+sumatoria_valor_extra+')', currency_format_GREY)
		worksheet1.write(row,col+2, '', format0)

		worksheet2.write(row,5, '=SUM('+sumatoria_valor+')', currency_format_GREY)
		worksheet2.write(row,col, '=SUM('+sumatoria_cant_extra+')', format4_GREY)
		worksheet2.write(row,col+1, '=SUM('+sumatoria_valor_extra+')', currency_format_GREY)
		worksheet2.write(row,col+2, '', format0)

		col=6
		# import pdb; pdb.set_trace()
		col_aux=0

		for semana in periodos:	
			
			sumatoria_cant_periodo = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)		
			worksheet1.write(row,col, '=SUM('+sumatoria_cant_periodo+')', format4_GREY)
			worksheet2.write(row,col, '=SUM('+sumatoria_cant_periodo+')', format4_GREY)
			col+=1
			str_sumatorias_periodo[col_aux]=str_sumatorias_periodo[col_aux]+'+{1}{0}'.format(row+1,xl_col_to_name(col))
			sumatoria_valor_periodo = '{1}{2}:{1}{0}'.format(row,xl_col_to_name(col),row_periodo)
			worksheet1.write(row,col, '=SUM('+sumatoria_valor_periodo+')', currency_format_GREY)
			worksheet2.write(row,col, '=SUM('+sumatoria_valor_periodo+')', currency_format_GREY)
			col+=1
			col_aux+=1


		row+=1
		#----------------- Fin tabla principal-----------------------
	row+=1
	

	
	worksheet1.merge_range('D{0}:E{0}'.format(row+1), 'TOTAL COSTOS DIRECTOS DEL PROYECTO', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+1), 'TOTAL COSTOS DIRECTOS DEL PROYECTO', merge_format_horizontal)
	worksheet1.write(row,5, '='+str_suma_subtotales, currency_format)
	worksheet2.write(row,5, '='+str_suma_subtotales, currency_format)

	k_aiu = EPresupuesto.objects.get(pk=presupuesto_id).aiu
	worksheet1.write(row+1,3, 'Administración', format2)
	worksheet2.write(row+1,3, 'Administración', format2)
	worksheet1.write(row+1,4, float(k_aiu-1)*60, format7)
	worksheet2.write(row+1,4, float(k_aiu-1)*60, format7)
	worksheet1.write(row+1,5, '=F{1}*E{0}/100'.format(row+2,row+1), currency_format)
	worksheet2.write(row+1,5, '=F{1}*E{0}/100'.format(row+2,row+1), currency_format)

	
	worksheet1.write(row+2,3, 'Imprevisto', format2)
	worksheet2.write(row+2,3, 'Imprevisto', format2)
	worksheet1.write(row+2,4, float(k_aiu-1)*20, format7)
	worksheet2.write(row+2,4, float(k_aiu-1)*20, format7)
	worksheet1.write(row+2,5, '=F{1}*E{0}/100'.format(row+3,row+1), currency_format)
	worksheet2.write(row+2,5, '=F{1}*E{0}/100'.format(row+3,row+1), currency_format)

	
	worksheet1.write(row+3,3, 'Utilidad', format2)
	worksheet2.write(row+3,3, 'Utilidad', format2)
	worksheet1.write(row+3,4, float(k_aiu-1)*20, format7)
	worksheet2.write(row+3,4, float(k_aiu-1)*20, format7)
	worksheet1.write(row+3,5, '=F{1}*E{0}/100'.format(row+4,row+1), currency_format)
	worksheet2.write(row+3,5, '=F{1}*E{0}/100'.format(row+4,row+1), currency_format)
	
	worksheet1.write(row+4,3, 'Iva sobre Utilidad', format2)
	worksheet2.write(row+4,3, 'Iva sobre Utilidad', format2)
	worksheet1.write(row+4,4, 19, format7)
	worksheet2.write(row+4,4, 19, format7)
	worksheet1.write(row+4,5, '=F{1}*E{0}/100'.format(row+5,row+4), currency_format)
	worksheet2.write(row+4,5, '=F{1}*E{0}/100'.format(row+5,row+4), currency_format)

	worksheet1.merge_range('D{0}:E{0}'.format(row+6), 'TOTAL COSTOS INDIRECTOS DEL PROYECTO', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+6), 'TOTAL COSTOS INDIRECTOS DEL PROYECTO', merge_format_horizontal)
	worksheet1.write(row+5,5, '=Sum(F{0}:F{1})'.format(row+2,row+5), currency_format)
	worksheet2.write(row+5,5, '=Sum(F{0}:F{1})'.format(row+2,row+5), currency_format)

	worksheet1.merge_range('D{0}:E{0}'.format(row+7), 'SUBTOTAL COSTOS DIRECTOS + COSTOS INDIRECTOS', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+7), 'SUBTOTAL COSTOS DIRECTOS + COSTOS INDIRECTOS', merge_format_horizontal)
	worksheet1.write(row+6,5, '=F{0}+F{1}'.format(row+1,row+6), currency_format)
	worksheet2.write(row+6,5, '=F{0}+F{1}'.format(row+1,row+6), currency_format)


	worksheet1.write(row+7,3, 'VALOR DEL SERVICIO RETIE', format2)
	worksheet2.write(row+7,3, 'VALOR DEL SERVICIO RETIE', format2)
	worksheet1.write(row+7,4, 1.5, format7)
	worksheet2.write(row+7,4, 1.5, format7)
	worksheet1.write(row+7,5, '=F{1}*E{0}/100'.format(row+8,row+1), currency_format)
	worksheet2.write(row+7,5, '=F{1}*E{0}/100'.format(row+8,row+1), currency_format)


	worksheet1.merge_range('D{0}:E{0}'.format(row+10), 'COSTO TOTAL DEL PROYECTO', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+10), 'COSTO TOTAL DEL PROYECTO', merge_format_horizontal)
	worksheet1.write(row+9,5, '=F{0}+F{1}'.format(row+7,row+8), currency_format)
	worksheet2.write(row+9,5, '=F{0}+F{1}'.format(row+7,row+8), currency_format)


	worksheet1.merge_range('D{0}:E{0}'.format(row+12), 'INDICE DE COSTO POR CLIENTE', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+12), 'INDICE DE COSTO POR CLIENTE', merge_format_horizontal)
	worksheet1.write(row+11,5, '=F{0}/F{1}'.format(row+10,row+15), currency_format)
	worksheet2.write(row+11,5, '=F{0}/F{1}'.format(row+10,row+15), currency_format)


	worksheet1.merge_range('D{0}:E{0}'.format(row+13), 'INDICE DE COSTO POR kVA DE TRANSFORMADORES', merge_format_horizontal)
	worksheet2.merge_range('D{0}:E{0}'.format(row+13), 'INDICE DE COSTO POR kVA DE TRANSFORMADORES', merge_format_horizontal)
	worksheet1.write(row+12,5, '=F{0}/F{1}'.format(row+10,row+16), currency_format)
	worksheet2.write(row+12,5, '=F{0}/F{1}'.format(row+10,row+16), currency_format)



	worksheet1.merge_range('D{0}:E{0}'.format(row+15), 'Número de Clientes Beneficiados', merge_format)
	worksheet2.merge_range('D{0}:E{0}'.format(row+15), 'Número de Clientes Beneficiados', merge_format)
	worksheet1.write(row+14,5, 0, format2)
	worksheet2.write(row+14,5, 0, format2)


	worksheet1.merge_range('D{0}:E{0}'.format(row+16), 'kVA de Transformadores Instalado', merge_format)
	worksheet2.merge_range('D{0}:E{0}'.format(row+16), 'kVA de Transformadores Instalado', merge_format)
	worksheet1.write(row+15,5, 0, format2)
	worksheet2.write(row+15,5, 0, format2)

	
	col = 6
	for item in str_sumatorias_periodo:
		worksheet1.write(row,col, '', merge_format_horizontal)
		worksheet2.write(row,col, '', merge_format_horizontal)
		col+=1
		worksheet1.write(row,col, item, currency_format)
		worksheet2.write(row,col, item, currency_format)

		worksheet1.write(row+1,col, '=E{0}*{2}{1}/100'.format(row+2, row+1, xl_col_to_name(col)), currency_format)
		worksheet2.write(row+1,col, '=E{0}*{2}{1}/100'.format(row+2, row+1, xl_col_to_name(col)), currency_format)

		worksheet1.write(row+2,col, '=E{0}*{2}{1}/100'.format(row+3, row+1, xl_col_to_name(col)), currency_format)
		worksheet2.write(row+2,col, '=E{0}*{2}{1}/100'.format(row+3, row+1, xl_col_to_name(col)), currency_format)

		worksheet1.write(row+3,col, '=E{0}*{2}{1}/100'.format(row+4, row+1, xl_col_to_name(col)), currency_format)
		worksheet2.write(row+3,col, '=E{0}*{2}{1}/100'.format(row+4, row+1, xl_col_to_name(col)), currency_format)
		col+=1

	worksheet1.write(row,col, '', merge_format_horizontal)
	worksheet2.write(row,col, '', merge_format_horizontal)
	col+=1
	worksheet1.write(row,col, str_suma_subtotales_periodos, currency_format)
	worksheet2.write(row,col, str_suma_subtotales_periodos, currency_format)

	worksheet1.write(row+1,col, '=E{0}*{2}{1}/100'.format(row+2, row+1, xl_col_to_name(col)), currency_format)
	worksheet2.write(row+1,col, '=E{0}*{2}{1}/100'.format(row+2, row+1, xl_col_to_name(col)), currency_format)

	worksheet1.write(row+2,col, '=E{0}*{2}{1}/100'.format(row+3, row+1, xl_col_to_name(col)), currency_format)
	worksheet2.write(row+2,col, '=E{0}*{2}{1}/100'.format(row+3, row+1, xl_col_to_name(col)), currency_format)

	worksheet1.write(row+3,col, '=E{0}*{2}{1}/100'.format(row+4, row+1, xl_col_to_name(col)), currency_format)
	worksheet2.write(row+3,col, '=E{0}*{2}{1}/100'.format(row+4, row+1, xl_col_to_name(col)), currency_format)
	col+=1

	col = 6
	for semana in periodos:
		worksheet1.write(row+9,col, '', merge_format_horizontal)
		worksheet2.write(row+9,col, '', merge_format_horizontal)
		col+=1
		worksheet1.write(row+9,col, '=Sum({2}{0}:{2}{1})'.format(row+1,row+4,xl_col_to_name(col)), currency_format)
		worksheet2.write(row+9,col, '=Sum({2}{0}:{2}{1})'.format(row+1,row+4,xl_col_to_name(col)), currency_format)
		col+=1

	worksheet1.write(row+9,col, '', merge_format_horizontal)
	worksheet2.write(row+9,col, '', merge_format_horizontal)
	col+=1
	worksheet1.write(row+9,col, '=Sum({2}{0}:{2}{1})'.format(row+1,row+4,xl_col_to_name(col)), currency_format)
	worksheet2.write(row+9,col, '=Sum({2}{0}:{2}{1})'.format(row+1,row+4,xl_col_to_name(col)), currency_format)
	col+=1
	worksheet1.write(row+9,col, '=({1}{0}/F{0})*100'.format(row+10,xl_col_to_name(col-1)), format7)
	worksheet2.write(row+9,col, '=({1}{0}/F{0})*100'.format(row+10,xl_col_to_name(col-1)), format7)
	#-------------------------------------------------------------------------------------------
	#--------------------Hoja 2----- formatos de la hoja----------------------------------------
	

	worksheet2.set_column('A:A', 5)
	worksheet2.set_column('B:B', 12)
	worksheet2.set_column('C:C', 30)
	worksheet2.set_column('D:D', 12)
	worksheet2.set_column('E:E', 16)
	worksheet2.set_column('F:F', 16)

	worksheet2.merge_range('A1:A3', 'Merged Range', merge_format)
	worksheet2.merge_range('B1:B3', 'Merged Range', merge_format)
	worksheet2.merge_range('C1:C3', 'Merged Range', merge_format)
	worksheet2.merge_range('D1:D3', 'Merged Range', merge_format)
	worksheet2.merge_range('E1:E3', 'Merged Range', merge_format)
	worksheet2.merge_range('F1:F3', 'Merged Range', merge_format)

	worksheet2.write('A1', 'Item', format1)
	worksheet2.write('B1', 'Cod. UUCC', format1)
	worksheet2.write('C1', 'Descripcion UUCC', format1)
	worksheet2.write('D1', 'Cant.', format1)	
	worksheet2.write('E1', 'Valor Unitario', format1)
	worksheet2.write('F1', 'Valor TOTAL', format1)

	worksheet2.freeze_panes(3, 6)
	
	#-------------------------------------------------------------------------------------------
	#--------------------Hoja 3-----------------------------------------------------------------
	worksheet3.set_column('A:A', 2)
	worksheet3.set_column('B:B', 20)
	worksheet3.set_column('F:K', 15)
	worksheet3.set_column('M:P', 15)
	worksheet3.set_column('D:D', 20)
	worksheet3.set_row(11,72)
	

	cronogrma = Cronograma.objects.get(pk=cronograma_id)
	worksheet3.merge_range('B2:H2', 'SEGUIMIENTO DE AL PLAN DE INVERSIÓN', merge_format_horizontal)
	worksheet3.write('B4', 'CONTRATO:', merge_format_horizontal)
	worksheet3.write('B5', 'OBJETO:', merge_format_horizontal)
	worksheet3.write('B6', 'CONTRATISTA:', merge_format_horizontal)
	worksheet3.write('B7', 'CONTRATANTE:', merge_format_horizontal)	
	worksheet3.write('B8', 'SUPERVISOR', merge_format_horizontal)
	worksheet3.write('B9', 'VALOR INICIAL', merge_format_horizontal)

	# import pdb; pdb.set_trace()

	worksheet3.merge_range('C4:H4', cronogrma.proyecto.mcontrato.nombre, format2)
	worksheet3.merge_range('C5:H5', '', format2)
	worksheet3.merge_range('C6:H6', cronogrma.proyecto.contrato.filter(tipo_contrato__id=8).first().contratista.nombre if cronogrma.proyecto.contrato.filter(tipo_contrato__id=8) else cronogrma.proyecto.mcontrato.nombre, format2)
	worksheet3.merge_range('C7:H7', 'AIR-E S.A. E.S.P.' , format2)
	worksheet3.merge_range('C8:H8', '', format2)
	worksheet3.merge_range('C9:E9', "=+'PROGRAMACION SEMANAL'!F{0}".format(row+10), currency_format)	
	worksheet3.write('F9', 'ANTICIPO:', merge_format_horizontal)
	worksheet3.merge_range('G9:H9', '=+(C9*40%)/100', currency_format)


	worksheet3.merge_range('B11:E12', 'PERIODO', merge_format_horizontal)
	worksheet3.merge_range('F11:H11', 'INVERSIÓN SEMANAL EJECUTADA', merge_format_horizontal)	
	worksheet3.merge_range('I11:K11', 'INVERSIÓN SEMANAL PROGRAMADA', merge_format_horizontal)
	worksheet3.merge_range('L11:L12', 'PLAZO', merge_format_horizontal)
	worksheet3.merge_range('M11:O11', 'INVERSIÓN MENSUAL EJECUTADA', merge_format_horizontal)

	worksheet3.write('F12', 'OBRA EJECUTADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet3.write('G12', 'OBRA EJECUTADA ACUMULADA', merge_format_horizontal)
	worksheet3.write('H12', 'OBRA EJECUTADA', merge_format_horizontal)
	worksheet3.write('I12', 'OBRA PROGRAMADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet3.write('J12', 'OBRA PROGRAMADA ACUMULADA', merge_format_horizontal)
	worksheet3.write('K12', 'OBRA PROGRAMADA', merge_format_horizontal)


	worksheet3.write('M12', 'OBRA EJECUTADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet3.write('N12', 'OBRA EJECUTADA ACUMULADA', merge_format_horizontal)
	worksheet3.write('O12', 'OBRA EJECUTADA', merge_format_horizontal)

	worksheet3.merge_range('P11:P12', 'SEMANAL Vs MENSUAL', merge_format_horizontal)

	worksheet3.write('B13', 'Desde', merge_format_horizontal)
	worksheet3.write('C13', 'DIA', merge_format_horizontal)
	worksheet3.write('D13', 'Hasta', merge_format_horizontal)
	worksheet3.write('E13', 'DIA', merge_format_horizontal)
	worksheet3.write('F13', '($)', merge_format_horizontal)
	worksheet3.write('G13', '($)', merge_format_horizontal)

	worksheet3.write('H13', '(%)', merge_format_horizontal)
	worksheet3.write('I13', '($)', merge_format_horizontal)
	worksheet3.write('J13', '($)', merge_format_horizontal)
	worksheet3.write('K13', '(%)', merge_format_horizontal)
	worksheet3.write('L13', '(%)', merge_format_horizontal)

	worksheet3.write('M13', '($)', merge_format_horizontal)
	worksheet3.write('N13', '($)', merge_format_horizontal)
	worksheet3.write('O13', '(%)', merge_format_horizontal)
	worksheet3.write('P13', 'DIFERENCIA', merge_format_horizontal)

	col_3=7
	row_3=14
	row_save = 14
	for semana in periodos:
		fechaDesde = Str_fecha(semana.fechaDesde)
		fechaHasta = Str_fecha(semana.fechaHasta)

		worksheet3.write('B{0}'.format(row_3), semana.fechaDesde, format_date)
		worksheet3.write('C{0}'.format(row_3), '=Text(B{0},"DDD")'.format(row_3), format2)

		worksheet3.write('D{0}'.format(row_3), semana.fechaHasta, format_date)
		worksheet3.write('E{0}'.format(row_3), '=Text(D{0},"DDD")'.format(row_3), format2)


		worksheet3.write('F{0}'.format(row_3), "='CANTIDADES SEMANAL'!{0}${1}".format(xl_col_to_name(col_3),row+10), currency_format)

		if row_3==14:
			worksheet3.write('G{0}'.format(row_3), '=+F{0}'.format(row_3), currency_format)
		else:
			worksheet3.write('G{0}'.format(row_3), '=+F{0}+G{1}'.format(row_3,row_3-1), currency_format)

		worksheet3.write('H{0}'.format(row_3), '=+(G{0}/$C$9)*100'.format(row_3), format7)



		worksheet3.write('I{0}'.format(row_3), "='PROGRAMACION SEMANAL'!{0}${1}".format(xl_col_to_name(col_3),row+10), currency_format)

		if row_3==14:
			worksheet3.write('J{0}'.format(row_3), '=+I{0}'.format(row_3), currency_format)
		else:
			worksheet3.write('J{0}'.format(row_3), '=+I{0}+J{1}'.format(row_3,row_3-1), currency_format)

		worksheet3.write('K{0}'.format(row_3), '=+(J{0}/$C$9)*100'.format(row_3), format7)

		worksheet3.write('L{0}'.format(row_3), '=(D{0}-$B${0})/($D${1}-$B${0})*100'.format(row_3,13+len(periodos)), format7)
		
		if row_3<13+len(periodos):
			if semana.fechaHasta.month!=periodos[row_3-13].fechaHasta.month:
				worksheet3.write('M{0}'.format(row_3), '=+Sum(F{0}:F{1})'.format(row_3,row_save), currency_format)
				if row_save==14:
					worksheet3.write('N{0}'.format(row_3), '=+M{0}+N{1}'.format(row_3,row_save), currency_format)
				else:
					worksheet3.write('N{0}'.format(row_3), '=+M{0}+N{1}'.format(row_3,row_save-1), currency_format)

				worksheet3.write('O{0}'.format(row_3), '=+(N{0}/$C$9)*100'.format(row_3), format7)
				worksheet3.write('P{0}'.format(row_3), '=+O{0}-H{0}'.format(row_3), format7)
				
				row_save=row_3+1
		else:			
			worksheet3.write('M{0}'.format(row_3), '=+Sum(F{0}:F{1})'.format(row_3,row_save), currency_format)
			
			if row_save==14:
				worksheet3.write('N{0}'.format(row_3), '=+M{0}+N{1}'.format(row_3,row_save), currency_format)
			else:
				worksheet3.write('N{0}'.format(row_3), '=+M{0}+N{1}'.format(row_3,row_save-1), currency_format)

			worksheet3.write('O{0}'.format(row_3), '=+(N{0}/$C$9)*100'.format(row_3), format7)
			worksheet3.write('P{0}'.format(row_3), '=+O{0}-H{0}'.format(row_3), format7)
			
			row_save=row_3+1


		col_3+=2
		row_3+=1


	chart = workbook.add_chart({'type': 'line'})	

	chart.add_series({
	    'name':       'Programado',
	    'categories': "='SEGUIMIENTO SEMANAL'!D14:D{0}".format(row_3-1),
	    'values':     "='SEGUIMIENTO SEMANAL'!J14:J{0}".format(row_3-1),
	})

	chart.add_series({
	    'name':       'Ejecutado',
	    'categories': "='SEGUIMIENTO SEMANAL'!D14:D{0}".format(row_3-1),
	    'values':     "='SEGUIMIENTO SEMANAL'!G14:G{0}".format(row_3-1),
	})
	chart.set_title({ 'name': 'CURVA DE PROGRAMACIÓN VS EJECUCIÓN PRESUPUESTAL'})
	chart.set_size({'width': 1500, 'height': 576})
	chart.set_legend({'position': 'none'})

	worksheet3.insert_chart('B{0}'.format(row_3+2), chart)

	workbook.close()

	return response


@login_required
def descargar_plantilla_reporte_trabajo(request):	

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="plantilla_reporte.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Reporte de Trabajo')

	format0=workbook.add_format({'border':1})
	format0.set_bg_color('#D5D8DC')

	format1=workbook.add_format({'border':1,'font_size':13,'bold':True})
	format1.set_text_wrap()

	format2=workbook.add_format({'border':1})
	format2.set_text_wrap()

	format3=workbook.add_format({'border':1})
	format3.set_align('center')

	format4=workbook.add_format({'border':1})
	format4.set_align('center')
	format4.set_locked(True)


	format5=workbook.add_format({'border':1,'font_size':13,'bold':True})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_align('center')
	format5.set_text_wrap()


	format6=workbook.add_format({'font_size':13,'bold':True})
	format6.set_align('center')
	format6.set_text_wrap()
	
	format7=workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format7.set_align('center')
	format7.set_locked(True)
	# format7.set_text_wrap()



	format_red = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_red.set_bg_color('#E6B0AA')
	format_red.set_font_color('#C70039')
	format_red.set_align('center')
	format_red.set_locked(True)

	format_yellow = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_yellow.set_bg_color('#F9E79F')
	format_yellow.set_font_color('#FFC300')
	format_yellow.set_align('center')
	format_yellow.set_locked(True)

	format_green =  workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_green.set_bg_color('#ABEBC6')
	format_green.set_font_color('#28B463')
	format_green.set_align('center')
	format_green.set_locked(True)


	worksheet.set_column('A:A', 5)
	worksheet.set_column('B:B', 15)
	worksheet.set_column('C:C', 30)
	worksheet.set_column('D:F', 16)


	cronograma_id= request.GET['cronograma_id']	
	presupuesto_id= request.GET['presupuesto_id']	
	
	fechaDesde= request.GET['fechaDesde']
	fechaHasta= request.GET['fechaHasta']
	

	merge_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})

	merge_format_horizontal = workbook.add_format({
    'bold': 1,
    'border': 1})    
	merge_format_horizontal.set_bg_color('#D5D8DC')

	worksheet.merge_range('A1:A2', 'Merged Range', merge_format)
	worksheet.merge_range('B1:B2', 'Merged Range', merge_format)
	worksheet.merge_range('C1:C2', 'Merged Range', merge_format)
	worksheet.merge_range('D1:D2', 'Merged Range', merge_format)
	worksheet.merge_range('E1:E2', 'Merged Range', merge_format)
	worksheet.merge_range('F1:F2', 'Merged Range', merge_format)
	# worksheet.merge_range('G1:G3', 'Merged Range', merge_format)
	# worksheet.merge_range('H1:H3', 'Merged Range', merge_format)

	worksheet.write('A1', 'Id', format1)
	# worksheet.write('B1', 'Hitos', format1)
	# worksheet.write('B1', 'Actividad', format1)
	worksheet.write('B1', 'Cod. UUCC', format1)
	worksheet.write('C1', 'Descripcion UUCC', format1)
	worksheet.write('D1', 'Cant. presupuestada', format1)
	worksheet.write('E1', 'Cant. ejecutada', format1)
	worksheet.write('F1', '% Cant. Reportada', format1)

	worksheet.freeze_panes(2, 6)
	

	# import pdb; pdb.set_trace()
	col=6
	p=1 

	if fechaDesde and fechaHasta:
		periodos = ReporteTrabajo.objects.filter(reporteCerrado=False,sinAvance=False,periodoProgramacion__cronograma_id=cronograma_id,fechaReporte__gte=fechaDesde, fechaReporte__lte=fechaHasta).order_by('fechaReporte')
	
	elif fechaDesde and not fechaHasta:
		periodos = ReporteTrabajo.objects.filter(reporteCerrado=False,sinAvance=False,periodoProgramacion__cronograma_id=cronograma_id, fechaReporte__gte=fechaDesde).order_by('fechaReporte')
	
	elif not fechaDesde and fechaHasta:
		periodos = ReporteTrabajo.objects.filter(reporteCerrado=False,sinAvance=False,periodoProgramacion__cronograma_id=cronograma_id, fechaReporte__lte=fechaHasta).order_by('fechaReporte')
	
	else:
		periodos = ReporteTrabajo.objects.filter(reporteCerrado=False,sinAvance=False,periodoProgramacion__cronograma_id=cronograma_id).order_by('fechaReporte')
	

	for periodo in periodos:	
		# import pdb; pdb.set_trace()

		# worksheet.write(0,col,'Reporte '+str(p), format6)
		worksheet.write(0,col, periodo.fechaReporte, format5)		
		worksheet.write(1,col,'Cantidad', format6)

		p+=1
		col+=1

	worksheet.set_column(6,col, 16)

	

	uucc_actividades = FDetallePresupuesto.objects.filter(presupuesto_id=presupuesto_id).order_by('id')
	actividades = uucc_actividades.filter().values('actividad__id').distinct()
	actividades_cap = CEsquemaCapitulosActividadesG.objects.filter(pk__in=actividades).order_by('id')

	row=2

	for cap in actividades_cap:
		# import pdb; pdb.set_trace()
		worksheet.write(row,0, '', format0)
		# worksheet.merge_range('B{0}:{1}{0}'.format(row,xl_col_to_name(col)), 'Merged Range', merge_format)
		worksheet.merge_range('B{0}:F{0}'.format(row+1), cap.nombre, merge_format_horizontal)

		col=6
		for periodo in periodos:				
			worksheet.write(row,col,'', format0)
			col+=1

		row+=1

		uucc = uucc_actividades.filter(actividad__id=cap.id).order_by('id')

		
		for item_act in uucc:
			worksheet.write(row,0, item_act.id, format2)
			# padre = CEsquemaCapitulosActividadesG.objects.get(pk=item_act.actividad.padre)
			# worksheet.write(row,1, padre.nombre, format2)
			# worksheet.write(row,1, item_act.actividad.nombre, format2)
			worksheet.write(row,1, item_act.codigoUC, format2)
			worksheet.write(row,2, item_act.descripcionUC, format2)
			worksheet.write(row,3, round(item_act.cantidad, 4), format4)	
			worksheet.write(row,4, '=SUM(G{0}:{1}{0})'.format(str(row+1),xl_col_to_name(col-1)), format4)
			worksheet.write(row,5, '=(E{0}/D{0})*100'.format(str(row+1)),format7)

			col=6
			for periodo in periodos:	

				cantidad = DetalleReporteTrabajo.objects.filter(reporteTrabajo_id=periodo.id,detallePresupuesto_id=item_act.id).aggregate(Sum('cantidad'))
				
				if cantidad:
					worksheet.write(row,col,cantidad['cantidad__sum'], format4)
				else:
					worksheet.write(row,col,0, format4)

				col+=1

			row+=1
		

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': '>',
		                                        'value':    100,
		                                        'format':   format_red})

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': 'between',
		                                        'minimum': 0,
		                                        'maximum': 99.99999,
		                                        'format':   format_yellow})

			worksheet.conditional_format('F{0}:F{0}'.format(row), {
												'type':     'cell',
		                                        'criteria': '==',
		                                        'value':    100,
		                                        'format':   format_green})



	workbook.close()

	return response



@login_required
@transaction.atomic
def guardar_reporte_archivo(request):
	# import pdb; pdb.set_trace()
	try:		
		soporte= request.FILES['archivo']
		# esquema_id= request.POST['esquema_id']
		cronograma_id= request.POST['cronograma_id']
		id_presupuesto= request.POST['presupuesto_id']

		fechaDesde= request.POST['fechaDesde']
		fechaHasta= request.POST['fechaHasta']

		doc = openpyxl.load_workbook(soporte,data_only=True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)

		
		if fechaDesde and fechaHasta:
			periodos = ReporteTrabajo.objects.filter(
				reporteCerrado=False,
				sinAvance=False,
				periodoProgramacion__cronograma_id=cronograma_id,
				fechaReporte__gte=fechaDesde, fechaReporte__lte=fechaHasta).order_by('fechaReporte')
		
		elif fechaDesde and not fechaHasta:
			periodos = ReporteTrabajo.objects.filter(
				reporteCerrado=False,
				sinAvance=False,
				periodoProgramacion__cronograma_id=cronograma_id,
				fechaReporte__gte=fechaDesde).order_by('fechaReporte')
		
		elif fechaHasta and not fechaDesde:
			periodos = ReporteTrabajo.objects.filter(
				reporteCerrado=False,
				sinAvance=False,
				periodoProgramacion__cronograma_id=cronograma_id,
				fechaReporte__lte=fechaHasta).order_by('fechaReporte')
		
		else:
			periodos = ReporteTrabajo.objects.filter(
				reporteCerrado=False,
				sinAvance=False,
				periodoProgramacion__cronograma_id=cronograma_id).order_by('fechaReporte')
		
		cant_periodos = periodos.count()
		if cant_periodos==0:
			return JsonResponse({'message':'No se ha podido encontrar reportes de trabajo segun las fechas establecidas','success':'fail',
	 		'data':''},status=status.HTTP_400_BAD_REQUEST)

		i=0
		contador=hoja.max_row - 1
		numeroFila = 1
		
		mensaje_acumulado = ''
		if int(contador) > 0:
			sid = transaction.savepoint()
			

			validad_positiva_actividad_inicial = False
			cont_si=0

			if float(cant_periodos+6) != hoja.max_column:
				return JsonResponse({'message':'La cantidad de reportes del archivo no coincide con la cantidad de reportes encontrados segun las fechas establecidas','success':'fail',
						'data':''},status=status.HTTP_400_BAD_REQUEST)


			for fila in hoja.rows:
				if numeroFila==1:
					aux_col = 0
					for col in fila:
						if aux_col>5 and col.value:
							
							# date_doc = datetime.strptime(col.value, '%d/%m/%y %H:%M:%S')

							# print('Fecha archivo:', type(col.value))
							# print('Fecha del reporte :', type(periodos[aux_col-6].fechaReporte))

							if col.value.date()!=periodos[aux_col-6].fechaReporte:
								return JsonResponse({'message':'La fecha del reporte de la columna "'+xl_col_to_name(aux_col)+'" no coincide con la fecha establecida en el sistema, '+\
									'esto puede deberse a que las fechas no se encuentran ordenadas o que se ha agregado un nuevo reporte en SININ.<br>'+\
									'Sugerimos descargar nuevamente la plantilla y realizar las modificaciones pertinentes',
									'success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
						aux_col+=1

				if fila[0].value and numeroFila>3:

					if FDetallePresupuesto.objects.filter(id=fila[0].value,presupuesto__id=id_presupuesto,codigoUC=fila[1].value).count() == 0:
						return JsonResponse({
							'message':'El Id indicado en la fila No.' + str(numeroFila) + ' no coincide con la UUCC'+ \
							'. Sugerimos descargar nuevamente la plantilla para verificar el Id en esta fila',
							'success':'error',
							'data':''})

					try:
						# import pdb; pdb.set_trace()
						if  float(fila[5].value)>0:
							mensaje_acumulado = mensaje_acumulado + '<br>El % Cant. reportada en la fila No ' + str(numeroFila) + ' es superior al 100%'
					except Exception as e:
						return JsonResponse({
							'message':'Hubo un error en fila No.' + str(numeroFila) + \
							'. Fue ingresado una letra en lugar de un numero en las cantidades de algun reporte en dicha fila '
							'. Sugerimos corregir la planilla gestionada y volverla a subir en SININ',
							'success':'error',
							'data':''})

					




				numeroFila+= 1

			transaction.savepoint_commit(sid)
			
			ids_periodos = periodos.all().values('id')
			DetalleReporteTrabajo.objects.filter(reporteTrabajo_id__in=ids_periodos).delete()

			numeroFila = 1
			for fila in hoja.rows:
				if fila[0].value and numeroFila>3:

					c = 0
					for column in fila:
						if c>5 and column.value:
							if float(column.value)>0:
								# print('Periodo_id : '+str(periodos[c-6].id)+' | Detalle_id :'+str(fila[0].value))
								# import pdb; pdb.set_trace()
								detalle_programacion = DetalleReporteTrabajo(
									detallePresupuesto_id=fila[0].value,
									reporteTrabajo_id=periodos[c-6].id,
									cantidad=column.value)
								detalle_programacion.save()
						c+=1

				numeroFila+= 1
								
			
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'avanceObraLite.Reporte(carga_masiva)')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



def excel_seguimiento_contractual(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Avance de Obra - Seguimiento contractual.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})	

	format0=workbook.add_format({'border':1})
	format0.set_bg_color('#D5D8DC')

	format1=workbook.add_format({'border':1,'font_size':12,'bold':True})
	format1.set_text_wrap()

	format_date = workbook.add_format(
		{'border':1, 'num_format': 'dd/mm/yyyy',
		'align': 'center','valign': 'vcenter'})
	format_date.set_text_wrap()

	format2=workbook.add_format({'border':1,'font_size':8})
	format2.set_text_wrap()

	format3=workbook.add_format({'border':1,'font_size':8})
	format3.set_align('center')

	format4=workbook.add_format({'border':1,'font_size':8})
	format4.set_align('center')
	format4.set_locked(True)

	currency_format = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8})
	currency_format.set_align('center')

	format4_GREY=workbook.add_format({'border':1,'font_size':8,'bold':True})
	format4_GREY.set_align('center')
	format4_GREY.set_bg_color('#D5D8DC')
	format4_GREY.set_locked(True)
	
	currency_format_GREY = workbook.add_format({'num_format': '$#,##0.00', 'border':1,'font_size':8,'bold':True})
	currency_format_GREY.set_align('center')
	currency_format_GREY.set_bg_color('#D5D8DC')
	
	format5=workbook.add_format({'border':1,'font_size':8})
	format5.set_num_format('yyyy-mm-dd')
	format5.set_text_wrap()


	format6=workbook.add_format({'font_size':12,'bold':True})
	format6.set_align('center')
	format6.set_text_wrap()
	
	format7=workbook.add_format({'border':1, 'num_format': '0.00"%"','font_size':8})
	format7.set_align('center')
	format7.set_locked(True)
	# format7.set_text_wrap()



	format_red = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_red.set_bg_color('#E6B0AA')
	format_red.set_font_color('#C70039')
	format_red.set_align('center')
	format_red.set_locked(True)

	

	format_yellow = workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_yellow.set_bg_color('#F9E79F')
	format_yellow.set_font_color('#FFC300')
	format_yellow.set_align('center')
	format_yellow.set_locked(True)

	format_green =  workbook.add_format({'border':1, 'num_format': '0.00"%"'})
	format_green.set_bg_color('#ABEBC6')
	format_green.set_font_color('#28B463')
	format_green.set_align('center')
	format_green.set_locked(True)

	merge_format = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})
	merge_format_horizontal = workbook.add_format({
    'bold': 1,
    'border': 1,
    'align': 'center',
    'valign': 'vcenter'})  
	merge_format_horizontal.set_text_wrap()
	merge_format_horizontal.set_align('center')

	worksheet1 = workbook.add_worksheet('Seguimiento contractual')

	worksheet1.set_column('A:A', 5)
	worksheet1.set_column('B:B', 30)
	worksheet1.set_column('C:S', 15)


	worksheet1.merge_range('B2:B4', 'ACTIVIDAD', merge_format_horizontal)
	worksheet1.merge_range('C2:C4', 'PESO ACTIVIDAD', merge_format_horizontal)

	worksheet1.merge_range('D1:G1', 'PLAZO', merge_format_horizontal)

	worksheet1.merge_range('D2:D4', 'FECHA INICIO', merge_format_horizontal)
	worksheet1.merge_range('E2:E4', 'FECHA FIN', merge_format_horizontal)
	worksheet1.merge_range('F2:F4', 'DIAS DE TRABAJO PROYECTADO', merge_format_horizontal)
	worksheet1.merge_range('G2:G4', 'DIAS CURSADOS', merge_format_horizontal)
	worksheet1.merge_range('H2:H4', 'CANT. SEGÚN DISEÑO', merge_format_horizontal)
	
	#----------------------------------------------------------------------------------
	worksheet1.merge_range('I1:J1', 'PERIODO DE ESTUDIO', merge_format_horizontal)

	worksheet1.merge_range('I2:I4', 'FECHA INICIO', merge_format_horizontal)
	worksheet1.merge_range('J2:J4', 'FECHA FIN', merge_format_horizontal)
	#----------------------------------------------------------------------------------
	worksheet1.merge_range('K1:M1', 'OBRA EJECUTADA', merge_format_horizontal)

	worksheet1.merge_range('K2:K4', 'OBRA EJECUTADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet1.merge_range('L2:L4', 'OBRA EJECUTADA ACUMULADA', merge_format_horizontal)
	worksheet1.merge_range('M2:M4', '% OBRA EJECUTADA ACUMULADA', merge_format_horizontal)
	#----------------------------------------------------------------------------------
	worksheet1.merge_range('N1:P1', 'OBRA PROGRAMADA', merge_format_horizontal)

	worksheet1.merge_range('N2:N4', 'OBRA PROGRAMADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet1.merge_range('O2:O4', 'OBRA PROGRAMADA ACUMULADA', merge_format_horizontal)
	worksheet1.merge_range('P2:P4', '% OBRA PROGRAMADA ACUMULADA', merge_format_horizontal)
	#----------------------------------------------------------------------------------
	worksheet1.merge_range('Q1:S1', 'DESVIACIÓN DE OBRA', merge_format_horizontal)

	worksheet1.merge_range('Q2:Q4', 'OBRA DESVIADA DURANTE EL PERIODO', merge_format_horizontal)
	worksheet1.merge_range('R2:R4', 'OBRA DESVIADA ACUMULADA', merge_format_horizontal)
	worksheet1.merge_range('S2:S4', '% OBRA DESVIADA ACUMULADA', merge_format_horizontal)

	cronograma = Cronograma.objects.filter(proyecto__id=request.GET['id_proyecto'])
	if cronograma:
		cronograma = cronograma.first()

	qset = (~Q(actividad__nombre='No aplica')) & (Q(presupuesto__cronograma__proyecto__id=request.GET['id_proyecto']))
	actividades = FDetallePresupuesto.objects.filter(qset).values(
		'actividad__id',
		'actividad__nombre',
		'actividad__peso').annotate(total=Sum('cantidad'))

	fecha_actual = datetime.today().date()


	# col = 
	row = 4
	for act in actividades:
		
		ejecutar = DetalleReporteTrabajo.objects.filter(
			reporteTrabajo__fechaReporte__lte=fecha_actual,
			detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))

		programar= DetallePeriodoProgramacion.objects.filter(
			periodoProgramacion__fechaDesde__lte=fecha_actual,
			detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))


		worksheet1.write(row,1,act['actividad__nombre'], format2)
		worksheet1.write(row,2,act['actividad__peso'], format7)		
		worksheet1.write(row,3,cronograma.fechaInicio, format_date)
		worksheet1.write(row,4,cronograma.fechaFinal, format_date)
		worksheet1.write(row,5,'=(E{0}-D{0}+1)-ROUNDDOWN((E{0}-D{0}+1)/7,0)'.format(row+1), format2)		
		worksheet1.write(row,6,'=IF(TODAY()<D{0},0,IF(AND(D{0}<=TODAY(),TODAY()<=E{0}),(TODAY()-D{0}+1)-ROUNDDOWN((TODAY()-D{0}+1)/7,0),F{0}))'.format(row+1), format2)
		
		worksheet1.write(row,7,act['total'] if act['total'] else 0,format2)

		worksheet1.write(row,8,request.GET['fecha_inicio'],format_date)
		worksheet1.write(row,9,request.GET['fecha_fin'],format_date)

		worksheet1.write(row,11,ejecutar['cantidad__sum'] if ejecutar['cantidad__sum'] else 0,format2)
		worksheet1.write(row,14,programar['cantidad__sum'] if programar['cantidad__sum'] else 0,format2)


		periodo_ejecutar = DetalleReporteTrabajo.objects.filter(
			reporteTrabajo__fechaReporte__gte=request.GET['fecha_inicio'],
			reporteTrabajo__fechaReporte__lte=request.GET['fecha_fin'],
			detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))

		periodo_programar= DetallePeriodoProgramacion.objects.filter(
			periodoProgramacion__fechaDesde__gte=request.GET['fecha_inicio'],
			periodoProgramacion__fechaHasta__lte=request.GET['fecha_fin'],
			detallePresupuesto__actividad__id=int(act['actividad__id'])).aggregate(Sum('cantidad'))


		worksheet1.write(row,10,periodo_ejecutar['cantidad__sum'] if periodo_ejecutar['cantidad__sum'] else 0,format2)
		worksheet1.write(row,13,periodo_programar['cantidad__sum'] if periodo_programar['cantidad__sum'] else 0,format2)


		worksheet1.write(row,12,'=IFERROR(($L{0}/$H{0})*$C{0},0)'.format(row+1), format7)	
		worksheet1.write(row,15,'=IFERROR(($O{0}/$H{0})*$C{0},0)'.format(row+1), format7)	

		worksheet1.write(row,16,'=+N{0}-K{0}'.format(row+1),format2)
		worksheet1.write(row,17,'=+O{0}-L{0}'.format(row+1),format2)
		worksheet1.write(row,18,'=-P{0}+M{0}'.format(row+1), format7)

		row+=1
	
	worksheet1.write(row,1,'AVANCE DE OBRA GENERAL', format2)
	worksheet1.write(row,2,'=Sum(C{0}:C{1})'.format(5,row), format7)
	worksheet1.write(row,3,'=Min(D{0}:D{1})'.format(5,row), format_date)
	worksheet1.write(row,4,'=Max(E{0}:E{1})'.format(5,row), format_date)

	worksheet1.merge_range('K{0}:L{0}'.format(row+1), ' % OBRA EJECUTADA', merge_format_horizontal)
	worksheet1.merge_range('N{0}:O{0}'.format(row+1), ' % OBRA PROGRAMADA', merge_format_horizontal)
	worksheet1.merge_range('Q{0}:R{0}'.format(row+1), ' % DESVIACIÓN DE OBRA ', merge_format_horizontal)

	worksheet1.write(row,12,'=Sum(M{0}:M{1})'.format(5,row), format7)
	worksheet1.write(row,15,'=Sum(P{0}:P{1})'.format(5,row), format7)
	worksheet1.write(row,18,'=Sum(S{0}:S{1})'.format(5,row), format7)


	workbook.close()

	return response


	
	

	
