from django.shortcuts import render,redirect
#,render_to_response
from django.urls import reverse
from rest_framework import viewsets, serializers
from django.db.models import Q
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import json
import xlsxwriter
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from tipo.models import Tipo
from estado.models import Estado
from tipo.views import TipoSerializer
from estado.views import EstadoSerializer
from contrato.models import Contrato
from contrato.views import ContratoSerializer
from proyecto.models import Proyecto
from empresa.models import Empresa
from usuario.models import Usuario,Persona
from usuario.views import UsuarioSerializer
from proyecto.views import ProyectoSerializer
from .enum import enumEstados,enumTipo
from .models import ASolicita,BUnidadConstructiva,CCambio,DMaterial,EUUCCMaterial,FCambioProyecto,GSoporte
from parametrizacion.models import Municipio , Departamento
from parametrizacion.views import  MunicipioSerializer , DepartamentoSerializer
from logs.models import Logs,Acciones
from django.db import transaction,connection
from django.db.models.deletion import ProtectedError
from django.contrib.auth.decorators import login_required
from django.db.models import Max
import openpyxl
from rest_framework.decorators import api_view, throttle_classes
from coasmedas.functions import functions
from adminMail.models import Mensaje
from adminMail.tasks import sendAsyncMail
from django.conf import settings


# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre')

# Serializer de proyecto
class ProyectoLiteSerializer(serializers.HyperlinkedModelSerializer):

	mcontrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	mcontrato=ContratoLiteSerializer(read_only=True)

	municipio = MunicipioSerializer(read_only = True)

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre','mcontrato_id','mcontrato','municipio')


#Api rest para solictia de control de cambios
class SolicitaSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = ASolicita
		fields=('id','nombre')


class SolicitaViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de quien solicita
	"""
	model=ASolicita
	queryset = model.objects.all()
	serializer_class = SolicitaSerializer
	nombre_modulo='control_cambios.solicita'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SolicitaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			qset=''

			if dato:
				qset =Q(nombre__icontains=dato)
						
			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SolicitaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():

					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.solicita',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SolicitaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.solicita',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para solicita de control de cambios



#Api rest para unidad constructiva de control de cambios
class UnidadConstructivaSerializer(serializers.HyperlinkedModelSerializer):

	proyecto_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all(),allow_null=True)
	proyecto=ProyectoLiteSerializer(read_only=True)

	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoLiteSerializer(read_only=True)

	class Meta:
		model = BUnidadConstructiva
		fields=('id','contrato','contrato_id','proyecto',
				'proyecto_id','codigo','descripcion',
				'valor_mano_obra','valor_materiales')


class UnidadConstructivaViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista las unidades constructivas del modulo de control de cambios.
	"""
	model=BUnidadConstructiva
	queryset = model.objects.all()
	serializer_class = UnidadConstructivaSerializer
	nombre_modulo='control_cambios.unidadConstructiva'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(UnidadConstructivaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			mcontrato= self.request.query_params.get('mcontrato',None)
			proyecto= self.request.query_params.get('proyecto',None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			id_empresa = request.user.usuario.empresa.id

			qset=''

			if (dato or mcontrato or proyecto or id_empresa):

				qset = Q(contrato__empresacontrato__empresa=id_empresa)


				if mcontrato and int(mcontrato)>0:
					qset = qset &(Q(contrato__id=mcontrato))

				if proyecto and int(proyecto)>0:
					qset = qset &(Q(proyecto__id=proyecto))

				if dato:
					qset = qset & (Q(codigo__icontains=dato) | Q(descripcion__icontains=dato))

			if qset != '':
				
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				unidad = BUnidadConstructiva.objects.filter(codigo=request.DATA['codigo'])

				if unidad:
					return Response({'message':'Ya existe el codigo de la unidad constructiva registrada','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


				serializer = UnidadConstructivaSerializer(data=request.DATA,context={'request': request})

				if request.DATA['proyecto_id']=='':
					request.DATA['proyecto_id']=None
				
				if serializer.is_valid():

					serializer.save(contrato_id=request.DATA['contrato_id'],proyecto_id=request.DATA['proyecto_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.UnidadConstructiva',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = UnidadConstructivaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if request.DATA['proyecto_id']=='':
					request.DATA['proyecto_id']=None

				if serializer.is_valid():
					serializer.save(contrato_id=request.DATA['contrato_id'],proyecto_id=request.DATA['proyecto_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.UnidadConstructiva',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para unidad constructiva



#Api rest para cambios de control de cambios
class CambioSerializer(serializers.HyperlinkedModelSerializer):

	proyecto_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Proyecto.objects.all())
	proyecto=ProyectoLiteSerializer(read_only=True)

	tipo = TipoSerializer(read_only=True)
	tipo_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='control_cambios'),allow_null=True,default=None)

	solicita_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=ASolicita.objects.all())
	solicita=SolicitaSerializer(read_only=True)

	usuario_revisa_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all(), allow_null=True)
	usuario_revisa=UsuarioSerializer(read_only=True)

	usuario=UsuarioSerializer(read_only=True)

	#solicitud_enviada = serializers.BooleanField(default=False)

	#cantidad_cambio=serializers.IntegerField(read_only=True)

	class Meta:
		model = CCambio
		fields=('id','proyecto',
				'proyecto_id','tipo','tipo_id','solicita_id','solicita',
				'usuario_revisa_id','usuario_revisa','usuario','fecha','numero_cambio',
				'motivo','cantidad_cambio','maximo_id_cambio','estado_cambio_proyecto')


class CambioViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de los cambios del modulo de control de cambios.
	"""
	model=CCambio
	queryset = model.objects.all()
	serializer_class = CambioSerializer
	nombre_modulo='control_cambios.cambio'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CambioViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			mcontrato= self.request.query_params.get('mcontrato',None)
			proyecto= self.request.query_params.get('proyecto',None)
			tipo= self.request.query_params.get('tipo',None)		
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			id_empresa = request.user.usuario.empresa.id
			usuario_id = self.request.query_params.get('usuario', None)

			qset=''

			if (dato or mcontrato or proyecto or id_empresa or tipo):

				qset = Q(proyecto__mcontrato__empresacontrato__empresa=id_empresa)


				if usuario_id and int(usuario_id)>0:
					qset = Q(usuario__id=usuario_id)

				if tipo and int(tipo)>0:
					qset = qset &(Q(tipo__id=tipo))


				if mcontrato and int(mcontrato)>0:
					qset = qset &(Q(proyecto__mcontrato__id=mcontrato))

				if proyecto and int(proyecto)>0:
					qset = qset &(Q(proyecto__id=proyecto))

				if dato:
					qset = qset & (Q(proyecto__nombre__icontains=dato) 
						| Q(usuario_revisa__persona__nombres__icontains=dato) 
						| Q(usuario_revisa__persona__apellidos__icontains=dato)
						| Q(solicita__nombre__icontains=dato))

			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if request.DATA['fecha']=='':
					request.DATA['fecha']=None

				if request.DATA['usuario_revisa_id'] is None or request.DATA['usuario_revisa_id']=='':
					request.DATA['usuario_revisa_id']=None 

				serializer = CambioSerializer(data=request.DATA,context={'request': request})

				guardarCambioProyecto=CCambio.objects.filter(proyecto__id=request.DATA['proyecto_id']).aggregate(Max('id'))

				listaCambioProyecto=FCambioProyecto.objects.filter(cambio_id=guardarCambioProyecto['id__max'])

				request.DATA['numero_cambio']=0
				tipo = enumTipo()				
				if int(request.DATA['tipo_id'])==tipo.Cambio:					
					cambio=CCambio.objects.filter(proyecto_id=request.DATA['proyecto_id'], tipo_id=tipo.Cambio).aggregate(Max('numero_cambio'))					
					request.DATA['numero_cambio']=cambio['numero_cambio__max']+1 if cambio['numero_cambio__max'] is not None else 1
				
				if serializer.is_valid():

					serializer.save(proyecto_id=request.DATA['proyecto_id'],tipo_id=request.DATA['tipo_id'],
									usuario_id=request.user.usuario.id,solicita_id=request.DATA['solicita_id'],
									usuario_revisa_id=request.DATA['usuario_revisa_id'])

					for item in list(listaCambioProyecto):

						logs_model=FCambioProyecto(cantidad=item.cantidad,comentario=item.comentario,cambio_id=serializer.data['id'],estado_id=item.estado.id,uucc_id=item.uucc.id)
						logs_model.save()


					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.cambios',id_manipulado=serializer.data['id'])
					logs_model.save()

					notificacion_cambio(request.DATA['proyecto_id'],request.DATA['usuario_revisa_id'])

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				serializer = CambioSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():

					serializer.save(proyecto_id=request.DATA['proyecto_id'],tipo_id=request.DATA['tipo_id'],
									usuario_id=request.DATA['usuario_id'],solicita_id=request.DATA['solicita_id'],
									usuario_revisa_id=request.DATA['usuario_revisa_id'] if request.DATA['usuario_revisa_id'] is not None else None)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.cambios',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para cambio



#Api rest para material de control de cambios
class MaterialSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = DMaterial
		fields=('id','codigo','descripcion','unidad')


class MaterialViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de materiales del modulo de control de cambios.
	"""
	model=DMaterial
	queryset = model.objects.all()
	serializer_class = MaterialSerializer
	nombre_modulo='control_cambios.material'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(MaterialViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			qset=''

			if dato:
				qset = qset & (Q(codigo__icontains=dato) | Q(descripcion__icontains=dato) | Q(unidad__icontains=dato))


			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = MaterialSerializer(data=request.DATA,context={'request': request})
				
				if serializer.is_valid():

					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.material',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = MaterialSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.material',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para material



#Api rest para uucc material de control de cambios
class UUCCMaterialSerializer(serializers.HyperlinkedModelSerializer):

	uucc_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=BUnidadConstructiva.objects.all())
	uucc=UnidadConstructivaSerializer(read_only=True)

	material_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=DMaterial.objects.all())
	material=MaterialSerializer(read_only=True)

	class Meta:
		model = EUUCCMaterial
		fields=('id','uucc_id','uucc','material_id','material','cantidad')


class UUCCMaterialViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de uucc materiales del modulo de control de cambios.
	"""
	model=EUUCCMaterial
	queryset = model.objects.all()
	serializer_class = UUCCMaterialSerializer
	nombre_modulo='control_cambios.uuccmaterial'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(UUCCMaterialViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			qset=''

			if dato:
				qset = qset & (Q(uucc__proyecto__nombre__icontains=dato) | Q(uucc__contrato__nombre__icontains=dato) | Q(material__codigo__icontains=dato))


			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = UUCCMaterialSerializer(data=request.DATA,context={'request': request})
				
				if serializer.is_valid():

					serializer.save(uucc_id=request.DATA['uucc_id'],material_id=request.DATA['material_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.UUCCMaterial',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					#print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				serializer = UUCCMaterialSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(uucc_id=request.DATA['uucc_id'],material_id=request.DATA['material_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.UUCCMaterial',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para uuccmaterial



#Api rest para cambio proyecto de control de cambios
class CambioProyectoSerializer(serializers.HyperlinkedModelSerializer):

	uucc_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=BUnidadConstructiva.objects.all())
	uucc=UnidadConstructivaSerializer(read_only=True)

	cambio_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=CCambio.objects.all())
	cambio=CambioSerializer(read_only=True)

	estado = EstadoSerializer(read_only=True)
	estado_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Estado.objects.filter(app='EstadoControlCambio'))

	class Meta:
		model = FCambioProyecto
		fields=('id','uucc_id','uucc','cambio_id','cambio','estado_id','estado','cantidad','comentario')


class CambioProyectoViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de cambio proyecto del modulo de control de cambios.
	"""
	model=FCambioProyecto
	queryset = model.objects.all()
	serializer_class = CambioProyectoSerializer
	nombre_modulo='control_cambios.cambioproyecto'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CambioProyectoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			cambio = self.request.query_params.get('cambio_id', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			qset=''
			
			if (dato or cambio):

				qset = Q(cambio__id=cambio)

				if dato:
					qset = qset & (Q(uucc__proyecto__nombre__icontains=dato) | Q(uucc__contrato__nombre__icontains=dato)
						| Q(uucc__codigo__icontains=dato) | Q(uucc__descripcion__icontains=dato))

			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CambioProyectoSerializer(data=request.DATA,context={'request': request})
				estado=enumEstados()
				request.DATA['estado_id']=estado.Pendiente

				if serializer.is_valid():

					cambio_proyecto = FCambioProyecto.objects.filter(uucc_id=request.DATA['uucc_id'],cambio_id=request.DATA['cambio_id']).first()

					if cambio_proyecto:

						return Response({'message':'La unidad constructiva ya existe!','success':'fail',
										'data':''},status=status.HTTP_400_BAD_REQUEST)

					else:

						serializer.save(uucc_id=request.DATA['uucc_id'],cambio_id=request.DATA['cambio_id'],estado_id=enumEstados.Pendiente)

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.CambioProyecto',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)

						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
							'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				serializer = CambioProyectoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(uucc_id=request.DATA['uucc_id'],cambio_id=request.DATA['cambio_id'],estado_id=request.DATA['estado_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.CambioProyecto',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para cambio proyecto



#Api rest para soporte de control de cambios
class SoporteSerializer(serializers.HyperlinkedModelSerializer):

	cambio_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=CCambio.objects.all())
	cambio=CambioSerializer(read_only=True)

	class Meta:
		model = GSoporte
		fields=('id','cambio_id','cambio','nombre','ruta')


class SoporteViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de soporte del modulo de control de cambios.
	"""
	model=GSoporte
	queryset = model.objects.all()
	serializer_class = SoporteSerializer
	nombre_modulo='control_cambios.soporte'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SoporteViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			cambio = self.request.query_params.get('cambio_id', None)

			qset=''

			qset = Q(cambio_id=cambio)

			if dato:
				qset = qset & (Q(cambio__proyecto__nombre__icontains=dato))


			if qset != '':
				queryset = self.model.objects.filter(qset)	


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SoporteSerializer(data=request.DATA,context={'request': request})

				if self.request.FILES.get('ruta') is not None:

					serializer.save(ruta=self.request.FILES.get('ruta'),cambio_id=request.DATA['cambio_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.soporte',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

				else:

					serializer.save(cambio_id=request.DATA['cambio_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='control_cambios.soporte',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

				return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
					'data':serializer.data},status=status.HTTP_201_CREATED)


			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()

				serializer = SoporteSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(cambio_id=request.DATA['cambio_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.soporte',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para cambio proyecto


#eliminar las unidades constructivas
@transaction.atomic
def eliminar_unidad_constructiva(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			BUnidadConstructiva.objects.filter(id=item['id']).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='control_cambios.UnidadConstructiva',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})		



#eliminar los administrar cambios uucc
@transaction.atomic
def eliminar_cambio_uucc(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			CCambio.objects.filter(id=item['id']).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='control_cambios.Cambios',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#eliminar los cambios proyecto
@transaction.atomic
def eliminar_cambio_proyecto(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			FCambioProyecto.objects.filter(id=item['id']).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='control_cambios.cambio_proyecto',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




#eliminar los administrar cambios uucc
@transaction.atomic
def eliminar_soporte_uucc_cambio(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			GSoporte.objects.filter(id=item['id']).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='control_cambios.soporte',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

#Actualiza la cantidad en cambio proyecto
@transaction.atomic
def actualizar_cantidad_cambio_proyecto(request):

	sid = transaction.savepoint()
	try:
		
		lista=request.POST['_content']
		respuesta= json.loads(lista)		
		id = respuesta['id']
		
		if respuesta['cantidad']:
			cantidad = respuesta['cantidad']

		else:
			cantidad=0

		cambio_proyecto=FCambioProyecto.objects.get(pk=id)
		cambio_proyecto.cantidad=cantidad
		cambio_proyecto.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.cambio_proyecto',id_manipulado=id)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		




#actualizar vista agregar uuc
@transaction.atomic
def actualizar_todo_agregar_uucc(request):

	sid = transaction.savepoint()

	mensaje=''
	status=''
	res=''

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		envio_correo= respuesta['envio_correo']
		usuario_revisa_id= respuesta['usuario_revisa']
		proyecto_id= respuesta['proyecto_id']

		for item in respuesta['lista']:

			object_cambio_proyecto=FCambioProyecto.objects.get(pk=item['id'])

			object_cambio_proyecto.cantidad=item['cantidad']
			object_cambio_proyecto.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.cambio_proyecto',id_manipulado=item['id'])
			logs_model.save()

			transaction.savepoint_commit(sid)


		if envio_correo==1:

			usuario= Usuario.objects.get(id=usuario_revisa_id)
			persona = Persona.objects.get(id=usuario.persona.id)
			proyecto= Proyecto.objects.get(id=proyecto_id)

			#inicio del codigo del envio de correo
			contenido='<h3>SININ - Sistema Integral de informacion</h3>'
			contenido = contenido + 'Estimado(a)'+persona.nombres +' '+ persona.apellidos +'<br/><br/>'
			contenido = contenido + 'Nos permitimos informale que se ha registrado un cambio en las cantidades de unidades constructivas del proyecto '+ proyecto.nombre +' en el macro-contrato '+ proyecto.mcontrato.nombre+'. Dicho cambio fue registrado por el usuario '+ persona.nombres +' '+ persona.apellidos+' de la empresa '
			contenido = contenido + 'No responder este mensaje, este correo es de uso informativo exclusivamente,<br/><br/>'
			contenido = contenido + 'Soporte SININ<br/>soporte@sinin.co'
			mail = Mensaje(
				remitente=settings.REMITENTE,
				destinatario=persona.correo,
				asunto='informativo de SININ',
				contenido=contenido,
				appLabel='Usuario',
				)			
			mail.save()
			res=sendAsyncMail(mail)
			mensaje='Notificacion de cambios en cantidades de unidades constructivas en obra.'
			status='ok'


		return JsonResponse({'message':'Los datos se actualizaron correctamente','success':'ok',
				'data':''})		
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#exportar las unidades constructivas
def export_excel_unidades_constructivas(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Unidades constructiva.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Unidades constructiva')
	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format2=workbook.add_format({'border':0})

	row=1
	col=0

	cursor = connection.cursor()

	id_empresa = request.user.usuario.empresa.id
	qset=None

	qset = Q(contrato__empresacontrato__empresa=id_empresa)
	
			
	unidadConstructiva = BUnidadConstructiva.objects.filter(qset)

	worksheet.write('A1', 'Codigo', format1)
	worksheet.write('B1', 'Descripcion', format1)
	worksheet.write('C1', 'Valor mano de obra', format1)
	worksheet.write('D1', 'Valor materiales', format1)

	for c in unidadConstructiva:

		worksheet.write(row, col,c.codigo,format2)
		worksheet.write(row, col+1,c.descripcion,format2)
		worksheet.write(row, col+2,c.valor_mano_obra,format2)
		worksheet.write(row, col+3,c.valor_materiales,format2)
		
		row +=1

	workbook.close()

	return response
    #return response

#Cargar soporte de la vista agregar uucc cambio
def cargar_soporte_uucc_cambio(request):

	try:

		if request.POST['cambio_id']:

			modelo=GSoporte(cambio_id=request.POST['cambio_id'],nombre=request.POST['nombre_documento'],ruta=request.FILES['archivo'])
			modelo.save()
		
			return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})

		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','success':'ok','data':''})
		
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



#funcion utilizada para cargar masivamente las cantidades en la vista agregar uucc
@transaction.atomic
def cargar_excel_masivo_cambio(request):

	try:
		cambio_id= request.POST['cambio_id']
		proyecto_id=request.POST['proyecto_id']
		contrato_id=request.POST['contrato_id']
		soporte= request.FILES['archivo']

		doc = openpyxl.load_workbook(soporte)

		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)

		index=0
		codigo_uucc=''
		cantidad=''
		qset=''

		# print hoja.rows
		for fila in hoja.rows:

			if index>=1:				
				codigo_uucc =fila[0].value
				cantidad =fila[1].value

				if contrato_id and int(contrato_id)>0 :		
					qset = Q(contrato__id=contrato_id)

				# if proyecto_id and int(proyecto_id)>0 :		
				# 	qset = Q(proyecto__id=proyecto_id)

				if codigo_uucc:
					qset = qset & (Q(codigo__icontains=codigo_uucc))		

				uucc = BUnidadConstructiva.objects.filter(qset).first()

				cambio_proyecto = FCambioProyecto.objects.filter(uucc__id=uucc.id).first()


				if cambio_proyecto:
					actualiza_cambio=FCambioProyecto.objects.get(pk=cambio_proyecto.id)
					actualiza_cambio.cantidad=cantidad
					actualiza_cambio.save()

				else:

					modelo=FCambioProyecto(uucc_id=uucc.id,estado_id=enumEstados.Pendiente,cambio_id=cambio_id,cantidad=cantidad)
					modelo.save()

			index=index+1

		return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response



#funcion para traer el listado de comparar
@api_view(['GET'])
def lista_comparar(request):
	if request.method == 'GET':	
		cursor = connection.cursor()
		try:

			dato = request.GET.get('dato', None)
			proyecto_id = request.GET['proyecto_id']

			cursor.callproc('[dbo].[control_cambios_consultar_comparar_cambios]', [proyecto_id, dato,])			
			columns = cursor.description

			if columns: 
				lista = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]
				encabezados = [i[0] for i in columns]

				cursor.callproc('[dbo].[control_cambios_consultar_columnas_comparar_cambios]', [proyecto_id])	
				columns = cursor.description 
				soportes = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]

				return Response({'message':'','success':'ok','data':{'lista':lista, 'encabezados':encabezados, 'soportes':soportes}})

			return Response({'message':'No se encontraron registros','success':'error','data':None})
				
		except Exception as e:

			#print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)	
		finally:
			cursor.close()


#exporta vista comparar
@api_view(['GET'])
def export_reporte_comparar(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Comparacion_cambios.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Comparacion_cambios')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})

	
	cursor = connection.cursor()

	dato = request.GET.get('dato', None)
	proyecto_id = request.GET['proyecto_id']

	cursor.callproc('[dbo].[control_cambios_consultar_comparar_cambios]', [proyecto_id, dato,])			
	columns = cursor.description 
	lista = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]
	encabezados = [i[0] for i in columns]

	c=0

	for column in encabezados:		
		worksheet.write(0, c, column, format1)
		c += 1

	row=1	
	for r in lista:
		col=0
		for column in encabezados:
			worksheet.write(row, col,r[column],format2)	
			col+=1
		row +=1


	workbook.close()

	return response
    #return response



#funcion utilizada para cargar masivamente de la vista carga constructivas
@transaction.atomic
def cargar_excel_unidades_constructivas(request):

	try:
		contrato_id= request.POST['contrato']
		soporte= request.FILES['archivo']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)

		index=0
		codigo=''
		descripcion=''
		valor_mano_obra=''
		valor_material=''

		# print hoja.rows
		for fila in hoja.rows:

			if fila[0].value is not None:

				if index>=1:

					unidades_constructivas = BUnidadConstructiva.objects.filter(codigo=fila[0].value).first()

					#print unidades_constructivas
					if unidades_constructivas is None:

						#print 'entro'

						codigo =fila[0].value
						descripcion =fila[1].value
						valor_mano_obra =fila[2].value
						valor_material =fila[3].value

						modelo=BUnidadConstructiva(proyecto_id=None,contrato_id=contrato_id,codigo=codigo,descripcion=descripcion,valor_mano_obra=valor_mano_obra,valor_materiales=valor_material)
						modelo.save()

			index=index+1

		return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response


#funcion para traer el listado de detalle de solicitud
@api_view(['GET'])
def lista_detalle_solicitud(request):
	if request.method == 'GET':	
		cursor = connection.cursor()
		try:

			dato = request.GET.get('dato', None)
			cambio_id = request.GET['cambio_id']
			proyecto_id = request.GET['proyecto_id']
			usuario_id = request.user.usuario.id

			cursor.callproc('[dbo].[control_cambios_consultar_detalle_solicitud]', [cambio_id,proyecto_id,dato,usuario_id])			
			columns = cursor.description 
			lista = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]
			encabezados = [i[0] for i in columns]


			cursor.callproc('[dbo].[control_cambios_consultar_columnas_comparar_cambios]', [proyecto_id])	
			columns = cursor.description 
			soportes = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]

			return Response({'message':'','success':'ok','data':{'lista':lista, 'encabezados':encabezados, 'soportes':soportes}})	
		except Exception as e:

			#print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)	
		finally:
			cursor.close()



#exporta vista comparar
@api_view(['GET'])
def export_reporte_detalle_solicitud(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Comparacion_cambios.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Comparacion_cambios')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})

	
	cursor = connection.cursor()

	dato = request.GET.get('dato', None)
	cambio_id = request.GET['cambio_id']
	proyecto_id = request.GET['proyecto_id']
	usuario_id = request.user.usuario.id

	cursor.callproc('[dbo].[control_cambios_consultar_detalle_solicitud]', [cambio_id,proyecto_id,dato,usuario_id])			
	columns = cursor.description 
	lista = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]
	encabezados = [i[0] for i in columns]

	c=0

	for column in encabezados:

		if column!='id_uucc' and column!='id_estado' and column!='comentario' and column!='bloquear':	
			worksheet.write(0, c, column, format1)
			c += 1

	row=1	
	for r in lista:
		col=0
		for column in encabezados:

			if column!='id_uucc' and column!='id_estado' and column!='comentario' and column!='bloquear':	

				worksheet.write(row, col,r[column],format2)	
				col+=1
		row +=1


	workbook.close()

	return response
    #return response


#Actualiza el estado de la vista detall solicitud
@transaction.atomic
def actualizar_estado_mis_solicitudes(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		aprovado=100
		rechazado=101

		for item in respuesta['lista']:
			object_cambio_proyecto=FCambioProyecto.objects.get(uucc_id=item['id'], cambio_id=respuesta['cambio_id'])

			object_cambio_proyecto.estado_id=respuesta['estado']

			#if respuesta['estado']== enumEstados.Aprovada:
			if respuesta['estado']== aprovado:
				object_cambio_proyecto.comentario=respuesta['comentario']
				
			object_cambio_proyecto.save()


			#if respuesta['estado']== enumEstados.Rechazado:
			if respuesta['estado']== rechazado:
				object_cambio=CCambio.objects.get(pk=respuesta['cambio_id'])
				object_cambio.motivo=respuesta['motivo']
				object_cambio.save()


			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='control_cambios.cambio_proyecto',id_manipulado=item['id'])
			logs_model.save()

			transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro se ha actualizo correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#envio de correo cuando se crea un cambio
def notificacion_cambio(proyecto_id,usuario_revisa_id):
	mensaje=''
	status=''
	res=''
	
	try:
		usuario= Usuario.objects.get(id=usuario_revisa_id)
		persona = Persona.objects.get(id=usuario.persona.id)
		proyecto= Proyecto.objects.get(id=proyecto_id)

		#print persona.correo

		#inicio del codigo del envio de correo
		contenido='<h3>SININ - Sistema Integral de informacion</h3>'
		contenido = contenido + 'Estimado(a)'+persona.nombres +' '+ persona.apellidos +'<br/><br/>'
		contenido = contenido + 'Nos permitimos informale que se ha registrado un cambio en las cantidades de unidades constructivas del proyecto '+ proyecto.nombre +' en el macro-contrato '+ proyecto.mcontrato.nombre+'. Dicho cambio fue registrado por el usuario '+ persona.nombres +' '+ persona.apellidos+' de la empresa '
		contenido = contenido + 'No responder este mensaje, este correo es de uso informativo exclusivamente,<br/><br/>'
		contenido = contenido + 'Soporte SININ<br/>soporte@sinin.co'
		mail = Mensaje(
			remitente=settings.REMITENTE,
			destinatario=persona.correo,
			asunto='informativo de SININ',
			contenido=contenido,
			appLabel='Usuario',
			)			
		mail.save()
		res=sendAsyncMail(mail)
		mensaje='Se ha enviado un correo a ' + persona.correo + ' para notificarle la creacion de cambio en control de cambios.'
		status='ok'			
			
	except Usuario.DoesNotExist:
		mensaje='El usuario ' + str(username) + ' No se encuentra registrado en el sistema'
		status='error'
		return Response({'message':mensaje,'success':status,'data':''},status=status.HTTP_400_BAD_REQUEST)


@login_required
def index_control_cambio(request):

	qsConteo=CCambio.objects.filter(usuario_id=request.user.usuario.id,tipo_id=81).count()
	
	return render(request, 'index.html',{'app':'control_cambios','model':'UnidadesConstructiva','conteo':qsConteo})


@login_required
def ConfigurarUnidadesConstructivas(request):
	return render(request, 'configurar_unidades_constructivas.html',{'app':'control_cambios','model':'UnidadesConstructivas'})


@login_required
def CambiosObra(request,id_cuenta=None):
	return render(request, 'cambios_obra.html',{'app':'control_cambios','model':'Cambio'})		


@login_required
def MisSolicitudes(request,id_cuenta=None):
	return render(request, 'mis_solicitudes.html',{'app':'control_cambios','model':'Cambio'})		


@login_required
def AdministrarUUCC(request,id_proyecto=None):

	#qsProyecto=CCambio.objects.filter(proyecto__id=id_proyecto).first()
	qsProyecto=Proyecto.objects.filter(id=id_proyecto).first()
	qsTipo=Tipo.objects.filter(app='control_cambios')
	qsSolicita=ASolicita.objects.all()

	return render(request, 'administrar_uucc_cambios.html',{'solicita':qsSolicita,'tipo':qsTipo,'proyecto':qsProyecto,'app':'control_cambios','model':'Cambio','id_proyecto':id_proyecto})		


@login_required
def AgregarUUCC(request,id_proyecto=None,id_cambio=None,id_contrato=None):

	qsProyecto=CCambio.objects.filter(proyecto__id=id_proyecto, id=id_cambio).first()

	return render(request, 'agregar_ucc_cambio.html',{'proyecto':qsProyecto,'app':'control_cambios','model':'Cambioproyecto','id_proyecto':id_proyecto,'id_cambio':id_cambio,'id_contrato':id_contrato})		


@login_required
def Comparar(request,id_proyecto=None):
	qsProyecto=Proyecto.objects.filter(id=id_proyecto).first()
	return render(request, 'comparar.html',{'proyecto':qsProyecto,'app':'control_cambios','model':'CambioProyecto','id_proyecto':id_proyecto})


@login_required
def DetalleSolicitud(request,id_proyecto=None,id_cambio=None):

	qsProyecto=CCambio.objects.filter(proyecto__id=id_proyecto, id=id_cambio).first()

	return render(request, 'detalle_solicitud.html',{'proyecto':qsProyecto,'app':'control_cambios','model':'Cambio','id_proyecto':id_proyecto,'id_cambio':id_cambio})		


#plantilla de la vista agregar uucc
@login_required
def descargar_plantilla_agregaruucc(request):
	return functions.exportarArchivoS3('plantillas/uucc/PlantillaCambioUUCC.xlsx')


#plantilla de la vista configurar unidades constructivas
@login_required
def descargar_plantilla_unidadesConstructiva(request):
	return functions.exportarArchivoS3('plantillas/uucc/FormatoMUC.xlsx')