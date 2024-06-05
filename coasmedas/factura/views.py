# -*- coding: utf-8 -*-
from django.shortcuts import render,redirect,render_to_response
# from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.template import RequestContext
from django.contrib.contenttypes.models import ContentType

from django.conf import settings
# from django.conf.urls.static import static

from rest_framework import viewsets, serializers
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
# from rest_framework.parsers import MultiPartParser, FormParser

# Para importaciones a excel y retornar Json a la vista
from django.http import HttpResponse, JsonResponse

# Para consultas SQL
from django.db.models import Q, Sum
from django.db import transaction, connection
from django.db.models.deletion import ProtectedError

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, NamedStyle, Side, Font, colors
from openpyxl.worksheet import *
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
# from openpyxl.worksheet.copier import WorksheetCopy
import openpyxl, re
import json
# from copy import copy
import copy
# from io import StringIO

from datetime import *
# import calendar

from contrato.enumeration import tipoC, estadoC#, tipoV
from logs.models import Logs, Acciones
from .models import Factura, MesCausado, Cesion, Descuento, Compensacion, DetalleCompensacion, FacturaProyecto
from contrato.models import Contrato, EmpresaContrato
from estado.models import Estado
from tipo.models import Tipo
from empresa.models import Empresa
from proyecto.models import Proyecto
from parametrizacion.models import Banco
from giros.models import CNombreGiro,DEncabezadoGiro,DetalleGiro
from giros.enum import enumEstados, enumTipoPagoAnticipo
from proceso.models import GProcesoRelacionDato, HSoporteProcesoRelacionDato, FProcesoRelacion, BItem, AProceso, CPermisoEmpresaItem
from multa.models import Solicitud

from seguimiento_factura.models import GestionOp

from estado.views import EstadoSerializer
from tipo.views import TipoSerializer
from empresa.views import EmpresaSerializer
from contrato.views import ContratoLiteSerializerByDidi #ContratoSerializer
from parametrizacion.views import BancoSerializer
from proyecto.views import ProyectoSerializer
from seguimiento_factura.views import GestionOpSerializer
from proceso.views import SoporteProcesoRelacionDatoSerializer

from .enumeration import estadoFactura, tablaForanea

from sinin4.functions import functions
#Comienza la creacion de las vistas de Factura.

#Api rest para Factura

class ContratistaLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Empresa
		fields=('id','nombre', 'nit' ,'codigo_acreedor')

class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	contratista = ContratistaLiteSerializer(read_only=True)
	contratista_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Empresa.objects.all())

	class Meta:
		model = Contrato
		fields=('id','nombre','numero','contratista','contratista_id')

class MesCausadoConsultaSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = MesCausado
		fields=('id','mes','ano')

class ProyectoLiteSerializer2(serializers.HyperlinkedModelSerializer):

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre')

class FacturaProyectoConsultaSerializer(serializers.HyperlinkedModelSerializer):

	proyecto = ProyectoLiteSerializer2(read_only=True)
	proyecto_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Proyecto.objects.all())

	class Meta:
		model = FacturaProyecto
		fields=('id','proyecto','proyecto_id','valor')

class GestionOpLiteSerializer(serializers.HyperlinkedModelSerializer):

	valor_gestion_op = serializers.SerializerMethodField()

	class Meta:
		model = GestionOp
		fields=('id','codigo','fecha_registro','fecha_pago', 'valor_gestion_op')



	def get_valor_gestion_op(self, obj):

		try:
			sumatoria = Factura.objects.filter(codigo_op__id=obj.id).aggregate(total=Sum('valor_contable'))
			total = sumatoria['total']
		except Exception as e:
			total = 0
		
		return total

class FacturaSerializer(serializers.HyperlinkedModelSerializer):
	tipo_c=tipoC()
	contrato = ContratoLiteSerializer(read_only=True)
	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	estado = EstadoSerializer(read_only=True)
	estado_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Estado.objects.filter(app='Factura'))

	codigo_op = GestionOpLiteSerializer(read_only=True)
	codigo_op_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = GestionOp.objects.all(), allow_null = True)

	mes_causado = MesCausadoConsultaSerializer(many=True,read_only=True)

	# proyecto = ProyectoLiteSerializer2(read_only = True, many=True)
	proyecto = FacturaProyectoConsultaSerializer(many=True,read_only=True)

	mcontrato = ContratoLiteSerializerByDidi(read_only=True)
	mcontrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.filter(tipo_contrato=tipo_c.m_contrato), allow_null = True)
	soloLectura = serializers.SerializerMethodField()
	class Meta:
		model = Factura
		fields=('id','referencia','radicado','numero','fecha','concepto','valor_factura','valor_contable',
			'valor_subtotal','soporte','contrato_id','contrato','estado_id','estado','mes_causado','proyecto',
			'codigo_op','codigo_op_id','pagada','bloqueo_factura','recursos_propios','orden_pago','fecha_reporte',
			'fecha_pago','motivo_anulacion','mcontrato','mcontrato_id', 'fecha_vencimiento', 'soloLectura')

	def get_soloLectura(self, obj):
		request = self.context.get("request")
		if request is None:
			return False

		contrato = None
		if obj.mcontrato:
			contrato = obj.mcontrato.id

		empresaContrato = EmpresaContrato.objects.filter(contrato__id=contrato, empresa__id=request.user.usuario.empresa.id).first()
		soloLectura = False if empresaContrato is not None and empresaContrato.edita else True
		return soloLectura	
			
class FacturaViewSet(viewsets.ModelViewSet):
	"""

	"""
	model=Factura
	queryset = model.objects.all()
	serializer_class = FacturaSerializer
	nombre_modulo = 'factura - FacturaViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(FacturaViewSet, self).get_queryset()
			referencia = self.request.query_params.get('referencia', None)
			referencia_isnull = self.request.query_params.get('referencia_isnull', None)

			numero = self.request.query_params.get('numero', None)
			num_ref_cont = self.request.query_params.get('dato', None)
			fecha_desde = self.request.query_params.get('fecha_desde', None)
			fecha_hasta = self.request.query_params.get('fecha_hasta', None)
			id_estado = self.request.query_params.get('id_estado').split(',') if self.request.query_params.get('id_estado') else None
			valor_cont = self.request.query_params.get('valor_cont', None)

			id_contrato = self.request.query_params.get('id_contrato', None)
			tipo_contrato = self.request.query_params.get('tipo_contrato', None)
			numero_contrato = self.request.query_params.get('numero_contrato', None)
			id_contratista = self.request.query_params.get('id_contratista', None)
			id_mcontrato = self.request.query_params.get('id_mcontrato', None)
			
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id

			orden_pago = self.request.query_params.get('orden_pago', None)
			bloqueo_factura = self.request.query_params.get('bloqueo_factura', None)
			pagada = self.request.query_params.get('pagada', None)
			
			recursos_propios = self.request.query_params.get('recursos_propios', None)
			codigo_op = self.request.query_params.get('codigo_op', None)

			# consulta las facturas por contabilizar
			# luis mendoza
			#this get true or false
			por_contabilizar = self.request.query_params.get('por_contabilizar', None)
			habilitadas_para_contabilizacion = self.request.query_params.get('habilitadas_para_contabilizacion', None)
			#this get number of rad
			radicado = self.request.query_params.get('radicado', None)
			# tags
			#this get true or false
			por_pagar_cuenta_bancaria = self.request.query_params.get('por_pagar_cuenta_bancaria', None)

			qset=(~Q(id=0))

			if num_ref_cont:
				qset = qset &(
					Q(referencia__icontains=num_ref_cont)|Q(radicado__icontains=num_ref_cont)|Q(numero__icontains=num_ref_cont)|Q(concepto__icontains=num_ref_cont)|Q(contrato__numero__icontains=num_ref_cont)
					)

			if referencia_isnull:
				qset = qset &(Q(referencia__exact=''))

			if referencia:
				qset = qset &(Q(referencia__icontains=referencia))

			if radicado:
				qset = qset &(Q(radicado__icontains=radicado))

			if numero:
				qset = qset &(Q(numero__icontains=numero))

			if fecha_desde:
				qset = qset &(Q(fecha__gte=fecha_desde))
			if fecha_hasta:
				qset = qset &(Q(fecha__lte=fecha_hasta))

			if valor_cont:
				qset = qset &(Q(valor_contable__gt=0))

			if id_contrato and int(id_contrato)>0:
				qset = qset &(Q(contrato__id=id_contrato))
			if tipo_contrato:
				qset = qset &(Q(contrato__tipo_contrato=tipo_contrato))
			if numero_contrato:
				qset = qset &(Q(contrato__numero=numero_contrato))
			if id_contratista and int(id_contratista)>0:
				qset = qset &(Q(contrato__contratista_id=id_contratista))
			if id_mcontrato and int(id_mcontrato)>0:
				qset = qset &(Q(contrato__mcontrato_id=id_mcontrato))

			if id_estado:
				qset = qset &(Q(estado__in=id_estado))

			if id_empresa:
				qset = qset &(Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1))

			#validacion para el modulo de seguimiento de factura 
			if orden_pago:
				qset = qset &(Q(orden_pago=orden_pago))

			if bloqueo_factura:
				qset = qset &(Q(bloqueo_factura=bloqueo_factura))

			if pagada:
				qset = qset &(Q(estado_id__in = [estadoFactura.compensada]))
				qset = qset &(Q(pagada = True))
				qset = qset &(Q(codigo_op_id__isnull = False))

			if por_pagar_cuenta_bancaria:
				qset = qset &(Q(estado_id__in = [estadoFactura.compensada]))
				qset = qset &(Q(pagada = False))
				qset = qset &(Q(codigo_op_id__isnull = False))
				qset = qset &(Q(recursos_propios = False))

			if recursos_propios:
				qset = qset &(Q(recursos_propios=recursos_propios))

			if codigo_op:
				qset = qset &(Q(codigo_op_id=codigo_op))
				
			if por_contabilizar:
				qset = qset &(Q(estado_id = estadoFactura.activa) & Q(fecha__gte = '2018-10-31'))
				qset = qset &(Q(referencia__exact = '') | Q(referencia__exact = '0'))

			if habilitadas_para_contabilizacion:
				qset = qset &(Q(estado_id = estadoFactura.activa) & Q(fecha__gte = '2018-10-31'))
				qset = qset &(Q(fecha_contabilizacion__isnull = False) & Q(fecha_vencimiento__isnull = False))
				qset = qset &(Q(pagada = False))
				qset = qset &(Q(bloqueo_factura = False))


			queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				existeFactura = Factura.objects.filter(contrato__id=request.DATA['contrato_id'],
														numero=request.DATA['numero']).exists()

				if existeFactura:
					return Response({'message':'La factura que intenta ingresar ya existe.','success':'fail','data':''},
										status=status.HTTP_400_BAD_REQUEST)

				serializer = FacturaSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if request.DATA['codigo_op_id'] == '':
					request.DATA['codigo_op_id']=None
				if request.DATA['mcontrato_id'] == '':
					request.DATA['mcontrato_id']=None
				if request.DATA['recursos_propios'] == 'true':
					request.DATA['recursos_propios'] = 1

				if serializer.is_valid():

					serializer.save(contrato_id=request.DATA['contrato_id'], estado_id=request.DATA['estado_id'],
						soporte=request.FILES['soporte'] if request.FILES.get('soporte') is not None else '',
						codigo_op_id=request.DATA['codigo_op_id'],
						mcontrato_id=request.DATA['mcontrato_id'])

					# Guardar Mes causado
					if request.DATA['mes_causado']:
						createMesCausadoLista(request.DATA['mes_causado'], serializer.data['id'])
						# respuesta= json.loads(request.DATA['mes_causado'])
						# # print myList
						# for item in respuesta:
						# 	# print item['mes']

					if request.DATA['proyecto']:
						createFacturaProyectoConLista(request.DATA['proyecto'], serializer.data['id'], request)

					#Guardar en Proceso
					if request.DATA['factura_final'] == 'true':
						# print "Es Factura Final: "+str(request.DATA['factura_final'])
						crearSoporteProcesoFactura(request, serializer.data['id'])
					# else:
					# 	print "No es Factura Final"

					# Crear un proceso Para Enelar
					db = settings.DATABASES['default']['NAME']
					# id_tabla = 140
					id_proceso = 3
					if db == 'sinin41_Enelar':
						# crearProesoEnelarFactura(int(request.DATA['contrato_id']), int(serializer.data['id']), int(id_proceso))
						crearProesoEnelarFactura(request, int(serializer.data['id']), int(id_proceso))
						crearPermisoProesoItem(request, id_proceso)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.facturas',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					# print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# #print request.DATA
				serializer = FacturaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if request.DATA['mcontrato_id'] == '':
					request.DATA['mcontrato_id']=None

				if serializer.is_valid():
					if request.DATA['codigo_op_id'] == '':
						request.DATA['codigo_op_id']=None

					serializer.save(
						soporte=self.request.FILES['soporte'] if self.request.FILES.get('soporte') is not None else instance.soporte,
						contrato_id=self.request.DATA['contrato_id'],
						estado_id=self.request.DATA['estado_id'],
						codigo_op_id=request.DATA['codigo_op_id'],
						mcontrato_id=request.DATA['mcontrato_id'])

					# Actualizar Mes causado
					# print ("mes: ", request.DATA['meses'] )
					# print ("ano: ", request.DATA['ano'] )
					# print ("id: ", serializer.data['id'] )
					# actualizarMesCausadoLista(request.DATA['meses'], request.DATA['ano'], serializer.data['id'])
					actualizarMesCausadoLista(request.DATA['mes_causado'], serializer.data['id'])

					updateSoporteProcesoFactura(request)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.facturas',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					# print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.facturas',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para Factura


#Api rest para Cesion
class CesionSerializer(serializers.HyperlinkedModelSerializer):

	contrato = ContratoLiteSerializer(read_only=True)
	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	beneficiario = EmpresaSerializer(read_only=True)
	beneficiario_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Empresa.objects.all())

	banco = BancoSerializer(read_only=True)
	banco_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Banco.objects.all())

	tipo_cuenta = TipoSerializer(read_only=True)
	tipo_cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='cuenta'))

	proceso_soporte = SoporteProcesoRelacionDatoSerializer(read_only=True)
	proceso_soporte_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = HSoporteProcesoRelacionDato.objects.all(), allow_null = True)

	class Meta:
		model = Cesion
		fields=('id','referencia','numero_cuenta','descripcion','fecha','valor','soporte',
			'contrato_id','contrato','beneficiario','beneficiario_id','banco','banco_id',
			'tipo_cuenta','tipo_cuenta_id','proceso_soporte','proceso_soporte_id')

class CesionViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=Cesion
	queryset = model.objects.all()
	serializer_class = CesionSerializer
	nombre_modulo = 'factura - CesionViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CesionViewSet, self).get_queryset()
			ref_desc = self.request.query_params.get('dato', None)
			referencia = self.request.query_params.get('referencia', None)
			descripcion = self.request.query_params.get('descripcion', None)
			fecha_desde = self.request.query_params.get('fecha_desde', None)
			fecha_hasta = self.request.query_params.get('fecha_hasta', None)
			id_contrato = self.request.query_params.get('id_contrato', None)
			num_contrato = self.request.query_params.get('numero_contrato', None)
			id_beneficiario = self.request.query_params.get('id_beneficiario', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id

			qset = (~Q(id=0))

			if ref_desc:
				qset = ( Q(referencia__icontains=ref_desc)|Q(descripcion__icontains=ref_desc)|Q(contrato__numero__icontains=ref_desc))

			if referencia:
				qset = qset &( Q(referencia__icontains=referencia) )
				
			if descripcion:
				qset = qset &(Q(descripcion__icontains=descripcion))

			if fecha_desde:
				qset = qset &(Q(fecha__gte=fecha_desde))
				
			if fecha_hasta:
				qset = qset &(Q(fecha__lte=fecha_hasta))

			if id_contrato:
				qset = qset &(Q(contrato__id=id_contrato))
				
			if num_contrato:
				qset = qset &(Q(contrato__numero__icontains=num_contrato))

			if id_beneficiario:
				qset = qset &( Q(beneficiario__id=id_beneficiario))
				
			if id_empresa:
				qset = qset &(
						#Q(empresa_contrato__empresa_id=id_empresa)
						Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
					)
				
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CesionSerializer(data=request.DATA,context={'request': request})

				if request.DATA['proceso_soporte_id'] == '':
					request.DATA['proceso_soporte_id']=None

				if serializer.is_valid():

					serializer.save(contrato_id=request.DATA['contrato_id'], beneficiario_id=request.DATA['beneficiario_id'],
						banco_id=request.DATA['banco_id'] if request.DATA['banco_id'] != '' else '',
						soporte=request.FILES['soporte'] if request.FILES.get('soporte') is not None else '',
						tipo_cuenta_id=request.DATA['tipo_cuenta_id'] if request.DATA['tipo_cuenta_id'] != '' else '',
						proceso_soporte_id=request.DATA['proceso_soporte_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.Cesion',id_manipulado=serializer.data['id'])
					logs_model.save()

					if int(request.DATA['nombre_giro']) > 0:
						crearDetalleGiro(request, serializer.data['id'])

					crearSoporteProceso(request, serializer.data['id'])

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# #print request.DATA
				serializer = CesionSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					if request.DATA['proceso_soporte_id'] == '':
						request.DATA['proceso_soporte_id']=None
					serializer.save(
						soporte=self.request.FILES['soporte'] if self.request.FILES.get('soporte') is not None else instance.soporte, 						
						contrato_id=self.request.DATA['contrato_id'],
						beneficiario_id=self.request.DATA['beneficiario_id'],
						banco_id=request.DATA['banco_id'], tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
						proceso_soporte_id=request.DATA['proceso_soporte_id'])

					updateDetalleGiro(request)
					updateSoporteProceso(request)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.Cesion',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.Cesion',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para Cesion


#Api rest para Descuento
class DescuentoSerializer(serializers.HyperlinkedModelSerializer):

	contrato = ContratoLiteSerializer(read_only=True)
	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	banco = BancoSerializer(read_only=True)
	banco_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Banco.objects.all())

	class Meta:
		model = Descuento
		fields=('id','referencia','numero_cuenta','concepto','valor','soporte',
			'contrato_id','contrato','banco','banco_id')

class DescuentoViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=Descuento
	queryset = model.objects.all()
	serializer_class = DescuentoSerializer
	nombre_modulo = 'factura - CesionViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DescuentoViewSet, self).get_queryset()
			ref_conc = self.request.query_params.get('dato', None)
			referencia = self.request.query_params.get('referencia', None)
			num_contrato = self.request.query_params.get('numero_contrato', None)
			concepto = self.request.query_params.get('concepto', None)
			id_contrato = self.request.query_params.get('id_contrato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id

			qset = (~Q(id=0)) 

			if ref_conc:
				qset = ( Q(referencia__icontains=ref_conc)|Q(concepto__icontains=ref_conc))

			if referencia:
				qset = qset &(Q(referencia__icontains=referencia))

			if concepto:
				qset = qset &(Q(concepto__icontains=concepto))

			if id_contrato:
				qset = qset &(Q(contrato__id=id_contrato))

			if num_contrato:
				qset = qset &(Q(contrato__numero__icontains=num_contrato))

			if id_empresa:
				qset = qset &(
						#Q(empresa_contrato__empresa_id=id_empresa)
						Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
					)
				
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DescuentoSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if serializer.is_valid():
					#serializer.save()

					serializer.save(contrato_id=request.DATA['contrato_id'],
						banco_id=request.DATA['banco_id'],
						soporte=request.FILES['soporte'] if request.FILES.get('soporte') is not None else '')

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.Descuento',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				#print request.DATA
				serializer = DescuentoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(
						soporte=self.request.FILES['soporte'] if self.request.FILES.get('soporte') is not None else instance.soporte, 						
						contrato_id=self.request.DATA['contrato_id'],
						banco_id=request.DATA['banco_id'] if request.DATA['banco_id'] is not None else None)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.Descuento',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				# print(serializer.errors)()
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.Descuento',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para Descuento


class ContentTypeSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = ContentType
		fields=('id','name','app_label','model')

class ContentTypeLiteSerializer(serializers.HyperlinkedModelSerializer):
    
	class Meta:
		model = ContentType
		fields=('id','model')

#Api rest para Compensacion
class DetalleCompensacionConsultaSerializer(serializers.HyperlinkedModelSerializer):

	tablaForanea = ContentTypeSerializer(read_only=True)
	tablaForanea_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = ContentType.objects.filter(app_label__in=['factura', 'giros'] ).order_by('app_label'))

	class Meta:
		model = DetalleCompensacion
		fields=('id','tablaForanea','tablaForanea_id','id_registro')

class DetalleCompensacionDetalleSerializer(serializers.HyperlinkedModelSerializer):

	tablaForanea = ContentTypeLiteSerializer(read_only=True)
	# tablaForanea_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = ContentType.objects.filter(app_label__in=['factura', 'giros'] ).order_by('app_label'))

	class Meta:
		model = DetalleCompensacion
		fields=('id','tablaForanea','id_registro')

class CompensacionSerializer(serializers.HyperlinkedModelSerializer):

	contrato = ContratoLiteSerializer(read_only=True)
	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	detalle_cruce = DetalleCompensacionConsultaSerializer(many=True,read_only=True)

	class Meta:
		model = Compensacion
		fields=('id','referencia','fecha','descripcion','valor','contrato_id','contrato','detalle_cruce')

class CompensacionLiteDetalleSerializer(serializers.HyperlinkedModelSerializer):
    
	contrato = ContratoLiteSerializer(read_only=True)
	## contrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	# detalle_cruce = DetalleCompensacionDetalleSerializer(many=True,read_only=True)

	giro = serializers.SerializerMethodField('_giro',read_only=True)
	factura = serializers.SerializerMethodField('_factura',read_only=True)
	cesion = serializers.SerializerMethodField('_cesion',read_only=True)
	descuento = serializers.SerializerMethodField('_descuento',read_only=True)
	multa = serializers.SerializerMethodField('_multa',read_only=True)

	def _giro(self, obj):

		model_detalleCompensacion = DetalleCompensacion.objects.filter(compensacion=obj.id)

		lista=[]
		row = 0
		# print obj.tablaForanea.model
		# print obj.id_registro

		#solicitado,autorizado o pagado
		estado_solicitada=Estado.objects.filter(app='EstadoGiro',codigo=1)
		estado_autorizada=Estado.objects.filter(app='EstadoGiro',codigo=2)
		estado_pagada=Estado.objects.filter(app='EstadoGiro',codigo=3)

		for item in model_detalleCompensacion:
			if item.tablaForanea.model == 'dencabezadogiro':
				model_EncabezadoGiro = DEncabezadoGiro.objects.filter(id=item.id_registro).values('id','referencia','nombre__nombre','soporte').first()

				sumatoria=DetalleGiro.objects.filter(encabezado_id=item.id_registro,estado_id__in=[estado_pagada[0].id,estado_autorizada[0].id,estado_solicitada[0].id]).aggregate(suma_detalle=Sum('valor_girar'))

				model_EncabezadoGiro.update({'valor':sumatoria['suma_detalle']})

				# lista[row] = model_EncabezadoGiro
				lista.append(model_EncabezadoGiro)
				row +=1
				# print model_EncabezadoGiro['nombre__nombre']

		return lista

	def _factura(self, obj):

		model_detalleCompensacion = DetalleCompensacion.objects.filter(compensacion=obj.id)

		lista=[]

		for item in model_detalleCompensacion:
			if item.tablaForanea.model == 'factura':
				model_factura = Factura.objects.filter(id=item.id_registro).values('id','referencia','fecha','valor_contable','soporte').first()

				lista.append(model_factura)

		return lista

	def _cesion(self, obj):
		model_detalleCompensacion = DetalleCompensacion.objects.filter(compensacion=obj.id)

		lista=[]

		for item in model_detalleCompensacion:
			if item.tablaForanea.model == 'cesion':
				model_cesion = Cesion.objects.filter(id=item.id_registro).values('id','referencia','fecha','valor','soporte').first()

				lista.append(model_cesion)

		return lista

	def _descuento(self, obj):
		model_detalleCompensacion = DetalleCompensacion.objects.filter(compensacion=obj.id)

		lista=[]

		for item in model_detalleCompensacion:
			if item.tablaForanea.model == 'descuento':
				model_descuento = Descuento.objects.filter(id=item.id_registro).values('id','referencia','concepto','valor','soporte').first()

				lista.append(model_descuento)

		return lista

	def _multa(self, obj):
		model_detalleCompensacion = DetalleCompensacion.objects.filter(compensacion=obj.id)

		lista=[]

		for item in model_detalleCompensacion:
			if item.tablaForanea.model == 'solicitud':
				model_multa = Solicitud.objects.filter(id=item.id_registro).values('id','codigoReferencia','fechaDiligencia','valorImpuesto','soporte').first()

				lista.append(model_multa)

		return lista

	class Meta:
		model = Compensacion
		fields=('id','referencia','fecha','descripcion','valor','contrato'
				,'giro', 'factura', 'cesion', 'descuento', 'multa')

class CompensacionViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=Compensacion
	queryset = model.objects.all()
	serializer_class = CompensacionSerializer
	nombre_modulo = 'factura - CompensacionViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()

			lite_detalle = self.request.query_params.get('lite_detalle',None)

			if lite_detalle:
				serializer = CompensacionLiteDetalleSerializer(instance, context={'request': request})
			else:
				serializer = self.get_serializer(instance)

			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CompensacionViewSet, self).get_queryset()
			ref_desc = self.request.query_params.get('dato', None)
			referencia = self.request.query_params.get('referencia', None)
			descripcion = self.request.query_params.get('descripcion', None)
			fecha = self.request.query_params.get('fecha', None)
			id_contrato = self.request.query_params.get('id_contrato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id

			qset = (~Q(id=0))

			if ref_desc:
				qset = (
					Q(referencia__icontains=ref_desc)|Q(descripcion__icontains=ref_desc)|Q(contrato__numero__icontains=ref_desc)
					)

			if referencia:
				qset = qset &(Q(referencia__icontains=referencia))

			if descripcion:
				qset = qset &( Q(descripcion__icontains=descripcion) )
				
			if fecha:
				qset = qset &(Q(fecha=fecha))				

			if id_contrato:
				qset = qset &(Q(contrato__id=id_contrato))				

			if id_empresa:
				qset = qset &(
						#Q(empresa_contrato__empresa_id=id_empresa)
						Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
					)
				
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})
			else: # Sin Paginacion
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CompensacionSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if serializer.is_valid():
					#serializer.save()

					serializer.save(contrato_id=request.DATA['contrato_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.Compensacion',id_manipulado=serializer.data['id'])
					logs_model.save()

					if request.DATA['ids_anticipo']:
						# print("anrticipo:", request.DATA['ids_anticipo'])
						createDetalleCompensacionLista(request.DATA['ids_anticipo'], tablaForanea.encabezadoGiros, serializer.data['id'])

					if request.DATA['ids_factura']:
						# print("ids_factura:", request.DATA['ids_factura'])
						createDetalleCompensacionLista(request.DATA['ids_factura'], tablaForanea.factura, serializer.data['id'])

					if request.DATA['ids_cesion']:
						# print("ids_cesion:", request.DATA['ids_cesion'])
						createDetalleCompensacionLista(request.DATA['ids_cesion'], tablaForanea.cesion, serializer.data['id'])

					if request.DATA['ids_descuento']:
						# print("ids_descuento:", request.DATA['ids_descuento'])
						createDetalleCompensacionLista(request.DATA['ids_descuento'], tablaForanea.descuento, serializer.data['id'])

					if request.DATA['ids_multa']:
    						# print("ids_multa:", request.DATA['ids_multa'])
						createDetalleCompensacionLista(request.DATA['ids_multa'], tablaForanea.multa, serializer.data['id'])

					resultado = None
					if float(request.DATA['valor']) < 0:
						resultado = crearDetalleGiro2(request, serializer.data['id'])

						if resultado != '1':
							transaction.savepoint_rollback(sid)
							return Response({'message':resultado,'success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

						# print "resul: "+ str(resultado)

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)()
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# #print request.DATA
				serializer = CompensacionSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					# print 'entro'
					serializer.save(contrato_id=self.request.DATA['contrato_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.Compensacion',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.Compensacion',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para Cesion


#Api rest para MesCausado
class MesCausadoSerializer(serializers.HyperlinkedModelSerializer):

	factura = FacturaSerializer(read_only=True)
	factura_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Factura.objects.all())

	class Meta:
		model = MesCausado
		fields=('id','factura','factura_id','mes','ano')

class MesCausadoViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=MesCausado
	queryset = model.objects.all()
	serializer_class = MesCausadoSerializer
	nombre_modulo = 'factura - MesCausadoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(MesCausadoViewSet, self).get_queryset()
			mes_ano = self.request.query_params.get('dato', None)
			mes = self.request.query_params.get('mes', None)
			ano = self.request.query_params.get('ano', None)
			id_factura = self.request.query_params.get('id_factura', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)

			qset = (~Q(id=0))
			if mes_ano:
				qset = (Q(mes__icontains=mes_ano)|Q(ano__icontains=mes_ano))

			if mes:
				qset = qset &(Q(mes__icontains=mes))
				
			if ano:
				qset = qset &(Q(ano__icontains=ano))

			if id_factura:
				qset = qset &(Q(factura__id=id_factura))
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else: # Sin Paginacion
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = MesCausadoSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if serializer.is_valid():
					#serializer.save()

					serializer.save(factura_id=request.DATA['factura_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.MesCausado',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				#print request.DATA
				serializer = MesCausadoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(factura_id=self.request.DATA['factura_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.MesCausado',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.MesCausado',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para MesCausado


#Api rest para FacturaProyecto
class ProyectoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre',
				 'No_cuenta' ,
				 'valor_adjudicado' ,
				 'fecha_inicio' , 'fecha_fin')

class FacturaProyectoSerializer(serializers.HyperlinkedModelSerializer):

	factura = FacturaSerializer(read_only=True)
	factura_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Factura.objects.all())

	proyecto = ProyectoLiteSerializer(read_only=True)
	proyecto_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Proyecto.objects.all())

	class Meta:
		model = FacturaProyecto
		fields=('id','factura','factura_id','proyecto','proyecto_id','valor')

		validators=[serializers.UniqueTogetherValidator(
								queryset=model.objects.all(),
								fields=('factura_id','proyecto_id'),
								message=('La factura ya tiene asignado un proyecto.') ) ]

class FacturaProyectoViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=FacturaProyecto
	queryset = model.objects.all()
	serializer_class = FacturaProyectoSerializer
	nombre_modulo = 'factura - FacturaProyectoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(FacturaProyectoViewSet, self).get_queryset()
			# mes_ano = self.request.query_params.get('dato', None)
			id_proyecto = self.request.query_params.get('id_proyecto', None)
			id_factura = self.request.query_params.get('id_factura', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)

			qset = (~Q(id=0))
			# if mes_ano:
			# 	qset = (
			# 		Q(mes__icontains=mes_ano)|Q(ano__icontains=mes_ano)
			# 		)

			if id_factura:
				qset = qset &(Q(factura__id=id_factura))				

			if id_proyecto:
				qset = qset &(Q(proyecto__id=id_proyecto))
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else: # Sin Paginacion
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = FacturaProyectoSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if serializer.is_valid():
					#serializer.save()

					serializer.save(factura_id=request.DATA['factura_id'], proyecto_id=request.DATA['proyecto_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.FacturaProyecto',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				#print request.DATA
				serializer = FacturaProyectoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					serializer.save(factura_id=self.request.DATA['factura_id'], proyecto_id=request.DATA['proyecto_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.FacturaProyecto',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					# print(serializer.errors)
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.FacturaProyecto',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para FacturaProyecto


#Api rest para DetalleCompensacion
class DetalleCompensacionSerializer(serializers.HyperlinkedModelSerializer):

	compensacion = CompensacionSerializer(read_only=True)
	compensacion_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Compensacion.objects.all())

	tablaForanea = ContentTypeSerializer(read_only=True)
	tablaForanea_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = ContentType.objects.filter(app_label__in=['factura', 'giros'] ).order_by('app_label'))

	class Meta:
		model = DetalleCompensacion
		fields=('id','compensacion','compensacion_id','tablaForanea','tablaForanea_id','id_registro')

class DetalleCompensacionViewSet(viewsets.ModelViewSet):
	"""
	
	"""
	model=DetalleCompensacion
	queryset = model.objects.all()
	serializer_class = DetalleCompensacionSerializer
	nombre_modulo = 'factura - DetalleCompensacionViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)

	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleCompensacionViewSet, self).get_queryset()
			# mes_ano = self.request.query_params.get('dato', None)
			id_compensacion = self.request.query_params.get('id_compensacion', None)
			id_tablaForanea = self.request.query_params.get('id_tablaForanea', None)
			id_registro = self.request.query_params.get('id_registro', None)
			id_contrato = self.request.query_params.get('id_contrato', None)
			sin_paginacion = self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id

			qset = (~Q(id=0))

			if id_compensacion:
				qset = qset &(Q(compensacion__id=id_compensacion))

			if id_tablaForanea:
				qset = qset &(Q(tablaForanea__id=id_tablaForanea))

			if id_registro:
				qset = qset &(Q(id_registro=id_registro))

			if id_contrato:
				qset = qset &(Q(compensacion__contrato__id=id_contrato))

			if id_empresa:
				qset = qset &(
						#Q(empresa_contrato__empresa_id=id_empresa)
						Q(compensacion__contrato__empresacontrato__empresa=id_empresa) &
						Q(compensacion__contrato__empresacontrato__participa=1) &
						Q(compensacion__contrato__activo=1)
					)
				
			if qset is not None:
				queryset = self.model.objects.filter(qset).order_by('-id')

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)

				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok','data':serializer.data})

			else: # Sin Paginacion
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DetalleCompensacionSerializer(data=request.DATA,context={'request': request})
				# print serializer

				if serializer.is_valid():
					#serializer.save()

					serializer.save(compensacion_id=request.DATA['compensacion_id'], tablaForanea_id=request.DATA['tablaForanea_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='factura.DetalleCompensacion',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				# #print request.DATA
				serializer = DetalleCompensacionSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				if serializer.is_valid():
					# print 'entro'
					serializer.save(compensacion_id=request.DATA['compensacion_id'], tablaForanea_id=request.DATA['tablaForanea_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.DetalleCompensacion',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)()
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			# self.perform_destroy(instance)
			self.model.objects.get(pk=instance.id).delete()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.DetalleCompensacion',id_manipulado=instance.id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para DetalleCompensacion


@login_required
def facturacion(request):
	return render(request, 'factura/inicio.html',{'model':'factura','app':'factura'})

@login_required
def factura(request):
	tipo_c = tipoC()
	querysetTipos=Tipo.objects.filter(app='contrato')
	id_empresa = request.user.usuario.empresa.id
	querysetMContrato=Contrato.objects.filter(empresacontrato__empresa=id_empresa, empresacontrato__participa=1, tipo_contrato=tipo_c.m_contrato, activo=1)
	return render(request, 'factura/factura.html',{'tipos':querysetTipos, 'mcontratos':querysetMContrato, 'tipo_mcontrato':tipo_c.m_contrato, 'model':'factura','app':'factura'})

@login_required
def cesion(request):
	querysetTipos=Tipo.objects.filter(app='contrato')
	querysetBanco=Banco.objects.all()
	return render(request, 'factura/cesion.html',{'tipos':querysetTipos, 'bancos':querysetBanco, 'model':'cesion','app':'factura'})

@login_required
def descuento(request):
	querysetTipos=Tipo.objects.filter(app='contrato')
	querysetBanco=Banco.objects.all()
	return render(request, 'factura/descuento.html',{'tipos':querysetTipos, 'bancos':querysetBanco, 'model':'descuento','app':'factura'})

@login_required
def compensacion(request):
	querysetTipos=Tipo.objects.filter(app='contrato')
	return render(request, 'factura/cruce.html',{'tipos':querysetTipos, 'model':'compensacion','app':'factura'})


# Guardar Mes caudado
def createMesCausadoLista(mes_causado, id_factura):
	try:

		respuesta= json.loads(mes_causado)
		# print respuesta
		for item in respuesta:
			model_mesCausado = MesCausado(mes=item['mes'],ano=item['ano'],factura_id=id_factura)
			model_mesCausado.save()
		# return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data': ''})
	except Exception as e:
		functions.toLog(e, 'Factura - createMesCausadoLista')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Actualizar mes Causado
def actualizarMesCausadoLista(mes_causado, id_factura):
	try:
		model_mesCausado1 = MesCausado.objects.filter(factura_id=id_factura)
		model_mesCausado1.delete()

		if mes_causado:

			respuesta = json.loads(mes_causado)
			# print respuesta
			for item in respuesta:
				model_mesCausado = MesCausado(mes=item['mes'],ano=item['ano'],factura_id=id_factura)
				model_mesCausado.save()

		# return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data': ''})
	except Exception as e:
		functions.toLog(e, 'Factura - actualizarMesCausadoLista')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	# try:
	# 	model_mesCausado1 = MesCausado.objects.filter(factura_id=id_factura)
	# 	model_mesCausado1.delete()

	# 	if meses != '' and ano != '':
	# 		myList = meses.split(',')

	# 		for item in myList:
	# 			if item:
	# 				model_mesCausado = MesCausado(mes=item,ano=ano,factura_id=id_factura)
	# 				model_mesCausado.save()

	# 	return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data': ''})
	# except Exception as e:
	# 	functions.toLog(e, 'Factura - actualizarMesCausadoLista')
	# 	return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Cambiar el Estado en factura
@transaction.atomic
def cambiarEstadoFactura(request):
		
	sid = transaction.savepoint()
	try:
		result=Factura.objects.get(pk=request.GET['id'])

		# Buscar si tiene sopote en proceso
		# result = Cesion.objects.get(pk = item['id'])
		if result.proceso_soporte_id:

			model_sprd = HSoporteProcesoRelacionDato.objects.get(pk = result.proceso_soporte_id)

			# ACTUALIZAR EL PROCESO_RELACION_DATO
			queryset_prd = GProcesoRelacionDato.objects.get(pk = model_sprd.procesoRelacionDato_id)
			queryset_prd.valor = 'Vacio'
			queryset_prd.estado = 0
			queryset_prd.save()

			result.estado_id=request.GET['estado']
			result.proceso_soporte_id = None
			result.motivo_anulacion=request.GET['motivo']
			result.save()

			model_sprd.delete()
		else:
			result.estado_id=request.GET['estado']
			result.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.facturas',id_manipulado=request.GET['id'])
		logs_model.save()

		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':''})

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - cambiarEstadoFactura')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@transaction.atomic
def cambiarEstadoFactura2(id, estado):
	sid = transaction.savepoint()
	try:
		result=Factura.objects.get(pk=id)
		result.estado_id=estado
		result.save()

		# print "cambio estado factura"

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - cambiarEstadoFactura2')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# exporta a excel factura
def exportReporteFactura(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_factura.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Factura')

	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 25)
	worksheet.set_column('C:E', 12)
	worksheet.set_column('F:I', 18)

	# cursor = connection.cursor()

	# encabezado_id= request.GET['encabezado_id']

	num_ref_cont = None
	tipo_contrato = None
	id_contratista = None
	numero_contrato = None
	referencia = None
	numero = None
	radicado = None
	fecha_desde = None
	fecha_hasta = None
	id_estado = estadoFactura.activa
	id_empresa = request.user.usuario.empresa.id

	if request.GET['dato']:
		num_ref_cont = request.GET['dato']

	if request.GET['radicado']:
		radicado = request.GET['radicado']

	if request.GET['tipo_contrato']:
		tipo_contrato = request.GET['tipo_contrato']

	if request.GET['id_contratista']:
		id_contratista = request.GET['id_contratista']

	if request.GET['numero_contrato'] and request.GET['numero_contrato'] != 'null':
		numero_contrato = request.GET['numero_contrato']

	if request.GET['referencia'] and request.GET['referencia'] != 'null':
		referencia = request.GET['referencia']

	if request.GET['numero'] and request.GET['numero'] != 'null':
		numero = request.GET['numero']

	if request.GET['fecha_desde'] and request.GET['fecha_desde'] != 'null':
		fecha_desde = request.GET['fecha_desde']

	if request.GET['fecha_hasta'] and request.GET['fecha_hasta'] != 'null':
		fecha_hasta = request.GET['fecha_hasta']

	qset=(~Q(id=0))
	if num_ref_cont:
		qset = qset &(Q(referencia__icontains=num_ref_cont)|Q(numero__icontains=num_ref_cont)|Q(concepto__icontains=num_ref_cont)|Q(contrato__numero__icontains=num_ref_cont))

	if referencia:
		qset = qset &(Q(referencia__icontains=referencia))

	if numero:
		qset = qset &(Q(numero__icontains=numero))

	if radicado:
		qset = qset &(Q(radicado=radicado))

	if fecha_desde:
		qset = qset &(Q(fecha__gte=fecha_desde))
	if fecha_hasta:
		qset = qset &(Q(fecha__lte=fecha_hasta))

	if tipo_contrato:
		qset = qset &(Q(contrato__tipo_contrato=tipo_contrato))
	if numero_contrato:
		qset = qset &(Q(contrato__numero=numero_contrato))
	if id_contratista:
		qset = qset &(Q(contrato__contratista_id=id_contratista))

	if id_empresa:
		qset = qset &(Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1))

	# print qset

	model=Factura
	formato_fecha = "%Y-%m-%d"
	if qset is not None:
		queryset = model.objects.filter(qset).order_by('-id')

		# serializer = self.get_serializer(queryset,many=True)
		# return Response({'message':'','success':'ok','data':serializer.data})

		# detalle = DetalleGiro.objects.filter(qset)

		worksheet.write('A1', 'Soporte', format1)
		worksheet.write('B1', 'Tipo de Servicio', format1)
		worksheet.write('C1', 'N Contrato', format1)
		worksheet.write('D1', 'Proyectos', format1)
		worksheet.write('E1', 'N factura', format1)
		worksheet.write('F1', 'Fecha', format1)
		worksheet.write('G1', 'Referencia', format1)
		worksheet.write('H1', 'Concepto', format1)
		worksheet.write('I1', 'Mes Causado', format1)
		worksheet.write('J1', 'Valor Factura', format1)
		worksheet.write('K1', 'Valor Pagar', format1)
		worksheet.write('L1', 'No. Radicado', format1)

	row=1
	col=0
	
	for factura in queryset:
		meses = ''
		nombre_p = ''
		cont_p = 0
		soporte = ''
		# if detalle.cuenta is not None:
		# 	cuenta_nombre= "Girar de la cuenta de %s No. %s ' - ' %s" % (detalle.cuenta.nombre , detalle.cuenta.numero, detalle.cuenta.fiduciaria) if detalle.cuenta.nombre is not None else ''

		# if detalle.cuenta is not None:
		# 	cuenta_referencia= detalle.encabezado.referencia  if detalle.encabezado.referencia is not None else ''	

		for mes_causado in factura.mes_causado():

			if meses != '':
				meses = meses+', '

			meses = meses+mes_causado.mes+'-'+mes_causado.ano

		# BUSCAR PROYECTOS DE LA FACTURA
		proyectos=factura.proyecto()
		cont_p=proyectos.count()
		if cont_p==1:
			nombre_p = proyectos[0]['proyecto__nombre']
		elif cont_p > 1:
			nombre_p = 'Varios Proyectos'
		else:
			nombre_p = 'Sin Proyectos'
		# for proyecto in factura.proyecto():
		# 	cont_p+=1
		# 	if nombre_p == '':
		# 		nombre_p = proyecto.proyecto.nombre

		# if cont_p > 1:
		# 	nombre_p = 'Varios Proyectos'
		# else:
		# 	nombre_p = 'Sin Proyectos'

		if factura.soporte:
			soporte = 'Cargado'
		else:
			soporte = 'Por cargar'

		worksheet.write(row, col,str(soporte),format2)
		worksheet.write(row, col+1,factura.contrato.tipo_contrato.nombre,format2)
		worksheet.write(row, col+2,factura.contrato.numero,format2)
		worksheet.write(row, col+3,nombre_p,format2)
		worksheet.write(row, col+4,factura.numero,format2)
		worksheet.write(row, col+5,factura.fecha,format5)
		worksheet.write(row, col+6,factura.referencia,format2)
		worksheet.write(row, col+7,factura.concepto,format2)
		worksheet.write(row, col+8,meses,format2)
		worksheet.write(row, col+9,factura.valor_factura,format2)
		worksheet.write(row, col+10,factura.valor_contable,format2)
		worksheet.write(row, col+11, factura.radicado if factura.radicado else '',format2)

		#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
		row +=1
	workbook.close()
	return response

# exporta a excel cesion
def exportReporteCesion(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_cesion.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cesion')

	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 14)
	worksheet.set_column('C:D', 30)
	worksheet.set_column('E:G', 15)
	worksheet.set_column('H:I', 20)

	# cursor = connection.cursor()

	ref_desc = None
	referencia = None
	id_beneficiario = None
	num_contrato = None
	fecha_desde = None
	fecha_hasta = None
	id_empresa = request.user.usuario.empresa.id

	if request.GET['dato']:
		ref_desc = request.GET['dato']

	if request.GET['referencia'] and request.GET['referencia'] != 'null':
		referencia = request.GET['referencia']

	if request.GET['id_beneficiario']:
		id_beneficiario = request.GET['id_beneficiario']

	if request.GET['numero_contrato'] and request.GET['numero_contrato'] != 'null':
		num_contrato = request.GET['numero_contrato']

	if request.GET['fecha_desde'] and request.GET['fecha_desde'] != 'null':
		fecha_desde = request.GET['fecha_desde']

	if request.GET['fecha_hasta'] and request.GET['fecha_hasta'] != 'null':
		fecha_hasta = request.GET['fecha_hasta']

	qset = (~Q(id=0))

	if ref_desc:
		qset = ( Q(referencia__icontains=ref_desc)|Q(descripcion__icontains=ref_desc) )

	if referencia:
		qset = qset &(Q(referencia__icontains=referencia))
		
	if fecha_desde:
		qset = qset &(Q(fecha__gte=fecha_desde))
		
	if fecha_hasta:
		qset = qset &(Q(fecha__lte=fecha_hasta))

	if num_contrato:
		qset = qset &(Q(contrato__numero__icontains=num_contrato))

	if id_beneficiario:
		qset = qset &(Q(beneficiario__id=id_beneficiario))

	if id_empresa:
		qset = qset &(
				#Q(empresa_contrato__empresa_id=id_empresa)
				Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
			)

	# print qset

	model=Cesion
	formato_fecha = "%Y-%m-%d"
	if qset is not None:
		queryset = model.objects.filter(qset).order_by('-id')

		# serializer = self.get_serializer(queryset,many=True)
		# return Response({'message':'','success':'ok','data':serializer.data})

		# detalle = DetalleGiro.objects.filter(qset)

		worksheet.write('A1', 'Soporte', format1)
		worksheet.write('B1', 'N Contrato', format1)
		worksheet.write('C1', 'Descripcion', format1)
		worksheet.write('D1', 'Beneficiario', format1)
		worksheet.write('E1', 'Referencia', format1)
		worksheet.write('F1', 'Valor', format1)
		worksheet.write('G1', 'Fecha', format1)
		worksheet.write('H1', 'Banco', format1)
		worksheet.write('I1', 'Numero de Cuenta', format1)

	row=1
	col=0
	
	for cesion in queryset:
		meses = ''
		# if detalle.cuenta is not None:
		# 	cuenta_nombre= "Girar de la cuenta de %s No. %s ' - ' %s" % (detalle.cuenta.nombre , detalle.cuenta.numero, detalle.cuenta.fiduciaria) if detalle.cuenta.nombre is not None else ''

		# if detalle.cuenta is not None:
		# 	cuenta_referencia= detalle.encabezado.referencia  if detalle.encabezado.referencia is not None else ''

		worksheet.write(row, col,str(cesion.soporte),format2)
		worksheet.write(row, col+1,cesion.contrato.numero,format2)
		worksheet.write(row, col+2,cesion.descripcion,format2)
		worksheet.write(row, col+3,cesion.beneficiario.nombre,format5)
		worksheet.write(row, col+4,cesion.referencia,format2)
		worksheet.write(row, col+5,cesion.valor,format2)
		worksheet.write(row, col+6,cesion.fecha,format5)
		if cesion.banco:
			worksheet.write(row, col+7,cesion.banco.nombre,format2)
		else:
			worksheet.write(row, col+7,'',format2)
		worksheet.write(row, col+8,cesion.numero_cuenta,format2)

		#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
		row +=1
	workbook.close()
	return response

# Eliminar cesion con una lista
@transaction.atomic
def destroyCesion(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']

			for item in myList:

				# Buscar si tiene Detalle_Giro
				m_d = DetalleGiro.objects.filter(cesion_id=item['id'])
				if m_d.exists():
					# print "Eliminanado detalle:"

					id_enc = m_d[0].encabezado.id
					id_detalle = m_d[0].id
					m_e = DetalleGiro.objects.filter(encabezado_id=id_enc)
					num_enc = +m_e.count()
					m_d[0].delete()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.detalle_giro',id_manipulado=id_detalle)
					logs_model.save()

					# Si el emcabezado del Giro tiene un solo Detalle, Eliminarlo.

					if num_enc == 1:
						# print "Eliminando encabezado"

						model_encabezado = DEncabezadoGiro.objects.get(pk=id_enc)
						id_encabezado = model_encabezado.id
						model_encabezado.delete()

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.encabezado_giro',id_manipulado=id_encabezado)
						logs_model.save()

				# Buscar si tiene sopote en proceso
				model_cesion = Cesion.objects.get(pk = item['id'])
				if model_cesion.proceso_soporte_id:

					model_sprd = HSoporteProcesoRelacionDato.objects.get(pk = model_cesion.proceso_soporte_id)

					# ACTUALIZAR EL PROCESO_RELACION_DATO
					queryset_prd = GProcesoRelacionDato.objects.get(pk = model_sprd.procesoRelacionDato_id)
					queryset_prd.valor = 'Vacio'
					queryset_prd.estado = 0
					queryset_prd.save()

					model_sprd.delete()

				# model_c = Cesion.objects.get(pk=item['id'])
				model_cesion.delete()

				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.Cesion',id_manipulado=item['id'])
				logs_model.save()

			# transaction.commit()
			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok','data': ''})

		except ProtectedError:
			return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
			transaction.savepoint_rollback(sid)

		except Exception as e:
			transaction.savepoint_rollback(sid)
			functions.toLog(e, 'Factura - destroyCesion')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Exporta a excel Descuento
def exportReporteDescuento(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_Descuento.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Descuento')

	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:A', 30)
	worksheet.set_column('B:B', 15)
	worksheet.set_column('C:C', 33)
	worksheet.set_column('D:E', 15)
	worksheet.set_column('F:G', 23)

	# cursor = connection.cursor()

	ref_conc = None
	referencia = None
	num_contrato = None
	concepto = None
	id_empresa = request.user.usuario.empresa.id

	if request.GET['dato']:
		ref_conc = request.GET['dato']

	if request.GET['referencia'] and request.GET['referencia'] != 'null':
		referencia = request.GET['referencia']

	if request.GET['numero_contrato'] and request.GET['numero_contrato'] != 'null':
		num_contrato = request.GET['numero_contrato']

	if request.GET['concepto'] and request.GET['concepto'] != 'null':
		concepto = request.GET['concepto']

	qset = (~Q(id=0))

	if ref_conc:
		qset = (
			Q(referencia__icontains=ref_conc)|Q(concepto__icontains=ref_conc)
			)

	if referencia:
		qset = qset &(Q(referencia__icontains=referencia))
		
	if concepto:
		qset = qset &(Q(concepto__icontains=concepto))

	if num_contrato:
		qset = qset &(Q(contrato__numero__icontains=num_contrato))

	if id_empresa:
		qset = qset &(
				#Q(empresa_contrato__empresa_id=id_empresa)
				Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
			)

	# print qset

	model=Descuento
	formato_fecha = "%Y-%m-%d"
	if qset is not None:
		queryset = model.objects.filter(qset).order_by('-id')

		# serializer = self.get_serializer(queryset,many=True)
		# return Response({'message':'','success':'ok','data':serializer.data})

		# detalle = DetalleGiro.objects.filter(qset)

		worksheet.write('A1', 'Soporte', format1)
		worksheet.write('B1', 'N Contrato', format1)
		worksheet.write('C1', 'Concepto', format1)
		worksheet.write('D1', 'Referencia', format1)
		worksheet.write('E1', 'Valor', format1)
		worksheet.write('F1', 'Banco', format1)
		worksheet.write('G1', 'Numero de Cuenta', format1)

	row=1
	col=0
	
	for descuento in queryset:
		meses = ''
		# if detalle.cuenta is not None:
		# 	cuenta_nombre= "Girar de la cuenta de %s No. %s ' - ' %s" % (detalle.cuenta.nombre , detalle.cuenta.numero, detalle.cuenta.fiduciaria) if detalle.cuenta.nombre is not None else ''

		# if detalle.cuenta is not None:
		# 	cuenta_referencia= detalle.encabezado.referencia  if detalle.encabezado.referencia is not None else ''

		worksheet.write(row, col,str(descuento.soporte),format2)
		worksheet.write(row, col+1,descuento.contrato.numero,format2)
		worksheet.write(row, col+2,descuento.concepto,format2)
		worksheet.write(row, col+3,descuento.referencia,format5)
		worksheet.write(row, col+4,descuento.valor,format2)
		if descuento.banco:
			worksheet.write(row, col+5,descuento.banco.nombre,format2)
		else:
			worksheet.write(row, col+5,'',format2)
		worksheet.write(row, col+6,descuento.numero_cuenta,format5)

		#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
		row +=1
	workbook.close()
	return response

# Eliminar Descuento con una lista
@transaction.atomic
def destroyDescuento(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']

			for item in myList:
				# # print item
				model_contrato = Descuento.objects.get(pk=item['id'])
				model_contrato.delete()

				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='factura.Descuento',id_manipulado=item['id'])
				logs_model.save()

			# transaction.commit()
			return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok','data': ''})
			transaction.savepoint_commit(sid)
		except Exception as e:
			transaction.savepoint_rollback(sid)
			functions.toLog(e, 'Factura - destroyDescuento')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Buscar Giros
def listAnticipo(request):
	cursor = connection.cursor()
	try:
		id_contrato=request.GET['id_contrato']
		cursor.callproc('[giro].[lista_giro]', [id_contrato,])
		#if cursor.return_value == 1:

		columns = cursor.description 
		result = [{columns[index][0]:column for index, column in enumerate(value)} for value in cursor.fetchall()]

		# result_set = cursor.fetchall()
		lista=[]
		for x in list(result):
			item={
				'id':x['id'],
				'soporte':x['soporte'],
				'referencia':x['referencia'],
				'nombre':x['nom_giro'],
				'valor':x['valor']
			}
			lista.append(item)
		return JsonResponse({'message':'','success':'ok','data':lista})
	except Exception as e:
		functions.toLog(e, 'Factura - listAnticipo')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})

# Crear detalle giro
@transaction.atomic
def crearDetalleGiro(request, id_cesion):
	sid = transaction.savepoint()
	try:
		model_encabezado = DEncabezadoGiro.objects.filter(contrato=request.DATA['contrato_id'], nombre=request.DATA['nombre_giro'])
		# model_encabezado = DEncabezadoGiro.objects.filter(contrato_id=5004, nombre_id=1002)

		if model_encabezado.exists():

			# Crear el Detalle del Giro
			m_d = DetalleGiro(encabezado_id=model_encabezado[0].id,
												contratista_id=request.DATA['beneficiario_id'],
												banco_id=request.DATA['banco_id'],
												no_cuenta=request.DATA['numero_cuenta'],
												tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
												valor_girar=request.DATA['valor'],
												estado_id=enumEstados.solicitado,
												test_op='',
												codigo_pago='',
												cesion_id=id_cesion)
			m_d.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.detalle_giro',id_manipulado=m_d.id)
			logs_model.save()
		else:
			# print "No hay encabezado"
			# Crear el Encabezado del Giro
			m_e = DEncabezadoGiro(nombre_id=request.DATA['nombre_giro'],
														contrato_id=request.DATA['contrato_id'],
														num_causacion='',
														numero_radicado='',
														pago_recurso_id = enumTipoPagoAnticipo.cuenta_bancaria)
			m_e.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.encabezado_giro',id_manipulado=m_e.id)
			logs_model.save()

			# Crear el Detalle del Giro
			m_d = DetalleGiro(encabezado_id=m_e.id,
												contratista_id=request.DATA['beneficiario_id'],
												banco_id=request.DATA['banco_id'],
												no_cuenta=request.DATA['numero_cuenta'],
												tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
												valor_girar=request.DATA['valor'],
												estado_id=enumEstados.solicitado,
												test_op='',
												codigo_pago='',
												cesion_id=id_cesion)
			m_d.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.detalle_giro',id_manipulado=m_d.id)
			logs_model.save()
		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearDetalleGiro')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Crear detalle giro2
@transaction.atomic
def crearDetalleGiro2(request, id_cruce):
	sid = transaction.savepoint()
	try:
		# Cargar info del encabeazdo del  giro
		model_encabezado = DEncabezadoGiro.objects.filter(contrato=request.DATA['contrato_id'])
		# model_encabezado = DEncabezadoGiro.objects.filter(contrato_id=5004, nombre_id=1002)

		# Cargar M-Contrato
		contrato = Contrato.objects.get(pk = request.DATA['contrato_id'])

		# Buscar el proyecto
		# proyecto = Proyecto.objects.filter()

		model_proyecto = Proyecto.objects.filter(contrato__id = contrato.id)
		vProyecto = 1
		retorno = '0'
		# print("No_cuenta:", model_proyecto[0].No_cuenta)
		# print("entidad_bancaria:", model_proyecto[0].entidad_bancaria)
		# print("tipo_cuenta:", model_proyecto[0].tipo_cuenta)

		# Validar datos bancarios del proyecto
		if not model_proyecto[0].entidad_bancaria:
			vProyecto = 0

		if not model_proyecto[0].tipo_cuenta:
			vProyecto = 0

		if vProyecto == 1:
			if contrato.mcontrato:

				# Buscar los nombres de giro del M-contrato
				nombreGiro = CNombreGiro.objects.filter(contrato = contrato.mcontrato.id)

				if nombreGiro:
					id_nombre_giro = 0
					retorno_id_nom_giro = 0
					id_anticipo = 0

					# Buscar el id de pago final
					for nomGiros in nombreGiro:
						# print("id_nombreGiro:", nomGiros.id)
						# print("NombreGiro:", nomGiros.nombre)

						cadena = nomGiros.nombre.lower()

						if cadena.find('final') >= 0:
						# if 'FInal' in str(nomGiros.nombre):
							# print "Encontro el pago final"
							id_nombre_giro = nomGiros.id
							retorno_id_nom_giro = 1

					if retorno_id_nom_giro == 1:

						if model_encabezado.exists():

							# Buscar los el ID del encabezado de pago final
							for encabezado in model_encabezado:
								if encabezado.nombre.id == id_nombre_giro:
									id_anticipo = encabezado.id

						if id_anticipo == 0:

							# Crear el Encabezado del Giro
							e_g = DEncabezadoGiro(nombre_id=id_nombre_giro,
																		contrato_id=request.DATA['contrato_id'],
																		num_causacion='',
																		numero_radicado='')
							e_g.save()
							# print "ID del encabezado: "+str(e_g.id)
							id_anticipo = e_g.id
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.encabezado_giro',id_manipulado = e_g.id)
							logs_model.save()

						if id_anticipo > 0:
							# print "Crear el Detalle del Giro"
							# Crear el Detalle del Giro
							id_banco = None
							id_tipo_cuenta = None
							valor = None

							if model_proyecto[0].entidad_bancaria:
								id_banco = model_proyecto[0].entidad_bancaria.id

							if model_proyecto[0].tipo_cuenta:
								id_tipo_cuenta = model_proyecto[0].tipo_cuenta.id

							valor = float(request.DATA['valor']) * -1

							m_d = DetalleGiro(encabezado_id=id_anticipo,
																contratista_id=contrato.contratista.id,
																banco_id=id_banco,
																no_cuenta=model_proyecto[0].No_cuenta,
																tipo_cuenta_id=id_tipo_cuenta,
																valor_girar=valor,
																estado_id=enumEstados.solicitado,
																test_op='',
																codigo_pago='',
																cruce_id=id_cruce)
							m_d.save()
							logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.detalle_giro',id_manipulado=m_d.id)
							logs_model.save()
							retorno = '1'
					else:
						# print "El M-Contrato no tiene nombre de Giros de pago final"
						retorno = "El M-Contrato no tiene nombre de giros de pago final."
				else:
					# print "El M-Contrato no tiene nombre de Giros"
					retorno = "El M-Contrato no tiene nombre de Giros."
			else:
				# print "El contrato Esta sin M-Contrato"
				retorno = "El contrato esta sin M-Contrato."
		else:
				# print "El proyecto no tiene los datos bancarios"
				retorno = "El proyecto no tiene los datos bancarios."

		return retorno
		# transaction.savepoint_commit(sid)
	except Exception as e:
		# transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearDetalleGiro2')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Crear Soporte Procesos para Cesion
@transaction.atomic
def crearSoporteProceso(request, id_autorizacion):
	sid = transaction.savepoint()
	try:
		id_empresa = request.user.usuario.empresa.id
		tipo_c = tipoC()
		querysetContrato=Contrato.objects.get(pk = request.DATA['contrato_id'])

		# print "t_c:"+str(querysetContrato.tipo_contrato.id)
		# print "t_c_enum:"+str(tipo_c.contratoProyecto)

		if querysetContrato.tipo_contrato.id == tipo_c.contratoProyecto:
			# print "es De Obra"

			querysetProyecto = Proyecto.objects.filter(contrato__id = request.DATA['contrato_id'],
																									contrato__empresacontrato__empresa=id_empresa,
																									contrato__empresacontrato__participa=1).first()
			# print "proyecto Mcontrato:"+str(querysetProyecto.mcontrato)

			# UPPER() CONVIERTE EL TEXTO EN MAYUSCULA
			str1 = querysetProyecto.mcontrato.nombre.upper()
			str2 = "PRONE"

			if str1.find(str2) >= 0:
				entro = 0
				if str1.find('2012') >= 0:
					id_item = 105
					id_proceso = 9
					entro = 1

				if str1.find('2013') >= 0 or str1.find('2014') >= 0:
					id_item = 47
					id_proceso = 1
					entro = 1
					# ejemplo
					# id_item = 1
					# id_proceso = 1

				if entro == 1:
					elemento_entidad = querysetProyecto.id

					# print "Nombre Proyecto:"+str(querysetProyecto.nombre)
					# print "id_proceso:"+str(id_proceso)
					# print "id_tabla_referencia:"+str(elemento_entidad)

					# BUSCAR EL ID DEL PROCESO_RELACION
					querysetProcesoRelacionDato = GProcesoRelacionDato.objects.filter(procesoRelacion__proceso = id_proceso,
																																						procesoRelacion__idApuntador = elemento_entidad,
																																						item = id_item).first()
					# print "id_relacion_datos:"+str(querysetProcesoRelacionDato)
					id_relacion_datos = querysetProcesoRelacionDato.id

					# ACTUALIZAR EL PROCESO_RELACION_DATO
					queryset_prd = GProcesoRelacionDato.objects.get(pk = id_relacion_datos)
					queryset_prd.valor = 'Si'
					queryset_prd.estado = 1
					queryset_prd.save()

					logs_model=Logs(
						usuario_id=request.user.usuario.id,
						accion=Acciones.accion_actualizar,
						nombre_modelo='Procesos.procesoRelacionDatos',
						id_manipulado=queryset_prd.id
					)
					logs_model.save()
					# GUARDAR EL SOPORTE DEL PROCESO_RELACION_DATO
					nombre = str(request.DATA['descripcion'])
					modelSoporteProcesoRelacionDato = HSoporteProcesoRelacionDato(procesoRelacionDato_id = id_relacion_datos,
																					nombre = nombre[0:99],
																					documento = request.FILES['soporte'])
					modelSoporteProcesoRelacionDato.save()

					# EDITAR LA SOLICITUD CREADA, PARA GUARDAR EL ID DEL SOPORTE_PROCESO_RELACION_DATO

					model_cesion = Cesion.objects.get(pk = id_autorizacion)
					model_cesion.proceso_soporte_id	= modelSoporteProcesoRelacionDato.id
					model_cesion.save()

					logs_model=''
					logs_model=Logs(
						usuario_id=request.user.usuario.id,
						accion=Acciones.accion_crear,
						nombre_modelo='Procesos.soporteProcesoRelacionDatos',
						id_manipulado=modelSoporteProcesoRelacionDato.id
					)
					logs_model.save()
				else:
					pass
					# print "No es 2012, 2013, 2014"
			else:
				pass
				# print "No es prone"

		else:
			pass
			# print "No es De Obra"

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearSoporteProceso')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Eliminar detalle giro
@transaction.atomic
def updateDetalleGiro(request):
	sid = transaction.savepoint()
	try:
		model_detalleGiro = DetalleGiro.objects.filter(cesion_id=request.data['id'])
		# model_encabezado = DEncabezadoGiro.objects.filter(contrato=request.DATA['contrato_id'], nombre=request.DATA['nombre_giro'])
		# # model_encabezado = DEncabezadoGiro.objects.filter(contrato_id=5004, nombre_id=1002)

		if model_detalleGiro.exists():
			m_d = DetalleGiro.objects.get(pk=model_detalleGiro[0].id)

			# Editar el Detalle del Giro
			m_d.contratista_id = request.DATA['beneficiario_id']
			m_d.banco_id = request.DATA['banco_id']
			m_d.no_cuenta = request.DATA['numero_cuenta']
			m_d.tipo_cuenta_id = request.DATA['tipo_cuenta_id']
			m_d.valor_girar = request.DATA['valor']
			m_d.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giro',id_manipulado=m_d.id)
			logs_model.save()

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - updateDetalleGiro')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@transaction.atomic
def updateSoporteProceso(request):
	sid = transaction.savepoint()
	try:
		model_cesion = Cesion.objects.get(pk = request.DATA['id'])
		if model_cesion.proceso_soporte_id:

			model_sprd = HSoporteProcesoRelacionDato.objects.get(pk = model_cesion.proceso_soporte_id)

			nombre = str(request.DATA['descripcion'])

			model_sprd.nombre = nombre[0:99]
			model_sprd.documento = request.FILES['soporte'] if request.FILES.get('soporte') is not None else model_sprd.documento
			model_sprd.save()

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - updateSoporteProceso')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# Eliminar compensacion con una lista
@transaction.atomic
def destroyCompensacion(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']

			for item in myList:
				# # print item['id']

				# Buscar si tiene Detalle_Giro
				m_d = DetalleGiro.objects.filter(cruce_id=item['id'])
				if m_d.exists():
					# print "Eliminanado detalle:"

					id_enc = m_d[0].encabezado.id
					id_detalle = m_d[0].id
					m_e = DetalleGiro.objects.filter(encabezado_id=id_enc)
					num_enc = +m_e.count()
					m_d[0].delete()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.detalle_giro',id_manipulado=id_detalle)
					logs_model.save()

					# Si el emcabezado del Giro tiene un solo Detalle, Eliminarlo.

					if num_enc == 1:
						# print "Eliminando encabezado"

						model_encabezado = DEncabezadoGiro.objects.get(pk=id_enc)
						id_encabezado = model_encabezado.id
						model_encabezado.delete()

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.encabezado_giro',id_manipulado=id_encabezado)
						logs_model.save()

				# Eliminar el detalle del cruce
				model_detalleCruce = DetalleCompensacion.objects.filter(compensacion_id=item['id'])

				# Buscar si tiene factura
				for detalleCruce in model_detalleCruce:
					if detalleCruce.tablaForanea == tablaForanea.factura:
						# Cambiar el estado de la factura de compensada a activa
						cambiarEstadoFactura2(detalleCruce.id, estadoFactura.activa)

				model_detalleCruce.delete()

				model_cruce = Compensacion.objects.get(pk=item['id'])
				model_cruce.delete()

			# transaction.commit()
			return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok','data': ''})
			transaction.savepoint_commit(sid)

		except ProtectedError:
			return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','status':'error','data':''})

		except Exception as e:
			transaction.savepoint_rollback(sid)
			functions.toLog(e, 'Factura - destroyCompensacion')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Guardar Detalle Compensacion
def createDetalleCompensacionLista(ids, tabla_foranea, id_compensacion):
	try:
		myList = ids.split(',')

		for item in myList:
			if item:
				model_detalleCompensacion = DetalleCompensacion(id_registro=item,compensacion_id=id_compensacion,tablaForanea_id=tabla_foranea)
				model_detalleCompensacion.save()

				if tabla_foranea == tablaForanea.factura:
					# print "es factura"
					cambiarEstadoFactura2(item, estadoFactura.compensada)
					# print "es factura2"

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data': ''})
	except Exception as e:
		functions.toLog(e, 'Factura - createDetalleCompensacionLista')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# exporta a excel Compensacion
def exportReporteCompensacion(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_Compensacion.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Compensacion')

	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:D', 14)
	worksheet.set_column('E:E', 33)
	worksheet.set_column('F:G', 14)
	worksheet.set_column('H:H', 30)
	worksheet.set_column('I:I', 15)
	worksheet.set_column('J:J', 33)

	# cursor = connection.cursor()

	ref_desc = None
	id_empresa = request.user.usuario.empresa.id

	if request.GET['dato']:
		ref_desc = request.GET['dato']

	qset = None
	if ref_desc:
		qset = (
			Q(referencia__icontains=ref_desc)|Q(descripcion__icontains=ref_desc)|Q(contrato__numero__icontains=ref_desc)
			)

	if id_empresa:
		if qset != None:
			qset = qset &(
				#Q(empresa_contrato__empresa_id=id_empresa)
				Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
			)
		else:
			qset = (
				#Q(empresa_contrato__empresa_id=id_empresa)
				Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1)
			)

	# print qset

	model=Compensacion
	formato_fecha = "%Y-%m-%d"
	if qset is not None:
		queryset = model.objects.filter(qset).order_by('-id')

		# serializer = self.get_serializer(queryset,many=True)
		# return Response({'message':'','success':'ok','data':serializer.data})

		worksheet.write('A1', 'Fecha', format1)
		worksheet.write('B1', 'Referencia', format1)
		worksheet.write('C1', 'Tipo Documento', format1)
		worksheet.write('D1', 'No. Factura', format1)
		worksheet.write('E1', 'Descripcion', format1)
		worksheet.write('F1', 'Valor', format1)
		worksheet.write('G1', 'Contrato', format1)
		worksheet.write('H1', 'Contratista', format1)
		worksheet.write('I1', 'Nit', format1)
		worksheet.write('J1', 'Tipo Contrato', format1)

	row=1
	col=0
	
	for cruce in queryset:

		# if detalle.cuenta is not None:
		# 	cuenta_nombre= "Girar de la cuenta de %s No. %s ' - ' %s" % (detalle.cuenta.nombre , detalle.cuenta.numero, detalle.cuenta.fiduciaria) if detalle.cuenta.nombre is not None else ''

		# if detalle.cuenta is not None:
		# 	cuenta_referencia= detalle.encabezado.referencia  if detalle.encabezado.referencia is not None else ''

		# Llenar Compensaciom
		worksheet.write(row, col,str(cruce.fecha), format2)
		worksheet.write(row, col+1,cruce.referencia, format2)
		worksheet.write(row, col+2,'Cruce', format2)
		worksheet.write(row, col+3,'N/A', format5)
		worksheet.write(row, col+4,cruce.descripcion, format2)
		worksheet.write(row, col+5,cruce.valor, format2)
		worksheet.write(row, col+6,cruce.contrato.numero, format5)
		worksheet.write(row, col+7,cruce.contrato.contratista.nombre, format2)
		worksheet.write(row, col+8,cruce.contrato.contratista.nit, format2)
		worksheet.write(row, col+9,cruce.contrato.tipo_contrato.nombre, format2)
		
		for item in cruce.detalle_cruce():

			if item.tablaForanea.id == tablaForanea.encabezadoGiros:
				
				row +=1
				giro = DEncabezadoGiro.objects.get(pk=item.id_registro)

				sumatoria=DetalleGiro.objects.filter(encabezado_id=item.id_registro).aggregate(suma_detalle=Sum('valor_girar'))
				# print("summ:",sumatoria)

				if giro:
					worksheet.write(row, col,str(giro.fecha_conta), format2)
					worksheet.write(row, col+1,giro.referencia, format2)
					worksheet.write(row, col+2,'Anticipo', format2)
					worksheet.write(row, col+3,'N/A', format5)
					worksheet.write(row, col+4,giro.nombre.nombre, format2)
					worksheet.write(row, col+5,sumatoria['suma_detalle'], format2)
					worksheet.write(row, col+6,giro.contrato.numero, format5)
					worksheet.write(row, col+7,giro.contrato.contratista.nombre, format2)
					worksheet.write(row, col+8,giro.contrato.contratista.nit, format2)
					worksheet.write(row, col+9,giro.contrato.tipo_contrato.nombre, format2)

			if item.tablaForanea.id == tablaForanea.factura:
				
				row +=1
				factura = Factura.objects.get(pk=item.id_registro)

				worksheet.write(row, col,str(factura.fecha), format2)
				worksheet.write(row, col+1,factura.referencia, format2)
				worksheet.write(row, col+2,'Factura', format2)
				worksheet.write(row, col+3,factura.numero, format5)
				worksheet.write(row, col+4,factura.concepto, format2)
				worksheet.write(row, col+5,factura.valor_contable, format2)
				worksheet.write(row, col+6,factura.contrato.numero, format5)
				worksheet.write(row, col+7,factura.contrato.contratista.nombre, format2)
				worksheet.write(row, col+8,factura.contrato.contratista.nit, format2)
				worksheet.write(row, col+9,factura.contrato.tipo_contrato.nombre, format2)

			if item.tablaForanea.id == tablaForanea.cesion:
				
				row +=1
				cesion = Cesion.objects.get(pk=item.id_registro)

				worksheet.write(row, col,str(cesion.fecha), format2)
				worksheet.write(row, col+1,cesion.referencia, format2)
				worksheet.write(row, col+2,'Autorización de Giro', format2)
				worksheet.write(row, col+3,'N/A', format5)
				worksheet.write(row, col+4,cesion.descripcion, format2)
				worksheet.write(row, col+5,cesion.valor, format2)
				worksheet.write(row, col+6,cesion.contrato.numero, format5)
				worksheet.write(row, col+7,cesion.contrato.contratista.nombre, format2)
				worksheet.write(row, col+8,cesion.contrato.contratista.nit, format2)
				worksheet.write(row, col+9,cesion.contrato.tipo_contrato.nombre, format2)

			if item.tablaForanea.id == tablaForanea.descuento:
				
				row +=1
				descuento = Descuento.objects.get(pk=item.id_registro)

				worksheet.write(row, col,'', format2)
				worksheet.write(row, col+1,descuento.referencia, format2)
				worksheet.write(row, col+2,'Descuento', format2)
				worksheet.write(row, col+3,'N/A', format5)
				worksheet.write(row, col+4,descuento.concepto, format2)
				worksheet.write(row, col+5,descuento.valor, format2)
				worksheet.write(row, col+6,descuento.contrato.numero, format5)
				worksheet.write(row, col+7,descuento.contrato.contratista.nombre, format2)
				worksheet.write(row, col+8,descuento.contrato.contratista.nit, format2)
				worksheet.write(row, col+9,descuento.contrato.tipo_contrato.nombre, format2)

			#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
		row +=1
	workbook.close()
	return response

# exporta a excel la planilla de proyecto_factura
def exportPlantilla(request):
	tipo_c = tipoC()
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Planilla_facturaProyecto.xlsx"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Proyectos')

	format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
	format1.set_align('center')
	format1.set_align('vcenter')
	format2=workbook.add_format({'border':0})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.set_column('A:B', 0)
	worksheet.set_column('C:C', 40)
	worksheet.set_column('D:E', 15)
	worksheet.set_column('F:G', 17)

	mcontrato = None
	id_empresa = request.user.usuario.empresa.id
	cursor = connection.cursor()

	if request.GET['mcontrato']:
		mcontrato = request.GET['mcontrato']

	# cursor.callproc('[factura].[planilla_factura_proyecto]', [mcontrato, tipo_c.interventoria])
	# result_set = cursor.fetchall()

	# model_factura = Factura.objects.filter(contrato__empresacontrato__empresa=id_empresa).values('id', 'numero', 'valor_factura', 'contrato')
	# result_set = model_factura.contrato.filter(tipo_contrato_id = tipo_c.interventoria)

	# print "tipo contrato: "+str(tipo_c.interventoria)
	model_proyecto = Proyecto.objects.filter(mcontrato_id=mcontrato,
											contrato__tipo_contrato_id = tipo_c.interventoria,
											contrato__empresacontrato__empresa=id_empresa).values('id', 'nombre',
																								'contrato','contrato__numero',
																								'contrato__fk_factura_contrato','contrato__fk_factura_contrato__numero',
																								'contrato__fk_factura_contrato__valor_factura')
	# print model_proyecto.query

	worksheet.write('A1', 'Cod. 1', format1)# Id del proyecto
	worksheet.write('B1', 'Cod. 2', format1)# Id de la factura
	worksheet.write('C1', 'Nombre Proyecto', format1)
	worksheet.write('D1', 'Num. Contrato', format1)
	worksheet.write('E1', 'Num. Factura', format1)
	worksheet.write('F1', 'Valor Factura', format1)
	worksheet.write('G1', 'Valor Proyecto', format1)

	row=1
	col=0
	for proyecto in model_proyecto:

		if(proyecto['contrato__fk_factura_contrato']):
			worksheet.write(row, col,proyecto['id'],format2)
			worksheet.write(row, col+1,proyecto['contrato__fk_factura_contrato'],format2)
			worksheet.write(row, col+2,proyecto['nombre'],format2)
			worksheet.write(row, col+3,proyecto['contrato__numero'],format5)
			worksheet.write(row, col+4,proyecto['contrato__fk_factura_contrato__numero'],format2)
			worksheet.write(row, col+5,proyecto['contrato__fk_factura_contrato__valor_factura'],format2)
			row +=1
			#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
	workbook.close()
	return response
	# return JsonResponse(list(model_proyecto), safe=False)

# exporta a excel Cruce - SOLICITUD DE CRUCE DE REGISTROS CONTABLES
def exportReporteCruce(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Reporte_Cruce.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cruce')

	# IMAGE
	worksheet.insert_image('J2', settings.BASE_DIR+'/static/images/icono_eca.png')
	# worksheet.insert_image('J2', 'C:/AppServ/www/sinin4/Scripts/SININWEB/static/images/img_eca.png')

	# print settings.BASE_DIR

	format1=workbook.add_format({'border':0,
															 'font_size':10,
															 'bold':True,
															 'align':'center',
														   'valign':'vcenter',})

	format1.set_font_name('DIN-Regular')

	format2=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':False,
															 # 'align':'center',
														   'valign':'vright',
														   'bg_color':'#1F497D',
														   'font_color': '#ffffff'})
	# format2.set_font_color('#1F497D')
	format2.set_font_name('DIN-Regular')

	format3=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':False,
														   'valign':'vright',})
	format3.set_font_name('DIN-Regular')

	titulo=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':False,
															 'align':'center',
														   'valign':'vright',
														   'bg_color':'#1F497D',
														   'font_color': '#ffffff'})
	titulo.set_font_name('DIN-Regular')

	titulo2=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':False,
															 'align':'center',
														   'valign':'vright',
														   'bg_color':'#DCE6F1'})
	titulo2.set_font_name('DIN-Regular')

	subtotal=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':False,
															 'align':'right',
														   'valign':'vright',
														   'bg_color':'#1F497D',
														   'font_color': '#ffffff'})
	subtotal.set_font_name('DIN-Regular')

	pregunta=workbook.add_format({'border':False,
															 'font_size':10,
															 'bold':False,
															 'align':'center',
														   'valign':'vright',
														   'bg_color':'#DCE6F1'})
	pregunta.set_font_name('DIN-Regular')

	firma=workbook.add_format({'border':False,
															 'font_size':10,
															 'bold':True,
															 'align':'center',
														   'valign':'vcenter',})
	firma.set_font_name('DIN-Regular')

	firma2=workbook.add_format({'border':False,
															 'font_size':10,
															 'bold':False,
															 'align':'center',
														   'valign':'vcenter',})
	firma2.set_font_name('DIN-Regular')

	firma3=workbook.add_format({'border':True,
															 'font_size':10,
															 'bold':True,
															 'align':'center',
														   'valign':'vcenter',})
	firma3.set_font_name('DIN-Regular')

	format_money=workbook.add_format({'border':True,
																		'font_size':10,
																		'bold':False,
																		'valign':'vright',
																		'num_format': '$#,##0'})
	format_money.set_font_name('DIN-Regular')

	merge_format = workbook.add_format({'bold':     True,
																			'border':   6,
																			'align':    'center',
																			'valign':   'vcenter',
																			'fg_color': '#D7E4BC',})

	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	worksheet.merge_range('B1:I1', 'ELECTRICARIBE S.A ESP',format1)
	worksheet.merge_range('B2:I2', 'UNIDAD NORMALIZACION DE REDES', format1)
	worksheet.merge_range('B3:I3', 'SOLICITUD DE CRUCE DE REGISTROS CONTABLES', format1)

	id_cruce = None
	# id_empresa = request.user.usuario.empresa.id

	if request.GET['dato']:
		id_cruce = request.GET['dato']

	model=Compensacion
	# queryset = model.objects.filter(qset).order_by('-id')
	queryset = model.objects.get(pk=id_cruce)

	# print("q:",queryset.valor)
	# print queryset

	hoy = date.today()
	formato_fecha = "%Y-%m-%d"
	# hoy = datetime.strptime(str(hoy), formato_fecha)

	worksheet.set_column('B:B', 12)
	worksheet.merge_range('B5:C5', 'FECHA SOLICITUD:', format2)
	worksheet.merge_range('B6:C6', 'NIT ACREEDOR:', format2)
	worksheet.merge_range('B7:C7', 'ACREEDOR SAP:', format2)
	worksheet.merge_range('B8:C8', 'NOMBRE ACREEDOR:', format2)
	worksheet.merge_range('B9:C9', 'CONTRATO:', format2)
	worksheet.merge_range('B10:C10', 'CONVENIO ECA-MME:', format2)
	worksheet.merge_range('B11:C11', 'CONCEPTO:', format2)

	worksheet.merge_range('D5:L5', str(hoy), format3)
	worksheet.merge_range('D6:L6', queryset.contrato.contratista.nit, format3)
	worksheet.merge_range('D7:L7', '0', format3)
	worksheet.merge_range('D8:L8', queryset.contrato.contratista.nombre, format3)
	worksheet.merge_range('D9:L9', queryset.contrato.numero, format3)
	# worksheet.merge_range('D10:J10', queryset.contrato.mcontrato.nombre, format3)
	worksheet.merge_range('D10:L10', queryset.contrato.mcontrato.nombre if queryset.contrato.mcontrato is not None else '',format3)
	worksheet.merge_range('D11:L11', queryset.descripcion, format3)

	# ANTICIPOS
	worksheet.merge_range('B13:L13', 'ANTICIPOS Y SALDOS A LEGALIZAR', titulo)

	worksheet.merge_range('B14:C14', 'Codigo SAP', titulo2)
	worksheet.merge_range('D14:J14', 'Nombre', titulo2)
	worksheet.merge_range('K14:L14', 'Valor Contabilizado', titulo2)

	row = 15
	total_anticipo = 0
	for item in queryset.detalle_cruce():

		if item.tablaForanea.id == tablaForanea.encabezadoGiros:

			giro = DEncabezadoGiro.objects.get(pk=item.id_registro)

			sumatoria=DetalleGiro.objects.filter(encabezado_id=item.id_registro).aggregate(suma_detalle=Sum('valor_girar'))

			worksheet.merge_range('B'+str(row)+':C'+str(row), giro.referencia, format3)
			worksheet.merge_range('D'+str(row)+':J'+str(row), giro.nombre.nombre, format3)
			worksheet.merge_range('K'+str(row)+':L'+str(row), sumatoria['suma_detalle'], format_money)
			total_anticipo += sumatoria['suma_detalle']
			row +=1

	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SUBTOTAL', subtotal)
	worksheet.merge_range('K'+str(row)+':L'+str(row), total_anticipo, format_money)
	
	# CESIONES
	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'CESIONES A PROVEEDORES', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':C'+str(row), 'Codigo SAP', titulo2)
	worksheet.merge_range('D'+str(row)+':J'+str(row), 'Descripcion', titulo2)
	worksheet.merge_range('K'+str(row)+':L'+str(row), 'Valor Contabilizado', titulo2)
	
	row +=1
	total_cesion = 0
	
	for item in queryset.detalle_cruce():
		if item.tablaForanea.id == tablaForanea.cesion:

			cesion = Cesion.objects.get(pk=item.id_registro)

			worksheet.merge_range('B'+str(row)+':C'+str(row), cesion.referencia, format3)
			worksheet.merge_range('D'+str(row)+':J'+str(row), cesion.descripcion, format3)
			worksheet.merge_range('K'+str(row)+':L'+str(row), cesion.valor, format_money)
			total_cesion += cesion.valor
			row +=1

	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SUBTOTAL', subtotal)
	worksheet.merge_range('K'+str(row)+':L'+str(row), total_cesion, format_money)
	
	# FACTURA
	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'FACTURAS PRESENTADAS', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':C'+str(row), 'Codigo SAP', titulo2)
	worksheet.merge_range('D'+str(row)+':J'+str(row), 'No. Factura', titulo2)
	worksheet.merge_range('K'+str(row)+':L'+str(row), 'Valor Contabilizado', titulo2)
	
	row +=1
	total_factura = 0
	
	for item in queryset.detalle_cruce():
		if item.tablaForanea.id == tablaForanea.factura:

			factura = Factura.objects.get(pk=item.id_registro)

			worksheet.merge_range('B'+str(row)+':C'+str(row), factura.referencia, format3)
			worksheet.merge_range('D'+str(row)+':J'+str(row), factura.numero, format3)
			worksheet.merge_range('K'+str(row)+':L'+str(row), factura.valor_contable, format_money)
			total_factura += factura.valor_contable
			row +=1

	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SUBTOTAL', subtotal)
	worksheet.merge_range('K'+str(row)+':L'+str(row), total_factura, format_money)

	# DESCUENTO
	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'DESCUENTOS', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':C'+str(row), 'Referencia', titulo2)
	worksheet.merge_range('D'+str(row)+':J'+str(row), 'Concepto', titulo2)
	worksheet.merge_range('K'+str(row)+':L'+str(row), 'Valor', titulo2)
	
	row +=1
	total_descuento = 0

	for item in queryset.detalle_cruce():
		if item.tablaForanea.id == tablaForanea.descuento:

			descuento = Descuento.objects.get(pk=item.id_registro)

			worksheet.merge_range('B'+str(row)+':C'+str(row), descuento.referencia, format3)
			worksheet.merge_range('D'+str(row)+':J'+str(row), descuento.concepto, format3)
			worksheet.merge_range('K'+str(row)+':L'+str(row), descuento.valor, format_money)
			total_descuento += descuento.valor
			row +=1

	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SUBTOTAL', subtotal)
	worksheet.merge_range('K'+str(row)+':L'+str(row), total_descuento, format_money)

	# MULTAS
	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'MULTAS Y SANCIONES', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':C'+str(row), 'Codigo SAP', titulo2)
	worksheet.merge_range('D'+str(row)+':J'+str(row), 'No. Factura', titulo2)
	worksheet.merge_range('K'+str(row)+':L'+str(row), 'Valor contabilizado', titulo2)
	
	row +=1
	total_multas = 0

	for item in queryset.detalle_cruce():
		if item.tablaForanea.id == tablaForanea.multa:

			multa = Solicitud.objects.get(pk=item.id_registro)

			worksheet.merge_range('B'+str(row)+':C'+str(row), multa.codigoReferencia, format3)
			worksheet.merge_range('D'+str(row)+':J'+str(row), 'Multa por incumplimiento', format3)
			worksheet.merge_range('K'+str(row)+':L'+str(row), multa.valorImpuesto, format_money)
			total_multas += multa.valorImpuesto
			row +=1

	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SUBTOTAL', subtotal)
	worksheet.merge_range('K'+str(row)+':L'+str(row), total_multas, format_money)
	
	# sumatoria
	# TOTAL DE (-)
	total_n = total_factura;
	# TOTAL DE (+)
	total_p = total_anticipo + total_cesion +  total_descuento + total_multas;
	total = total_p - total_n;

	row +=2
	if total < 0:
		msn = 'SALDO A FAVOR DEL ACREEDOR'
	elif total > 0:
		msn = 'SALDO EN CONTRA DEL ACREEDOR'
	else:
		msn = 'SIN SALDO PENDIENTE'

	worksheet.merge_range('K'+str(row)+':L'+str(row), total, format3)
	worksheet.merge_range('B'+str(row)+':J'+str(row), 'SALDO A FAVOR DEL ACREEDOR', subtotal)

	row +=3
	worksheet.merge_range('B'+str(row)+':H'+str(row), 'GIRAR SALDO A FAVOR DEL ACREEDOR', format2)
	# worksheet.merge_range('I'+str(row)+':I'+str(row), 'Si', format2)
	# worksheet.merge_range('J'+str(row)+':J'+str(row), 'No', format2)
	if total < 0:
		worksheet.write('I'+str(row)+':I', 'Si', pregunta)
		worksheet.write('J'+str(row)+':J', 'No')
	else:
		worksheet.write('I'+str(row)+':I', 'Si')
		worksheet.write('J'+str(row)+':J', 'No', pregunta)

	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'ORIGEN DE LOS RECURSOS', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Entidad:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Tipo:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'No. De Cuenta / Encargo Fid.', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)

	row +=2
	worksheet.merge_range('B'+str(row)+':L'+str(row), 'DATOS BANCARIOS DE LA CUENTA DESTINO DEL SALDO', titulo)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'NIT:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), queryset.contrato.contratista.nit, format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Beneficiario:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), queryset.contrato.contratista.nombre, format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Entidad Bancaria:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Tipo de Cuenta::', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'No. De Cuenta:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)
	row +=1
	worksheet.merge_range('B'+str(row)+':F'+str(row), 'Saldo neto a girar:', format3)
	worksheet.merge_range('G'+str(row)+':L'+str(row), '', format3)


	row +=3
	i = row
	i += 2
	worksheet.merge_range('B'+str(row)+':C'+str(i), 'ELABORA', firma)
	worksheet.merge_range('D'+str(row)+':H'+str(i), 'Carlos Acosta Castillo', firma2)
	worksheet.merge_range('I'+str(row)+':L'+str(i), '', format3)

	row +=4
	i = row
	i += 2
	worksheet.merge_range('B'+str(row)+':C'+str(i), 'REVISA', firma)
	worksheet.merge_range('D'+str(row)+':H'+str(i), 'Jorge Fierro Correa', firma2)
	worksheet.merge_range('I'+str(row)+':L'+str(i), '', format3)

	row +=4
	i = row
	i += 2
	worksheet.merge_range('B'+str(row)+':C'+str(i), 'VoBo', firma)
	worksheet.merge_range('D'+str(row)+':H'+str(i), 'Victor Paternina Novoa', firma2)
	worksheet.merge_range('I'+str(row)+':L'+str(i), '', format3)

	row +=4
	i = row
	i += 2
	worksheet.merge_range('B'+str(row)+':F'+str(i), 'AUTORIZACION AGENTE ESPECIAL ELECTRICARIBE', firma)
	worksheet.merge_range('G'+str(row)+':I'+str(i), 'Javier Lastra Fuscaldo', firma2)
	worksheet.merge_range('J'+str(row)+':L'+str(i), '', format3)

	#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)
	workbook.close()
	return response

def insert_rows(self, row_idx, cnt, above=True, copy_style=True, fill_formulae=False):
	"""Inserts new (empty) rows into worksheet at specified row index.

	:param row_idx: Row index specifying where to insert new rows.
	:param cnt: Number of rows to insert.
	:param above: Set True to insert rows above specified row index.
	:param copy_style: Set True if new rows should copy style of immediately above row.
	:param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.

	Usage:

	* insert_rows(2, 10, above=True, copy_style=False)

	"""
	CELL_RE  = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

	row_idx = row_idx - 1 if above else row_idx

	def replace(m):
		row = m.group('row')
		prefix = "$" if row.find("$") != -1 else ""
		row = int(row.replace("$",""))
		row += cnt if row > row_idx else 0
		return m.group('col') + prefix + str(row)

	# First, we shift all cells down cnt rows...
	old_cells = set()
	old_fas   = set()
	new_cells = dict()
	new_fas   = dict()
	for c in self._cells.values():

		old_coor = c.coordinate

		# Shift all references to anything below row_idx
		if c.data_type == Cell.TYPE_FORMULA:
			c.value = CELL_RE.sub(
				replace,
				c.value
			)
			# Here, we need to properly update the formula references to reflect new row indices
			if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
				self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
					replace,
					self.formula_attributes[old_coor]['ref']
				)

		# Do the magic to set up our actual shift
		if c.row > row_idx:
			old_coor = c.coordinate
			old_cells.add((c.row,c.col_idx))
			c.row += cnt
			new_cells[(c.row,c.col_idx)] = c
			if old_coor in self.formula_attributes:
				old_fas.add(old_coor)
				fa = self.formula_attributes[old_coor].copy()
				new_fas[c.coordinate] = fa

	for coor in old_cells:
		del self._cells[coor]
	self._cells.update(new_cells)

	for fa in old_fas:
		del self.formula_attributes[fa]
	self.formula_attributes.update(new_fas)

	# Next, we need to shift all the Row Dimensions below our new rows down by cnt...
	for row in range(len(self.row_dimensions)-1+cnt,row_idx+cnt,-1):
		new_rd = copy.copy(self.row_dimensions[row-cnt])
		new_rd.index = row
		self.row_dimensions[row] = new_rd
		del self.row_dimensions[row-cnt]

	# Now, create our new rows, with all the pretty cells
	row_idx += 1
	for row in range(row_idx,row_idx+cnt):
		# Create a Row Dimension for our new row
		new_rd = copy.copy(self.row_dimensions[row-1])
		new_rd.index = row
		self.row_dimensions[row] = new_rd
		for col in range(1,self.max_column):
			col = get_column_letter(col)
			cell = self.cell('%s%d'%(col,row))
			cell.value = None
			source = self.cell('%s%d'%(col,row-1))
			if copy_style:
				cell.number_format = source.number_format
				cell.font      = source.font.copy()
				cell.alignment = source.alignment.copy()
				cell.border    = source.border.copy()
				cell.fill      = source.fill.copy()
			if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
				s_coor = source.coordinate
				if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
					fa = self.formula_attributes[s_coor].copy()
					self.formula_attributes[cell.coordinate] = fa
				# print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
				cell.value = re.sub(
					"(\$?[A-Z]{1,3}\$?)%d"%(row - 1),
					lambda m: m.group(1) + str(row),
					source.value
				)   
				cell.data_type = Cell.TYPE_FORMULA

	# Check for Merged Cell Ranges that need to be expanded to contain new cells
	for cr_idx, cr in enumerate(self.merged_cell_ranges):
		self.merged_cell_ranges[cr_idx] = CELL_RE.sub(
			replace,
			cr
		)
# Worksheet.insert_rows = insert_rows

# exporta a excel Cruce - SOLICITUD DE CRUCE DE REGISTROS CONTABLES
def exportReporteCruce2(request):
	try:
		# wb = load_workbook(filename = settings.BASE_DIR+'/static/documentos/Solicitud de cruce de registros contables.xlsx', read_only=False)
		wb = load_workbook(filename = settings.BASE_DIR+'/static/documentos/Solicitud de cruce de registros contables.xltx', read_only=False)
		# ws = wb['Worksheet 1']
		ws = wb.worksheets[0]

		# wb = load_workbook(filename = settings.BASE_DIR+'/static/documentos/Libro1.xlsx', read_only=False)
		# ws = wb['Hoja1']

		# ws = wb.active

		# worksheet.merge_range('B1:I1', 'ELECTRICARIBE S.A ESP',format1)
		# ws.write('A1', 'Fehca')

		# IMAGE
		img = Image(settings.BASE_DIR+'/static/images/img_eca.png')
		ws.add_image(img, 'J2')

		id_cruce = None
		# id_empresa = request.user.usuario.empresa.id

		if request.GET['dato']:
			id_cruce = request.GET['dato']

		model=Compensacion
		# queryset = model.objects.filter(qset).order_by('-id')
		queryset = model.objects.get(pk=id_cruce)

		# print("q:",queryset.valor)
		# print queryset

		hoy = date.today()
		formato_fecha = "%Y-%m-%d"
		# hoy = datetime.strptime(str(hoy), formato_fecha)
		
		ws['C5'] = str(hoy)
		ws['C6'] = queryset.contrato.contratista.nit
		ws['C8'] = queryset.contrato.contratista.nombre
		ws['C9'] = queryset.contrato.numero
		ws['C10'] = queryset.contrato.mcontrato.nombre if queryset.contrato.mcontrato is not None else ''
		ws['C11'] = queryset.descripcion

		# highlight = NamedStyle(name="highlight")
		# bd = Side(style='thin', color=colors.RED)
		# highlight = Border(left=bd, top=bd, right=bd, bottom=bd)
		# ws.cell('D5').border = highlight
		# # ws.unmerge_cells('C5:H5')

		# ws.merge_cells('C5:H5')

		# ws.append([row[1], row[3]])


		# insert_rows(self, 2, 3, above=False, copy_style=True, fill_formulae=True)
		ws.insert_rows(15, 3, above=True, copy_style=True)
		# ws.insert_rows = insert_rows(2, 10, above=True, copy_style=False)

		wb.template = False
		response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
		response['Content-Disposition'] ='attachment; filename="Reporte_Cruce.xlsx"'
		return response
	except Exception as e:
		functions.toLog(e, 'Factura - exportReporteCruce2')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# def insert_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
	# 	RE_CELL  = re.compile("(?P<col>[A-Z]+)(?P<row>\d+)")
	# 	RE_RANGE = re.compile("(?P<s_col>[A-Z]+)(?P<s_row>\d+):(?P<e_col>[A-Z]+)(?P<e_row>\d+)")

	# 	row_idx = row_idx - 1 if above else row_idx

	# 	# First, we shift all cells down cnt rows...
	# 	old_cells = set()
	# 	old_fas   = set()
	# 	new_cells = dict()
	# 	new_fas   = dict()
	# 	for c in self._cells.values():
	# 		if c.row > row_idx:
	# 			old_coor = c.coordinate
	# 			old_cells.add((c.row,c.col_idx))
	# 			c.row += cnt
	# 			new_cells[(c.row,c.col_idx)] = c
	# 			if c.data_type == Cell.TYPE_FORMULA:
	# 				#print("Updating formula in cell %s%d to match %s%d"%(c.column,c.row-cnt,c.column,c.row))
	# 				c.value = re.sub(
	# 					"(\$?[A-Z]{1,3})\$?%d"%(c.row-cnt),
	# 					lambda m: m.group(1) + str(c.row),
	# 					c.value
	# 				)
	# 				# Here, we need to properly update the formula references to reflect new row indices
	# 				if old_coor in self.formula_attributes:
	# 					old_fas.add(old_coor)
	# 					fa = self.formula_attributes[old_coor].copy()
	# 					if 'ref' in fa:
	# 						# Handle case where cell references itself!
	# 						if fa['ref'] == old_coor:
	# 							fa['ref'] = c.coordinate
	# 						# Otherwise, we need to shift the range reference down by cnt
	# 						else:
	# 							g = RE_RANGE.search(fa['ref']).groupdict()
	# 							fa['ref'] = g['s_col'] + str(int(g['s_row'])+cnt) + ":" + g['e_col'] + str(int(g['e_row'])+cnt)
	# 					new_fas[c.coordinate] = fa

	# 	for coor in old_cells:
	# 		del self._cells[coor]
	# 	self._cells.update(new_cells)

	# 	for fa in old_fas:
	# 		del self.formula_attributes[fa]
	# 	self.formula_attributes.update(new_fas)

	# 	# Next, we need to shift all the Row Dimensions below out new rows down by cnt...
	# 	for row in range(len(self.row_dimensions)-1+cnt,row_idx+cnt,-1):
	# 		new_rd = copy.copy(self.row_dimensions[row-cnt])
	# 		new_rd.index = row
	# 		self.row_dimensions[row] = new_rd
	# 		del self.row_dimensions[row-cnt]

	# 	# Now, create our new rows, with all the pretty cells
	# 	row_idx += 1
	# 	for row in range(row_idx,row_idx+cnt):
	# 		# Create a Row Dimension for our new row
	# 		new_rd = copy.copy(self.row_dimensions[row-1])
	# 		new_rd.index = row
	# 		self.row_dimensions[row] = new_rd
	# 		for col in range(1,ws.max_column):
	# 			col = get_column_letter(col)
	# 			cell = self.cell('%s%d'%(col,row))
	# 			cell.value = None
	# 			source = self.cell('%s%d'%(col,row-1))
	# 			if copy_style:
	# 				cell.number_format = source.number_format
	# 				cell.font      = source.font.copy()
	# 				cell.alignment = source.alignment.copy()
	# 				cell.border    = source.border.copy()
	# 				cell.fill      = source.fill.copy()
	# 			if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
	# 				if source.value == "=":
	# 					if source.coordinate in self.formula_attributes:
	# 						fa = self.formula_attributes[source.coordinate].copy()
	# 						self.formula_attributes[cell.coordinate] = fa
	# 				else:
	# 					# print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
	# 					cell.value = re.sub(
	# 						"(\$?[A-Z]{1,3})\$?%d"%(row - 1),
	# 						lambda m: m.group(1) + str(row),
	# 						source.value
	# 					)   
	# 				cell.data_type = Cell.TYPE_FORMULA

	# 	# Check for Merged Cell Ranges that need to be expanded to contain new cells
	# 	for cr_idx, cr in enumerate(self.merged_cell_ranges):
	# 		g = RE_RANGE.search(cr).groupdict()
	# 		#print("s_row = %s, e_row = %s, row = %d"%(g['s_row'], g['e_row'],row))
	# 		if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
	# 			#print("Expanding merged cell range '%s' by %d row(s) to '%d'"%(cr,cnt,int(g['e_row'])+cnt))
	# 			self.merged_cell_ranges[cr_idx] = g['s_col'] + g['s_row'] + ":" + g['e_col'] + str(int(g['e_row'])+cnt)

	# 	# Check for Formula Attributes that need reference ranges expanded to include new rows
	# 	for k,v in self.formula_attributes.items():
	# 		if 'ref' in v:
	# 			ref = v['ref']
	# 			if ":" in ref:
	# 				g = RE_RANGE.search(v['ref']).groupdict()
	# 				if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
	# 					# print("Expanding cell range '%s' in formula attribute by %d row(s) to '%d'"%(ref,cnt,int(g['e_row'])+cnt))
	# 					self.formula_attributes[k]['ref'] = g['s_col'] + g['s_row'] + ":" + g['e_col'] + str(int(g['e_row'])+cnt)

# Worksheet.insert_rows = insert_rows

# Guardar Factura-Proyecto
def guardarFacturaProyecto(request):
	try:
		soporte = request.FILES['soporte']
		wb = load_workbook(soporte)
		# ws = wb['Proyectos']
		ws = wb.worksheets[0]

		# print ws['C1'].value
		# print " "
		# print ws.max_row
		id_proyecto = []
		val_proyecto = []
		suma = 0
		id_factura = 0
		val_factura = 0
		i = 0

		if int(ws.max_row) > 1:

			for fila in ws.rows:
				if fila[6].value != "Valor Proyecto":
					if fila[0].value and fila[1].value and fila[6].value:
						# print fila[2].value+" "+fila[3].value+" "+str(fila[6].value)

						if id_factura == 0:
							id_factura = fila[1].value #Guardar el id de la factura
							# print "id factura "

						if fila[1].value == id_factura:
							suma = suma + int(fila[6].value)
							id_proyecto.append(int(fila[0].value))
							val_proyecto.append(int(fila[6].value))
							val_factura = fila[5].value
							# print "suma:"+str(suma)
						else:
							if suma == val_factura:
								suma = 0
								# print "guarda "

								for p in id_proyecto:
									print ("ids proy:", p)
									print ("val proy:", val_proyecto[i])

									model_facturaProyecto = FacturaProyecto(valor=val_proyecto[i],proyecto_id=p,factura_id=id_factura)
									model_facturaProyecto.save()
									i +=1

								id_proyecto = []
								val_proyecto = []
								suma = 0
								id_factura = 0
								val_factura = 0
								i = 0

								id_factura = fila[1].value #Guardar el id de la factura
								suma = suma + int(fila[6].value)
								id_proyecto.append(int(fila[0].value))
								val_proyecto.append(int(fila[6].value))
								val_factura = fila[5].value

			if len(val_proyecto) > 0:
				if suma == val_factura:
					i = 0
					for p in id_proyecto:

						model_facturaProyecto = FacturaProyecto(valor=val_proyecto[i],proyecto_id=p,factura_id=id_factura)
						model_facturaProyecto.save()
						i +=1

		else:
			return JsonResponse({'message':'Plantilla sin registros','success':'error','data':''})

		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok','data':''})
	except Exception as e:
		functions.toLog(e, 'Factura - guardarFacturaProyecto')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Guardar responsabilidades del funcionario
@transaction.atomic
def createFacturaProyectoConLista(lista, id_factura, request):
	# if request.method == 'POST':
	sid = transaction.savepoint()
	try:
		myList = lista.split(',')
		for i in myList:

			model_fp = FacturaProyecto(factura_id=id_factura, proyecto_id=i)
			model_fp.save()

		insert_list_log = []
		fp = FacturaProyecto.objects.filter(factura_id=id_factura, proyecto_id__in = myList)

		for i in fp:
			insert_list_log.append(Logs(usuario_id=request.user.usuario.id
															,accion=Acciones.accion_crear
															,nombre_modelo='factura.FacturaProyecto'
															,id_manipulado=i.id))
		Logs.objects.bulk_create(insert_list_log)

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data': ''})
	except Exception as e:
		functions.toLog(e, 'Factura - createFacturaProyectoConLista')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Crear Soporte en Procesos para factura
@transaction.atomic
def crearSoporteProcesoFactura(request, id_factura):
	sid = transaction.savepoint()
	try:

		id_empresa = request.user.usuario.empresa.id
		tipo_c = tipoC()

		# Traer datos del contrato
		querysetContrato=Contrato.objects.get(pk = request.DATA['contrato_id'])

		# print "t_c:"+str(querysetContrato.tipo_contrato.id)
		# print "t_c_enum:"+str(tipo_c.contratoProyecto)

		if querysetContrato.tipo_contrato.id == tipo_c.contratoProyecto:
			# print "es De Obra"

			querysetProyecto = Proyecto.objects.filter(contrato__id = request.DATA['contrato_id'],
																									contrato__empresacontrato__empresa=id_empresa,
																									contrato__empresacontrato__participa=1).first()
			# print "proyecto Mcontrato:"+str(querysetProyecto.mcontrato.nombre)

			# UPPER() CONVIERTE EL TESTO EN MAYUSCULA
			str1 = querysetProyecto.mcontrato.nombre.upper()
			str2 = "PRONE"
			str3 = "FAER"

			if str1.find(str2) >= 0:
				entro = 0
				# print "Es PRONE"

				if str1.find('2012') >= 0:
					id_item = 96
					id_proceso = 9
					# print "Es prone 2012"
					entro = 1
					# ejemplo
					# id_item = 2
					# id_proceso = 1

				if str1.find('2013') >= 0 or str1.find('2014') >= 0:
					id_item = 28
					id_proceso = 1
					# print "Es prone 2013 - 2014"
					entro = 1
					# ejemplo
					# id_item = 2
					# id_proceso = 1

			elif str1.find(str3) >= 0:
				entro = 0
				# print "Es FAER"

				if str1.find('2012') >= 0 or str1.find('2013') >= 0 or str1.find('2014') >= 0:
					id_item = 172
					id_proceso = 15
					# print "Es FAER 2012 - 2013 - 2014"
					entro = 1
					# ejemplo
					# id_item = 2
					# id_proceso = 1

			else:
				pass
				# print "No es PRONE NI FAER"

			if entro == 1:
				elemento_entidad = querysetProyecto.id

				# BUSCAR EL ID DEL PROCESO_RELACION
				querysetProcesoRelacionDato = GProcesoRelacionDato.objects.filter(procesoRelacion__proceso = id_proceso,
																					procesoRelacion__idApuntador = elemento_entidad,
																					item = id_item).first()
				# print "id_relacion_datos:"+str(querysetProcesoRelacionDato)
				id_relacion_datos = querysetProcesoRelacionDato[0].id

				# ACTUALIZAR EL PROCESO_RELACION_DATO
				queryset_prd = GProcesoRelacionDato.objects.get(pk = id_relacion_datos)
				queryset_prd.valor = 'Si'
				queryset_prd.estado = 1
				queryset_prd.save()

				logs_model=Logs(
					usuario_id=request.user.usuario.id,
					accion=Acciones.accion_actualizar,
					nombre_modelo='Procesos.procesoRelacionDatos',
					id_manipulado=queryset_prd.id
				)
				logs_model.save()

				# GUARDAR EL SOPORTE DEL PROCESO_RELACION_DATO
				modelSoporteProcesoRelacionDato = HSoporteProcesoRelacionDato(procesoRelacionDato_id = id_relacion_datos,
																				nombre = request.DATA['concepto'],
																				documento = request.FILES['soporte'])
				modelSoporteProcesoRelacionDato.save()

				# EDITAR LA FACTURA CREADA, PARA GUARDAR EL ID DEL SOPORTE_PROCESO_RELACION_DATO
				model_factura = Factura.objects.get(pk = id_factura)
				model_factura.proceso_soporte_id	= modelSoporteProcesoRelacionDato.id
				model_factura.save()

				logs_model=''
				logs_model=Logs(
					usuario_id=request.user.usuario.id,
					accion=Acciones.accion_crear,
					nombre_modelo='Procesos.soporteProcesoRelacionDatos',
					id_manipulado=modelSoporteProcesoRelacionDato.id
				)
				logs_model.save()
		else:
			pass
			# print "No es De Obra"
		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearSoporteProcesoFactura')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Editar Soporte en Procesos para factura
@transaction.atomic
def updateSoporteProcesoFactura(request):
	sid = transaction.savepoint()
	try:
		model_factura = Factura.objects.get(pk = request.DATA['id'])
		if model_factura.proceso_soporte_id:

			model_sprd = HSoporteProcesoRelacionDato.objects.get(pk = model_factura.proceso_soporte_id)

			model_sprd.nombre = request.DATA['concepto']
			model_sprd.documento = request.FILES['soporte'] if request.FILES.get('soporte') is not None else model_sprd.documento
			model_sprd.save()

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - updateSoporteProcesoFactura')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Crear Proceso de cuentas por pagar de Enalar
@transaction.atomic
def crearProesoEnelarFactura(request, id_factura, id_proceso):
	sid = transaction.savepoint()
	try:
		model_log=Logs
		model_acciones=Acciones
		nombre_modulo1='Procesos.procesoRelacion'
		nombre_modulo2='Procesos.procesoRelacionDatos'
		nombre_modulo3='Procesos.soporteProcesoRelacionDatos'

		# Guardar el Proceso relacion
		modelProcesoRelacion = FProcesoRelacion(proceso_id = id_proceso,
												idApuntador = request.DATA['contrato_id'], #id_contrato,
												idTablaReferencia = id_factura)
		modelProcesoRelacion.save()
		transaction.savepoint_commit(sid)

		logs_model=model_log(
						usuario_id=request.user.usuario.id,
						accion=model_acciones.accion_crear,
						nombre_modelo=nombre_modulo1,
						id_manipulado=modelProcesoRelacion.id
					)
		logs_model.save()
		transaction.savepoint_commit(sid)

		# buscar los item del porceso para guardarlos
		querysetItem = BItem.objects.filter(proceso = id_proceso)

		for item in querysetItem:
			sid = transaction.savepoint()
			if item.id == 42:
				id_estado = 1
				valor = 'Si'
			else:
				id_estado = 0
				valor = 'Vacio'

			# Guardar el Proceso relacion dato
			modelProcesoRelacionDato = GProcesoRelacionDato(procesoRelacion_id = modelProcesoRelacion.id,
															item_id = item.id,
															estado = id_estado,
															valor = valor,
															observacion = None,
															fechaVencimiento = None)
			modelProcesoRelacionDato.save()
			transaction.savepoint_commit(sid)

			logs_model=model_log(
						usuario_id=request.user.usuario.id,
						accion=model_acciones.accion_crear,
						nombre_modelo=nombre_modulo2,
						id_manipulado=modelProcesoRelacionDato.id
					)
			logs_model.save()
			transaction.savepoint_commit(sid)

			if item.id == 42:
				id_relacion_datos = modelProcesoRelacionDato.id

				# GUARDAR EL SOPORTE DEL PROCESO_RELACION_DATO del item 42
				modelSoporteProcesoRelacionDato = HSoporteProcesoRelacionDato(procesoRelacionDato_id = id_relacion_datos,
																			nombre = request.DATA['numero'],
																			documento = request.FILES['soporte'])
				modelSoporteProcesoRelacionDato.save()
				transaction.savepoint_commit(sid)

				logs_model=model_log(
						usuario_id=request.user.usuario.id,
						accion=model_acciones.accion_crear,
						nombre_modelo=nombre_modulo3,
						id_manipulado=modelSoporteProcesoRelacionDato.id
					)
				logs_model.save()
				transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearProesoEnelarFactura')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# crearProesoEnelarFactura = transaction.atomic(crearProesoEnelarFactura)
# busca si la empresa tiene permiso para el preceso y si no lo tiene lo crea igual que los item
@transaction.atomic
def crearPermisoProesoItem(request, id_proceso):
	sid = transaction.savepoint()
	try:
		id_empresa = request.user.usuario.empresa.id

		model_proceso = AProceso.objects.get(pk=id_proceso)
		proceso_empresa = model_proceso.empresas.filter(id=id_empresa).values('id')#.first()
		# print "permiso empresa proceso: "+str(proceso_empresa.count())#+" Id: "+str(proceso_empresa[0]['id'])

		if proceso_empresa.count() == 0:
			# print "La Empresa no tiene permisos para el Proceso."
			model_proceso.empresas.add(id_empresa)

		# buscar los item del porceso
		querysetItem = BItem.objects.filter(proceso = id_proceso)
		queryPermisoEmpresaItem = CPermisoEmpresaItem.objects.filter(empresa = id_empresa, item__in = querysetItem)
		# print "permiso empresa Item: "+str(queryPermisoEmpresaItem.count())+" - Numero de Item: "+str(querysetItem.count())

		if queryPermisoEmpresaItem.count() < querysetItem.count():

			if queryPermisoEmpresaItem.count() > 0:
				queryPermisoEmpresaItem.delete()
			# 	print "No tine todos los permisos."
			# else:
			# 	print "No tine ninguno de los permisos."

			insert_list_item = []
			for item in querysetItem:
				if item.id == 42:
					escritura = True
				else:
					escritura = False

				insert_list_item.append(CPermisoEmpresaItem(lectura = True,
															escritura = escritura,
															empresa_id = id_empresa,
															item_id = item.id))
			CPermisoEmpresaItem.objects.bulk_create(insert_list_item)

		transaction.savepoint_commit(sid)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		functions.toLog(e, 'Factura - crearPermisoProesoItem')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#guardar el pago de las factura
@transaction.atomic
def guardar_pago_factura(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			object_factura=Factura.objects.get(pk=item['id'])

			object_factura.fecha_pago=respuesta['fecha_pago']
			object_factura.pagada=1
			#object_detalle.estado_id=9
			object_factura.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.facturas',id_manipulado=item['id'])
			logs_model.save()

			transaction.savepoint_commit(sid)


		return JsonResponse({'message':'El registro se ha actualizo correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e, 'Factura - guardar_pago_factura')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

#guardar Pago recursos Propios
@transaction.atomic
def pagoRecursosPropios(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			object_factura=Factura.objects.get(pk=item['id'])

			# object_factura.fecha_pago=respuesta['fecha_pago']
			object_factura.recursos_propios=1
			object_factura.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='factura.facturas',id_manipulado=item['id'])
			logs_model.save()

			transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro se ha actualizo correctamente','success':'ok','data':''})
		
	except Exception as e:
		functions.toLog(e, 'Factura - pagoRecursosPropios')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})


@login_required
def VerSoporte(request):
	if request.method == 'GET':
		try:
			
			archivo = Factura.objects.get(pk=request.GET['id'])
			
			# filename = ""+str(archivo.soporte)+""
			# extension = filename[filename.rfind('.'):]
			# nombre = re.sub('[^A-Za-z0-9 ]+', '', archivo.nombre)
			return functions.exportarArchivoS3(str(archivo.soporte))

		except Exception as e:
			functions.toLog(e,'factura.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def VerSoporteCesion(request):
	if request.method == 'GET':
		try:
			
			archivo = Cesion.objects.get(pk=request.GET['id'])
			
			# filename = ""+str(archivo.soporte)+""
			# extension = filename[filename.rfind('.'):]
			# nombre = re.sub('[^A-Za-z0-9 ]+', '', archivo.nombre)
			return functions.exportarArchivoS3(str(archivo.soporte))

		except Exception as e:
			functions.toLog(e,'factura.VerSoporteCesion')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def VerSoporteDescuento(request):
	if request.method == 'GET':
		try:
			
			archivo = Descuento.objects.get(pk=request.GET['id'])
			
			# filename = ""+str(archivo.soporte)+""
			# extension = filename[filename.rfind('.'):]
			# nombre = re.sub('[^A-Za-z0-9 ]+', '', archivo.nombre)
			return functions.exportarArchivoS3(str(archivo.soporte))

		except Exception as e:
			functions.toLog(e,'factura.VerSoporteCesion')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
