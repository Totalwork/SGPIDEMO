from django.shortcuts import render,redirect,render_to_response
from django.urls import reverse
from django.core.serializers import serialize
from rest_framework import viewsets, serializers
from .models import AFondo,ACampana,ASolicitante,AUnidadMedida,CampanaEmpresa,CSolicitud,DatoDiseno,Diseno,SoporteSolicitud,DVersionesDiseno
from .models import DocumentoEstado,EstadoDiseno,InfoDiseno,MapaDiseno,PermisoDiseno,SoporteEstado,SoporteEstadoComentario,TComentarioDiseno

from empresa.views import EmpresaSerializer
from empresa.models import Empresa

from estado.models import Estado
from estado.views import EstadoSerializer

from usuario.views import UsuarioSerializer
from usuario.models import Usuario

from parametrizacion.models import Municipio,Departamento
from parametrizacion.views import MunicipioSerializer

from logs.models import Logs,Acciones
from django.db import transaction
from django.db.models import Q,Max

from django.contrib.auth.decorators import login_required

from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination

import json
import xlsxwriter
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.deletion import ProtectedError

import openpyxl
from django.db import connection

from .enumeration import EstadoFondo

from django.core.paginator import Paginator
from datetime import *
from sinin4.functions import functions

from .tasks import envioNotificacionReporteProyecto,envioNotificacionReporteExitoso,envioNotificacionReporteInsconsitencia
# Create your views here.

#Api rest para fondo
class FondoSerializer(serializers.HyperlinkedModelSerializer):

	empresa=EmpresaSerializer(read_only=True)
	empresa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app='Gestion_proyecto_fondo'))

	class Meta:
		model = AFondo
		fields=('id','nombre','empresa','empresa_id','estado_id','estado',)

class FondoViewSet(viewsets.ModelViewSet):
	
	model=AFondo
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.fondo'
	queryset = model.objects.all()
	serializer_class = FondoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(FondoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(Q(empresa_id=request.user.usuario.empresa.id))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)
			
			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = FondoSerializer(data=request.DATA,context={'request': request})
				estado = Estado()
				codigo=request.DATA['estado_id']
				request.DATA['estado_id']=estado.ObtenerID('Gestion_proyecto_fondo',codigo)
				# print 'error3'

				if serializer.is_valid():
					serializer.save(estado_id=request.DATA['estado_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = FondoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(estado_id=request.DATA['estado_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para fondo


#Api rest para campana
class CampanaSerializer(serializers.HyperlinkedModelSerializer):

	totalDocumento=serializers.SerializerMethodField()

	class Meta:
		model = ACampana
		fields=('id','nombre','totalDocumento',)

	def get_totalDocumento(self, obj):
		documentos=DocumentoEstado.objects.filter(campana_id=obj.id)		
		return len(documentos)

class CampanaViewSet(viewsets.ModelViewSet):
	
	model=ACampana
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.campana'
	queryset = model.objects.all()
	serializer_class = CampanaSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CampanaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CampanaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					empresa=CampanaEmpresa(campana_id=serializer.data['id'],empresa_id=request.user.usuario.empresa.id,propietario=True)
					empresa.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='gestion_empresa.campana_empresa',id_manipulado=empresa.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CampanaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para campana

#Api rest para solicitante
class SolicitanteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = ASolicitante
		fields=('id','nombre',)

class SolicitanteViewSet(viewsets.ModelViewSet):
	
	model=ASolicitante
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.solicitante'
	queryset = model.objects.all()
	serializer_class = SolicitanteSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SolicitanteViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SolicitanteSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SolicitanteSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para solicitante


#Api rest para unidad de medida
class UnidadMedidaSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = AUnidadMedida
		fields=('id','nombre',)

class UnidadMedidaViewSet(viewsets.ModelViewSet):
	
	model=AUnidadMedida
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.unidad_medida'
	queryset = model.objects.all()
	serializer_class = UnidadMedidaSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(UnidadMedidaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = UnidadMedidaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = UnidadMedidaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para unidad de medida

#Api rest para campana empresa
class CampanaEmpresaSerializer(serializers.HyperlinkedModelSerializer):

	empresa=EmpresaSerializer(read_only=True)
	empresa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	campana=CampanaSerializer(read_only=True)
	campana_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=ACampana.objects.all())

	class Meta:
		model = CampanaEmpresa
		fields=('id','empresa','empresa_id','campana_id','campana','propietario',)

class CampanaEmpresaViewSet(viewsets.ModelViewSet):
	
	model=CampanaEmpresa
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.campana_empresa'
	queryset = model.objects.all()
	serializer_class = CampanaEmpresaSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(CampanaEmpresaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			id_empresa= self.request.query_params.get('filtro_id_empresa',None)
			id_campana= self.request.query_params.get('id_campana',None)

			qset=(~Q(id=0))
			if id_empresa is not None:
				qset=qset &(Q(empresa_id=request.user.usuario.empresa.id))
			if dato:
				qset = qset &(
					Q(campana__nombre__icontains=dato)
					)
			if id_campana:
				qset = qset &(
					Q(campana_id=id_campana)
					)

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = CampanaEmpresaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(campana_id=request.DATA['campana_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = CampanaEmpresaSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(campana_id=request.DATA['campana_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para campana empresa


#Serializador para soporte solicitud
class SoporteSolicitudLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = SoporteSolicitud
		fields=('id','nombre','ruta',)

#Fin de serializador de soporte solicitud

#Api rest para solicitud
class SolicitudSerializer(serializers.HyperlinkedModelSerializer):


	soportes=SoporteSolicitudLiteSerializer(read_only=True,many=True)

	class Meta:
		model = CSolicitud
		fields=('id','nombre','fecha','entidad','visita','fecha_visita','fecha_respuesta','descripcion_visita','soportes',)
	

class SolicitudDisenoViewSet(viewsets.ModelViewSet):
	
	model=CSolicitud
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.solicitud'
	queryset = model.objects.all()
	serializer_class = SolicitudSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SolicitudDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if request.DATA['fecha_visita']=='':
					request.DATA['fecha_visita']=None
				serializer = SolicitudSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					soportes=request.FILES.getlist('soporte[]')
					
					if soportes:	
						fecha=datetime.now().strftime('%Y-%m-%d')	
						for sp in soportes:					
							s= SoporteSolicitud(ruta=sp, nombre=sp.name, fecha=fecha,usuario_id=request.user.usuario.id,solicitud_id=serializer.data['id'])
							s.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='gestion_proyecto.soporte_solicitud',id_manipulado=s.pk)
							logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				if request.DATA['fecha_visita']=='' or request.DATA['fecha_visita']=='null':
					request.DATA['fecha_visita']=None
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SolicitudSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					soportes=request.FILES.getlist('soporte[]')
					
					if soportes:	
						fecha=datetime.now().strftime('%Y-%m-%d')	
						for sp in soportes:					
							s= SoporteSolicitud(ruta=sp, nombre=sp.name, fecha=fecha,usuario_id=request.user.usuario.id,solicitud_id=serializer.data['id'])
							s.save()
							logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='gestion_proyecto.soporte_solicitud',id_manipulado=s.pk)
							logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para solicitud

#Api rest para dato diseno
class DatoDisenoSerializer(serializers.HyperlinkedModelSerializer):

	unidad_medida=UnidadMedidaSerializer(read_only=True)
	unidad_medida_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=AUnidadMedida.objects.all())

	class Meta:
		model = DatoDiseno
		fields=('id','nombre','orden','unidad_medida_id','unidad_medida',)

class DatoDisenoViewSet(viewsets.ModelViewSet):
	
	model=DatoDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.dato_diseno'
	queryset = model.objects.all()
	serializer_class = DatoDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DatoDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				datos=DatoDiseno.objects.all().aggregate(Max('orden'))
				if datos['orden__max'] is None:
					request.DATA['orden']=1
				else:
					request.DATA['orden']=datos['orden__max']+1
				serializer = DatoDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():					
					serializer.save(unidad_medida_id=request.DATA['unidad_medida_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DatoDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(unidad_medida_id=request.DATA['unidad_medida_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para dato diseno


class SoporteEstadoLiteSerializer(serializers.HyperlinkedModelSerializer):

	usuario=UsuarioSerializer(read_only=True)
	usuario_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	class Meta:
		model = SoporteEstado
		fields=('id','usuario','usuario_id','fecha','ruta','nombre',)

#Api rest para diseno
class DisenoSerializer(serializers.HyperlinkedModelSerializer):

	municipio=MunicipioSerializer(read_only=True)
	municipio_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Municipio.objects.all())

	fondo=FondoSerializer(read_only=True)
	fondo_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=AFondo.objects.all())

	solicitante=SolicitanteSerializer(read_only=True)
	solicitante_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=ASolicitante.objects.all())

	campana=CampanaSerializer(read_only=True)
	campana_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=ACampana.objects.all())

	propietaria=EmpresaSerializer(read_only=True)
	propietaria_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	disenadores=EmpresaSerializer(read_only=True,many=True)
	
	solicitudes=SolicitudSerializer(read_only=True,many=True)

	class Meta:
		model = Diseno
		fields=('id','municipio_id','municipio','fondo_id','fondo','solicitante_id','solicitante',
			'campana_id','campana','propietaria_id','propietaria','activado','costo_proyecto','costo_diseno',
			'disenadores','solicitudes','nombre','estado')


class DisenoViewSet(viewsets.ModelViewSet):
	
	model=Diseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.diseno'
	queryset = model.objects.all()
	serializer_class = DisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			departamento_id= self.request.query_params.get('departamento_id',None)
			municipio_id= self.request.query_params.get('municipio_id',None)
			fondo_id= self.request.query_params.get('fondo_id',None)
			campana_id= self.request.query_params.get('campana_id',None)
			solicitante_id= self.request.query_params.get('solicitante_id',None)
			estado_id= self.request.query_params.get('estado_id',None)
			disenadores_id= self.request.query_params.get('disenadores_id',None)
			filtro_solicitud=self.request.query_params.get('filtro_solicitud',None)
			solicitud_id=self.request.query_params.get('solicitud_id',None)
			filtro_tablero=self.request.query_params.get('filtro_tablero',None)
			diseno_id=self.request.query_params.get('diseno_id',None)
			filtro_permiso=self.request.query_params.get('filtro_permiso',None)
			empresa_id=self.request.query_params.get('empresa_id',None)

			qset=(Q(activado=True)&Q(permiso_diseno_diseno__empresa__id=request.user.usuario.empresa.id))
			if dato:
				qset = qset&(
					Q(nombre__icontains=dato)
					)

			if departamento_id and int(departamento_id)>0:
				qset = qset&(
					Q(municipio__departamento__id=departamento_id)
					)

			if municipio_id and int(municipio_id)>0:
				qset = qset&(
					Q(municipio__id=municipio_id)
					)

			if fondo_id and int(fondo_id)>0:
				qset = qset&(
					Q(fondo__id=fondo_id)
					)

			if campana_id and int(campana_id)>0:
				qset = qset&(
					Q(campana__id=campana_id)
					)

			if solicitante_id and int(solicitante_id)>0:
				qset = qset&(
					Q(solicitante__id=solicitante_id)
					)

			if estado_id and int(estado_id)>0:
				lista_id=[]
				estado=Diseno.objects.filter(activado=True,permiso_diseno_diseno__empresa__id=request.user.usuario.empresa.id)
				for item in estado:
					if item.estado and int(item.estado['estado__id'])==int(estado_id):
						lista_id.append(item.id)
				
				qset = qset & (
					Q(id__in=(lista_id))
					)

			if filtro_solicitud and int(filtro_solicitud)>0:
				lista_id=[]
				solicitud=Diseno.objects.filter(solicitudes=filtro_solicitud)
				if len(solicitud)>0:
					for item in solicitud:
						lista_id.append(item.id)
					qset = qset & (
						~Q(id__in=(lista_id))
						)

			if disenadores_id and int(disenadores_id)>0:
					qset=qset&(Q(disenadores=disenadores_id))

			if solicitud_id and int(solicitud_id)>0:
					qset=qset&(Q(solicitudes=solicitud_id))			

			if filtro_tablero is not None and diseno_id:
				ListEmpresas=PermisoDiseno.objects.filter(diseno_id=diseno_id).values('empresa__nombre','consultar','editar')
				Lista=[]
				ListEstado=Estado.objects.filter(app='Gestion_proyecto_documento').values('id','nombre').order_by('codigo')
				for item in ListEstado:
					fecha=''
					id_fecha=0
					ListaSoporte=[]
					valor=EstadoDiseno.objects.filter(diseno_id=diseno_id,estado_id=item['id']).values('fecha','id')
					if len(valor)>0:
						fecha=valor[0]['fecha']
						id_fecha=valor[0]['id']	
					Lista.append({'id':id_fecha,'id_estado':item['id'],'nombre':item['nombre'],'fecha':fecha})
				
				Lista2=[]
				ListDatos=DatoDiseno.objects.all().order_by('orden')
				for item in ListDatos:
					valor_dato=0
					id_valor=0
					valor=InfoDiseno.objects.filter(diseno_id=diseno_id,dato_diseno_id=item.id).values('valor','id')
					if len(valor) >0:
						valor_dato=valor[0]['valor']
						id_valor=valor[0]['id']	
					Lista2.append({'id':id_valor,'id_dato':item.id,'nombre':item.nombre,'unidad':item.unidad_medida.nombre,'valor':valor_dato});

				return Response({'message':'','success':'ok',
					'data':{'empresas':ListEmpresas,'estados':Lista,'datos':Lista2}})


			if filtro_permiso is not None and empresa_id:
				ListDiseno=self.model.objects.filter(qset)
				Lista=[]

				# print len(ListDiseno)
				for item in ListDiseno:
					permiso=PermisoDiseno.objects.filter(empresa_id=empresa_id,diseno_id=item.id)
					consultar=False
					editar=False
					if len(permiso)>0:
						consultar=permiso[0].consultar
						editar=permiso[0].editar

					Lista.append({'id':item.id,'nombre':item.nombre,'fondo':item.fondo.nombre,'municipio':item.municipio.nombre,'departamento':item.municipio.departamento.nombre,'consultar':consultar,'editar':editar});

					return Response({'message':'','success':'ok','data':Lista})

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				if int(request.DATA['disenadores_id'])==0:
					request.DATA['disenadores_id']=None

				if int(request.DATA['solicitudes_id'])==0:
					request.DATA['solicitudes_id']=None

				serializer = DisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(municipio_id=request.DATA['municipio_id'],fondo_id=request.DATA['fondo_id'],campana_id=request.DATA['campana_id'],
						solicitante_id=request.DATA['solicitante_id'],propietaria_id=request.DATA['propietaria_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					permiso=PermisoDiseno(diseno_id=serializer.data['id'],empresa_id=request.user.usuario.empresa.id,consultar=True,editar=True)
					permiso.save()
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo='gestion_proyecto.permiso_diseno',id_manipulado=permiso.id)
					logs_model.save()

					objecto_numero=Diseno.objects.get(pk=serializer.data['id'])
					if request.DATA['disenadores_id']>0:
						objecto_numero.disenadores.add(request.DATA['disenadores_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()


					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(municipio_id=request.DATA['municipio_id'],fondo_id=request.DATA['fondo_id'],campana_id=request.DATA['campana_id'],
						solicitante_id=request.DATA['solicitante_id'],propietaria_id=request.DATA['propietaria_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					objecto_numero=Diseno.objects.get(pk=instance.id)
					if int(request.DATA['disenadores_id'])>0:
						objecto_numero.disenadores.clear()
						objecto_numero.disenadores.add(request.DATA['disenadores_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
						logs_model.save()
					else:
						objecto_numero.disenadores.clear()


					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para diseno



#Api rest para version de diseno
class VersionDisenoSerializer(serializers.HyperlinkedModelSerializer):

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	estado_reporte=EstadoSerializer(read_only=True)
	estado_reporte_id=serializers.PrimaryKeyRelatedField(write_only=True,allow_null=True,queryset=Estado.objects.all())	

	class Meta:
		model = DVersionesDiseno
		fields=('id','nombre','diseno','diseno_id','fecha','fecha_format','estado_reporte','estado_reporte_id','reportar_satisfaccion','reportar_diseno')

class VersionDisenoViewSet(viewsets.ModelViewSet):
	
	model=DVersionesDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.version_diseno'
	queryset = model.objects.all()
	serializer_class = VersionDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(VersionDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			diseno_id = self.request.query_params.get('diseno_id', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))

			if diseno_id:
				qset = qset &(
					Q(diseno_id=diseno_id)
					)


			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)
			
			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:

				if int(request.DATA['estado_reporte_id'])==0:
					request.DATA['estado_reporte_id']=None

				serializer = VersionDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],estado_reporte_id=request.DATA['estado_reporte_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = VersionDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],estado_reporte_id=request.DATA['estado_reporte_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para fondo



#Api rest para documento estado
class DocumentoEstadoSerializer(serializers.HyperlinkedModelSerializer):

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app='Gestion_proyecto_documento'))

	campana=CampanaSerializer(read_only=True)
	campana_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=ACampana.objects.all())

	class Meta:
		model = DocumentoEstado
		fields=('id','nombre','estado_id','estado','campana_id','campana',)

class DocumentoEstadoViewSet(viewsets.ModelViewSet):
	
	model=DocumentoEstado
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.documento_estado'
	queryset = model.objects.all()
	serializer_class = DocumentoEstadoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DocumentoEstadoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			id_campana= self.request.query_params.get('id_campana',None)
			id_estado= self.request.query_params.get('id_estado',None)
			filtro_estado= self.request.query_params.get('filtro_estado',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if id_campana:
				qset = qset &(
					Q(campana_id=id_campana)
					)

			if id_estado:
				qset = qset &(
					Q(estado_id=id_estado)
					)

			if filtro_estado is not None:
				Listado=[]
				queryset= Estado.objects.filter(app='Gestion_proyecto_documento')

				for item in queryset:
					documentos=DocumentoEstado.objects.filter(campana_id=id_campana,estado_id=item.id).values('id','nombre')
					Listado.append({'nombre':item.nombre,'id':item.id,'listado_documentos':documentos})

				return Response({'message':'','success':'ok',
					'data':Listado})
			else:			
				queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = DocumentoEstadoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(campana_id=request.DATA['campana_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DocumentoEstadoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(campana_id=request.DATA['campana_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		sid = transaction.savepoint()
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			transaction.savepoint_commit(sid)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_201_CREATED)

		except ProtectedError:
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			transaction.savepoint_rollback(sid)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para documento dato

#Api rest para estado diseno
class EstadoDisenoSerializer(serializers.HyperlinkedModelSerializer):

	estado=EstadoSerializer(read_only=True)
	estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Estado.objects.filter(app='Gestion_proyecto_documento'))

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	class Meta:
		model = EstadoDiseno
		fields=('id','fecha','estado_id','estado','diseno_id','diseno',)


class EstadoDisenoViewSet(viewsets.ModelViewSet):
	
	model=EstadoDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.estado_diseno'
	queryset = model.objects.all()
	serializer_class = EstadoDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(EstadoDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			campana_id= self.request.query_params.get('campana_id',None)

			qset=(~Q(id=0))
			if dato:
				qset = (
					Q(diseno__nombre__icontains=dato)
					)

			if campana_id and campana_id>0:
				qset = (
					Q(diseno__campana__id=campana_id)
					)

			queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = EstadoDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EstadoDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],estado_id=request.DATA['estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para estado diseno

#Api rest para info diseno
class InfoDisenoSerializer(serializers.HyperlinkedModelSerializer):

	dato_diseno=DatoDisenoSerializer(read_only=True)
	dato_diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=DatoDiseno.objects.all())

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	class Meta:
		model = InfoDiseno
		fields=('id','valor','dato_diseno_id','dato_diseno','diseno_id','diseno',)

class InfoDisenoViewSet(viewsets.ModelViewSet):
	
	model=InfoDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.info_diseno'
	queryset = model.objects.all()
	serializer_class = InfoDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(InfoDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			if dato:
				qset = (
					Q(diseno__nombre__icontains=dato)
					)
				queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = InfoDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],dato_diseno_id=request.DATA['dato_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = InfoDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],dato_diseno_id=request.DATA['dato_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para info diseno

#Api rest para mapa diseno
class MapaDisenoSerializer(serializers.HyperlinkedModelSerializer):

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	class Meta:
		model = MapaDiseno
		fields=('id','longitud','latitud','diseno_id','diseno','nombre',)

class MapaDisenoViewSet(viewsets.ModelViewSet):
	
	model=MapaDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.mapa_diseno'
	queryset = model.objects.all()
	serializer_class = MapaDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(MapaDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			diseno_id = self.request.query_params.get('diseno_id', None)
			version_diseno_id = self.request.query_params.get('version_diseno_id', None)
			lista_diseno = self.request.query_params.get('lista_diseno', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if diseno_id and diseno_id>0:
				qset = qset &(
					Q(diseno_id=diseno_id)
					)

			if version_diseno_id and version_diseno_id>0:
				qset = qset &(
					Q(version_diseno_id=version_diseno_id)
					)

			if lista_diseno:
				lista=lista_diseno.split(',')
				qset = qset &(
					Q(diseno_id__in=lista)
					)
				
			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = MapaDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],version_diseno_id=request.DATA['version_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = MapaDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],version_diseno_id=request.DATA['version_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para mapa diseno

#Api rest para permiso diseno
class PermisoDisenoSerializer(serializers.HyperlinkedModelSerializer):

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	empresa=EmpresaSerializer(read_only=True)
	empresa_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())

	class Meta:
		model = PermisoDiseno
		fields=('id','empresa_id','empresa','diseno_id','diseno','consultar','editar',)

class PermisoDisenoViewSet(viewsets.ModelViewSet):
	
	model=PermisoDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.permiso_diseno'
	queryset = model.objects.all()
	serializer_class = PermisoDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(PermisoDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			departamento_id = self.request.query_params.get('departamento_id', None)
			municipio_id= self.request.query_params.get('municipio_id', None)
			fondo_id= self.request.query_params.get('fondo_id', None)

			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset & (
					Q(diseno__nombre__icontains=dato)
					)
			if departamento_id and departamento_id>0:
				qset = qset & (
					Q(diseno__municipio__departamento__id=departamento_id)
					)

			if municipio_id and municipio_id>0:
				qset = qset & (
					Q(diseno__municipio__id=municipio_id)
					)

			if fondo_id and fondo_id>0:
				qset = qset & (
					Q(diseno__fondo__id=fondo_id)
					)
			

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = PermisoDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = PermisoDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(diseno_id=request.DATA['diseno_id'],empresa_id=request.DATA['empresa_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para permiso diseno

#Api rest para soporte estado
class SoporteEstadoSerializer(serializers.HyperlinkedModelSerializer):

	estado_diseno=EstadoDisenoSerializer(read_only=True)
	estado_diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=EstadoDiseno.objects.all())

	documento_estado=DocumentoEstadoSerializer(read_only=True)
	documento_estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=DocumentoEstado.objects.all())

	usuario=UsuarioSerializer(read_only=True)
	usuario_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	class Meta:
		model = SoporteEstado
		fields=('id','estado_diseno_id','estado_diseno','documento_estado_id','documento_estado',
			'usuario','usuario_id','fecha','ruta','nombre','cantidad_comentarios',)

class SoporteEstadoViewSet(viewsets.ModelViewSet):
	
	model=SoporteEstado
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.soporte_estado'
	queryset = model.objects.all()
	serializer_class = SoporteEstadoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SoporteEstadoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)
			estado_id= self.request.query_params.get('estado_id',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(documento_estado__nombre__icontains=dato)
					)
			if estado_id:
				qset = qset &(
					Q(estado_diseno_id=estado_id)
					)

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SoporteEstadoSerializer(data=request.DATA,context={'request': request})
				if serializer.is_valid():
					serializer.save(ruta=self.request.FILES.get('ruta'),usuario_id=request.DATA['usuario_id'],estado_diseno_id=request.DATA['estado_diseno_id'],documento_estado_id=request.DATA['documento_estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SoporteEstadoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					if self.request.FILES.get('ruta')=='' or self.request.FILES.get('ruta') is None:
						soporte=SoporteEstado.objects.get(pk=instance.id)
						serializer.save(ruta=soporte.ruta,usuario_id=request.DATA['usuario_id'],estado_diseno_id=request.DATA['estado_diseno_id'],documento_estado_id=request.DATA['documento_estado_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
						logs_model.save()
					else:
						serializer.save(ruta=self.request.FILES.get('ruta'),usuario_id=request.DATA['usuario_id'],estado_diseno_id=request.DATA['estado_diseno_id'],documento_estado_id=request.DATA['documento_estado_id'])
						logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
						logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			SoporteEstadoComentario.objects.filter(soporte_estado_id=instance.id).delete()
			logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_borrar,nombre_modelo='gestion_proyecto.comentario_soporte',id_manipulado=instance.id)
			logs_model.save()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para soporte estado

#Api rest para soporte estado comentario
class SoporteEstadoComentarioSerializer(serializers.HyperlinkedModelSerializer):

	soporte_estado=SoporteEstadoSerializer(read_only=True)
	soporte_estado_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=SoporteEstado.objects.all())

	usuario=UsuarioSerializer(read_only=True)
	usuario_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	class Meta:
		model = SoporteEstadoComentario
		fields=('id','soporte_estado_id','soporte_estado','usuario','usuario_id','comentario','fecha','fecha_format',)

class SoporteEstadoComentarioViewSet(viewsets.ModelViewSet):
	
	model=SoporteEstadoComentario
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.soporte_estado_comentario'
	queryset = model.objects.all()
	serializer_class = SoporteEstadoComentarioSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SoporteEstadoComentarioViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			soporte_estado_id = self.request.query_params.get('soporte_estado_id', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(comentario__icontains=dato)
					)

			if soporte_estado_id and soporte_estado_id>0:
				qset = qset &(
					Q(soporte_estado_id=soporte_estado_id)
					)

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SoporteEstadoComentarioSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],soporte_estado_id=request.DATA['soporte_estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SoporteEstadoComentarioSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],soporte_estado_id=request.DATA['soporte_estado_id'],documento_estado_id=request.DATA['documento_estado_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para soporte estado comentario


#Api rest para soporte solicitud
class SoporteSolicitudSerializer(serializers.HyperlinkedModelSerializer):

	solicitud=SolicitudSerializer(read_only=True)
	solicitud_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=CSolicitud.objects.all())

	usuario=UsuarioSerializer(read_only=True)
	usuario_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	class Meta:
		model = SoporteSolicitud
		fields=('id','solicitud_id','solicitud','usuario','usuario_id','ruta','fecha','nombre',)

class SoporteSolicitudViewSet(viewsets.ModelViewSet):
	
	model=SoporteSolicitud
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.soporte_solicitud'
	queryset = model.objects.all()
	serializer_class = SoporteSolicitudSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(SoporteSolicitudViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			solicitud_id = self.request.query_params.get('solicitud_id', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(nombre__icontains=dato)
					)

			if solicitud_id:
				qset = qset &(
					Q(solicitud_id=solicitud_id)
					)

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = SoporteSolicitudSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],solicitud_id=request.DATA['solicitud_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = SoporteSolicitudSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],solicitud_id=request.DATA['solicitud_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para soporte solicitud


#Api rest para soporte estado comentario
class ComentarioDisenoSerializer(serializers.HyperlinkedModelSerializer):

	diseno=DisenoSerializer(read_only=True)
	diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Diseno.objects.all())

	version_diseno=VersionDisenoSerializer(read_only=True)
	version_diseno_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=DVersionesDiseno.objects.all())

	usuario=UsuarioSerializer(read_only=True)
	usuario_id=serializers.PrimaryKeyRelatedField(write_only=True,queryset=Usuario.objects.all())

	class Meta:
		model = TComentarioDiseno
		fields=('id','diseno_id','diseno','usuario','usuario_id','comentario','fecha','fecha_format','version_diseno','version_diseno_id')

class ComentarioDisenoViewSet(viewsets.ModelViewSet):
	
	model=TComentarioDiseno
	model_log=Logs
	model_acciones=Acciones
	nombre_modulo='gestion_proyecto.comentario_diseno'
	queryset = model.objects.all()
	serializer_class = ComentarioDisenoSerializer

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(ComentarioDisenoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			diseno_id = self.request.query_params.get('diseno_id', None)
			version_diseno_id = self.request.query_params.get('version_diseno_id', None)
			sin_paginacion= self.request.query_params.get('sin_paginacion',None)

			qset=(~Q(id=0))
			if dato:
				qset = qset &(
					Q(comentario__icontains=dato)
					)

			if diseno_id and diseno_id>0:
				qset = qset &(
					Q(diseno_id=diseno_id)
					)


			if version_diseno_id and version_diseno_id>0:
				qset = qset &(
					Q(version_diseno_id=version_diseno_id)
					)

			queryset = self.model.objects.filter(qset)


			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = ComentarioDisenoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],diseno_id=request.DATA['diseno_id'],version_diseno_id=request.DATA['version_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_crear,nombre_modelo=self.nombre_modulo,id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = ComentarioDisenoSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(usuario_id=request.DATA['usuario_id'],diseno_id=request.DATA['diseno_id'],version_diseno_id=request.DATA['version_diseno_id'])
					logs_model=self.model_log(usuario_id=request.user.usuario.id,accion=self.model_acciones.accion_actualizar,nombre_modelo=self.nombre_modulo,id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_201_CREATED)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	

#Fin Api rest para soporte estado comentario


@login_required
@transaction.atomic
def eliminar_fondos(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			AFondo.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.fondo',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.fondos')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def cambio_estado(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		estado=Estado()
		estado_fondo=EstadoFondo()
		
		for item in respuesta['lista']:
			diseno=Diseno.objects.filter(fondo_id=item['id'])
			if len(diseno)==0 or int(estado_fondo.activo)==int(estado.ObtenerID('Gestion_proyecto_fondo',respuesta['estado_id'])):
				fondo=AFondo.objects.get(id=item['id'])
				fondo.estado_id=estado.ObtenerID('Gestion_proyecto_fondo',respuesta['estado_id'])
				fondo.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.fondo',id_manipulado=item['id'])
				logs_model.save()
			else:
				transaction.savepoint_rollback(sid)
				return JsonResponse({'message':'No es posible cambiar el estado, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})
			
		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro ha sido actualizado exitosamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.estado')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
def export_excel_fondo(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="fondos.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Fondos')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0

	dato= request.GET['dato']
	estado_fondo=EstadoFondo()
	qset=(Q(empresa_id=request.user.usuario.empresa.id)&Q(estado_id=estado_fondo.activo))

	if (len(dato)>0):

		qset = qset &(
				Q(nombre__icontains=dato)|
				Q(codigo_bancario__icontains=dato)
				)
				
	fondos = AFondo.objects.filter(qset)

	worksheet.write('A1', 'Nombre', format1)
	worksheet.write('B1', 'Estado', format1)

	for item in fondos:
		worksheet.write(row, col,item.nombre,format2)
		worksheet.write(row, col+1,item.estado.nombre,format2)

		row +=1


	workbook.close()

	return response

@login_required
@transaction.atomic
def incluir_disenadores(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			empresa=Empresa.objects.get(id=item['id'])
			empresa.esDisenador=True
			empresa.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='empresa.empresa',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.disenadores')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def quitar_disenadores(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			diseno=Diseno.objects.filter(disenadores=item['id'])
			if len(diseno)==0:
				empresa=Empresa.objects.get(id=item['id'])
				empresa.esDisenador=False
				empresa.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='empresa.empresa',id_manipulado=item['id'])
				logs_model.save()
			else:
				transaction.savepoint_rollback(sid)
				return JsonResponse({'message':'No es posible quitar el disenador, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.disenadores')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def eliminar_datos_disenos(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DatoDiseno.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.dato_diseno',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.disenos')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def actualizar_orden(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			dato=DatoDiseno.objects.get(id=item['id'])
			dato.orden=item['orden']
			dato.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.dato_diseno',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.orden')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_campanas(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			ACampana.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.campanas',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.campanas')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def clonar_campana(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		datos=DocumentoEstado.objects.filter(campana_id=respuesta['id_campana_clonar'])

		for item in datos:
			documento=DocumentoEstado(campana_id=respuesta['campana_id'],estado_id=item.estado.id,nombre=item.nombre)
			documento.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.estado_documentos',id_manipulado=documento.id)
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
					'data':''})

	except Exception as e:
		functions.toLog(e,'gestion_proyecto.campanas')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def incluir_empresas(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			campana=CampanaEmpresa.objects.filter(campana_id=respuesta['campana_id'],empresa_id=item['id'])
			if len(campana)==0:
				empresa=CampanaEmpresa(campana_id=respuesta['campana_id'],empresa_id=item['id'],propietario=False)
				empresa.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.campana_empresa',id_manipulado=item['id'])
				logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.empresas')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def quitar_empresas(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		
		for item in respuesta['lista']:
			CampanaEmpresa.objects.get(pk=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.campana_empresa',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.empresa')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def eliminar_solicitantes(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			ASolicitante.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.solicitante',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.solicitantes')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def export_excel_solicitante(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="solicitantes.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Solicitantes')
	format1=workbook.add_format({'border':1,'font_size':15,'bold':True})
	format2=workbook.add_format({'border':1})

	row=1
	col=0
	solicitantes=None
	dato= request.GET['dato']

	if (len(dato)>0):

		qset = (
				Q(nombre__icontains=dato)
				)
				
		solicitantes = ASolicitante.objects.filter(qset)

	else:
		solicitantes=ASolicitante.objects.all()

	worksheet.write('A1', 'Nombre', format1)

	for item in solicitantes:
		worksheet.write(row, col,item.nombre,format2)

		row +=1


	workbook.close()

	return response


@login_required
def deshabilitardiseno(request):
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			diseno=Diseno.objects.get(pk=item['id'])
			diseno.activado=False
			diseno.save()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.diseno',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})

	except Exception as e:
		functions.toLog(e,'gestion_proyecto.disenos')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def eliminar_solicitudes(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			CSolicitud.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.solicitud',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.solicitudes')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def agregar_solicitudes(request):

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			diseno=Diseno.objects.get(id=item['id'])
			diseno.solicitudes.add(respuesta['id_solicitud'])
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.diseno',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.solicitudes')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
def actualizar_fechas(request):

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta:
			if int(item['id'])==0:
				estado=EstadoDiseno(estado_id=item['id_estado'],diseno_id=item['id_diseno'],fecha=item['fecha'],version_diseno_id=item['version_diseno_id'])
				estado.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.estado_diseno',id_manipulado=estado.id)
				logs_model.save()
			else:
				estado=EstadoDiseno.objects.get(id=item['id'])
				estado.fecha=item['fecha']
				estado.save()	
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.estado_diseno',id_manipulado=item['id'])
				logs_model.save()		

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def actualizar_info(request):

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta:
			if int(item['id'])==0:
				info=InfoDiseno(dato_diseno_id=item['id_dato'],diseno_id=item['id_diseno'],valor=item['valor'],version_diseno_id=item['version_diseno_id'])
				info.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.info_diseno',id_manipulado=info.id)
				logs_model.save()
			else:
				info=InfoDiseno.objects.get(id=item['id'])
				info.valor=item['valor']
				info.save()	
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.info_diseno',id_manipulado=item['id'])
				logs_model.save()		

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.info_diseno')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def export_excel_diseno(request):
	
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="informe_diseno.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Disenos')
		format1=workbook.add_format({'border':1,'font_size':8,'bold':True})
		format2=workbook.add_format({'border':1})

		disenos=None
		dato=request.GET.get('dato',None)
		fondo_id=request.GET.get('fondo_id',None)
		departamento_id=request.GET.get('departamento_id',None)
		municipio_id=request.GET.get('municipio_id',None)
		campana_id=request.GET.get('campana_id',None)
		solicitante_id=request.GET.get('solicitante_id',None)
		estado_id=request.GET.get('estado_id',None)
		disenadores_id=request.GET.get('disenadores_id',None)

		qset=(Q(activado=True)&Q(permiso_diseno_diseno__empresa__id=request.user.usuario.empresa.id))

		if dato and (len(dato)>0):
			qset =qset&(Q(nombre__icontains=dato))
		
		if fondo_id and int(fondo_id)>0:
			qset = qset&(Q(fondo__id=fondo_id))

		if departamento_id and int(departamento_id)>0:
			qset = qset&(Q(municipio__departamento__id=departamento_id))

		if municipio_id and int(municipio_id)>0:
			qset = qset&(Q(municipio__id=municipio_id))

		if campana_id and int(campana_id)>0:
			qset = qset&(Q(campana__id=campana_id))

		if solicitante_id and int(solicitante_id)>0:
			qset = qset&(Q(solicitante__id=solicitante_id))

		if estado_id and int(estado_id)>0:
			lista_id=[]
			estado=Diseno.objects.filter(activado=True)
			for item in estado:
				if item.estado and item.estado['estado__id']==estado_id:
						lista_id.append(item.id)
			qset = qset & (Q(id__in=(lista_id)))

		if disenadores_id and int(disenadores_id)>0:
			qset=qset&(Q(disenadores=disenadores_id))

	
		disenos=Diseno.objects.filter(qset)

		worksheet.write('A1', 'CAMPANA', format1)
		worksheet.write('B1', 'FONDO', format1)
		worksheet.write('C1', 'DEPARTAMENTO', format1)
		worksheet.write('D1', 'MUNICIPIO', format1)
		worksheet.write('E1', 'NOMBRE PROYECTO', format1)
		worksheet.write('F1', 'COSTO DE PROYECTO', format1)
		worksheet.write('G1', 'COSTO DE DISENO', format1)
		worksheet.write('H1', 'ESTADO', format1)

		estados=DatoDiseno.objects.all().order_by('orden')

		row=0
		col=8
		for item in estados:
			worksheet.write(row,col,item.nombre ,format1)
			col +=1

		row=1
		col=0

		for item2 in disenos:
			worksheet.write(row, col,item2.campana.nombre,format2)
			worksheet.write(row, col+1,item2.fondo.nombre,format2)
			worksheet.write(row, col+2,item2.municipio.departamento.nombre,format2)
			worksheet.write(row, col+3,item2.municipio.nombre,format2)
			worksheet.write(row, col+4,item2.nombre,format2)
			worksheet.write(row, col+5,item2.costo_proyecto,format2)
			worksheet.write(row, col+6,item2.costo_diseno,format2)
			if item2.estado is not None:
				worksheet.write(row, col+7,item2.estado['estado__nombre'],format2)
			else:
				worksheet.write(row, col+7,'',format2)

			estados=DatoDiseno.objects.all().order_by('orden')

			col=8
			for item3 in estados:
				info=InfoDiseno.objects.filter(diseno_id=item2.id,dato_diseno_id=item3.id)
				if len(info)>0:
					worksheet.write(row, col,info[0].valor,format2)
				else:
					worksheet.write(row, col,0,format2)

				col +=1

			row +=1
			col=0


		workbook.close()

		return response

	except Exception as e:
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	



@login_required
def export_excel_convocatoria(request):
	
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="informe_convocatoria.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Disenos')
		format1=workbook.add_format({'border':0,'font_size':9,'bold':True,'align':'center'})
		format2=workbook.add_format({'border':0,'font_size':9})

		disenos=None
		campana_id=request.GET.get('campana_id',None)

		qset=(Q(activado=True))
	
		if campana_id and int(campana_id)>0:
			qset = qset&(Q(campana__id=campana_id))

	
		disenos=Diseno.objects.filter(qset)

		worksheet.write('A2', 'Convocatoria', format1)
		worksheet.write('B2', 'Departamento', format1)
		worksheet.write('C2', 'Municipio', format1)
		worksheet.write('D2', 'Nombre del Proyecto', format1)
		worksheet.merge_range('E1:N1', 'Tipo de Archivo',format1)


		documentos=DocumentoEstado.objects.filter(campana__id=campana_id).order_by('estado__id')

		row=1
		col=4
		for item in documentos:
			worksheet.write(row,col,item.nombre ,format1)
			col +=1

		row=2
		col=0
		for item in disenos:
			worksheet.write(row, col,item.campana.nombre,format2)
			worksheet.write(row, col+1,item.municipio.departamento.nombre,format2)
			worksheet.write(row, col+2,item.municipio.nombre,format2)
			worksheet.write(row, col+3,item.nombre,format2)

			documentos=DocumentoEstado.objects.filter(campana__id=campana_id).order_by('estado__id')
			col=4
			for item2 in documentos:
				soporte=SoporteEstado.objects.filter(documento_estado_id=item2.id,estado_diseno__diseno__id=item.id)
				if len(soporte)>0:
					worksheet.write(row, col,"x",format2)
				else:
					worksheet.write(row, col,"",format2)

				col +=1


			row +=1
			col=0


		workbook.close()

		return response

	except Exception as e:
		functions.toLog(e,'gestion_proyecto.informe')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})	


@login_required
@transaction.atomic
def eliminar_puntos(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			MapaDiseno.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.mapa',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.puntos')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def guardar_puntos_soporte(request):

	try:		
		soporte= request.FILES['archivo']
		doc = openpyxl.load_workbook(soporte)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0


		contador=hoja.max_row - 1

		if int(contador) > 0:

			for fila in hoja.rows:
				if fila[0].value and fila[1].value and fila[2].value:
					if i == 0:
						i=1
					else:
						punto=MapaDiseno(nombre=fila[0].value,latitud=fila[1].value,longitud=fila[2].value,diseno_id=int(request.POST['id_diseno']),version_diseno_id=int(request.POST['version_diseno_id']))
						punto.save()
						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.mapa',id_manipulado=punto.id)
						logs_model.save()


		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.puntos')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def guardar_permisos(request):

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			lista=PermisoDiseno.objects.filter(diseno_id=item['id_diseno'],empresa_id=respuesta['id_empresa'])
			if len(lista)>0:
				permiso=PermisoDiseno.objects.get(pk=lista[0].id)
				permiso.consultar=item['consultar']
				permiso.editar=item['editar']
				permiso.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.permiso_diseno',id_manipulado=lista[0].id)
			else:
				permiso=PermisoDiseno(diseno_id=item['id_diseno'],empresa_id=respuesta['id_empresa'],consultar=item['consultar'],editar=item['editar'])
				permiso.save()
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.permiso_diseno',id_manipulado=permiso.id)
				logs_model.save()		
			

		#return HttpResponse(str('0'), content_type="text/plain")
		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})

	except Exception as e:
		functions.toLog(e,'gestion_proyecto.permisos')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


def cantidades_estado(request):

	try: 
		campana_id=request.GET.get('campana_id',None)
		disenos=None
		cursor = connection.cursor()
		if campana_id is not None:
			cursor.execute(""";with query as (select Max(estado_diseno.id) as id,estado_diseno.diseno_id 
							from gestion_proyecto_estado_diseno estado_diseno 
							group by estado_diseno.diseno_id)
							select count(estado_diseno.estado_id) as cont,esta.id,esta.nombre
							from gestion_proyecto_estado_diseno estado_diseno 
							inner join estado_estado esta on esta.id=estado_diseno.estado_id
							inner join query q on q.id=estado_diseno.id
							inner join gestion_proyecto_diseno diseno on diseno.id=estado_diseno.diseno_id
							inner join gestion_proyecto_permiso_diseno permiso on permiso.diseno_id=diseno.id
							where diseno.activado=1  and permiso.empresa_id="""+str(request.user.usuario.empresa.id)+"""
							and diseno.campana_id="""+str(campana_id)+"""
							group by esta.id,esta.nombre""")
				#disenos=Diseno.objects.filter(activado=True,permiso_diseno_diseno__empresa__id=request.user.usuario.empresa.id,campana_id=campana_id)		
		else:
			cursor.execute(""";with query as (select Max(estado_diseno.id) as id,estado_diseno.diseno_id 
							from gestion_proyecto_estado_diseno estado_diseno 
							group by estado_diseno.diseno_id)
							select count(estado_diseno.estado_id) as cont,esta.id,esta.nombre
							from gestion_proyecto_estado_diseno estado_diseno 
							inner join estado_estado esta on esta.id=estado_diseno.estado_id
							inner join query q on q.id=estado_diseno.id
							inner join gestion_proyecto_diseno diseno on diseno.id=estado_diseno.diseno_id
							inner join gestion_proyecto_permiso_diseno permiso on permiso.diseno_id=diseno.id
							where diseno.activado=1  and permiso.empresa_id="""+str(request.user.usuario.empresa.id)+"""
							group by esta.id,esta.nombre""")
			#disenos=Diseno.objects.filter(activado=True,permiso_diseno_diseno__empresa__id=request.user.usuario.empresa.id)		
		
		#estados=Estado.objects.filter(app='Gestion_proyecto_documento')
		List=[]

		for item in cursor.fetchall():
			# cont=0
			# for item2 in disenos:
			# 	if item2.estado is not None:
			# 		if item2.estado['estado__id']==item.id:
			# 			cont=cont+1

			# if cont>0:
			List.append({'id':item[1],'nombre':item[2],'cantidad_diseno':item[0]})
					
		
		return JsonResponse({'message':'','success':'ok','data':List})

	except Exception as e:
		print(e)
		functions.toLog(e,'gestion_proyecto.estados')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def reportar_reporte_diseno(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado_inicial=Estado.objects.filter(app='Gestion_proyecto_diseno_reporte',codigo=1)
		estado_correcion=Estado.objects.filter(app='Gestion_proyecto_diseno_reporte',codigo=2)

		version_diseno=DVersionesDiseno.objects.get(pk=respuesta['version_diseno_id'])
		version_diseno.reportar_diseno=True
		if version_diseno.estado_reporte is None:
			version_diseno.estado_reporte_id=estado_inicial[0].id
		else:
			version_diseno.estado_reporte_id=estado_correcion[0].id
		version_diseno.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.version_diseno',id_manipulado=respuesta['version_diseno_id'])
		logs_model.save()

		envioNotificacionReporteProyecto.delay(respuesta['version_diseno_id'],request.user.usuario.empresa.nombre)

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'gestion_proyecto.diseno')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
@transaction.atomic
def reportar_reporte_exitoso(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado_exitoso=Estado.objects.filter(app='Gestion_proyecto_diseno_reporte',codigo=4)	

		version_diseno=DVersionesDiseno.objects.get(pk=respuesta['version_diseno_id'])
		version_diseno.reportar_satisfaccion=True
		version_diseno.estado_reporte_id=estado_exitoso[0].id
		version_diseno.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.version_diseno',id_manipulado=respuesta['version_diseno_id'])
		logs_model.save()
		nombre_completo=request.user.usuario.persona.nombres+" "+request.user.usuario.persona.apellidos
		envioNotificacionReporteExitoso.delay(respuesta['version_diseno_id'],request.user.usuario.empresa.nombre,nombre_completo)

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'gestion_proyecto.diseno')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
@transaction.atomic
def reportar_insconsistencia(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		estado_novedad=Estado.objects.filter(app='Gestion_proyecto_diseno_reporte',codigo=3)	

		version_diseno=DVersionesDiseno.objects.get(pk=respuesta['version_diseno_id'])
		version_diseno.estado_reporte_id=estado_novedad[0].id
		version_diseno.reportar_diseno=False
		version_diseno.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='gestion_proyecto.version_diseno',id_manipulado=respuesta['version_diseno_id'])
		logs_model.save()

		fecha=datetime.now().strftime('%Y-%m-%d')	
		comentario=TComentarioDiseno(version_diseno_id=respuesta['version_diseno_id'],diseno_id=respuesta['diseno_id'],fecha=fecha,usuario_id=request.user.usuario.id,comentario=respuesta['comentario'])
		comentario.save()
		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='gestion_proyecto.comentario_diseno',id_manipulado=respuesta['diseno_id'])
		logs_model.save()

		nombre_completo=request.user.usuario.persona.nombres+" "+request.user.usuario.persona.apellidos
		envioNotificacionReporteInsconsitencia.delay(respuesta['version_diseno_id'],request.user.usuario.empresa.nombre,nombre_completo,respuesta['comentario'])

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'gestion_proyecto.comentario_diseno')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def eliminar_versiones(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DVersionesDiseno.objects.filter(id=item['id']).delete()
			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='gestion_proyecto.version_diseno',id_manipulado=item['id'])
			logs_model.save()

		#return HttpResponse(str('0'), content_type="text/plain")
		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error',
			'data':''})	
		
	except Exception as e:
		functions.toLog(e,'gestion_proyecto.version_diseno')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
@transaction.atomic
def consultar_tablero(request):
	sid = transaction.savepoint()
	try:
		diseno_id=request.GET.get('diseno_id',None)
		version_diseno_id=request.GET.get('version_diseno_id',None)

		ListEmpresas=PermisoDiseno.objects.filter(diseno_id=diseno_id).values('empresa__nombre','consultar','editar')
		Lista=[]
		ListEstado=Estado.objects.filter(app='Gestion_proyecto_documento').values('id','nombre').order_by('codigo')
		for item in ListEstado:
			fecha=''
			id_fecha=0
			ListaSoporte=[]
			valor=EstadoDiseno.objects.filter(diseno_id=diseno_id,version_diseno_id=version_diseno_id,estado_id=item['id']).values('fecha','id')
			if len(valor)>0:
				fecha=valor[0]['fecha']
				id_fecha=valor[0]['id']	
			
			Lista.append({'id':id_fecha,'id_estado':item['id'],'nombre':item['nombre'],'fecha':fecha})
				
			Lista2=[]
			ListDatos=DatoDiseno.objects.all().order_by('orden')
			for item in ListDatos:
				valor_dato=0
				id_valor=0
				valor=InfoDiseno.objects.filter(diseno_id=diseno_id,version_diseno_id=version_diseno_id,dato_diseno_id=item.id).values('valor','id')
				if len(valor) >0:
					valor_dato=valor[0]['valor']
					id_valor=valor[0]['id']	
				
				Lista2.append({'id':id_valor,'id_dato':item.id,'nombre':item.nombre,'unidad':item.unidad_medida.nombre,'valor':valor_dato});

		
		return JsonResponse({'message':'','success':'ok','data':{'empresas':list(ListEmpresas),'estados':Lista,'datos':Lista2}})
		
	except Exception as e:
		print(e)
		functions.toLog(e,'gestion_proyecto.version_diseno')
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@login_required
def descargar_plantilla(request):
	return functions.exportarArchivoS3("/plantillas/GestionProyectos/Formato_Localizacion_diseno.xlsx")	


@login_required
def gestion_proyecto(request):
	Campanas=CampanaEmpresa.objects.filter(empresa_id=request.user.usuario.empresa.id)
	return render(request, 'gestion_proyecto/index.html',{'model':'diseno','app':'gestion_proyecto','campanas':Campanas,})


@login_required
def opciones(request):
	return render(request, 'gestion_proyecto/opciones.html',{'model':'diseno','app':'gestion_proyecto',})


@login_required
def fondo(request):
	return render(request, 'gestion_proyecto/fondo.html',{'model':'afondo','app':'gestion_proyecto',})

@login_required
def disenadores(request):
	return render(request, 'gestion_proyecto/disenadores.html',{'model':'disenadores','app':'gestion_proyecto',})

@login_required
def datos_diseno(request):
	ListaUnidades=AUnidadMedida.objects.all()
	return render(request, 'gestion_proyecto/dato_diseno.html',{'model':'datodiseno','app':'gestion_proyecto','unidades':ListaUnidades,})

@login_required
def campana(request):
	return render(request, 'gestion_proyecto/campana.html',{'model':'acampana','app':'gestion_proyecto',})

@login_required
def documentos(request,id_campana):
	Campana=ACampana.objects.get(pk=id_campana)
	ListEstado=Estado.objects.filter(app='Gestion_proyecto_documento')
	ListClonacion=ACampana.objects.filter(~Q(id=id_campana))
	return render(request, 'gestion_proyecto/documentos.html',{'model':'documentoestado','app':'gestion_proyecto','campana':Campana,'estados':ListEstado,'campanas':ListClonacion,})

@login_required
def campana_empresa(request,id_campana):
	Campana=ACampana.objects.get(pk=id_campana)
	return render(request, 'gestion_proyecto/campana_empresa.html',{'model':'campanaempresa','app':'gestion_proyecto','campana':Campana,})

@login_required
def solicitante(request):
	return render(request, 'gestion_proyecto/solicitante.html',{'model':'asolicitante','app':'gestion_proyecto',})

@login_required
def permisos(request):
	ListEmpresas=Empresa.objects.all()
	ListDep=Departamento.objects.all()
	estado_fondo=EstadoFondo()
	ListFondos=AFondo.objects.filter(Q(empresa_id=request.user.usuario.empresa.id)&Q(estado_id=estado_fondo.activo))
	return render(request, 'gestion_proyecto/permisos.html',{'model':'permisodiseno','app':'gestion_proyecto','empresa':ListEmpresas,'departamentos':ListDep,'fondos':ListFondos,})


@login_required
def diseno(request):
	ListSolicitante=ASolicitante.objects.all()
	ListCampanas=CampanaEmpresa.objects.filter(empresa_id=request.user.usuario.empresa.id)
	ListDep=Departamento.objects.all()
	estado_fondo=EstadoFondo()
	ListFondos=AFondo.objects.filter(Q(empresa_id=request.user.usuario.empresa.id)&Q(estado_id=estado_fondo.activo))
	ListEmpresas=Empresa.objects.filter(esDisenador=True)
	ListEstado=Estado.objects.filter(app='Gestion_proyecto_documento')
	return render(request, 'gestion_proyecto/diseno.html',{'model':'diseno','app':'gestion_proyecto','solicitantes':ListSolicitante,'campanas':ListCampanas,'departamentos':ListDep,'fondos':ListFondos,'disenadores':ListEmpresas,'estado':ListEstado,})


@login_required
def tablero(request,version_diseno_id):
	version_diseno=DVersionesDiseno.objects.get(pk=version_diseno_id)
	diseno=Diseno.objects.get(pk=version_diseno.diseno.id)
	permiso=PermisoDiseno.objects.filter(diseno_id=diseno.id,empresa_id=request.user.usuario.empresa.id)
	return render(request, 'gestion_proyecto/tablero.html',{'model':'diseno','app':'gestion_proyecto','version':version_diseno,'version_diseno_id':version_diseno_id,'diseno_id':diseno.id,'diseno':diseno,'permiso':permiso[0]})


@login_required
def solicitudes(request):
	return render(request, 'gestion_proyecto/solicitud.html',{'model':'csolicitud','app':'gestion_proyecto',})


@login_required
def solicitud_diseno(request,id_solicitud):
	solicitud=CSolicitud.objects.get(pk=id_solicitud)
	ListDep=Departamento.objects.all()	
	estado_fondo=EstadoFondo()
	ListFondos=AFondo.objects.filter(Q(empresa_id=request.user.usuario.empresa.id)&Q(estado_id=estado_fondo.activo))
	return render(request, 'gestion_proyecto/solicitud_diseno.html',{'model':'gestion_proyecto','app':'gestion_proyecto','solicitud':solicitud,'departamentos':ListDep,'fondos':ListFondos,})


@login_required
def mapa(request):
	ListDep=Departamento.objects.all()
	estado_fondo=EstadoFondo()
	ListFondos=AFondo.objects.filter(Q(empresa_id=request.user.usuario.empresa.id)&Q(estado_id=estado_fondo.activo))
	return render(request, 'gestion_proyecto/mapa.html',{'model':'mapa_diseno','app':'gestion_proyecto','departamentos':ListDep,'fondos':ListFondos})

@login_required
def mapa_grande(request,id_diseno):
	version_diseno_id=DVersionesDiseno.objects.filter(diseno_id=id_diseno).values('id')[0]
	version_diseno_id = version_diseno_id['id']
	return render(request, 'gestion_proyecto/mapa_grande.html',{'model':'gestion_proyecto','app':'gestion_proyecto','version_diseno_id':version_diseno_id})


@login_required
def version_diseno(request,id_diseno):
	diseno=Diseno.objects.get(pk=id_diseno)
	return render(request, 'gestion_proyecto/version_diseno.html',{'model':'dversionesdiseno','app':'gestion_proyecto','diseno':diseno,'diseno_id':id_diseno})

@login_required
def VerSoporte(request):
	if request.method == 'GET':
		try:
			
			archivo = SoporteEstado.objects.get(pk=request.GET['id'])
			
			# filename = str(archivo.ruta)
			# extension = filename[filename.rfind('.'):]
			# nombre = re.sub('[^A-Za-z0-9 ]+', '', archivo.nombre)
			return functions.exportarArchivoS3(str(archivo.ruta))

		except Exception as e:
			functions.toLog(e,'gestion_proyecto.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def VerSoporteSolicitud(request):
	if request.method == 'GET':
		try:
			archivo = SoporteSolicitud.objects.get(pk=request.GET['id'])

			# filename = str(archivo.ruta)
			# extension = filename[filename.rfind('.'):]
			# nombre = re.sub('[^A-Za-z0-9 ]+', '', archivo.nombre)
			return functions.exportarArchivoS3(str(archivo.ruta))

		except Exception as e:
			functions.toLog(e,'gestion_proyecto.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)