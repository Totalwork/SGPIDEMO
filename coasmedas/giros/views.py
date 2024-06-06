# -*- coding: utf-8 -*-
from django.shortcuts import render,redirect,render_to_response
from django.urls import reverse
from rest_framework import viewsets, serializers
from django.db.models import Q
from estado.views import EstadoSerializer
from tipo.views import TipoSerializer
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import xlsxwriter
import json
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from .models import CNombreGiro,DEncabezadoGiro,DetalleGiro,RechazoGiro
from contrato.models import Contrato, EmpresaContrato
from contrato.views import ContratoSerializer
from cesion_economica.models import CesionEconomica
from financiero.models import FinancieroCuenta,FinancieroCuentaMovimiento
from financiero.views import FinancieroCuentaSerializer
from empresa.views import EmpresaSerializer
from parametrizacion.views import BancoSerializer
from usuario.models import Usuario
from parametrizacion.models import Banco
from empresa.models import Empresa
from estado.models import Estado
from tipo.models import Tipo
from correspondencia.models import CorrespondenciaConsecutivo , CorrespondenciaEnviada , CorresPfijo, CorrespondenciaSoporte
from proceso.models import FProcesoRelacion,BItem,GProcesoRelacionDato,HSoporteProcesoRelacionDato
from proyecto.models import Proyecto,Proyecto_empresas

from proceso.models import AProceso
from factura.models import Cesion, Compensacion
from logs.models import Logs,Acciones
from django.db import connection
from datetime import *
from django.db import transaction
from django.db.models.deletion import ProtectedError
from .enum import enumEstados,enumTipo, enumTipoPagoAnticipo
from django.contrib.auth.decorators import login_required
from django.conf import settings
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from proyecto.views import ProyectoSerializer
from coasmedas.functions import functions

from django.db.models import F, FloatField, Sum
from django.contrib.contenttypes.models import ContentType

from django.contrib.contenttypes.models import ContentType

import sys, os

# Serializador de contratista
class ProyectoContratoSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Proyecto.contrato
		fields=('id','id_contrato','id_proyecto')

# Serializador de contratista
class MacroContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre')


# Serializador de contratista
class ContratistaLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Empresa
		fields=('id','nombre')

# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	contratista = ContratistaLiteSerializer(read_only=True)
	contratista_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Empresa.objects.all())

	contratante = ContratistaLiteSerializer(read_only=True)
	contratante_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Empresa.objects.all())

	mcontrato = MacroContratoLiteSerializer(read_only=True)
	mcontrato_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Contrato.objects.all())

	# proyecto = ProyectoContratoSerializer(read_only=True)

	class Meta:
		model = Contrato
		fields=('id','nombre','contratista','contratista_id','mcontrato','mcontrato_id','contratante','contratante_id','numero')

		
#Api rest para nombre del giro
class NombreGiroSerializer(serializers.HyperlinkedModelSerializer):
	tipo = TipoSerializer(read_only=True)
	tipo_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='TipoGiro'))

	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoLiteSerializer(read_only=True)

	class Meta:
		model = CNombreGiro
		fields=('id','nombre','contrato','contrato_id','tipo','tipo_id')


class NombreGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de nombres de los giros utilizados en encabezado giro, se puede filtrar por dato, mcontrato y permite sin paginacion.
	"""
	model=CNombreGiro
	queryset = model.objects.all()
	serializer_class = NombreGiroSerializer
	nombre_modulo='giros.nombreGiros'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(NombreGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			mcontrato= self.request.query_params.get('contrato',None)
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			id_empresa = request.user.usuario.empresa.id

			if (dato or mcontrato or id_empresa):

				qset = Q(contrato__empresacontrato__empresa=id_empresa)

				if dato:

					qset = qset &(
						Q(nombre__icontains=dato)
					)


				if mcontrato and int(mcontrato)>0:

					qset = qset &(
						Q(contrato__id=mcontrato)
					)


			if qset != '':
				queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()

			try:
				serializer = NombreGiroSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(tipo_id=request.DATA['tipo_id'],contrato_id=request.DATA['contrato_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.nombre_giro',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = NombreGiroSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					serializer.save(tipo_id=request.DATA['tipo_id'],contrato_id=request.DATA['contrato_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.nombre_giro',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para nombre del giro



#Api rest para encabezado del giro
class EncabezadoGiroSerializer(serializers.HyperlinkedModelSerializer):
	contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoLiteSerializer(read_only=True)

	nombre_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=CNombreGiro.objects.all())
	nombre=NombreGiroSerializer(read_only=True)
	
	#nombre = serializers.PrimaryKeyRelatedField(queryset=NombreGiro.objects.all())

	porcentaje = serializers.SerializerMethodField('_porcentaje',read_only=True)
	sol_giro_suma_detalle = serializers.SerializerMethodField('_sol_giro_suma_detalle',read_only=True)

	pago_recurso = TipoSerializer(read_only=True)

	def _porcentaje(self,obj):
		# retorno=0
		# objProceso=0
		# proyecto = Proyecto.objects.filter(contrato=obj.contrato)

		# if len(AProceso.objects.filter(id=2))>0:
		# 	objProceso = AProceso.objects.get(pk=2)

		# qset2 = Q(proceso=objProceso)
		# listaProcesoRelacion = FProcesoRelacion.objects.filter(qset2).order_by('id').values()
		# usuarioApp = Usuario.objects.get(user=self.context['request'].user)
		# qset2 = Q(empresa=usuarioApp.empresa.id)
		# qset2 = qset2 & (Q(proyecto__id_in=proyecto.values('id')))
		# proyectoEmpresas = Proyecto_empresas.objects.filter(qset2).values('proyecto__id')							
		# for pr in listaProcesoRelacion:
		# 	qset2=Q(proyecto__id_in=proyectoEmpresas) & Q(empresa=usuarioApp.empresa.id)
		# 	qset2 = qset2 & (Q(proyecto__id=pr['idApuntador']))	
		# 	proy = Proyecto_empresas.objects.filter(qset2).values('proyecto__id',
		# 		'proyecto__mcontrato__nombre',
		# 		'proyecto__municipio__departamento__nombre',
		# 		'proyecto__municipio__nombre','proyecto__nombre')
		# 	if proy:	
		# 		if objProceso.tablaForanea:
		# 			modeloReferencia = ContentType.objects.get(pk=objProceso.tablaForanea.id).model_class()
		# 			elemento = modeloReferencia.objects.filter(
		# 				id=pr['idTablaReferencia']).values(objProceso.etiqueta)
		# 		else:
		# 			elemento = proy
		# 		qsprd = GProcesoRelacionDato.objects.filter(procesoRelacion=
		# 				FProcesoRelacion.objects.get(id=pr['id']))
		# 		tareas= float(qsprd.count())
		# 		tareasCumplidas = float(qsprd.filter(estado=1).count())
		# 		porcentaje = round((tareasCumplidas / tareas)*100,2)
		# 		retorno=porcentaje

		return 'retorno'

	# se suma los detalle solicitados y autorizado para solicitud de giro
	def _sol_giro_suma_detalle(self,obj):

		total_detalle = DetalleGiro.objects.filter(encabezado_id=obj.id, estado_id__in=[1,2]).aggregate(sol_suma_detalle=Sum('valor_girar'))
		return total_detalle['sol_suma_detalle']

	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre','nombre_id','contrato_id','contrato','soporte','referencia','num_causacion','texto_documento_sap','fecha_conta','disparar_flujo','numero_radicado','suma_detalle','flujo_test','porcentaje','sol_giro_suma_detalle' , 'pago_recurso' )


class EncabezadoGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de encabezado de los giros.
	"""
	model=DEncabezadoGiro
	queryset = model.objects.all()
	serializer_class = EncabezadoGiroSerializer
	nombre_modulo='giros.encabezadoGiro'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:

			queryset = super(EncabezadoGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			contrato_filtro= self.request.query_params.get('contrato_filtro',None)
			mcontrato_filtro= self.request.query_params.get('mcontrato_filtro',None)
			contratista_filtro= self.request.query_params.get('contratista_filtro',None)
			encabezado_id= self.request.query_params.get('encabezado_id',None)
			sin_paginacion=self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id
			referencia = self.request.query_params.get('referencia',None)
			flujo_test= self.request.query_params.get('flujotest',None)
			detallescompletos= self.request.query_params.get('detallescompletos',None)
			usuario = self.request.query_params.get('usuario', request.user.id)
				

			if (dato or contrato_filtro or contratista_filtro or id_empresa):
				# print "entro"
				qset = (Q(contrato__empresacontrato__empresa=id_empresa))
	
				if dato:
					qset = qset &(Q(contrato__nombre__icontains=dato) | Q(contrato__numero__icontains=dato))
	
	
				if encabezado_id and int(encabezado_id)>0:
					qset = qset &(Q(id=encabezado_id))
	
	
				if mcontrato_filtro and int(mcontrato_filtro)>0:
					qset = qset & (Q(contrato__mcontrato__id=mcontrato_filtro))
	
				if contrato_filtro and int(contrato_filtro)>0:
	
					qset = qset & (Q(contrato_id=contrato_filtro))
	
	
				if contratista_filtro and int(contratista_filtro)>0:
						
					qset = qset & (Q(contrato__contratista__id=contratista_filtro))
	
				if referencia:
					if int(referencia)==3:
						qset = qset & (Q(referencia=''))
					elif int(referencia)==1:
						qset = qset & (Q(referencia__gt=''))
					elif int(referencia)==2:
						qset = qset
							
				if flujo_test:
					qset = qset & (Q(flujo_test=flujo_test))
					
				if detallescompletos:
					if detallescompletos=='false':
						# qset = qset & (Q(detalle_encabezado_giro__test_op__lte=''))
						qset = qset & (Q(detalle_encabezado_giro__test_op='') | Q(detalle_encabezado_giro__test_op__isnull=True) )
					else:
						# qset = qset & (Q(detalle_encabezado_giro__test_op__gt=''))
						qset = qset & (Q(detalle_encabezado_giro__test_op__isnull=False) & Q(detalle_encabezado_giro__fecha_pago_esperada__isnull=False))


			#print qset
			if qset != '':
				if detallescompletos:
					if detallescompletos=='false':
						queryset = self.model.objects.filter(qset).distinct()
					else:
						queryset = self.model.objects.filter(qset).distinct().exclude(detalle_encabezado_giro__test_op='')
				else:
					queryset = self.model.objects.filter(qset)
	
			if sin_paginacion is None:	
		
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					#print serializer
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
					
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
				
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
					'data':serializer.data})
											
		except Exception as e:
			#print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':

			sid = transaction.savepoint()

			try:
				serializer = EncabezadoGiroSerializer(data=request.DATA,context={'request': request})
				#print serializer
				if serializer.is_valid():

					if self.request.FILES.get('soporte') is not None:
						#print self.request.FILES.get('soporte')
						serializer.save(soporte=self.request.FILES.get('soporte'),nombre_id=request.DATA['nombre_id'],contrato_id=request.DATA['contrato_id'] , pago_recurso_id = enumTipoPagoAnticipo.cuenta_bancaria)

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.encabezado_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
					else:
						#print self.request.FILES.get('soporte')
						serializer.save(soporte='',nombre_id=request.DATA['nombre_id'],contrato_id=request.DATA['contrato_id'] , pago_recurso_id = enumTipoPagoAnticipo.cuenta_bancaria  )

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.encabezado_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 			'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)
		

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':

			sid = transaction.savepoint()


			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = EncabezadoGiroSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					
					if self.request.FILES.get('soporte') is not None:
						#print self.request.FILES.get('soporte')
						serializer.save(soporte=self.request.FILES.get('soporte'),nombre_id=request.DATA['nombre_id'],contrato_id=request.DATA['contrato_id'])

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.encabezado_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
					else:
						encabezado=self.model.objects.get(pk=instance.id)
						serializer.save(soporte=encabezado.soporte,nombre_id=request.DATA['nombre_id'],contrato_id=request.DATA['contrato_id'])

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.encabezado_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)

					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	


#Fin api rest para encabezado del giro

#Serializer encabezado
class EncabezadoLiteGiroSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre')

#Serializer encabezado
class EncabezadoLiteSolGiroGiroSerializer(serializers.HyperlinkedModelSerializer):

	nombre=NombreGiroSerializer(read_only=True)

	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre','referencia')		


#Serializer financiero cuenta
class FinancieroListCuentaSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = FinancieroCuenta
		fields=('id','numero','nombre')



class CorresPfijoLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = CorresPfijo
		fields=('id', 'nombre')

#Serializer correspondencia consecutivo
class CorrespondenciaEnviadaLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	prefijo = CorresPfijoLiteSerializer(read_only=True)

	class Meta:
		model = CorrespondenciaEnviada
		fields=('id','consecutivo' , 'prefijo' , 'anoEnvio' )

#Api rest para detalle del giro
class DetalleGiroSolicitudSerializer(serializers.HyperlinkedModelSerializer):

	#encabezado= serializers.PrimaryKeyRelatedField(queryset=DEncabezadoGiro.objects.all())
	encabezado_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=DEncabezadoGiro.objects.all())
	encabezado=EncabezadoLiteSolGiroGiroSerializer(read_only=True)

	banco_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Banco.objects.all())
	banco=BancoSerializer(read_only=True)

	contratista_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())
	contratista=EmpresaSerializer(read_only=True)

	cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=FinancieroCuenta.objects.all(), allow_null=True)
	cuenta=FinancieroListCuentaSerializer(read_only=True)

	estado = EstadoSerializer(read_only=True)
	estado_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Estado.objects.filter(app='EstadoGiro'))

	tipo_cuenta = TipoSerializer(read_only=True)
	tipo_cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='cuenta'))

	rechazo = serializers.SerializerMethodField('_pagoRechazado',read_only=True)

	def _pagoRechazado(self,obj):

		rechazo = RechazoGiro.objects.filter(detalle=obj.id).first()

		if rechazo is not None:
			retorno = rechazo.atendido
		else:
			retorno = None

		return retorno



	class Meta:
		model = DetalleGiro
		fields=('id','encabezado','encabezado_id','contratista_id','contratista',
			'banco_id','banco','no_cuenta','tipo_cuenta','tipo_cuenta_id','valor_girar',
			'estado','fecha_pago','cuenta','cuenta_id','test_op',
			'fecha_pago_esperada','codigo_pago','estado_id','rechazo')		


#Serializer cesion
class CesionLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = Cesion
		fields=('id',)

#Serializer cesion
class CompensacionLiteSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = Compensacion
		fields=('id',)

#Api rest para detalle del giro
class DetalleGiroSerializer(serializers.HyperlinkedModelSerializer):

	#encabezado= serializers.PrimaryKeyRelatedField(queryset=DEncabezadoGiro.objects.all())
	encabezado_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=DEncabezadoGiro.objects.all())
	encabezado=EncabezadoLiteGiroSerializer(read_only=True)

	banco_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Banco.objects.all())
	banco=BancoSerializer(read_only=True)

	contratista_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())
	contratista=EmpresaSerializer(read_only=True)

	cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=FinancieroCuenta.objects.all(), allow_null=True)
	cuenta=FinancieroListCuentaSerializer(read_only=True)

	estado = EstadoSerializer(read_only=True)
	estado_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Estado.objects.filter(app='EstadoGiro'))

	tipo_cuenta = TipoSerializer(read_only=True)
	tipo_cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='cuenta'))

	carta_autorizacion_id = serializers.PrimaryKeyRelatedField(write_only=True, allow_null=True, queryset=CorrespondenciaEnviada.objects.all())
	carta_autorizacion=CorrespondenciaEnviadaLiteSerializer(read_only=True)

	cesion_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Cesion.objects.all(), allow_null=True)
	cesion=CesionLiteSerializer(read_only=True)

	cruce_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Compensacion.objects.all(), allow_null=True)
	cruce=CompensacionLiteSerializer(read_only=True)

	contratistaContrato = serializers.SerializerMethodField()

	class Meta:
		model = DetalleGiro
		fields=('id','encabezado','encabezado_id','contratista_id','contratista',
			'banco_id','banco','no_cuenta','tipo_cuenta','tipo_cuenta_id','valor_girar',
			'estado','fecha_pago','cuenta','cuenta_id','test_op',
			'fecha_pago_esperada','codigo_pago','estado_id',
			'carta_autorizacion','carta_autorizacion_id','soporte_consecutivo_desabilitado',
			'cesion_id','cesion','cruce_id','cruce', 'contratistaContrato')

	def get_contratistaContrato(self,obj):
		resul = DEncabezadoGiro.objects.filter(id=obj.encabezado_id).values('id','contrato__contratista').first()

		return resul['contrato__contratista']

@login_required
def VerSoporteConsecutivoDeshabilitado(request):
	if request.method == 'GET':
		try:
			
			archivo = DetalleGiro.objects.get(pk=request.GET['id'])			
			return functions.exportarArchivoS3(str(archivo.soporte_consecutivo_desabilitado))

		except Exception as e:
			functions.toLog(e,'DetalleGiro.VerSoporteDeshabilitado')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			


class DetalleGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de los detalles del giros de cada encabezado.
	"""
	model=DetalleGiro
	queryset = model.objects.all()
	serializer_class = DetalleGiroSerializer
	nombre_modulo='giros.detalleGiro'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			encabezado_id= self.request.query_params.get('encabezado_id',None)
			sin_pago = self.request.query_params.get('sinpago',None)
			liteversion=self.request.query_params.get('lite',None)
			mcontrato=self.request.query_params.get('mcontrato_filtro',None)
			contratista=self.request.query_params.get('contratista_filtro',None)
			fechadesde=self.request.query_params.get('fechadesde',None)
			fechahasta=self.request.query_params.get('fechahasta',None)
			fechapago=self.request.query_params.get('fechapago',None)
			referencia=self.request.query_params.get('referencia',None)
			id_empresa = request.user.usuario.empresa.id
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)

			# se usa para solicitud de giros
			sol_giro_solicitados_autorizado = self.request.query_params.get('sol_giro_solicitados_autorizado', None)

			qset=None

			if (dato or encabezado_id or id_empresa or sol_giro_solicitados_autorizado):

				qset = (Q(encabezado__contrato__empresacontrato__empresa=id_empresa))

				if dato:
					qset = qset &(
								Q(contratista__nombre__icontains=dato) | Q(banco__nombre__icontains=dato) | Q(no_cuenta__icontains=dato)
						)

				if encabezado_id:
					qset = qset &(Q(encabezado__id=encabezado_id))

				if sin_pago:
					qset = qset &(Q(test_op__gt=''))

				if mcontrato:
					qset = qset &(Q(encabezado__nombre__contrato__id=mcontrato))

				if contratista:
					qset = qset &(Q(contratista__id=contratista))

				if referencia:
					qset = qset &(Q(encabezado__referencia__gt=''))

				if fechadesde:
					qset = qset &(Q(fecha_pago_esperada__gte=fechadesde))

				if fechahasta:					
					qset = qset &(Q(fecha_pago_esperada__lte=fechahasta))

				if fechapago:
					qset = qset &(Q(fecha_pago=None))

				if sol_giro_solicitados_autorizado:
					# giros solicitados y autorizados
					qset = qset &(Q(estado_id__in=[1,2]))

			
			if qset!=None:
				queryset = self.model.objects.filter(qset)


			if liteversion:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = DetalleGiroSolicitudSerializer(page,many=True,context={'request':request})	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
				serializer = DetalleGiroSolicitudSerializer(queryset,many=True,context={'request':request})
				return Response({'message':'','success':'ok',
						'data':serializer.data})

			
			if sin_paginacion is None:	

				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

			else:

				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
				

		except Exception as e:
			# print(e);
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':

			sid = transaction.savepoint()

			try:

				serializer = DetalleGiroSerializer(data=request.DATA,context={'request': request})
 				#print serializer
				if serializer.is_valid():

						serializer.save(cesion_id=request.DATA['cesion_id'] if request.DATA['cesion_id'] !='' else None,soporte_consecutivo_desabilitado=self.request.FILES.get('soporte_consecutivo_desabilitado') if self.request.FILES.get('soporte_consecutivo_desabilitado') is not None else '',
							estado_id=request.DATA['estado_id'],contratista_id=request.DATA['contratista_id'],
							banco_id=request.DATA['banco_id'],tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
							encabezado_id=request.DATA['encabezado_id'],
							carta_autorizacion_id=request.DATA['carta_autorizacion_id'] if request.DATA['carta_autorizacion_id'] !='' else None,
							cuenta_id=request.DATA['cuenta_id'] if request.DATA['cuenta_id'] !='' else None,
							cruce_id=request.DATA['cruce_id'] if request.DATA['cruce_id'] !='' else None)

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.detalle_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					#print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 			'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)



	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':

			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = DetalleGiroSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					detalle=self.model.objects.get(pk=instance.id)

					serializer.save(cesion_id=request.DATA['cesion_id'] if request.DATA['cesion_id'] !='' else None,soporte_consecutivo_desabilitado=self.request.FILES.get('soporte_consecutivo_desabilitado') if self.request.FILES.get('soporte_consecutivo_desabilitado') is not None else '',
							estado_id=request.DATA['estado_id'],contratista_id=request.DATA['contratista_id'],
							banco_id=request.DATA['banco_id'],tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
							encabezado_id=request.DATA['encabezado_id'],
							cuenta_id=request.DATA['cuenta_id'] if request.DATA['cuenta_id'] !='' else None,
							cruce_id=request.DATA['cruce_id'] if request.DATA['cruce_id'] !='' else None)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giro',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)


	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	



#Api rest para detalle del giro
class RechazoGiroSerializer(serializers.HyperlinkedModelSerializer):


	detalle= DetalleGiroSerializer(read_only=True)

	class Meta:
		model = RechazoGiro
		fields=('id','detalle','fecha','motivo','atendido')


class RechazoGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de los detalles del giros de cada encabezado.
	"""
	model=RechazoGiro
	queryset = model.objects.all()
	serializer_class = RechazoGiroSerializer
	nombre_modulo='giros.rechazoGiro'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(RechazoGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			desde = self.request.query_params.get('desde', None)
			hasta = self.request.query_params.get('hasta', None)
			contratista = self.request.query_params.get('contratista', None)
			testop = self.request.query_params.get('testop', None)
			rechazados = self.request.query_params.get('rechazados', None)


			id_empresa = request.user.usuario.empresa.id
			qset=None

			if (dato or id_empresa):

				qset = (Q(detalle__encabezado__contrato__empresacontrato__empresa=id_empresa))

				if dato:
					qset = qset &(
								Q(contratista__nombre__icontains=dato) | Q(banco__nombre__icontains=dato) | Q(no_cuenta__icontains=dato)
						)

				if rechazados:
					qset = qset &(Q(atendido=0))				

				if desde:
					qset = qset &(Q(fecha__gte=desde))

				if hasta:
					qset = qset &(Q(fecha__lte=hasta))

				if contratista:
					qset = qset &(Q(detalle__contratista__id=contratista))

				if testop:
					qset = qset &(Q(detalle__codigo_pago=testop))					

			
			if qset!=None:
				queryset = self.model.objects.filter(qset)

			page = self.paginate_queryset(queryset)
			if page is not None:
				serializer = self.get_serializer(page,many=True)	
				return self.get_paginated_response({'message':'','success':'ok',
				'data':serializer.data})
	
			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
					'data':serializer.data})			
		except Exception as e:
			#print(e);
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':

			sid = transaction.savepoint()

			try:

				serializer = RechazoGiroSerializer(data=request.DATA,context={'request': request})
 				#print serializer
				if serializer.is_valid():

						serializer.save(cesion_id=request.DATA['cesion_id'] if request.DATA['cesion_id'] !='' else None,soporte_consecutivo_desabilitado=self.request.FILES.get('soporte_consecutivo_desabilitado') if self.request.FILES.get('soporte_consecutivo_desabilitado') is not None else '',
							estado_id=request.DATA['estado_id'],contratista_id=request.DATA['contratista_id'],
							banco_id=request.DATA['banco_id'],tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
							encabezado_id=request.DATA['encabezado_id'],
							carta_autorizacion_id=request.DATA['carta_autorizacion_id'] if request.DATA['carta_autorizacion_id'] !='' else None,
							cuenta_id=request.DATA['cuenta_id'] if request.DATA['cuenta_id'] !='' else None,
							cruce_id=request.DATA['cruce_id'] if request.DATA['cruce_id'] !='' else None)

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='giros.detalle_giro',id_manipulado=serializer.data['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)
						
						return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					#print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 			'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)



	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':

			sid = transaction.savepoint()

			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = RechazoGiroSerializer(instance,data=request.DATA,context={'request': request},partial=partial)
				if serializer.is_valid():
					detalle=self.model.objects.get(pk=instance.id)

					serializer.save(cesion_id=request.DATA['cesion_id'] if request.DATA['cesion_id'] !='' else None,soporte_consecutivo_desabilitado=self.request.FILES.get('soporte_consecutivo_desabilitado') if self.request.FILES.get('soporte_consecutivo_desabilitado') is not None else '',
							estado_id=request.DATA['estado_id'],contratista_id=request.DATA['contratista_id'],
							banco_id=request.DATA['banco_id'],tipo_cuenta_id=request.DATA['tipo_cuenta_id'],
							encabezado_id=request.DATA['encabezado_id'],
							cuenta_id=request.DATA['cuenta_id'] if request.DATA['cuenta_id'] !='' else None,
							cruce_id=request.DATA['cruce_id'] if request.DATA['cruce_id'] !='' else None)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giro',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
		  				

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)				


#Fin api rest para encabezado del giro

@login_required
def descargar_plantilla(request):
	return functions.exportarArchivoS3('plantillas/verificar_giro/plantillaGiro.xlsx')
	

@login_required
def giros(request):
	return render(request, 'giros/encabezado_giro.html',{'app':'giros','model':'dencabezadogiro'})


@login_required
def detalle_giros(request,id_encabezado=None,mcontrato=None,contrato=None):
	return render(request, 'giros/detalle_giro.html',{'app':'giros','model':'detallegiro','id_encabezado':id_encabezado,'mcontrato':mcontrato,'contrato':contrato})		


@login_required
def detalle_giros_solo_lectura(request,id_encabezado=None,mcontrato=None,contrato=None):
	return render(request, 'giros/detalle_giro_solo_lectura.html',{'app':'giros','model':'detallegiro','id_encabezado':id_encabezado,'mcontrato':mcontrato,'contrato':contrato})		

#eliminar los encabezados giros
@transaction.atomic
def eliminar_varios_id(request):
	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')
		
		for item in respuesta['lista']:
			DEncabezadoGiro.objects.filter(id=item['id']).delete()

		#return HttpResponse(str('0'), content_type="text/plain")

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.encabezado_giro',id_manipulado=item['id'])
			logs_model.save()

		transaction.savepoint_commit(sid)
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})			 
		


#Actualiza el numero del radicado y la fecha en la pestana no.radicado
@transaction.atomic
def actualizar_radicado(request):

	sid = transaction.savepoint()
	try:
		
		lista=request.POST['_content']
		respuesta= json.loads(lista)		
		id = respuesta['id']
		numero_radicado = respuesta['numero_radicado']
		fecha_conta = respuesta['fecha_conta']

		if fecha_conta=='null':
			fecha_conta=None

		giros=DEncabezadoGiro.objects.get(pk=id)
		giros.numero_radicado=numero_radicado
		giros.fecha_conta=fecha_conta
		giros.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.encabezado_giro',id_manipulado=id)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		


#eliminar el detalle de los giros
@transaction.atomic
def eliminar_varios_id_detalle(request):

	sid = transaction.savepoint()
	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		#print respuesta['lista'].split()

		#lista=respuesta['lista'].split(',')

		for item in respuesta['lista']:

			if item['estado_id']== enumEstados.solicitado:

				# DetalleGiro.objects.filter(id=item['id']).delete()
				model_d = DetalleGiro.objects.get(id=item['id'])
				if not model_d.cesion_id and not model_d.cruce_id:

					model_d.delete()
					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='giros.detalle_giro',id_manipulado=item['id'])
					logs_model.save()
				else:
					# transaction.savepoint_rollback(sid)

					return JsonResponse({'message':'No se puede eliminar los giros que tiene cesión o están cruzados.','success':'error','data':''})

				transaction.savepoint_commit(sid)
			else:
				return JsonResponse({'message':'Solo puede eliminar los giros en estado solicitado','success':'error','data':''})
								
		return JsonResponse({'message':'El registro se ha eliminado correctamente','success':'ok','data':''})

	except ProtectedError:
		return JsonResponse({'message':'No es posible eliminar el registro, se esta utilizando en otra seccion del sistema','success':'error','data':''})	

	except Exception as e:
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#funcion para traer el encabezado del detalle del giro
def encabezado_detalle_giro(request):

	cursor = connection.cursor()
	try:
		encabezado_id = request.GET['encabezado_id']
		contrato = request.GET['contrato']
		proyecto = request.GET['proyecto']

		cursor.callproc('[dbo].[consultar_encabezado_detalle_giro]', [encabezado_id,contrato,proyecto])
		#if cursor.return_value == 1:			
		result_set = cursor.fetchall()
		lista=[]
		for x in list(result_set):
			item={
				'nombre_contratante':x[0],
				'nombre_contratista':x[1],
				'numero_contrato':x[2],
				'nombre_proyecto':x[3],
				'giro_nombre':x[4],
				'suma_valor_detalles':x[5],
				'soporte_giro':x[6],
				'nombre_contrato':x[7],
				'soporte':x[8],
				'departamento':x[9],
				'municipio':x[10],
				'contrato_id':x[11],

			}
			lista.append(item)
		
		return JsonResponse({'message':'','success':'ok','data':lista})	
	except Exception as e:

		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)	
	finally:
		cursor.close()


#guardar el pago de los detalles de los giros
@transaction.atomic
def guardar_pago_detalle(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		valida_registro_autorizacion = 0

		insert_list_cuenta_movimiento = []
		insert_list_detalle_giros = []

		for item in respuesta['lista']:
			object_detalle=DetalleGiro.objects.get(pk=item['id'])
			fecha_pa=datetime.strptime(respuesta['fecha_pago'], '%Y-%m-%d').date()
			estado=object_detalle.estado.id
			#carta_autorizacion=object_detalle.carta_autorizacion.id
			# ddiff=fecha_autor-fecha_pa
			# print ddiff

			if estado==enumEstados.autorizado:

				object_detalle.cuenta_id=respuesta['cuenta']
				object_detalle.fecha_pago=respuesta['fecha_pago']
				object_detalle.estado_id=enumEstados.pagado
				#object_detalle.estado_id=9
				object_detalle.save()

				
				insert_list_detalle_giros.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item['id']))
	

				object_encabezado_giro = DEncabezadoGiro.objects.get(pk = object_detalle.encabezado_id)

				# solo se hace egreso a las cuentas bancarias en caso que el pago del anticipo sea con cuenta bancaria
				if object_encabezado_giro.pago_recurso_id == enumTipoPagoAnticipo.cuenta_bancaria:
					finan=FinancieroCuentaMovimiento(cuenta_id=object_detalle.cuenta.id,tipo_id=enumTipo.egreso,valor=object_detalle.valor_girar,fecha=object_detalle.fecha_pago,descripcion='contrato numero '+str(respuesta['numero_contrato'])+'-  '+str(respuesta['anticipo'])+'- Beneficiario '+str(object_detalle.contratista.nombre))
					finan.save()
					insert_list_cuenta_movimiento.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='financiero.cuenta_movimiento',id_manipulado=finan.id))			


				object_cesion=CesionEconomica.objects.filter(
					contrato_id=object_encabezado_giro.contrato_id,
					nombre_giro_id=object_encabezado_giro.nombre_id,
					proveedor_id=object_detalle.contratista_id,
					banco_id=object_detalle.banco_id).first()
				if object_cesion:
					object_cesion.estado_id=142
					object_cesion.save()

			elif valida_registro_autorizacion==0 :
				valida_registro_autorizacion = 1

		# se hace una sola transaccion con bulk create
		if insert_list_detalle_giros:
			Logs.objects.bulk_create(insert_list_detalle_giros)

		if insert_list_cuenta_movimiento:
			Logs.objects.bulk_create(insert_list_cuenta_movimiento)

		transaction.savepoint_commit(sid)
		if valida_registro_autorizacion==1 :

			return JsonResponse({'message':'Solo se pueden pagar giros en estado de autorizacion','success':'error',
					'data':''})
		else:

			return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#exporta todos los detalles del giro
def export_excel_detalle_giro(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="detalle_giro.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Detalle')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':1})
	format3=workbook.add_format({'border':0,'font_size':12})
	format5=workbook.add_format()
	format5.set_num_format('yyyy-mm-dd')

	row=12
	col=0

	cursor = connection.cursor()

	encabezado_id= request.GET['encabezado_id']
	contrato= request.GET['contrato']
	proyecto= request.GET['proyecto']

	if encabezado_id:
	
		qset = (
				Q(encabezado_id=encabezado_id)
			)

				
		detalle = DetalleGiro.objects.filter(qset)

		cursor.callproc('[dbo].[consultar_encabezado_detalle_giro]', [encabezado_id,contrato,proyecto])
		#if cursor.return_value == 1:			
		result_set = cursor.fetchall()
		lista=[]
		for x in list(result_set):
			item={
				'nombre_contratante':x[0],
				'nombre_contratista':x[1],
				'numero_contrato':x[2],
				'nombre_proyecto':x[3],
				'giro_nombre':x[4],
				'suma_valor_detalles':x[5],

			}
			lista.append(item)

		worksheet.write('A5', 'Contratante:', format3)
		worksheet.write('B5', lista[0]['nombre_contratante'], format3)

		worksheet.write('A6', 'Proyecto', format3)
		worksheet.write('B6', lista[0]['nombre_proyecto'], format3)

		worksheet.write('A7', 'Contratista', format3)
		worksheet.write('B7', lista[0]['nombre_contratista'], format3)

		worksheet.write('A8', 'Convenio', format3)
		worksheet.write('B8', lista[0]['numero_contrato'], format3)

		worksheet.write('A9', 'Nombre anticipo', format3)
		worksheet.write('B9', lista[0]['giro_nombre'], format3)

		worksheet.write('A12', 'Beneficiario/Proveedor', format1)
		worksheet.write('B12', 'Nit', format1)
		worksheet.write('C12', 'Entidad bancaria', format1)
		worksheet.write('D12', 'No. cuenta destino', format1)
		worksheet.write('E12', 'Tipo cuenta destino', format1)
		worksheet.write('F12', 'Valor neto a girar', format1)
		worksheet.write('G12', 'Estado', format1)
		worksheet.write('H12', 'Carta autorizacion', format1)
		worksheet.write('I12', 'Fecha pago', format1)

		worksheet.write('A17', 'Total anticipos', format1)
		worksheet.write('B17', lista[0]['suma_valor_detalles'], format3)

		worksheet.set_column('A:A', 24)
		worksheet.set_column('B:B', 18)
		worksheet.set_column('C:C', 18)
		worksheet.set_column('D:D', 19)
		worksheet.set_column('E:E', 24)
		worksheet.set_column('F:F', 18)
		worksheet.set_column('G:G', 10)
		worksheet.set_column('H:H', 18)
		worksheet.set_column('I:I', 12)

		for detalle in detalle:
			worksheet.write(row, col,detalle.contratista.nombre,format2)
			worksheet.write(row, col+1,detalle.contratista.nit,format2)
			worksheet.write(row, col+2,detalle.banco.nombre,format2)
			worksheet.write(row, col+3,detalle.no_cuenta,format2)
			worksheet.write(row, col+4,detalle.tipo_cuenta.nombre,format2)
			worksheet.write(row, col+5,detalle.valor_girar,format2)
			worksheet.write(row, col+6,detalle.estado.nombre,format2)
			
			if detalle.carta_autorizacion:
				worksheet.write(row, col+7,detalle.carta_autorizacion.numero,format2)

			else:

				worksheet.write(row, col+7,'No asociado',format2)

			worksheet.write(row, col+8,detalle.fecha_pago,format5)

			row +=1


	workbook.close()

	return response
    #return response							


#reversar el detalle de los giros
@transaction.atomic
def reversar_giros(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		valida_registro_pagado = 0
		insert_list_cuenta_movimiento = []
		insert_list_detalle_giros = []

		for item in respuesta['lista']:
			object_detalle=DetalleGiro.objects.get(pk=item['id'])
			estado=object_detalle.estado_id


			if estado == enumEstados.pagado:

				#object_detalle.estado_id=7
				object_detalle.estado_id = enumEstados.reversado
				object_detalle.save()

				insert_list_detalle_giros.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item['id']))

				object_encabezado_giro = DEncabezadoGiro.objects.get(pk = object_detalle.encabezado_id)

				# solo se hace egreso a las cuentas bancarias en caso que el pago del anticipo sea con cuenta bancaria
				if object_encabezado_giro.pago_recurso_id == enumTipoPagoAnticipo.cuenta_bancaria:
					finan = FinancieroCuentaMovimiento(cuenta_id=object_detalle.cuenta.id, tipo_id=enumTipo.ingreso, valor=object_detalle.valor_girar, fecha=object_detalle.fecha_pago, descripcion='giro reversado : contrato numero '+str(object_encabezado_giro.contrato.numero)+'-  '+str(object_encabezado_giro.contrato.nombre)+'- Beneficiario '+str(object_detalle.contratista.nombre))
					finan.save()
					insert_list_cuenta_movimiento.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='financiero.cuenta_movimiento',id_manipulado=finan.id))

				transaction.savepoint_commit(sid)

			elif valida_registro_pagado==0 :
				valida_registro_pagado = 1


		# se hace una sola transaccion con bulk create
		if insert_list_detalle_giros:
			Logs.objects.bulk_create(insert_list_detalle_giros)

		if insert_list_cuenta_movimiento:
			Logs.objects.bulk_create(insert_list_cuenta_movimiento)


		transaction.savepoint_commit(sid)
		if valida_registro_pagado==1 :

			return JsonResponse({'message':'Solo se pueden reversar giros en estado pagado','success':'error',
					'data':''})
		else:

			return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
					'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


#autorizar el detalle de los giros
@transaction.atomic
def autorizar_giros(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)


		numero= respuesta['numero']
		ano= respuesta['ano']
		empresa= respuesta['empresa']
		prefijo= respuesta['prefijo']
		qset=''

		if numero and int(numero)>0:
			qset = (Q(consecutivo=int(numero)))

		if ano and int(ano)>0:
			qset = qset &(Q(anoEnvio = ano))

		if empresa and int(empresa)>0:
			qset = qset &(Q(prefijo__empresa__id = empresa))	

		if prefijo and int(prefijo)>0:
			qset = qset &(Q(prefijo_id = prefijo))	

		correspondencia = CorrespondenciaEnviada.objects.filter(qset)

		pr=None
		if correspondencia:
			soporteCorrespondencia=CorrespondenciaSoporte.objects.filter(
									correspondencia__id=correspondencia.first().id
									).values('id')

			if soporteCorrespondencia:

				for item in respuesta['lista']:
					object_detalle=DetalleGiro.objects.get(pk=item['id'])
					estado=object_detalle.estado_id

					if object_detalle.estado_id==enumEstados.solicitado:
						#object_detalle.estado_id=7
						object_detalle.estado_id=enumEstados.autorizado
						object_detalle.carta_autorizacion_id=correspondencia.first().id
						object_detalle.save()

						if pr is None:
							#inicio codigo para guardar en el seguimiento del giro la carta de autorizacion de giro.
							#validar q este implementado el proceso con el encabezadogiro de object_detalle
							pr = FProcesoRelacion.objects.filter(proceso__id=2,
																idApuntador=object_detalle.encabezado.contrato.id,
																idTablaReferencia=object_detalle.encabezado.id
																).first()
							if pr is None:
								#implementar el proceso si no esta implementado
								pr = FProcesoRelacion(proceso=AProceso.objects.get(id=2),
													idApuntador=object_detalle.encabezado.contrato.id,
													idTablaReferencia=object_detalle.encabezado.id)
								pr.save()
								#creo los registos en procesoRelacionDato
								items = BItem.objects.filter(proceso_id=2).values('id')
								for item in items:
									prd=GProcesoRelacionDato(procesoRelacion=pr,
															item=BItem.objects.get(id=item['id'])
															)
									if item['id']==41:
										prdObj=prd

									prd.save()
							else:
								#identificar el procesoRelacion, el procesoRelacionDato de la implementacion
								prdObj = GProcesoRelacionDato.objects.filter(procesoRelacion__id=pr.id,
																			item__id=41).first()

							prdObj.estado='1'
							prdObj.valor=str(correspondencia.first().fechaEnvio)
							prdObj.save()
							#crear el registro en el soporteProcesoRelacionDato tomando el correspondenciaSoporte
							#del objecto correspondencia
							#import pdb; pdb.set_trace()						

							for soporte in soporteCorrespondencia:
								carta = CorrespondenciaSoporte.objects.get(id=soporte['id'])
								soportePrd = HSoporteProcesoRelacionDato(procesoRelacionDato=prdObj,
																	nombre=carta.nombre,
																	documento=carta.soporte)
								soportePrd.save()


							#fin codigo para guardar en el seguimiento del giro la carta de autorizacion de giro. 

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item['id'])
						logs_model.save()

						transaction.savepoint_commit(sid)

				return JsonResponse({'message':'Los registros se han actualizado correctamente','success':'ok',
							'data':''})
			else:
				return JsonResponse({'message':'El consecutivo ingresado no tiene cargado soportes, por favor\
					cargue el soporte en el modulo de correspondencia enviada',
					'success':'error',
					'data':''})	

		return JsonResponse({'message':'No se encontraron consecutivos con los parametros ingresados','success':'error',
		'data':''})	

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




#desautorizar los giros
@transaction.atomic
def desautorizar_giros(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			object_detalle=DetalleGiro.objects.get(pk=item['id'])
			estado=object_detalle.estado_id

			if object_detalle.estado_id==enumEstados.autorizado:

				soporte=str(object_detalle.soporte_consecutivo_desabilitado)

				object_detalle.estado_id=enumEstados.solicitado
				object_detalle.soporte_consecutivo_desabilitado=None
				object_detalle.carta_autorizacion_id=None
				object_detalle.save()

				# functions.eliminarArchivoS3(soporte)

				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item['id'])
				logs_model.save()

				transaction.savepoint_commit(sid)

			else:

				return JsonResponse({'message':'Solo se puede desautorizar giros en estado autorizado','success':'error','data':''})


		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


		
#exporta todos los detalles del giro
# def export_excel_encabezado_giro(request):
	
# 	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
# 	response['Content-Disposition'] = 'attachment; filename="Encabezado del giro.xls"'
	
# 	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
# 	worksheet = workbook.add_worksheet('Giros')
# 	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
# 	format3=workbook.add_format({'border':0,'font_size':12})

# 	row=1
# 	col=0

# 	cursor = connection.cursor()

# 	id_empresa = request.user.usuario.empresa.id
# 	contrato = request.GET['contrato']

# 	cursor.callproc('[dbo].[generar_informe_encabezado_giro3]', [id_empresa,contrato])
				
# 	result_set = cursor.fetchall()
# 	lista=[]

# 	worksheet.write('A1', 'Contrato macro', format1)
# 	worksheet.write('B1', 'Interventor', format1)
# 	worksheet.write('C1', 'Contratista', format1)
# 	worksheet.write('D1', 'No. contrato', format1)
# 	worksheet.write('E1', 'Nombre anticipo', format1)
# 	worksheet.write('F1', 'Beneficiario / Proveedor', format1)
# 	worksheet.write('G1', 'Nit', format1)
# 	worksheet.write('H1', 'Entidad bancaria', format1)
# 	worksheet.write('I1', 'No. cuenta destino', format1)
# 	worksheet.write('J1', 'Tipo cuenta destino', format1)
# 	worksheet.write('K1', 'Valor neto', format1)
# 	worksheet.write('L1', 'Estado', format1)
# 	worksheet.write('M1', 'Fecha radicado', format1)
# 	worksheet.write('N1', 'Carta autorizacion', format1)
# 	worksheet.write('O1', 'Fecha pago', format1)
# 	worksheet.write('P1', 'Linea', format1)
# 	worksheet.write('Q1', 'Cuenta origen', format1)
# 	worksheet.write('R1', 'Rerencia', format1)
# 	worksheet.write('S1', 'Codigo pago', format1)


# 	worksheet.set_column('A:A', 20)
# 	worksheet.set_column('B:B', 18)
# 	worksheet.set_column('C:C', 18)
# 	worksheet.set_column('D:D', 18)
# 	worksheet.set_column('E:E', 24)
# 	worksheet.set_column('F:F', 18)
# 	worksheet.set_column('G:G', 18)
# 	worksheet.set_column('H:H', 18)	
# 	worksheet.set_column('I:I', 18)
# 	worksheet.set_column('J:J', 14)
# 	worksheet.set_column('K:K', 18)
# 	worksheet.set_column('L:L', 18)
# 	worksheet.set_column('M:M', 18)
# 	worksheet.set_column('N:N', 14)
# 	worksheet.set_column('O:O', 14)
# 	worksheet.set_column('P:P', 18)
# 	worksheet.set_column('Q:Q', 18)
# 	worksheet.set_column('R:R', 18)
# 	worksheet.set_column('S:S', 18)


# 	for x in list(result_set):

# 			worksheet.write(row, col,x[0],format3)
# 			worksheet.write(row, col+1,x[1],format3)
# 			worksheet.write(row, col+2,x[2],format3)
# 			worksheet.write(row, col+3,x[3],format3)
# 			worksheet.write(row, col+4,x[4],format3)
# 			worksheet.write(row, col+5,x[5],format3)
# 			worksheet.write(row, col+6,x[6],format3)
# 			worksheet.write(row, col+7,x[7],format3)
# 			worksheet.write(row, col+8,x[8],format3)
# 			worksheet.write(row, col+9,x[9],format3)
# 			worksheet.write(row, col+10,x[10],format3)
# 			worksheet.write(row, col+11,x[11],format3)
# 			worksheet.write(row, col+12,x[12],format3)
# 			worksheet.write(row, col+13,x[13],format3)
# 			# worksheet.write(row, col+14,x[14],format3)
# 			worksheet.write(row, col+14,x[14],format3)
# 			worksheet.write(row, col+15,x[15],format3)
# 			worksheet.write(row, col+16,x[16],format3)
# 			worksheet.write(row, col+17,x[17],format3)
# 			worksheet.write(row, col+18,x[18],format3)


# 			row +=1


# 	workbook.close()

# 	return response
    #return response


def export_excel_encabezado_giro(request):

	try:
	
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Encabezado del giro.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Giros')
		format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
		format3=workbook.add_format({'border':0,'font_size':12})
		format_date=workbook.add_format({'num_format': 'yyyy-mm-dd'})

		row=1
		col=0

		cursor = connection.cursor()

		id_empresa = request.user.usuario.empresa.id
		contrato = request.GET['contrato']
		qset=''
		#import pdb; pdb.set_trace()

		if contrato and int(contrato)>0:

			qset = (Q(encabezado__contrato__mcontrato__id=contrato))

			detalle = DetalleGiro.objects.filter(qset).values(
				'carta_autorizacion__prefijo__nombre','carta_autorizacion__consecutivo',
				'encabezado__contrato__id','encabezado__contrato__mcontrato__nombre', 
				'encabezado__contrato__contratista__nombre', 'encabezado__contrato__numero',
				'encabezado__nombre__nombre','contratista__nombre','contratista__nit',
				'banco__nombre','no_cuenta','tipo_cuenta__nombre','valor_girar','estado__nombre',
				'fecha_pago','id','cuenta__nombre','encabezado__referencia',
				'encabezado__contrato__contratista__codigo_acreedor', 
				'encabezado__pago_recurso__nombre','carta_autorizacion__fechaEnvio',
				'encabezado__contrato__nombre')

		else:
			mcontratosDisponibles = EmpresaContrato.objects.filter(
										contrato__tipo_contrato_id=12,
										empresa__id=id_empresa).values('contrato__id')
			qset = (Q(encabezado__contrato__mcontrato__id__in=mcontratosDisponibles))
			detalle = DetalleGiro.objects.filter(qset).values(
				'carta_autorizacion__prefijo__nombre','carta_autorizacion__consecutivo',
				'encabezado__contrato__id','encabezado__contrato__mcontrato__nombre', 
				'encabezado__contrato__contratista__nombre', 'encabezado__contrato__numero',
				'encabezado__nombre__nombre','contratista__nombre','contratista__nit',
				'banco__nombre','no_cuenta','tipo_cuenta__nombre','valor_girar','estado__nombre',
				'fecha_pago','id','cuenta__nombre','encabezado__referencia',
				'encabezado__contrato__contratista__codigo_acreedor',
				'encabezado__pago_recurso__nombre','carta_autorizacion__fechaEnvio',
				'encabezado__contrato__nombre')

		worksheet.write('A1', 'Contrato macro', format1)
		worksheet.write('B1', 'Interventor', format1)
		worksheet.write('C1', 'Contratista', format1)
		worksheet.write('D1', 'No. contrato', format1)
		worksheet.write('E1', 'Nombre contrato', format1)
		worksheet.write('F1', 'Nombre anticipo', format1)
		worksheet.write('G1', 'Beneficiario / Proveedor', format1)
		worksheet.write('H1', 'Nit', format1)
		worksheet.write('I1', 'Entidad bancaria', format1)
		worksheet.write('J1', 'No. cuenta destino', format1)
		worksheet.write('K1', 'Tipo cuenta destino', format1)
		worksheet.write('L1', 'Valor neto', format1)
		worksheet.write('M1', 'Estado', format1)
		worksheet.write('N1', 'Fecha radicado', format1)
		worksheet.write('O1', 'Carta autorizacion', format1)
		worksheet.write('P1', 'Fecha Carta autorizacion', format1)
		worksheet.write('Q1', 'Fecha pago', format1)
		worksheet.write('R1', 'Origen de recursos', format1)
		worksheet.write('S1', 'Linea', format1)
		worksheet.write('T1', 'Cuenta origen', format1)
		worksheet.write('U1', 'Referencia', format1)
		worksheet.write('V1', 'Codigo Acreedor', format1)



		worksheet.set_column('A:A', 20)
		worksheet.set_column('B:B', 18)
		worksheet.set_column('C:C', 18)
		worksheet.set_column('D:D', 18)
		worksheet.set_column('E:E', 24)
		worksheet.set_column('F:F', 18)
		worksheet.set_column('G:G', 18)
		worksheet.set_column('H:H', 18)	
		worksheet.set_column('I:I', 18)
		worksheet.set_column('J:J', 14)
		worksheet.set_column('K:K', 18)
		worksheet.set_column('L:L', 18)
		worksheet.set_column('M:M', 18)
		worksheet.set_column('N:N', 14)
		worksheet.set_column('O:O', 14)
		worksheet.set_column('P:V', 18)

		nombreContrato=''
		nombre_contratista_interventor=''
		for deta in detalle:

			#print deta.encabezado.contrato.id
			idProyecto= Proyecto.objects.filter(contrato__id=deta['encabezado__contrato__id']).values('id')

			if idProyecto.count()==1:

				contratosDelProyecto= Proyecto.objects.filter(id=idProyecto).values('contrato__id')
				nombre_contratista_interventor=Contrato.objects.filter(id__in=contratosDelProyecto,tipo_contrato=9).values('contratista__nombre')

			else:
				nombre_contratista_interventor='N/A'

			worksheet.write(row, col,deta['encabezado__contrato__mcontrato__nombre'])

			if idProyecto.count()==1:

				if nombre_contratista_interventor:

					worksheet.write(row, col+1,nombre_contratista_interventor[0]['contratista__nombre'])

			else:

				worksheet.write(row, col+1,nombre_contratista_interventor)

			worksheet.write(row, col+2,deta['encabezado__contrato__contratista__nombre'])
			worksheet.write(row, col+3,deta['encabezado__contrato__numero'])
			worksheet.write(row, col+4,deta['encabezado__contrato__nombre'])
			worksheet.write(row, col+5,deta['encabezado__nombre__nombre'])
			worksheet.write(row, col+6,deta['contratista__nombre'])
			worksheet.write(row, col+7,deta['contratista__nit'])
			worksheet.write(row, col+8,deta['banco__nombre'])
			worksheet.write(row, col+9,deta['no_cuenta'])
			worksheet.write(row, col+10,deta['tipo_cuenta__nombre'])
			worksheet.write(row, col+11,deta['valor_girar'])
			worksheet.write(row, col+12,deta['estado__nombre'])
			# worksheet.write(row, col+12,x[12],format3)
			if deta['carta_autorizacion__prefijo__nombre'] and deta['carta_autorizacion__consecutivo']:
				worksheet.write(row, col+14, deta['carta_autorizacion__prefijo__nombre'] + ' - ' + str(deta['carta_autorizacion__consecutivo']))
				worksheet.write(row, col+15,deta['carta_autorizacion__fechaEnvio'],format_date)
			if deta['fecha_pago']:
				worksheet.write(row, col+16,deta['fecha_pago'],format_date)

			worksheet.write(row,col+17,deta['encabezado__pago_recurso__nombre'])

				
			worksheet.write(row, col+18,deta['id'])
			worksheet.write(row, col+19,deta['cuenta__nombre'])
			worksheet.write(row, col+20,deta['encabezado__referencia'])
			worksheet.write(row, col+21,deta['encabezado__contrato__contratista__codigo_acreedor'])

			# worksheet.write(row, col+18,x[18],format3)

			row +=1
		
		workbook.close()

		return response

	except Exception as e:
		# print idProyecto.count()
		# print deta['encabezado__contrato__id']
		# print nombre_contratista_interventor
		# exc_type, exc_obj, exc_tb = sys.exc_info()
		# fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
		# print(exc_type, fname, exc_tb.tb_lineno, e)
		functions.toLog(e,'giros')
		return Response({'message':'Se presentaron errores de comunicacion con el servidor',
			'status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR) 								


#exporta todos los detalles del giro
def export_reporte_giro(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Cuentas_por_pagar.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cuentas_por_pagar')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})

	row=1
	col=0

	cursor = connection.cursor()

	encabezado_id= request.GET['encabezado_id']

	if encabezado_id:
	
		qset = (
				Q(encabezado_id=encabezado_id)
			)

				
		detalle = DetalleGiro.objects.filter(qset)

		worksheet.write('A1', 'Nit', format1)
		worksheet.write('B1', 'Nombre acreedor', format1)
		worksheet.write('C1', 'Entidad bancaria', format1)
		worksheet.write('D1', 'Tipo cuenta', format1)
		worksheet.write('E1', '# Cuenta', format1)
		worksheet.write('F1', 'Importe en mil', format1)
		worksheet.write('G1', 'Numero sap', format1)
		worksheet.write('H1', 'De donde sale', format1)

		worksheet.set_column('A:A', 10)
		worksheet.set_column('B:B', 18)
		worksheet.set_column('C:C', 18)
		worksheet.set_column('D:D', 12)
		worksheet.set_column('E:E', 12)
		worksheet.set_column('F:F', 16)
		worksheet.set_column('G:G', 18)
		worksheet.set_column('H:H', 40)


		for deta in detalle:

			cuenta_nombre=''
			cuenta_referencia=''

			if deta.encabezado.pago_recurso.id==84 or deta.encabezado.pago_recurso.id==85:
				cuenta_nombre='Recursos propios'
			else:

				financier=FinancieroCuenta.objects.filter(contrato_id=deta.encabezado.contrato.mcontrato.id).first()
				cuenta_nombre=financier.nombre


			# if detalle.cuenta is not None:
			# 	cuenta_nombre= "Girar de la cuenta de %s No. %s ' - ' %s" % (detalle.cuenta.nombre , detalle.cuenta.numero, detalle.cuenta.fiduciaria) if detalle.cuenta.nombre is not None else ''
			# else:
			# 	cuenta_nombre='El contrato no tiene una cuenta bancaria asociada'


			if deta.cuenta is not None:
				cuenta_referencia= deta.encabezado.referencia  if deta.encabezado.referencia is not None else ''	


			worksheet.write(row, col,deta.contratista.nit,format2)
			worksheet.write(row, col+1,deta.contratista.nombre,format2)
			worksheet.write(row, col+2,deta.banco.nombre,format2)
			worksheet.write(row, col+3,deta.tipo_cuenta.nombre,format2)
			worksheet.write(row, col+4,deta.no_cuenta,format2)
			worksheet.write(row, col+5,deta.valor_girar,format2)
			worksheet.write(row, col+6,cuenta_referencia,format2)
			worksheet.write(row, col+7,"{}-{}-{}".format(cuenta_nombre, financier.numero, financier.fiduciaria),format2)
			#worksheet.write(row, col+7,detalle.cuenta.nombre if detalle.cuenta is not None else '',format2)

			row +=1


	workbook.close()

	return response
    #return response



#recorre el archivo en excel
def consulta_excel(request):

	try:

		fecha_inicio= request.POST['fecha_inicio']
		fecha_fin= request.POST['fecha_fin']
		cuenta= request.POST['cuenta']
		soporte= request.FILES['archivo']
		qset=''


		doc = openpyxl.load_workbook(soporte)

		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0
	
	
		lista_errores=[]

		contador=hoja.max_row - 1

		if int(contador) > 0:

			validacionTotal=False

			for fila in hoja.rows:

				if i == 0 :
					i=1
	
				else:

					#valida que los campos del archivo no vengan vacios
					if fila[0].value and fila[1].value and fila[2].value:

						fecha_excel =  fila[0].value.strftime("%Y-%m-%d")
						valor =fila[1].value
						no_cuenta =fila[2].value
						validacion=False
						mensaje_modelo_detalle=False


						if no_cuenta and int(no_cuenta)>0:						
							qset = (Q(no_cuenta=no_cuenta))


						if(fecha_inicio and fecha_fin):
							#qset = (Q(fecha_pago_esperada__gte=fecha_inicio) and Q(fecha_pago_esperada__lte=fecha_fin))
							qset = qset &(Q(fecha_pago_esperada__range=(fecha_inicio,fecha_fin)))

						if cuenta and int(cuenta)>0:
							#qset = (Q(cuenta__id=cuenta))
							qset = qset &(Q(cuenta__id=cuenta))


						#consulta al modelo detalle del giro
						detalle = DetalleGiro.objects.filter(qset)

						for item in list(detalle):
							mensaje_modelo_detalle=True


							#compara que los datos del excel conicidan con los de la base de datos y actualiza
							if (str(item.fecha_pago_esperada) == str(fecha_excel) and item.valor_girar == valor):

								validacion=True
								validacionTotal=True

								## print item.id

								item.estado_id=enumEstados.pagado
								item.fecha_pago=fecha_excel
								item.save()

								modelo=FinancieroCuentaMovimiento(valor=item.valor_girar,descripcion='Esta es una prueba',cuenta_id=cuenta,tipo_id=enumTipo.egreso)
								modelo.save()
								mensaje_conflicto=True


							if validacion==False:		

								#guarda la lista de las fecha que no coicidan
								if (str(item.fecha_pago_esperada) != str(fecha_excel)):

									lista_errores.append(

										{		
											'archivo_recibido':fecha_excel,
											'registro_sistema':item.fecha_pago_esperada,
											'descripcion':'Conflito en la fecha de pago esperada',
											'validacion':1
										}

									)

									mensaje_conflicto=False


								#guarda la lista de los valores que no coicidan
								if (item.valor_girar != valor):
										
									lista_errores.append(

										{		
											'archivo_recibido':valor,
											'registro_sistema':item.valor_girar,
											'descripcion':'Conflito en el valor a girar',
											'validacion':2
										}

									)

									mensaje_conflicto=False



						if mensaje_modelo_detalle==False:

							return JsonResponse({'message':'No se encontraron registros con los parametros ingresados','success':'ok','data':lista_errores})

					else:
						return JsonResponse({'message':'Datos incompletos en el archivo','success':'ok','data':''})

			if validacionTotal== False:

				return JsonResponse({'message':'La informacion del archivo recibido no coiciden con los del registro del sistema','success':'ok','data':lista_errores})


			if mensaje_conflicto==False:

				return JsonResponse({'message':'El registro se actualizo exitosamente y se encontraron conflictos en el archivo de excel','success':'ok','data':lista_errores})

			else:

				return JsonResponse({'message':'El registro se actualizo exitosamente','success':'ok','data':''})

		
		return JsonResponse({'message':'No se encontraron registros en el archivo','success':'ok','data':''})

	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response


def consultar_procesos(request):
	try:
		#import pdb; pdb.set_trace()
		idapuntador = int(request.GET['idApuntador'] if request.GET['idApuntador'] else 0)

		qsetc = (Q(contrato__tipo_contrato=12))& (Q(edita=1)) &(Q(empresa=4) & Q(participa=1))
		ListMacro = EmpresaContrato.objects.filter(qsetc).values('contrato_id').distinct().order_by("contrato_id")

		qset = (Q(mcontrato__id__in=ListMacro))
		qset = qset & (Q(id=idapuntador))

		model_contrato = Contrato.objects.filter(qset).exists()
		idproceso = 2
		if model_contrato == True:
			idproceso = 19

		return JsonResponse({'message':'','success':'ok','data':idproceso})

	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# guarda en la tabla proceso relacion  y en la tabla proceso relacion datos con los parametros enviados desde giros
def guardar_proceso(request):
	#import pdb; pdb.set_trace()
	try:

		lista=request.POST['_content']
		respuesta = json.loads(lista)

		if respuesta['idProcesoRelacion']==0:

			# print respuesta['proceso']
			# print respuesta['idApuntador']
			# print respuesta['idTablaReferencia']


			modelo=FProcesoRelacion(proceso_id=respuesta['proceso'],idApuntador=respuesta['idApuntador'],idTablaReferencia=respuesta['idTablaReferencia'])
			modelo.save()

			listado_item = BItem.objects.filter(proceso=respuesta['proceso'])

			for item in listado_item:

				modelo2=GProcesoRelacionDato(procesoRelacion_id=modelo.id,item_id=item.id,valor='Vacio')
				modelo2.save()
		
			return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':modelo.id})

		
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)




# Actualiza el estado del detalle del giro solo si el consecutivo de la empresa esta desabilitado
@transaction.atomic
def Actualizar_detalle_giro_segun_consecutivo(request):

	sid = transaction.savepoint()

	try:

		soporte= request.FILES['archivo']

		lista=request.POST['lista']
		listado=lista.split(',')

		for item in listado:

			## print item
			
			object_detalle=DetalleGiro.objects.get(pk=item)

			object_detalle.estado_id=enumEstados.autorizado
			object_detalle.soporte_consecutivo_desabilitado=soporte
			object_detalle.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item)
			logs_model.save()

			transaction.savepoint_commit(sid)

		return JsonResponse({'message':'Los registros se han actualizado correctamente','success':'ok',
							'data':''})
				
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)



#genera el reporte de solicitud de los anticipos
def generar_reporte_solicitud_anticipo(request):

	try:

		empresa_id = request.user.usuario.empresa.id
		#newpath = r'static/papelera/'
		filename = "plantillas/giros/solicitud_anticipos.xlsx"
		extension = filename[filename.rfind('.'):]
		nombre = 'empresa'+str(empresa_id)+'-'+str(request.user.usuario.id)+extension

		ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'

		functions.descargarArchivoS3(str(filename), ruta, nombre)
		

		wb = load_workbook(filename = '{}/{}'.format(ruta, nombre), read_only=False)
		ws = wb.worksheets[0]

		cursor = connection.cursor()
		id_anticipo = request.GET['encabezado_id']
		# mcontrato_id = request.GET['mcontrato_id']
		# contrato_obra = request.GET['contra']
		id_empresa = request.user.usuario.empresa.id
		qset=''
		nombre_giros=''
		nombre_proyecto=''

		row=35
		col=0

		#Estilo para los bordes del archivo
		thin_border = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			top=Side(style='thin',color='BDBDBD'),
			bottom=Side(style='thin',color='BDBDBD'))

		thin_border2 = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			bottom=Side(style='thin',color='BDBDBD'))

		thin_border3 = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			top=Side(style='thin',color='BDBDBD'))


		#agrega las imagenes al archivo de excel
		img = Image(settings.BASE_DIR+'/static/images/electricaribe.jpg')
		ws.add_image(img, 'W3')

		img2 = Image(settings.BASE_DIR+'/static/images/moneda2.jpg')
		ws.add_image(img2, 'C28')


		#borde derecho y izquierdo general de la plantilla
		for y in xrange(2,83):
			ws.cell(row=y, column=2).border = thin_border.left
			ws.cell(row=y, column=28).border = thin_border.right

		#borde top general de la plantilla
		for x in xrange(2,28):
			ws.cell(row=2, column=x).border = thin_border3

		#borde bottom general de la plantillan
		for x in xrange(2,28):
			ws.cell(row=82, column=x).border = thin_border2



		#borde para el nombre de la contratista
		for x in xrange(4,27):
			ws.cell(row=17, column=x).border = thin_border



		#borde del nit de la empresa
		for x in xrange(4,6):
			ws.cell(row=19, column=x).border = thin_border

		#borde del cuadro del lado derecho de nit
		ws.cell(row=19, column=7).border = thin_border

		#borde del cod Acreedor
		for x in xrange(12,14):
			ws.cell(row=19, column=x).border = thin_border

		#borde contacto de proveedor
		for x in xrange(19,23):
			ws.cell(row=19, column=x).border = thin_border

		#borde del tel
		for x in xrange(24,27):
			ws.cell(row=19, column=x).border = thin_border



		#borde derecho y izquierdo para la justificacion
		for y in xrange(21,26):
			ws.cell(row=y, column=4).border = thin_border.left
			ws.cell(row=y, column=27).border = thin_border.right

		#borde top para la justifiacion
		for x in xrange(4,27):
			ws.cell(row=21, column=x).border = thin_border3

		#borde bottom para la justifacion
		for x in xrange(4,27):
			ws.cell(row=25, column=x).border = thin_border2



		#borde de detalle de anticipo
		for x in xrange(9,14):
			ws.cell(row=28, column=3).border = thin_border

		#borde de detalle de anticipo
		for x in xrange(4,6):
			ws.cell(row=28, column=x).border = thin_border

		#borde de detalle de anticipo
		for x in xrange(6,9):
			ws.cell(row=28, column=x).border = thin_border

		#borde de pesos colombianos
		for x in xrange(9,14):
			ws.cell(row=28, column=x).border = thin_border

		#borde de us dolares
		for x in xrange(14,19):
			ws.cell(row=28, column=x).border = thin_border

		#borde de euros
		for x in xrange(19,22):
			ws.cell(row=28, column=x).border = thin_border

		#borde tasa de cambio
		for x in xrange(22,24):
			ws.cell(row=28, column=x).border = thin_border

		#borde total en pesos
		for x in xrange(24,27):
			ws.cell(row=28, column=x).border = thin_border



		#borde top para la descripcion de la forma de armortizacion del anticipo
		for x in xrange(3,27):
			ws.cell(row=32, column=x).border = thin_border3

		#borde derecho y izquierdo  para la descripcion de la forma de armortizacion del anticipo
		for y in xrange(33,56):
			ws.cell(row=y, column=3).border = thin_border.left
			ws.cell(row=y, column=27).border = thin_border.right


		#borde bottom para la descripcion de la forma de armortizacion del anticipo
		for x in xrange(3,27):
			ws.cell(row=55, column=x).border = thin_border2



		#borde medio de pago
		for x in xrange(6,11):
			ws.cell(row=60, column=x).border = thin_border

		#borde de tipo de cuenta
		for x in xrange(21,23):
			ws.cell(row=60, column=x).border = thin_border



		#borde de entidad bancaria
		for x in xrange(6,16):
			ws.cell(row=62, column=x).border = thin_border

		#borde no de cuenta
		for x in xrange(21,27):
			ws.cell(row=62, column=x).border = thin_border



		#borde nonbre solicitante del anticipo
		for x in xrange(6,16):
			ws.cell(row=64, column=x).border = thin_border

		#borde ext
		ws.cell(row=64, column=20).border = thin_border

		#borde area
		for x in xrange(23,27):
			ws.cell(row=64, column=x).border = thin_border



		#borde voBo
		for x in xrange(6,16):
			ws.cell(row=70, column=x).border = thin_border

		#borde autoriza
		for x in xrange(20,27):
			ws.cell(row=70, column=x).border = thin_border



		#borde firma del lado izquierdo
		for y in xrange(72,76):

			for x in xrange(6,16):
				ws.cell(row=y, column=x).border = thin_border


		#borde firma del lado derecho
		for y in xrange(72,76):

			for x in xrange(20,27):
				ws.cell(row=y, column=x).border = thin_border



		#contra=Contrato.objects.filter(id=mcontrato_id).first()
		giros = DEncabezadoGiro.objects.filter(id=id_anticipo)
		detalles = DetalleGiro.objects.filter(encabezado__id=id_anticipo)
		#cuenta=FinancieroCuenta.objects.filter(contrato_id=mcontrato_id,estado_id=87).first()
		#proyectos = Proyecto.objects.filter(mcontrato_id=mcontrato_id)
		#proyectos = Proyecto_empresas.objects.filter(proyecto__mcontrato__id=mcontrato_id,empresa_id=id_empresa)
		con=''
		validacion=0
		entidad_financiera=''
		no_cuenta=''
		tipo_cuenta=''
		cuenta_financiero=''
		numero_financiero=''
		financiero=''



		#concateno el nombre de los giros
		for giro in giros:

			# giro_encabezado = DEncabezadoGiro.objects.filter(contrato_id=giro.contrato.id)

			# for nombregiro in giro_encabezado:

			nombre_giros= nombre_giros+giro.nombre.nombre+' ,'


			cuenta=FinancieroCuenta.objects.filter(contrato_id=giro.contrato.mcontrato.id,estado_id=87).first()

			#Valido si del modelo de financiero cuenta llegan registros
			if cuenta:

				cuenta_financiero=cuenta.numero
				numero_financiero=cuenta.fiduciaria
				financiero="{} {} - {}".format(" No. ",cuenta_financiero,numero_financiero)

			else:
				financiero='No se encontro cuenta asociada al contrato'


			proyectos = Proyecto_empresas.objects.filter(proyecto__mcontrato__id=giro.contrato.mcontrato.id,empresa_id=id_empresa)

			#concateno el nombre de los proyectos y realizo las validaciones para los campos entidad financiera, tipo de cuenta y no de cuenta
			for proy in proyectos:

				nombre_proyecto= nombre_proyecto+proy.proyecto.nombre+' - '

				if proy.proyecto.No_cuenta and proy.proyecto.entidad_bancaria_id and proy.proyecto.tipo_cuenta_id:

					validacion=validacion+1


				if validacion ==1:

					if proy.proyecto.entidad_bancaria:

						entidad_financiera=proy.proyecto.entidad_bancaria.nombre

					else:

						entidad_financiera='No se encontro entidad bancaria asociada al proyecto'


					if proy.proyecto.No_cuenta:

						no_cuenta=proy.proyecto.No_cuenta

					else:

						no_cuenta='No se encontro cuenta asociada al proyecto'


					if proy.proyecto.tipo_cuenta:

						tipo_cuenta=proy.proyecto.tipo_cuenta.nombre

					else:

						tipo_cuenta='No se encontro tipo de cuenta asociada al proyecto'

				else:

					entidad_financiera='Se encontro mas de una entidad financiera asociada en los proyectos'
					no_cuenta='Se encontro mas de una cuenta asociada en los proyectos'
					tipo_cuenta='Se encontro mas de un tipo de cuenta asociada en los proyectos'



			ws['D17'] = giro.contrato.contratista.nombre
			ws['D19'] = giro.contrato.contratista.nit
			ws['D21'] = "{} - {} : {}".format(giro.contrato.mcontrato.nombre,nombre_giros.strip(','),nombre_proyecto.strip(' - ').encode('utf-8'))
			ws['D22'] = "RECURSOS GOBIERNO NACIONAL"
			ws['C32'] = "{} - {} {} ".format("Girar de la cuenta de Ahorros ",giro.contrato.mcontrato.nombre,financiero)

		ws['C34'] = 'PROVEEDOR'
		ws['C34'].font = Font(underline="single",bold=True)
		ws['K34'] = 'NIT'
		ws['K34'].font = Font(underline="single",bold=True)
		ws['N34'] = 'No. DE CUENTA'
		ws['N34'].font = Font(underline="single",bold=True)
		ws['S34'] = 'TIPO DE CUENTA - ENTIDAD FINANCIERA'
		ws['S34'].font = Font(underline="single",bold=True)
		ws['X34'] = 'VALOR A GIRAR'
		ws['X34'].font = Font(underline="single",bold=True)

		# alignment=Alignment(
		# 	horizontal='left',
		# 	vertical='top',
		# 	text_rotation=90,
		# 	wrap_text=False,
		# 	shrink_to_fit=False,
		# 	indent=0)

		ws['A7'] = 'solicitud anticipo  a proveedores'
		ws['A7'].alignment = Alignment(text_rotation=90,vertical='top',horizontal='left')
		#ws['A7'].alignment = alignment
		ws['A7'].font = Font(name='Calibri',size=72,color='BDBDBD')
		ws['F62'] = entidad_financiera
		ws['U62'] = no_cuenta
		ws['U60'] = tipo_cuenta


		for detalle in detalles:

			ws['C'+str(row)]=detalle.contratista.nombre
			ws['K'+str(row)]=detalle.contratista.nit
			ws['N'+str(row)]=detalle.no_cuenta
			ws['S'+str(row)]="{}/{}".format(detalle.tipo_cuenta.nombre, detalle.banco.nombre)
			ws['X'+str(row)]=detalle.valor_girar

			row+=1

		wb.template = False
		response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
		response['Content-Disposition'] ='attachment; filename="solicitud_anticipos.xlsx"'
		return response
	except Exception as e:
		#print(e)
		functions.toLog(e,'giros.generar_reporte_solicitud_anticipo')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
		


#genera el reporte de solicitud de los anticipos
def generar_reporte_solicitud_anticipo2(request):

	try:

		empresa_id = request.user.usuario.empresa.id
		#newpath = r'static/papelera/'
		filename = "plantillas/giros/anticiposProveedores.xlsx"
		extension = filename[filename.rfind('.'):]
		nombre = 'empresa'+str(empresa_id)+'-'+str(request.user.usuario.id)+extension

		ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'

		functions.descargarArchivoS3(str(filename), ruta, nombre)
		

		wb = load_workbook(filename = '{}/{}'.format(ruta, nombre), read_only=False)
		ws = wb.worksheets[0]

		cursor = connection.cursor()
		id_anticipo = request.GET['encabezado_id']
		# mcontrato_id = request.GET['mcontrato_id']
		# contrato_obra = request.GET['contra']
		id_empresa = request.user.usuario.empresa.id
		qset=''
		nombre_giros=''
		nombre_proyecto=''

		row=35
		col=0

		#Estilo para los bordes del archivo
		thin_border = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			top=Side(style='thin',color='BDBDBD'),
			bottom=Side(style='thin',color='BDBDBD'))

		thin_border2 = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			bottom=Side(style='thin',color='BDBDBD'))

		thin_border3 = Border(
			left=Side(style='thin',color='BDBDBD'),
			right=Side(style='thin',color='BDBDBD'),
			top=Side(style='thin',color='BDBDBD'))

		thin_border4 = Border(
			top=Side(style='thin',color='BDBDBD'))


		#agrega las imagenes al archivo de excel
		img = Image(settings.BASE_DIR+'/static/images/imagenElictricaribe.jpg')
		ws.add_image(img, 'S3')

		img2 = Image(settings.BASE_DIR+'/static/images/moneda2.jpg')
		ws.add_image(img2, 'C28')


		#borde derecho y izquierdo general de la plantilla
		for y in xrange(2,74):
			ws.cell(row=y, column=2).border = thin_border.left
			ws.cell(row=y, column=26).border = thin_border.right

		#borde top general de la plantilla
		for x in xrange(2,26):
			ws.cell(row=2, column=x).border = thin_border3

		#borde bottom general de la plantillan
		for x in xrange(2,26):
			ws.cell(row=74, column=x).border = thin_border2



		#borde para el nombre de la contratista
		for x in xrange(4,20):
			ws.cell(row=17, column=x).border = thin_border



		#borde del nit de la empresa
		for x in xrange(4,6):
			ws.cell(row=19, column=x).border = thin_border

		#borde del cuadro del lado derecho de nit
		ws.cell(row=19, column=7).border = thin_border

		#borde del cod Acreedor
		for x in xrange(23,25):
			ws.cell(row=17, column=x).border = thin_border

		#borde contacto de proveedor
		for x in xrange(14,21):
			ws.cell(row=19, column=x).border = thin_border

		#borde del tel
		for x in xrange(22,25):
			ws.cell(row=19, column=x).border = thin_border



		#borde derecho y izquierdo para la justificacion
		for y in xrange(21,25):
			ws.cell(row=y, column=4).border = thin_border.left
			ws.cell(row=y, column=25).border = thin_border.right

		#borde top para la justifiacion
		for x in xrange(4,25):
			ws.cell(row=21, column=x).border = thin_border3

		#borde bottom para la justifacion
		# for x in xrange(4,25):
		# 	ws.cell(row=25, column=x).border = thin_border2



		# #borde de detalle de anticipo
		# for x in xrange(2,4):
		# 	ws.cell(row=28, column=3).border = thin_border

		# #borde de detalle de anticipo
		# for x in xrange(4,6):
		# 	ws.cell(row=28, column=x).border = thin_border

		# #borde de detalle de anticipo
		# for x in xrange(6,9):
		# 	ws.cell(row=28, column=x).border = thin_border

		# #borde de pesos colombianos
		for x in xrange(9,14):
			ws.cell(row=28, column=x).border = thin_border

		# #borde de us dolares
		for x in xrange(14,19):
			ws.cell(row=28, column=x).border = thin_border

		#borde de euros
		for x in xrange(19,22):
			ws.cell(row=28, column=x).border = thin_border

		#borde tasa de cambio
		for x in xrange(22,24):
			ws.cell(row=28, column=x).border = thin_border

		#borde total en pesos
		for x in xrange(24,25):
			ws.cell(row=28, column=x).border = thin_border



		# #borde top para la descripcion de la forma de armortizacion del anticipo
		for x in xrange(3,25):
			ws.cell(row=32, column=x).border = thin_border3

		# #borde derecho y izquierdo  para la descripcion de la forma de armortizacion del anticipo
		for y in xrange(33,49):
			ws.cell(row=y, column=3).border = thin_border.left
			ws.cell(row=y, column=25).border = thin_border.right

	
		#borde bottom para la descripcion de la forma de armortizacion del anticipo
		for x in xrange(3,25):

			ws.cell(row=49, column=x).border = thin_border4



		#borde medio de pago
		for x in xrange(6,16):
			ws.cell(row=54, column=x).border = thin_border

		#borde de tipo de cuenta
		for x in xrange(19,21):
			ws.cell(row=54, column=x).border = thin_border



		#borde de entidad bancaria
		for x in xrange(6,16):
			ws.cell(row=56, column=x).border = thin_border

		#borde no de cuenta
		for x in xrange(19,25):
			ws.cell(row=56, column=x).border = thin_border



		#borde nonbre solicitante del anticipo
		for x in xrange(6,16):
			ws.cell(row=58, column=x).border = thin_border

		#borde ext
		# ws.cell(row=52, column=20).border = thin_border

		#borde area
		for x in xrange(22,25):
			ws.cell(row=58, column=x).border = thin_border



		#borde voBo
		for x in xrange(6,16):
			ws.cell(row=64, column=x).border = thin_border

		#borde autoriza
		for x in xrange(18,25):
			ws.cell(row=64, column=x).border = thin_border



		# #borde firma del lado izquierdo
		for y in xrange(66,69):

			for x in xrange(6,16):
				ws.cell(row=y, column=x).border = thin_border
				ws.cell(row=69, column=x).border = thin_border2


		#borde firma del lado derecho
		for y in xrange(66,69):

			for x in xrange(18,25):
				ws.cell(row=y, column=x).border = thin_border
				ws.cell(row=69, column=x).border = thin_border2

		#contra=Contrato.objects.filter(id=mcontrato_id).first()
		giros = DEncabezadoGiro.objects.filter(id=id_anticipo)
		detalles = DetalleGiro.objects.filter(encabezado__id=id_anticipo)
		#cuenta=FinancieroCuenta.objects.filter(contrato_id=mcontrato_id,estado_id=87).first()
		#proyectos = Proyecto.objects.filter(mcontrato_id=mcontrato_id)
		#proyectos = Proyecto_empresas.objects.filter(proyecto__mcontrato__id=mcontrato_id,empresa_id=id_empresa)
		con=''
		validacion=0
		entidad_financiera=''
		no_cuenta=''
		tipo_cuenta=''
		cuenta_financiero=''
		numero_financiero=''
		financiero=''



		#concateno el nombre de los giros
		for giro in giros:

			nombre_giros= nombre_giros+giro.nombre.nombre+' ,'

			cuenta=FinancieroCuenta.objects.filter(contrato_id=giro.contrato.mcontrato.id,estado_id=87).first()

			#Valido si del modelo de financiero cuenta llegan registros
			if cuenta:

				cuenta_financiero=cuenta.numero
				numero_financiero=cuenta.fiduciaria
				financiero="{} {} - {}".format(" No. ",cuenta_financiero,numero_financiero)

			else:
				financiero='No se encontro cuenta asociada al contrato'


			proyecto_l= Proyecto.objects.filter(contrato__id=giro.contrato.id).values('id','nombre','entidad_bancaria__nombre','No_cuenta','tipo_cuenta__nombre').first()


			if proyecto_l['entidad_bancaria__nombre']:

				entidad_financiera=proyecto_l['entidad_bancaria__nombre']

			else:

				entidad_financiera='No se encontro entidad bancaria asociada al proyecto'


			if proyecto_l['No_cuenta']:

				no_cuenta=proyecto_l['No_cuenta']

			else:

				no_cuenta='No se encontro cuenta asociada al proyecto'



			if proyecto_l['tipo_cuenta__nombre']:

				tipo_cuenta=proyecto_l['tipo_cuenta__nombre']

			else:

				tipo_cuenta='No se encontro tipo de cuenta asociada al proyecto'



			ws['D17'] = giro.contrato.contratista.nombre
			ws['D19'] = giro.contrato.contratista.nit
			ws['D21'] = "{} - {} : {}".format(giro.contrato.mcontrato.nombre,nombre_giros.strip(','),proyecto_l['nombre'].strip(' - ').encode('utf-8'))
			ws['D22'] = "RECURSOS GOBIERNO NACIONAL"
			ws['C32'] = "{} - {} {} ".format("Girar de la cuenta de Ahorros ",giro.contrato.mcontrato.nombre,financiero)

		
		ws['C34'] = 'PROVEEDOR'
		ws['C34'].font = Font(underline="single",bold=True)
		ws['K34'] = 'NIT'
		ws['K34'].font = Font(underline="single",bold=True)
		ws['N34'] = 'No. DE CUENTA'
		ws['N34'].font = Font(underline="single",bold=True)
		ws['S34'] = 'TIPO DE CUENTA - ENTIDAD FINANCIERA'
		ws['S34'].font = Font(underline="single",bold=True)
		ws['X34'] = 'VALOR A GIRAR'
		ws['X34'].font = Font(underline="single",bold=True)


		ws['A7'] = 'solicitud anticipo  a proveedores'
		ws['A7'].alignment = Alignment(text_rotation=90,vertical='top',horizontal='left')
		#ws['A7'].alignment = alignment
		ws['A7'].font = Font(name='Calibri',size=72,color='BDBDBD')
		ws['F56'] = entidad_financiera
		ws['S56'] = no_cuenta
		ws['S54'] = tipo_cuenta


		for detalle in detalles:

			ws['C'+str(row)]=detalle.contratista.nombre
			ws['K'+str(row)]=detalle.contratista.nit
			ws['M'+str(row)]=detalle.no_cuenta
			ws['Q'+str(row)]="{}/{}".format(detalle.tipo_cuenta.nombre, detalle.banco.nombre)
			ws['U'+str(row)]=detalle.valor_girar

			row+=1


		wb.template = False
		response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
		response['Content-Disposition'] ='attachment; filename="solicitud_anticipos.xlsx"'
		return response
	except Exception as e:
		#print(e)
		#functions.toLog(e,'giros.generar_reporte_solicitud_anticipo')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
		


def cargar_soporte_proceso(request):

	try:

		if request.POST['procesoRelacionDato_id']:

			modelo=HSoporteProcesoRelacionDato(procesoRelacionDato_id=request.POST['procesoRelacionDato_id'],nombre=request.POST['nombre_documento'],documento=request.FILES['archivo'])
			modelo.save()
		
			return JsonResponse({'message':'Se registro exitosamente','success':'ok','data':''})

		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','success':'ok','data':''})
		
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# actualizar tipo de pago del giro
#Api rest para update_tipo_pago_del_giro
@transaction.atomic
def updateTipoPagoGiro(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			pago_recurso_id = request.POST['pago_recurso_id']
			myList = request.POST['id'].split(',')
			insert_list = []
			for item in myList:

				DEncabezadoGiro.objects.filter(id = item).update(pago_recurso_id = pago_recurso_id)

				insert_list.append(Logs(usuario_id=request.user.usuario.id
										,accion=Acciones.accion_actualizar
										,nombre_modelo='giros.DEncabezadoGiro'
										,id_manipulado=item)
										)

			Logs.objects.bulk_create(insert_list)
			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok','data': ''})
		except Exception as e:
			transaction.savepoint_rollback(sid)	
			print(e)
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# actualizar disparar flujo
#Api rest para update_disparar_flujo
@transaction.atomic
def updateDispararFlujo(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			disparar_flujo = request.POST['disparar_flujo']
			encabezado_id = request.POST['id']

			if int(disparar_flujo)==0:
				disparar_flujo = False 

			DEncabezadoGiro.objects.filter(id = encabezado_id).update(disparar_flujo = disparar_flujo)

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.DEncabezadoGiro',id_manipulado = encabezado_id)
			logs_model.save()

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok','data': ''})
		except Exception as e:
			transaction.savepoint_rollback(sid)	
			print(e)
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def validar_estado_cuenta(request):
	try:


		encabezado_id = request.GET["encabezado_id"]

		encabezado_giro = DEncabezadoGiro.objects.get(pk = encabezado_id)

		if encabezado_giro.pago_recurso_id == enumTipoPagoAnticipo.cuenta_bancaria:
			mcontrato_id = encabezado_giro.nombre.contrato_id
			cuenta = FinancieroCuenta.objects.filter(contrato_id = mcontrato_id)[:1].get()

			valor_ingreso= FinancieroCuentaMovimiento.objects.filter(cuenta_id= cuenta.id, tipo_id=31).aggregate(suma_ingreso=Sum('valor'))	
			valor_egreso= FinancieroCuentaMovimiento.objects.filter(cuenta_id= cuenta.id, tipo_id=29).aggregate(suma_egreso=Sum('valor'))
			valor_rendimiento= FinancieroCuentaMovimiento.objects.filter(cuenta_id= cuenta.id, tipo_id=32).aggregate(suma_rendimiento=Sum('valor'))	

			total_detalle = DetalleGiro.objects.filter(encabezado_id = encabezado_id).aggregate(suma_detalle=Sum('valor_girar'))	

			total = float(valor_ingreso["suma_ingreso"]) + float(valor_rendimiento["suma_rendimiento"]) - float(valor_egreso["suma_egreso"])
			
			total = float(total) - float(total_detalle["suma_detalle"])

			cuenta_nombre = cuenta.nombre
		else:
			total = 1
			cuenta_nombre = ' '

		return JsonResponse({'message':' ','success':'ok','data': { 'total': total, 'cuenta': cuenta_nombre } })


	except Exception as e:
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def VerSoporte(request):
	if request.method == 'GET':
		try:
						
			archivo = DEncabezadoGiro.objects.get(pk=request.GET['id'])
			
			return functions.exportarArchivoS3(str(archivo.soporte))
			
		except Exception as e:
			functions.toLog(e,'Giros.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def validacion_giro(request):
	if request.method == 'GET':
		try:
			contrato_id = int(request.GET['contrato_id'])
			mcontrato_id = int(request.GET['mcontrato_id'])
			contrato_model = Contrato.objects.get(pk=contrato_id)
			proyectos = Proyecto.objects.filter(contrato__id=contrato_id,mcontrato__id=mcontrato_id).exists()

			if proyectos:				

				proyectos = Proyecto.objects.filter(contrato__id=contrato_id,mcontrato__id=mcontrato_id)
				validacion_no_cuenta=True
				string_return_proyectos = '('
				num=1

				#import pdb; pdb.set_trace()
				for proyecto in proyectos:
					if not proyecto.No_cuenta or proyecto.No_cuenta=='Null':
						if len(proyecto.nombre)>29:
							string_return_proyectos=string_return_proyectos+proyecto.municipio.nombre+'-'+proyecto.mcontrato.nombre
						else:
							string_return_proyectos=string_return_proyectos+proyecto.nombre+'-'+proyecto.mcontrato.nombre
						validacion_no_cuenta=False
						if proyectos.count()==num:
							string_return_proyectos=string_return_proyectos+')'
						else:
							string_return_proyectos=string_return_proyectos+', '
					num+=1

				if validacion_no_cuenta:
					return JsonResponse({'message':'','success':'ok','data':[]})
				else:
					return JsonResponse({'message':'No se puede registrar un giro con este contrato, los siguientes proyectos asociados a este contrato, no cuentan con un No. cuenta '+string_return_proyectos,'success':'fail','data':[]})
			else:				
				return JsonResponse({'message':'No se puede registrar un giro con este contrato debido a que no se encuentra asociado a ningun proyecto con el MME','success':'fail','data':[]})

		except Exception as e:
			functions.toLog(e,'Giros.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
