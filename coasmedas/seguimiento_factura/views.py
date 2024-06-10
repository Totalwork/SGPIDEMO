from django.shortcuts import render,redirect
#,render_to_response
from django.urls import reverse
from rest_framework import viewsets, serializers
from django.db.models import Q
from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.views import exception_handler
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status 
from rest_framework.pagination import PageNumberPagination
import xlsxwriter
import uuid
import json
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from tipo.models import Tipo
from empresa.views import EmpresaSerializer
from contrato.views import ContratoSerializer
from contrato.models import Contrato
from contrato.enumeration import tipoC
from empresa.models import Empresa,EmpresaCuenta
from factura.models import Factura
from financiero.models import FinancieroCuenta
from .models import GestionOp
from logs.models import Logs,Acciones
from django.db import connection
from django.db import transaction
from django.db.models.deletion import ProtectedError
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view
from decimal import *
from django.db.models import F, FloatField, Sum
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from coasmedas.functions import functions
from django.conf import settings
from datetime import datetime, date, time, timedelta
from factura.enumeration import estadoFactura, notificacion
import collections



# from factura.views import FacturaSerializer

#Api rest para gestion op
class GestionOpSerializer(serializers.HyperlinkedModelSerializer):

	beneficiario = serializers.SerializerMethodField()
	contrato = serializers.SerializerMethodField()
	valor = serializers.SerializerMethodField()
	financiero_cuenta = serializers.SerializerMethodField()

	class Meta:
		model = GestionOp
		fields=('id','contrato', 'beneficiario', 'valor',
			'codigo','fecha_registro','fecha_pago','soporte','financiero_cuenta','pagado_recursos_propios','soporte_pago')


	def get_beneficiario(self, obj):
		try:
			factura = Factura.objects.filter(codigo_op_id=obj.id)[:1].get()
			nombre = factura.contrato.contratista.nombre
		except Factura.DoesNotExist:
			nombre = ''

		return nombre

	def get_contrato(self, obj):

		try:
			factura = Factura.objects.filter(codigo_op_id=obj.id)[:1].get()
			nombre = factura.contrato.nombre

		except Exception as e:
			nombre = ''
		
		return nombre

	def get_valor(self, obj):

		try:
			sumatoria = Factura.objects.filter(codigo_op_id=obj.id).aggregate(total=Sum('valor_contable'))
			total = sumatoria['total']
		except Exception as e:
			total = 0
		
		return total

	def get_financiero_cuenta(self, obj):

		try:
			factura = Factura.objects.filter(codigo_op_id=obj.id)[:1].get()
			cuenta = FinancieroCuenta.objects.filter(contrato_id = factura.contrato.mcontrato.id).values('id','nombre').first()
		except Exception as e:
			cuenta = 0
		
		return cuenta

class GestionOpViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de gestion op.
	"""
	model=GestionOp
	queryset = model.objects.all()
	serializer_class = GestionOpSerializer
	nombre_modulo='seguimiento_factura.gestionOp'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(GestionOpViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			id_pago= self.request.query_params.get('id_pago',None)
			id_empresa = request.user.usuario.empresa.id
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			contrato= self.request.query_params.get('contrato',None)
			contratista= self.request.query_params.get('contratista',None)
			cuenta= self.request.query_params.get('cuenta',None)
			desde= self.request.query_params.get('desde',None)
			hasta= self.request.query_params.get('hasta',None)
			recursos= self.request.query_params.get('recursos', None)
			
			recursos = True if recursos=='1' else False	
			qset = (~Q(id=0))
			qset = qset &( Q(anulado = False) )

			if (dato or id_pago or contrato or contratista or id_empresa or cuenta or desde or hasta or recursos):

				if dato:
					qset = qset &( Q(codigo__icontains=dato) )

				if id_pago and int(id_pago)>0:
					qset = qset &( Q(id=id_pago) )

				if contrato and int(contrato)>0:
					qset = qset &(Q(contrato__id=contrato)	)

				if contratista and int(contratista)>0:
					qset = qset &(Q(beneficiario__id=contratista))

				if cuenta and int(cuenta)>0:
					qset = qset &( Q(contrato__id=cuenta))

				if desde:
					qset = qset &( Q(fecha_registro__gte=desde))
				if hasta and desde:
					qset = qset &( Q(fecha_registro__gte=desde) and Q(fecha_registro__lte=hasta))
			
				if recursos != None:
					qset = qset &( Q(pagado_recursos_propios=recursos)	)			
				queryset = self.model.objects.filter(qset)

			if sin_paginacion is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

		except Exception as e:
			print(e)
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':

			sid = transaction.savepoint()

			try:

				serializer = GestionOpSerializer(data=request.DATA,context={'request': request})

				try:
					
					gestion_op = GestionOp.objects.get(codigo = self.request.DATA['codigo'])
					return Response({'message':'El codigo TEST-OP digitado ya existe','success':'fail',
			 			'data':''},status=status.HTTP_400_BAD_REQUEST)
				except Exception as e:
					pass
				


				##print request.DATA
 				#print serializer
				if serializer.is_valid():

					listado_factura=json.loads(self.request.DATA['lista_factura'])

					serializer.save(soporte=self.request.FILES.get('soporte') if self.request.FILES.get('soporte') is not None else '',
						soporte_pago=self.request.FILES.get('soporte_pago') if self.request.FILES.get('soporte_pago') is not None else '')

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='seguimiento_factura.gestion_op',id_manipulado=serializer.data['id'])
					logs_model.save()
					
					
					for item in listado_factura:
						
						## print item['id_factura']
						object_factura = Factura.objects.get(pk= item["id_factura"])

						object_factura.codigo_op_id = serializer.data['id']
						object_factura.estado_id = estadoFactura.compensada
						object_factura.save()

						logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='seguimiento_factura.factura',id_manipulado=item['id_factura'])
						logs_model.save()						
					
					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok',
					'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 			'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
			  		'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':

			sid = transaction.savepoint()
			# print "mendoza"
			try:
				partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = GestionOpSerializer(instance,data=request.DATA,context={'request': request},partial=partial)

				##print request.DATA
				if serializer.is_valid():
					detalle=self.model.objects.get(pk=instance.id)

					serializer.save(soporte=self.request.FILES.get('soporte') if self.request.FILES.get('soporte') is not None else instance.soporte
							,soporte_pago=self.request.FILES.get('soporte_pago') if self.request.FILES.get('soporte_pago') is not None else instance.soporte_pago)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='seguimiento_factura.gestion_op',id_manipulado=serializer.data['id'])
					logs_model.save()

					fecha_pago= self.request.DATA['fecha_pago'] if 'fecha_pago' in request.DATA else None;
					if fecha_pago:
						Factura.objects.filter(codigo_op_id = serializer.data['id']).update(fecha_pago = fecha_pago, pagada = 1)

					transaction.savepoint_commit(sid)
					
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok',
						'data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					#print(serializer.errors)
					return Response({'message':'datos requeridos no fueron recibidos','success':'fail',
			 		'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				#print(e)
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error',
					'data':''},status=status.HTTP_400_BAD_REQUEST)
		  				

	def destroy(self,request,*args,**kwargs):
		try:
			instance = self.get_object()
			self.perform_destroy(instance)
			return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''},status=status.HTTP_400_BAD_REQUEST)	
#Fin api rest para gestion op


#actualiza el campo orden de pago de las facturas
@transaction.atomic
def actualizar_orden_factura(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			object_factura=Factura.objects.get(pk=item['id'])

			object_factura.orden_pago=respuesta['orden_pago']
			object_factura.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='seguimiento_factura.factura',id_manipulado=item['id'])
			logs_model.save()

			transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#exporta todos las facturas habilitadas para test op
def export_factura_vencidas(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Facturas_vencidas.xls"'

	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Factuas_vencidas')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})

	row=1
	col=0

	cursor = connection.cursor()

	id_empresa = request.user.usuario.empresa.id
	orden_pago= request.GET['orden_pago'] if 'orden_pago' in request.GET else None
	bloqueo_factura= request.GET['bloqueo_factura'] if 'bloqueo_factura' in request.GET else None
	pagada= request.GET['pagada'] if 'pagada' in request.GET else None
	id_contrato= request.GET['id_contrato'] if 'id_contrato' in request.GET else None
	id_contratista= request.GET['id_contratista'] if 'id_contratista' in request.GET else None
	qset=(~Q(id=0))
	qset = qset &(Q(estado_id = estadoFactura.activa))
	qset = qset &(Q(fecha_reporte__isnull = True ))
	qset = qset &(Q(codigo_op_id__isnull = True ))
	

	if id_empresa:
		qset = qset & (Q(contrato__empresacontrato__empresa = id_empresa) & Q(contrato__empresacontrato__participa = 1) & Q(contrato__activo = 1))

	if id_contrato and int(id_contrato)>0:
		qset = qset &(Q(contrato__mcontrato__id = id_contrato))

	if id_contratista and int(id_contratista)>0:
		qset = qset &(Q(contrato__contratista_id = id_contratista))

	if pagada:	
		qset = qset &(Q(pagada = pagada))

	if bloqueo_factura:
		qset = qset &(Q(bloqueo_factura = bloqueo_factura))

	if orden_pago:
		qset = qset &(Q(orden_pago = orden_pago))

	factura = Factura.objects.filter(qset)
	worksheet.write('A1', 'Contrato', format1)
	worksheet.write('B1', 'Proveedor', format1)
	worksheet.write('C1', 'No. factura', format1)
	worksheet.write('D1', 'Referencia', format1)
	worksheet.write('E1', 'Valor a pagar', format1)
	worksheet.write('F1', 'Fecha reporte', format1)
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 18)
	worksheet.set_column('C:C', 18)
	worksheet.set_column('D:D', 12)
	worksheet.set_column('E:E', 12)
	worksheet.set_column('F:F', 16)

	for factura in factura:

		worksheet.write(row, col,factura.contrato.nombre,format2)
		worksheet.write(row, col+1,factura.contrato.contratista.nombre,format2)
		worksheet.write(row, col+2,factura.numero,format2)
		worksheet.write(row, col+3,factura.referencia,format2)
		worksheet.write(row, col+4,factura.valor_factura,format2)
		row +=1

	workbook.close()

	return response
  

@login_required
def test_op(request):

	try:

		id_empresa = request.user.usuario.empresa.id
			
		contrato = Contrato.objects.all()

		validacion=False
		lista_datos=[]
		
		for item in contrato:

			qset=(~Q(id=0))

			if item.empresa_contrato()[0].empresa.id == id_empresa:

				if item.id and int(item.id)>0:

					## print item.id
					qset = qset &( Q(contrato__id=item.id) )
					
				factura = Factura.objects.filter(qset)

				factura = Factura.objects.filter(qset)

				for fac in factura:

			 		#print fac.id
			 		## print item.contratista.id
			 		## print item.contratista.nombre

					if fac.id:

						lista_datos.append(

							{		
								'id_contratista':item.contratista.id,
								'nombre_contratista':item.contratista.nombre,
								'referencia':fac.referencia,
								'fecha_reporte':fac.fecha_reporte,
								'valor_a_pgar':fac.valor_factura
							}
						)

						validacion=True


		if validacion==True:

			return JsonResponse({'message':'','success':'ok','data':lista_datos})

		else:

			return JsonResponse({'message':'No se encontraron registro','success':'error',
				'data':''})
		
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def listado_contrato(request):

	try:

		id_empresa = request.user.usuario.empresa.id
		orden_pago= request.GET['orden_pago']
		bloqueo_factura= request.GET['bloqueo_factura']
		pagada= request.GET['pagada']
		recursos_propios= request.GET['recursos_propios']
		dato= request.GET['dato']

		qset = (~Q(id=0))
		fecha_actual = datetime.now().strftime('%Y-%m-%d')

		qset = qset &(Q(fecha_vencimiento__lte = fecha_actual))
	
		if dato:
			qset = qset &( Q(contrato__contratista__nombre__icontains=dato) | Q(contrato__nombre__icontains=dato) )

		if orden_pago:
			qset = qset &( Q(orden_pago = orden_pago)	)
		
		if bloqueo_factura:
			qset = qset &( Q(bloqueo_factura = bloqueo_factura) )

		if recursos_propios:
			qset = qset &(Q(recursos_propios = recursos_propios))

		if id_empresa:
			qset = qset &(Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1))

		if pagada:
			qset = qset &(Q(pagada = pagada))

		qset = qset &(Q(codigo_op_id__isnull = True) & Q(fecha__gte = '2018-10-31'))
			
		factura = Factura.objects.filter(qset).exclude(estado_id__in = [estadoFactura.compensada]).exclude(referencia__exact = '').exclude(referencia__exact = '0')
		validacion = False
		lista_datos = []
		lista_id = []

		for item in factura:
			# # print item.id
			#if item.contrato.empresa_contrato()[0].empresa.id == id_empresa:

			if (item.referencia!='' and int(item.referencia)>0):
				
				if item.contrato.id and int(item.contrato.id)>0:
	 					
					if  item.contrato.id not in lista_id:						

						lista_id.append(item.contrato.id)

						fact = Factura.objects.filter(fecha_vencimiento__lte = fecha_actual,estado_id=estadoFactura.activa,contrato__id=item.contrato.id,orden_pago=orden_pago,recursos_propios=recursos_propios,pagada=pagada,fecha__gte = '2018-10-31').values('id','numero','referencia','fecha_reporte','valor_contable')

						if item.mcontrato:

							cuenta = FinancieroCuenta.objects.filter(contrato__id=item.mcontrato.id,estado__id=87).values('id','nombre','numero','fiduciaria').first()
	 					
						else:
							cuenta=''

						f=[]

						valor_total=Factura.objects.filter(contrato__id=item.contrato.id,orden_pago=orden_pago,recursos_propios=recursos_propios,pagada=pagada).aggregate(suma_ingreso=Sum('valor_contable'))	

						for item2 in fact:

							if item2['referencia']!='' and int(item2['referencia'])>0:

								f.append(
		 							{	
		 								'id_factura':item2['id'],
		 								'numero':item2['numero'],
		 								'referencia':item2['referencia'],
		 								'fecha_reporte':item2['fecha_reporte'],
		 								'valor_factura':item2['valor_contable']
		 									
		 							}
		 						)


						idcontrato=''
						nombrecontrato=''

						if item.mcontrato:
							idcontrato=item.mcontrato.id
							nombrecontrato=item.mcontrato.nombre


						lista_datos.append(

							{		
								'id_contratista':item.contrato.contratista.id,
								'nombre_contratista':item.contrato.contratista.nombre,
								'contrato_id':idcontrato,
								'contrato_nombre':nombrecontrato, 
								'listado_factura' : f,
								'valor_total':valor_total,
								'cuenta':cuenta				

							}

						)

						validacion=True


		if validacion==True:

			return JsonResponse({'message':'','success':'ok','data':lista_datos})

		else:

			return JsonResponse({'message':'No se encontraron registro','success':'error',
				'data':''})
		
	except Exception as e:
		# print(e)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



@login_required
def listado_de_pagos(request):

	try:

		id_pago= request.GET['id_pago']

		qset = (~Q(id=0))

		if id_pago:
			qset = qset &( Q(id=id_pago) )

		if id_pago and int(id_pago)>0:
			gestionOp = GestionOp.objects.filter(qset)

		else:
			# print 'entro'
			gestionOp = GestionOp.objects.all()

		validacion=False
		lista_datos=[]
	
		for item in gestionOp:

			fact=Factura.objects.filter(codigo_op__id=item.id).first()
			cuenta=FinancieroCuenta.objects.filter(contrato__id=fact.contrato.id).values('nombre').first()

			lista_datos.append(

				{		
					'beneficiario':item.beneficiario.nombre,
					'codigo_compensacion':item.codigo,
					'valor':item.valor,
					'contrato_nombre':fact.contrato.nombre,
					'cuenta_nombre':cuenta,
					'id':item.id
			
				}

			)

			validacion=True


		if validacion==True:

			return JsonResponse({'message':'','success':'ok','data':lista_datos})

		else:

			return JsonResponse({'message':'No se encontraron registro','success':'error',
				'data':''})
		
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#exporta gestion op
def export_gestion_op(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Cuentas_por_pagar.xls"'

	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cuentas_por_pagar')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})
	format_date=workbook.add_format({'num_format': 'yyyy-mm-dd'})
	qset=''
	row=1
	col=0

	cursor = connection.cursor()

	recursos = request.GET['recursos']

	if recursos:
	
		qset = (
				Q(pagado_recursos_propios=recursos)
			)

	gestionOp = GestionOp.objects.filter(qset)

	worksheet.write('A1', 'Contrato', format1)
	worksheet.write('B1', 'Cuenta origen', format1)
	worksheet.write('C1', 'Numero cuenta origen', format1)
	worksheet.write('D1', 'Tipo cuenta origen', format1)
	worksheet.write('E1', 'Numero cuenta destino', format1)
	worksheet.write('F1', 'Tipo cuenta destino', format1)
	worksheet.write('G1', 'Entidad bancaria', format1)
	worksheet.write('H1', 'Beneficiario', format1)
	worksheet.write('I1', 'NO. acreedor', format1)
	worksheet.write('J1', 'TEST/OP', format1)
	worksheet.write('K1', 'Valor', format1)
	worksheet.write('L1', 'NO. factura', format1)
	worksheet.write('M1', 'Fecha pago TEST/OP', format1)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 18)
	worksheet.set_column('C:C', 22)
	worksheet.set_column('D:D', 12)
	worksheet.set_column('E:E', 18)
	worksheet.set_column('F:F', 12)
	worksheet.set_column('G:G', 17)
	worksheet.set_column('H:H', 12)
	worksheet.set_column('I:I', 13)
	worksheet.set_column('J:J', 12)
	worksheet.set_column('K:K', 12)
	worksheet.set_column('L:L', 12)
	worksheet.set_column('M:M', 15)

	for item in gestionOp:

		factura = Factura.objects.filter(codigo_op_id = item.id)[:1].get()
		
		cuenta=FinancieroCuenta.objects.filter(contrato__id=factura.contrato.mcontrato.id).values('tipo__nombre','numero','nombre').first()
		empreCuen=EmpresaCuenta.objects.filter(empresa__id=item.beneficiario.id).values('entidad_bancaria','tipo_cuenta__nombre','numero_cuenta').first()

		fact=Factura.objects.filter(codigo_op__id=item.id).values('numero')
		numFa=''
		for f in fact:

			numFa=numFa +str(f['numero'])+','

		worksheet.write(row, col,item.contrato.nombre,format2)
		worksheet.write(row, col+1,cuenta['nombre'] if cuenta is not None else '',format2)
		worksheet.write(row, col+2,cuenta['numero'],format2)
		worksheet.write(row, col+3,cuenta['tipo__nombre'],format2)

		if empreCuen:

			worksheet.write(row, col+4,empreCuen['numero_cuenta'],format2)
			worksheet.write(row, col+5,empreCuen['tipo_cuenta__nombre'],format2)
			worksheet.write(row, col+6,empreCuen['entidad_bancaria'],format2)

		worksheet.write(row, col+7,item.beneficiario.nombre,format2)
		worksheet.write(row, col+8,item.beneficiario.codigo_acreedor,format2)
		worksheet.write(row, col+9,item.codigo,format2)
		worksheet.write(row, col+10,item.valor,format2)
		worksheet.write(row, col+11,numFa,format2)
		worksheet.write(row, col+12,item.fecha_pago,format_date)
			
		row +=1

	workbook.close()

	return response
    #return response

#exporta gestion op
def exportar_gestion_op_recursos_propios(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Cuentas_por_pagar.xls"'

	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Cuentas_por_pagar')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})
	format_date=workbook.add_format({'num_format': 'yyyy-mm-dd'})
	qset=''
	row=1
	col=0

	cursor = connection.cursor()

	recursos = request.GET['recursos']

	if recursos:
	
		qset = (
				Q(pagado_recursos_propios=recursos)
			)

	gestionOp = GestionOp.objects.filter(qset)

	worksheet.write('A1', 'Codigo compensacion', format1)
	worksheet.write('B1', 'Valor a pagar', format1)
	worksheet.write('C1', 'No. factura', format1)
	worksheet.write('D1', 'Fecha compensacion', format1)

	worksheet.set_column('A:A', 17)
	worksheet.set_column('B:B', 12)
	worksheet.set_column('C:C', 12)
	worksheet.set_column('D:D', 17)

	for item in gestionOp:

		factura = Factura.objects.filter(codigo_op_id = item.id)[:1].get()
		beneficiario = factura.contrato.contratista
		
		cuenta=FinancieroCuenta.objects.filter(contrato__id=factura.contrato.mcontrato.id).values('tipo__nombre','numero','nombre').first()
		fact=Factura.objects.filter(codigo_op__id=item.id).values('numero')

		valor_gestion_op = Factura.objects.filter(codigo_op_id=item.id, pagada = True).aggregate(total=Sum('valor_contable'))
		 	
		worksheet.write(row, col,item.codigo,format2)
		worksheet.write(row, col+1, str(valor_gestion_op['total']),format2)
		worksheet.write(row, col+2,','.join(str(e['numero']) for e in fact),format2)
		worksheet.write(row, col+3,item.fecha_pago,format_date)
			
		row +=1

	workbook.close()

	return response
    #return response


#recorre el archivo en excel
def consulta_excel(request):

	try:
		soporte = request.FILES['archivo']
		qset = ''
		doc = openpyxl.load_workbook(soporte)
		nombrehoja = doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i = 2
		sw = 0
		contador = hoja.max_row - 1
		fila_lista = []
		if int(contador) > 0:	
			filas = [f for f in hoja.rows]
			for fila in filas[1:]:
				# print fila
				# Numero de factura = fila[0].value
				# Documento sap o referencia = fila[1].value
				# Valor pagado = fila[2].value
				# Cod acreedor = fila[3].value
				# Fecha de vencimiento = fila[4].value
				# Fecha de contabilizacion = fila[5].value
				# print fila[0].value
				# print fila[1].value
				# print fila[2].value
				# print fila[3].value
				# print fila[4].value
				# print fila[5].value
				#valida que los campos del archivo no vengan vacios
				if fila[0].value and fila[1].value and fila[2].value and fila[3].value and fila[4].value and fila[5].value:
					# print 'entro'
					#fecha_excel =  fila[0].value.strftime("%Y-%m-%d")
					num_factura= fila[0].value
					referencia= fila[1].value
					valor_pagado= (float(fila[2].value))*(-1)
					codigo_acreedor= fila[3].value
					fecha_vencimiento= fila[4].value
					fecha_vencimiento= fecha_vencimiento.replace(".", "-")
					fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%d-%m-%Y").strftime("%Y-%m-%d")
					fecha_contabilizacion= fila[5].value
					fecha_contabilizacion= fecha_contabilizacion.replace(".", "-")
					fecha_contabilizacion = datetime.strptime(fecha_contabilizacion, "%d-%m-%Y").strftime("%Y-%m-%d")
					mensaje_modelo_detalle= False

					qset = Q(estado__id= estadoFactura.activa)
					qset = qset & (Q(pagada = False))
					qset = qset & (Q(referencia__exact = '') | Q(referencia__exact = '0'))
					qset = qset & (Q(fecha__gte = '2018-10-31') )

					if num_factura:
						qset = qset &(Q(numero__exact = num_factura))

					if codigo_acreedor and int(codigo_acreedor)>0:
						qset = qset &(Q(contrato__contratista__codigo_acreedor__exact = codigo_acreedor))


					try:
						factura= Factura.objects.get(qset)
						object_factura = Factura.objects.get(qset)
						object_factura.fecha_contabilizacion = fecha_contabilizacion
						object_factura.fecha_vencimiento = fecha_vencimiento
						object_factura.referencia = referencia
						object_factura.valor_contable = valor_pagado
						object_factura.save()

					except Factura.DoesNotExist:
						# print "Factura.DoesNotExist"
						sw = 1
						fila_lista.append(i)


				else:
					fila_lista.append(i)

				i+=1


			respuesta  = JsonResponse({'message': 'Las facturas se han actualizado exitosamente','success':'ok','data': ''}) if sw == 0 else JsonResponse({'message':'Las siguientes filas del documento excel no fueron actualizadas: <br>'+'<br>'.join('-'+str(e) for e in fila_lista),'success':'fail','data':''});
			return respuesta

		return JsonResponse({'message':'No se encontraron registros en el archivo','success':'ok','data':''})				

	
	except Exception as e:
		functions.toLog(e,'seguimiento_factura.consulta_excel')
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

    #return response


#actualiza la informacion de factura segun el archivo excel cargado
@transaction.atomic
def actualizar_excel(request):

	# sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)

		for item in respuesta['lista']:
			# print item
			object_factura=Factura.objects.get(pk=item['id_factura'])

			factura=Factura.objects.filter(id=item['id_factura']).values('id','numero','referencia','valor_contable','fecha_contabilizacion','pagada').first()
			
			object_factura.fecha_contabilizacion = item["fecha_contabilizacion_archi"]
			object_factura.fecha_vencimiento = item["fecha_vencimiento"]

			# if item['validacion']==1:
			# 	object_factura.orden_pago=1

			# if item['eliminado']==True:
			# 	if item['validacion']!=4:
			# 		object_factura.orden_pago=1
										
			# if item['procesar']==True:

			# 	if item['validacion']==2:
			# 		object_factura.orden_pago= 1
			# 		object_factura.referencia= item['referencia']

			# 	if item['validacion']==3:

			# 		object_factura.orden_pago= 1
			# 		object_factura.valor_contable= item['valor_pagado']

			
			# object_factura.save()


			# logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='seguimiento_factura.factura',id_manipulado=item['id_factura'])
			# logs_model.save()

			# transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
		
	except Exception as e:
		#print(e)
		# transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



# actualiza el archivo excel segun la opcion que se escoja
@transaction.atomic
def actualizar_conflicto(request):

	sid = transaction.savepoint()

	try:

		lista=request.POST['_content']
		respuesta= json.loads(lista)

		id_factura= respuesta['id_factura']
		validacion= respuesta['validacion']
		referencia_archivo= respuesta['referencia']
		referencia_factura= respuesta['referencia2']
		valor_archivo= respuesta['valor']
		valor2_factura= respuesta['valor2']
		orden_pago= respuesta['orden_pago']
		fecha_reporte= respuesta['fecha_reporte']
		
			
		object_factura=Factura.objects.get(pk=id_factura)

		if validacion==0:
			
			object_factura.orden_pago=orden_pago
			object_factura.fecha_reporte=fecha_reporte
			object_factura.save()

		if validacion==1:

			object_factura.orden_pago=orden_pago
			object_factura.fecha_reporte=fecha_reporte
			object_factura.referencia=referencia_archivo
			object_factura.save()


		if validacion==2:

			object_factura.orden_pago=orden_pago
			object_factura.fecha_reporte=fecha_reporte
			object_factura.valor_contable=valor_archivo
			object_factura.save()


		if validacion==4:

			object_factura.fecha_reporte=fecha_reporte
			object_factura.save()

		if validacion==5:

			factura=Factura.objects.filter(id=id_factura).first()

			object_factura.fecha_reporte=factura.fecha_reporte
			object_factura.save() 

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='seguimiento_factura.factura',id_manipulado=id_factura)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'Los registros se han actualizado correctamente','success':'ok',
							'data':''})
			

	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@login_required
def index_seguimiento(request):

	qset_por_contabilizar = (Q(estado_id = estadoFactura.activa))
	qset_por_contabilizar = qset_por_contabilizar & (Q(pagada = False) )
	qset_por_contabilizar = qset_por_contabilizar & (Q(referencia__exact = '') | Q(referencia__exact = '0'))
	qset_por_contabilizar = qset_por_contabilizar & (Q(fecha__gte = '2018-10-31') )

	cantidadFacturaPorContabilizar = Factura.objects.filter(qset_por_contabilizar).count()
	
	fecha_actual = datetime.now().strftime('%Y-%m-%d')
	query_test_recursos = Factura.objects.filter(fecha_vencimiento__lte = fecha_actual, orden_pago = True ,pagada = False, codigo_op_id__isnull = True, estado_id = estadoFactura.activa, fecha__gte = '2018-10-31' ).exclude(referencia__exact = '').exclude(referencia__exact = '0')
	# facturas TEST/OP
	cantidadFacturaTestOp = query_test_recursos.filter(recursos_propios = False).count()
	# cantidad recursos propios
	cantidadTetstRecursosPropios = query_test_recursos.filter(recursos_propios = True).count()

	# facturas que se pagan con cuentas
	cantidadFacturaPorPagar = Factura.objects.filter(estado_id = estadoFactura.compensada, pagada = False, recursos_propios = False, codigo_op_id__isnull = False).count()

	return render(request, 'seguimiento_factura/index.html'
		,{'cantidadFacturaPorPagar': cantidadFacturaPorPagar, 'cantidadFacturaPorContabilizar': cantidadFacturaPorContabilizar, 'cantidadTetst': cantidadFacturaTestOp ,'cantidadTetstRecursosPropios':cantidadTetstRecursosPropios,'app':'seguimiento_factura'}
		)


@login_required
def filtro_consultar_contrato_pago(request):

	try:

		id_empresa = request.user.usuario.empresa.id
		recursos= request.GET['recursos']
		lista_datos=[]
			
		gestion = GestionOp.objects.filter(pagado_recursos_propios=recursos,contrato__empresacontrato__empresa=id_empresa).values('contrato__id','contrato__nombre').distinct()

		for item in gestion:

			lista_datos.append(

				{		
					'contrato_id':item['contrato__id'],
					'contrato_nombre':item['contrato__nombre']
				}

			)

		return JsonResponse({'message':'','success':'ok','data':lista_datos})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})




@login_required
def filtro_consultar_contratista_pago(request):

	try:

		id_empresa = request.user.usuario.empresa.id
		recursos= request.GET['recursos']
		lista_datos=[]
			
		gestion = GestionOp.objects.filter(pagado_recursos_propios=recursos,contrato__empresacontrato__empresa=id_empresa).values('beneficiario__id','beneficiario__nombre').distinct()

		for item in gestion:

			lista_datos.append(

				{		
					'contratista_id':item['beneficiario__id'],
					'contratista_nombre':item['beneficiario__nombre']
				}

			)

		return JsonResponse({'message':'','success':'ok','data':lista_datos})
	
	except Exception as e:
		#print(e)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})



#exporta gestion op
def export_factura_pagadas(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="facturas_pagadas.xls"'

	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Facturas')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})
	format_date=workbook.add_format({'num_format': 'yyyy-mm-dd'})
	
	qset=''
	row=1
	col=0

	cursor = connection.cursor()

	bloqueo_factura = request.GET['bloqueo_factura']
	recursos_propios = request.GET['recursos_propios']
	pagada = request.GET['pagada']
	codigo_op = request.GET['codigo_op']

	qset = Q(codigo_op_id=codigo_op)

	if bloqueo_factura:
		qset = qset &(Q(bloqueo_factura=bloqueo_factura))

	if pagada:
		qset = qset &(Q(pagada=pagada))

	if recursos_propios:
		qset = qset &(Q(recursos_propios=recursos_propios))

	factura = Factura.objects.filter(qset)

	worksheet.write('A1', 'Doc. SAP', format1)
	worksheet.write('B1', 'Fecha reporte', format1)
	worksheet.write('C1', 'No. factura', format1)
	worksheet.write('D1', 'Cod. acreedor', format1)
	worksheet.write('E1', 'Nombre acreedor. ', format1)
	worksheet.write('F1', 'Valor a pagar', format1)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 18)
	worksheet.set_column('C:C', 18)
	worksheet.set_column('D:D', 15)
	worksheet.set_column('E:E', 20)
	worksheet.set_column('F:F', 15)


	for item in factura:

		worksheet.write(row, col,item.referencia,format2)
		worksheet.write(row, col+1,item.fecha_reporte,format_date)
		worksheet.write(row, col+2,item.numero,format2)
		worksheet.write(row, col+3,item.contrato.contratista.codigo_acreedor,format2)
		worksheet.write(row, col+4,item.contrato.contratista.nombre,format2)
		worksheet.write(row, col+5,item.valor_contable,format2)
			
		row +=1

	workbook.close()

	return response
    #return response


def export_factura(request):
	
	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="facturas_pagadas.xls"'

	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Facturas')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})
	format_date=workbook.add_format({'num_format': 'yyyy-mm-dd'})
	
	qset=''
	row=1
	col=0

	cursor = connection.cursor()

	bloqueo_factura = request.GET['bloqueo_factura']
	recursos_propios = request.GET['recursos_propios']
	codigo_op = request.GET['codigo_op']

	qset = Q(codigo_op_id=codigo_op)

	if bloqueo_factura:
		qset = qset &(Q(bloqueo_factura=bloqueo_factura))

	if recursos_propios:
		qset = qset &(Q(recursos_propios=recursos_propios))

	factura = Factura.objects.filter(qset)

	worksheet.write('A1', 'Doc. SAP', format1)
	worksheet.write('B1', 'Fecha reporte', format1)
	worksheet.write('C1', 'NO. factura', format1)
	worksheet.write('D1', 'Valor a pagar', format1)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 18)
	worksheet.set_column('C:C', 18)
	worksheet.set_column('D:D', 12)


	for item in factura:

		worksheet.write(row, col,item.referencia,format2)
		worksheet.write(row, col+1,item.fecha_reporte,format_date)
		worksheet.write(row, col+2,item.numero,format2)
		worksheet.write(row, col+3,item.valor_contable,format2)
			
		row +=1

	workbook.close()

	return response
    #return response


#genera el reporte de las plantillas fiduciaria BANCO DE BOGOTA; BANCOLOMBIA
def generar_reporte(request):

	try:

		testop_id = request.GET['testop_id']
		orden_pago = request.GET['orden_pago']
		bloqueo_factura = request.GET['bloqueo_factura']

		fact = Factura.objects.filter(codigo_op__id=testop_id, orden_pago=orden_pago, bloqueo_factura = bloqueo_factura)
		
		factura_first = Factura.objects.filter(codigo_op__id=testop_id, orden_pago=orden_pago, bloqueo_factura = bloqueo_factura).first()

		valor_total_factura= Factura.objects.filter(codigo_op__id=testop_id, orden_pago=orden_pago, bloqueo_factura=bloqueo_factura).aggregate(suma_ingreso=Sum('valor_factura'))	

		cuenta_empresa= EmpresaCuenta.objects.filter(empresa__id = factura_first.contrato.contratista_id, estado__id=87).values('entidad_bancaria','numero_cuenta','tipo_cuenta__id','empresa__nit','empresa__nombre','tipo_cuenta__nombre').first()

		# datos de la cuenta
		cuenta = FinancieroCuenta.objects.filter(contrato_id = factura_first.contrato.mcontrato_id) 
		codfideico = ''
		codfideico_antiguo = ''
		nombrefideico = ''
		nombrefiduciaria = ''
		macro_contrato = ''

		for item2 in cuenta:

			nocuenta = item2.numero 
			# codigo fidecomiso
			if item2.codigo_fidecomiso:
				codfideico = item2.codigo_fidecomiso

			if item2.codigo_fidecomiso_a:
				codfideico_antiguo = item2.codigo_fidecomiso_a
			# nombre de fidecomiso

			if item2.nombre_fidecomiso :
				nombrefideico= item2.nombre_fidecomiso

			tipofideico = item2.tipo.nombre

			if item2.fiduciaria:
				nombrefiduciaria = item2.fiduciaria

			if item2.contrato:
				macro_contrato = item2.contrato.nombre


		ruta = settings.STATICFILES_DIRS[0]
		newpath = ruta + '/papelera/'
		empresa_id = request.user.usuario.empresa.id
		ahora = datetime.now()

		if nombrefiduciaria=='FIDUBOGOTA':				

			filename = "plantillas/solicituGiros/FIDUBOGOTACREATE.xlsx"
			extension = filename[filename.rfind('.'):]
			nombre = 'empresa'+str(empresa_id)+''+str(request.user.usuario.id)+extension		

			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )

			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre
			# wb = load_workbook(filename = ruta, keep_vba=True)
			wb = load_workbook(filename = ruta, read_only=False)

			inicio = wb.worksheets[0]

			# descargar imagen para la plantilla  bogota
			filename = "plantillas/solicituGiros/excel.png"
			extension = filename[filename.rfind('.'):]
			nombre = 'fiduciariabogotainicio'+extension		
			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )
			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre

			img_inicio = Image(ruta)
			inicio.add_image(img_inicio, 'A1')


			ws = wb.worksheets[1]
			ws_datos = wb.worksheets[2]
			# ws = wb['10112015']	



			# #Estilo para los bordes del archivo
			thick_border = Border(
				left=Side(style='medium',color='000000'),
				right=Side(style='medium',color='000000'),
				top=Side(style='medium',color='000000'),
				bottom=Side(style='medium',color='000000'))
	
			thin_border2 = Border(
				left=Side(style='thin',color='A6A5A5'),
				right=Side(style='thin',color='A6A5A5'),
				bottom=Side(style='thin',color='A6A5A5'))
	
			thin_border3 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='thin',color='000000'),
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thick_border4 = Border(
				bottom=Side(style='medium',color='000000'))

			thick_border5 = Border(
				right=Side(style='medium',color='000000'),		
				bottom=Side(style='medium',color='000000'))

			thick_border6 = Border(
				right=Side(style='medium',color='000000'))

			thin_border7 = Border(
				bottom=Side(style='thin',color='000000'))

			thin_border8 = Border(
				right=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border9 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='medium',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border10 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='medium',color='000000'),
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border11 = Border(
				right=Side(style='medium',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border12 = Border(
				right=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))

			thin_border13 = Border(
				top=Side(style='thin',color='000000'),
				right=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border14 = Border(
				top=Side(style='thin',color='000000'),
				left=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thick_border15 = Border(
				left=Side(style='medium',color='000000'))

			thick_border16 = Border(
				left=Side(style='medium',color='000000'),		
				bottom=Side(style='medium',color='000000'))

			thin_border17 = Border(
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border18 = Border(
				left=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))

			thick_border19 = Border(
				right=Side(style='medium',color='000000'),
				left=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))




			# order de operacion sifi COLUM R TO COLUM Z
			for x in xrange(16,27):
				ws.cell(row=1, column=x).border = thick_border4

			for x in xrange(29,33):
				ws.cell(row=1, column=x).border = thick_border4


			for x in xrange(17,27):
				ws.cell(row=8, column=x).border = thin_border7


			for x in xrange(4,6):
				ws.cell(row=9, column=x).border = thin_border17

			for x in xrange(7,10):
				ws.cell(row=9, column=x).border = thin_border17

			for x in xrange(12,27):
				ws.cell(row=9, column=x).border = thin_border7

			ws.cell(row=9, column=9).border = thin_border13

			ws.cell(row=1, column=33).border = thick_border5

			ws.cell(row=8, column=33).border = thick_border5
			ws.cell(row=9, column=33).border = thick_border5
			ws.cell(row=10, column=33).border = thick_border5
			ws.cell(row=11, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=12, column=x).border = thick_border4

			ws.cell(row=12, column=33).border = thick_border5	

			ws.cell(row=13, column=31).border = thin_border7


			for x in xrange(8,10):
				ws.cell(row=14, column=x).border = thin_border17	

			ws.cell(row=14, column=9).border = thin_border13	

			for x in xrange(3,33):
				ws.cell(row=15, column=x).border = thick_border4

			ws.cell(row=15, column=33).border = thick_border5
			ws.cell(row=15, column=31).border = thin_border12
			ws.cell(row=15, column=26).border = thin_border18

			ws.cell(row=15, column=33).border = thick_border19

			for x in xrange(3,33):
				ws.cell(row=16, column=x).border = thick_border4

			ws.cell(row=16, column=33).border = thick_border5

			# LIQUIDACION DE LA OPERACION

			for x in xrange(12,18):
				ws.cell(row=17, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=17, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=18, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=18, column=x).border = thin_border7

			ws.cell(row=18, column=18).border = thin_border9

			for x in xrange(27,33):
				ws.cell(row=18, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=24, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=24, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=24, column=x).border = thick_border4

			for x in xrange(12,18):
				ws.cell(row=25, column=x).border = thick_border4

			for x in xrange(27,33):
				ws.cell(row=25, column=x).border = thick_border4

			ws.cell(row=18, column=33).border = thin_border10
			ws.cell(row=19, column=33).border = thin_border9
			ws.cell(row=20, column=33).border = thin_border9
			ws.cell(row=21, column=33).border = thin_border9
			ws.cell(row=22, column=33).border = thin_border9
			ws.cell(row=23, column=33).border = thin_border9

			# fin liquidacion de la operacion


			for x in xrange(5,33):
				ws.cell(row=26, column=x).border = thin_border7

			for x in xrange(3,33):
				ws.cell(row=27, column=x).border = thin_border7

			ws.cell(row=26, column=33).border = thin_border11
			ws.cell(row=27, column=33).border = thin_border11


			for x in xrange(3,33):
				ws.cell(row=28, column=x).border = thick_border4

			ws.cell(row=28, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=29, column=x).border = thick_border4

			ws.cell(row=29, column=33).border = thick_border5

			for x in xrange(6,33):
				ws.cell(row=30, column=x).border = thin_border7

			ws.cell(row=30, column=33).border = thin_border11

			for x in xrange(3,33):
				ws.cell(row=32, column=x).border = thick_border4


			# formacion de operacion
			for x in xrange(3,33):
				ws.cell(row=33, column=x).border = thick_border4

			ws.cell(row=33, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=72, column=x).border = thick_border4
			ws.cell(row=72, column=33).border = thick_border5


			for x in xrange(26,32):
				ws.cell(row=39, column=x).border = thin_border7

			for x in xrange(14,32):
				ws.cell(row=40, column=x).border = thin_border7

			for x in xrange(13,19):
				ws.cell(row=41, column=x).border = thin_border7

			for x in xrange(22,32):
				ws.cell(row=41, column=x).border = thin_border7

			# observaciones
			for x in xrange(3,33):
				ws.cell(row=43, column=x).border = thick_border4
			ws.cell(row=43, column=33).border = thick_border5

			ws.cell(row=44, column=2).border = thick_border15
			ws.cell(row=45, column=2).border = thick_border15
			ws.cell(row=46, column=2).border = thick_border15
			ws.cell(row=47, column=2).border = thick_border15
			ws.cell(row=48, column=2).border = thick_border15
			ws.cell(row=49, column=2).border = thick_border16

			ws.cell(row=44, column=33).border = thick_border6
			ws.cell(row=45, column=33).border = thick_border6
			ws.cell(row=46, column=33).border = thick_border6
			ws.cell(row=47, column=33).border = thick_border6
			ws.cell(row=48, column=33).border = thick_border6

			# condiciones de manejo
			for x in xrange(3,33):
				ws.cell(row=49, column=x).border = thick_border4
			ws.cell(row=49, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=50, column=x).border = thick_border4
			ws.cell(row=50, column=33).border = thick_border5

			for x in xrange(3,13):
				ws.cell(row=53, column=x).border = thin_border7

			for x in xrange(14,24):
				ws.cell(row=53, column=x).border = thin_border7

			for x in xrange(25,33):
				ws.cell(row=53, column=x).border = thin_border7

			ws.cell(row=52, column=2).border = thick_border15
			ws.cell(row=53, column=2).border = thick_border15

			for x in xrange(6,13):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(6,13):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(6,13):
				ws.cell(row=57, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=57, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=57, column=x).border = thin_border7

			ws.cell(row=51, column=33).border = thick_border6
			ws.cell(row=52, column=33).border = thick_border6
			ws.cell(row=53, column=33).border = thin_border11


			ws.cell(row=54, column=33).border = thick_border6
			ws.cell(row=55, column=33).border = thin_border11
			ws.cell(row=56, column=33).border = thin_border11
			ws.cell(row=57, column=33).border = thin_border11		



			# uso exclusivo fidubogota
			for x in xrange(3,33):
				ws.cell(row=58, column=x).border = thick_border4
			ws.cell(row=58, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=59, column=x).border = thick_border4
			ws.cell(row=59, column=33).border = thick_border5

			
			for x in xrange(9,19):
				ws.cell(row=61, column=x).border = thin_border7

			for x in xrange(22,33):
				ws.cell(row=61, column=x).border = thin_border7

			for x in xrange(13,19):
				ws.cell(row=62, column=x).border = thin_border7

			for x in xrange(25,33):
				ws.cell(row=62, column=x).border = thin_border7

			for x in xrange(4,6):
				ws.cell(row=65, column=x).border = thin_border7

			for x in xrange(16,19):
				ws.cell(row=65, column=x).border = thin_border7

			for x in xrange(22,24):
				ws.cell(row=65, column=x).border = thin_border7


			ws.cell(row=61, column=19).border = thin_border11
			ws.cell(row=62, column=19).border = thin_border11
			ws.cell(row=65, column=19).border = thin_border11

			ws.cell(row=60, column=33).border = thick_border6
			ws.cell(row=61, column=33).border = thin_border11
			ws.cell(row=62, column=33).border = thin_border11

			ws.cell(row=60, column=19).border = thick_border6
			

			# informacion de procesos especiales
			for x in xrange(3,33):
				ws.cell(row=66, column=x).border = thick_border4
			ws.cell(row=66, column=33).border = thick_border5
			ws.cell(row=66, column=19).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=67, column=x).border = thick_border4
			ws.cell(row=67, column=33).border = thick_border5


			for x in xrange(8,12):
				ws.cell(row=68, column=x).border = thin_border7

			for x in xrange(16,25):
				ws.cell(row=68, column=x).border = thin_border7

			for x in xrange(31,33):
				ws.cell(row=68, column=x).border = thin_border7


			for x in xrange(8,12):
				ws.cell(row=69, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=69, column=x).border = thin_border7

			for x in xrange(8,12):
				ws.cell(row=70, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=70, column=x).border = thin_border7

			for x in xrange(8,12):
				ws.cell(row=71, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=71, column=x).border = thin_border7

			for x in xrange(30,33):
				ws.cell(row=71, column=x).border = thin_border7

			ws.cell(row=68, column=33).border = thick_border6
			ws.cell(row=71, column=33).border = thin_border11


			# informacion de procesos generales
			for x in xrange(3,33):
				ws.cell(row=73, column=x).border = thick_border4
			ws.cell(row=73, column=33).border = thick_border5

			for x in xrange(28,33):
				ws.cell(row=75, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=76, column=x).border = thin_border7

			for x in xrange(28,33):
				ws.cell(row=76, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=77, column=x).border = thin_border7

			for x in xrange(14,26):
				ws.cell(row=77, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=78, column=x).border = thin_border7

			for x in xrange(17,23):
				ws.cell(row=78, column=x).border = thin_border7

			for x in xrange(28,33):
				ws.cell(row=78, column=x).border = thin_border7

			ws.cell(row=78, column=33).border = thin_border11

			ws.cell(row=77, column=8).border = thin_border13
			ws.cell(row=78, column=8).border = thin_border13
			ws.cell(row=79, column=8).border = thin_border13

			ws.cell(row=77, column=5).border = thin_border14
			ws.cell(row=78, column=5).border = thin_border14
			ws.cell(row=79, column=5).border = thin_border14

			

			for x in xrange(6,8):
				ws.cell(row=79, column=x).border = thin_border7

			for x in xrange(8,20):
				ws.cell(row=83, column=x).border = thin_border7

			for x in xrange(26,33):
				ws.cell(row=83, column=x).border = thin_border7
			

			for x in xrange(3,33):
				ws.cell(row=84, column=x).border = thick_border4
			ws.cell(row=84, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=85, column=x).border = thick_border4
			ws.cell(row=85, column=33).border = thick_border5
			ws.cell(row=86, column=33).border = thick_border6
			ws.cell(row=87, column=33).border = thick_border6
			ws.cell(row=88, column=33).border = thick_border6
			ws.cell(row=89, column=33).border = thick_border6
			ws.cell(row=90, column=33).border = thick_border6


			for x in xrange(3,6):
				ws.cell(row=86, column=x).border = thin_border7
			
			for x in xrange(3,6):
				ws.cell(row=87, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=88, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=89, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=90, column=x).border = thin_border7

			for x in xrange(7,14):
				ws.cell(row=87, column=x).border = thin_border7

			for x in xrange(7,14):
				ws.cell(row=89, column=x).border = thin_border7

			ws.cell(row=86, column=6).border = thin_border8
			ws.cell(row=87, column=6).border = thin_border8
			ws.cell(row=88, column=6).border = thin_border8
			ws.cell(row=89, column=6).border = thin_border8
			ws.cell(row=90, column=6).border = thin_border8

			ws.cell(row=87, column=14).border = thin_border11
			ws.cell(row=88, column=14).border = thick_border5
			ws.cell(row=89, column=14).border = thin_border11
			ws.cell(row=90, column=14).border = thick_border5
			

			ws.cell(row=87, column=27).border = thick_border6
			ws.cell(row=88, column=27).border = thick_border6
			ws.cell(row=89, column=27).border = thick_border6
			ws.cell(row=90, column=27).border = thick_border6			


			for x in xrange(3,33):
				ws.cell(row=91, column=x).border = thick_border4

			
			ws.cell(row=91, column=6).border = thin_border12
			ws.cell(row=91, column=33).border = thick_border5
			ws.cell(row=91, column=14).border = thick_border5
			ws.cell(row=91, column=27).border = thick_border5

			for x in xrange(6,20):
				ws.cell(row=93, column=x).border = thin_border7

			for x in xrange(6,20):
				ws.cell(row=94, column=x).border = thin_border7

			for x in xrange(6,20):
				ws.cell(row=95, column=x).border = thin_border7

			for x in xrange(6,12):
				ws.cell(row=96, column=x).border = thin_border7

			for x in xrange(16,21):
				ws.cell(row=96, column=x).border = thin_border7


			for x in xrange(27,33):
				ws.cell(row=93, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=94, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=95, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=927, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=96, column=x).border = thin_border7

			ws.cell(row=92, column=33).border = thick_border6
			ws.cell(row=93, column=33).border = thin_border11
			ws.cell(row=94, column=33).border = thin_border11
			ws.cell(row=95, column=33).border = thin_border11
			ws.cell(row=96, column=33).border = thin_border11


			# hoja datos
			for x in xrange(3,9):
				ws_datos.cell(row=11, column=x).border = thin_border7

			for x in xrange(3,5):
				ws_datos.cell(row=16, column=x).border = thin_border7

			for x in xrange(7,11):
				ws_datos.cell(row=16, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=18, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=19, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=20, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=21, column=x).border = thin_border7


			#nombre macro contrato
			ws['K9'] = macro_contrato
			
			ws['AB9'] = codfideico[:1]
			ws['AD9'] = codfideico[2:3]
			ws['AF9'] = codfideico[-5:]

			if codfideico_antiguo:

				ws['AB11'] = codfideico_antiguo[:1]
				ws['AD11'] = codfideico_antiguo[2:3]
				ws['AF11'] = codfideico_antiguo[-5:]

			ws['E14'] = str(ahora.day)
			ws['F14'] = str(ahora.month)
			ws['G14'] = str(ahora.year)


			valor_total_letras = 0
			row = 2

			ws['K18'] = valor_total_factura['suma_ingreso']

			if cuenta_empresa:
				ws['M36'] = cuenta_empresa['entidad_bancaria']
				ws['M37'] = cuenta_empresa['numero_cuenta']

				if cuenta_empresa['tipo_cuenta__id']==7:
					ws['Q39'] = 'X'
				else:
					ws['U39'] = 'X'

				ws['V41'] = cuenta_empresa['empresa__nit']
				
			numero_factura=''

			fila = 2
			for f in fact:

				try:
					empresa_cuenta = EmpresaCuenta.objects.get(empresa_id = f.contrato.contratista_id)
					numero_cuenta_beneficiario = str(empresa_cuenta.numero_cuenta)
					tipo_cuenta_beneficiario = ''
					if str(empresa_cuenta.tipo_cuenta.nombre).lower() == 'ahorro':
						tipo_cuenta_beneficiario = 'CH'
					else:
						tipo_cuenta_beneficiario = 'CC'

				except EmpresaCuenta.DoesNotExist:
					numero_cuenta_beneficiario = ''


				nit_titular = str(f.contrato.contratista.nit)
				nombre_titular = str(f.contrato.contratista.nombre)

				ws_datos['C'+str(fila)] = nit_titular
				ws_datos['D'+str(fila)] = nombre_titular
				ws_datos['I'+str(fila)] = f.valor_contable
				# ws_datos['J'+str(fila)] = str(f.fecha_pago)
				ws_datos['J'+str(fila)] = ''
				# ws_datos['K'+str(fila)] = str(codfideico)
				# ws_datos['M'+str(fila)] = str(nit_titular)
				# ws_datos['N'+str(fila)] = str(nombre_titular)
				ws_datos['P'+str(fila)] = ''
				ws_datos['Q'+str(fila)] = numero_cuenta_beneficiario
				ws_datos['R'+str(fila)] = tipo_cuenta_beneficiario
				numero_factura = numero_factura +str(f.numero)+','
				fila+=1


			ws['B44'] = 'EFECTUAR LAS TRANSFERENCIAS CONFORME A LA RELACION EXCEL ADJUNTA. ME PERMITO CERTIFICAR QUE EN CONCORDANCIA CON LOS TERMINOS ACORDADOS EN EL CONTRATO FIDUCIARIO, LOS CONCEPTOS DE LOS GIROS SOLICITADOS MEDIANTE ESTA ORDEN, FUERON CAUSADOS, JUNTO CON LAS RESPECTIVAS RETENCIONES DE RENTA, IVA E ICA Y DEMAS OBLIGACIONES TRIBUTARIAS, EN NUESTRA CONTABILIDAD, LA DECLARACION, PRESENTACION Y PAGO DE ESTAS RETENCIONES SON DE ABSOLUTA RESPONSABILIDAD NUESTRA Y SE HAN REALIZADO DE ACUERDO A TODAS LAS NORMAS Y REGLAMENTACIONES TRIBUTARIAS VIGENTES, TAL COMO CONSTA EN LA CERTIFICACION EXPEDIDA CON RADICADO TES 0094-16 DE FECHA 15 DE FEBRERO-2016 POR CONTADOR. Facturas: '+numero_factura

			# descargar imagen para la plantilla  bogota
			filename = "plantillas/solicituGiros/fiduciariabogota.png"
			extension = filename[filename.rfind('.'):]
			nombre = 'fiduciariabogota'+extension		
			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )
			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre

			img2 = Image(ruta)
			ws.add_image(img2, 'B1')


			wb.template = False
			response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
			response['Content-Disposition'] ='attachment; filename="Reportedeanticipos.xlsx"'
			return response		

		


	except Exception as e:
		print(e)
		functions.toLog(e,'seguimiento_factura.generar_reporte')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
			



@login_required
def habilitar_testOp(request):

	tipo_c = tipoC()
	querysetTipos=Tipo.objects.filter(app='contrato')
	id_empresa = request.user.usuario.empresa.id
	querysetMContrato = Contrato.objects.filter(empresacontrato__empresa=id_empresa, empresacontrato__participa=1, tipo_contrato=tipo_c.m_contrato, activo=1)

	return render(request, 'seguimiento_factura/habilitar_test_op.html'
		,{'tipos':querysetTipos, 'mcontratos':querysetMContrato, 'app':'seguimiento_factura'})

@login_required
def deshabilitar_testOp(request):
	return render(request, 'seguimiento_factura/deshabilitar_test_op.html',{'app':'seguimiento_factura'})

@login_required
def gestion_op(request):

	return render(request, 'seguimiento_factura/gestion_op.html',{'app':'seguimiento_factura'})

@login_required
def pago_factura(request):
	return render(request, 'seguimiento_factura/pago_factura.html',{'app':'seguimiento_factura'})

@login_required
def gestion_op_recursos_propios(request):
	return render(request, 'seguimiento_factura/gestion_op_recursos_propios.html',{'app':'seguimiento_factura'})

@login_required
def consulta_factura_pagada(request):
	return render(request, 'seguimiento_factura/consultar_factura.html',{'app':'seguimiento_factura'})

@login_required
def listado_factura_pagada(request,id_pago=None):
	return render(request, 'seguimiento_factura/listado_factura_pagada.html',{'app':'seguimiento_factura','id_pago':id_pago})		

@login_required
def consulta_pago_factura(request):
	return render(request, 'seguimiento_factura/consulta_pago_factura.html',{'app':'seguimiento_factura'})

@login_required
def consulta_pago_factura_recursos_propios(request):
	return render(request, 'seguimiento_factura/consulta_pago_factura_recursos_propios.html',{'app':'seguimiento_factura'})

@login_required
def carga_masiva(request):
	return render(request, 'seguimiento_factura/carga_masiva.html',{'app':'seguimiento_factura'})

@login_required
def descargar_plantilla(request):
	return functions.exportarArchivoS3('plantillas/seguimiento_factura/seguimientoCarga.xlsx')

@login_required
def administrador_cuenta(request):
	return render(request, 'seguimiento_factura/empresa_cuenta.html',{'app':'seguimiento_factura','model':'EmpresaCuenta',})

# LUIS MENDOZA
@login_required
def facturas_por_contabilizar(request):
	tipo_c = tipoC()
	querysetTipos=Tipo.objects.filter(app='contrato')
	id_empresa = request.user.usuario.empresa.id
	querysetMContrato = Contrato.objects.filter(empresacontrato__empresa=id_empresa, empresacontrato__participa=1, tipo_contrato=tipo_c.m_contrato, activo=1)

	return render(request, 'seguimiento_factura/facturas_por_contabilizar.html'
		,{'tipos':querysetTipos, 'mcontratos':querysetMContrato, 'app':'seguimiento_factura','model':'Factura'}
		)

@login_required
def facturas_por_pagar(request):
	tipo_c = tipoC()
	querysetTipos=Tipo.objects.filter(app='contrato')
	id_empresa = request.user.usuario.empresa.id
	querysetMContrato = Contrato.objects.filter(empresacontrato__empresa=id_empresa, empresacontrato__participa=1, tipo_contrato=tipo_c.m_contrato, activo=1)

	return render(request, 'seguimiento_factura/facturas_por_pagar.html'
		,{'tipos':querysetTipos, 'mcontratos':querysetMContrato, 'app':'seguimiento_factura','model':'Factura'}
		)

def exportar_facturas_por_contabilizar(request):
	try:
		
		# cantidadFacturaPorContabilizar = Factura.objects.filter(estado_id = estadoFactura.activa, referencia__exact = '', fecha__gte = '2018-10-31').count()
		qset = (Q(estado_id=estadoFactura.activa) & Q(fecha__gte = '2018-10-31') )
		qset = qset &(Q(referencia__exact='') | Q(referencia__exact='0'))

		id_mcontrato = request.GET['id_mcontrato'] if 'id_mcontrato' in request.GET else None;
		numero = request.GET['numero'] if 'numero' in request.GET else None;
		tipo_contrato = request.GET['tipo_contrato'] if 'tipo_contrato' in request.GET else None;
		numero_contrato = request.GET['numero_contrato'] if 'numero_contrato' in request.GET else None;
		id_contratista = request.GET['id_contratista'] if 'id_contratista' in request.GET else None;
		id_empresa = request.user.usuario.empresa.id
		qset = qset &(Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1))

		if numero:
			qset = qset &(Q(numero__icontains=numero))
		if tipo_contrato:
			qset = qset &(Q(contrato__tipo_contrato=tipo_contrato))
		if numero_contrato:
			qset = qset &(Q(contrato__numero=numero_contrato))
		if id_contratista and int(id_contratista)>0:
			qset = qset &(Q(contrato__contratista_id=id_contratista))
		if id_mcontrato and int(id_mcontrato)>0:
			qset = qset &(Q(contrato__mcontrato_id=id_mcontrato))


		queryset = Factura.objects.filter(qset)

		#   # INICIO - Crear el Excel
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] ='attachment; filename="Reporte_facturas_por_contabilizar.xlsx"'
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})

		worksheet = workbook.add_worksheet('Facturas por contabilizar')

		worksheet.set_column('A:A', 25)
		worksheet.set_column('B:B', 15)
		worksheet.set_column('C:C', 15)
		worksheet.set_column('D:D', 20)
		worksheet.set_column('E:E', 50)
		worksheet.set_column('F:F', 35)



		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
		format1.set_align('center')
		format1.set_align('vcenter')
		format2=workbook.add_format({'border':0})
		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format_money=workbook.add_format({'border':False,
			'font_size':11,
			'bold':False,
			'valign':'vright',
			'num_format': '$#,##0'})

		row=1
		col=0

		worksheet.write('A1', 'Radicado', format1)
		worksheet.write('B1', 'No Factura', format1)
		worksheet.write('C1', 'Nit', format1)
		worksheet.write('D1', 'Cod. Acreedor', format1)
		worksheet.write('E1', 'Nombre Acreedor', format1)
		worksheet.write('F1', 'Valor factura', format1)

		for objecto in queryset:

			meses = ''
			fecha = ''

			worksheet.write(row, col  , str(objecto.radicado) ,format2)
			worksheet.write(row, col+1,str(objecto.numero) ,format2)
			worksheet.write(row, col+2,str(objecto.contrato.contratista.nit) ,format2)
			worksheet.write(row, col+3, str(objecto.contrato.contratista.codigo_acreedor) ,format2)
			worksheet.write(row, col+4, str(objecto.contrato.contratista.nombre ),format2)
			worksheet.write(row, col+5, objecto.valor_factura ,format_money)
			
			row +=1

		workbook.close()# FIN - Crear el Excel		
		return response

	except Exception as e:
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@transaction.atomic
def actualizar_codigo_compensacion(request):

	try:

		documento = request.FILES.get('soporte')
		excel_document = openpyxl.load_workbook(documento)
		sheetList = excel_document.sheetnames
		hoja = excel_document.get_sheet_by_name(sheetList[0])
		filas = hoja.rows
		fecha_hoy = date.today()
		sw = 0
		# codigo acreedor = fila[0].value
		# codigo sap o referencia = fila[1].value
		# codigo compensacion = fila[2].value
		# valor contable = fila[3].value
		# numero de la factura = fila[4].value

		filas = [f for f in filas]
		codigo_acreedor = [c[0].value for c in filas[1:]]
		codigo_compensacion = [c[2].value for c in filas[1:]]
		array_verificar = zip(codigo_acreedor, codigo_compensacion)

		array_repetidos = [x for x, y in collections.Counter(codigo_compensacion).items() if y > 1]
		if len(array_repetidos)>0:
			sw = validar_cod_compensacion(array_repetidos, codigo_compensacion, array_verificar)

			if sw == 1:
				return JsonResponse({'message': 'Existen codigos de compensacion aignados a mas de un acreedor diferente','success':'fail','data': ''}) 

		i = 2
		fila_lista = []

		for fila in filas[1:]:

			sid = transaction.savepoint()	
			try:
				gestion_op = GestionOp.objects.get(codigo = fila[2].value, pagado_recursos_propios= True)

			except GestionOp.DoesNotExist:
					gestion_op = GestionOp(codigo = fila[2].value, pagado_recursos_propios= True, fecha_registro = fecha_hoy, fecha_pago = fecha_hoy)
					gestion_op.save()

			try:
				object_factura = Factura.objects.get(estado_id = estadoFactura.activa, contrato__contratista__codigo_acreedor= fila[0].value,referencia = fila[1].value, valor_contable = float(fila[3].value), numero = fila[4].value)
				object_factura.codigo_op_id = gestion_op.id
				object_factura.estado_id = estadoFactura.compensada
				object_factura.pagada = True
				object_factura.save()

			except Factura.DoesNotExist:
				sw = 1
				fila_lista.append(i)
				transaction.savepoint_rollback(sid)
			transaction.savepoint_commit(sid)	
							
			i+=1	
				
		
		respuesta  = JsonResponse({'message': 'El registro ha sido guardado exitosamente','success':'ok','data': ''}) if sw == 0 else JsonResponse({'message':'Las siguientes filas del documento excel no fueron actualizadas: <br>'+','.join(str(e) for e in fila_lista),'success':'fail','data':''});
		return respuesta

	except Exception as e:
		functions.toLog(e,'seguimiento_factura.actualizar_codigo_compensacion')
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# valida si el codigo de compensacion esta asignado a diferentes acreedor
def validar_cod_compensacion(repetidos, codigo_compensacion, array_verificar):

	sw = 0

	for c in repetidos:

		m = [i for i,x in enumerate(codigo_compensacion) if x==c]
		# print m
		lista_verifica = []
		for i in m:
			# print i
			# print array_verificar[i]
			if not lista_verifica:
				lista_verifica.append(array_verificar[i][0])
				# print array_verificar[i][0]
			else:
				if not array_verificar[i][0] in lista_verifica: 
					# print "mendoza"
					sw = 1


	return sw


def exportar_facturas_recursos_propios(request):
	try:
		
		# cantidadFacturaPorContabilizar = Factura.objects.filter(estado_id = estadoFactura.activa, referencia__exact = '', fecha__gte = '2018-10-31').count()
		qset = (Q(estado_id=estadoFactura.activa) & Q(fecha__gte = '2018-10-31') )
		qset = qset &(Q(referencia__exact='') | Q(referencia__exact='0'))

		id_empresa = request.user.usuario.empresa.id
		orden_pago= request.GET['orden_pago'] if 'orden_pago' in request.GET else None;
		bloqueo_factura= request.GET['bloqueo_factura'] if 'bloqueo_factura' in request.GET else None;
		pagada= request.GET['pagada'] if 'pagada' in request.GET else None;
		recursos_propios= request.GET['recursos_propios'] if 'recursos_propios' in request.GET else None;
		dato= request.GET['dato'] if 'dato' in request.GET else None;

		qset = (Q(codigo_op_id__isnull = True) & Q(fecha__gte = '2018-10-31'))

		if dato:
			qset = qset &( Q(contrato__contratista__nombre__icontains=dato) | Q(contrato__nombre__icontains=dato) )

		if orden_pago:
			qset = qset &( Q(orden_pago = orden_pago)	)
		
		if bloqueo_factura:
			qset = qset &( Q(bloqueo_factura = bloqueo_factura) )

		if recursos_propios:
			qset = qset &(Q(recursos_propios = recursos_propios))

		if id_empresa:
			qset = qset &(Q(contrato__empresacontrato__empresa=id_empresa) & Q(contrato__empresacontrato__participa=1) & Q(contrato__activo=1))

		if pagada:
			qset = qset &(Q(pagada = pagada))

		queryset = Factura.objects.filter(qset).exclude(estado_id__in = [estadoFactura.compensada]).exclude(referencia__exact = '').exclude(referencia__exact = '0')

		#   # INICIO - Crear el Excel
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] ='attachment; filename="Facturas por compensar.xlsx"'
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})

		worksheet = workbook.add_worksheet('Facturas por contabilizar')

		worksheet.set_column('A:A', 25)
		worksheet.set_column('B:B', 15)
		worksheet.set_column('C:C', 25)
		worksheet.set_column('D:D', 20)
		worksheet.set_column('E:E', 20)

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
		format1.set_align('center')
		format1.set_align('vcenter')
		format2=workbook.add_format({'border':0})
		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')
		format_money=workbook.add_format({'border':False,
			'font_size':11,
			'bold':False,
			'valign':'vright',
			'num_format': '$#,##0'})

		row=1
		col=0

		worksheet.write('A1', 'Codigo acreedor', format1)
		worksheet.write('B1', 'codigo SAP', format1)
		worksheet.write('C1', 'Codigo de compensacion', format1)
		worksheet.write('D1', 'Valor contable', format1)
		worksheet.write('E1', 'numero de factura', format1)

		for objecto in queryset:

			meses = ''
			fecha = ''

			worksheet.write(row, col  , str(objecto.contrato.contratista.codigo_acreedor) ,format2)
			worksheet.write(row, col+1,str(objecto.referencia) ,format2)
			worksheet.write(row, col+2, '' ,format2)
			worksheet.write(row, col+3, float(objecto.valor_contable) ,format2)
			worksheet.write(row, col+4, str(objecto.numero ),format2)
			
			row +=1

		workbook.close()# FIN - Crear el Excel		
		return response
	except Exception as e:
		functions.toLog(e,'seguimiento_factura.exportar_facturas_recursos_propios')
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@transaction.atomic
def actualizar_pago_factura(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:	

			
			lista = request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']
			fecha_pago = respuesta['fecha_pago']
			# print myList
			facturas = Factura.objects.filter(id__in = myList)
			facturas.update(pagada = True)

			for item in facturas:

				try:
					gestion = GestionOp.objects.get(pk = item.codigo_op_id)
					gestion.fecha_pago = fecha_pago
					gestion.save()
				except GestionOp.DoesNotExist:
					pass				

			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok','data': ''})
	
		except Exception as e:
			print(e)
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@transaction.atomic
def anular_pago_compensacion(request):
	if request.method == 'POST':
		sid = transaction.savepoint()
		try:
			
			lista = request.POST['_content']
			respuesta= json.loads(lista)
			gestion_op_id = respuesta['gestion_op_id']

			# Factura.objects.filter(gestion_op_id = gestion_op_id)
			facturas = Factura.objects.filter(codigo_op_id = gestion_op_id)
			facturas_id = ','.join([str(factura.id) for factura in facturas])
			facturas.update(codigo_op_id = None, pagada = False, estado_id = estadoFactura.activa)
			gestionCompensacion = GestionOp.objects.get(pk = gestion_op_id)
			gestionCompensacion.anulado = True
			gestionCompensacion.facturas_anuladas = facturas_id
			gestionCompensacion.save()
			transaction.savepoint_commit(sid)
			return JsonResponse({'message':'El registro se ha actualizado correctamente','success':'ok','data': ''})
		except Exception as e:
			print(e)
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)