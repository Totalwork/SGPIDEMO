from django.shortcuts import render, render_to_response
from django.urls import reverse

from rest_framework import viewsets, serializers, response
from django.db.models import Q

from django.db import IntegrityError,transaction
from django.http import HttpResponse,JsonResponse
from .models import Servidumbre_expediente, Servidumbre_grupo_documento, Servidumbre_grupo_documento, Servidumbre_documento, Servidumbre_persona, Servidumbre_predio, Servidumbre_predio_documento, Servidumbre_predio_georeferencia

from .serializers import Servidumbre_grupo_documentoSerializer
from .serializers import Servidumbre_expedienteSerializer
from .serializers import Servidumbre_documentoSerializer
from .serializers import Servidumbre_personaSerializer, Servidumbre_personaLiteSerializer
from .serializers import Servidumbre_predioSerializer
from .serializers import Servidumbre_predio_documentoSerializer
from .serializers import Servidumbre_grupo_documentoLiteSerializer
from .serializers import ExpedienteLiteEditarSerializer
from .serializers import Servidumbre_predio_georeferenciaSerializer
from .serializers import Servidumbre_predio_georeferenciaWriteSerializer
from .serializers import Servidumbre_predioGeoSerializer
from django.db.models import Q

from proyecto.models import Proyecto, Proyecto_empresas
from proyecto.views import ProyectoSerializer

from .enum import enumEstados,enumTipo

from usuario.models import Usuario
from usuario.views import *

from tipo.models import Tipo
from tipo.views import *

from estado.models import Estado
from estado.views import *


from django.template import RequestContext
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework import status
from rest_framework.pagination import PageNumberPagination

from rest_framework.decorators import api_view

from logs.models import Logs, Acciones
from coasmedas.functions import functions
from django.db import transaction

import xlsxwriter
import json
import openpyxl

@login_required
def expediente(request):
	return render(request, 'servidumbre/gestion_servidumbres.html',{'model':'servidumbre_expediente','app':'servidumbre'})

@login_required
def configuracion(request):
	return render(request, 'servidumbre/configuracion.html',{'model':'servidumbre_documento','app':'servidumbre'})
	

@login_required
def nuevo_expediente(request):
	return render(request, 'servidumbre/nuevo_expediente.html',{'model':'servidumbre_expediente','app':'servidumbre'})

@login_required
def expedientes(request,id):
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/expediente.html',
		{'id':id,'expediente':expediente,'model':'servidumbre_expediente','app':'servidumbre'},
		context_instance=RequestContext(request))
	
@login_required
def editar_expediente (request,id):
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/editar_expediente.html',
		{'id':id,'expediente':expediente,'model':'servidumbre_expediente','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def predios (request,id):
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/predios.html',
		{'id':id,'estado':int(expediente.estado.codigo),'expediente':expediente,'model':'servidumbre_predio','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def predio (request,id,pk):
	predio = Servidumbre_predio.objects.get(id=pk)
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/predio.html',
		{'id':id,'expediente':expediente, 'predio':predio,'model':'servidumbre_predio','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def editar_predio (request, id, pk):
	predio = Servidumbre_predio.objects.get(id=pk)
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/editar_predio.html',
		{'id':id,'expediente':expediente, 'predio':predio,'model':'servidumbre_predio','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def documentos (request, id, pk):
	predio = Servidumbre_predio.objects.get(id=pk)
	expediente = Servidumbre_expediente.objects.get(id=id)

	return render(request, 'servidumbre/documentos.html',
		{'expediente':expediente, 'estado':int(expediente.estado.codigo), 'predio':predio,'model':'Servidumbre_predio_documento','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def nuevo_predio(request, id):
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/nuevo_predio.html',
		{
		'expediente':id,
		'municipio':expediente.proyecto.municipio.nombre,
		'departamento':expediente.proyecto.municipio.departamento.nombre,
		'proyecto':expediente.proyecto.nombre,
		'mcontrato':expediente.proyecto.mcontrato.nombre,
		'model':'servidumbre_predio',
		'app':'servidumbre',
		},
		context_instance=RequestContext(request)
	)
		

@login_required
def expediente_georeferencias(request,id):
	expediente = Servidumbre_expediente.objects.get(id=id)
	return render(request, 'servidumbre/expediente_georeferencias.html',
		{'id':id,'expediente':expediente,'estado':int(expediente.estado.codigo),'model':'servidumbre_predio_georeferencia','app':'servidumbre'},
		context_instance=RequestContext(request))


@login_required
def predio_georeferencias (request, id, pk):
	#import pdb; pdb.set_trace()
	predio = Servidumbre_predio.objects.get(id=pk)
	expediente = Servidumbre_expediente.objects.get(id=id)

	return render(request, 'servidumbre/predio_georeferencias.html',
		{'expediente':expediente, 'estado':int(expediente.estado.codigo), 'predio':predio,'model':'Servidumbre_predio_documento','app':'servidumbre'},
		context_instance=RequestContext(request))

@login_required
def graficas(request):
	return render(request, 'servidumbre/graficas_servidumbre.html',{'model':'servidumbre_expediente','app':'servidumbre'})



# Serializer de proyecto
class ProyectoLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta: 
		model = Proyecto
		fields=( 'id','nombre')
# Serializer de	usuario
class UsuarioLiteSerializer(serializers.HyperlinkedModelSerializer):    

    persona = PersonaLiteSerializer(read_only=True)

    class Meta:
        model = Usuario
        fields=('id','persona') 





#Api rest para Servidumbre_expediente
##Metodos
class Servidumbre_expedienteViewSet(viewsets.ModelViewSet):
	model=Servidumbre_expediente
	queryset = model.objects.all()
	serializer_class = Servidumbre_expedienteSerializer
	nombre_modulo = 'Servidumbre - Servidumbre_expedienteViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()

			lite_editar = self.request.query_params.get('lite_editar',None)

			if lite_editar:
				serializer = ExpedienteLiteEditarSerializer(instance, context={'request': request})
			else:
				serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_expedienteViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			page = self.request.query_params.get('page', None)			
			ID = self.request.query_params.get('id', None)

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(proyecto__nombre__icontains=dato) | 
							  (Q(estado__nombre__icontains=dato) | 
							  (Q(usuario_creador__persona__nombres__icontains=dato) |
							  (Q(usuario_creador__persona__apellidos__icontains=dato)
						  	) ) ) )
			if ID:
				qset = qset & Q(id=ID)

			# qsetc = (Q(edita=1)) &(Q(empresa=request.user.usuario.empresa.id) & Q(participa=1))
			# ListContratosValidos = EmpresaContrato.objects.filter(qsetc).values('contrato_id').distinct().order_by("contrato_id")
						
			# qset = qset & (Q(proyecto__mcontrato__id__in=ListContratosValidos))

			ListProyectosValidos = Proyecto_empresas.objects.filter(empresa__id=request.user.usuario.empresa.id).values('proyecto__id').distinct()
			qset = qset & (Q(proyecto__id__in=ListProyectosValidos))

			queryset = self.model.objects.filter(qset).order_by('-id')

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_expedienteSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(proyecto_id=request.DATA['proyecto_id'],
					 estado_id=Estado.objects.get(app='servidumbre_expediente',codigo=159).id, 
					 usuario_creador_id=request.user.usuario.id)
					
					logs_model=Logs(usuario_id=request.user.usuario.id,
						accion=Acciones.accion_crear,
						nombre_modelo='servidumbre.Servidumbre_expediente',
						id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':	
			sid = transaction.savepoint()			
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_expediente',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)
				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = Servidumbre_expedienteSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					
					serializer.save(proyecto_id=request.DATA['proyecto_id'])
					
					

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_expediente',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)











#Api rest para Servidumbre_grupo_documento
##Metodos
class Servidumbre_grupo_documentoViewSet(viewsets.ModelViewSet):
	model=Servidumbre_grupo_documento
	queryset = model.objects.all()
	serializer_class = Servidumbre_grupo_documentoSerializer
	nombre_modulo = 'Servidumbre - Servidumbre_grupo_documentoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_grupo_documentoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			page = self.request.query_params.get('page', None)			
			ID = self.request.query_params.get('id', None)

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(nombre__icontains=dato))	
			if ID:
				qset = qset & Q(id=ID)

			queryset = self.model.objects.filter(qset)

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_grupo_documentoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_grupo_documento',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':		
			sid = transaction.savepoint()		
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_grupo_documento',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)
				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer  = Servidumbre_grupo_documentoSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					
					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_grupo_documento',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)


    

#Api rest para Servidumbre_documento
##Metodos

class Servidumbre_documentoViewSet(viewsets.ModelViewSet):
	model=Servidumbre_documento
	queryset = model.objects.all()
	serializer_class = Servidumbre_documentoSerializer
	nombre_modulo = 'Servidumbre - Servidumbre_documentoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_documentoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			page = self.request.query_params.get('page', None)	
			grupo_documento_id = request.GET['grupo_documento_id'] if 'grupo_documento_id' in request.GET else 0;		
			ID = self.request.query_params.get('id', None)
			

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(nombre__icontains=dato) | 
							  (Q(grupo_documento__nombre__icontains=dato)  
							  ) )	
			if grupo_documento_id:
				qset = qset & (Q(grupo_documento__id=grupo_documento_id))
			if ID:
				qset = qset & Q(id=ID)

			

			queryset = self.model.objects.filter(qset)

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_documentoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(grupo_documento_id=request.DATA['grupo_documento_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_documento',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':
			sid = transaction.savepoint()				
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_documento',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)
				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer  = Servidumbre_documentoSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					
					serializer.save(grupo_documento_id=request.DATA['grupo_documento_id'])
					

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_documento',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)






#Api rest para Servidumbre_persona
##Metodos

class Servidumbre_personaViewSet(viewsets.ModelViewSet):
	model=Servidumbre_persona
	queryset = model.objects.all()
	serializer_class = Servidumbre_personaSerializer
	nombre_modulo = 'Servidumbre - Servidumbre_personaViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_personaViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			page = self.request.query_params.get('page', None)			
			ID = self.request.query_params.get('id', None)
			lite = self.request.query_params.get('lite', None)

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(nombres__icontains=dato) | 
							  (Q(apellidos__icontains=dato) | 
							  (Q(cedula__icontains=dato) ) ) )	
			if ID:
				qset = qset & Q(id=ID)

			queryset = self.model.objects.filter(qset)

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					if lite:
						serializer = Servidumbre_personaLiteSerializer(queryset,many=True)	
					else:
						serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			if lite:
				serializer = Servidumbre_personaLiteSerializer(queryset,many=True)
			else:
				serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_personaSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_persona',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					msg='Se presentaron los siguientes errores: '
					for k,v in serializer.errors.items():
						msg = msg +("%s : %s" %(k,v[0]))
					return Response({'message':msg,'success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':	
			sid = transaction.savepoint()			
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_persona',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)
				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = Servidumbre_personaSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					
					serializer.save()
					

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_persona',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)






#Api rest para Servidumbre_predio
##Metodos

class Servidumbre_predioViewSet(viewsets.ModelViewSet):
	model=Servidumbre_predio
	queryset = model.objects.all()
	serializer_class = Servidumbre_predioSerializer
	nombre_modulo = 'Servidumbre - Servidumbre_predioViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()

			
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_predioViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			page = self.request.query_params.get('page', None)	
			expediente = self.request.query_params.get('expediente',None)		
			ID = self.request.query_params.get('id', None)

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(expediente__proyecto__nombre__icontains=dato) | 
							  (Q(persona__nombres__icontains=dato) | 
							  (Q(persona__apellidos__icontains=dato) |
							  (Q(nombre_direccion__icontains=dato) 	|
							  (Q(grupo_documento__nombre__icontains=dato) 	|
							  (Q(tipo__nombre__icontains=dato) 	
								)	) ) ) )	)
			if expediente:
				qset = qset & Q(expediente__id=expediente)
					
			if ID:
				qset = qset & Q(id=ID)

			queryset = self.model.objects.filter(qset).order_by('-id')



			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				#import pdb; pdb.set_trace()
				serializer = Servidumbre_predioSerializer(data=request.DATA,context={'request': request})



				if serializer.is_valid():
					serializer.save(expediente_id=request.DATA['expediente_id'], persona_id=request.DATA['persona_id'],	tipo_id=request.DATA['tipo_id'], grupo_documento_id=request.DATA['grupo_documento_id'])
					
					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_predio',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
    
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':		
			sid = transaction.savepoint()	
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_predio',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)
				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_204_NO_CONTENT)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = Servidumbre_predioSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					
					serializer.save(expediente_id=request.DATA['expediente_id'], persona_id=request.DATA['persona_id'],	tipo_id=request.DATA['tipo_id'], grupo_documento_id=request.DATA['grupo_documento_id'])
										

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_predio',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)



#Api rest para Servidumbre_predio_documento
##Metodos

class Servidumbre_predio_documentoViewSet(viewsets.ModelViewSet):

	model=Servidumbre_predio_documento
	queryset = model.objects.all()
	serializer_class = Servidumbre_predio_documentoSerializer
	parser_classes=(FormParser, MultiPartParser,)
	nombre_modulo = 'Servidumbre - Servidumbre_predio_documentoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_predio_documentoViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			predio_id = self.request.query_params.get('predio_id',None)
			documento_id = self.request.query_params.get('documento_id',None)
			page = self.request.query_params.get('page', None)			
			ID = self.request.query_params.get('id', None)

			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(predio__nombre_direccion__icontains=dato) | 
							  (Q(documento__nombre__icontains=dato) | 
							  (Q(archivo__icontains=dato) ) ) )	
			if predio_id:
				qset = qset & Q(predio__id=predio_id)
			if documento_id:
				qset = qset & Q(documento__id=documento_id)
			if ID:
				qset = qset & Q(id=ID)

			queryset = self.model.objects.filter(qset)

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})

			serializer = self.get_serializer(queryset,many=True)
			return Response({'message':'','success':'ok',
				'data':serializer.data})
		except Exception as e:
			print(e)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

		except Exception as e:
			#print (e)
			respuesta=Estructura.error500()
			return Response(respuesta, status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_predio_documentoSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					serializer.save(predio_id=request.DATA['predio_id'], 
						documento_id=request.DATA['documento_id'],
						archivo=self.request.FILES.get('archivo')
						)
					
					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':		
			sid = transaction.savepoint()	
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)

				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_200_OK)
			except Exception as e:
				transaction.savepoint_rollback(sid)
				functions.toLog(e,self.nombre_modulo)
				return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			    'data':''},status=status.HTTP_400_BAD_REQUEST)
	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = Servidumbre_predio_documentoSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					valores=Servidumbre_predio_documento.objects.get(id=instance.id)
					if self.request.FILES.get('archivo') is not None:
						serializer.save(predio_id=request.DATA['predio_id'], 
							documento_id=request.DATA['documento_id'],
							archivo=self.request.FILES.get('archivo')
							)

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)



class Servidumbre_georeferenciaViewSet(viewsets.ModelViewSet):

	model=Servidumbre_predio_georeferencia
	queryset = model.objects.all()
	serializer_class = Servidumbre_predio_georeferenciaSerializer
	parser_classes=(FormParser, MultiPartParser,)
	nombre_modulo = 'Servidumbre - Servidumbre_predio_documentoViewSet'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	@transaction.atomic
	def list(self, request, *args, **kwargs):
		try:
			queryset = super(Servidumbre_georeferenciaViewSet, self).get_queryset()
			predio_id = self.request.query_params.get('predio_id',None)
			expediente_id = self.request.query_params.get('expediente_id',None)
			dato = self.request.query_params.get('dato',None)
			ID = self.request.query_params.get('ID',None)

			#import pdb; pdb.set_trace()
			qset = (~Q(id = 0))
			if dato:
				qset = qset & (Q(predio__nombre_direccion__icontains=dato) )
				
			if expediente_id:
				qset = qset & (Q(predio__expediente__id=int(expediente_id)) )

			if predio_id:
				qset = qset & Q(predio__id=int(predio_id))

			if ID:
				qset = qset & Q(id=int(ID))

			queryset = self.model.objects.filter(qset).order_by('orden')

			ignorePagination= self.request.query_params.get('ignorePagination',None)
			if ignorePagination is None:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
			else:
				serializer = self.get_serializer(queryset,many=True)				
				return Response({'message':'','success':'ok','data':serializer.data})

		except Exception as e:
				functions.toLog(e,self.nombre_modulo)
				transaction.savepoint_rollback(sid)
				return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)


	@transaction.atomic
	def create(self, request, *args, **kwargs):
		if request.method == 'POST':
			#import pdb; pdb.set_trace()
			sid = transaction.savepoint()
			try:
				serializer = Servidumbre_predio_georeferenciaWriteSerializer(data=request.DATA,context={'request': request})

				if serializer.is_valid():
					last_model= Servidumbre_predio_georeferencia.objects.filter(predio__id=int(request.DATA['predio_id'])).order_by('orden').last()
					if last_model:
						serializer.save(predio_id=request.DATA['predio_id'],orden=int(last_model.orden)+1)
					else:
						serializer.save(predio_id=request.DATA['predio_id'],orden=1)
					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=serializer.data['id'])
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
						
				else:
					print(serializer.errors)
					return Response({'message':'Datos requeridos no fueron recibidos','success':'fail','data':''},status=status.HTTP_400_BAD_REQUEST)
			
			except Exception as e:
					functions.toLog(e,self.nombre_modulo)
					transaction.savepoint_rollback(sid)
					return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def update(self,request,*args,**kwargs):
		if request.method == 'PUT':
			sid = transaction.savepoint()
			try:
				# partial = kwargs.pop('partial', False)
				instance = self.get_object()
				serializer = Servidumbre_predio_georeferenciaSerializer(instance,data=request.DATA,context={'request': request})
				if serializer.is_valid():
					serializer.save(predio_id=request.DATA['predio_id'])

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=instance.id)
					logs_model.save()

					transaction.savepoint_commit(sid)
					return Response({'message':'El registro ha sido actualizado exitosamente','success':'ok','data':serializer.data},status=status.HTTP_201_CREATED)
				
				else:
					if serializer.errors['non_field_errors']:
						mensaje = serializer.errors['non_field_errors']
					else:
						mensaje='datos requeridos no fueron recibidos'
					return Response({'message':mensaje,'success':'fail', 'data':''},status=status.HTTP_400_BAD_REQUEST)
			except Exception as e:
					functions.toLog(e,self.nombre_modulo)
					transaction.savepoint_rollback(sid)
					return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

	@transaction.atomic
	def destroy(self,request,*args,**kwargs):
		if request.method == 'DELETE':		
			sid = transaction.savepoint()	
			try:
				instance = self.get_object()
				self.perform_destroy(instance)
				serializer = self.get_serializer(instance)
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_borrar,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=instance.id)
				logs_model.save()

				transaction.savepoint_commit(sid)

				return Response({'message':'El registro se ha eliminado correctamente','success':'ok',
				'data':''},status=status.HTTP_200_OK)
			except Exception as e:
					functions.toLog(e,self.nombre_modulo)
					transaction.savepoint_rollback(sid)
					return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

@transaction.atomic
def cerrar_expedientes(request):

	sid = transaction.savepoint()

	try:
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		valida_registro_abierto = 0
		insert_list_expedientes = []

		for item in respuesta['lista']:
			object_detalle=Servidumbre_expediente.objects.get(pk=item['id'])
			estado=object_detalle.estado_id


			if estado == enumEstados.abierto:

				#object_detalle.estado_id=7
				object_detalle.estado_id = enumEstados.cerrado
				object_detalle.save()

				insert_list_expedientes.append(Logs(usuario_id=request.user.usuario.id,
					accion=Acciones.accion_actualizar,
					nombre_modelo='servidumbre.expediente',
					id_manipulado=item['id']))

				
				transaction.savepoint_commit(sid)

			elif valida_registro_abierto==0 :
				valida_registro_abierto = 1


		# se hace una sola transaccion con bulk create
		if insert_list_expedientes:
			Logs.objects.bulk_create(insert_list_expedientes)

		

		transaction.savepoint_commit(sid)
		if valida_registro_abierto==1 :

			return JsonResponse({'message':'Solo se pueden cerrar expedientes en estado abierto','success':'error',
					'data':''})
		else:

			return JsonResponse({'message':'Registro(s) cerrado(s) correctamente','success':'ok',
					'data':''})
		
	except Exception as e:
		#print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@transaction.atomic
def reabrir(request,pk):

		sid = transaction.savepoint()	
		try:
			instance = Servidumbre_expediente.objects.get(pk=pk)	
			estado = instance.estado_id
			valida_registro = 0

			if estado == enumEstados.cerrado:

				instance.estado_id=enumEstados.abierto
				instance.save()

				logs_model=Logs(usuario_id=request.user.usuario.id,
					accion=Acciones.accion_actualizar,
					nombre_modelo='servidumbre.expediente',
					id_manipulado=instance.id)
				logs_model.save()

			elif valida_registro==0 :
				valida_registro = 1



				transaction.savepoint_commit(sid)

			if valida_registro==1 :
				return JsonResponse({'message':'Solo se puede reabrir un expediente en estado cerrado','success':'error',
						'data':''})

			else:
				return JsonResponse({'message':'El expediente se ha reabrierto correctamente','success':'ok',
						'data':''})

		except Exception as e:
			#print(e)
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})


@transaction.atomic
def cerrar(request,pk):
	
		sid = transaction.savepoint()	
		try:
			instance = Servidumbre_expediente.objects.get(pk=pk)
			estado = instance.estado_id
			valida_registro = 0

			if estado == enumEstados.abierto:

				instance.estado_id=enumEstados.cerrado
				instance.save()

				logs_model=Logs(usuario_id=request.user.usuario.id,
					accion=Acciones.accion_actualizar,
					nombre_modelo='servidumbre.expediente',
					id_manipulado=instance.id)
				logs_model.save()

			elif valida_registro==0 :
				valida_registro = 1


				transaction.savepoint_commit(sid)

			if valida_registro==1 :
				return JsonResponse({'message':'Solo se puede cerrar un expediente en estado abierto','success':'error',
						'data':''})

			else:
				return JsonResponse({'message':'El expediente se ha cerrado correctamente','success':'ok',
						'data':''})

		except Exception as e:
			#print(e)
			transaction.savepoint_rollback(sid)
			return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
				'data':''})


@transaction.atomic
def eliminar_grupos(request):
	if request.method == 'POST':
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']
			
			try:
				sid = transaction.savepoint()
				for item in myList:
					
					
					Servidumbre_grupo_documento.objects.get(id = item['id'] ).delete()


					logs_model=Logs(usuario_id=request.user.usuario.id,
						accion=Acciones.accion_borrar,
						nombre_modelo='servidumbre.Servidumbre_grupo_documento',
						id_manipulado=item['id'])
					logs_model.save()


				transaction.savepoint_commit(sid)
				return JsonResponse({'message':'Registro(s) eliminado(s) correctamente','success':'ok','data': ''})
			except Exception as e:
				transaction.savepoint_rollback(sid)	
				return JsonResponse({'message':'Algunos de los grupos seleccionados estan asociados a datos tecnicos. ','status':'error','data':''})							

		except Exception as e:
			
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@transaction.atomic
def eliminar_documentos(request):
	if request.method == 'POST':
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']
			
			try:
				sid = transaction.savepoint()
				for item in myList:

				
					Servidumbre_documento.objects.get(id = item['id'] ).delete()		

					logs_model=Logs(usuario_id=request.user.usuario.id,
						accion=Acciones.accion_borrar,
						nombre_modelo='servidumbre.Servidumbre_documento',
						id_manipulado=item['id'])
					logs_model.save()		

				transaction.savepoint_commit(sid)
				return JsonResponse({'message':'Registro(s) eliminado(s) correctamente','success':'ok','data': ''})
			except Exception as e:
				transaction.savepoint_rollback(sid)	
				return JsonResponse({'message':'Algunos de los documentos seleccionados estan asociados a datos tecnicos. ','status':'error','data':''})							

		except Exception as e:
			
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@transaction.atomic
def eliminar_predios(request):
	if request.method == 'POST':
		try:
			lista=request.POST['_content']
			respuesta= json.loads(lista)
			myList = respuesta['lista']
			
			try:
				sid = transaction.savepoint()
				for item in myList:

				
					Servidumbre_predio.objects.get(id = item['id'] ).delete()		

					logs_model=Logs(usuario_id=request.user.usuario.id,
						accion=Acciones.accion_borrar,
						nombre_modelo='servidumbre.Servidumbre_predio',
						id_manipulado=item['id'])
					logs_model.save()		

				transaction.savepoint_commit(sid)
				return JsonResponse({'message':'Registro(s) eliminado(s) correctamente','success':'ok','data': ''})
			except Exception as e:
				transaction.savepoint_rollback(sid)	
				return JsonResponse({'message':'Algunos de los predios seleccionados estan asociados a datos tecnicos. ','status':'error','data':''})							

		except Exception as e:
			
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)




def exportReporteExpedientes(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Expedientes.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Expedientes')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':1})
	format_date=workbook.add_format({'border':1,'num_format': 'yyyy-mm-dd'})

	row=1
	col=0

	cursor = connection.cursor()

					
	ServidumbreExpediente = Servidumbre_expediente.objects.filter()

	worksheet.write('A1', 'Numero', format1)
	worksheet.write('B1', 'Macro contrato', format1)
	worksheet.write('C1', 'Departamento', format1)
	worksheet.write('D1', 'Municipio', format1)
	worksheet.write('E1', 'Proyecto - Servicio', format1)
	worksheet.write('F1', 'Usuario creador', format1)
	worksheet.write('G1', 'Fecha creacion', format1)
	worksheet.write('H1', 'Estado', format1)

	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 25)
	worksheet.set_column('C:C', 25)
	worksheet.set_column('D:D', 25)
	worksheet.set_column('E:E', 40)
	worksheet.set_column('F:F', 25)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)


	for Ser in ServidumbreExpediente:

			
		worksheet.write(row, col,Ser.id,format2)
		worksheet.write(row, col+1,Ser.proyecto.mcontrato.nombre,format2)
		worksheet.write(row, col+2,Ser.proyecto.municipio.departamento.nombre,format2)
		worksheet.write(row, col+3,Ser.proyecto.municipio.nombre,format2)
		worksheet.write(row, col+4,Ser.proyecto.nombre,format2)
		worksheet.write(row, col+5,Ser.usuario_creador.persona.nombres + ' ' + Ser.usuario_creador.persona.apellidos,format2)
		worksheet.write(row, col+6,Ser.fecha_creacion,format_date)
		worksheet.write(row, col+7,Ser.estado.nombre,format2)
			

		row +=1


	workbook.close()

	return response
    
def exportReportePredios(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
	response['Content-Disposition'] = 'attachment; filename="Predios.xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Predios')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':1})
	number_format = workbook.add_format({'num_format': '#,##0.00','border':1})
	format_date=workbook.add_format({'border':1,'num_format': 'yyyy-mm-dd'})

	row=1
	col=0

	cursor = connection.cursor()

	ServidumbrePredio = Servidumbre_predio.objects.filter()

	worksheet.write('A1', 'No. Expediente', format1)
	worksheet.write('B1', 'Proyecto', format1)
	worksheet.write('C1', 'Macro contrato', format1)
	worksheet.write('D1', 'Departamento', format1)
	worksheet.write('E1', 'Municipio', format1)	
	worksheet.write('F1', 'Nombre del predio', format1)
	worksheet.write('G1', 'Propietario', format1)
	worksheet.write('H1', 'Tipo predio', format1)
	worksheet.write('I1', 'Grupo documentos', format1)
	worksheet.write('J1', '% documentos', format1)


	worksheet.set_column('A:A', 15)
	worksheet.set_column('B:B', 25)
	worksheet.set_column('C:C', 25)
	worksheet.set_column('D:D', 25)
	worksheet.set_column('E:E', 40)
	worksheet.set_column('F:F', 25)
	worksheet.set_column('G:G', 20)
	worksheet.set_column('H:H', 20)
	worksheet.set_column('I:I', 20)
	worksheet.set_column('J:J', 20)


	for Ser in ServidumbrePredio:

		total_documentos = float(Servidumbre_documento.objects.filter(
					grupo_documento__id=Ser.grupo_documento.id).count())
		documentos_cargados = float (Servidumbre_predio_documento.objects.filter(
					predio__id=Ser.id).count())
		porcentaje = round((documentos_cargados / total_documentos)*100,2)


			
		worksheet.write(row, col,Ser.expediente.id,format2)
		worksheet.write(row, col+1,Ser.expediente.proyecto.nombre,format2)
		worksheet.write(row, col+2,Ser.expediente.proyecto.mcontrato.nombre,format2)
		worksheet.write(row, col+3,Ser.expediente.proyecto.municipio.departamento.nombre,format2)
		worksheet.write(row, col+4,Ser.expediente.proyecto.municipio.nombre,format2)
		worksheet.write(row, col+5,Ser.nombre_direccion,format2)
		worksheet.write(row, col+6,Ser.persona.nombres + ' ' + Ser.persona.apellidos,format2)
		worksheet.write(row, col+7,Ser.tipo.nombre,format_date)
		worksheet.write(row, col+8,Ser.grupo_documento.nombre,format2)
		worksheet.write(row, col+9,porcentaje,number_format)
			

		row +=1


	workbook.close()

	return response



@api_view(['GET'])
def select_create_update_predio(request):
	if request.method == 'GET':
		try:
			qsetTipos = Tipo.objects.filter(app='Servidumbre_predio')
			tipo = TipoLiteSerializer(qsetTipos,many=True).data

			qsetGrupoDocumentos = Servidumbre_grupo_documento.objects.all()
			grupo_documento = Servidumbre_grupo_documentoLiteSerializer(qsetGrupoDocumentos,many=True).data

			return JsonResponse({'message':'','success':'ok','data':{
										'tipo_id' : tipo,
										'grupo_documento_id' : grupo_documento
									 }})


		except Exception as e:
			return JsonResponse(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'status':'error',
				'data':''
				},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def documentos_predio(request):
	if request.method == 'GET':
		try:
			id = request.query_params.get('id', None)
			if id:
				predio = Servidumbre_predio.objects.get(pk=id)
				# definir la variable porcentaje:
				total_documentos = float(Servidumbre_documento.objects.filter(
					grupo_documento__id=predio.grupo_documento.id).count())

				auxiliar =0
				count = 0

				Servidumbredocumento = Servidumbre_documento.objects.filter(grupo_documento__id=predio.grupo_documento.id)

				for doc in Servidumbredocumento:
				
					auxiliar = float (Servidumbre_predio_documento.objects.filter(
					predio__id=predio.id, documento__id =doc.id ).count())

					if (auxiliar>0):
						count = count +1;	

				documentos_cargados = count

				porcentaje = round((documentos_cargados / total_documentos)*100,2)
				
				# definir las variables de proyecto, departamento, municipio y mcontrato
				# nombre_proyecto = predio.expediente.proyecto.nombre
				# departamento = predio.expediente.proyecto.municipio.departamento.nombre
				# municipio = predio.expediente.proyecto.municipio.nombre
				# mcontrato = predio.expediente.proyecto.mcontrato.nombre

				# definir el json de proyecto
				documentos = []
				docsDelGrupo=Servidumbre_documento.objects.filter(
					grupo_documento__id=predio.grupo_documento.id).values('id','nombre')

				for docDelGrupo in docsDelGrupo:
					cantidadDocumentos = Servidumbre_predio_documento.objects.filter(
						predio__id=predio.id,
						documento__id=docDelGrupo['id']
						).count()
					documentos.append({
						'id': docDelGrupo['id'],
						'nombre':docDelGrupo['nombre'],
						'cantidad': cantidadDocumentos
						})

				return JsonResponse({'message':'','success':'ok','data':{
											'porcentaje' : porcentaje,
											#'nombre_proyecto' : nombre_proyecto,
											#'departamento': departamento,
											#'municipio': municipio,
											#'mcontrato':mcontrato,
											'documentos':documentos,
											'predio_id':id
										 }
									})
			else:
				return JsonResponse({'message':'Debe indicar el predio a consultar','success':'ok','data':{}})	
		except Exception as e:
			return JsonResponse(
				{'message':'Se presentaron errores de comunicacion con el servidor',
				'status':'error',
				'data':''
				},
				status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
def VerSoporte(request):
	if request.method == 'GET':
		try:
						
			archivo = Servidumbre_predio_documento.objects.get(pk=request.GET['id'])
			
			return functions.exportarArchivoS3(str(archivo.archivo))
			
		except Exception as e:
			functions.toLog(e,'Servidumbre.VerSoporte')
			return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@transaction.atomic
@api_view(['POST'])
def guardar_archivo(request):

	sid = transaction.savepoint()

	try:
		predio = request.POST['predio_id']
		nombre = request.POST['nombre']
		archivo = request.FILES['archivo']
		documento = request.POST['documento_id']

		PredioDoc=Servidumbre_predio_documento(predio_id=predio,nombre=nombre,archivo=archivo,documento_id=documento)
		PredioDoc.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.Servidumbre_predio_documento',id_manipulado=PredioDoc.id)
		logs_model.save()

		transaction.savepoint_commit(sid)
		return Response({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''},status=status.HTTP_201_CREATED)
		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

@login_required
def descargar_plantilla_georeferencias(request):
	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')
		response['Content-Disposition'] = 'attachment; filename="Plantilla_georeferencias.xls"'

		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Puntos GPS')

		format1=workbook.add_format({'border':0,'font_size':12,'bold':True})
		format1.set_align('center')
		format1.set_align('vcenter')
		format2=workbook.add_format({'border':0})
		format3=workbook.add_format({'border':0,'font_size':12})
		format5=workbook.add_format()
		format5.set_num_format('yyyy-mm-dd')

		worksheet.set_column('A:A', 20)
		worksheet.set_column('B:B', 20)
		worksheet.set_column('C:H', 20)

		row=1
		col=0

		worksheet.write('A1', 'Orden', format1)
		worksheet.write('B1', 'Longitud', format1)
		worksheet.write('C1', 'Latitud', format1)

		worksheet.write('A2', 1, format1)
		worksheet.write('A3', 2, format1)
		worksheet.write('A4', 3, format1)
		worksheet.write('A5', 4, format1)

		workbook.close()
		return response

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return Response({'message':'Se presentaron errores al procesar los datos','success':'error','data':''},status=status.HTTP_400_BAD_REQUEST)

@login_required
def guardar_coordenadas_archivo(request):
	sid = transaction.savepoint()

	try:		
		soporte= request.FILES['archivo']
		predio_id= int(request.POST['predio_id'])
		doc = openpyxl.load_workbook(soporte,data_only = True)
		nombrehoja=doc.get_sheet_names()[0]
		hoja = doc.get_sheet_by_name(nombrehoja)
		i=0

		transaction.savepoint_commit(sid)
		revisar_detalle=Servidumbre_predio_georeferencia.objects.filter(predio__id=predio_id)


		if len(revisar_detalle) > 0:
			Servidumbre_predio_georeferencia.objects.filter(predio__id=predio_id).delete()


		contador=hoja.max_row - 1
		numeroFila = 1
		numeroColumna = 1
		numeroCelda = 1

		list_orden = []
		columnas=hoja.columns
		for columna in columnas:
			if numeroColumna==1:			

				for celdas in columna:
					if numeroCelda>1:
						try:
							list_orden.append(int(celdas.value))
						except Exception as e:
							return JsonResponse({
							'message':'El dato de Orden en la fila No.' + str(numeroCelda) +' no es un numero'+ \
							'. Sugerimos corregir la plantilla e ingresarla nuevamente',
							'success':'error',
							'data':''})
						
					numeroCelda+=1
			numeroColumna = numeroColumna + 1
		

		if int(contador) > 0:
			
			for fila in hoja.rows:
				if numeroFila > 1:
					
					if not fila[0].value :
						return JsonResponse({
							'message':'No se encontro el Orden en la fila No.' + str(numeroFila) + \
							'. Sugerimos corregir la plantilla e ingresarla nuevamente',
							'success':'error',
							'data':''})

					if not fila[1].value :
						return JsonResponse({
							'message':'No se encontro la Longitud en la fila No.' + str(numeroFila) + \
							'. Sugerimos corregir la plantilla e ingresarla nuevamente',
							'success':'error',
							'data':''})

					if not fila[2].value :
						return JsonResponse({
							'message':'No se encontro la Latitud en la fila No.' + str(numeroFila) + \
							'. Sugerimos corregir la plantilla e ingresarla nuevamente',
							'success':'error',
							'data':''})

					if list_orden.count(fila[0].value)>1:
						return JsonResponse({
							'message':'El numero ' + str(fila[0].value)+' se encuentra repetido en la columna de Orden' + \
							'. Sugerimos corregir la plantilla e ingresarla nuevamente',
							'success':'error',
							'data':''})

					try:
						aux = float(fila[1].value)
					except Exception as e:
						return JsonResponse({
						'message':'El dato de Longitud en la fila No.' + str(numeroCelda) +' no es un numero valido'+ \
						'. Sugerimos corregir la plantilla e ingresarla nuevamente',
						'success':'error',
						'data':''})

					try:
						aux = float(fila[2].value)
					except Exception as e:
						return JsonResponse({
						'message':'El dato de Latitud en la fila No.' + str(numeroCelda) +' no es un numero valido'+ \
						'. Sugerimos corregir la plantilla e ingresarla nuevamente',
						'success':'error',
						'data':''})
			

				numeroFila = numeroFila + 1	
									
			numeroFila = 1
			transaction.savepoint_commit(sid)

			#import pdb; pdb.set_trace()

			for fila in hoja.rows:
				if numeroFila>1:
					model_coordenada=Servidumbre_predio_georeferencia(
						predio_id=predio_id,
						orden=int(fila[0].value),
						longitud=str(fila[1].value),
						latitud=str(fila[2].value)
						)
					model_coordenada.save()

					logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='servidumbre.predio_georeferencias',id_manipulado=model_coordenada.id)
					logs_model.save()

				numeroFila = numeroFila + 1	

		
		
		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})

		
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'servidumbre.predio_georeferencias')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@api_view(['POST',])
@transaction.atomic
def guardar_cambio_cantidades(request):
	sid = transaction.savepoint()
	try:
		
		respuesta=request.data
		base_auxiliar = 99999
		transaction.savepoint_rollback(sid)
		list_orden = []
		for item in respuesta['lista']:			

			if item['orden']==0:
				return JsonResponse({
					'message':'El numero de Orden no puede ser "' + str(item['orden'])+'"' + \
					'. Sugerimos corregir la informacion e ingresarla nuevamente',
					'success':'error',
					'data':''})

			list_orden.append(int(item['orden']))

		for item in respuesta['lista']:
			#import pdb; pdb.set_trace()
			if list_orden.count(int(item['orden']))>1:
				return JsonResponse({
					'message':'El numero ' + str(item['orden'])+' se encuentra repetido en la columna de Orden' + \
					'. Sugerimos corregir la informacion e ingresarla nuevamente',
					'success':'error',
					'data':''})

			model_coordenada = Servidumbre_predio_georeferencia.objects.get(pk=item['id'])
			model_coordenada.orden=base_auxiliar
			model_coordenada.save()
			base_auxiliar=base_auxiliar-1

		
		for item in respuesta['lista']:
			model_coordenada_2 = Servidumbre_predio_georeferencia.objects.get(pk=item['id'])
			model_coordenada_2.orden=item['orden']
			model_coordenada_2.save()
				

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='servidumbre.predio_georeferencias',id_manipulado=model_coordenada.id)
			logs_model.save()


		return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
				'data':''})
	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		functions.toLog(e,'servidumbre.predio_georeferencias')
		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})


@api_view(['GET',])
def consultar_predios_coordenadas(request):
	try:
		#import pdb; pdb.set_trace()
		expediente_id = request.query_params.get('id', None)
		if expediente_id:
			predios = []

			predios_objs=Servidumbre_predio.objects.filter(expediente__id=expediente_id)
			predios_serializer= Servidumbre_predioGeoSerializer(predios_objs,many=True,context={'request': request})

			for predio in predios_serializer.data:
				predios.append(predio)

			return JsonResponse({'data':list(predios),'success':'ok','message':''})
		else:
			return JsonResponse({'data':{},'success':'ok','message':'Debe indicar el expediente a consultar'})
	
	except Exception as e:
		print(e)
		functions.toLog(e,'servidumbre.expediente_georeferencias')		
		return JsonResponse({'message':'Se presentaron errores al procesar la solicitud','success':'error',
			'data':''})

@api_view(['GET'])
def get_dataGraph(request):
	if request.method == 'GET':
		try:
			datos = []
			#cantidad de expedientes por estado
			
			ListProyectosValidos = Proyecto_empresas.objects.filter(empresa__id=request.user.usuario.empresa.id).values('proyecto__id').distinct()
			qset = (Q(proyecto__id__in=ListProyectosValidos))

			qsetabierto = qset & (Q(estado__codigo=159)) & (Q(estado__app='servidumbre_expediente'))
			qsetcerrado = qset & (Q(estado__codigo=160)) & (Q(estado__app='servidumbre_expediente'))
	
			
			datagrafica = []
			total_abierto = Servidumbre_expediente.objects.filter(qsetabierto).count()
			total_cerrado = Servidumbre_expediente.objects.filter(qsetcerrado).count()

			total = total_abierto + total_cerrado

			
			datagrafica.append(['Abierto',
				round((float(total_abierto) / float(total))*100,2)])

			datagrafica.append(['Cerrado',
				round((float(total_cerrado) / float(total))*100,2)])

			datos.append(
				{
					'grafica' : 'Graficas expedientes',
					'datagrafica' : datagrafica
				}
			)

			#cantidad de predios por docuentacion			
			
			qset2 = (Q(expediente__proyecto__id__in=ListProyectosValidos))
			queryset2 = Servidumbre_predio.objects.filter(qset2)

			serializer = Servidumbre_predioGeoSerializer(queryset2,many=True)

			datagrafica2 = []
			total2 = queryset2.count()

			data_con_documentos = 0
			data_sin_documentos = 0

			#import pdb; pdb.set_trace()

			for obj in serializer.data:
				if int(obj['porcentajedocumentos'])==100:
					data_con_documentos+=1

				elif int(obj['porcentajedocumentos'])<100:
					data_sin_documentos+=1
			
		
			datagrafica2.append(['Documentacion completa',
				round((data_con_documentos / float(total2))*100,2)])

			
			datagrafica2.append(['Documentacion incompleta',
				round((data_sin_documentos / float(total2))*100,2)])

			datos.append(
				{
					'grafica' : 'Graficas predios',
					'datagrafica' : datagrafica2
				}
			)

			return Response({'message':'','success':'ok','data':datos})	
					
		except Exception as e:
			functions.toLog(e,'servidumbre.datosGraficas')
			return Response({'message':'Se presentaron errores al procesar la solicitud','success':'error','data':''})	
