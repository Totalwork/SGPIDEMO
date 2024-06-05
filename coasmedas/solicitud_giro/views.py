from django.shortcuts import render,render_to_response
from django.contrib.auth.decorators import login_required
from django.template import RequestContext
from django.db import transaction
from django.http import HttpResponse,JsonResponse
from rest_framework.response import Response
from django.db.models import Q
from rest_framework import status

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import xlsxwriter
import json
from django.db.models import F, FloatField, Sum
from rest_framework import viewsets, serializers
from giros.models import DEncabezadoGiro
from logs.models import Logs,Acciones
from tipo.models import Tipo
from contrato.models import Contrato,EmpresaContrato
from contrato.enumeration import tipoC
from giros.models import CNombreGiro,DEncabezadoGiro,DetalleGiro,RechazoGiro
from contrato.models import Contrato
from empresa.models import Empresa
from proyecto.models import Proyecto, Proyecto_empresas
from proceso.models import AProceso, FProcesoRelacion, GProcesoRelacionDato
from financiero.models import FinancieroCuenta
from datetime import datetime, date, time, timedelta
import calendar
from sinin4.functions import functions
from numbertoletters import number_to_letters
from tipo.views import TipoSerializer
from parametrizacion.models import Banco
from parametrizacion.views import BancoSerializer
from empresa.views import EmpresaSerializer
from estado.models import Estado
from estado.views import EstadoSerializer
from giros.enum import enumEstados,enumTipo, enumTipoPagoAnticipo
from financiero.models import FinancieroCuenta,FinancieroCuentaMovimiento

from django.conf import settings
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from sinin4.functions import functions
from usuario.models import Usuario
from django.contrib.contenttypes.models import ContentType
# SERIALIZER LITE
class EmpresaLiteSerializer(serializers.HyperlinkedModelSerializer):
	"""docstring for ClassName"""	
	class Meta:
		model = Empresa
		fields=('id' , 'nombre'	, 'nit')



@login_required
def solicitud_inicio(request):
	return render(request, 'solicitud_giro/solicitud_inicio.html',{'app':'giros','model':'DEncabezadoGiro'})

@login_required
def solicitud_referencia(request):
	tipo_contrato = tipoC()
	qsetMcontratos = Contrato.objects.filter(activo=True 
											, empresacontrato__participa = True 
											, empresacontrato__empresa = request.user.usuario.empresa.id 
											, tipo_contrato_id = tipo_contrato.m_contrato).values('id', 'nombre').distinct()

	context = {
		'mcontrato':list(qsetMcontratos),
		'model':'giros',
		'app':'sol_giros'
	}	
	return render(request, 'solicitud_giro/codigo_sap.html',context)

@login_required
def detalle_giros(request,id_encabezado=None,mcontrato=None,contrato=None):
	return render(request, 'solicitud_giro/detalleAnticipo.html',{'app':'giros','model':'detallegiro','id_encabezado':id_encabezado,'mcontrato':mcontrato,'contrato':contrato})			

@login_required
def actualizar_testop(request):
	
	# inicio consulta de macro contratos y contratistas de la obra
	tipo_contrato = tipoC()
	qsetMcontratos = Contrato.objects.filter(activo=True 
											, empresacontrato__participa = True 
											, empresacontrato__empresa = request.user.usuario.empresa.id 
											, tipo_contrato_id = tipo_contrato.m_contrato).values('id', 'nombre').distinct()

	context = {
		'mcontrato':list(qsetMcontratos),
		'model':'giros',
		'app':'sol_giros'
	}	
	return render(request, 'solicitud_giro/solRevisar.html',context)

@login_required
def solTestOP(request):

	# inicio consulta de macro contratos y contratistas de la obra
	tipo_contrato = tipoC()
	qsetMcontratos = Contrato.objects.filter(activo=True 
											, empresacontrato__participa = True 
											, empresacontrato__empresa = request.user.usuario.empresa.id 
											, tipo_contrato_id = tipo_contrato.m_contrato).values('id', 'nombre').distinct()

	context = {
		'mcontrato':list(qsetMcontratos),
		'model':'giros',
		'app':'sol_giros'
	}	
	return render(request, 'solicitud_giro/solSinTest.html',context)


@login_required
def rechazados(request):
	tipo_contrato = tipoC()

	qset = (Q(empresacontrato__participa = True))
	qset = qset & (Q(empresacontrato__empresa = request.user.usuario.empresa.id))
	qset = qset & (Q(tipo_contrato_id = tipo_contrato.contratoProyecto))

	contratistas = Contrato.objects.filter(qset).values_list('contratista_id').distinct()

	# CONTRATISTAS DEL PROYECTO
	qsetEmpresas = Empresa.objects.filter(id__in = contratistas , esContratista = True)
	empresasData = EmpresaLiteSerializer(qsetEmpresas,many=True).data


	context = {
		'empresa':empresasData,
		'model':'giros',
		'app':'sol_giros'
	}
	return render(request, 'solicitud_giro/pagos_rechazados.html',context)	


@login_required
def solSinTestPagar(request):

	tipo_contrato = tipoC()
	qsetMcontratos = Contrato.objects.filter(activo=True 
											,empresacontrato__participa = True 
											, empresacontrato__empresa = request.user.usuario.empresa.id 
											, tipo_contrato_id = tipo_contrato.m_contrato ).values('id', 'nombre').distinct()

	context = {
		'mcontrato':list(qsetMcontratos),
		'model':'giros',
		'app':'sol_giros'
	}	
	return render(request, 'solicitud_giro/solSinPagar.html',context)	

@login_required
def registrarCodPago(request):
	tipo_contrato = tipoC()
	qsetMcontratos = Contrato.objects.filter(activo=True 
											,empresacontrato__participa = True 
											, empresacontrato__empresa = request.user.usuario.empresa.id 
											, tipo_contrato_id = tipo_contrato.m_contrato ).values('id', 'nombre').distinct()

	context = {
		'mcontrato':list(qsetMcontratos),
		'model':'giros',
		'app':'sol_giros'
	}	
	return render(request, 'solicitud_giro/regCodigoPago.html',context)


#funcion para traer el encabezado del detalle del giro
def conteoSolPendientes(request):
	try:
		id_empresa = request.user.usuario.empresa.id

		sinreferencia = DEncabezadoGiro.objects.filter(referencia='' ,disparar_flujo = True ,flujo_test=False ,contrato__empresacontrato__empresa=id_empresa ).count()

		lista1 = sinreferencia if  sinreferencia>0 else None

		por_revisar = DEncabezadoGiro.objects.filter(disparar_flujo = True, flujo_test=False, contrato__empresacontrato__empresa=id_empresa).exclude(referencia__exact='').count()

		lista2 = por_revisar if por_revisar>0 else None

		# este contador son las solicitudes que no tenga TEST-OP y tengan flujo 1 
		actualizar = DEncabezadoGiro.objects.filter(disparar_flujo = True, flujo_test=True, contrato__empresacontrato__empresa=id_empresa,detalle_encabezado_giro__test_op='' ).exclude(referencia__exact='').distinct()
		subquery_detalle= DetalleGiro.objects.filter(
					estado_id__in=[enumEstados.solicitado,enumEstados.autorizado]
					,encabezado_id__in=actualizar
					,test_op=''
				).values("encabezado_id").annotate(id=Sum('id')).values("encabezado_id").exclude(contratista_id__in=[39])
		actualizar = actualizar.filter(id__in = subquery_detalle).count()
		# actualizar = DetalleGiro.objects.filter(test_op__lte='',encabezado__flujo_test=True).count()

		lista3 = actualizar if actualizar>0 else None

		sol_sinpagar = DetalleGiro.objects.filter(encabezado__disparar_flujo=True, encabezado__flujo_test=True, fecha_pago=None, estado_id__in=[1,2]).exclude(test_op__exact='').count()

		lista4 = sol_sinpagar if sol_sinpagar>0 else None

		sol_sincodigo = DetalleGiro.objects.filter(test_op__gte='',codigo_pago='',fecha_pago=None).count()

		lista5 = sol_sincodigo if sol_sincodigo>0 else None

		return JsonResponse({'count':'','results':{'message':'','success':'ok','data':{
				'sinreferencia':lista1
				,'por_revisar':lista2
				,'actualizar':lista3
				,'sol_sinpagar':lista4
				,'sol_codigo':lista5

				}}})
					
	except Exception as e:
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)	



@transaction.atomic
def actualizar_flujotest(request):

	sid = transaction.savepoint()
	try:
		
		lista=request.POST['_content']
		respuesta= json.loads(lista)		
		id = respuesta['id']
		flujo_test = respuesta['flujo_test']

		giros=DEncabezadoGiro.objects.get(pk=id)
		giros.flujo_test=flujo_test
		giros.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.encabezado_giro',id_manipulado=id)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			


@transaction.atomic
def actualizar_referencia(request):

	sid = transaction.savepoint()
	try:
		
		lista=request.POST['_content']
		respuesta= json.loads(lista)		
		id = respuesta['id']
		referencia = respuesta['referencia']
		fecha_conta = respuesta['fecha_conta']
		texto_documento_sap = respuesta['texto_documento_sap']

		if fecha_conta=='null' or fecha_conta == '':
			fecha_conta=None

		giros=DEncabezadoGiro.objects.get(pk=id)
		giros.referencia=referencia
		giros.fecha_conta=fecha_conta
		giros.texto_documento_sap=texto_documento_sap
		giros.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.encabezado_giro',id_manipulado=id)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		

@transaction.atomic
def actualizar_detalle_testop(request):

	sid = transaction.savepoint()
	try:
				
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		for item in respuesta['lista']:		

			test_op = respuesta['test_op']

			if respuesta['fecha_pago_esperada']=='':

				fecha_pago_esperada=None

			else:

				fecha_pago_esperada = respuesta['fecha_pago_esperada']

			dgiro=DetalleGiro.objects.get(pk=item)
			dgiro.test_op=test_op
			dgiro.fecha_pago_esperada=fecha_pago_esperada
			dgiro.save()

			logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giro',id_manipulado=item)
			logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@transaction.atomic
def establecer_fechapago(request):

	sid = transaction.savepoint()
	try:
		
				
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		fecha_pago = respuesta['fechapago']
		cuenta_id = respuesta['cuenta']
		if cuenta_id=="":
			return JsonResponse({'message':'Debe seleccionar la cuenta','success':'error',
					'data':''})

		valida_registro_autorizacion = 0
		valida_registro_cuenta = 0
		insert_list_cuenta_movimiento = []
		insert_list_detalle_giros = []

		for item in respuesta['id']:		

			rechazopago=RechazoGiro.objects.filter(detalle_id=item).first() 
			object_detalle=DetalleGiro.objects.get(pk=item)
			estado=object_detalle.estado.id
			object_encabezado_giro = DEncabezadoGiro.objects.get(pk = object_detalle.encabezado_id)
			object_financiero_cuenta = FinancieroCuenta.objects.get(pk = cuenta_id)

			if object_encabezado_giro.nombre.contrato_id==object_financiero_cuenta.contrato_id:

				if rechazopago is None:
					
					if estado==enumEstados.autorizado:
						
						object_detalle.cuenta_id=cuenta_id
						object_detalle.fecha_pago=fecha_pago
						object_detalle.estado_id=enumEstados.pagado
						#object_detalle.estado_id=9
						object_detalle.save()

						insert_list_detalle_giros.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item))

						# solo se hace egreso a las cuentas bancarias en caso que el pago del anticipo sea con cuenta bancaria
						if object_encabezado_giro.pago_recurso_id == enumTipoPagoAnticipo.cuenta_bancaria:
							
							finan=FinancieroCuentaMovimiento(cuenta_id=object_detalle.cuenta.id,tipo_id=enumTipo.egreso,valor=object_detalle.valor_girar,fecha=object_detalle.fecha_pago,descripcion='contrato numero '+str(object_encabezado_giro.contrato.numero)+'-  '+str(object_encabezado_giro.contrato.nombre)+'- Beneficiario '+str(object_detalle.contratista.nombre))
							finan.save()
							insert_list_cuenta_movimiento.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='financiero.cuenta_movimiento',id_manipulado=finan.id))			

					elif valida_registro_autorizacion==0 :
						valida_registro_autorizacion = 1

				
				else:
					rechazo=RechazoGiro.objects.get(detalle_id=item)
					rechazo.atendido=1
					rechazo.save()

					if estado==enumEstados.autorizado:
						object_detalle.cuenta_id=cuenta_id
						object_detalle.fecha_pago=fecha_pago
						object_detalle.estado_id=enumEstados.pagado
						#object_detalle.estado_id=9
						object_detalle.save()

						insert_list_detalle_giros.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giros',id_manipulado=item))
		
						# solo se hace egreso a las cuentas bancarias en caso que el pago del anticipo sea con cuenta bancaria
						if object_encabezado_giro.pago_recurso_id == enumTipoPagoAnticipo.cuenta_bancaria:
							finan=FinancieroCuentaMovimiento(cuenta_id=object_detalle.cuenta.id,tipo_id=enumTipo.egreso,valor=object_detalle.valor_girar,fecha=object_detalle.fecha_pago,descripcion='contrato numero '+str(object_encabezado_giro.contrato.numero)+'-  '+str(object_encabezado_giro.contrato.nombre)+'- Beneficiario '+str(object_detalle.contratista.nombre))
							finan.save()
							insert_list_cuenta_movimiento.append(Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_crear,nombre_modelo='financiero.cuenta_movimiento',id_manipulado=finan.id))			

					elif valida_registro_autorizacion==0 :
						valida_registro_autorizacion = 1
			else:
				valida_registro_cuenta = 1


		# se hace una sola transaccion con bulk create
		if insert_list_detalle_giros:
			Logs.objects.bulk_create(insert_list_detalle_giros)

		if insert_list_cuenta_movimiento:
			Logs.objects.bulk_create(insert_list_cuenta_movimiento)

		transaction.savepoint_commit(sid)

		if valida_registro_cuenta==1:
			return JsonResponse({'message':'Algunos giros no coinciden con la cuenta que se esta pagando.','success':'error',
						'data':''})

		else:

			if valida_registro_autorizacion==1 :

				return JsonResponse({'message':'Solo se pueden pagar giros en estado de autorizacion','success':'error',
						'data':''})
			else:

				return JsonResponse({'message':'El registro se ha guardado correctamente','success':'ok',
						'data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			


@transaction.atomic
def establecer_rechazo(request):

	sid = transaction.savepoint()
	try:
		
				
		lista=request.POST['_content']
		respuesta= json.loads(lista)
		fecha_rechazo = respuesta['rechazo']
		motivo = respuesta['motivo']

		for item in respuesta['id']:		
			object_detalle=DetalleGiro.objects.get(pk=item)
			rechazopago=RechazoGiro.objects.filter(detalle_id=item).first() 
			if rechazopago is None:
				
				rechazo=RechazoGiro()
				rechazo.fecha=fecha_rechazo
				rechazo.motivo=motivo
				rechazo.atendido=0
				rechazo.detalle=object_detalle
				rechazo.save()

				object_detalle.estado_id=enumEstados.rechazado
	
				logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.rechazo',id_manipulado=item)
				logs_model.save()

			else:
				transaction.savepoint_rollback(sid)
				return JsonResponse({'message':'El anticipo con referencia numero :'+str(object_detalle.encabezado.referencia)+' se encuentra rechazado','success':'warnning','data':''})

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)			


@transaction.atomic
def codigo_pago(request):

	sid = transaction.savepoint()
	try:
		
		lista=request.POST['_content']
		respuesta= json.loads(lista)		
		id = respuesta['id']
		codigo_pago = respuesta['codigo_pago']


		giros=DetalleGiro.objects.get(pk=id)
		giros.codigo_pago=codigo_pago
		giros.save()

		logs_model=Logs(usuario_id=request.user.usuario.id,accion=Acciones.accion_actualizar,nombre_modelo='giros.detalle_giro',id_manipulado=id)
		logs_model.save()

		transaction.savepoint_commit(sid)

		return JsonResponse({'message':'El registro ha sido guardado exitosamente','success':'ok','data':''})

	except Exception as e:
		print(e)
		transaction.savepoint_rollback(sid)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)		


def mes_en_letra(mes):
	
	if mes==1:
		return 'ene'
	elif mes==2:
		return 'feb'
	elif mes ==3:
		return 'mar'
	elif mes == 4:
		return 'abr'
	elif mes==5:
		return 'may'
	elif mes==6:
		return 'jun'
	elif mes==7:
		return 'jul'
	elif mes==8:
		return 'ago'
	elif mes==9:
		return 'sep'
	elif mes==10:
		return 'oct'
	elif mes==11:
		return 'nov'
	else:
		return 'dic'


#Genera excel de test realizados
def report_testOP(request):

	try:
		response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')

		response['Content-Disposition'] = 'attachment; filename="Reporte_test_op'+str(datetime.now())+'.xls"'
		
		workbook = xlsxwriter.Workbook(response, {'in_memory': True})
		worksheet = workbook.add_worksheet('Hoja1')
		format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
		format2=workbook.add_format({'border':0})
		format3 = workbook.add_format({'num_format': 'dd-mm-yyyy','border':0})

		row=1
		col=0

		referencia = request.GET['referencia'] if 'referencia' in request.GET else 1;
		mcontrato_filtro = request.GET['mcontrato_filtro'] if 'mcontrato_filtro' in request.GET else None;
		contratista_filtro = request.GET['contratista_filtro'] if 'contratista_filtro' in request.GET else None; 
		dato = request.GET['dato'] if 'dato' in request.GET else None; 
		pago_recurso = request.GET['pago_recurso'] if 'pago_recurso' in request.GET else None;
		detallescompletos = request.GET['detallescompletos'] if 'detallescompletos' in request.GET else None;
		usuario = request.GET['usuario'] if 'usuario' in request.GET else request.user.id;
			
		id_empresa = request.user.usuario.empresa.id
		
		qset = ( Q(disparar_flujo=True) & Q(flujo_test=True) )
		qset = qset & (Q(contrato__empresacontrato__empresa=id_empresa))

		if dato:
			qset = qset &(Q(contrato__nombre__icontains=dato) | Q(contrato__numero__icontains=dato))

		if mcontrato_filtro and int(mcontrato_filtro)>0:
			qset = qset & (Q(nombre__contrato__id=mcontrato_filtro))

		if contratista_filtro and int(contratista_filtro)>0:
				
			qset = qset & (Q(contrato__contratista__id=contratista_filtro))

		if referencia:
			if int(referencia)==3:
				qset = qset & (Q(referencia=''))
			elif int(referencia)==1:
				qset = qset & (Q(referencia__gt=''))
			elif int(referencia)==2:
				qset = qset
					
		if detallescompletos:
			if detallescompletos=='false':
				qset = qset & (Q(detalle_encabezado_giro__test_op__lte=''))
			else:
				qset = qset & (Q(detalle_encabezado_giro__test_op__gt=''))

		if pago_recurso:
			qset = qset & (Q(pago_recurso=pago_recurso))

		encabezado = DEncabezadoGiro.objects.filter(qset).exclude(detalle_encabezado_giro__contratista_id__in=[39]).distinct()
		if detallescompletos=='true':
			subquery_detalle= DetalleGiro.objects.filter(
						estado_id__in=[enumEstados.solicitado,enumEstados.autorizado]
						,encabezado_id__in=encabezado
					).values("encabezado_id").annotate(id=Sum('id')).values("encabezado_id").exclude(contratista_id__in=[39]).exclude(test_op__exact='')
			encabezado = encabezado.filter(id__in = subquery_detalle)	
		else:
			subquery_detalle= DetalleGiro.objects.filter(
						estado_id__in=[enumEstados.solicitado,enumEstados.autorizado]
						,encabezado_id__in=encabezado
						,test_op=''
					).values("encabezado_id").annotate(id=Sum('id')).values("encabezado_id").exclude(contratista_id__in=[39])# 39 es el contratista vega energy
			encabezado = encabezado.filter(id__in = subquery_detalle)

		if encabezado:

			worksheet.write('A1', 'Fecha cont', format1)
			worksheet.write('B1', 'Documento SAP', format1)
			worksheet.write('C1', 'macro-contrato', format1)
			worksheet.write('D1', 'Proyecto', format1)
			worksheet.write('E1', 'No Contrato', format1)
			worksheet.write('F1', 'Contratista', format1)
			worksheet.write('G1', 'Giro', format1)
			worksheet.write('H1', 'Valor a girar', format1)
			worksheet.write('I1', 'Fecha de pago', format1)
			worksheet.write('J1', 'Cod. Acreedor', format1)
			worksheet.write('K1', 'No. Radicado', format1)
			worksheet.write('L1', 'Cuenta origen', format1)
			worksheet.write('M1', 'No. Orden (TEST/OP)', format1)


			worksheet.set_column('A:A', 10)
			worksheet.set_column('B:B', 18)
			worksheet.set_column('C:C', 18)
			worksheet.set_column('D:D', 12)
			worksheet.set_column('E:E', 12)
			worksheet.set_column('F:F', 16)
			worksheet.set_column('G:G', 18)
			worksheet.set_column('H:H', 18)
			worksheet.set_column('I:I', 18)
			worksheet.set_column('J:J', 18)
			worksheet.set_column('K:K', 18)
			worksheet.set_column('L:L', 18)
			worksheet.set_column('M:M', 18)

			total=0
			
			for enc in encabezado:

				worksheet.write(row, col,enc.fecha_conta,format3)
				worksheet.write(row, col+1,enc.referencia,format2)
				worksheet.write(row, col+2,enc.nombre.contrato.nombre,format2)
				worksheet.write(row, col+3,enc.contrato.nombre,format2)
				worksheet.write(row, col+4,enc.contrato.numero,format2)
				worksheet.write(row, col+5,enc.contrato.contratista.nombre,format2)
				worksheet.write(row, col+6,enc.nombre.nombre,format2)
				worksheet.write(row, col+8,' ',format2)
				worksheet.write(row, col+9,enc.contrato.contratista.codigo_acreedor,format2)
				worksheet.write(row, col+10,enc.numero_radicado,format2)
				cuentaOrigen=FinancieroCuenta.objects.filter(
					id__in=DetalleGiro.objects.filter(encabezado=enc,estado_id=3).values('cuenta_id')).values(
					'numero','nombre','fiduciaria')
				#import pdb; pdb.set_trace()
				if cuentaOrigen.count()==0:
					cuentaOrigen=FinancieroCuenta.objects.filter(
						contrato=enc.contrato.mcontrato).last()
					worksheet.write(row, col+11,
						cuentaOrigen.fiduciaria + ' - ' + 
						cuentaOrigen.nombre + ' - No. ' + 
						cuentaOrigen.numero ,
						format2)
				else:					
					worksheet.write(row, col+11,
						cuentaOrigen[0]['fiduciaria'] + ' - ' + 
						cuentaOrigen[0]['nombre'] + ' - No. ' + 
						cuentaOrigen[0]['numero'] ,
						format2)


				ops= DetalleGiro.objects.filter(encabezado=enc).values('test_op').distinct()
				if ops:
					if ops.count()>1:
						optest=''
						sep=''
						for op in ops:
							optest=optest + sep +  op['test_op']
							sep=', '
					else:
						optest=ops[0]['test_op']

				worksheet.write(row, col+12,optest,format2)

				total_detalle = DetalleGiro.objects.filter(encabezado_id=enc.id).aggregate(sol_suma_detalle=Sum('valor_girar'))
				total_detalle = total_detalle['sol_suma_detalle'] if total_detalle['sol_suma_detalle']>0 else "0"
				worksheet.write(row, col+7,str("$")+str(total_detalle),format2)

				row +=1

		else:
			worksheet.write('A1', 'No se encontraron registros', format2)


		workbook.close()

		return response
	except Exception as e:
		functions.toLog(e,'Solicitud_giro')
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	
#Genera excel de test realizados
def test_realizados(request):

	response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8')

	response['Content-Disposition'] = 'attachment; filename="Reporte_test_op"'+str(datetime.now())+'".xls"'
	
	workbook = xlsxwriter.Workbook(response, {'in_memory': True})
	worksheet = workbook.add_worksheet('Hoja1')
	format1=workbook.add_format({'border':1,'font_size':12,'bold':True, 'bg_color':'#4a89dc','font_color':'white'})
	format2=workbook.add_format({'border':0})
	format3 = workbook.add_format({'num_format': 'dd/mm/yyyy','border':0})

	row=1
	col=0


	mcontrato=request.GET.get('mcontrato')

	contratista=request.GET.get('contratista')

	fecha_estimada_desde=request.GET.get('fecha_e_desde')

	fecha_estimada_hasta=request.GET.get('fecha_e_hasta')

	qset = (Q(test_op__gt=''))

	if mcontrato:
		qset = qset &(Q(encabezado__nombre__contrato__id=mcontrato))

	if contratista:
		qset = qset &(Q(contratista__id=contratista))

	if fecha_estimada_desde:
		qset = qset &(Q(fecha_pago_esperada__gte=fecha_estimada_desde))

	if fecha_estimada_hasta:
		qset = qset &(Q(fecha_pago_esperada__lte=fecha_estimada_hasta))

	# if fecha_generacion_desde:
	# 	qset = qset &(Q(contratista__id=fecha_generacion_desde))

	# if fecha_generacion_hasta:
	# 	qset = qset &(Q(contratista__id=fecha_generacion_hasta))
				
	detalle = DetalleGiro.objects.filter(qset)

	if detalle:

		worksheet.write('A1', 'Fecha contabilizacion', format1)
		worksheet.write('B1', 'Fecha estimada de pago', format1)
		worksheet.write('C1', 'Documento SAP', format1)
		worksheet.write('D1', 'macro-contrato', format1)
		worksheet.write('E1', 'Proyecto', format1)
		worksheet.write('F1', 'No Contrato', format1)
		worksheet.write('G1', 'Contratista', format1)
		worksheet.write('H1', 'Beneficiario/Provedor', format1)
		worksheet.write('I1', 'Nit', format1)
		worksheet.write('J1', 'Entidad bancaria', format1)
		worksheet.write('K1', 'No. Cuenta destino', format1)
		worksheet.write('L1', 'Tipo cuenta destino', format1)
		worksheet.write('M1', 'Valor a girar', format1)
		worksheet.write('N1', 'TEST-OP', format1)
		worksheet.write('O1', 'Estado', format1)

		worksheet.set_column('A:A', 19)
		worksheet.set_column('B:B', 18)
		worksheet.set_column('C:C', 18)
		worksheet.set_column('D:D', 12)
		worksheet.set_column('E:E', 12)
		worksheet.set_column('F:F', 16)
		worksheet.set_column('G:G', 18)
		worksheet.set_column('H:H', 18)
		worksheet.set_column('I:I', 18)
		worksheet.set_column('J:J', 18)
		worksheet.set_column('K:K', 18)
		worksheet.set_column('L:L', 18)
		worksheet.set_column('M:M', 18)
		worksheet.set_column('N:N', 18)
		worksheet.set_column('O:O', 18)

		total=0

		for details in detalle:


			worksheet.write(row, col,details.encabezado.fecha_conta,format3)
			worksheet.write(row, col+1,details.fecha_pago_esperada,format3)
			worksheet.write(row, col+2,details.encabezado.referencia,format2)
			worksheet.write(row, col+3,details.encabezado.nombre.contrato.nombre,format2)
			worksheet.write(row, col+4,details.encabezado.contrato.nombre,format2)
			worksheet.write(row, col+5,details.encabezado.contrato.numero,format2)
			worksheet.write(row, col+6,details.encabezado.contrato.contratista.nombre,format2)
			worksheet.write(row, col+7,details.contratista.nombre,format2)
			worksheet.write(row, col+8,details.contratista.nit,format2)
			worksheet.write(row, col+9,details.banco.nombre,format2)
			worksheet.write(row, col+10,details.no_cuenta,format2)
			worksheet.write(row, col+11,details.tipo_cuenta.nombre,format2)
			worksheet.write(row, col+12,details.valor_girar,format2)
			worksheet.write(row, col+13,details.test_op,format2)
			worksheet.write(row, col+14,details.estado.nombre,format2)


			row +=1

	else:
		worksheet.write('A1', 'No se encontraron registros', format2)



	workbook.close()

	return response	


#genera el reporte de las plantillas fiduciaria BANCO DE BOGOTA; BANCOLOMBIA
def generar_reporte_fiduciario(request):

	try:
		id_anticipo = request.GET['encabezado_id']
		giros_detalle = request.GET['detalles']
		

		encabezadogiro = DEncabezadoGiro.objects.get(pk=id_anticipo)

		detalles = DetalleGiro.objects.filter(encabezado__id=id_anticipo ,  id__in = giros_detalle.split (','))

		cuenta=FinancieroCuenta.objects.filter(contrato=encabezadogiro.nombre.contrato) 

		codfideico = ''
		codfideico_antiguo = ''
		nombrefideico = ''
		nombrefiduciaria = ''
		macro_contrato = ''

		for item2 in cuenta:

			nocuenta = item2.numero 
			# codigo fidecomiso
			if item2.codigo_fidecomiso:
				codfideico = item2.codigo_fidecomiso

			if item2.codigo_fidecomiso_a:
				codfideico_antiguo = item2.codigo_fidecomiso_a
			# nombre de fidecomiso

			if item2.nombre_fidecomiso :
				nombrefideico= item2.nombre_fidecomiso

			tipofideico = item2.tipo.nombre

			if item2.fiduciaria:
				nombrefiduciaria = item2.fiduciaria

			if item2.contrato:
				macro_contrato = item2.contrato.nombre

		ruta = settings.STATICFILES_DIRS[0]
		newpath = ruta + '/papelera/'
		empresa_id = request.user.usuario.empresa.id
		ahora = datetime.now()

		if nombrefiduciaria=='FIDUBOGOTA':				

			filename = "plantillas/solicituGiros/FIDUBOGOTACREATE.xlsx"
			extension = filename[filename.rfind('.'):]
			nombre = 'empresa'+str(empresa_id)+''+str(request.user.usuario.id)+extension		

			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )

			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre
			# wb = load_workbook(filename = ruta, keep_vba=True)
			wb = load_workbook(filename = ruta, read_only=False)

			inicio = wb.worksheets[0]

			# descargar imagen para la plantilla  bogota
			filename = "plantillas/solicituGiros/excel.png"
			extension = filename[filename.rfind('.'):]
			nombre = 'fiduciariabogotainicio'+extension		
			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )
			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre

			img_inicio = Image(ruta)
			inicio.add_image(img_inicio, 'A1')


			ws = wb.worksheets[1]
			ws_datos = wb.worksheets[2]
			# ws = wb['10112015']	



			# #Estilo para los bordes del archivo
			thick_border = Border(
				left=Side(style='medium',color='000000'),
				right=Side(style='medium',color='000000'),
				top=Side(style='medium',color='000000'),
				bottom=Side(style='medium',color='000000'))
	
			thin_border2 = Border(
				left=Side(style='thin',color='A6A5A5'),
				right=Side(style='thin',color='A6A5A5'),
				bottom=Side(style='thin',color='A6A5A5'))
	
			thin_border3 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='thin',color='000000'),
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thick_border4 = Border(
				bottom=Side(style='medium',color='000000'))

			thick_border5 = Border(
				right=Side(style='medium',color='000000'),		
				bottom=Side(style='medium',color='000000'))

			thick_border6 = Border(
				right=Side(style='medium',color='000000'))

			thin_border7 = Border(
				bottom=Side(style='thin',color='000000'))

			thin_border8 = Border(
				right=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border9 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='medium',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border10 = Border(
				left=Side(style='thin',color='000000'),
				right=Side(style='medium',color='000000'),
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border11 = Border(
				right=Side(style='medium',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border12 = Border(
				right=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))

			thin_border13 = Border(
				top=Side(style='thin',color='000000'),
				right=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border14 = Border(
				top=Side(style='thin',color='000000'),
				left=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thick_border15 = Border(
				left=Side(style='medium',color='000000'))

			thick_border16 = Border(
				left=Side(style='medium',color='000000'),		
				bottom=Side(style='medium',color='000000'))

			thin_border17 = Border(
				top=Side(style='thin',color='000000'),
				bottom=Side(style='thin',color='000000'))

			thin_border18 = Border(
				left=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))

			thick_border19 = Border(
				right=Side(style='medium',color='000000'),
				left=Side(style='thin',color='000000'),
				bottom=Side(style='medium',color='000000'))




			# order de operacion sifi COLUM R TO COLUM Z
			for x in xrange(16,27):
				ws.cell(row=1, column=x).border = thick_border4

			for x in xrange(29,33):
				ws.cell(row=1, column=x).border = thick_border4


			for x in xrange(17,27):
				ws.cell(row=8, column=x).border = thin_border7


			for x in xrange(4,6):
				ws.cell(row=9, column=x).border = thin_border17

			for x in xrange(7,10):
				ws.cell(row=9, column=x).border = thin_border17

			for x in xrange(12,27):
				ws.cell(row=9, column=x).border = thin_border7

			ws.cell(row=9, column=9).border = thin_border13

			ws.cell(row=1, column=33).border = thick_border5

			ws.cell(row=8, column=33).border = thick_border5
			ws.cell(row=9, column=33).border = thick_border5
			ws.cell(row=10, column=33).border = thick_border5
			ws.cell(row=11, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=12, column=x).border = thick_border4

			ws.cell(row=12, column=33).border = thick_border5	

			ws.cell(row=13, column=31).border = thin_border7


			for x in xrange(8,10):
				ws.cell(row=14, column=x).border = thin_border17	

			ws.cell(row=14, column=9).border = thin_border13	

			for x in xrange(3,33):
				ws.cell(row=15, column=x).border = thick_border4

			ws.cell(row=15, column=33).border = thick_border5
			ws.cell(row=15, column=31).border = thin_border12
			ws.cell(row=15, column=26).border = thin_border18

			ws.cell(row=15, column=33).border = thick_border19

			for x in xrange(3,33):
				ws.cell(row=16, column=x).border = thick_border4

			ws.cell(row=16, column=33).border = thick_border5

			# LIQUIDACION DE LA OPERACION

			for x in xrange(12,18):
				ws.cell(row=17, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=17, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=18, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=18, column=x).border = thin_border7

			ws.cell(row=18, column=18).border = thin_border9

			for x in xrange(27,33):
				ws.cell(row=18, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=19, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=20, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=21, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(23,25):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=22, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=23, column=x).border = thin_border7

			for x in xrange(7,9):
				ws.cell(row=24, column=x).border = thin_border7

			for x in xrange(12,18):
				ws.cell(row=24, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=24, column=x).border = thick_border4

			for x in xrange(12,18):
				ws.cell(row=25, column=x).border = thick_border4

			for x in xrange(27,33):
				ws.cell(row=25, column=x).border = thick_border4

			ws.cell(row=18, column=33).border = thin_border10
			ws.cell(row=19, column=33).border = thin_border9
			ws.cell(row=20, column=33).border = thin_border9
			ws.cell(row=21, column=33).border = thin_border9
			ws.cell(row=22, column=33).border = thin_border9
			ws.cell(row=23, column=33).border = thin_border9

			# fin liquidacion de la operacion


			for x in xrange(5,33):
				ws.cell(row=26, column=x).border = thin_border7

			for x in xrange(3,33):
				ws.cell(row=27, column=x).border = thin_border7

			ws.cell(row=26, column=33).border = thin_border11
			ws.cell(row=27, column=33).border = thin_border11


			for x in xrange(3,33):
				ws.cell(row=28, column=x).border = thick_border4

			ws.cell(row=28, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=29, column=x).border = thick_border4

			ws.cell(row=29, column=33).border = thick_border5

			for x in xrange(6,33):
				ws.cell(row=30, column=x).border = thin_border7

			ws.cell(row=30, column=33).border = thin_border11

			for x in xrange(3,33):
				ws.cell(row=32, column=x).border = thick_border4


			# formacion de operacion
			for x in xrange(3,33):
				ws.cell(row=33, column=x).border = thick_border4

			ws.cell(row=33, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=72, column=x).border = thick_border4
			ws.cell(row=72, column=33).border = thick_border5


			for x in xrange(26,32):
				ws.cell(row=39, column=x).border = thin_border7

			for x in xrange(14,32):
				ws.cell(row=40, column=x).border = thin_border7

			for x in xrange(13,19):
				ws.cell(row=41, column=x).border = thin_border7

			for x in xrange(22,32):
				ws.cell(row=41, column=x).border = thin_border7

			# observaciones
			for x in xrange(3,33):
				ws.cell(row=43, column=x).border = thick_border4
			ws.cell(row=43, column=33).border = thick_border5

			ws.cell(row=44, column=2).border = thick_border15
			ws.cell(row=45, column=2).border = thick_border15
			ws.cell(row=46, column=2).border = thick_border15
			ws.cell(row=47, column=2).border = thick_border15
			ws.cell(row=48, column=2).border = thick_border15
			ws.cell(row=49, column=2).border = thick_border16

			ws.cell(row=44, column=33).border = thick_border6
			ws.cell(row=45, column=33).border = thick_border6
			ws.cell(row=46, column=33).border = thick_border6
			ws.cell(row=47, column=33).border = thick_border6
			ws.cell(row=48, column=33).border = thick_border6

			# condiciones de manejo
			for x in xrange(3,33):
				ws.cell(row=49, column=x).border = thick_border4
			ws.cell(row=49, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=50, column=x).border = thick_border4
			ws.cell(row=50, column=33).border = thick_border5

			for x in xrange(3,13):
				ws.cell(row=53, column=x).border = thin_border7

			for x in xrange(14,24):
				ws.cell(row=53, column=x).border = thin_border7

			for x in xrange(25,33):
				ws.cell(row=53, column=x).border = thin_border7

			ws.cell(row=52, column=2).border = thick_border15
			ws.cell(row=53, column=2).border = thick_border15

			for x in xrange(6,13):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=55, column=x).border = thin_border7

			for x in xrange(6,13):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=56, column=x).border = thin_border7

			for x in xrange(6,13):
				ws.cell(row=57, column=x).border = thin_border7

			for x in xrange(16,24):
				ws.cell(row=57, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=57, column=x).border = thin_border7

			ws.cell(row=51, column=33).border = thick_border6
			ws.cell(row=52, column=33).border = thick_border6
			ws.cell(row=53, column=33).border = thin_border11


			ws.cell(row=54, column=33).border = thick_border6
			ws.cell(row=55, column=33).border = thin_border11
			ws.cell(row=56, column=33).border = thin_border11
			ws.cell(row=57, column=33).border = thin_border11		



			# uso exclusivo fidubogota
			for x in xrange(3,33):
				ws.cell(row=58, column=x).border = thick_border4
			ws.cell(row=58, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=59, column=x).border = thick_border4
			ws.cell(row=59, column=33).border = thick_border5

			
			for x in xrange(9,19):
				ws.cell(row=61, column=x).border = thin_border7

			for x in xrange(22,33):
				ws.cell(row=61, column=x).border = thin_border7

			for x in xrange(13,19):
				ws.cell(row=62, column=x).border = thin_border7

			for x in xrange(25,33):
				ws.cell(row=62, column=x).border = thin_border7

			for x in xrange(4,6):
				ws.cell(row=65, column=x).border = thin_border7

			for x in xrange(16,19):
				ws.cell(row=65, column=x).border = thin_border7

			for x in xrange(22,24):
				ws.cell(row=65, column=x).border = thin_border7


			ws.cell(row=61, column=19).border = thin_border11
			ws.cell(row=62, column=19).border = thin_border11
			ws.cell(row=65, column=19).border = thin_border11

			ws.cell(row=60, column=33).border = thick_border6
			ws.cell(row=61, column=33).border = thin_border11
			ws.cell(row=62, column=33).border = thin_border11

			ws.cell(row=60, column=19).border = thick_border6
			

			# informacion de procesos especiales
			for x in xrange(3,33):
				ws.cell(row=66, column=x).border = thick_border4
			ws.cell(row=66, column=33).border = thick_border5
			ws.cell(row=66, column=19).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=67, column=x).border = thick_border4
			ws.cell(row=67, column=33).border = thick_border5


			for x in xrange(8,12):
				ws.cell(row=68, column=x).border = thin_border7

			for x in xrange(16,25):
				ws.cell(row=68, column=x).border = thin_border7

			for x in xrange(31,33):
				ws.cell(row=68, column=x).border = thin_border7


			for x in xrange(8,12):
				ws.cell(row=69, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=69, column=x).border = thin_border7

			for x in xrange(8,12):
				ws.cell(row=70, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=70, column=x).border = thin_border7

			for x in xrange(8,12):
				ws.cell(row=71, column=x).border = thin_border7

			for x in xrange(19,25):
				ws.cell(row=71, column=x).border = thin_border7

			for x in xrange(30,33):
				ws.cell(row=71, column=x).border = thin_border7

			ws.cell(row=68, column=33).border = thick_border6
			ws.cell(row=71, column=33).border = thin_border11


			# informacion de procesos generales
			for x in xrange(3,33):
				ws.cell(row=73, column=x).border = thick_border4
			ws.cell(row=73, column=33).border = thick_border5

			for x in xrange(28,33):
				ws.cell(row=75, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=76, column=x).border = thin_border7

			for x in xrange(28,33):
				ws.cell(row=76, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=77, column=x).border = thin_border7

			for x in xrange(14,26):
				ws.cell(row=77, column=x).border = thin_border7

			for x in xrange(5,8):
				ws.cell(row=78, column=x).border = thin_border7

			for x in xrange(17,23):
				ws.cell(row=78, column=x).border = thin_border7

			for x in xrange(28,33):
				ws.cell(row=78, column=x).border = thin_border7

			ws.cell(row=78, column=33).border = thin_border11

			ws.cell(row=77, column=8).border = thin_border13
			ws.cell(row=78, column=8).border = thin_border13
			ws.cell(row=79, column=8).border = thin_border13

			ws.cell(row=77, column=5).border = thin_border14
			ws.cell(row=78, column=5).border = thin_border14
			ws.cell(row=79, column=5).border = thin_border14

			

			for x in xrange(6,8):
				ws.cell(row=79, column=x).border = thin_border7

			for x in xrange(8,20):
				ws.cell(row=83, column=x).border = thin_border7

			for x in xrange(26,33):
				ws.cell(row=83, column=x).border = thin_border7
			

			for x in xrange(3,33):
				ws.cell(row=84, column=x).border = thick_border4
			ws.cell(row=84, column=33).border = thick_border5

			for x in xrange(3,33):
				ws.cell(row=85, column=x).border = thick_border4
			ws.cell(row=85, column=33).border = thick_border5
			ws.cell(row=86, column=33).border = thick_border6
			ws.cell(row=87, column=33).border = thick_border6
			ws.cell(row=88, column=33).border = thick_border6
			ws.cell(row=89, column=33).border = thick_border6
			ws.cell(row=90, column=33).border = thick_border6


			for x in xrange(3,6):
				ws.cell(row=86, column=x).border = thin_border7
			
			for x in xrange(3,6):
				ws.cell(row=87, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=88, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=89, column=x).border = thin_border7

			for x in xrange(3,6):
				ws.cell(row=90, column=x).border = thin_border7

			for x in xrange(7,14):
				ws.cell(row=87, column=x).border = thin_border7

			for x in xrange(7,14):
				ws.cell(row=89, column=x).border = thin_border7

			ws.cell(row=86, column=6).border = thin_border8
			ws.cell(row=87, column=6).border = thin_border8
			ws.cell(row=88, column=6).border = thin_border8
			ws.cell(row=89, column=6).border = thin_border8
			ws.cell(row=90, column=6).border = thin_border8

			ws.cell(row=87, column=14).border = thin_border11
			ws.cell(row=88, column=14).border = thick_border5
			ws.cell(row=89, column=14).border = thin_border11
			ws.cell(row=90, column=14).border = thick_border5
			

			ws.cell(row=87, column=27).border = thick_border6
			ws.cell(row=88, column=27).border = thick_border6
			ws.cell(row=89, column=27).border = thick_border6
			ws.cell(row=90, column=27).border = thick_border6			


			for x in xrange(3,33):
				ws.cell(row=91, column=x).border = thick_border4

			
			ws.cell(row=91, column=6).border = thin_border12
			ws.cell(row=91, column=33).border = thick_border5
			ws.cell(row=91, column=14).border = thick_border5
			ws.cell(row=91, column=27).border = thick_border5

			for x in xrange(6,20):
				ws.cell(row=93, column=x).border = thin_border7

			for x in xrange(6,20):
				ws.cell(row=94, column=x).border = thin_border7

			for x in xrange(6,20):
				ws.cell(row=95, column=x).border = thin_border7

			for x in xrange(6,12):
				ws.cell(row=96, column=x).border = thin_border7

			for x in xrange(16,21):
				ws.cell(row=96, column=x).border = thin_border7


			for x in xrange(27,33):
				ws.cell(row=93, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=94, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=95, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=927, column=x).border = thin_border7

			for x in xrange(27,33):
				ws.cell(row=96, column=x).border = thin_border7

			ws.cell(row=92, column=33).border = thick_border6
			ws.cell(row=93, column=33).border = thin_border11
			ws.cell(row=94, column=33).border = thin_border11
			ws.cell(row=95, column=33).border = thin_border11
			ws.cell(row=96, column=33).border = thin_border11


			# hoja datos
			for x in xrange(3,9):
				ws_datos.cell(row=11, column=x).border = thin_border7

			for x in xrange(3,5):
				ws_datos.cell(row=16, column=x).border = thin_border7

			for x in xrange(7,11):
				ws_datos.cell(row=16, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=18, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=19, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=20, column=x).border = thin_border7

			for x in xrange(9,11):
				ws_datos.cell(row=21, column=x).border = thin_border7


			#nombre macro contrato
			ws['K9'] = macro_contrato
			
			ws['AB9'] = codfideico[:1]
			ws['AD9'] = codfideico[2:3]
			ws['AF9'] = codfideico[-5:]

			if codfideico_antiguo:

				ws['AB11'] = codfideico_antiguo[:1]
				ws['AD11'] = codfideico_antiguo[2:3]
				ws['AF11'] = codfideico_antiguo[-5:]

			ws['E14'] = str(ahora.day)
			ws['F14'] = str(ahora.month)
			ws['G14'] = str(ahora.year)


			valor_total_letras = 0
			row = 2
			for item in detalles:
	
				fecha_generacion = DetalleGiro.objects.get(pk=item.id)
	
				fecha_generacion.fecha_transaccion_test=str(ahora.year)+"-"+str(ahora.month)+"-"+str(ahora.day)
	
				fecha_generacion.save()
	
				ws_datos['C'+str(row)] = item.contratista.nit
				ws_datos['D'+str(row)] = item.contratista.nombre
				ws_datos['I'+str(row)] = item.valor_girar
				ws_datos['J'+str(row)] = item.fecha_pago_esperada
				ws_datos['N'+str(row)] = item.contratista.nombre
				ws_datos['P'+str(row)] = item.banco.codigo_bancario
				ws_datos['Q'+str(row)] = item.no_cuenta

				nombre_tipo = item.tipo_cuenta.nombre

				if  nombre_tipo.lower() == 'corriente':
					
					ws_datos['R'+str(row)] = 'CC'
				else:

					ws_datos['R'+str(row)] = 'CH'
				
				
				row = row+1
				valor_total_letras = valor_total_letras+item.valor_girar

			observacion = encabezadogiro.nombre.nombre+' - '+encabezadogiro.contrato.nombre

			ws_datos['D13'] = observacion			


			# descargar imagen para la plantilla  bogota
			filename = "plantillas/solicituGiros/fiduciariabogota.png"
			extension = filename[filename.rfind('.'):]
			nombre = 'fiduciariabogota'+extension	
			# print nombre	
			functions.descargarArchivoS3(str(filename), str(newpath) , nombre )
			ruta = settings.STATICFILES_DIRS[0]+ '/papelera/'+nombre

			img2 = Image(ruta)
			ws.add_image(img2, 'B1')


			wb.template = False
			response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
			response['Content-Disposition'] ='attachment; filename="Formato Orden de Operacion '+str(macro_contrato)+'.xlsx"'
			return response		

		else:#nombrefiduciaria=='BANCOLOMBIA':
			pass

	except Exception as e:
		print(e)
		return JsonResponse({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
			
#Serializer financiero cuenta
class FinancieroListCuentaSerializer(serializers.HyperlinkedModelSerializer):
	
	class Meta:
		model = FinancieroCuenta
		fields=('id','numero','nombre')

# Serializador de contratista
class ContratistaLiteSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Empresa
		fields=('id','nombre')

# Serializador de contrato
class ContratoLiteSerializer(serializers.HyperlinkedModelSerializer):

	contratista = ContratistaLiteSerializer(read_only=True)
	# contratista_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Empresa.objects.all())
	class Meta:
		model = Contrato
		fields=('id','nombre','contratista','numero')

class ContratoInterAdministrativoSerializer(serializers.HyperlinkedModelSerializer):

	class Meta:
		model = Contrato
		fields=('id','nombre','numero')

#serializador de nombre giro
class NombreGiroLiteSerializer(serializers.HyperlinkedModelSerializer):
	# contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoInterAdministrativoSerializer(read_only=True)
	class Meta:
		model = CNombreGiro
		fields=('id','nombre','contrato')

#Api rest para encabezado del giro
class EncabezadoGiroSerializer(serializers.HyperlinkedModelSerializer):
	# contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoLiteSerializer(read_only=True)
	# nombre_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=CNombreGiro.objects.all())
	nombre=NombreGiroLiteSerializer(read_only=True)	
	#nombre = serializers.PrimaryKeyRelatedField(queryset=NombreGiro.objects.all())
	porcentaje = serializers.SerializerMethodField('_porcentajeV2',read_only=True)
	sol_giro_suma_detalle = serializers.SerializerMethodField('_sol_giro_suma_detalle',read_only=True)

	def _porcentajeV2(self,obj):
		retorno=0
		#identificar el proceso q se va a analizar:
		#import pdb; pdb.set_trace()
		nombreMcontrato=obj.nombre.contrato.nombre
		objProceso=None
		if nombreMcontrato.find('PRONE')>=0 and nombreMcontrato.find('2012')>=0:
			objProceso=AProceso.objects.get(pk=9)
		elif nombreMcontrato.find('PRONE')>=0 and (nombreMcontrato.find('2013')>=0 or nombreMcontrato.find('2014')>=0 ):
			objProceso=AProceso.objects.get(pk=1)
		elif nombreMcontrato.find('FAER')>=0:
			objProceso=AProceso.objects.get(pk=15)

		if objProceso:	
			#identificar el proyecto que se esta revisando:
			proyecto=Proyecto.objects.filter(contrato=obj.contrato)[:1].values('id')

			items = GProcesoRelacionDato.objects.filter(procesoRelacion__proceso=objProceso,
				procesoRelacion__idApuntador=proyecto[0]['id'])
			totalItems = items.count()
			totalItemsCumplidos = items.filter(estado='1').count()
			retorno = round((float(totalItemsCumplidos) / float(totalItems))*100,2)
		
		return retorno

	# se suma los detalle solicitados y autorizado para solicitud de giro
	def _sol_giro_suma_detalle(self,obj):
		total_detalle = DetalleGiro.objects.filter(encabezado_id=obj.id, estado_id__in=[1,2]).aggregate(sol_suma_detalle=Sum('valor_girar'))
		return total_detalle['sol_suma_detalle']

	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre','contrato','referencia','fecha_conta','disparar_flujo','numero_radicado','suma_detalle','porcentaje','sol_giro_suma_detalle','texto_documento_sap')

		
#serializer encabezado giro lite
class EncabezadoLiteSerializer(serializers.HyperlinkedModelSerializer):
	# contrato_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Contrato.objects.all())
	contrato=ContratoLiteSerializer(read_only=True)

	# nombre_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=CNombreGiro.objects.all())
	nombre=NombreGiroLiteSerializer(read_only=True)
	
	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre','contrato','soporte','referencia', 'fecha_conta' ,'suma_detalle')


# SOL SIN TEST-OP
class EncabezadoGiroSolSinTestOpSerializer(serializers.HyperlinkedModelSerializer):
	contrato=ContratoLiteSerializer(read_only=True)
	nombre=NombreGiroLiteSerializer(read_only=True)	
	sol_giro_suma_detalle = serializers.SerializerMethodField('_sol_giro_suma_detalle',read_only=True)

	# se suma los detalle solicitados y autorizado para solicitud de giro
	def _sol_giro_suma_detalle(self,obj):
		total_detalle = DetalleGiro.objects.filter(encabezado_id=obj.id).aggregate(sol_suma_detalle=Sum('valor_girar'))
		return total_detalle['sol_suma_detalle']

	class Meta:
		model = DEncabezadoGiro
		fields=('id','nombre','contrato','referencia','fecha_conta','disparar_flujo','numero_radicado','suma_detalle','sol_giro_suma_detalle','texto_documento_sap')

#api encabezado giro lite
class EncabezadoSolicitudGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de encabezado de los giros.
	"""
	model=DEncabezadoGiro
	queryset = model.objects.all()
	serializer_class = EncabezadoGiroSerializer
	nombre_modulo='Solicitud_giro'
	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:

			queryset = super(EncabezadoSolicitudGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			contrato_filtro= self.request.query_params.get('contrato_filtro',None)
			mcontrato_filtro= self.request.query_params.get('mcontrato_filtro',None)
			contratista_filtro= self.request.query_params.get('contratista_filtro',None)
			encabezado_id= self.request.query_params.get('encabezado_id',None)
			sin_paginacion=self.request.query_params.get('sin_paginacion',None)
			id_empresa = request.user.usuario.empresa.id
			referencia = self.request.query_params.get('referencia',None)	
			detallescompletos= self.request.query_params.get('detallescompletos',None)
			usuario = self.request.query_params.get('usuario', request.user.id)
			disparar_flujo = self.request.query_params.get('disparar_flujo', 1 )
			pago_recurso = self.request.query_params.get('pago_recurso', None )
			flujo_test= self.request.query_params.get('flujotest',None)
			
				
			# SERIALIZADORES
			sol_sin_test_op = self.request.query_params.get('solsintestop',None)

			qset = (Q(contrato__empresacontrato__empresa=id_empresa))
			if (dato or contrato_filtro or contratista_filtro or referencia or disparar_flujo):
								
	
				if dato:
					qset = qset &(Q(contrato__nombre__icontains=dato) | Q(contrato__numero__icontains=dato)| Q(numero_radicado__icontains=dato))
		
				if encabezado_id and int(encabezado_id)>0:
					qset = qset &(Q(id=encabezado_id))
	
				if mcontrato_filtro and int(mcontrato_filtro)>0:
					qset = qset & (Q(nombre__contrato__id=mcontrato_filtro))
	
				if contrato_filtro and int(contrato_filtro)>0:
	
					qset = qset & (Q(contrato_id=contrato_filtro))
	
	
				if contratista_filtro and int(contratista_filtro)>0:
						
					qset = qset & (Q(contrato__contratista__id=contratista_filtro))
	
				if referencia:
					if int(referencia)==3:
						qset = qset & (Q(referencia=''))
					elif int(referencia)==1:
						qset = qset & (~Q(referencia=''))
					elif int(referencia)==2:
						qset = qset
							
				if flujo_test:
					qset = qset & (Q(flujo_test=flujo_test))

				if disparar_flujo :						
					qset = qset & (Q(disparar_flujo = bool(disparar_flujo)))

				if detallescompletos:
					if detallescompletos=='false':
						qset = qset & (Q(detalle_encabezado_giro__test_op=''))
					else:
						qset = qset & (~Q(detalle_encabezado_giro__test_op=''))

			#print qset
			#import pdb; pdb.set_trace()

			queryset = self.model.objects.filter(qset).exclude(detalle_encabezado_giro__contratista_id__in=[39]).distinct()

			if detallescompletos:					
				if detallescompletos=='true':
					subquery_detalle= DetalleGiro.objects.filter(
						encabezado_id__in=queryset).values("encabezado_id").annotate(
						id=Sum('id')).values("encabezado_id").exclude(
						contratista_id__in=[39]).exclude(test_op__exact='')
					queryset = queryset.filter(id__in = subquery_detalle)
				else:
					subquery_detalle= DetalleGiro.objects.filter(
								estado_id__in=[enumEstados.solicitado,enumEstados.autorizado]
								,encabezado_id__in=queryset
								,test_op=''
							).values("encabezado_id").annotate(id=Sum('id')).values("encabezado_id").exclude(contratista_id__in=[39])# 39 es el contratista vega energy
					queryset = queryset.filter(id__in = subquery_detalle)	
			else:
				subquery_detalle= DetalleGiro.objects.filter(
							encabezado_id__in=queryset
							,test_op=''
						).values("encabezado_id").annotate(id=Sum('id')).values("encabezado_id").exclude(contratista_id__in=[39])# 39 es el contratista vega energy
				queryset = queryset.filter(id__in = subquery_detalle)


			if sin_paginacion is None:	
		
				page = self.paginate_queryset(queryset)
				if page is not None:

					serializer_context = {
						'request': request
					}

					if sol_sin_test_op:
						serializer = EncabezadoGiroSolSinTestOpSerializer(page,many=True,context=serializer_context)
					else:	
						serializer = self.get_serializer(page,many=True)	
					#print serializer
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
					
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
				
			else:
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
					'data':serializer.data})
											
		except Exception as e:
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#Api rest para detalle del giro
class DetalleGiroSolicitudSerializer(serializers.HyperlinkedModelSerializer):

	#encabezado= serializers.PrimaryKeyRelatedField(queryset=DEncabezadoGiro.objects.all())
	encabezado_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=DEncabezadoGiro.objects.all())
	encabezado = EncabezadoLiteSerializer(read_only=True)

	banco_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Banco.objects.all())
	banco = BancoSerializer(read_only=True)

	contratista_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=Empresa.objects.all())
	contratista = EmpresaSerializer(read_only=True)

	cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True,queryset=FinancieroCuenta.objects.all(), allow_null=True)
	cuenta = FinancieroListCuentaSerializer(read_only=True)

	estado = EstadoSerializer(read_only=True)
	estado_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Estado.objects.filter(app='EstadoGiro'))

	tipo_cuenta = TipoSerializer(read_only=True)
	tipo_cuenta_id = serializers.PrimaryKeyRelatedField(write_only=True, queryset = Tipo.objects.filter(app='cuenta'))

	rechazo = serializers.SerializerMethodField('_pagoRechazado',read_only=True)

	def _pagoRechazado(self,obj):

		rechazo = RechazoGiro.objects.filter(detalle=obj.id).first()

		if rechazo is not None:
			retorno = rechazo.atendido
		else:
			retorno = None

		return retorno

	class Meta:
		model = DetalleGiro
		fields=('id','encabezado','encabezado_id','contratista_id','contratista',
			'banco_id','banco','no_cuenta','tipo_cuenta','tipo_cuenta_id','valor_girar',
			'estado','fecha_pago','cuenta','cuenta_id','test_op',
			'fecha_pago_esperada','codigo_pago','estado_id','rechazo')	

# api detalle giro lite
class DetalleGiroSolicitudGiroViewSet(viewsets.ModelViewSet):
	"""
	Retorna una lista de los detalles del giros de cada encabezado.
	"""
	model=DetalleGiro
	queryset = model.objects.all()
	serializer_class = DetalleGiroSolicitudSerializer
	nombre_modulo='giros.detalleGiro'

	def retrieve(self,request,*args, **kwargs):
		try:
			instance = self.get_object()
			serializer = self.get_serializer(instance)
			return Response({'message':'','status':'success','data':serializer.data})
		except Exception as e:
			return Response({'message':'No se encontraron datos','success':'fail','data':''},status=status.HTTP_404_NOT_FOUND)


	def list(self, request, *args, **kwargs):
		try:
			queryset = super(DetalleGiroSolicitudGiroViewSet, self).get_queryset()
			dato = self.request.query_params.get('dato', None)
			encabezado_id= self.request.query_params.get('encabezado_id',None)
			sin_pago = self.request.query_params.get('sinpago',None)
			liteversion=self.request.query_params.get('lite',None)
			mcontrato=self.request.query_params.get('mcontrato_filtro',None)
			contratista=self.request.query_params.get('contratista_filtro',None)
			fechadesde=self.request.query_params.get('fechadesde',None)
			fechahasta=self.request.query_params.get('fechahasta',None)
			fechapago=self.request.query_params.get('fechapago',None)
			referencia=self.request.query_params.get('referencia',None)
			id_empresa = request.user.usuario.empresa.id
			sin_paginacion = self.request.query_params.get('sin_paginacion', None)
			disparar_flujo = self.request.query_params.get('disparar_flujo', 1 )
			flujo_test = self.request.query_params.get('test_op', 1)


			test_op_icontains = self.request.query_params.get('test_op_busqueda', None)
			referencia_contrato_icontains = self.request.query_params.get('referencia_contrato_busqueda', None)			

			# se usa para solicitud de giros
			sol_giro_solicitados_autorizado = self.request.query_params.get('sol_giro_solicitados_autorizado', None)
			sol_giro_autorizado = self.request.query_params.get('sol_giro_autorizado', None)

			qset = (Q(encabezado__contrato__empresacontrato__empresa = id_empresa))

			if dato:
				qset = qset &( Q(contratista__nombre__icontains=dato) | Q(banco__nombre__icontains=dato) | Q(no_cuenta__icontains=dato))

			if test_op_icontains:
				qset = qset &( Q(test_op__icontains = test_op_icontains) )

			if referencia_contrato_icontains:
				qset = qset &( Q(encabezado__contrato__nombre__icontains = referencia_contrato_icontains) | Q(encabezado__referencia__nombre__icontains = referencia_contrato_icontains ) )

			if encabezado_id:
				qset = qset &(Q(encabezado__id=encabezado_id))

			if mcontrato:
				qset = qset &(Q(encabezado__nombre__contrato__id=mcontrato))

			if contratista:
				qset = qset &(Q(contratista__id=contratista))

			if referencia:
				qset = qset &(Q(encabezado__referencia__gte=''))

			if fechadesde:
				qset = qset &(Q(fecha_pago_esperada__gte=fechadesde))

			if fechahasta:					
				qset = qset &(Q(fecha_pago_esperada__lte=fechahasta))

			if fechapago:
				qset = qset &(Q(fecha_pago=None))

			if sol_giro_solicitados_autorizado:
				# giros solicitados y autorizados
				qset = qset &(Q(estado_id__in=[enumEstados.autorizado,enumEstados.solicitado]))

			if sol_giro_autorizado:
				# giros solicitados y autorizados
				qset = qset &(Q(estado_id=enumEstados.autorizado))

			if disparar_flujo :						
				qset = qset & (Q(encabezado__disparar_flujo = bool(disparar_flujo)))

			if flujo_test :					
				qset = qset & (Q(encabezado__flujo_test= bool(flujo_test)))

				
			
			if sin_pago:
				queryset = self.model.objects.filter(qset).exclude(test_op__exact='')
			else:
				queryset = self.model.objects.filter(qset)	


			if liteversion:
				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = DetalleGiroSolicitudSerializer(page,many=True,context={'request':request})	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
	
				serializer = DetalleGiroSolicitudSerializer(queryset,many=True,context={'request':request})
				return Response({'message':'','success':'ok',
						'data':serializer.data})

			
			if sin_paginacion is None:	

				page = self.paginate_queryset(queryset)
				if page is not None:
					serializer = self.get_serializer(page,many=True)	
					return self.get_paginated_response({'message':'','success':'ok',
					'data':serializer.data})
		
				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})

			else:

				serializer = self.get_serializer(queryset,many=True)
				return Response({'message':'','success':'ok',
						'data':serializer.data})
				

		except Exception as e:
			# print(e);
			functions.toLog(e,self.nombre_modulo)
			return Response({'message':'Se presentaron errores de comunicacion con el servidor','status':'error','data':''},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
